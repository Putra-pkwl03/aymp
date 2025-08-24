from flask import Blueprint, render_template, request, redirect, url_for, flash, session, jsonify, send_file, abort, make_response
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import tempfile
from utils import (
    group_anggaran_data, flatten_anggaran_data, create_standardized_file_path,
    allowed_file, convert_decimal_for_json
)

# Import yang diperlukan akan dilakukan di dalam fungsi untuk menghindari circular import
import MySQLdb.cursors
import os
import math
import re
from datetime import datetime, timedelta
from decimal import Decimal
import locale
from calendar import monthrange
import shutil
import traceback
import logging

# Setup logging untuk mahasiswa blueprint (tanpa file handler)
logging.basicConfig(
    level=logging.DEBUG,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler()  # Hanya console handler, tidak ada file handler
    ]
)
logger = logging.getLogger('mahasiswa')
import traceback

# Fungsi helper untuk mendapatkan referensi ke app.py
def get_app_functions():
    """Mendapatkan referensi ke fungsi-fungsi dari app.py untuk menghindari circular import"""
    import app
    return {
        'mysql': app.mysql,
        'log_mahasiswa_activity': app.log_mahasiswa_activity,
        'get_mahasiswa_info_from_session': app.get_mahasiswa_info_from_session,
        'update_related_data_on_bahan_baku_change': app.update_related_data_on_bahan_baku_change,
        'update_penjualan_on_bahan_baku_change': app.update_penjualan_on_bahan_baku_change,
        'update_laba_rugi_on_produksi_change': app.update_laba_rugi_on_produksi_change,
        'update_arus_kas_otomatis': app.update_arus_kas_otomatis,
        'calculate_and_save_laba_rugi': app.calculate_and_save_laba_rugi,
        'generate_excel_laba_rugi': app.generate_excel_laba_rugi,
        'generate_pdf_laba_rugi': app.generate_pdf_laba_rugi,
        'generate_word_laba_rugi': app.generate_word_laba_rugi,
        'generate_excel_arus_kas': app.generate_excel_arus_kas,
        'generate_pdf_arus_kas': app.generate_pdf_arus_kas,
        'generate_word_arus_kas': app.generate_word_arus_kas,
        'generate_excel_neraca': app.generate_excel_neraca,
        'generate_pdf_neraca': app.generate_pdf_neraca,
        'generate_word_neraca': app.generate_word_neraca,
        'hitung_neraca_real_time': app.hitung_neraca_real_time,
        'hapus_semua_file_proposal': app.hapus_semua_file_proposal,
        'hapus_file_laporan_kemajuan': app.hapus_file_laporan_kemajuan,
        'hapus_file_laporan_akhir': app.hapus_file_laporan_akhir
    }


# Import fungsi hitung_neraca_real_time akan dilakukan secara lazy di dalam fungsi

mahasiswa_bp = Blueprint('mahasiswa', __name__)


@mahasiswa_bp.route('/edit_proposal/<int:proposal_id>', methods=['POST'])
def edit_proposal(proposal_id):
    logger.info(f"Edit proposal dipanggil untuk proposal_id: {proposal_id}")
    if 'user_type' not in session or session['user_type'] != 'mahasiswa':
        logger.warning(f"Akses ditolak untuk edit proposal: {session.get('user_type', 'None')}")
        return jsonify({'success': False, 'message': 'Akses ditolak!'})
    try:
        app_funcs = get_app_functions()
        cursor = app_funcs['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        # Ambil data proposal lama
        cursor.execute('SELECT file_path FROM proposal WHERE id=%s', (proposal_id,))
        old_data = cursor.fetchone()
        old_file_path = old_data['file_path'] if old_data else None

        # Ambil data proposal lama untuk tahapan usaha (readonly)
        cursor.execute('SELECT tahapan_usaha FROM proposal WHERE id=%s', (proposal_id,))
        old_proposal_data = cursor.fetchone()
        if not old_proposal_data:
            return jsonify({'success': False, 'message': 'Data proposal tidak ditemukan!'})
        
                # Ambil data dari form
        judul_usaha = request.form.get('edit_judul_usaha', '').strip()
        kategori = request.form.get('edit_kategori', '').strip()
        tahapan_usaha = old_proposal_data['tahapan_usaha']  # Gunakan data lama (readonly)
        merk_produk = request.form.get('edit_merk_produk', '').strip()
        
        # Handle field opsional - ubah string kosong menjadi None untuk database
        nib = request.form.get('edit_nib', '').strip() or None
        tahun_nib = request.form.get('edit_tahun_nib', '').strip() or None
        platform_penjualan = request.form.get('edit_platform_penjualan', '').strip().title() or None
        
        # Jika tahun_nib ada, konversi ke integer
        if tahun_nib:
            try:
                tahun_nib = int(tahun_nib)
            except ValueError:
                tahun_nib = None
        
        dosen_pembimbing = request.form.get('edit_dosen_pembimbing', '').strip()
        nid_dosen = request.form.get('edit_nid_dosen', '').strip()
        program_studi_dosen = request.form.get('edit_program_studi_dosen', '').strip()

        # Validasi field yang required
        if not judul_usaha:
            return jsonify({'success': False, 'message': 'Judul usaha harus diisi!'})
        if not kategori:
            return jsonify({'success': False, 'message': 'Kategori harus dipilih!'})
        if not merk_produk:
            return jsonify({'success': False, 'message': 'Merk/nama produk harus diisi!'})
        if not dosen_pembimbing:
            return jsonify({'success': False, 'message': 'Dosen pembimbing harus dipilih!'})

        # Handle file upload
        file_path = old_file_path
        file = request.files.get('edit_file_proposal')
        if file and file.filename:
            # Validasi ekstensi dan ukuran
            allowed_ext = ['.pdf', '.doc', '.docx']
            ext = os.path.splitext(file.filename)[1].lower()
            if ext not in allowed_ext:
                return jsonify({'success': False, 'message': 'File harus PDF, DOC, atau DOCX.'})
            file.seek(0, os.SEEK_END)
            if file.tell() > 16 * 1024 * 1024:
                return jsonify({'success': False, 'message': 'Ukuran file maksimal 16MB.'})
            file.seek(0)
            
            # Ambil data proposal untuk membuat path yang standar
            cursor.execute('SELECT nim, judul_usaha FROM proposal WHERE id = %s', (proposal_id,))
            proposal_data = cursor.fetchone()
            
            if not proposal_data:
                return jsonify({'success': False, 'message': 'Data proposal tidak ditemukan!'})
            
            # Ambil data mahasiswa untuk perguruan tinggi dan nama
            cursor.execute('SELECT perguruan_tinggi, nama_ketua FROM mahasiswa WHERE nim = %s', (proposal_data['nim'],))
            mahasiswa_data = cursor.fetchone()
            
            if not mahasiswa_data:
                return jsonify({'success': False, 'message': 'Data mahasiswa tidak ditemukan!'})
            
            # Hapus file lama jika ada
            if old_file_path and os.path.exists(old_file_path):
                os.remove(old_file_path)
            
            # Buat path upload yang sama seperti tambah_proposal
            safe_judul = re.sub(r'[^\w\s-]', '', proposal_data['judul_usaha']).strip().replace(' ', '_')
            safe_nama_ketua = re.sub(r'[^\w\s-]', '', mahasiswa_data['nama_ketua']).strip().replace(' ', '_')
            upload_dir = os.path.join('static', 'uploads', 'Proposal', safe_judul)
            os.makedirs(upload_dir, exist_ok=True)
            file_extension = file.filename.rsplit('.', 1)[1].lower()
            filename = f"Proposal_{safe_judul}_{safe_nama_ketua}.{file_extension}"
            file_path = os.path.join(upload_dir, filename)
            
            # Simpan file baru
            file.save(file_path)

        # Update proposal di database (tanpa mengubah tanggal_buat dan tanggal_kirim)
        cursor.execute('''
            UPDATE proposal SET
                judul_usaha=%s,
                kategori=%s,
                tahapan_usaha=%s,
                merk_produk=%s,
                nib=%s,
                tahun_nib=%s,
                platform_penjualan=%s,
                dosen_pembimbing=%s,
                nid_dosen=%s,
                program_studi_dosen=%s,
                file_path=%s,
                komentar_pembimbing=NULL,
                tanggal_komentar_pembimbing=NULL
            WHERE id=%s
        ''', (
            judul_usaha, kategori, tahapan_usaha, merk_produk, nib, tahun_nib,
            platform_penjualan, dosen_pembimbing, nid_dosen, program_studi_dosen,
            file_path, proposal_id
        ))
        app_funcs['mysql'].connection.commit()
        cursor.close()
        return jsonify({'success': True, 'message': 'Data proposal berhasil diperbarui!'})
    except Exception as e:
        logger.error(f"Error dalam edit_proposal: {str(e)}")
        logger.error(f"Traceback: {traceback.format_exc()}")
        print(f"ERROR dalam edit_proposal: {str(e)}")
        traceback.print_exc()
        return jsonify({'success': False, 'message': str(e)})

@mahasiswa_bp.route('/pengajuan_anggaran_awal_mahasiswa')
def pengajuan_anggaran_awal_mahasiswa():
    app_funcs = get_app_functions()
    print(f"DEBUG: Accessing pengajuan_anggaran_awal_mahasiswa with args: {request.args}")
    if 'user_type' not in session or session['user_type'] != 'mahasiswa':
        flash('Anda harus login sebagai mahasiswa!', 'danger')
        return redirect(url_for('index'))
    
    # Ambil proposal_id dari query parameter
    proposal_id = request.args.get('proposal_id')
    
    if not proposal_id:
        flash('ID Proposal tidak ditemukan!', 'danger')
        return redirect(url_for('mahasiswa.proposal'))
    
    # Cek status mahasiswa
    if not hasattr(app_funcs['mysql'], 'connection') or app_funcs['mysql'].connection is None:
        flash('Koneksi ke database gagal. Cek konfigurasi database!', 'danger')
        return render_template('mahasiswa/pengajuan anggaran awal.html', anggaran_data=[], proposal_id=proposal_id, proposal_info=None)
    
    try:
        # Otomasi penolakan setelah jadwal selesai untuk akun ini (guard saat membuka halaman Anggaran Awal)
        auto_reject_expired_proposals_for_current_student()

        cursor = app_funcs["mysql"].connection.cursor(MySQLdb.cursors.DictCursor)

        # Cek status mahasiswa
        cursor.execute('SELECT status FROM mahasiswa WHERE nim = %s', (session['nim'],))
        mahasiswa_data = cursor.fetchone()

        if not mahasiswa_data:
            flash('Data mahasiswa tidak ditemukan!', 'danger')
            cursor.close()
            return redirect(url_for('mahasiswa.proposal'))

        if mahasiswa_data['status'] == 'proses':
            flash('Akun Anda belum diverifikasi. Anda tidak dapat mengakses halaman anggaran!', 'danger')
            cursor.close()
            return redirect(url_for('mahasiswa.proposal'))
        elif mahasiswa_data['status'] == 'tolak':
            flash('Akun Anda ditolak. Anda tidak dapat mengakses halaman anggaran!', 'danger')
            cursor.close()
            return redirect(url_for('mahasiswa.proposal'))
        # Mahasiswa status "selesai" bisa melihat data dalam mode read-only

        # Bangun SELECT dinamis berdasarkan kolom yang tersedia pada tabel anggaran_awal
        cursor.execute("DESCRIBE anggaran_awal")
        table_structure = cursor.fetchall()
        available_columns = {col['Field'] for col in table_structure}

        base_columns = [
            'id', 'kegiatan_utama', 'kegiatan', 'penanggung_jawab', 'target_capaian',
            'nama_barang', 'kuantitas', 'satuan', 'harga_satuan', 'jumlah', 'keterangan', 'status'
        ]
        optional_columns = []
        if 'nilai_bantuan' in available_columns:
            optional_columns.append('nilai_bantuan')
        if 'status_reviewer' in available_columns:
            optional_columns.append('status_reviewer')

        select_columns = [col for col in base_columns if col in available_columns] + optional_columns
        select_clause = ", ".join(select_columns)

        # Ambil data anggaran untuk proposal ini
        cursor.execute(f'''
            SELECT {select_clause}
            FROM anggaran_awal
            WHERE id_proposal = %s
            ORDER BY kegiatan_utama, kegiatan, nama_barang
        ''', (proposal_id,))

        anggaran_data = cursor.fetchall()

        # Urutkan sesuai urutan kegiatan utama yang diinginkan
        urutan_kegiatan = [
            "Pengembangan Produk/Riset",
            "Produksi",
            "Legalitas, perizinan, sertifikasi, dan standarisasi",
            "Belanja ATK dan Penunjang"
        ]
        urutan_kegiatan_lower = [k.strip().lower() for k in urutan_kegiatan]
        anggaran_data = sorted(
            anggaran_data,
            key=lambda x: urutan_kegiatan_lower.index(x['kegiatan_utama'].strip().lower()) if x.get('kegiatan_utama') and x['kegiatan_utama'].strip().lower() in urutan_kegiatan_lower else 99
        )

        # Ambil data proposal untuk informasi dan status
        cursor.execute('SELECT judul_usaha, tahapan_usaha, status FROM proposal WHERE id = %s', (proposal_id,))
        proposal_info = cursor.fetchone()

        if not proposal_info:
            flash('Proposal tidak ditemukan!', 'danger')
            cursor.close()
            return redirect(url_for('mahasiswa.proposal'))

        cursor.close()

        grouped_data = group_anggaran_data(anggaran_data)
        anggaran_data_flat = flatten_anggaran_data(anggaran_data)

        # Hitung total nilai bantuan (hanya yang sudah direview jika kolom tersedia)
        try:
            if anggaran_data and 'status_reviewer' in anggaran_data[0]:
                total_nilai_bantuan = sum(
                    (row.get('nilai_bantuan', 0) or 0)
                    for row in anggaran_data
                    if row.get('status_reviewer') == 'sudah_direview'
                )
            else:
                total_nilai_bantuan = sum((row.get('nilai_bantuan', 0) or 0) for row in anggaran_data)
        except Exception:
            total_nilai_bantuan = 0

        # Ambil pengaturan batas anggaran (min/max) untuk UI indikator
        try:
            cursor2 = app_funcs["mysql"].connection.cursor(MySQLdb.cursors.DictCursor)
            cursor2.execute(
                """
                CREATE TABLE IF NOT EXISTS pengaturan_anggaran (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    min_total_awal BIGINT DEFAULT 0,
                    max_total_awal BIGINT DEFAULT 0,
                    min_total_bertumbuh BIGINT DEFAULT 0,
                    max_total_bertumbuh BIGINT DEFAULT 0,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP
                )
                """
            )
            app_funcs["mysql"].connection.commit()
            cursor2.execute('SELECT * FROM pengaturan_anggaran ORDER BY id DESC LIMIT 1')
            batas_row = cursor2.fetchone() or {}
            batas_min_awal = int(batas_row.get('min_total_awal') or 0)
            batas_max_awal = int(batas_row.get('max_total_awal') or 0)
            cursor2.close()
        except Exception:
            batas_min_awal = 0
            batas_max_awal = 0

        return render_template('mahasiswa/pengajuan anggaran awal.html',
                             anggaran_data=anggaran_data,
                             grouped_data=grouped_data,
                             anggaran_data_flat=anggaran_data_flat,
                             proposal_id=proposal_id,
                             proposal_info=proposal_info,
                             total_nilai_bantuan=total_nilai_bantuan,
                             batas_min_awal=batas_min_awal,
                             batas_max_awal=batas_max_awal)

    except Exception as e:
        flash(f'Error saat mengambil data anggaran: {str(e)}', 'danger')
        return render_template('mahasiswa/pengajuan anggaran awal.html',
                             anggaran_data=[],
                             proposal_id=proposal_id,
                             proposal_info=None,
                             total_nilai_bantuan=0)

@mahasiswa_bp.route('/pengajuan_anggaran_bertumbuh_mahasiswa')
def pengajuan_anggaran_bertumbuh_mahasiswa():
    app_funcs = get_app_functions()
    print(f"DEBUG: Accessing pengajuan_anggaran_bertumbuh_mahasiswa with args: {request.args}")
    if 'user_type' not in session or session['user_type'] != 'mahasiswa':
        flash('Anda harus login sebagai mahasiswa!', 'danger')
        return redirect(url_for('index'))
    
    # Ambil proposal_id dari query parameter
    proposal_id = request.args.get('proposal_id')
    
    if not proposal_id:
        flash('ID Proposal tidak ditemukan!', 'danger')
        return redirect(url_for('mahasiswa.proposal'))
    
    # Cek status mahasiswa
    if not hasattr(app_funcs["mysql"], 'connection') or app_funcs["mysql"].connection is None:
        flash('Koneksi ke database gagal. Cek konfigurasi database!', 'danger')
        return render_template('mahasiswa/pengajuan anggaran bertumbuh.html', anggaran_data=[], proposal_id=proposal_id, proposal_info=None)
    
    try:
        # Otomasi penolakan setelah jadwal selesai untuk akun ini (guard saat membuka halaman Anggaran Bertumbuh)
        auto_reject_expired_proposals_for_current_student()

        cursor = app_funcs["mysql"].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Cek status mahasiswa
        cursor.execute('SELECT status FROM mahasiswa WHERE nim = %s', (session['nim'],))
        mahasiswa_data = cursor.fetchone()
        
        if not mahasiswa_data:
            flash('Data mahasiswa tidak ditemukan!', 'danger')
            cursor.close()
            return redirect(url_for('mahasiswa.proposal'))
        
        if mahasiswa_data['status'] == 'proses':
            flash('Akun Anda belum diverifikasi. Anda tidak dapat mengakses halaman anggaran!', 'danger')
            cursor.close()
            return redirect(url_for('mahasiswa.proposal'))
        elif mahasiswa_data['status'] == 'tolak':
            flash('Akun Anda ditolak. Anda tidak dapat mengakses halaman anggaran!', 'danger')
            cursor.close()
            return redirect(url_for('mahasiswa.proposal'))
        # Mahasiswa status "selesai" bisa melihat data dalam mode read-only
        
        # Cek apakah tabel anggaran_bertumbuh ada
        cursor.execute("SHOW TABLES LIKE 'anggaran_bertumbuh'")
        table_exists = cursor.fetchone()
        print(f"DEBUG: Table anggaran_bertumbuh exists: {table_exists is not None}")
        
        if table_exists:
            # Cek struktur tabel
            cursor.execute("DESCRIBE anggaran_bertumbuh")
            table_structure = cursor.fetchall()
            print(f"DEBUG: Table structure: {[col['Field'] for col in table_structure]}")
            
            # Cek apakah ada data untuk proposal ini
            cursor.execute("SELECT COUNT(*) as count FROM anggaran_bertumbuh WHERE id_proposal = %s", (proposal_id,))
            count_result = cursor.fetchone()
            print(f"DEBUG: Total records for proposal {proposal_id}: {count_result['count']}")

        # Ambil data anggaran untuk proposal ini
        print(f"DEBUG: Querying anggaran_bertumbuh for proposal_id: {proposal_id}")
        cursor.execute('''
            SELECT id, kegiatan_utama, kegiatan, penanggung_jawab, target_capaian, 
                   nama_barang, kuantitas, satuan, harga_satuan, jumlah, keterangan, status, nilai_bantuan
            FROM anggaran_bertumbuh 
            WHERE id_proposal = %s 
            ORDER BY kegiatan_utama, kegiatan, nama_barang
        ''', (proposal_id,))
        
        anggaran_data = cursor.fetchall()
        print(f"DEBUG: Found {len(anggaran_data)} anggaran records")
        
        # Debug: print first few records
        for i, record in enumerate(anggaran_data[:3]):
            print(f"DEBUG: Record {i+1}: {record}")
        
        # Urutkan sesuai urutan kegiatan utama yang diinginkan
        urutan_kegiatan = [
            "Pengembangan Pasar dan Saluran Distribusi",
            "Pengembangan Produk/Riset",
            "Produksi",
            "Legalitas, perizinan, sertifikasi, dan standarisasi",
            "Belanja ATK dan Penunjang"
        ]
        urutan_kegiatan_lower = [k.strip().lower() for k in urutan_kegiatan]
        anggaran_data = sorted(
            anggaran_data,
            key=lambda x: urutan_kegiatan_lower.index(x['kegiatan_utama'].strip().lower()) if x['kegiatan_utama'] and x['kegiatan_utama'].strip().lower() in urutan_kegiatan_lower else 99
        )
        
        # Ambil data proposal untuk informasi dan status
        cursor.execute('SELECT judul_usaha, tahapan_usaha, status FROM proposal WHERE id = %s', (proposal_id,))
        proposal_info = cursor.fetchone()
        
        if not proposal_info:
            flash('Proposal tidak ditemukan!', 'danger')
            cursor.close()
            return redirect(url_for('mahasiswa.proposal'))
        
        cursor.close()
        
        grouped_data = group_anggaran_data(anggaran_data)
        anggaran_data_flat = flatten_anggaran_data(anggaran_data)
        
        # Hitung total nilai bantuan
        total_nilai_bantuan = sum(row.get('nilai_bantuan', 0) or 0 for row in anggaran_data)
        
        print(f"DEBUG: grouped_data length: {len(grouped_data) if grouped_data else 0}")
        print(f"DEBUG: anggaran_data_flat length: {len(anggaran_data_flat) if anggaran_data_flat else 0}")
        print(f"DEBUG: total_nilai_bantuan: {total_nilai_bantuan}")
        print(f"DEBUG: proposal_info: {proposal_info}")
        
        # Ambil pengaturan batas anggaran (min/max) untuk UI badge
        try:
            cursor2 = app_funcs["mysql"].connection.cursor(MySQLdb.cursors.DictCursor)
            cursor2.execute('SELECT * FROM pengaturan_anggaran ORDER BY id DESC LIMIT 1')
            batas_row = cursor2.fetchone() or {}
            batas_min_bertumbuh = int(batas_row.get('min_total_bertumbuh') or 0)
            batas_max_bertumbuh = int(batas_row.get('max_total_bertumbuh') or 0)
            cursor2.close()
        except Exception:
            batas_min_bertumbuh = 0
            batas_max_bertumbuh = 0

        return render_template('mahasiswa/pengajuan anggaran bertumbuh.html', 
                             anggaran_data=anggaran_data, 
                             grouped_data=grouped_data,
                             anggaran_data_flat=anggaran_data_flat,
                             proposal_id=proposal_id,
                             proposal_info=proposal_info,
                             total_nilai_bantuan=total_nilai_bantuan,
                             batas_min_bertumbuh=batas_min_bertumbuh,
                             batas_max_bertumbuh=batas_max_bertumbuh)
        
    except Exception as e:
        flash(f'Error saat mengambil data anggaran: {str(e)}', 'danger')
        return render_template('mahasiswa/pengajuan anggaran bertumbuh.html', 
                             anggaran_data=[], 
                             proposal_id=proposal_id,
                             proposal_info=None,
                             total_nilai_bantuan=0)
@mahasiswa_bp.route('/dashboard_mahasiswa')
def dashboard_mahasiswa():
    logger.info("Dashboard mahasiswa dipanggil")
    app_funcs = get_app_functions()
    if 'user_type' not in session or session['user_type'] != 'mahasiswa':
        logger.warning(f"Akses ditolak untuk dashboard mahasiswa: {session.get('user_type', 'None')}")
        flash('Anda harus login sebagai mahasiswa!', 'danger')
        return redirect(url_for('index'))
    
    if not hasattr(app_funcs["mysql"], 'connection') or app_funcs["mysql"].connection is None:
        logger.error("Koneksi database tidak tersedia untuk dashboard mahasiswa")
        flash('Koneksi ke database gagal. Cek konfigurasi database!', 'danger')
        return render_template('mahasiswa/index mahasiswa.html', 
                             proposals=[], 
                             anggota_tim_list=[],
                             total_pendapatan=0,
                             total_biaya_produksi=0,
                             total_laba_kotor=0,
                             total_laba_bersih=0,
                             chart_dates=[],
                             chart_produksi=[],
                             chart_penjualan=[],
                             chart_weekly_dates=[],
                             chart_weekly_produksi=[],
                             chart_weekly_penjualan=[],
                             chart_monthly_dates=[],
                             chart_monthly_produksi=[],
                             chart_monthly_penjualan=[])
    
    try:
        cursor = app_funcs["mysql"].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Ambil data proposal mahasiswa
        cursor.execute('''
            SELECT p.*, 
                   COUNT(at.id) as jumlah_anggota
            FROM proposal p 
            LEFT JOIN anggota_tim at ON p.id = at.id_proposal
            WHERE p.nim = %s 
            GROUP BY p.id 
            ORDER BY p.tanggal_buat DESC
        ''', (session['nim'],))
        
        proposals = cursor.fetchall()
        
        # Ambil data anggota tim untuk proposal yang aktif
        anggota_tim_list = []
        if proposals:
            # Ambil proposal yang paling baru atau yang aktif
            active_proposal = proposals[0]  # Ambil proposal pertama (paling baru)
            
            cursor.execute('''
                SELECT at.*, p.judul_usaha
                FROM anggota_tim at
                JOIN proposal p ON at.id_proposal = p.id
                WHERE at.id_proposal = %s
                ORDER BY at.tanggal_tambah ASC
            ''', (active_proposal['id'],))
            
            anggota_tim_list = cursor.fetchall()
        
        # Ambil data laba rugi untuk dashboard
        total_pendapatan = 0
        total_biaya_produksi = 0
        total_laba_kotor = 0
        total_laba_bersih = 0
        
        # Data untuk chart (default kosong)
        from datetime import datetime, timedelta, date
        import math
        from calendar import monthrange
        
        today = datetime.now()
        days_since_monday = today.weekday()
        monday = today - timedelta(days=days_since_monday)
        
        nama_hari = ['Senin', 'Selasa', 'Rabu', 'Kamis', 'Jumat', 'Sabtu', 'Minggu']
        chart_dates = []
        for i in range(7):
            tanggal = monday + timedelta(days=i)
            hari = nama_hari[i]
            date_str = f"{hari} ({tanggal.strftime('%d/%m')})"
            chart_dates.append(date_str)
        
        chart_produksi = [0, 0, 0, 0, 0, 0, 0]
        chart_penjualan = [0, 0, 0, 0, 0, 0, 0]
        
        # Data untuk chart mingguan (kosong, akan diisi jika ada proposal)
        chart_weekly_dates = []
        chart_weekly_produksi = []
        chart_weekly_penjualan = []
        
        # Data untuk chart bulanan (kosong, akan diisi jika ada proposal)
        chart_monthly_dates = []
        chart_monthly_produksi = []
        chart_monthly_penjualan = []
        
        if proposals:
            # Ambil proposal yang aktif
            active_proposal = proposals[0]
            
            cursor.execute('''
                SELECT 
                    SUM(pendapatan) as total_pendapatan,
                    SUM(total_biaya_produksi) as total_biaya_produksi,
                    SUM(laba_rugi_kotor) as total_laba_kotor,
                    SUM(laba_rugi_bersih) as total_laba_bersih
                FROM laba_rugi 
                WHERE proposal_id = %s
            ''', (active_proposal['id'],))
            
            laba_rugi_data = cursor.fetchone()
            
            if laba_rugi_data:
                total_pendapatan = laba_rugi_data['total_pendapatan'] or 0
                total_biaya_produksi = laba_rugi_data['total_biaya_produksi'] or 0
                total_laba_kotor = laba_rugi_data['total_laba_kotor'] or 0
                total_laba_bersih = laba_rugi_data['total_laba_bersih'] or 0
            
            # Ambil data untuk chart (7 hari terakhir dari tanggal terakhir data)
            cursor.execute('''
                SELECT 
                    DATE(tanggal_produksi) as tanggal,
                    SUM(jumlah_produk) as total_produksi
                FROM produksi 
                WHERE proposal_id = %s 
                AND tanggal_produksi >= (
                    SELECT DATE_SUB(MAX(tanggal_produksi), INTERVAL 6 DAY)
                    FROM produksi 
                    WHERE proposal_id = %s
                )
                GROUP BY DATE(tanggal_produksi)
                ORDER BY tanggal_produksi
            ''', (active_proposal['id'], active_proposal['id']))
            
            produksi_data = cursor.fetchall()
            
            cursor.execute('''
                SELECT 
                    DATE(tanggal_penjualan) as tanggal,
                    SUM(quantity) as total_penjualan
                FROM penjualan 
                WHERE proposal_id = %s 
                AND tanggal_penjualan >= (
                    SELECT DATE_SUB(MAX(tanggal_penjualan), INTERVAL 6 DAY)
                    FROM penjualan 
                    WHERE proposal_id = %s
                )
                GROUP BY DATE(tanggal_penjualan)
                ORDER BY tanggal_penjualan
            ''', (active_proposal['id'], active_proposal['id']))
            
            penjualan_data = cursor.fetchall()
            
            # Buat data untuk chart (urutan Senin-Minggu)
            from datetime import datetime, timedelta
            
            # Urutan hari tetap Senin-Minggu
            nama_hari = ['Senin', 'Selasa', 'Rabu', 'Kamis', 'Jumat', 'Sabtu', 'Minggu']
            hari_map = {0: 'Senin', 1: 'Selasa', 2: 'Rabu', 3: 'Kamis', 4: 'Jumat', 5: 'Sabtu', 6: 'Minggu'}
            produksi_per_hari = {h: 0 for h in nama_hari}
            penjualan_per_hari = {h: 0 for h in nama_hari}

            # Isi data produksi berdasarkan hari
            for prod in produksi_data:
                # Konversi tanggal ke datetime jika perlu
                tanggal = prod['tanggal']
                if isinstance(tanggal, date):
                    tanggal = datetime.combine(tanggal, datetime.min.time())
                
                if hasattr(tanggal, 'weekday'):
                    hari_idx = tanggal.weekday()  # 0=Senin, 6=Minggu
                    hari = hari_map[hari_idx]
                    produksi_per_hari[hari] = float(prod['total_produksi'])

            # Isi data penjualan berdasarkan hari
            for penj in penjualan_data:
                # Konversi tanggal ke datetime jika perlu
                tanggal = penj['tanggal']
                if isinstance(tanggal, date):
                    tanggal = datetime.combine(tanggal, datetime.min.time())
                
                if hasattr(tanggal, 'weekday'):
                    hari_idx = tanggal.weekday()
                    hari = hari_map[hari_idx]
                    penjualan_per_hari[hari] = float(penj['total_penjualan'])

            # Buat chart_dates dengan format "Hari (dd/mm)"
            chart_dates = []
            chart_produksi = []
            chart_penjualan = []
            
            # Tentukan tanggal awal berdasarkan data yang ada
            if produksi_data or penjualan_data:
                # Ambil tanggal terakhir dari data yang ada
                all_dates = []
                for prod in produksi_data:
                    all_dates.append(prod['tanggal'])
                for penj in penjualan_data:
                    all_dates.append(penj['tanggal'])
                
                if all_dates:
                    # Ambil tanggal terakhir dan hitung 7 hari ke belakang
                    last_date = max(all_dates)
                    if isinstance(last_date, date):
                        last_date = datetime.combine(last_date, datetime.min.time())
                    
                    # Hitung hari Senin dari minggu yang berisi tanggal terakhir
                    days_since_monday = last_date.weekday()
                    monday = last_date - timedelta(days=days_since_monday)
                else:
                    # Fallback ke minggu ini jika tidak ada data
                    today = datetime.now()
                    days_since_monday = today.weekday()
                    monday = today - timedelta(days=days_since_monday)
            else:
                # Jika tidak ada data sama sekali, gunakan minggu ini
                today = datetime.now()
                days_since_monday = today.weekday()
                monday = today - timedelta(days=days_since_monday)
            
            for i in range(7):
                tanggal = monday + timedelta(days=i)
                hari = nama_hari[i]
                date_str = f"{hari} ({tanggal.strftime('%d/%m')})"
                chart_dates.append(date_str)
                chart_produksi.append(produksi_per_hari[hari])
                chart_penjualan.append(penjualan_per_hari[hari])
            
            # Data untuk chart mingguan (berdasarkan bulan saat ini)
            
            # Hitung jumlah minggu dalam bulan saat ini
            current_year = datetime.now().year
            current_month = datetime.now().month
            _, days_in_month = monthrange(current_year, current_month)
            
            # Hitung jumlah minggu (membagi jumlah hari dengan 7 dan membulatkan ke atas)
            total_weeks = math.ceil(days_in_month / 7)
            
            # Ambil data produksi untuk bulan saat ini
            cursor.execute('''
                SELECT 
                    DAY(tanggal_produksi) as hari,
                    SUM(jumlah_produk) as total_produksi
                FROM produksi 
                WHERE proposal_id = %s 
                AND YEAR(tanggal_produksi) = %s 
                AND MONTH(tanggal_produksi) = %s
                GROUP BY DAY(tanggal_produksi)
                ORDER BY hari
            ''', (active_proposal['id'], current_year, current_month))
            
            produksi_weekly_data = cursor.fetchall()
            
            # Ambil data penjualan untuk bulan saat ini
            cursor.execute('''
                SELECT 
                    DAY(tanggal_penjualan) as hari,
                    SUM(quantity) as total_penjualan
                FROM penjualan 
                WHERE proposal_id = %s 
                AND YEAR(tanggal_penjualan) = %s 
                AND MONTH(tanggal_penjualan) = %s
                GROUP BY DAY(tanggal_penjualan)
                ORDER BY hari
            ''', (active_proposal['id'], current_year, current_month))
            
            penjualan_weekly_data = cursor.fetchall()
            
            # Generate minggu berdasarkan bulan saat ini
            for week_num in range(1, total_weeks + 1):
                week_str = f"Minggu {week_num}"
                chart_weekly_dates.append(week_str)
                
                # Hitung range hari untuk minggu ini
                start_day = (week_num - 1) * 7 + 1
                end_day = min(week_num * 7, days_in_month)
                
                # Hitung total produksi untuk minggu ini
                produksi_weekly_value = 0
                for prod in produksi_weekly_data:
                    # Pastikan prod['hari'] adalah integer
                    hari = int(prod['hari']) if prod['hari'] is not None else 0
                    if start_day <= hari <= end_day:
                        produksi_weekly_value += float(prod['total_produksi'])
                chart_weekly_produksi.append(produksi_weekly_value)
                
                # Hitung total penjualan untuk minggu ini
                penjualan_weekly_value = 0
                for penj in penjualan_weekly_data:
                    # Pastikan penj['hari'] adalah integer
                    hari = int(penj['hari']) if penj['hari'] is not None else 0
                    if start_day <= hari <= end_day:
                        penjualan_weekly_value += float(penj['total_penjualan'])
                chart_weekly_penjualan.append(penjualan_weekly_value)
            
            # Data untuk chart bulanan (12 bulan terakhir)
            cursor.execute('''
                SELECT 
                    YEAR(tanggal_produksi) as tahun,
                    MONTH(tanggal_produksi) as bulan,
                    SUM(jumlah_produk) as total_produksi
                FROM produksi 
                WHERE proposal_id = %s 
                AND tanggal_produksi >= DATE_SUB(CURDATE(), INTERVAL 12 MONTH)
                GROUP BY YEAR(tanggal_produksi), MONTH(tanggal_produksi)
                ORDER BY tahun, bulan
            ''', (active_proposal['id'],))
            
            produksi_monthly_data = cursor.fetchall()
            
            cursor.execute('''
                SELECT 
                    YEAR(tanggal_penjualan) as tahun,
                    MONTH(tanggal_penjualan) as bulan,
                    SUM(quantity) as total_penjualan
                FROM penjualan 
                WHERE proposal_id = %s 
                AND tanggal_penjualan >= DATE_SUB(CURDATE(), INTERVAL 12 MONTH)
                GROUP BY YEAR(tanggal_penjualan), MONTH(tanggal_penjualan)
                ORDER BY tahun, bulan
            ''', (active_proposal['id'],))
            
            penjualan_monthly_data = cursor.fetchall()
            
            # Generate 12 bulan (Jan sampai Des) untuk tahun saat ini
            nama_bulan = ['Jan', 'Feb', 'Mar', 'Apr', 'Mei', 'Jun', 
                         'Jul', 'Ags', 'Sep', 'Okt', 'Nov', 'Des']
            
            current_year = datetime.now().year
            
            # Generate semua bulan dari Jan sampai Des
            for bulan in range(1, 13):
                month_name = nama_bulan[bulan - 1]
                month_str = f"{month_name} {current_year}"
                chart_monthly_dates.append(month_str)
                
                # Cari data produksi untuk bulan ini
                produksi_monthly_value = 0
                for prod in produksi_monthly_data:
                    if prod['tahun'] == current_year and prod['bulan'] == bulan:
                        produksi_monthly_value = float(prod['total_produksi'])
                        break
                chart_monthly_produksi.append(produksi_monthly_value)
                
                # Cari data penjualan untuk bulan ini
                penjualan_monthly_value = 0
                for penj in penjualan_monthly_data:
                    if penj['tahun'] == current_year and penj['bulan'] == bulan:
                        penjualan_monthly_value = float(penj['total_penjualan'])
                        break
                chart_monthly_penjualan.append(penjualan_monthly_value)
        
        # Jika tidak ada proposal, isi data default untuk chart
        if not proposals:
            # Data default untuk chart mingguan (berdasarkan bulan saat ini)
            
            current_year = datetime.now().year
            current_month = datetime.now().month
            _, days_in_month = monthrange(current_year, current_month)
            total_weeks = math.ceil(days_in_month / 7)
            
            chart_weekly_dates = [f"Minggu {i+1}" for i in range(total_weeks)]
            chart_weekly_produksi = [0] * total_weeks
            chart_weekly_penjualan = [0] * total_weeks
            
            # Data default untuk chart bulanan
            current_year = datetime.now().year
            nama_bulan = ['Jan', 'Feb', 'Mar', 'Apr', 'Mei', 'Jun', 'Jul', 'Ags', 'Sep', 'Okt', 'Nov', 'Des']
            chart_monthly_dates = [f"{bulan} {current_year}" for bulan in nama_bulan]
            chart_monthly_produksi = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
            chart_monthly_penjualan = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
        else:
            # Pastikan semua variabel chart terdefinisi jika ada proposal
            if not chart_weekly_dates:
                current_year = datetime.now().year
                current_month = datetime.now().month
                _, days_in_month = monthrange(current_year, current_month)
                total_weeks = math.ceil(days_in_month / 7)
                chart_weekly_dates = [f"Minggu {i+1}" for i in range(total_weeks)]
                chart_weekly_produksi = [0] * total_weeks
                chart_weekly_penjualan = [0] * total_weeks
            
            if not chart_monthly_dates:
                current_year = datetime.now().year
                nama_bulan = ['Jan', 'Feb', 'Mar', 'Apr', 'Mei', 'Jun', 'Jul', 'Ags', 'Sep', 'Okt', 'Nov', 'Des']
                chart_monthly_dates = [f"{bulan} {current_year}" for bulan in nama_bulan]
                chart_monthly_produksi = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
                chart_monthly_penjualan = [0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0]
        
        cursor.close()
        
        return render_template('mahasiswa/index mahasiswa.html', 
                             proposals=proposals, 
                             anggota_tim_list=anggota_tim_list,
                             total_pendapatan=total_pendapatan,
                             total_biaya_produksi=total_biaya_produksi,
                             total_laba_kotor=total_laba_kotor,
                             total_laba_bersih=total_laba_bersih,
                             chart_dates=chart_dates,
                             chart_produksi=chart_produksi,
                             chart_penjualan=chart_penjualan,
                             chart_weekly_dates=chart_weekly_dates,
                             chart_weekly_produksi=chart_weekly_produksi,
                             chart_weekly_penjualan=chart_weekly_penjualan,
                             chart_monthly_dates=chart_monthly_dates,
                             chart_monthly_produksi=chart_monthly_produksi,
                             chart_monthly_penjualan=chart_monthly_penjualan)
        
    except Exception as e:
        logger.error(f"Error dalam dashboard_mahasiswa: {str(e)}")
        flash(f'Error saat mengambil data proposal: {str(e)}', 'danger')
        return render_template('mahasiswa/index mahasiswa.html', 
                             proposals=[], 
                             anggota_tim_list=[],
                             total_pendapatan=0,
                             total_biaya_produksi=0,
                             total_laba_kotor=0,
                             total_laba_bersih=0,
                             chart_dates=[],
                             chart_produksi=[],
                             chart_penjualan=[],
                             chart_weekly_dates=[],
                             chart_weekly_produksi=[],
                             chart_weekly_penjualan=[],
                             chart_monthly_dates=[],
                             chart_monthly_produksi=[],
                             chart_monthly_penjualan=[])
@mahasiswa_bp.route('/laporan_kemajuan')
def laporan_kemajuan():
    app_funcs = get_app_functions()
    if 'user_type' not in session or session['user_type'] != 'mahasiswa':
        flash('Anda harus login sebagai mahasiswa!', 'danger')
        return redirect(url_for('index'))
    
    if not hasattr(app_funcs["mysql"], 'connection') or app_funcs["mysql"].connection is None:
        flash('Koneksi ke database gagal. Cek konfigurasi database!', 'danger')
        return render_template(
            'mahasiswa/laporan_kemajuan_awal_bertumbuh.html', 
            anggaran_data=[], 
            grouped_data=[],
            anggaran_data_flat=[],
            mahasiswa_info=None, 
            tabel_laporan='',
            proposal_id=None,
            tabel_anggaran='',
            total_nilai_bantuan=0,
            kemajuan_sudah_mulai=False,
            file_laporan=None
        )
    
    try:
        # Auto-reject laporan kemajuan yang melewati jadwal selesai untuk akun ini
        auto_reject_expired_kemajuan_for_current_student()
        
        # Restore laporan kemajuan yang ditolak sistem jika jadwal belum selesai
        restore_prematurely_rejected_kemajuan_for_current_student()
        
        # Pastikan event scheduler untuk auto-reject ada
        ensure_kemajuan_auto_reject_events()
        
        cursor = app_funcs["mysql"].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Ambil data mahasiswa dan proposal
        cursor.execute('''
            SELECT m.*, p.judul_usaha, p.tahapan_usaha, p.status as status_proposal, p.id as proposal_id
            FROM mahasiswa m 
            LEFT JOIN proposal p ON m.nim = p.nim
            WHERE m.nim = %s
        ''', (session['nim'],))
        
        mahasiswa_info = cursor.fetchone()
        if not mahasiswa_info:
            flash('Data mahasiswa tidak ditemukan!', 'danger')
            cursor.close()
            return redirect(url_for('index'))
        
        # Cek status mahasiswa - status "selesai" tetap bisa melihat data dalam mode read-only
        
        # Tentukan jenis anggaran berdasarkan tahapan usaha
        tahapan_usaha = mahasiswa_info.get('tahapan_usaha', '').lower()
        if 'bertumbuh' in tahapan_usaha:
            tabel_anggaran = 'anggaran_bertumbuh'
            tabel_laporan = 'laporan_kemajuan_bertumbuh'
        else:
            tabel_anggaran = 'anggaran_awal'
            tabel_laporan = 'laporan_kemajuan_awal'
        
        proposal_id = mahasiswa_info.get('proposal_id')
        
        if not proposal_id:
            flash('Proposal tidak ditemukan!', 'danger')
            cursor.close()
            return render_template(
                'mahasiswa/laporan_kemajuan_awal_bertumbuh.html', 
                anggaran_data=[], 
                grouped_data=[],
                anggaran_data_flat=[],
                mahasiswa_info=mahasiswa_info, 
                tabel_laporan=tabel_laporan,
                proposal_id=None,
                tabel_anggaran=tabel_anggaran,
                total_nilai_bantuan=0,
                kemajuan_sudah_mulai=False,
                file_laporan=None
            )
        
        # Buat tabel laporan kemajuan jika belum ada
        if tabel_laporan == 'laporan_kemajuan_awal':
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS laporan_kemajuan_awal (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    id_proposal INT NOT NULL,
                    kegiatan_utama VARCHAR(255),
                    kegiatan VARCHAR(255) NOT NULL,
                    penanggung_jawab VARCHAR(100) NOT NULL,
                    target_capaian TEXT NOT NULL,
                    nama_barang VARCHAR(255) NOT NULL,
                    kuantitas INT NOT NULL,
                    satuan VARCHAR(50) NOT NULL,
                    harga_satuan DECIMAL(15,2) NOT NULL,
                    jumlah DECIMAL(15,2) NOT NULL,
                    keterangan TEXT,
                    tanggal_buat TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    status ENUM('draf','diajukan','disetujui','ditolak','revisi') DEFAULT 'draf',
                    tanggal_review TIMESTAMP NULL,
                    nilai_bantuan DECIMAL(15,2) DEFAULT 0.00,
                    FOREIGN KEY (id_proposal) REFERENCES proposal(id) ON DELETE CASCADE
                )
            ''')
        else:
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS laporan_kemajuan_bertumbuh (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    id_proposal INT NOT NULL,
                    kegiatan_utama VARCHAR(255) NOT NULL,
                    kegiatan VARCHAR(255) NOT NULL,
                    penanggung_jawab VARCHAR(100) NOT NULL,
                    target_capaian TEXT NOT NULL,
                    nama_barang VARCHAR(255) NOT NULL,
                    kuantitas INT NOT NULL,
                    satuan VARCHAR(50) NOT NULL,
                    harga_satuan DECIMAL(15,2) NOT NULL,
                    jumlah DECIMAL(15,2) NOT NULL,
                    keterangan TEXT,
                    tanggal_buat TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    status ENUM('draf','diajukan','disetujui','ditolak','revisi') DEFAULT 'draf',
                    tanggal_review TIMESTAMP NULL,
                    nilai_bantuan DECIMAL(15,2) DEFAULT 0.00,
                    FOREIGN KEY (id_proposal) REFERENCES proposal(id) ON DELETE CASCADE
                )
            ''')
        
        # Buat tabel file laporan kemajuan terpisah
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS file_laporan_kemajuan (
                id INT AUTO_INCREMENT PRIMARY KEY,
                id_proposal INT NOT NULL,
                nama_file VARCHAR(255) NOT NULL,
                file_path VARCHAR(500) NOT NULL,
                status ENUM('draf', 'diajukan', 'disetujui', 'revisi') DEFAULT 'draf',
                komentar_pembimbing TEXT,
                tanggal_upload TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                tanggal_update TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                FOREIGN KEY (id_proposal) REFERENCES proposal(id) ON DELETE CASCADE
            )
        ''')
        
        # Ambil jadwal kemajuan (mulai & selesai)
        kemajuan_mulai_dt = None
        kemajuan_selesai_dt = None
        try:
            cursor.execute('SELECT kemajuan_mulai, kemajuan_selesai FROM pengaturan_jadwal ORDER BY id DESC LIMIT 1')
            jadwal_row = cursor.fetchone()
            if jadwal_row:
                kemajuan_mulai_dt = jadwal_row.get('kemajuan_mulai')
                kemajuan_selesai_dt = jadwal_row.get('kemajuan_selesai')
        except Exception:
            kemajuan_mulai_dt = None
            kemajuan_selesai_dt = None

        # Tentukan window jadwal menggunakan evaluasi langsung di DB agar bebas offset zona waktu
        try:
            cursor.execute(
                '''
                SELECT 
                    NOW()             AS server_now,
                    (kemajuan_mulai IS NOT NULL AND NOW() >= kemajuan_mulai) AS sudah_mulai_db,
                    (kemajuan_selesai IS NOT NULL AND NOW() > (kemajuan_selesai + INTERVAL 2 MINUTE)) AS after_window_db,
                    (kemajuan_mulai IS NOT NULL AND NOW() >= kemajuan_mulai 
                        AND (kemajuan_selesai IS NULL OR NOW() <= (kemajuan_selesai + INTERVAL 2 MINUTE))) AS in_window_db
                FROM pengaturan_jadwal 
                ORDER BY id DESC LIMIT 1
                '''
            )
            guard_row = cursor.fetchone() or {}
            now = guard_row.get('server_now') or datetime.now()
            kemajuan_sudah_mulai = bool(guard_row.get('sudah_mulai_db'))
            after_window = bool(guard_row.get('after_window_db'))
            in_window = bool(guard_row.get('in_window_db'))
        except Exception:
            # Fallback ke perhitungan lokal jika query gagal
            now = datetime.now()
            in_window = bool(
                kemajuan_mulai_dt and now >= kemajuan_mulai_dt and (not kemajuan_selesai_dt or now <= kemajuan_selesai_dt)
            )
            after_window = bool(kemajuan_selesai_dt and now > kemajuan_selesai_dt)
            kemajuan_sudah_mulai = bool(kemajuan_mulai_dt and now >= kemajuan_mulai_dt)

        # Pastikan event scheduler DB untuk auto-reject sudah dibuat (best-effort, abaikan error privilese)
        try:
            ensure_kemajuan_auto_reject_events()
        except Exception:
            pass

        # Jalankan auto-reject on-view hanya bila jadwal sudah lewat (fallback jika event DB belum berjalan)
        if after_window:
            auto_reject_expired_kemajuan_for_current_student()
        else:
            # Jika BELUM lewat jadwal, pastikan tidak ada status yang terlanjur 'ditolak' otomatis
            restore_prematurely_rejected_kemajuan_for_current_student()

        # Cek apakah sudah ada data laporan kemajuan untuk proposal ini
        cursor.execute(f'''
            SELECT COUNT(*) as count FROM {tabel_laporan} WHERE id_proposal = %s
        ''', (proposal_id,))
        existing_count = cursor.fetchone()['count']
        
        # Jika belum ada data laporan kemajuan, salin data dari anggaran yang sudah direview oleh reviewer
        # HANYA jika berada dalam window jadwal (mulai <= sekarang <= selesai)
        can_copy_from_schedule = in_window

        if existing_count == 0 and can_copy_from_schedule:
            cursor.execute(f'''
                SELECT id, kegiatan_utama, kegiatan, penanggung_jawab, target_capaian, 
                       nama_barang, satuan, status, nilai_bantuan
                FROM {tabel_anggaran} 
                WHERE id_proposal = %s AND status_reviewer = 'sudah_direview'
                ORDER BY kegiatan_utama, kegiatan, nama_barang
            ''', (proposal_id,))
            
            anggaran_data = cursor.fetchall()
            
            if anggaran_data:
                for anggaran in anggaran_data:
                    cursor.execute(f'''
                        INSERT INTO {tabel_laporan} (
                            id_proposal, kegiatan_utama, kegiatan, penanggung_jawab, target_capaian,
                            nama_barang, kuantitas, satuan, harga_satuan, jumlah, keterangan, status, nilai_bantuan
                        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    ''', (
                        proposal_id, anggaran['kegiatan_utama'], anggaran['kegiatan'], 
                        anggaran['penanggung_jawab'], anggaran['target_capaian'], anggaran['nama_barang'],
                        0, anggaran['satuan'], 0, 
                        0, '', 'draf', anggaran.get('nilai_bantuan', 0)
                    ))
                
                app_funcs["mysql"].connection.commit()
        
        # Ambil data laporan kemajuan (jika sudah ada)
        cursor.execute(f'''
            SELECT id, kegiatan_utama, kegiatan, penanggung_jawab, target_capaian, 
                   nama_barang, kuantitas, satuan, harga_satuan, jumlah, keterangan, status, nilai_bantuan
            FROM {tabel_laporan} 
            WHERE id_proposal = %s
            ORDER BY kegiatan_utama, kegiatan, nama_barang
        ''', (proposal_id,))
        laporan_data = cursor.fetchall()
        
        # Jika belum ada data laporan kemajuan, gunakan data anggaran yang sudah direview sebagai referensi
        if not laporan_data:
            cursor.execute(f'''
                SELECT id, kegiatan_utama, kegiatan, penanggung_jawab, target_capaian, 
                       nama_barang, 0 as kuantitas, satuan, 0 as harga_satuan, 0 as jumlah, '' as keterangan, 'belum_mulai' as status, nilai_bantuan
                FROM {tabel_anggaran} 
                WHERE id_proposal = %s AND status_reviewer = 'sudah_direview'
                ORDER BY kegiatan_utama, kegiatan, nama_barang
            ''', (proposal_id,))
            laporan_data = cursor.fetchall()
        
        # Urutkan data sesuai urutan kegiatan utama
        if 'bertumbuh' in tahapan_usaha:
            urutan_kegiatan = [
                "Pengembangan Pasar dan Saluran Distribusi",
                "Pengembangan Produk/Riset",
                "Produksi",
                "Legalitas, perizinan, sertifikasi, dan standarisasi",
                "Belanja ATK dan Penunjang"
            ]
        else:
            urutan_kegiatan = [
                "Pengembangan Produk/Riset",
                "Produksi",
                "Legalitas, perizinan, sertifikasi, dan standarisasi",
                "Belanja ATK dan Penunjang"
            ]
        
        laporan_data = sorted(
            laporan_data,
            key=lambda x: urutan_kegiatan.index(x['kegiatan_utama']) if x['kegiatan_utama'] in urutan_kegiatan else 99
        )
        
        grouped_data = group_anggaran_data(laporan_data)
        laporan_data_flat = flatten_anggaran_data(laporan_data)
        
        # Hitung total nilai bantuan dari data laporan kemajuan
        total_nilai_bantuan = 0
        try:
            total_nilai_bantuan = int(sum((item.get('nilai_bantuan', 0) or 0) for item in laporan_data))
        except Exception:
            total_nilai_bantuan = 0

        # Fallback: jika belum ada data laporan (mis. sebelum jadwal mulai) atau total masih 0,
        # ambil total nilai bantuan langsung dari tabel anggaran yang sudah direview
        if not total_nilai_bantuan:
            try:
                cursor2 = app_funcs["mysql"].connection.cursor(MySQLdb.cursors.DictCursor)
                # Cek apakah kolom status_reviewer ada
                cursor2.execute(f"SHOW COLUMNS FROM {tabel_anggaran} LIKE 'status_reviewer'")
                has_status_reviewer = cursor2.fetchone() is not None
                if has_status_reviewer:
                    cursor2.execute(
                        f"""
                        SELECT COALESCE(SUM(nilai_bantuan), 0) AS total
                        FROM {tabel_anggaran}
                        WHERE id_proposal = %s AND status_reviewer = 'sudah_direview'
                        """,
                        (proposal_id,)
                    )
                else:
                    cursor2.execute(
                        f"""
                        SELECT COALESCE(SUM(nilai_bantuan), 0) AS total
                        FROM {tabel_anggaran}
                        WHERE id_proposal = %s
                        """,
                        (proposal_id,)
                    )
                row_total = cursor2.fetchone() or {}
                cursor2.close()
                total_nilai_bantuan = int(row_total.get('total') or 0)
            except Exception:
                # Abaikan jika tabel/kolom belum ada; biarkan total tetap 0
                total_nilai_bantuan = total_nilai_bantuan or 0
        
        # Ambil data file laporan kemajuan
        cursor.execute('''
            SELECT * FROM file_laporan_kemajuan 
            WHERE id_proposal = %s
            ORDER BY tanggal_upload DESC 
            LIMIT 1
        ''', (proposal_id,))
        file_laporan = cursor.fetchone()
        
        cursor.close()
        
        # Debug: Log data yang akan dikirim ke template
        print(f"DEBUG: Total data laporan kemajuan: {len(laporan_data)}")
        print(f"DEBUG: Total data flat: {len(laporan_data_flat)}")
        print(f"DEBUG: Total nilai bantuan: {total_nilai_bantuan}")
        if laporan_data:
            print(f"DEBUG: Sample data: {laporan_data[0]}")
        
        # Get jadwal kemajuan status
        jadwal_kemajuan = get_jadwal_kemajuan_status()
        print(f"DEBUG: Jadwal kemajuan yang dikirim ke template: {jadwal_kemajuan}")
        
        return render_template('mahasiswa/laporan_kemajuan_awal_bertumbuh.html', 
                             anggaran_data=laporan_data, 
                             grouped_data=grouped_data,
                             anggaran_data_flat=laporan_data_flat,
                             mahasiswa_info=mahasiswa_info,
                             tabel_laporan=tabel_laporan,
                             proposal_id=proposal_id,
                             tabel_anggaran=tabel_anggaran,
                             total_nilai_bantuan=total_nilai_bantuan,
                             kemajuan_sudah_mulai=kemajuan_sudah_mulai,
                             file_laporan=file_laporan,
                             jadwal_kemajuan=jadwal_kemajuan)
        
    except Exception as e:
        flash(f'Error saat mengambil data laporan kemajuan: {str(e)}', 'danger')
        return render_template('mahasiswa/laporan_kemajuan_awal_bertumbuh.html', 
                             anggaran_data=[], 
                             grouped_data=[],
                             anggaran_data_flat=[],
                             mahasiswa_info=None,
                             tabel_laporan='',
                             proposal_id=None,
                             tabel_anggaran='',
                             total_nilai_bantuan=0,
                             kemajuan_sudah_mulai=False,
                             file_laporan=None)
@mahasiswa_bp.route('/laporan_akhir')
def laporan_akhir():
    app_funcs = get_app_functions()
    if 'user_type' not in session or session['user_type'] != 'mahasiswa':
        flash('Anda harus login sebagai mahasiswa!', 'danger')
    
    if not hasattr(app_funcs["mysql"], 'connection') or app_funcs["mysql"].connection is None:
        flash('Koneksi ke database gagal. Cek konfigurasi database!', 'danger')
        return render_template('mahasiswa/laporan_akhir_awal_bertumbuh.html', anggaran_data=[], mahasiswa_info=None, proposal_id=None)
    
    try:
        # Auto-reject laporan akhir yang melewati jadwal selesai untuk akun ini
        auto_reject_expired_akhir_for_current_student()
        
        # Restore laporan akhir yang ditolak sistem jika jadwal belum selesai
        restore_prematurely_rejected_akhir_for_current_student()

        cursor = app_funcs["mysql"].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Ambil data mahasiswa dan proposal
        cursor.execute('''
            SELECT m.*, p.judul_usaha, p.tahapan_usaha, p.status as status_proposal, p.id as proposal_id
            FROM mahasiswa m 
            LEFT JOIN proposal p ON m.nim = p.nim
            WHERE m.nim = %s
        ''', (session['nim'],))
        
        mahasiswa_info = cursor.fetchone()
        if not mahasiswa_info:
            flash('Data mahasiswa tidak ditemukan!', 'danger')
            cursor.close()
            return redirect(url_for('index'))
        
        # Cek status mahasiswa - status "selesai" tetap bisa melihat data dalam mode read-only
        
        # Tentukan jenis laporan berdasarkan tahapan usaha
        tahapan_usaha = mahasiswa_info.get('tahapan_usaha', '').lower()
        if 'bertumbuh' in tahapan_usaha:
            tabel_laporan_kemajuan = 'laporan_kemajuan_bertumbuh'
            tabel_laporan_akhir = 'laporan_akhir_bertumbuh'
        else:
            tabel_laporan_kemajuan = 'laporan_kemajuan_awal'
            tabel_laporan_akhir = 'laporan_akhir_awal'
        
        proposal_id = mahasiswa_info.get('proposal_id')
        
        if not proposal_id:
            flash('Proposal tidak ditemukan!', 'danger')
            cursor.close()
            return render_template('mahasiswa/laporan_akhir_awal_bertumbuh.html', 
                                 anggaran_data=[], mahasiswa_info=mahasiswa_info, proposal_id=None)
        
        # Buat tabel laporan akhir jika belum ada
        if tabel_laporan_akhir == 'laporan_akhir_awal':
            # Tambahkan kolom nilai_bantuan jika belum ada
            try:
                cursor.execute('ALTER TABLE laporan_akhir_awal ADD COLUMN nilai_bantuan DECIMAL(15,2) DEFAULT 0.00')
                app_funcs["mysql"].connection.commit()
                print(f"DEBUG: Berhasil menambahkan kolom nilai_bantuan ke tabel laporan_akhir_awal")
            except Exception as e:
                if "Duplicate column name" in str(e):
                    print(f"DEBUG: Kolom nilai_bantuan sudah ada di tabel laporan_akhir_awal")
                else:
                    print(f"DEBUG: Error menambahkan kolom nilai_bantuan: {str(e)}")
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS laporan_akhir_awal (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    id_proposal INT NOT NULL,
                    kegiatan_utama VARCHAR(255),
                    kegiatan VARCHAR(255) NOT NULL,
                    penanggung_jawab VARCHAR(100) NOT NULL,
                    target_capaian TEXT NOT NULL,
                    nama_barang VARCHAR(255) NOT NULL,
                    kuantitas INT NOT NULL,
                    satuan VARCHAR(50) NOT NULL,
                    harga_satuan DECIMAL(15,2) NOT NULL,
                    jumlah DECIMAL(15,2) NOT NULL,
                    keterangan TEXT,
                    tanggal_buat TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    status ENUM('draf','diajukan','disetujui','ditolak','revisi') DEFAULT 'draf',
                    tanggal_review TIMESTAMP NULL,
                    nilai_bantuan DECIMAL(15,2) DEFAULT 0.00,
                    FOREIGN KEY (id_proposal) REFERENCES proposal(id) ON DELETE CASCADE
                )
            ''')
        else:
            # Samakan skema dengan tabel awal: pastikan kolom nilai_bantuan ada
            try:
                cursor.execute('ALTER TABLE laporan_akhir_bertumbuh ADD COLUMN nilai_bantuan DECIMAL(15,2) DEFAULT 0.00')
                app_funcs["mysql"].connection.commit()
                print("DEBUG: Berhasil menambahkan kolom nilai_bantuan ke tabel laporan_akhir_bertumbuh")
            except Exception as e:
                if "Duplicate column name" in str(e):
                    print("DEBUG: Kolom nilai_bantuan sudah ada di tabel laporan_akhir_bertumbuh")
                else:
                    print(f"DEBUG: Error menambahkan kolom nilai_bantuan (bertumbuh): {str(e)}")

            cursor.execute('''
                CREATE TABLE IF NOT EXISTS laporan_akhir_bertumbuh (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    id_proposal INT NOT NULL,
                    kegiatan_utama VARCHAR(255) NOT NULL,
                    kegiatan VARCHAR(255) NOT NULL,
                    penanggung_jawab VARCHAR(100) NOT NULL,
                    target_capaian TEXT NOT NULL,
                    nama_barang VARCHAR(255) NOT NULL,
                    kuantitas INT NOT NULL,
                    satuan VARCHAR(50) NOT NULL,
                    harga_satuan DECIMAL(15,2) NOT NULL,
                    jumlah DECIMAL(15,2) NOT NULL,
                    keterangan TEXT,
                    tanggal_buat TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    status ENUM('draf','diajukan','disetujui','ditolak','revisi') DEFAULT 'draf',
                    tanggal_review TIMESTAMP NULL,
                    nilai_bantuan DECIMAL(15,2) DEFAULT 0.00,
                    FOREIGN KEY (id_proposal) REFERENCES proposal(id) ON DELETE CASCADE
                )
            ''')
        
        # Cek apakah sudah ada data laporan akhir untuk proposal ini
        cursor.execute(f'''
            SELECT COUNT(*) as count FROM {tabel_laporan_akhir} WHERE id_proposal = %s
        ''', (proposal_id,))
        existing_count = cursor.fetchone()['count']
        print(f"DEBUG: Ditemukan {existing_count} data laporan akhir yang sudah ada")
        
        # Jika belum ada data laporan akhir, buat dari laporan kemajuan
        # Catatan: Penyalinan HANYA dilakukan setelah jadwal laporan akhir dimulai (akhir_mulai)
        can_copy_final_from_schedule = False
        try:
            cursor.execute('SELECT akhir_mulai FROM pengaturan_jadwal ORDER BY id DESC LIMIT 1')
            jadwal_row = cursor.fetchone()
            if jadwal_row and jadwal_row.get('akhir_mulai'):
                akhir_mulai = jadwal_row['akhir_mulai']
                now = datetime.now()
                if now >= akhir_mulai:
                    can_copy_final_from_schedule = True
        except Exception:
            can_copy_final_from_schedule = False

        # Hanya salin data laporan kemajuan ke laporan akhir setelah jadwal Laporan Akhir dimulai
        if existing_count == 0 and can_copy_final_from_schedule:
            # Tambahkan kolom rekomendasi_pendanaan jika belum ada
            try:
                cursor.execute('ALTER TABLE penilaian_laporan_kemajuan ADD COLUMN rekomendasi_pendanaan VARCHAR(50) DEFAULT NULL')
                app_funcs["mysql"].connection.commit()
                print("DEBUG: Berhasil menambahkan kolom rekomendasi_pendanaan ke tabel penilaian_laporan_kemajuan")
            except Exception as e:
                if "Duplicate column name" in str(e):
                    print("DEBUG: Kolom rekomendasi_pendanaan sudah ada di tabel penilaian_laporan_kemajuan")
                else:
                    print(f"DEBUG: Error menambahkan kolom rekomendasi_pendanaan: {str(e)}")
            
            # Cek status rekomendasi pendanaan dari penilaian laporan kemajuan
            cursor.execute('''
                SELECT rekomendasi_pendanaan 
                FROM penilaian_laporan_kemajuan 
                WHERE id_proposal = %s AND rekomendasi_pendanaan IS NOT NULL
                ORDER BY id DESC 
                LIMIT 1
            ''', (proposal_id,))
            
            penilaian_result = cursor.fetchone()
            rekomendasi_pendanaan = penilaian_result['rekomendasi_pendanaan'] if penilaian_result else None
            
            print(f"DEBUG: Rekomendasi pendanaan untuk proposal {proposal_id}: {rekomendasi_pendanaan}")
            
            # Jika rekomendasi adalah 'berhenti pendanaan', tidak buat laporan akhir
            if rekomendasi_pendanaan == 'berhenti pendanaan':
                print(f"DEBUG: Rekomendasi pendanaan 'berhenti pendanaan' untuk proposal {proposal_id}. Laporan akhir tidak dibuat.")
                # Siapkan nilai default agar template tidak error ketika memakai variabel lain
                cursor.execute('''
                    SELECT 0 AS total_laporan_kemajuan
                ''')
                total_laporan_kemajuan_disetujui = 0
                total_nilai_bantuan = 0
                file_laporan_akhir = None
                cursor.close()
                return render_template(
                    'mahasiswa/laporan_akhir_awal_bertumbuh.html',
                    anggaran_data=[],
                    grouped_data=[],
                    anggaran_data_flat=[],
                    mahasiswa_info=mahasiswa_info,
                    tabel_laporan=tabel_laporan_akhir,
                    proposal_id=proposal_id,
                    total_laporan_kemajuan_disetujui=total_laporan_kemajuan_disetujui,
                    total_nilai_bantuan=total_nilai_bantuan,
                    file_laporan_akhir=file_laporan_akhir
                )
            
            # Jika rekomendasi adalah 'lanjutkan pendanaan' atau belum ada penilaian, cek laporan kemajuan yang disetujui
            if rekomendasi_pendanaan == 'lanjutkan pendanaan' or rekomendasi_pendanaan is None:
                # Tentukan tabel laporan kemajuan
                if 'bertumbuh' in tahapan_usaha:
                    tabel_laporan_kemajuan = 'laporan_kemajuan_bertumbuh'
                else:
                    tabel_laporan_kemajuan = 'laporan_kemajuan_awal'
                print(f"DEBUG: Mengecek laporan kemajuan di tabel {tabel_laporan_kemajuan} untuk proposal {proposal_id}")
                cursor.execute(f'''
                    SELECT COUNT(*) as count FROM {tabel_laporan_kemajuan} WHERE id_proposal = %s AND status = 'disetujui'
                ''', (proposal_id,))
                approved_count = cursor.fetchone()['count']
                print(f"DEBUG: Ditemukan {approved_count} data laporan kemajuan dengan status 'disetujui'")
                
                # Jika ada data kemajuan yang disetujui, salin dengan reset field tertentu
                if approved_count > 0:
                    print(f"DEBUG: Menyalin {approved_count} data dari laporan kemajuan ke laporan akhir")
                    print(f"DEBUG: Field yang akan di-reset: kuantitas=0, harga_satuan=0, jumlah=0, keterangan=''")
                    cursor.execute(f'''
                        SELECT id, kegiatan_utama, kegiatan, penanggung_jawab, target_capaian, 
                               nama_barang, satuan, status
                        FROM {tabel_laporan_kemajuan} 
                        WHERE id_proposal = %s AND status = 'disetujui'
                        ORDER BY kegiatan_utama, kegiatan, nama_barang
                    ''', (proposal_id,))
                    laporan_kemajuan_data = cursor.fetchall()
                    if laporan_kemajuan_data:
                        for laporan in laporan_kemajuan_data:
                            cursor.execute(f'''
                                INSERT INTO {tabel_laporan_akhir} (
                                    id_proposal, kegiatan_utama, kegiatan, penanggung_jawab, target_capaian,
                                    nama_barang, kuantitas, satuan, harga_satuan, jumlah, keterangan, status, nilai_bantuan
                                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                            ''', (
                                proposal_id, laporan['kegiatan_utama'], laporan['kegiatan'], 
                                laporan['penanggung_jawab'], laporan['target_capaian'], laporan['nama_barang'],
                                0, laporan['satuan'], 0, 0, '', 'draf', 0
                            ))
                        app_funcs["mysql"].connection.commit()
                        print(f"DEBUG: Berhasil menyalin {len(laporan_kemajuan_data)} data ke laporan akhir")
                        print(f"DEBUG: Field yang disalin: kegiatan_utama, kegiatan, penanggung_jawab, target_capaian, nama_barang, satuan")
                        print(f"DEBUG: Field yang di-reset: kuantitas=0, harga_satuan=0, jumlah=0, keterangan=''")
                    else:
                        print(f"DEBUG: Tidak ada data laporan kemajuan yang disetujui, tidak menyalin apapun")
                else:
                    print(f"DEBUG: Tidak ada data laporan kemajuan yang disetujui, tidak menyalin apapun")
            else:
                print(f"DEBUG: Rekomendasi pendanaan bukan 'lanjutkan', tidak membuat laporan akhir")
        else:
            print(f"DEBUG: Data laporan akhir sudah ada, tidak perlu membuat ulang")
        
        # Ambil data laporan akhir untuk ditampilkan (UI tetap di-guard oleh jadwal untuk aksi)
        cursor.execute(f'''
            SELECT id, kegiatan_utama, kegiatan, penanggung_jawab, target_capaian, 
                   nama_barang, kuantitas, satuan, harga_satuan, jumlah, keterangan, status, nilai_bantuan
            FROM {tabel_laporan_akhir} 
            WHERE id_proposal = %s
            ORDER BY kegiatan_utama, kegiatan, nama_barang
        ''', (proposal_id,))
        laporan_data = cursor.fetchall()
        
        # Urutkan data sesuai urutan kegiatan utama
        if 'bertumbuh' in tahapan_usaha:
            urutan_kegiatan = [
                "Pengembangan Pasar dan Saluran Distribusi",
                "Pengembangan Produk/Riset",
                "Produksi",
                "Legalitas, perizinan, sertifikasi, dan standarisasi",
                "Belanja ATK dan Penunjang"
            ]
        else:
            urutan_kegiatan = [
                "Pengembangan Produk/Riset",
                "Produksi",
                "Legalitas, perizinan, sertifikasi, dan standarisasi",
                "Belanja ATK dan Penunjang"
            ]
        
        laporan_data = sorted(
            laporan_data,
            key=lambda x: urutan_kegiatan.index(x['kegiatan_utama']) if x['kegiatan_utama'] in urutan_kegiatan else 99
        )
        
        grouped_data = group_anggaran_data(laporan_data)
        laporan_data_flat = flatten_anggaran_data(laporan_data)
        
        # Hitung total nilai bantuan dari anggaran yang sudah direview
        if 'bertumbuh' in tahapan_usaha:
            tabel_anggaran = 'anggaran_bertumbuh'
        else:
            tabel_anggaran = 'anggaran_awal'
            
        cursor.execute(f'''
            SELECT SUM(nilai_bantuan) as total_nilai_bantuan
            FROM {tabel_anggaran} 
            WHERE id_proposal = %s AND status_reviewer = 'sudah_direview'
        ''', (proposal_id,))
        
        nilai_bantuan_result = cursor.fetchone()
        total_nilai_bantuan = nilai_bantuan_result['total_nilai_bantuan'] or 0
        
        # Hitung total laporan kemajuan yang disetujui
        cursor.execute(f'''
            SELECT SUM(jumlah) as total_laporan_kemajuan
            FROM {tabel_laporan_kemajuan} 
            WHERE id_proposal = %s AND status = 'disetujui'
        ''', (proposal_id,))
        
        laporan_kemajuan_result = cursor.fetchone()
        total_laporan_kemajuan_disetujui = laporan_kemajuan_result['total_laporan_kemajuan'] or 0
        
        # Buat tabel file_laporan_akhir jika belum ada
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS file_laporan_akhir (
                id INT AUTO_INCREMENT PRIMARY KEY,
                id_proposal INT NOT NULL,
                nama_file VARCHAR(255) NOT NULL,
                file_path VARCHAR(500) NOT NULL,
                status ENUM('draf', 'diajukan', 'disetujui', 'revisi') DEFAULT 'draf',
                komentar_pembimbing TEXT,
                tanggal_upload TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                tanggal_update TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                FOREIGN KEY (id_proposal) REFERENCES proposal(id) ON DELETE CASCADE
            )
        ''')
        
        # Ambil data file laporan akhir
        cursor.execute('''
            SELECT * FROM file_laporan_akhir 
            WHERE id_proposal = %s
        ''', (proposal_id,))
        
        file_laporan_akhir = cursor.fetchone()
        
        cursor.close()
        
        # Get jadwal akhir status
        jadwal_akhir = get_jadwal_akhir_status()
        
        return render_template('mahasiswa/laporan_akhir_awal_bertumbuh.html', 
                             anggaran_data=laporan_data, 
                             grouped_data=grouped_data,
                             anggaran_data_flat=laporan_data_flat,
                             mahasiswa_info=mahasiswa_info,
                             tabel_laporan=tabel_laporan_akhir,
                             proposal_id=proposal_id,
                             total_laporan_kemajuan_disetujui=total_laporan_kemajuan_disetujui,
                             total_nilai_bantuan=total_nilai_bantuan,
                             file_laporan_akhir=file_laporan_akhir,
                             jadwal_akhir=jadwal_akhir)
        
    except Exception as e:
        flash(f'Error saat mengambil data laporan akhir: {str(e)}', 'danger')
        return render_template('mahasiswa/laporan_akhir_awal_bertumbuh.html', 
                             anggaran_data=[], 
                             mahasiswa_info=None,
                             proposal_id=None,
                             total_laporan_kemajuan_disetujui=0,
                             total_nilai_bantuan=0,
                             file_laporan_akhir=None)

@mahasiswa_bp.route('/get_pengaturan_jadwal')
def get_pengaturan_jadwal_mhs():
    """Mahasiswa mengambil pengaturan jadwal untuk guard UI"""
    app_funcs = get_app_functions()
    if not hasattr(app_funcs["mysql"], 'connection') or app_funcs["mysql"].connection is None:
        return jsonify({'success': False, 'message': 'Koneksi ke database gagal!'})
    try:
        cursor = app_funcs["mysql"].connection.cursor(MySQLdb.cursors.DictCursor)
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS pengaturan_jadwal (
                id INT AUTO_INCREMENT PRIMARY KEY,
                proposal_mulai DATETIME NULL,
                proposal_selesai DATETIME NULL,
                kemajuan_mulai DATETIME NULL,
                kemajuan_selesai DATETIME NULL,
                akhir_mulai DATETIME NULL,
                akhir_selesai DATETIME NULL,
                pembimbing_nilai_mulai DATETIME NULL,
                pembimbing_nilai_selesai DATETIME NULL,
                reviewer_proposal_mulai DATETIME NULL,
                reviewer_proposal_selesai DATETIME NULL,
                reviewer_kemajuan_mulai DATETIME NULL,
                reviewer_kemajuan_selesai DATETIME NULL,
                reviewer_akhir_mulai DATETIME NULL,
                reviewer_akhir_selesai DATETIME NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP
            )
        ''')
        # Ensure kolom baru tersedia (kompatibel versi lama)
        def ensure_column(column_name: str, add_sql: str) -> None:
            cursor.execute("SHOW COLUMNS FROM pengaturan_jadwal LIKE %s", (column_name,))
            if cursor.fetchone() is None:
                cursor.execute(f"ALTER TABLE pengaturan_jadwal ADD COLUMN {add_sql}")
                app_funcs["mysql"].connection.commit()

        ensure_column('pembimbing_nilai_mulai', 'pembimbing_nilai_mulai DATETIME NULL')
        ensure_column('pembimbing_nilai_selesai', 'pembimbing_nilai_selesai DATETIME NULL')
        ensure_column('reviewer_proposal_mulai', 'reviewer_proposal_mulai DATETIME NULL')
        ensure_column('reviewer_proposal_selesai', 'reviewer_proposal_selesai DATETIME NULL')
        ensure_column('reviewer_kemajuan_mulai', 'reviewer_kemajuan_mulai DATETIME NULL')
        ensure_column('reviewer_kemajuan_selesai', 'reviewer_kemajuan_selesai DATETIME NULL')
        ensure_column('reviewer_akhir_mulai', 'reviewer_akhir_mulai DATETIME NULL')
        ensure_column('reviewer_akhir_selesai', 'reviewer_akhir_selesai DATETIME NULL')
        app_funcs["mysql"].connection.commit()
        cursor.execute('SELECT * FROM pengaturan_jadwal ORDER BY id DESC LIMIT 1')
        row = cursor.fetchone()
        cursor.close()
        def fmt(dt):
            try:
                return dt.strftime('%Y-%m-%dT%H:%M') if dt else None
            except Exception:
                return None
        data = {
            'proposal_mulai': fmt(row.get('proposal_mulai')) if row else None,
            'proposal_selesai': fmt(row.get('proposal_selesai')) if row else None,
            'kemajuan_mulai': fmt(row.get('kemajuan_mulai')) if row else None,
            'kemajuan_selesai': fmt(row.get('kemajuan_selesai')) if row else None,
            'akhir_mulai': fmt(row.get('akhir_mulai')) if row else None,
            'akhir_selesai': fmt(row.get('akhir_selesai')) if row else None,
            # jadwal penilaian
            'pembimbing_nilai_mulai': fmt(row.get('pembimbing_nilai_mulai')) if row else None,
            'pembimbing_nilai_selesai': fmt(row.get('pembimbing_nilai_selesai')) if row else None,
            'reviewer_proposal_mulai': fmt(row.get('reviewer_proposal_mulai')) if row else None,
            'reviewer_proposal_selesai': fmt(row.get('reviewer_proposal_selesai')) if row else None,
            'reviewer_kemajuan_mulai': fmt(row.get('reviewer_kemajuan_mulai')) if row else None,
            'reviewer_kemajuan_selesai': fmt(row.get('reviewer_kemajuan_selesai')) if row else None,
            'reviewer_akhir_mulai': fmt(row.get('reviewer_akhir_mulai')) if row else None,
            'reviewer_akhir_selesai': fmt(row.get('reviewer_akhir_selesai')) if row else None,
        }
        return jsonify({'success': True, 'data': data})
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})


def get_jadwal_kemajuan_status():
    """Get status jadwal untuk upload laporan kemajuan"""
    try:
        app_funcs = get_app_functions()
        if not hasattr(app_funcs["mysql"], 'connection') or app_funcs["mysql"].connection is None:
            print("DEBUG: Database connection not available")
            return {
                'bisa_upload': True,
                'status': 'Error koneksi',
                'pesan': 'Bisa upload kapan saja',
                'warna': 'info'
            }
        
        cursor = app_funcs["mysql"].connection.cursor(MySQLdb.cursors.DictCursor)
        cursor.execute('SELECT kemajuan_mulai, kemajuan_selesai FROM pengaturan_jadwal ORDER BY id DESC LIMIT 1')
        row = cursor.fetchone()
        cursor.close()
        
        print(f"DEBUG: Jadwal row from database: {row}")
        
        if not row or not row.get('kemajuan_mulai') or not row.get('kemajuan_selesai'):
            print("DEBUG: Tidak ada jadwal yang valid, mengizinkan upload")
            return {
                'bisa_upload': True,
                'status': 'Tidak ada jadwal',
                'pesan': 'Bisa upload kapan saja',
                'warna': 'info'
            }
        
        now = datetime.now()
        mulai = row['kemajuan_mulai']
        selesai = row['kemajuan_selesai']
        
        print(f"DEBUG: Sekarang: {now}")
        print(f"DEBUG: Jadwal mulai: {mulai}")
        print(f"DEBUG: Jadwal selesai: {selesai}")
        
        if now < mulai:
            result = {
                'bisa_upload': False,
                'status': 'Belum waktunya',
                'pesan': f'Upload dibuka pada {mulai.strftime("%Y-%m-%d %H:%M")}',
                'warna': 'warning',
                'mulai': mulai.strftime('%Y-%m-%d %H:%M'),
                'selesai': selesai.strftime('%Y-%m-%d %H:%M')
            }
            print(f"DEBUG: Hasil jadwal - sebelum mulai: {result}")
            return result
        elif now > selesai:
            result = {
                'bisa_upload': False,
                'status': 'Jadwal berakhir',
                'pesan': f'Upload ditutup pada {selesai.strftime("%Y-%m-%d %H:%M")}',
                'warna': 'danger',
                'mulai': mulai.strftime('%Y-%m-%d %H:%M'),
                'selesai': selesai.strftime('%Y-%m-%d %H:%M')
            }
            print(f"DEBUG: Hasil jadwal - setelah selesai: {result}")
            return result
        else:
            result = {
                'bisa_upload': True,
                'status': 'Jadwal aktif',
                'pesan': f'Upload aktif sampai {selesai.strftime("%Y-%m-%d %H:%M")}',
                'warna': 'success',
                'mulai': mulai.strftime('%Y-%m-%d %H:%M'),
                'selesai': selesai.strftime('%Y-%m-%d %H:%M')
            }
            print(f"DEBUG: Hasil jadwal - dalam rentang: {result}")
            return result
    except Exception as e:
        print(f"Error getting jadwal kemajuan status: {e}")
        return {
            'bisa_upload': True,
            'status': 'Error jadwal',
            'pesan': 'Bisa upload kapan saja',
            'warna': 'info'
        }

def get_jadwal_akhir_status():
    """Get status jadwal untuk upload laporan akhir"""
    try:
        app_funcs = get_app_functions()
        if not hasattr(app_funcs["mysql"], 'connection') or app_funcs["mysql"].connection is None:
            return {
                'bisa_upload': True,
                'status': 'Error koneksi',
                'pesan': 'Bisa upload kapan saja',
                'warna': 'info'
            }
        
        cursor = app_funcs["mysql"].connection.cursor(MySQLdb.cursors.DictCursor)
        cursor.execute('SELECT akhir_mulai, akhir_selesai FROM pengaturan_jadwal ORDER BY id DESC LIMIT 1')
        row = cursor.fetchone()
        cursor.close()
        
        if not row or not row.get('akhir_mulai') or not row.get('akhir_selesai'):
            return {
                'bisa_upload': True,
                'status': 'Tidak ada jadwal',
                'pesan': 'Bisa upload kapan saja - Status: Lanjutkan Pendanaan',
                'warna': 'info'
            }
        
        now = datetime.now()
        mulai = row['akhir_mulai']
        selesai = row['akhir_selesai']
        
        if now < mulai:
            return {
                'bisa_upload': False,
                'status': 'Belum waktunya',
                'pesan': f'Upload dibuka pada {mulai.strftime("%Y-%m-%d %H:%M")} - Status: Lanjutkan Pendanaan',
                'warna': 'warning',
                'mulai': mulai.strftime('%Y-%m-%d %H:%M'),
                'selesai': selesai.strftime('%Y-%m-%d %H:%M')
            }
        elif now > selesai:
            return {
                'bisa_upload': False,
                'status': 'Jadwal berakhir',
                'pesan': f'Upload ditutup pada {selesai.strftime("%Y-%m-%d %H:%M")} - Status: Tidak Melakukan Upload File',
                'warna': 'danger',
                'mulai': mulai.strftime('%Y-%m-%d %H:%M'),
                'selesai': selesai.strftime('%Y-%m-%d %H:%M')
            }
        else:
            return {
                'bisa_upload': True,
                'status': 'Jadwal aktif',
                'pesan': f'Upload aktif sampai {selesai.strftime("%Y-%m-%d %H:%M")} - Status: Lanjutkan Pendanaan',
                'warna': 'success',
                'mulai': mulai.strftime('%Y-%m-%d %H:%M'),
                'selesai': selesai.strftime('%Y-%m-%d %H:%M')
            }
    except Exception as e:
        print(f"Error getting jadwal akhir status: {e}")
        return {
            'bisa_upload': True,
            'status': 'Error jadwal',
            'pesan': 'Bisa upload kapan saja - Status: Lanjutkan Pendanaan',
            'warna': 'info'
        }


# Helper: Otomatis menolak proposal dan anggaran yang masih "draf/diajukan/revisi" jika sudah melewati jadwal proposal selesai
def auto_reject_expired_proposals_for_current_student():
    app_funcs = get_app_functions()
    if not hasattr(app_funcs["mysql"], 'connection') or app_funcs["mysql"].connection is None:
        return
    try:
        cursor = app_funcs["mysql"].connection.cursor(MySQLdb.cursors.DictCursor)
        # Ambil jadwal terbaru
        cursor.execute('SELECT proposal_selesai FROM pengaturan_jadwal ORDER BY id DESC LIMIT 1')
        row = cursor.fetchone()
        if not row or not row.get('proposal_selesai'):
            cursor.close()
            return
        batas_selesai = row['proposal_selesai']
        now = datetime.now()
        if not batas_selesai or now <= batas_selesai:
            cursor.close()
            return

        # Update status proposal mahasiswa ini menjadi 'ditolak' jika masih draf/diajukan/revisi
        cursor.execute('''
            UPDATE proposal 
            SET status = 'ditolak'
            WHERE nim = %s AND status IN ('draf','diajukan','revisi')
        ''', (session['nim'],))

        # Update status anggaran_awal terkait
        cursor.execute('''
            UPDATE anggaran_awal aa
            JOIN proposal p ON aa.id_proposal = p.id
            SET aa.status = 'ditolak'
            WHERE p.nim = %s AND aa.status IN ('draf','diajukan','revisi')
        ''', (session['nim'],))

        # Update status anggaran_bertumbuh terkait (jika tabel ada)
        cursor.execute("SHOW TABLES LIKE 'anggaran_bertumbuh'")
        if cursor.fetchone():
            cursor.execute('''
                UPDATE anggaran_bertumbuh ab
                JOIN proposal p ON ab.id_proposal = p.id
                SET ab.status = 'ditolak'
                WHERE p.nim = %s AND ab.status IN ('draf','diajukan','revisi')
            ''', (session['nim'],))

        if hasattr(app_funcs["mysql"], 'connection') and app_funcs["mysql"].connection is not None:
            app_funcs["mysql"].connection.commit()
        cursor.close()
    except Exception:
        try:
            cursor.close()
        except Exception:
            pass


# Helper: Otomatis menolak laporan kemajuan (awal/bertumbuh) yang masih "draf/diajukan/revisi" jika sudah melewati jadwal kemajuan selesai
def auto_reject_expired_kemajuan_for_current_student():
    app_funcs = get_app_functions()
    if not hasattr(app_funcs["mysql"], 'connection') or app_funcs["mysql"].connection is None:
        return
    try:
        cursor = app_funcs["mysql"].connection.cursor(MySQLdb.cursors.DictCursor)
        # Ambil jadwal kemajuan selesai terbaru
        cursor.execute('SELECT kemajuan_selesai FROM pengaturan_jadwal ORDER BY id DESC LIMIT 1')
        row = cursor.fetchone()
        if not row or not row.get('kemajuan_selesai'):
            cursor.close()
            return
        batas_selesai = row['kemajuan_selesai']
        # Gunakan waktu server database untuk membandingkan jadwal
        try:
            cursor.execute('SELECT NOW() AS server_now')
            _now_row = cursor.fetchone() or {}
            now = _now_row.get('server_now') or datetime.now()
        except Exception:
            now = datetime.now()

        if not batas_selesai or now <= batas_selesai:
            cursor.close()
            return

        # Update laporan kemajuan (awal) milik mahasiswa ini menjadi 'ditolak' jika masih draf/diajukan/revisi
        # KECUALI status sudah 'disetujui'
        cursor.execute('''
            UPDATE laporan_kemajuan_awal lk
            JOIN proposal p ON lk.id_proposal = p.id
            SET lk.status = 'ditolak'
            WHERE p.nim = %s
              AND lk.status IN ('draf','diajukan','revisi')
              AND lk.status != 'disetujui'
        ''', (session['nim'],))

        # Update laporan kemajuan (bertumbuh) jika tabel ada
        cursor.execute("SHOW TABLES LIKE 'laporan_kemajuan_bertumbuh'")
        if cursor.fetchone():
            cursor.execute('''
                UPDATE laporan_kemajuan_bertumbuh lk
                JOIN proposal p ON lk.id_proposal = p.id
                SET lk.status = 'ditolak'
                WHERE p.nim = %s
                  AND lk.status IN ('draf','diajukan','revisi')
                  AND lk.status != 'disetujui'
            ''', (session['nim'],))

        if hasattr(app_funcs["mysql"], 'connection') and app_funcs["mysql"].connection is not None:
            app_funcs["mysql"].connection.commit()
        cursor.close()
    except Exception:
        try:
            cursor.close()
        except Exception:
            pass


# Helper: Mengembalikan status yang terlanjur 'ditolak' menjadi 'diajukan' saat jadwal BELUM selesai
def restore_prematurely_rejected_kemajuan_for_current_student():
    app_funcs = get_app_functions()
    if not hasattr(app_funcs["mysql"], 'connection') or app_funcs["mysql"].connection is None:
        return
    try:
        cursor = app_funcs["mysql"].connection.cursor(MySQLdb.cursors.DictCursor)
        # Guard: hanya jalan bila NOW() <= kemajuan_selesai
        cursor.execute('SELECT kemajuan_selesai FROM pengaturan_jadwal ORDER BY id DESC LIMIT 1')
        row = cursor.fetchone()
        if not row or not row.get('kemajuan_selesai'):
            cursor.close()
            return
        kemajuan_selesai = row['kemajuan_selesai']
        try:
            cursor.execute('SELECT NOW() AS server_now')
            now_row = cursor.fetchone() or {}
            now = now_row.get('server_now') or datetime.now()
        except Exception:
            now = datetime.now()

        # Tambah toleransi 2 menit untuk menghindari drift
        if now > (kemajuan_selesai + timedelta(minutes=2)):
            cursor.close()
            return

        # Pulihkan yang 'ditolak' menjadi 'diajukan' hanya jika ditolak sistem karena jadwal, bukan oleh pembimbing/admin.
        # Kita batasi pulihkan hanya baris yang TIDAK memiliki tanggal_review (indikasi belum di-review)
        cursor.execute('''
            UPDATE laporan_kemajuan_awal lk
            JOIN proposal p ON lk.id_proposal = p.id
            SET lk.status = 'diajukan'
            WHERE p.nim = %s 
              AND lk.status = 'ditolak'
              AND (lk.tanggal_review IS NULL OR lk.tanggal_review = '0000-00-00 00:00:00')
        ''', (session['nim'],))

        cursor.execute("SHOW TABLES LIKE 'laporan_kemajuan_bertumbuh'")
        if cursor.fetchone():
            cursor.execute('''
                UPDATE laporan_kemajuan_bertumbuh lk
                JOIN proposal p ON lk.id_proposal = p.id
                SET lk.status = 'diajukan'
                WHERE p.nim = %s 
                  AND lk.status = 'ditolak'
                  AND (lk.tanggal_review IS NULL OR lk.tanggal_review = '0000-00-00 00:00:00')
            ''', (session['nim'],))

        if hasattr(app_funcs["mysql"], 'connection') and app_funcs["mysql"].connection is not None:
            app_funcs["mysql"].connection.commit()
        cursor.close()
    except Exception:
        try:
            cursor.close()
        except Exception:
            pass


# Best-effort: Pastikan event scheduler untuk auto-reject kemajuan ada (MariaDB/MySQL)
def ensure_kemajuan_auto_reject_events():
    app_funcs = get_app_functions()
    if not hasattr(app_funcs["mysql"], 'connection') or app_funcs["mysql"].connection is None:
        return
    cursor = None
    try:
        cursor = app_funcs["mysql"].connection.cursor()
        # Aktifkan event scheduler (nyala di sesi ini; untuk global permanen set di server config)
        try:
            cursor.execute('SET SESSION event_scheduler = ON')
        except Exception:
            pass

        # Buat event untuk laporan_kemajuan_awal
        cursor.execute('''
CREATE EVENT IF NOT EXISTS auto_reject_kemajuan_awal
ON SCHEDULE EVERY 10 MINUTE
DO
  UPDATE laporan_kemajuan_awal lk
  JOIN (SELECT kemajuan_selesai FROM pengaturan_jadwal ORDER BY id DESC LIMIT 1) j
    ON 1=1
  SET lk.status = 'ditolak'
  WHERE lk.status IN ('draf','diajukan','revisi')
    AND lk.status != 'disetujui'
    AND j.kemajuan_selesai IS NOT NULL
    AND NOW() > j.kemajuan_selesai
    AND (lk.tanggal_review IS NULL OR lk.tanggal_review = '0000-00-00 00:00:00');
        ''')

        # Buat event untuk laporan_kemajuan_bertumbuh jika tabel ada
        try:
            cursor.execute("SHOW TABLES LIKE 'laporan_kemajuan_bertumbuh'")
            if cursor.fetchone():
                cursor.execute('''
CREATE EVENT IF NOT EXISTS auto_reject_kemajuan_bertumbuh
ON SCHEDULE EVERY 10 MINUTE
DO
  UPDATE laporan_kemajuan_bertumbuh lk
  JOIN (SELECT kemajuan_selesai FROM pengaturan_jadwal ORDER BY id DESC LIMIT 1) j
    ON 1=1
  SET lk.status = 'ditolak'
  WHERE lk.status IN ('draf','diajukan','revisi')
    AND lk.status != 'disetujui'
    AND j.kemajuan_selesai IS NOT NULL
    AND NOW() > j.kemajuan_selesai
    AND (lk.tanggal_review IS NULL OR lk.tanggal_review = '0000-00-00 00:00:00');
                ''')
        except Exception:
            pass

        try:
            app_funcs["mysql"].connection.commit()
        except Exception:
            pass
    except Exception:
        if cursor:
            try:
                cursor.close()
            except Exception:
                pass
        raise
    finally:
        if cursor:
            try:
                cursor.close()
            except Exception:
                pass


# Helper: Mengembalikan status yang terlanjur 'ditolak' menjadi 'diajukan' saat jadwal BELUM selesai
def restore_prematurely_rejected_akhir_for_current_student():
    app_funcs = get_app_functions()
    if not hasattr(app_funcs["mysql"], 'connection') or app_funcs["mysql"].connection is None:
        return
    try:
        cursor = app_funcs["mysql"].connection.cursor(MySQLdb.cursors.DictCursor)
        # Guard: hanya jalan bila NOW() <= akhir_selesai
        cursor.execute('SELECT akhir_selesai FROM pengaturan_jadwal ORDER BY id DESC LIMIT 1')
        row = cursor.fetchone()
        if not row or not row.get('akhir_selesai'):
            cursor.close()
            return
        akhir_selesai = row['akhir_selesai']
        try:
            cursor.execute('SELECT NOW() AS server_now')
            now_row = cursor.fetchone() or {}
            now = now_row.get('server_now') or datetime.now()
        except Exception:
            now = datetime.now()

        # Tambah toleransi 2 menit untuk menghindari drift
        if now > (akhir_selesai + timedelta(minutes=2)):
            cursor.close()
            return

        # Pulihkan yang 'ditolak' menjadi 'diajukan' hanya jika ditolak sistem karena jadwal, bukan oleh pembimbing/admin.
        # Kita batasi pulihkan hanya baris yang TIDAK memiliki tanggal_review (indikasi belum di-review)
        cursor.execute('''
            UPDATE laporan_akhir_awal la
            JOIN proposal p ON la.id_proposal = p.id
            SET la.status = 'diajukan'
            WHERE p.nim = %s 
              AND la.status = 'ditolak'
              AND (la.tanggal_review IS NULL OR la.tanggal_review = '0000-00-00 00:00:00')
        ''')

        cursor.execute("SHOW TABLES LIKE 'laporan_akhir_bertumbuh'")
        if cursor.fetchone():
            cursor.execute('''
                UPDATE laporan_akhir_bertumbuh la
                JOIN proposal p ON la.id_proposal = p.id
                SET la.status = 'diajukan'
                WHERE p.nim = %s 
                  AND la.status = 'ditolak'
                  AND (la.tanggal_review IS NULL OR la.tanggal_review = '0000-00-00 00:00:00')
            ''')

        if hasattr(app_funcs["mysql"], 'connection') and app_funcs["mysql"].connection is not None:
            app_funcs["mysql"].connection.commit()
        cursor.close()
    except Exception:
        try:
            cursor.close()
        except Exception:
            pass


# Best-effort: Pastikan event scheduler untuk auto-reject laporan akhir ada (MariaDB/MySQL)
def ensure_akhir_auto_reject_events():
    app_funcs = get_app_functions()
    if not hasattr(app_funcs["mysql"], 'connection') or app_funcs["mysql"].connection is None:
        return
    cursor = None
    try:
        cursor = app_funcs["mysql"].connection.cursor()
        # Aktifkan event scheduler (nyala di sesi ini; untuk global permanen set di server config)
        try:
            cursor.execute('SET SESSION event_scheduler = ON')
        except Exception:
            pass

        # Buat event untuk laporan_akhir_awal
        cursor.execute('''
CREATE EVENT IF NOT EXISTS auto_reject_akhir_awal
ON SCHEDULE EVERY 10 MINUTE
DO
  UPDATE laporan_akhir_awal la
  JOIN (SELECT akhir_selesai FROM pengaturan_jadwal ORDER BY id DESC LIMIT 1) j
    ON 1=1
  SET la.status = 'ditolak'
  WHERE la.status IN ('draf','diajukan','revisi')
    AND la.status != 'disetujui'
    AND j.akhir_selesai IS NOT NULL
    AND NOW() > j.akhir_selesai
    AND (la.tanggal_review IS NULL OR la.tanggal_review = '0000-00-00 00:00:00');
        ''')

        # Buat event untuk laporan_akhir_bertumbuh jika tabel ada
        try:
            cursor.execute("SHOW TABLES LIKE 'laporan_akhir_bertumbuh'")
            if cursor.fetchone():
                cursor.execute('''
CREATE EVENT IF NOT EXISTS auto_reject_akhir_bertumbuh
ON SCHEDULE EVERY 10 MINUTE
DO
  UPDATE laporan_akhir_bertumbuh la
  JOIN (SELECT akhir_selesai FROM pengaturan_jadwal ORDER BY id DESC LIMIT 1) j
    ON 1=1
  SET la.status = 'ditolak'
  WHERE la.status IN ('draf','diajukan','revisi')
    AND la.status != 'disetujui'
    AND j.akhir_selesai IS NOT NULL
    AND NOW() > j.akhir_selesai
    AND (la.tanggal_review IS NULL OR la.tanggal_review = '0000-00-00 00:00:00');
                ''')
        except Exception:
            pass

        try:
            app_funcs["mysql"].connection.commit()
        except Exception:
            pass
    except Exception:
        if cursor:
            try:
                cursor.close()
            except Exception:
                pass
        raise
    finally:
        if cursor:
            try:
                cursor.close()
            except Exception:
                pass

# Helper: Otomatis menolak laporan akhir (awal/bertumbuh) yang masih "draf/diajukan/revisi" jika sudah melewati jadwal akhir selesai
def auto_reject_expired_akhir_for_current_student():
    app_funcs = get_app_functions()
    if not hasattr(app_funcs["mysql"], 'connection') or app_funcs["mysql"].connection is None:
        return
    try:
        cursor = app_funcs["mysql"].connection.cursor(MySQLdb.cursors.DictCursor)
        # Ambil jadwal akhir selesai terbaru
        cursor.execute('SELECT akhir_selesai FROM pengaturan_jadwal ORDER BY id DESC LIMIT 1')
        row = cursor.fetchone()
        if not row or not row.get('akhir_selesai'):
            cursor.close()
            return
        batas_selesai = row['akhir_selesai']
        
        # Gunakan waktu server database untuk membandingkan jadwal
        try:
            cursor.execute('SELECT NOW() AS server_now')
            _now_row = cursor.fetchone() or {}
            now = _now_row.get('server_now') or datetime.now()
        except Exception:
            now = datetime.now()
            
        if not batas_selesai or now <= batas_selesai:
            cursor.close()
            return

        # Update laporan akhir (awal) milik mahasiswa ini menjadi 'ditolak' jika masih draf/diajukan/revisi
        # KECUALI status sudah 'disetujui'
        cursor.execute('''
            UPDATE laporan_akhir_awal la
            JOIN proposal p ON la.id_proposal = p.id
            SET la.status = 'ditolak'
            WHERE p.nim = %s 
              AND la.status IN ('draf','diajukan','revisi')
              AND la.status != 'disetujui'
        ''', (session['nim'],))

        # Update laporan akhir (bertumbuh) jika tabel ada
        cursor.execute("SHOW TABLES LIKE 'laporan_akhir_bertumbuh'")
        if cursor.fetchone():
            cursor.execute('''
                UPDATE laporan_akhir_bertumbuh la
                JOIN proposal p ON la.id_proposal = p.id
                SET la.status = 'ditolak'
                WHERE p.nim = %s 
                  AND la.status IN ('draf','diajukan','revisi')
                  AND la.status != 'disetujui'
            ''', (session['nim'],))

        if hasattr(app_funcs["mysql"], 'connection') and app_funcs["mysql"].connection is not None:
            app_funcs["mysql"].connection.commit()
        cursor.close()
    except Exception:
        try:
            cursor.close()
        except Exception:
            pass

@mahasiswa_bp.route('/proposal')
def proposal():
    app_funcs = get_app_functions()
    if 'user_type' not in session or session['user_type'] != 'mahasiswa':
        flash('Anda harus login sebagai mahasiswa!', 'danger')
        return redirect(url_for('index'))
    
    if not hasattr(app_funcs["mysql"], 'connection') or app_funcs["mysql"].connection is None:
        flash('Koneksi ke database gagal. Cek konfigurasi database!', 'danger')
        return render_template('mahasiswa/proposal.html', proposals=[], status_mahasiswa='proses', daftar_dosen=[])
    try:
        # Otomasi penolakan setelah jadwal selesai untuk akun ini (guard ringan saat membuka halaman Proposal)
        auto_reject_expired_proposals_for_current_student()
        
        # Pastikan event scheduler untuk auto-reject ada
        ensure_kemajuan_auto_reject_events()
        ensure_akhir_auto_reject_events()

        cursor = app_funcs["mysql"].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Ambil status mahasiswa
        cursor.execute('SELECT status FROM mahasiswa WHERE nim = %s', (session['nim'],))
        mahasiswa_status = cursor.fetchone()
        
        if not mahasiswa_status:
            flash('Data mahasiswa tidak ditemukan!', 'danger')
            cursor.close()
            return redirect(url_for('index'))
        
        status_mahasiswa = mahasiswa_status['status']
        
        # Ambil data proposal mahasiswa
        cursor.execute('''
            SELECT p.*, 
                   COUNT(at.id) as jumlah_anggota
            FROM proposal p 
            LEFT JOIN anggota_tim at ON p.id = at.id_proposal
            WHERE p.nim = %s 
            GROUP BY p.id 
            ORDER BY p.tanggal_buat DESC
        ''', (session['nim'],))
        
        proposals = cursor.fetchall()
        
        # Ambil daftar dosen pembimbing dengan informasi kuota
        cursor.execute('''
            SELECT p.nama, p.nip, p.program_studi, p.kuota_mahasiswa,
                   COUNT(DISTINCT CASE WHEN pr.status != 'selesai' THEN pr.nim END) as jumlah_mahasiswa_bimbing
            FROM pembimbing p
            LEFT JOIN proposal pr ON p.nama = pr.dosen_pembimbing
            WHERE p.status = 'aktif'
            GROUP BY p.id, p.nama, p.nip, p.program_studi, p.kuota_mahasiswa
            ORDER BY p.nama
        ''')
        
        daftar_dosen = cursor.fetchall()
        
        # Tambahkan informasi ketersediaan kuota
        for dosen in daftar_dosen:
            dosen['sisa_kuota'] = dosen['kuota_mahasiswa'] - dosen['jumlah_mahasiswa_bimbing']
            dosen['status_kuota'] = 'tersedia' if dosen['sisa_kuota'] > 0 else 'penuh'
        
        # Ambil jumlah anggota untuk setiap proposal
        jumlah_anggota_per_proposal = {}
        for proposal in proposals:
            cursor.execute('SELECT COUNT(*) as jumlah FROM anggota_tim WHERE id_proposal = %s', (proposal['id'],))
            jumlah = cursor.fetchone()['jumlah']
            jumlah_anggota_per_proposal[proposal['id']] = jumlah
        
        # Ambil daftar program studi untuk dropdown
        cursor.execute('SELECT nama_program_studi FROM program_studi ORDER BY nama_program_studi')
        program_studi_list = cursor.fetchall()
        
        cursor.close()
        
        return render_template('mahasiswa/proposal.html', 
                             proposals=proposals, 
                             status_mahasiswa=status_mahasiswa,
                             daftar_dosen=daftar_dosen,
                             jumlah_anggota_per_proposal=jumlah_anggota_per_proposal,
                             program_studi_list=program_studi_list)
        
    except Exception as e:
        flash(f'Error saat mengambil data proposal: {str(e)}', 'danger')
        return render_template('mahasiswa/proposal.html', 
                             proposals=[], 
                             status_mahasiswa='proses',
                             daftar_dosen=[],
                             program_studi_list=[])

@mahasiswa_bp.route('/pengelolaan_bahan_baku_mahasiswa')
def pengelolaan_bahan_baku_mahasiswa():
    app_funcs = get_app_functions()
    if 'user_type' not in session or session['user_type'] != 'mahasiswa':
        flash('Anda harus login sebagai mahasiswa!', 'danger')
        return redirect(url_for('index'))
    
    if not hasattr(app_funcs["mysql"], 'connection') or app_funcs["mysql"].connection is None:
        flash('Koneksi ke database gagal. Cek konfigurasi database!', 'danger')
        return render_template('mahasiswa/bahan_baku.html', proposal_data={})
    
    try:
        cursor = app_funcs["mysql"].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Ambil data mahasiswa yang login
        cursor.execute('SELECT * FROM mahasiswa WHERE nim = %s', (session['nim'],))
        mahasiswa = cursor.fetchone()
        
        if not mahasiswa:
            flash('Data mahasiswa tidak ditemukan', 'error')
            return redirect(url_for('index'))
        
        # Ambil proposal mahasiswa yang sudah lolos
        cursor.execute('''
            SELECT p.*, bb.id as bahan_baku_id, bb.nama_bahan, bb.quantity, bb.satuan, 
                   bb.harga_satuan, bb.total_harga, bb.tanggal_beli
            FROM proposal p
            LEFT JOIN bahan_baku bb ON p.id = bb.proposal_id
            WHERE p.nim = %s AND p.status_admin = 'lolos'
            ORDER BY p.tanggal_buat DESC, bb.tanggal_beli DESC
        ''', (session['nim'],))
        
        proposals = cursor.fetchall()
        
        # Group data bahan baku berdasarkan proposal
        proposal_data = {}
        for row in proposals:
            proposal_id = row['id']
            if proposal_id not in proposal_data:
                proposal_data[proposal_id] = {
                    'proposal': {
                        'id': row['id'],
                        'judul_usaha': row['judul_usaha'],
                        'kategori': row['kategori'],
                        'status': row['status'],
                        'tahun_nib': row['tahun_nib']
                    },
                    'bahan_baku': []
                }
            
            if row['bahan_baku_id']:
                proposal_data[proposal_id]['bahan_baku'].append({
                    'id': row['bahan_baku_id'],
                    'nama_bahan': row['nama_bahan'],
                    'quantity': row['quantity'],
                    'satuan': row['satuan'],
                    'harga_satuan': row['harga_satuan'],
                    'total_harga': row['total_harga'],
                    'tanggal_beli': row['tanggal_beli']
                })
        
        # Ambil status mahasiswa
        status_mahasiswa = mahasiswa['status']
        
        cursor.close()
        return render_template('mahasiswa/bahan_baku.html', proposal_data=proposal_data, status_mahasiswa=status_mahasiswa)
        
    except Exception as e:
        flash(f'Error saat mengambil data: {str(e)}', 'danger')
        return render_template('mahasiswa/bahan_baku.html', proposal_data={}, status_mahasiswa='proses')

# Route untuk CRUD Bahan Baku
@mahasiswa_bp.route('/tambah_bahan_baku', methods=['POST'])
def tambah_bahan_baku():
    app_funcs = get_app_functions()
    if 'user_type' not in session or session['user_type'] != 'mahasiswa':
        return jsonify({'success': False, 'message': 'Anda harus login sebagai mahasiswa'})
    
    if not hasattr(app_funcs["mysql"], 'connection') or app_funcs["mysql"].connection is None:
        return jsonify({'success': False, 'message': 'Koneksi ke database gagal'})
    
    try:
        cursor = app_funcs["mysql"].connection.cursor(MySQLdb.cursors.DictCursor)
        
        proposal_id = request.form.get('proposal_id')
        nama_bahan = request.form.get('nama_bahan')
        quantity = request.form.get('quantity')
        satuan = request.form.get('satuan')
        harga_satuan = request.form.get('harga_satuan')
        tanggal_beli = request.form.get('tanggal_beli')
        
        # Validasi input
        if not all([proposal_id, nama_bahan, quantity, satuan, harga_satuan, tanggal_beli]):
            return jsonify({'success': False, 'message': 'Semua field harus diisi'})
        
        # Validasi dan konversi nilai numerik
        try:
            quantity_float = float(quantity) if quantity else 0
            harga_satuan_float = float(harga_satuan) if harga_satuan else 0
        except (ValueError, TypeError):
            return jsonify({'success': False, 'message': 'Nilai quantity dan harga satuan harus berupa angka'})
        
        # Hitung total harga
        total_harga = quantity_float * harga_satuan_float
        
        # Normalisasi nama_bahan dan satuan
        nama_bahan = (nama_bahan or '').strip()
        satuan = (satuan or '').strip()
        nama_bahan = ' '.join([w.capitalize() for w in nama_bahan.split()])
        
        # Normalisasi satuan - ubah ke lowercase dulu, lalu capitalize huruf pertama
        satuan = satuan.lower()
        if satuan:
            satuan = satuan[0].upper() + satuan[1:] if len(satuan) > 1 else satuan.upper()
        
        # Insert ke database
        cursor.execute('''
            INSERT INTO bahan_baku (proposal_id, nama_bahan, quantity, satuan, harga_satuan, total_harga, tanggal_beli)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
        ''', (proposal_id, nama_bahan, quantity_float, satuan, harga_satuan_float, total_harga, tanggal_beli))
        
        bahan_baku_id = cursor.lastrowid
        app_funcs["mysql"].connection.commit()
        
        # Log aktivitas
        mahasiswa_info = app_funcs['get_mahasiswa_info_from_session']()
        if mahasiswa_info:
            data_baru = {
                'id': bahan_baku_id,
                'nama_bahan': nama_bahan,
                'quantity': quantity_float,
                'satuan': satuan,
                'harga_satuan': harga_satuan_float,
                'total_harga': total_harga,
                'tanggal_beli': tanggal_beli
            }
            app_funcs['log_mahasiswa_activity'](
                mahasiswa_info['id'], 
                mahasiswa_info['nim'], 
                mahasiswa_info['nama_ketua'], 
                'tambah', 
                'bahan_baku', 
                f'id_{bahan_baku_id}', 
                f'Menambahkan bahan baku: {nama_bahan}',
                None,
                data_baru
            )
        
        cursor.close()
        
        return jsonify({'success': True, 'message': 'Bahan baku berhasil ditambahkan'})
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@mahasiswa_bp.route('/get_bahan_baku/<int:bahan_baku_id>')
def get_bahan_baku(bahan_baku_id):
    app_funcs = get_app_functions()
    if 'user_type' not in session or session['user_type'] != 'mahasiswa':
        return jsonify({'success': False, 'message': 'Anda harus login sebagai mahasiswa'})
    
    if not hasattr(app_funcs["mysql"], 'connection') or app_funcs["mysql"].connection is None:
        return jsonify({'success': False, 'message': 'Koneksi ke database gagal'})
    
    try:
        cursor = app_funcs["mysql"].connection.cursor(MySQLdb.cursors.DictCursor)
        
        cursor.execute('SELECT * FROM bahan_baku WHERE id = %s', (bahan_baku_id,))
        bahan_baku = cursor.fetchone()
        
        cursor.close()
        
        if bahan_baku:
            return jsonify({'success': True, 'data': bahan_baku})
        else:
            return jsonify({'success': False, 'message': 'Data bahan baku tidak ditemukan'})
            
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@mahasiswa_bp.route('/update_bahan_baku', methods=['POST'])
def update_bahan_baku():
    app_funcs = get_app_functions()
    if 'user_type' not in session or session['user_type'] != 'mahasiswa':
        return jsonify({'success': False, 'message': 'Anda harus login sebagai mahasiswa'})
    
    if not hasattr(app_funcs["mysql"], 'connection') or app_funcs["mysql"].connection is None:
        return jsonify({'success': False, 'message': 'Koneksi ke database gagal'})
    
    try:
        cursor = app_funcs["mysql"].connection.cursor(MySQLdb.cursors.DictCursor)
        
        bahan_baku_id = request.form.get('bahan_baku_id')
        nama_bahan = request.form.get('nama_bahan')
        quantity = request.form.get('quantity')
        satuan = request.form.get('satuan')
        harga_satuan = request.form.get('harga_satuan')
        tanggal_beli = request.form.get('tanggal_beli')
        
        # Validasi input
        if not all([bahan_baku_id, nama_bahan, quantity, satuan, harga_satuan, tanggal_beli]):
            return jsonify({'success': False, 'message': 'Semua field harus diisi'})
        
        # Validasi dan konversi nilai numerik
        try:
            quantity_float = float(quantity) if quantity else 0
            harga_satuan_float = float(harga_satuan) if harga_satuan else 0
        except (ValueError, TypeError):
            return jsonify({'success': False, 'message': 'Nilai quantity dan harga satuan harus berupa angka'})
        
        # Hitung total harga
        total_harga = quantity_float * harga_satuan_float
        
        # Normalisasi nama_bahan dan satuan
        nama_bahan = (nama_bahan or '').strip()
        satuan = (satuan or '').strip()
        nama_bahan = ' '.join([w.capitalize() for w in nama_bahan.split()])
        
        # Normalisasi satuan - ubah ke lowercase dulu, lalu capitalize huruf pertama
        satuan = satuan.lower()
        if satuan:
            satuan = satuan[0].upper() + satuan[1:] if len(satuan) > 1 else satuan.upper()
        
        # Ambil data lama untuk logging
        cursor.execute('SELECT * FROM bahan_baku WHERE id = %s', (bahan_baku_id,))
        data_lama_record = cursor.fetchone()
        data_lama = dict(data_lama_record) if data_lama_record else None
        
        # Update database
        cursor.execute('''
            UPDATE bahan_baku 
            SET nama_bahan = %s, quantity = %s, satuan = %s, harga_satuan = %s, total_harga = %s, tanggal_beli = %s
            WHERE id = %s
        ''', (nama_bahan, quantity_float, satuan, harga_satuan_float, total_harga, tanggal_beli, bahan_baku_id))
        
        app_funcs["mysql"].connection.commit()
        
        # Log aktivitas
        mahasiswa_info = app_funcs['get_mahasiswa_info_from_session']()
        if mahasiswa_info and bahan_baku_id:
            data_baru = {
                'id': int(bahan_baku_id),
                'nama_bahan': nama_bahan,
                'quantity': quantity_float,
                'satuan': satuan,
                'harga_satuan': harga_satuan_float,
                'total_harga': total_harga,
                'tanggal_beli': tanggal_beli
            }
            app_funcs['log_mahasiswa_activity'](
                mahasiswa_info['id'], 
                mahasiswa_info['nim'], 
                mahasiswa_info['nama_ketua'], 
                'edit', 
                'bahan_baku', 
                f'id_{bahan_baku_id}', 
                f'Mengubah bahan baku: {nama_bahan}',
                data_lama,
                data_baru
            )
        
        # Update data terkait otomatis
        try:
            app_funcs['update_related_data_on_bahan_baku_change'](bahan_baku_id)
            
            # Update penjualan terkait
            try:
                app_funcs['update_penjualan_on_bahan_baku_change'](bahan_baku_id)
            except Exception as e:
                print(f"Warning: Error updating penjualan: {str(e)}")
                
        except Exception as e:
            print(f"Warning: Error updating related data: {str(e)}")
        
        cursor.close()
        
        return jsonify({'success': True, 'message': 'Bahan baku berhasil diupdate dan data terkait telah diperbarui otomatis'})
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})
@mahasiswa_bp.route('/delete_bahan_baku', methods=['POST'])
def delete_bahan_baku():
    app_funcs = get_app_functions()
    if 'user_type' not in session or session['user_type'] != 'mahasiswa':
        return jsonify({'success': False, 'message': 'Anda harus login sebagai mahasiswa'})
    
    if not hasattr(app_funcs["mysql"], 'connection') or app_funcs["mysql"].connection is None:
        return jsonify({'success': False, 'message': 'Koneksi ke database gagal'})
    
    try:
        cursor = app_funcs["mysql"].connection.cursor(MySQLdb.cursors.DictCursor)
        
        bahan_baku_id = request.form.get('bahan_baku_id')
        
        if not bahan_baku_id:
            return jsonify({'success': False, 'message': 'ID bahan baku tidak valid'})
        
        # Ambil data bahan baku untuk logging
        cursor.execute('SELECT * FROM bahan_baku WHERE id = %s', (bahan_baku_id,))
        bahan_baku_data = cursor.fetchone()
        data_lama = dict(bahan_baku_data) if bahan_baku_data else None
        
        # Cek apakah bahan baku digunakan dalam produksi
        cursor.execute('''
            SELECT DISTINCT p.proposal_id 
            FROM produk_bahan_baku pbb
            JOIN produksi p ON pbb.produksi_id = p.id
            WHERE pbb.bahan_baku_id = %s
        ''', (bahan_baku_id,))
        
        affected_proposals = cursor.fetchall()
        
        # Delete dari database
        cursor.execute('DELETE FROM bahan_baku WHERE id = %s', (bahan_baku_id,))
        
        app_funcs["mysql"].connection.commit()
        
        # Log aktivitas
        mahasiswa_info = app_funcs['get_mahasiswa_info_from_session']()
        if mahasiswa_info and data_lama:
            app_funcs['log_mahasiswa_activity'](
                mahasiswa_info['id'], 
                mahasiswa_info['nim'], 
                mahasiswa_info['nama_ketua'], 
                'hapus', 
                'bahan_baku', 
                f'id_{bahan_baku_id}', 
                f'Menghapus bahan baku: {data_lama.get("nama_bahan", "Unknown")}',
                data_lama,
                None
            )
        
        # Update laba rugi dan penjualan untuk proposal yang terpengaruh
        for proposal in affected_proposals:
            try:
                app_funcs['update_laba_rugi_on_produksi_change'](proposal['proposal_id'])
            except Exception as e:
                print(f"Warning: Error updating laba rugi for proposal {proposal['proposal_id']}: {str(e)}")
        
        cursor.close()
        
        return jsonify({'success': True, 'message': 'Bahan baku berhasil dihapus dan data terkait telah diperbarui otomatis'})
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})
# Route untuk CRUD Produksi
@mahasiswa_bp.route('/get_bahan_baku_for_produksi/<int:proposal_id>')
def get_bahan_baku_for_produksi(proposal_id):
    app_funcs = get_app_functions()
    if 'user_type' not in session or session['user_type'] != 'mahasiswa':
        return jsonify({'success': False, 'message': 'Anda harus login sebagai mahasiswa'})
    
    if not hasattr(app_funcs["mysql"], 'connection') or app_funcs["mysql"].connection is None:
        return jsonify({'success': False, 'message': 'Koneksi ke database gagal'})
    
    try:
        cursor = app_funcs["mysql"].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Ambil bahan baku yang tersedia untuk proposal ini
        cursor.execute('''
            SELECT id, nama_bahan, quantity, satuan, harga_satuan, total_harga
            FROM bahan_baku 
            WHERE proposal_id = %s AND quantity > 0
            ORDER BY nama_bahan
        ''', (proposal_id,))
        
        bahan_baku_list = cursor.fetchall()
        
        # Konversi nilai desimal ke integer untuk frontend
        for bahan in bahan_baku_list:
            bahan['harga_satuan'] = int(bahan['harga_satuan'])
            bahan['total_harga'] = int(bahan['total_harga'])
            bahan['quantity'] = int(bahan['quantity'])
        
        cursor.close()
        
        return jsonify({'success': True, 'data': bahan_baku_list})
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@mahasiswa_bp.route('/tambah_produksi', methods=['POST'])
def tambah_produksi():
    app_funcs = get_app_functions()
    from decimal import Decimal, InvalidOperation
    import json
    if 'user_type' not in session or session['user_type'] != 'mahasiswa':
        return jsonify({'success': False, 'message': 'Anda harus login sebagai mahasiswa'})
    
    if not hasattr(app_funcs["mysql"], 'connection') or app_funcs["mysql"].connection is None:
        return jsonify({'success': False, 'message': 'Koneksi ke database gagal'})
    
    try:
        cursor = app_funcs["mysql"].connection.cursor(MySQLdb.cursors.DictCursor)
        
        proposal_id = request.form.get('proposal_id')
        nama_produk = request.form.get('nama_produk')
        jumlah_produk = request.form.get('jumlah_produk')
        persentase_laba = request.form.get('persentase_laba')
        tanggal_produksi = request.form.get('tanggal_produksi')
        bahan_baku_data = request.form.get('bahan_baku_data')  # JSON string
        
        # Validasi input
        if not all([proposal_id, nama_produk, jumlah_produk, persentase_laba, tanggal_produksi, bahan_baku_data]):
            return jsonify({'success': False, 'message': 'Semua field harus diisi'})
        
        # Validasi proposal_id adalah angka
        if not proposal_id:
            return jsonify({'success': False, 'message': 'ID proposal tidak boleh kosong'})
        try:
            proposal_id = int(proposal_id)
        except (ValueError, TypeError):
            return jsonify({'success': False, 'message': 'ID proposal tidak valid'})
        
        # Cek apakah proposal ada dan milik mahasiswa yang login
        cursor.execute('SELECT id FROM proposal WHERE id = %s AND nim = %s', (proposal_id, session['nim']))
        proposal_exists = cursor.fetchone()
        if not proposal_exists:
            return jsonify({'success': False, 'message': 'Proposal tidak ditemukan atau bukan milik Anda'})
        
        # Validasi tanggal produksi
        if not tanggal_produksi:
            return jsonify({'success': False, 'message': 'Tanggal produksi tidak boleh kosong'})
        try:
            from datetime import datetime
            datetime.strptime(tanggal_produksi, '%Y-%m-%d')
        except ValueError:
            return jsonify({'success': False, 'message': 'Format tanggal produksi tidak valid'})
        
        # Parse bahan baku data
        if bahan_baku_data is None:
            return jsonify({'success': False, 'message': 'Data bahan baku tidak boleh kosong'})
        try:
            bahan_baku_list = json.loads(bahan_baku_data)
        except Exception:
            return jsonify({'success': False, 'message': 'Format data bahan baku tidak valid'})
        
        if not bahan_baku_list:
            return jsonify({'success': False, 'message': 'Pilih minimal satu bahan baku'})
        
        # Validasi tidak ada duplikasi bahan baku
        bahan_ids = [bahan['bahan_baku_id'] for bahan in bahan_baku_list]
        if len(bahan_ids) != len(set(bahan_ids)):
            return jsonify({'success': False, 'message': 'Tidak boleh ada bahan baku yang duplikat'})
        
        # Validasi dan konversi nilai numerik
        try:
            if not jumlah_produk or not persentase_laba:
                return jsonify({'success': False, 'message': 'Jumlah produk dan persentase laba tidak boleh kosong'})
            
            if not jumlah_produk or jumlah_produk.strip() == '':
                return jsonify({'success': False, 'message': 'Jumlah produk tidak boleh kosong'})
                
            if not persentase_laba or persentase_laba.strip() == '':
                return jsonify({'success': False, 'message': 'Persentase laba tidak boleh kosong'})
                
            jumlah_produk_int = int(jumlah_produk)
            persentase_laba_dec = Decimal(persentase_laba.replace('.', '').replace(',', '.'))
        except (ValueError, TypeError, InvalidOperation):
            return jsonify({'success': False, 'message': 'Nilai jumlah produk dan persentase laba harus berupa angka'})
        if jumlah_produk_int <= 0 or persentase_laba_dec < 0:
            return jsonify({'success': False, 'message': 'Jumlah produk dan persentase laba harus lebih dari 0'})
        
        # Hitung total biaya (tanpa validasi stok)
        total_biaya = Decimal('0')
        for bahan in bahan_baku_list:
            try:
                bahan_id = bahan['bahan_baku_id']
                quantity_digunakan = Decimal(str(bahan['quantity_digunakan']))
            except (KeyError, InvalidOperation, ValueError, TypeError):
                return jsonify({'success': False, 'message': 'Data bahan baku tidak valid'})
            
            # Validasi quantity tidak boleh negatif
            if quantity_digunakan < 0:
                return jsonify({'success': False, 'message': f'Quantity tidak boleh negatif'})
            
            # Validasi quantity tidak boleh nol
            if quantity_digunakan == 0:
                return jsonify({'success': False, 'message': f'Quantity tidak boleh 0'})
            
            # Ambil harga satuan bahan baku
            cursor.execute('SELECT harga_satuan FROM bahan_baku WHERE id = %s', (bahan_id,))
            stok_data = cursor.fetchone()
            
            if not stok_data:
                return jsonify({'success': False, 'message': f'Bahan baku tidak ditemukan'})
            
            harga_satuan = Decimal(str(stok_data['harga_satuan']))
            
            subtotal = quantity_digunakan * harga_satuan
            total_biaya += subtotal
        
        # Hitung HPP Dasar = Total Biaya / Jumlah Produksi
        hpp_dasar = total_biaya / jumlah_produk_int if jumlah_produk_int > 0 else Decimal('0')
        
        # Hitung Harga Jual = HPP + (Persentase Laba × HPP)
        harga_jual_calculated = hpp_dasar + (persentase_laba_dec / 100 * hpp_dasar)
        
        # Validasi nama_produk
        if not nama_produk or nama_produk.strip() == '':
            return jsonify({'success': False, 'message': 'Nama produk tidak boleh kosong'})
        
        # Normalisasi nama_produk
        nama_produk = nama_produk.strip()
        nama_produk = ' '.join([w.capitalize() for w in nama_produk.split()])
        
        # Insert ke tabel produksi
        cursor.execute('''
            INSERT INTO produksi (proposal_id, nama_produk, jumlah_produk, harga_jual, persentase_laba, total_biaya, hpp, tanggal_produksi)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
        ''', (proposal_id, nama_produk, jumlah_produk_int, float(harga_jual_calculated), float(persentase_laba_dec), float(total_biaya), float(hpp_dasar), tanggal_produksi))
        
        produksi_id = cursor.lastrowid
        
        # Insert ke tabel produk_bahan_baku (tanpa mengurangi stok)
        for bahan in bahan_baku_list:
            bahan_id = bahan['bahan_baku_id']
            quantity_digunakan = Decimal(str(bahan['quantity_digunakan']))
            
            # Ambil data bahan baku untuk harga
            cursor.execute('SELECT harga_satuan FROM bahan_baku WHERE id = %s', (bahan_id,))
            stok_data = cursor.fetchone()
            harga_satuan = Decimal(str(stok_data['harga_satuan']))
            
            # Validasi quantity tidak boleh negatif
            if quantity_digunakan < 0:
                return jsonify({'success': False, 'message': f'Quantity bahan baku ID {bahan_id} tidak boleh negatif'})
            
            # Validasi quantity tidak boleh 0
            if quantity_digunakan == 0:
                return jsonify({'success': False, 'message': f'Quantity bahan baku ID {bahan_id} tidak boleh 0'})
            
            subtotal = quantity_digunakan * harga_satuan
            
            # Insert ke produk_bahan_baku
            cursor.execute('''
                INSERT INTO produk_bahan_baku (produksi_id, bahan_baku_id, quantity_digunakan, harga_satuan, subtotal)
                VALUES (%s, %s, %s, %s, %s)
            ''', (produksi_id, bahan_id, float(quantity_digunakan), float(harga_satuan), float(subtotal)))
            
            # Tidak mengurangi stok bahan baku saat produksi
            # Stok bahan baku tetap utuh untuk keperluan pencatatan saja
        
        app_funcs["mysql"].connection.commit()
        
        # Log aktivitas
        mahasiswa_info = app_funcs['get_mahasiswa_info_from_session']()
        if mahasiswa_info:
            data_baru = {
                'id': produksi_id,
                'nama_produk': nama_produk,
                'jumlah_produk': jumlah_produk_int,
                'harga_jual': float(harga_jual_calculated),
                'total_biaya': float(total_biaya),
                'hpp': float(hpp_dasar),
                'tanggal_produksi': tanggal_produksi,
                'persentase_laba': float(persentase_laba_dec)
            }
            app_funcs['log_mahasiswa_activity'](
                mahasiswa_info['id'], 
                mahasiswa_info['nim'], 
                mahasiswa_info['nama_ketua'], 
                'tambah', 
                'produksi', 
                f'id_{produksi_id}', 
                f'Menambahkan produksi: {nama_produk}',
                None,
                data_baru
            )
        
        # Hitung dan simpan data laba rugi otomatis
        try:
            app_funcs['update_laba_rugi_on_produksi_change'](proposal_id)
        except Exception as e:
            print(f"Warning: Error updating laba rugi: {str(e)}")
        
        # Update arus kas otomatis untuk bulan produksi
        if tanggal_produksi:
            bulan_produksi = datetime.strptime(tanggal_produksi, '%Y-%m-%d').strftime('%Y-%m')
            app_funcs['update_arus_kas_otomatis'](proposal_id, bulan_produksi)
        
        cursor.close()
        
        return jsonify({'success': True, 'message': 'Produksi berhasil ditambahkan dan laba rugi telah diperbarui otomatis'})
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@mahasiswa_bp.route('/get_produksi/<int:produksi_id>')
def get_produksi(produksi_id):
    app_funcs = get_app_functions()
    if 'user_type' not in session or session['user_type'] != 'mahasiswa':
        return jsonify({'success': False, 'message': 'Anda harus login sebagai mahasiswa'})
    
    if not hasattr(app_funcs["mysql"], 'connection') or app_funcs["mysql"].connection is None:
        return jsonify({'success': False, 'message': 'Koneksi ke database gagal'})
    
    try:
        cursor = app_funcs["mysql"].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Ambil data produksi dan validasi kepemilikan
        cursor.execute('''
            SELECT p.* FROM produksi p 
            JOIN proposal pr ON p.proposal_id = pr.id 
            WHERE p.id = %s AND pr.nim = %s
        ''', (produksi_id, session['nim']))
        produksi = cursor.fetchone()
        
        if not produksi:
            return jsonify({'success': False, 'message': 'Data produksi tidak ditemukan atau bukan milik Anda'})
        
        # Format tanggal_produksi ke YYYY-MM-DD jika ada
        if produksi.get('tanggal_produksi'):
            try:
                # Jika sudah string, parsing dulu
                if isinstance(produksi['tanggal_produksi'], str):
                    from datetime import datetime
                    dt = datetime.strptime(produksi['tanggal_produksi'], '%Y-%m-%d')
                    produksi['tanggal_produksi'] = dt.strftime('%Y-%m-%d')
                else:
                    produksi['tanggal_produksi'] = produksi['tanggal_produksi'].strftime('%Y-%m-%d')
            except Exception:
                pass
        
        # Ambil data bahan baku yang digunakan
        cursor.execute('''
            SELECT pbb.*, bb.nama_bahan, bb.satuan
            FROM produk_bahan_baku pbb
            JOIN bahan_baku bb ON pbb.bahan_baku_id = bb.id
            WHERE pbb.produksi_id = %s
        ''', (produksi_id,))
        
        bahan_baku_list = cursor.fetchall()
        
        # Konversi nilai desimal ke integer untuk frontend
        for bahan in bahan_baku_list:
            bahan['quantity_digunakan'] = int(bahan['quantity_digunakan'])
            bahan['harga_satuan'] = int(bahan['harga_satuan'])
            bahan['subtotal'] = int(bahan['subtotal'])
        
        # Konversi nilai produksi ke integer
        produksi['total_biaya'] = int(produksi['total_biaya'])
        produksi['hpp'] = int(produksi['hpp'])
        produksi['harga_jual'] = int(produksi['harga_jual'])
        if 'persentase_laba' in produksi and produksi['persentase_laba'] is not None:
            produksi['persentase_laba'] = int(produksi['persentase_laba'])
        else:
            produksi['persentase_laba'] = 0
        
        produksi['bahan_baku'] = bahan_baku_list
        
        cursor.close()
        
        return jsonify({'success': True, 'data': produksi})
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@mahasiswa_bp.route('/update_produksi', methods=['POST'])
def update_produksi():
    app_funcs = get_app_functions()
    from decimal import Decimal, InvalidOperation
    import json
    if 'user_type' not in session or session['user_type'] != 'mahasiswa':
        return jsonify({'success': False, 'message': 'Anda harus login sebagai mahasiswa'})
    
    if not hasattr(app_funcs["mysql"], 'connection') or app_funcs["mysql"].connection is None:
        return jsonify({'success': False, 'message': 'Koneksi ke database gagal'})
    
    try:
        cursor = app_funcs["mysql"].connection.cursor(MySQLdb.cursors.DictCursor)
        
        produksi_id = request.form.get('produksi_id')
        proposal_id = request.form.get('proposal_id')
        nama_produk = request.form.get('nama_produk')
        jumlah_produk = request.form.get('jumlah_produk')
        persentase_laba = request.form.get('persentase_laba')
        tanggal_produksi = request.form.get('tanggal_produksi')
        bahan_baku_data = request.form.get('bahan_baku_data')
        
        # Validasi input
        if not all([produksi_id, proposal_id, nama_produk, jumlah_produk, persentase_laba, tanggal_produksi, bahan_baku_data]):
            return jsonify({'success': False, 'message': 'Semua field harus diisi'})
        
        # Validasi produksi_id adalah angka
        if not produksi_id:
            return jsonify({'success': False, 'message': 'ID produksi tidak boleh kosong'})
        try:
            produksi_id = int(produksi_id)
        except (ValueError, TypeError):
            return jsonify({'success': False, 'message': 'ID produksi tidak valid'})
        
        # Validasi proposal_id adalah angka
        if not proposal_id:
            return jsonify({'success': False, 'message': 'ID proposal tidak boleh kosong'})
        try:
            proposal_id = int(proposal_id)
        except (ValueError, TypeError):
            return jsonify({'success': False, 'message': 'ID proposal tidak valid'})
        
        # Cek apakah produksi ada dan milik mahasiswa yang login
        cursor.execute('''
            SELECT p.* FROM produksi p 
            JOIN proposal pr ON p.proposal_id = pr.id 
            WHERE p.id = %s AND pr.nim = %s
        ''', (produksi_id, session['nim']))
        produksi_exists = cursor.fetchone()
        if not produksi_exists:
            return jsonify({'success': False, 'message': 'Produksi tidak ditemukan atau bukan milik Anda'})
        
        # Validasi tanggal produksi
        if not tanggal_produksi:
            return jsonify({'success': False, 'message': 'Tanggal produksi tidak boleh kosong'})
        try:
            from datetime import datetime
            datetime.strptime(tanggal_produksi, '%Y-%m-%d')
        except ValueError:
            return jsonify({'success': False, 'message': 'Format tanggal produksi tidak valid'})
        
        # Parse bahan baku data
        if bahan_baku_data is None:
            return jsonify({'success': False, 'message': 'Data bahan baku tidak boleh kosong'})
        try:
            bahan_baku_list = json.loads(bahan_baku_data)
        except Exception:
            return jsonify({'success': False, 'message': 'Format data bahan baku tidak valid'})
        
        if not bahan_baku_list:
            return jsonify({'success': False, 'message': 'Pilih minimal satu bahan baku'})
        
        # Validasi tidak ada duplikasi bahan baku
        bahan_ids = [bahan['bahan_baku_id'] for bahan in bahan_baku_list]
        if len(bahan_ids) != len(set(bahan_ids)):
            return jsonify({'success': False, 'message': 'Tidak boleh ada bahan baku yang duplikat'})
        
        # Validasi dan konversi nilai numerik
        try:
            if not jumlah_produk or not persentase_laba:
                return jsonify({'success': False, 'message': 'Jumlah produk dan persentase laba tidak boleh kosong'})
            
            if not jumlah_produk or jumlah_produk.strip() == '':
                return jsonify({'success': False, 'message': 'Jumlah produk tidak boleh kosong'})
                
            if not persentase_laba or persentase_laba.strip() == '':
                return jsonify({'success': False, 'message': 'Persentase laba tidak boleh kosong'})
                
            jumlah_produk_int = int(jumlah_produk)
            persentase_laba_dec = Decimal(str(persentase_laba).replace('.', '').replace(',', '.'))
        except (ValueError, TypeError, InvalidOperation):
            return jsonify({'success': False, 'message': 'Nilai jumlah produk dan persentase laba harus berupa angka'})
        if jumlah_produk_int <= 0 or persentase_laba_dec < 0:
            return jsonify({'success': False, 'message': 'Jumlah produk dan persentase laba harus lebih dari 0'})
        
        # Validasi nama_produk
        if not nama_produk or nama_produk.strip() == '':
            return jsonify({'success': False, 'message': 'Nama produk tidak boleh kosong'})
        
        # Normalisasi nama_produk
        nama_produk = nama_produk.strip()
        nama_produk = ' '.join([w.capitalize() for w in nama_produk.split()])
        
        # Hitung total biaya (tanpa validasi stok)
        total_biaya = Decimal('0')
        for bahan in bahan_baku_list:
            try:
                bahan_id = bahan['bahan_baku_id']
                quantity_digunakan = Decimal(str(bahan['quantity_digunakan']))
            except (KeyError, InvalidOperation, ValueError, TypeError):
                return jsonify({'success': False, 'message': 'Data bahan baku tidak valid'})
            
            # Validasi quantity tidak boleh negatif
            if quantity_digunakan < 0:
                return jsonify({'success': False, 'message': f'Quantity bahan baku tidak boleh negatif'})
            
            # Validasi quantity tidak boleh nol
            if quantity_digunakan == 0:
                return jsonify({'success': False, 'message': f'Quantity bahan baku tidak boleh 0'})
            
            # Ambil harga satuan bahan baku
            cursor.execute('SELECT harga_satuan FROM bahan_baku WHERE id = %s', (bahan_id,))
            stok_data = cursor.fetchone()
            
            if not stok_data:
                return jsonify({'success': False, 'message': f'Bahan baku tidak ditemukan'})
            
            harga_satuan = Decimal(str(stok_data['harga_satuan']))
            
            subtotal = quantity_digunakan * harga_satuan
            total_biaya += subtotal
        
        # Hitung HPP Dasar = Total Biaya / Jumlah Produksi
        hpp_dasar = total_biaya / jumlah_produk_int if jumlah_produk_int > 0 else Decimal('0')
        
        # Hitung Harga Jual = HPP + (Persentase Laba × HPP)
        harga_jual_calculated = hpp_dasar + (persentase_laba_dec / 100 * hpp_dasar)
        
        # Tidak perlu mengembalikan stok lama karena produksi tidak mengurangi stok
        
        # Update data produksi
        cursor.execute('''
            UPDATE produksi 
            SET nama_produk = %s, jumlah_produk = %s, harga_jual = %s, persentase_laba = %s,
                total_biaya = %s, hpp = %s, tanggal_produksi = %s
            WHERE id = %s
        ''', (nama_produk, jumlah_produk_int, float(harga_jual_calculated), float(persentase_laba_dec),
              float(total_biaya), float(hpp_dasar), tanggal_produksi, produksi_id))
        
        # Hapus data produk_bahan_baku lama
        cursor.execute('DELETE FROM produk_bahan_baku WHERE produksi_id = %s', (produksi_id,))
        
        # Insert data produk_bahan_baku baru (tanpa mengurangi stok)
        for bahan in bahan_baku_list:
            bahan_id = bahan['bahan_baku_id']
            quantity_digunakan = Decimal(str(bahan['quantity_digunakan']))
            
            # Ambil data bahan baku untuk harga
            cursor.execute('SELECT harga_satuan FROM bahan_baku WHERE id = %s', (bahan_id,))
            stok_data = cursor.fetchone()
            harga_satuan = Decimal(str(stok_data['harga_satuan']))
            
            # Validasi quantity tidak boleh negatif
            if quantity_digunakan < 0:
                return jsonify({'success': False, 'message': f'Quantity bahan baku ID {bahan_id} tidak boleh negatif'})
            
            # Validasi quantity tidak boleh 0
            if quantity_digunakan == 0:
                return jsonify({'success': False, 'message': f'Quantity bahan baku ID {bahan_id} tidak boleh 0'})
            
            subtotal = quantity_digunakan * harga_satuan
            
            # Insert ke produk_bahan_baku
            cursor.execute('''
                INSERT INTO produk_bahan_baku (produksi_id, bahan_baku_id, quantity_digunakan, harga_satuan, subtotal)
                VALUES (%s, %s, %s, %s, %s)
            ''', (produksi_id, bahan_id, float(quantity_digunakan), float(harga_satuan), float(subtotal)))
            
            # Tidak mengurangi stok bahan baku saat produksi
            # Stok bahan baku tetap utuh untuk keperluan pencatatan saja
        
        app_funcs["mysql"].connection.commit()
        
        # Log aktivitas
        mahasiswa_info = app_funcs['get_mahasiswa_info_from_session']()
        if mahasiswa_info:
            data_baru = {
                'id': produksi_id,
                'nama_produk': nama_produk,
                'jumlah_produk': jumlah_produk_int,
                'tanggal_produksi': tanggal_produksi,
                'persentase_laba': float(persentase_laba_dec)
            }
            app_funcs['log_mahasiswa_activity'](
                mahasiswa_info['id'], 
                mahasiswa_info['nim'], 
                mahasiswa_info['nama_ketua'], 
                'edit', 
                'produksi', 
                f'id_{produksi_id}', 
                f'Mengubah produksi: {nama_produk}',
                None,  # Data lama tidak diambil untuk performa
                data_baru
            )
        
        # Hitung dan simpan data laba rugi otomatis
        try:
            app_funcs['update_laba_rugi_on_produksi_change'](proposal_id)
        except Exception as e:
            print(f"Warning: Error updating laba rugi: {str(e)}")
        
        cursor.close()
        
        return jsonify({'success': True, 'message': 'Produksi berhasil diupdate dan laba rugi telah diperbarui otomatis'})
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@mahasiswa_bp.route('/delete_produksi', methods=['POST'])
def delete_produksi():
    app_funcs = get_app_functions()
    if 'user_type' not in session or session['user_type'] != 'mahasiswa':
        return jsonify({'success': False, 'message': 'Anda harus login sebagai mahasiswa'})
    
    if not hasattr(app_funcs["mysql"], 'connection') or app_funcs["mysql"].connection is None:
        return jsonify({'success': False, 'message': 'Koneksi ke database gagal'})
    
    try:
        cursor = app_funcs["mysql"].connection.cursor(MySQLdb.cursors.DictCursor)
        
        produksi_id = request.form.get('produksi_id')
        
        if not produksi_id:
            return jsonify({'success': False, 'message': 'ID produksi tidak valid'})
        
        # Validasi produksi_id adalah angka
        try:
            produksi_id = int(produksi_id)
        except (ValueError, TypeError):
            return jsonify({'success': False, 'message': 'ID produksi tidak valid'})
        
        # Cek apakah produksi ada dan milik mahasiswa yang login
        cursor.execute('''
            SELECT p.* FROM produksi p 
            JOIN proposal pr ON p.proposal_id = pr.id 
            WHERE p.id = %s AND pr.nim = %s
        ''', (produksi_id, session['nim']))
        produksi_exists = cursor.fetchone()
        if not produksi_exists:
            return jsonify({'success': False, 'message': 'Produksi tidak ditemukan atau bukan milik Anda'})
        
        # Tidak perlu mengembalikan stok karena produksi tidak mengurangi stok
        # Stok bahan baku tetap utuh
        
        # Delete dari database
        cursor.execute('DELETE FROM produksi WHERE id = %s', (produksi_id,))
        
        app_funcs["mysql"].connection.commit()
        
        # Log aktivitas
        mahasiswa_info = app_funcs['get_mahasiswa_info_from_session']()
        if mahasiswa_info and produksi_exists:
            data_lama = dict(produksi_exists)
            app_funcs['log_mahasiswa_activity'](
                mahasiswa_info['id'], 
                mahasiswa_info['nim'], 
                mahasiswa_info['nama_ketua'], 
                'hapus', 
                'produksi', 
                f'id_{produksi_id}', 
                f'Menghapus produksi: {data_lama.get("nama_produk", "Unknown")}',
                data_lama,
                None
            )
        
        # Hitung dan simpan data laba rugi otomatis
        try:
            app_funcs['update_laba_rugi_on_produksi_change'](produksi_exists['proposal_id'])
        except Exception as e:
            print(f"Warning: Error updating laba rugi: {str(e)}")
        
        cursor.close()
        
        return jsonify({'success': True, 'message': 'Produksi berhasil dihapus dan laba rugi telah diperbarui otomatis'})
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})
@mahasiswa_bp.route('/aktivitas_produksi_mahasiswa')
def aktivitas_produksi_mahasiswa():
    app_funcs = get_app_functions()
    if 'user_type' not in session or session['user_type'] != 'mahasiswa':
        flash('Anda harus login sebagai mahasiswa!', 'danger')
        return redirect(url_for('index'))
    
    if not hasattr(app_funcs["mysql"], 'connection') or app_funcs["mysql"].connection is None:
        flash('Koneksi ke database gagal. Cek konfigurasi database!', 'danger')
        return render_template('mahasiswa/produksi.html', proposal_data={})
    
    try:
        cursor = app_funcs["mysql"].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Ambil data mahasiswa yang login
        cursor.execute('SELECT * FROM mahasiswa WHERE nim = %s', (session['nim'],))
        mahasiswa = cursor.fetchone()
        
        if not mahasiswa:
            flash('Data mahasiswa tidak ditemukan', 'error')
            return redirect(url_for('index'))
        
        # Ambil proposal mahasiswa yang sudah lolos
        cursor.execute('''
            SELECT p.*, pr.id as produksi_id, pr.nama_produk, pr.jumlah_produk, 
                   pr.harga_jual, pr.total_biaya, pr.hpp, pr.tanggal_produksi,
                   pbb.id as produk_bahan_baku_id, pbb.quantity_digunakan, pbb.harga_satuan as harga_satuan_digunakan,
                   pbb.subtotal, bb.nama_bahan, bb.satuan
            FROM proposal p
            LEFT JOIN produksi pr ON p.id = pr.proposal_id
            LEFT JOIN produk_bahan_baku pbb ON pr.id = pbb.produksi_id
            LEFT JOIN bahan_baku bb ON pbb.bahan_baku_id = bb.id
            WHERE p.nim = %s AND p.status_admin = 'lolos'
            ORDER BY p.tanggal_buat DESC, pr.tanggal_produksi DESC
        ''', (session['nim'],))
        
        proposals = cursor.fetchall()
        
        # Group data produksi berdasarkan proposal
        proposal_data = {}
        for row in proposals:
            proposal_id = row['id']
            if proposal_id not in proposal_data:
                proposal_data[proposal_id] = {
                    'proposal': {
                        'id': row['id'],
                        'judul_usaha': row['judul_usaha'],
                        'kategori': row['kategori'],
                        'status': row['status'],
                        'tahun_nib': row['tahun_nib']
                    },
                    'produksi': []
                }
            
            if row['produksi_id']:
                # Cek apakah produksi sudah ada di list
                produksi_exists = False
                for existing_produksi in proposal_data[proposal_id]['produksi']:
                    if existing_produksi['id'] == row['produksi_id']:
                        # Tambahkan bahan baku ke produksi yang sudah ada
                        if row['produk_bahan_baku_id']:
                            existing_produksi['bahan_baku'].append({
                                'nama_bahan': row['nama_bahan'],
                                'quantity_digunakan': row['quantity_digunakan'],
                                'harga_satuan': row['harga_satuan_digunakan'],
                                'subtotal': row['subtotal'],
                                'satuan': row['satuan']
                            })
                        produksi_exists = True
                        break
                
                if not produksi_exists:
                    # Buat produksi baru
                    produksi_data = {
                        'id': row['produksi_id'],
                        'nama_produk': row['nama_produk'],
                        'jumlah_produk': row['jumlah_produk'],
                        'harga_jual': row['harga_jual'],
                        'total_biaya': row['total_biaya'],
                        'hpp': row['hpp'],
                        'tanggal_produksi': row['tanggal_produksi'],
                        'bahan_baku': []
                    }
                    
                    if row['produk_bahan_baku_id']:
                        produksi_data['bahan_baku'].append({
                            'nama_bahan': row['nama_bahan'],
                            'quantity_digunakan': row['quantity_digunakan'],
                            'harga_satuan': row['harga_satuan_digunakan'],
                            'subtotal': row['subtotal'],
                            'satuan': row['satuan']
                        })
                    
                    proposal_data[proposal_id]['produksi'].append(produksi_data)
        
        # Ambil status mahasiswa
        cursor.execute('SELECT status FROM mahasiswa WHERE nim = %s', (session['nim'],))
        mahasiswa_status = cursor.fetchone()
        status_mahasiswa = mahasiswa_status['status'] if mahasiswa_status else 'proses'
        
        cursor.close()
        return render_template('mahasiswa/produksi.html', proposal_data=proposal_data, status_mahasiswa=status_mahasiswa)
        
    except Exception as e:
        flash(f'Error saat mengambil data: {str(e)}', 'danger')
        return render_template('mahasiswa/produksi.html', proposal_data={}, status_mahasiswa='proses')

@mahasiswa_bp.route('/pembelian_alat_mahasiswa')
def pembelian_alat_mahasiswa():
    app_funcs = get_app_functions()
    if 'user_type' not in session or session['user_type'] != 'mahasiswa':
        flash('Anda harus login sebagai mahasiswa', 'danger')
        return redirect(url_for('index'))
    
    if not hasattr(app_funcs["mysql"], 'connection') or app_funcs["mysql"].connection is None:
        flash('Koneksi ke database gagal', 'danger')
        return redirect(url_for('index'))
    
    try:
        cursor = app_funcs["mysql"].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Ambil proposal yang sudah disetujui (status_admin = 'lolos')
        cursor.execute('''
            SELECT * FROM proposal 
            WHERE nim = %s AND status_admin = 'lolos'
            ORDER BY tanggal_buat DESC
        ''', (session['nim'],))
        
        proposals = cursor.fetchall()
        
        if not proposals:
            flash('Anda harus memiliki proposal yang sudah disetujui untuk mengelola alat produksi', 'warning')
            return render_template('mahasiswa/alat_produksi.html', proposal_data={})
        
        # Ambil data alat produksi untuk setiap proposal
        proposal_data = {}
        for proposal in proposals:
            proposal_id = proposal['id']
            
            # Ambil data alat produksi
            cursor.execute('''
                SELECT * FROM alat_produksi 
                WHERE proposal_id = %s 
                ORDER BY tanggal_beli DESC
            ''', (proposal_id,))
            
            alat_produksi = cursor.fetchall()
            
            proposal_data[proposal_id] = {
                'proposal': proposal,
                'alat_produksi': alat_produksi
            }
        
        # Ambil status mahasiswa
        cursor.execute('SELECT status FROM mahasiswa WHERE nim = %s', (session['nim'],))
        mahasiswa_status = cursor.fetchone()
        status_mahasiswa = mahasiswa_status['status'] if mahasiswa_status else 'proses'
        
        cursor.close()
        return render_template('mahasiswa/alat_produksi.html', proposal_data=proposal_data, status_mahasiswa=status_mahasiswa)
        
    except Exception as e:
        flash(f'Error saat mengambil data: {str(e)}', 'danger')
        return render_template('mahasiswa/alat_produksi.html', proposal_data={}, status_mahasiswa='proses')

@mahasiswa_bp.route('/alat_produksi_mahasiswa')
def alat_produksi_mahasiswa():
    app_funcs = get_app_functions()
    if 'user_type' not in session or session['user_type'] != 'mahasiswa':
        flash('Anda harus login sebagai mahasiswa', 'danger')
        return redirect(url_for('index'))
    
    if not hasattr(app_funcs["mysql"], 'connection') or app_funcs["mysql"].connection is None:
        flash('Koneksi ke database gagal', 'danger')
        return redirect(url_for('index'))
    
    try:
        cursor = app_funcs["mysql"].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Ambil proposal yang sudah disetujui (status_admin = 'lolos')
        cursor.execute('''
            SELECT * FROM proposal 
            WHERE nim = %s AND status_admin = 'lolos'
            ORDER BY tanggal_buat DESC
        ''', (session['nim'],))
        
        proposals = cursor.fetchall()
        
        if not proposals:
            flash('Anda harus memiliki proposal yang sudah disetujui untuk mengelola alat produksi', 'warning')
            return render_template('mahasiswa/alat_produksi.html', proposal_data={})
        
        # Ambil data alat produksi untuk setiap proposal
        proposal_data = {}
        for proposal in proposals:
            proposal_id = proposal['id']
            
            # Ambil data alat produksi
            cursor.execute('''
                SELECT * FROM alat_produksi 
                WHERE proposal_id = %s 
                ORDER BY tanggal_beli DESC
            ''', (proposal_id,))
            
            alat_produksi = cursor.fetchall()
            
            proposal_data[proposal_id] = {
                'proposal': proposal,
                'alat_produksi': alat_produksi
            }
        
        # Ambil status mahasiswa
        cursor.execute('SELECT status FROM mahasiswa WHERE nim = %s', (session['nim'],))
        mahasiswa_status = cursor.fetchone()
        status_mahasiswa = mahasiswa_status['status'] if mahasiswa_status else 'proses'
        
        cursor.close()
        return render_template('mahasiswa/alat_produksi.html', proposal_data=proposal_data, status_mahasiswa=status_mahasiswa)
        
    except Exception as e:
        flash(f'Error saat mengambil data: {str(e)}', 'danger')
        return render_template('mahasiswa/alat_produksi.html', proposal_data={}, status_mahasiswa='proses')

# Route untuk CRUD alat produksi
@mahasiswa_bp.route('/tambah_alat_produksi', methods=['POST'])
def tambah_alat_produksi():
    app_funcs = get_app_functions()
    from decimal import Decimal, InvalidOperation
    from datetime import datetime
    
    if 'user_type' not in session or session['user_type'] != 'mahasiswa':
        return jsonify({'success': False, 'message': 'Anda harus login sebagai mahasiswa'})
    
    if not hasattr(app_funcs["mysql"], 'connection') or app_funcs["mysql"].connection is None:
        return jsonify({'success': False, 'message': 'Koneksi ke database gagal'})
    
    try:
        cursor = app_funcs["mysql"].connection.cursor(MySQLdb.cursors.DictCursor)
        
        proposal_id = request.form.get('proposal_id')
        nama_alat = request.form.get('nama_alat')
        quantity = request.form.get('quantity')
        harga = request.form.get('harga')
        harga_jual = request.form.get('harga_jual', '0')
        tanggal_beli = request.form.get('tanggal_beli')
        keterangan = request.form.get('keterangan', '')
        
        # Validasi input
        if not all([proposal_id, nama_alat, quantity, harga, tanggal_beli]):
            return jsonify({'success': False, 'message': 'Semua field wajib diisi'})
        
        # Validasi proposal_id
        if not proposal_id:
            return jsonify({'success': False, 'message': 'ID proposal tidak boleh kosong'})
        try:
            proposal_id = int(proposal_id)
        except (ValueError, TypeError):
            return jsonify({'success': False, 'message': 'ID proposal tidak valid'})
        
        # Cek apakah proposal ada dan milik mahasiswa yang login
        cursor.execute('SELECT id FROM proposal WHERE id = %s AND nim = %s AND status_admin = "lolos"', (proposal_id, session['nim']))
        proposal_exists = cursor.fetchone()
        if not proposal_exists:
            return jsonify({'success': False, 'message': 'Proposal tidak ditemukan atau belum disetujui'})
        
        # Validasi dan konversi nilai numerik
        if not quantity or not harga:
            return jsonify({'success': False, 'message': 'Quantity dan harga tidak boleh kosong'})
        try:
            quantity_int = int(quantity)
            harga_dec = Decimal(str(harga).replace('.', '').replace(',', '.'))
        except (ValueError, TypeError, InvalidOperation):
            return jsonify({'success': False, 'message': 'Quantity dan harga harus berupa angka'})
        
        if quantity_int <= 0 or harga_dec <= 0:
            return jsonify({'success': False, 'message': 'Quantity dan harga harus lebih dari 0'})
        
        # Validasi tanggal
        if not tanggal_beli:
            return jsonify({'success': False, 'message': 'Tanggal beli tidak boleh kosong'})
        try:
            datetime.strptime(tanggal_beli, '%Y-%m-%d')
        except ValueError:
            return jsonify({'success': False, 'message': 'Format tanggal tidak valid'})
        
        # Normalisasi nama_alat
        if not nama_alat:
            return jsonify({'success': False, 'message': 'Nama alat tidak boleh kosong'})
        nama_alat = nama_alat.strip()
        nama_alat = ' '.join([w.capitalize() for w in nama_alat.split()])

                # Normalisasi keterangan (jika ada)
        if keterangan:
            keterangan = keterangan.strip()
            keterangan = ' '.join([w.capitalize() for w in keterangan.split()])
        
        # Hitung total alat produksi
        total_alat_produksi = quantity_int * float(harga_dec)
        
        # Konversi harga_jual
        harga_jual_dec = Decimal(str(harga_jual).replace('.', '').replace(',', '.')) if harga_jual and harga_jual != '0' else Decimal('0')
        
        # Insert ke database
        cursor.execute('''
            INSERT INTO alat_produksi (proposal_id, nama_alat, quantity, harga, harga_jual, total_alat_produksi, tanggal_beli, keterangan)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
        ''', (proposal_id, nama_alat, quantity_int, float(harga_dec), float(harga_jual_dec), total_alat_produksi, tanggal_beli, keterangan))
        
        alat_id = cursor.lastrowid
        app_funcs["mysql"].connection.commit()
        
        # Log aktivitas
        mahasiswa_info = app_funcs['get_mahasiswa_info_from_session']()
        if mahasiswa_info:
            data_baru = {
                'id': alat_id,
                'nama_alat': nama_alat,
                'quantity': quantity_int,
                'harga': float(harga_dec),
                'harga_jual': float(harga_jual_dec),
                'total_alat_produksi': total_alat_produksi,
                'tanggal_beli': tanggal_beli,
                'keterangan': keterangan
            }
            app_funcs['log_mahasiswa_activity'](
                mahasiswa_info['id'], 
                mahasiswa_info['nim'], 
                mahasiswa_info['nama_ketua'], 
                'tambah', 
                'alat_produksi', 
                f'id_{alat_id}', 
                f'Menambahkan alat produksi: {nama_alat}',
                None,
                data_baru
            )
        
        # Hitung dan simpan data laba rugi
        app_funcs['calculate_and_save_laba_rugi'](proposal_id)
        
        # Update arus kas otomatis untuk bulan pembelian
        if tanggal_beli:
            bulan_pembelian = datetime.strptime(tanggal_beli, '%Y-%m-%d').strftime('%Y-%m')
            app_funcs['update_arus_kas_otomatis'](proposal_id, bulan_pembelian)
        
        cursor.close()
        
        return jsonify({'success': True, 'message': 'Alat produksi berhasil ditambahkan'})
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@mahasiswa_bp.route('/get_alat_produksi/<int:alat_id>')
def get_alat_produksi(alat_id):
    app_funcs = get_app_functions()
    if 'user_type' not in session or session['user_type'] != 'mahasiswa':
        return jsonify({'success': False, 'message': 'Anda harus login sebagai mahasiswa'})
    
    if not hasattr(app_funcs["mysql"], 'connection') or app_funcs["mysql"].connection is None:
        return jsonify({'success': False, 'message': 'Koneksi ke database gagal'})
    
    try:
        cursor = app_funcs["mysql"].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Ambil data alat produksi dan validasi kepemilikan
        cursor.execute('''
            SELECT ap.* FROM alat_produksi ap 
            JOIN proposal p ON ap.proposal_id = p.id 
            WHERE ap.id = %s AND p.nim = %s
        ''', (alat_id, session['nim']))
        
        alat = cursor.fetchone()
        
        if not alat:
            return jsonify({'success': False, 'message': 'Data alat produksi tidak ditemukan atau bukan milik Anda'})
        
        # Format tanggal
        if alat.get('tanggal_beli'):
            alat['tanggal_beli'] = alat['tanggal_beli'].strftime('%Y-%m-%d')
        
        # Konversi ke integer untuk frontend
        alat['quantity'] = int(alat['quantity'])
        alat['harga'] = int(alat['harga'])
        alat['harga_jual'] = int(alat['harga_jual']) if alat.get('harga_jual') else 0
        alat['total_alat_produksi'] = int(alat['total_alat_produksi'])
        
        cursor.close()
        return jsonify({'success': True, 'data': alat})
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@mahasiswa_bp.route('/update_alat_produksi', methods=['POST'])
def update_alat_produksi():
    app_funcs = get_app_functions()
    from decimal import Decimal, InvalidOperation
    from datetime import datetime
    
    if 'user_type' not in session or session['user_type'] != 'mahasiswa':
        return jsonify({'success': False, 'message': 'Anda harus login sebagai mahasiswa'})
    
    if not hasattr(app_funcs["mysql"], 'connection') or app_funcs["mysql"].connection is None:
        return jsonify({'success': False, 'message': 'Koneksi ke database gagal'})
    
    try:
        cursor = app_funcs["mysql"].connection.cursor(MySQLdb.cursors.DictCursor)
        
        alat_id = request.form.get('alat_id')
        nama_alat = request.form.get('nama_alat')
        quantity = request.form.get('quantity')
        harga = request.form.get('harga')
        harga_jual = request.form.get('harga_jual', '0')
        tanggal_beli = request.form.get('tanggal_beli')
        keterangan = request.form.get('keterangan', '')
        
        # Validasi input
        if not all([alat_id, nama_alat, quantity, harga, tanggal_beli]):
            return jsonify({'success': False, 'message': 'Semua field wajib diisi'})
        
        # Validasi alat_id
        if not alat_id:
            return jsonify({'success': False, 'message': 'ID alat tidak boleh kosong'})
        try:
            alat_id = int(alat_id)
        except (ValueError, TypeError):
            return jsonify({'success': False, 'message': 'ID alat tidak valid'})
        
        # Cek apakah alat ada dan milik mahasiswa yang login
        cursor.execute('''
            SELECT ap.* FROM alat_produksi ap 
            JOIN proposal p ON ap.proposal_id = p.id 
            WHERE ap.id = %s AND p.nim = %s
        ''', (alat_id, session['nim']))
        
        alat_exists = cursor.fetchone()
        if not alat_exists:
            return jsonify({'success': False, 'message': 'Alat produksi tidak ditemukan atau bukan milik Anda'})
        
        # Validasi dan konversi nilai numerik
        if not quantity or not harga:
            return jsonify({'success': False, 'message': 'Quantity dan harga tidak boleh kosong'})
        try:
            quantity_int = int(quantity)
            harga_dec = Decimal(str(harga).replace('.', '').replace(',', '.'))
        except (ValueError, TypeError, InvalidOperation):
            return jsonify({'success': False, 'message': 'Quantity dan harga harus berupa angka'})
        
        if quantity_int <= 0 or harga_dec <= 0:
            return jsonify({'success': False, 'message': 'Quantity dan harga harus lebih dari 0'})
        
        # Validasi tanggal
        if not tanggal_beli:
            return jsonify({'success': False, 'message': 'Tanggal beli tidak boleh kosong'})
        try:
            datetime.strptime(tanggal_beli, '%Y-%m-%d')
        except ValueError:
            return jsonify({'success': False, 'message': 'Format tanggal tidak valid'})
        
        # Normalisasi nama_alat
        if not nama_alat:
            return jsonify({'success': False, 'message': 'Nama alat tidak boleh kosong'})
        nama_alat = nama_alat.strip()
        nama_alat = ' '.join([w.capitalize() for w in nama_alat.split()])
                # Normalisasi keterangan (jika ada)
        if keterangan:
            keterangan = keterangan.strip()
            keterangan = ' '.join([w.capitalize() for w in keterangan.split()])
        
        # Hitung total alat produksi
        total_alat_produksi = quantity_int * float(harga_dec)
        
        # Konversi harga_jual
        harga_jual_dec = Decimal(str(harga_jual).replace('.', '').replace(',', '.')) if harga_jual and harga_jual != '0' else Decimal('0')
        
        # Update data
        cursor.execute('''
            UPDATE alat_produksi 
            SET nama_alat = %s, quantity = %s, harga = %s, harga_jual = %s, total_alat_produksi = %s, tanggal_beli = %s, keterangan = %s
            WHERE id = %s
        ''', (nama_alat, quantity_int, float(harga_dec), float(harga_jual_dec), total_alat_produksi, tanggal_beli, keterangan, alat_id))
        
        app_funcs["mysql"].connection.commit()
        
        # Log aktivitas
        mahasiswa_info = app_funcs['get_mahasiswa_info_from_session']()
        if mahasiswa_info and alat_id:
            data_baru = {
                'id': int(alat_id),
                'nama_alat': nama_alat,
                'quantity': quantity_int,
                'harga': float(harga_dec),
                'harga_jual': float(harga_jual_dec),
                'total_alat_produksi': total_alat_produksi,
                'tanggal_beli': tanggal_beli,
                'keterangan': keterangan
            }
            app_funcs['log_mahasiswa_activity'](
                mahasiswa_info['id'], 
                mahasiswa_info['nim'], 
                mahasiswa_info['nama_ketua'], 
                'edit', 
                'alat_produksi', 
                f'id_{alat_id}', 
                f'Mengubah alat produksi: {nama_alat}',
                dict(alat_exists) if alat_exists else None,
                data_baru
            )
        
        # Hitung dan simpan data laba rugi
        app_funcs['calculate_and_save_laba_rugi'](alat_exists['proposal_id'])
        
        # Update arus kas otomatis untuk bulan pembelian
        if tanggal_beli:
            bulan_pembelian = datetime.strptime(tanggal_beli, '%Y-%m-%d').strftime('%Y-%m')
            app_funcs['update_arus_kas_otomatis'](alat_exists['proposal_id'], bulan_pembelian)
        
        cursor.close()
        
        return jsonify({'success': True, 'message': 'Alat produksi berhasil diupdate'})
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@mahasiswa_bp.route('/delete_alat_produksi', methods=['POST'])
def delete_alat_produksi():
    app_funcs = get_app_functions()
    if 'user_type' not in session or session['user_type'] != 'mahasiswa':
        return jsonify({'success': False, 'message': 'Anda harus login sebagai mahasiswa'})
    
    if not hasattr(app_funcs["mysql"], 'connection') or app_funcs["mysql"].connection is None:
        return jsonify({'success': False, 'message': 'Koneksi ke database gagal'})
    
    try:
        cursor = app_funcs["mysql"].connection.cursor(MySQLdb.cursors.DictCursor)
        
        alat_id = request.form.get('alat_id')
        
        if not alat_id:
            return jsonify({'success': False, 'message': 'ID alat tidak boleh kosong'})
        
        try:
            alat_id = int(alat_id)
        except (ValueError, TypeError):
            return jsonify({'success': False, 'message': 'ID alat tidak valid'})
        
        # Cek apakah alat ada dan milik mahasiswa yang login
        cursor.execute('''
            SELECT ap.* FROM alat_produksi ap 
            JOIN proposal p ON ap.proposal_id = p.id 
            WHERE ap.id = %s AND p.nim = %s
        ''', (alat_id, session['nim']))
        
        alat_exists = cursor.fetchone()
        if not alat_exists:
            return jsonify({'success': False, 'message': 'Alat produksi tidak ditemukan atau bukan milik Anda'})
        
        # Hapus data
        cursor.execute('DELETE FROM alat_produksi WHERE id = %s', (alat_id,))
        
        app_funcs["mysql"].connection.commit()
        
        # Log aktivitas
        mahasiswa_info = app_funcs['get_mahasiswa_info_from_session']()
        if mahasiswa_info and alat_exists:
            data_lama = dict(alat_exists)
            app_funcs['log_mahasiswa_activity'](
                mahasiswa_info['id'], 
                mahasiswa_info['nim'], 
                mahasiswa_info['nama_ketua'], 
                'hapus', 
                'alat_produksi', 
                f'id_{alat_id}', 
                f'Menghapus alat produksi: {data_lama.get("nama_alat", "Unknown")}',
                data_lama,
                None
            )
        
        # Hitung dan simpan data laba rugi
        app_funcs['calculate_and_save_laba_rugi'](alat_exists['proposal_id'])
        
        # Update arus kas otomatis untuk bulan pembelian yang dihapus
        if alat_exists.get('tanggal_beli'):
            bulan_pembelian = alat_exists['tanggal_beli'].strftime('%Y-%m')
            app_funcs['update_arus_kas_otomatis'](alat_exists['proposal_id'], bulan_pembelian)
        
        cursor.close()
        
        return jsonify({'success': True, 'message': 'Alat produksi berhasil dihapus'})
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})
# Route untuk Biaya Lain-lain
@mahasiswa_bp.route('/biaya_non_operasional_mahasiswa')
def biaya_non_operasional_mahasiswa():
    app_funcs = get_app_functions()
    if 'user_type' not in session or session['user_type'] != 'mahasiswa':
        flash('Anda harus login sebagai mahasiswa!', 'danger')
        return redirect(url_for('index'))
    
    if not hasattr(app_funcs["mysql"], 'connection') or app_funcs["mysql"].connection is None:
        flash('Koneksi ke database gagal. Cek konfigurasi database!', 'danger')
        return render_template('mahasiswa/biaya_non_operasional.html', proposal_data={})

    try:
        cursor = app_funcs["mysql"].connection.cursor(MySQLdb.cursors.DictCursor)
        print('DEBUG: About to execute query with session[nim] =', repr(session.get('nim')))
        print('DEBUG: session[nim] type =', type(session.get('nim')))
        print('DEBUG: session[nim] == "1" =', session.get('nim') == "1")
        print('DEBUG: session[nim] == 1 =', session.get('nim') == 1)
        
        # Ambil data proposal yang lolos untuk mahasiswa ini
        cursor.execute('''
            SELECT p.id, p.judul_usaha, p.kategori, p.tahun_nib, p.status_admin
            FROM proposal p
            WHERE p.nim = %s AND p.status_admin = 'lolos'
            ORDER BY p.tanggal_buat DESC
        ''', (session['nim'],))
                
        proposals = cursor.fetchall()
        print('DEBUG: Query result proposals =', proposals)
        print('DEBUG: len(proposals) =', len(proposals))
        proposal_data = {}

        for proposal in proposals:
            proposal_id = proposal['id']
            print(f'DEBUG: Processing proposal_id = {proposal_id}')
            
            # Buat tabel biaya_non_operasional jika belum ada
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS biaya_non_operasional (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    proposal_id INT NOT NULL,
                    nama_biaya VARCHAR(255) NOT NULL,
                    quantity INT NOT NULL,
                    harga_satuan DECIMAL(15,2) NOT NULL,
                    total_harga DECIMAL(15,2) NOT NULL,
                    tanggal_transaksi DATE,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                    FOREIGN KEY (proposal_id) REFERENCES proposal(id) ON DELETE CASCADE
                )
            ''')
            
            # Ambil data Biaya Lain-lain untuk proposal ini
            cursor.execute('''
                SELECT id, nama_biaya, quantity, harga_satuan, total_harga, tanggal_transaksi
                FROM biaya_non_operasional
                WHERE proposal_id = %s
                ORDER BY created_at DESC
            ''', (proposal_id,))
            
            biaya_non_operasional = cursor.fetchall()
            print(f'DEBUG: biaya_non_operasional for proposal {proposal_id} = {biaya_non_operasional}')
            
            proposal_data[proposal_id] = {
                'proposal': proposal,
                'biaya_non_operasional': biaya_non_operasional
            }
            print(f'DEBUG: Added to proposal_data[{proposal_id}]')
        
        # Ambil status mahasiswa
        cursor.execute('SELECT status FROM mahasiswa WHERE nim = %s', (session['nim'],))
        mahasiswa_status = cursor.fetchone()
        status_mahasiswa = mahasiswa_status['status'] if mahasiswa_status else 'proses'

        cursor.close()
        
        print('DEBUG: type session[nim] =', type(session.get('nim')))
        print('DEBUG: session[nim] =', session.get('nim'))
        print('DEBUG: proposal_data =', proposal_data)
        
        return render_template('mahasiswa/biaya_non_operasional.html', proposal_data=proposal_data, status_mahasiswa=status_mahasiswa)
        
    except Exception as e:
        flash(f'Error: {str(e)}', 'danger')
        print('DEBUG: Exception occurred:', str(e))
        return render_template('mahasiswa/biaya_non_operasional.html', proposal_data={}, status_mahasiswa='proses')
# Route untuk biaya operasional
@mahasiswa_bp.route('/biaya_operasional_mahasiswa')
def biaya_operasional_mahasiswa():
    app_funcs = get_app_functions()
    if 'user_type' not in session or session['user_type'] != 'mahasiswa':
        flash('Anda harus login sebagai mahasiswa!', 'danger')
        return redirect(url_for('index'))
    
    if not hasattr(app_funcs["mysql"], 'connection') or app_funcs["mysql"].connection is None:
        flash('Koneksi ke database gagal. Cek konfigurasi database!', 'danger')
        return render_template('mahasiswa/biaya_operasional.html', proposal_data={})
    
    try:
        cursor = app_funcs["mysql"].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Ambil data proposal yang lolos untuk mahasiswa ini
        cursor.execute('''
            SELECT p.id, p.judul_usaha, p.kategori, p.tahun_nib, p.status_admin
            FROM proposal p
            WHERE p.nim = %s AND p.status_admin = 'lolos'
            ORDER BY p.tanggal_buat DESC
        ''', (session['nim'],))
        
        proposals = cursor.fetchall()
        proposal_data = {}
        
        for proposal in proposals:
            proposal_id = proposal['id']
            
            # Ambil data biaya operasional untuk proposal ini
            cursor.execute('''
                SELECT id, nama_biaya, estimasi_hari_habis, estimasi_aktif_digunakan, quantity, harga_satuan, total_harga, tanggal_beli
                FROM biaya_operasional
                WHERE proposal_id = %s
                ORDER BY created_at DESC
            ''', (proposal_id,))
            
            biaya_operasional = cursor.fetchall()
            
            proposal_data[proposal_id] = {
                'proposal': proposal,
                'biaya_operasional': biaya_operasional
            }
        
        # Ambil status mahasiswa
        cursor.execute('SELECT status FROM mahasiswa WHERE nim = %s', (session['nim'],))
        mahasiswa_status = cursor.fetchone()
        status_mahasiswa = mahasiswa_status['status'] if mahasiswa_status else 'proses'
        
        cursor.close()
        return render_template('mahasiswa/biaya_operasional.html', proposal_data=proposal_data, status_mahasiswa=status_mahasiswa)
        
    except Exception as e:
        flash(f'Error: {str(e)}', 'danger')
        return render_template('mahasiswa/biaya_operasional.html', proposal_data={}, status_mahasiswa='proses')

@mahasiswa_bp.route('/tambah_biaya_operasional', methods=['POST'])
def tambah_biaya_operasional():
    app_funcs = get_app_functions()
    from decimal import Decimal, InvalidOperation
    
    if 'user_type' not in session or session['user_type'] != 'mahasiswa':
        return jsonify({'success': False, 'message': 'Anda harus login sebagai mahasiswa'})
    
    if not hasattr(app_funcs["mysql"], 'connection') or app_funcs["mysql"].connection is None:
        return jsonify({'success': False, 'message': 'Koneksi ke database gagal'})
    
    try:
        cursor = app_funcs["mysql"].connection.cursor(MySQLdb.cursors.DictCursor)
        
        proposal_id = request.form.get('proposal_id')
        nama_biaya = request.form.get('nama_biaya')
        estimasi_hari_habis = request.form.get('estimasi_hari_habis')
        estimasi_aktif_digunakan = request.form.get('estimasi_aktif_digunakan')
        quantity = request.form.get('quantity')
        harga_satuan = request.form.get('harga_satuan')
        tanggal_beli = request.form.get('tanggal_beli')
        
        # Validasi input
        if not all([proposal_id, nama_biaya, estimasi_hari_habis, estimasi_aktif_digunakan, quantity, harga_satuan, tanggal_beli]):
            return jsonify({'success': False, 'message': 'Semua field wajib diisi'})
        
        # Validasi proposal_id
        if not proposal_id:
            return jsonify({'success': False, 'message': 'ID proposal tidak boleh kosong'})
        try:
            proposal_id = int(proposal_id)
        except (ValueError, TypeError):
            return jsonify({'success': False, 'message': 'ID proposal tidak valid'})
        
        # Cek apakah proposal ada dan milik mahasiswa yang login
        cursor.execute('SELECT id FROM proposal WHERE id = %s AND nim = %s AND status_admin = "lolos"', (proposal_id, session['nim']))
        proposal_exists = cursor.fetchone()
        if not proposal_exists:
            return jsonify({'success': False, 'message': 'Proposal tidak ditemukan atau belum disetujui'})
        
        # Validasi dan konversi nilai numerik
        if not estimasi_hari_habis or not estimasi_aktif_digunakan or not quantity or not harga_satuan or not tanggal_beli:
            return jsonify({'success': False, 'message': 'Estimasi hari habis, estimasi aktif digunakan, quantity, harga satuan, dan tanggal beli tidak boleh kosong'})
        try:
            estimasi_hari_habis_int = int(estimasi_hari_habis)
            estimasi_aktif_digunakan_int = int(estimasi_aktif_digunakan)
            quantity_int = int(quantity)
            harga_satuan_dec = Decimal(str(harga_satuan).replace('.', '').replace(',', '.'))
        except (ValueError, TypeError, InvalidOperation):
            return jsonify({'success': False, 'message': 'Estimasi hari habis, estimasi aktif digunakan, quantity, dan harga satuan harus berupa angka'})
        
        if estimasi_hari_habis_int <= 0 or estimasi_aktif_digunakan_int <= 0 or quantity_int <= 0 or harga_satuan_dec <= 0:
            return jsonify({'success': False, 'message': 'Estimasi hari habis, estimasi aktif digunakan, quantity, dan harga satuan harus lebih dari 0'})
        
        # Normalisasi nama_biaya
        if not nama_biaya:
            return jsonify({'success': False, 'message': 'Nama biaya tidak boleh kosong'})
        nama_biaya = nama_biaya.strip()
        nama_biaya = ' '.join([w.capitalize() for w in nama_biaya.split()])
        
        # Hitung total harga
        total_harga = quantity_int * float(harga_satuan_dec)
        
        # Insert ke database
        cursor.execute('''
            INSERT INTO biaya_operasional (proposal_id, nama_biaya, estimasi_hari_habis, estimasi_aktif_digunakan, quantity, harga_satuan, total_harga, tanggal_beli)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
        ''', (proposal_id, nama_biaya, estimasi_hari_habis_int, estimasi_aktif_digunakan_int, quantity_int, float(harga_satuan_dec), total_harga, tanggal_beli))
        
        biaya_id = cursor.lastrowid
        app_funcs["mysql"].connection.commit()
        
        # Log aktivitas
        mahasiswa_info = app_funcs['get_mahasiswa_info_from_session']()
        if mahasiswa_info:
            data_baru = {
                'id': biaya_id,
                'nama_biaya': nama_biaya,
                'estimasi_hari_habis': estimasi_hari_habis_int,
                'estimasi_aktif_digunakan': estimasi_aktif_digunakan_int,
                'quantity': quantity_int,
                'harga_satuan': float(harga_satuan_dec),
                'total_harga': total_harga,
                'tanggal_beli': tanggal_beli
            }
            app_funcs['log_mahasiswa_activity'](
                mahasiswa_info['id'], 
                mahasiswa_info['nim'], 
                mahasiswa_info['nama_ketua'], 
                'tambah', 
                'biaya_operasional', 
                f'id_{biaya_id}', 
                f'Menambahkan biaya operasional: {nama_biaya}',
                None,
                data_baru
            )
        
        # Hitung dan simpan data laba rugi
        app_funcs['calculate_and_save_laba_rugi'](proposal_id)
        
        # Update arus kas otomatis untuk bulan pembelian
        if tanggal_beli:
            bulan_pembelian = datetime.strptime(tanggal_beli, '%Y-%m-%d').strftime('%Y-%m')
            app_funcs['update_arus_kas_otomatis'](proposal_id, bulan_pembelian)
        
        cursor.close()
        
        return jsonify({'success': True, 'message': 'Biaya operasional berhasil ditambahkan'})
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@mahasiswa_bp.route('/get_biaya_operasional/<int:biaya_id>')
def get_biaya_operasional(biaya_id):
    app_funcs = get_app_functions()
    if 'user_type' not in session or session['user_type'] != 'mahasiswa':
        return jsonify({'success': False, 'message': 'Anda harus login sebagai mahasiswa'})
    
    if not hasattr(app_funcs["mysql"], 'connection') or app_funcs["mysql"].connection is None:
        return jsonify({'success': False, 'message': 'Koneksi ke database gagal'})
    
    try:
        cursor = app_funcs["mysql"].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Ambil data biaya operasional dan validasi kepemilikan
        cursor.execute('''
            SELECT bo.* FROM biaya_operasional bo 
            JOIN proposal p ON bo.proposal_id = p.id 
            WHERE bo.id = %s AND p.nim = %s
        ''', (biaya_id, session['nim']))
        
        biaya = cursor.fetchone()
        
        if not biaya:
            return jsonify({'success': False, 'message': 'Data biaya operasional tidak ditemukan atau bukan milik Anda'})
        
        # Konversi ke integer untuk frontend
        biaya['estimasi_hari_habis'] = int(biaya['estimasi_hari_habis'])
        biaya['estimasi_aktif_digunakan'] = int(biaya['estimasi_aktif_digunakan'])
        biaya['quantity'] = int(biaya['quantity'])
        biaya['harga_satuan'] = int(biaya['harga_satuan'])
        biaya['total_harga'] = int(biaya['total_harga'])
        
        # Format tanggal untuk frontend
        if biaya['tanggal_beli']:
            biaya['tanggal_beli'] = biaya['tanggal_beli'].strftime('%Y-%m-%d')
        else:
            biaya['tanggal_beli'] = ''
        
        cursor.close()
        return jsonify({'success': True, 'data': biaya})
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})
@mahasiswa_bp.route('/update_biaya_operasional', methods=['POST'])
def update_biaya_operasional():
    app_funcs = get_app_functions()
    from decimal import Decimal, InvalidOperation
    
    if 'user_type' not in session or session['user_type'] != 'mahasiswa':
        return jsonify({'success': False, 'message': 'Anda harus login sebagai mahasiswa'})
    
    if not hasattr(app_funcs["mysql"], 'connection') or app_funcs["mysql"].connection is None:
        return jsonify({'success': False, 'message': 'Koneksi ke database gagal'})
    
    try:
        cursor = app_funcs["mysql"].connection.cursor(MySQLdb.cursors.DictCursor)
        
        biaya_id = request.form.get('biaya_id')
        nama_biaya = request.form.get('nama_biaya')
        estimasi_hari_habis = request.form.get('estimasi_hari_habis')
        estimasi_aktif_digunakan = request.form.get('estimasi_aktif_digunakan')
        quantity = request.form.get('quantity')
        harga_satuan = request.form.get('harga_satuan')
        tanggal_beli = request.form.get('tanggal_beli')
        
        # Validasi input
        if not all([biaya_id, nama_biaya, estimasi_hari_habis, estimasi_aktif_digunakan, quantity, harga_satuan, tanggal_beli]):
            return jsonify({'success': False, 'message': 'Semua field wajib diisi'})
        
        # Validasi biaya_id
        if not biaya_id:
            return jsonify({'success': False, 'message': 'ID biaya tidak boleh kosong'})
        try:
            biaya_id = int(biaya_id)
        except (ValueError, TypeError):
            return jsonify({'success': False, 'message': 'ID biaya tidak valid'})
        
        # Cek apakah biaya ada dan milik mahasiswa yang login
        cursor.execute('''
            SELECT bo.* FROM biaya_operasional bo 
            JOIN proposal p ON bo.proposal_id = p.id 
            WHERE bo.id = %s AND p.nim = %s
        ''', (biaya_id, session['nim']))
        
        biaya_exists = cursor.fetchone()
        if not biaya_exists:
            return jsonify({'success': False, 'message': 'Biaya operasional tidak ditemukan atau bukan milik Anda'})
        
        # Validasi dan konversi nilai numerik
        if not estimasi_hari_habis or not estimasi_aktif_digunakan or not quantity or not harga_satuan or not tanggal_beli:
            return jsonify({'success': False, 'message': 'Estimasi hari habis, estimasi aktif digunakan, quantity, harga satuan, dan tanggal beli tidak boleh kosong'})
        try:
            estimasi_hari_habis_int = int(estimasi_hari_habis)
            estimasi_aktif_digunakan_int = int(estimasi_aktif_digunakan)
            quantity_int = int(quantity)
            harga_satuan_dec = Decimal(str(harga_satuan).replace('.', '').replace(',', '.'))
        except (ValueError, TypeError, InvalidOperation):
            return jsonify({'success': False, 'message': 'Estimasi hari habis, estimasi aktif digunakan, quantity, dan harga satuan harus berupa angka'})
        
        if estimasi_hari_habis_int <= 0 or estimasi_aktif_digunakan_int <= 0 or quantity_int <= 0 or harga_satuan_dec <= 0:
            return jsonify({'success': False, 'message': 'Estimasi hari habis, estimasi aktif digunakan, quantity, dan harga satuan harus lebih dari 0'})
        
        # Normalisasi nama_biaya
        if not nama_biaya:
            return jsonify({'success': False, 'message': 'Nama biaya tidak boleh kosong'})
        nama_biaya = nama_biaya.strip()
        nama_biaya = ' '.join([w.capitalize() for w in nama_biaya.split()])
        
        # Hitung total harga
        total_harga = quantity_int * float(harga_satuan_dec)
        
        # Update data
        cursor.execute('''
            UPDATE biaya_operasional 
            SET nama_biaya = %s, estimasi_hari_habis = %s, estimasi_aktif_digunakan = %s, quantity = %s, harga_satuan = %s, total_harga = %s, tanggal_beli = %s
            WHERE id = %s
        ''', (nama_biaya, estimasi_hari_habis_int, estimasi_aktif_digunakan_int, quantity_int, float(harga_satuan_dec), total_harga, tanggal_beli, biaya_id))
        
        app_funcs["mysql"].connection.commit()
        
        # Log aktivitas
        mahasiswa_info = app_funcs['get_mahasiswa_info_from_session']()
        if mahasiswa_info and biaya_id:
            data_baru = {
                'id': int(biaya_id),
                'nama_biaya': nama_biaya,
                'estimasi_hari_habis': estimasi_hari_habis_int,
                'estimasi_aktif_digunakan': estimasi_aktif_digunakan_int,
                'quantity': quantity_int,
                'harga_satuan': float(harga_satuan_dec),
                'total_harga': total_harga,
                'tanggal_beli': tanggal_beli
            }
            app_funcs['log_mahasiswa_activity'](
                mahasiswa_info['id'], 
                mahasiswa_info['nim'], 
                mahasiswa_info['nama_ketua'], 
                'edit', 
                'biaya_operasional', 
                f'id_{biaya_id}', 
                f'Mengubah biaya operasional: {nama_biaya}',
                dict(biaya_exists) if biaya_exists else None,
                data_baru
            )
        
        # Hitung dan simpan data laba rugi
        app_funcs['calculate_and_save_laba_rugi'](biaya_exists['proposal_id'])
        
        # Update arus kas otomatis untuk bulan pembelian
        if tanggal_beli:
            bulan_pembelian = datetime.strptime(tanggal_beli, '%Y-%m-%d').strftime('%Y-%m')
            app_funcs['update_arus_kas_otomatis'](biaya_exists['proposal_id'], bulan_pembelian)
        
        cursor.close()
        
        return jsonify({'success': True, 'message': 'Biaya operasional berhasil diupdate'})
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@mahasiswa_bp.route('/delete_biaya_operasional', methods=['POST'])
def delete_biaya_operasional():
    app_funcs = get_app_functions()
    if 'user_type' not in session or session['user_type'] != 'mahasiswa':
        return jsonify({'success': False, 'message': 'Anda harus login sebagai mahasiswa'})
    
    if not hasattr(app_funcs["mysql"], 'connection') or app_funcs["mysql"].connection is None:
        return jsonify({'success': False, 'message': 'Koneksi ke database gagal'})
    
    try:
        cursor = app_funcs["mysql"].connection.cursor(MySQLdb.cursors.DictCursor)
        
        biaya_id = request.form.get('biaya_id')
        
        if not biaya_id:
            return jsonify({'success': False, 'message': 'ID biaya tidak boleh kosong'})
        
        try:
            biaya_id = int(biaya_id)
        except (ValueError, TypeError):
            return jsonify({'success': False, 'message': 'ID biaya tidak valid'})
        
        # Cek apakah biaya ada dan milik mahasiswa yang login
        cursor.execute('''
            SELECT bo.* FROM biaya_operasional bo 
            JOIN proposal p ON bo.proposal_id = p.id 
            WHERE bo.id = %s AND p.nim = %s
        ''', (biaya_id, session['nim']))
        
        biaya_exists = cursor.fetchone()
        if not biaya_exists:
            return jsonify({'success': False, 'message': 'Biaya operasional tidak ditemukan atau bukan milik Anda'})
        
        # Hapus data
        cursor.execute('DELETE FROM biaya_operasional WHERE id = %s', (biaya_id,))
        
        app_funcs["mysql"].connection.commit()
        
        # Log aktivitas
        mahasiswa_info = app_funcs['get_mahasiswa_info_from_session']()
        if mahasiswa_info and biaya_exists:
            data_lama = dict(biaya_exists)
            app_funcs['log_mahasiswa_activity'](
                mahasiswa_info['id'], 
                mahasiswa_info['nim'], 
                mahasiswa_info['nama_ketua'], 
                'hapus', 
                'biaya_operasional', 
                f'id_{biaya_id}', 
                f'Menghapus biaya operasional: {data_lama.get("nama_biaya", "Unknown")}',
                data_lama,
                None
            )
        
        # Hitung dan simpan data laba rugi
        app_funcs['calculate_and_save_laba_rugi'](biaya_exists['proposal_id'])
        
        # Update arus kas otomatis untuk bulan pembelian yang dihapus
        if biaya_exists.get('tanggal_beli'):
            bulan_pembelian = biaya_exists['tanggal_beli'].strftime('%Y-%m')
            app_funcs['update_arus_kas_otomatis'](biaya_exists['proposal_id'], bulan_pembelian)
        
        cursor.close()
        
        return jsonify({'success': True, 'message': 'Biaya operasional berhasil dihapus'})
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@mahasiswa_bp.route('/laporan_penjualan_mahasiswa')
def laporan_penjualan_mahasiswa():
    app_funcs = get_app_functions()
    if 'user_type' not in session or session['user_type'] != 'mahasiswa':
        flash('Anda harus login sebagai mahasiswa!', 'danger')
        return redirect(url_for('index'))
    
    if not hasattr(app_funcs["mysql"], 'connection') or app_funcs["mysql"].connection is None:
        flash('Koneksi ke database gagal. Cek konfigurasi database!', 'danger')
        return render_template('mahasiswa/laporan_penjualan.html', proposal_data={})
    
    try:
        cursor = app_funcs["mysql"].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Ambil data proposal yang lolos untuk mahasiswa ini
        cursor.execute('''
            SELECT p.id, p.judul_usaha, p.kategori, p.tahun_nib, p.status_admin
            FROM proposal p
            WHERE p.nim = %s AND p.status_admin = 'lolos'
            ORDER BY p.tanggal_buat DESC
        ''', (session['nim'],))
        
        proposals = cursor.fetchall()
        proposal_data = {}
        
        for proposal in proposals:
            # Ambil data penjualan untuk proposal ini
            cursor.execute('''
                SELECT id, tanggal_penjualan, nama_produk, quantity, harga_jual, total
                FROM penjualan
                WHERE proposal_id = %s
                ORDER BY tanggal_penjualan DESC
            ''', (proposal['id'],))
            
            penjualan_data = cursor.fetchall()
            
            proposal_data[proposal['id']] = {
                'proposal': proposal,
                'penjualan': penjualan_data
            }
        
        # Ambil status mahasiswa
        cursor.execute('SELECT status FROM mahasiswa WHERE nim = %s', (session['nim'],))
        mahasiswa_status = cursor.fetchone()
        status_mahasiswa = mahasiswa_status['status'] if mahasiswa_status else 'proses'
        
        cursor.close()
        has_lolos_proposal = len(proposal_data) > 0
        return render_template('mahasiswa/laporan_penjualan.html', proposal_data=proposal_data, status_mahasiswa=status_mahasiswa, has_lolos_proposal=has_lolos_proposal)
        
    except Exception as e:
        flash(f'Error: {str(e)}', 'danger')
        return render_template('mahasiswa/laporan_penjualan.html', proposal_data={}, status_mahasiswa='proses', has_lolos_proposal=False)

@mahasiswa_bp.route('/get_produk_by_proposal/<int:proposal_id>')
def get_produk_by_proposal(proposal_id):
    app_funcs = get_app_functions()
    if 'user_type' not in session or session['user_type'] != 'mahasiswa':
        return jsonify({'success': False, 'message': 'Anda harus login sebagai mahasiswa'})
    if not hasattr(app_funcs["mysql"], 'connection') or app_funcs["mysql"].connection is None:
        return jsonify({'success': False, 'message': 'Koneksi ke database gagal'})
    try:
        cursor = app_funcs["mysql"].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Validasi proposal milik mahasiswa ini
        cursor.execute('SELECT id FROM proposal WHERE id = %s AND nim = %s AND status_admin = "lolos"', (proposal_id, session['nim']))
        if not cursor.fetchone():
            cursor.close()
            return jsonify({'success': False, 'message': 'Proposal tidak ditemukan atau tidak memiliki akses'})
        
        cursor.execute('SELECT id, nama_produk, harga_jual, jumlah_produk, tanggal_produksi FROM produksi WHERE proposal_id = %s ORDER BY tanggal_produksi DESC', (proposal_id,))
        produk_list = cursor.fetchall()
        cursor.close()
        
        return jsonify({'success': True, 'data': produk_list})
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@mahasiswa_bp.route('/get_penjualan/<int:penjualan_id>')
def get_penjualan(penjualan_id):
    app_funcs = get_app_functions()
    if 'user_type' not in session or session['user_type'] != 'mahasiswa':
        return jsonify({'success': False, 'message': 'Anda harus login sebagai mahasiswa!'})
    
    if not hasattr(app_funcs["mysql"], 'connection') or app_funcs["mysql"].connection is None:
        return jsonify({'success': False, 'message': 'Koneksi ke database gagal!'})
    
    try:
        cursor = app_funcs["mysql"].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Ambil data penjualan dengan validasi kepemilikan
        cursor.execute('''
            SELECT p.* FROM penjualan p
            JOIN proposal pr ON p.proposal_id = pr.id
            WHERE p.id = %s AND pr.nim = %s
        ''', (penjualan_id, session['nim']))
        
        penjualan = cursor.fetchone()
        cursor.close()
        
        if not penjualan:
            return jsonify({'success': False, 'message': 'Data penjualan tidak ditemukan!'})
        
        return jsonify({'success': True, 'data': penjualan})
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})
# Tambah penjualan: laporan penjualan untuk pencatatan
@mahasiswa_bp.route('/tambah_penjualan', methods=['POST'])
def tambah_penjualan():
    app_funcs = get_app_functions()
    if 'user_type' not in session or session['user_type'] != 'mahasiswa':
        return jsonify({'success': False, 'message': 'Anda harus login sebagai mahasiswa!'})
    if not hasattr(app_funcs["mysql"], 'connection') or app_funcs["mysql"].connection is None:
        return jsonify({'success': False, 'message': 'Koneksi ke database gagal!'})
    try:
        proposal_id = request.form.get('proposal_id')
        tanggal_penjualan = request.form.get('tanggal_penjualan')
        nama_produk = request.form.get('nama_produk')
        quantity = request.form.get('quantity')
        harga_jual = request.form.get('harga_jual')
        if not all([proposal_id, tanggal_penjualan, nama_produk, quantity, harga_jual]):
            return jsonify({'success': False, 'message': 'Semua field harus diisi!'})
        try:
            quantity = int(quantity) if quantity else 0
            harga_jual = float(harga_jual) if harga_jual else 0
        except ValueError:
            return jsonify({'success': False, 'message': 'Quantity harus berupa angka bulat dan harga jual harus berupa angka!'})
        if quantity <= 0 or harga_jual <= 0:
            return jsonify({'success': False, 'message': 'Quantity dan harga jual harus lebih dari 0!'})
        # Hitung total
        total = quantity * harga_jual
        cursor = app_funcs["mysql"].connection.cursor(MySQLdb.cursors.DictCursor)
        # Validasi proposal milik mahasiswa ini
        cursor.execute('''SELECT id FROM proposal WHERE id = %s AND nim = %s AND status_admin = 'lolos' ''', (proposal_id, session['nim']))
        if not cursor.fetchone():
            cursor.close()
            return jsonify({'success': False, 'message': 'Proposal tidak ditemukan atau tidak memiliki akses!'})
        # Ambil harga jual terbaru dari produksi
        cursor.execute('''SELECT harga_jual FROM produksi WHERE proposal_id = %s AND nama_produk = %s''', (proposal_id, nama_produk))
        prod = cursor.fetchone()
        if not prod:
            cursor.close()
            return jsonify({'success': False, 'message': 'Produk tidak ditemukan di produksi!'})
        
        # Gunakan harga jual terbaru dari produksi
        harga_jual_terbaru = float(prod['harga_jual'])
        total_terbaru = quantity * harga_jual_terbaru
        
        # Insert data penjualan dengan harga jual terbaru
        cursor.execute('''INSERT INTO penjualan (proposal_id, tanggal_penjualan, nama_produk, quantity, harga_jual, total) VALUES (%s, %s, %s, %s, %s, %s)''', (proposal_id, tanggal_penjualan, nama_produk, quantity, harga_jual_terbaru, total_terbaru))
        # Tidak mengurangi stok produksi - laporan penjualan hanya untuk pencatatan
        
        penjualan_id = cursor.lastrowid
        app_funcs["mysql"].connection.commit()
        
        # Log aktivitas
        mahasiswa_info = app_funcs['get_mahasiswa_info_from_session']()
        if mahasiswa_info:
            data_baru = {
                'id': penjualan_id,
                'nama_produk': nama_produk,
                'quantity': quantity,
                'harga_jual': harga_jual_terbaru,
                'total': total_terbaru,
                'tanggal_penjualan': tanggal_penjualan
            }
            app_funcs['log_mahasiswa_activity'](
                mahasiswa_info['id'], 
                mahasiswa_info['nim'], 
                mahasiswa_info['nama_ketua'], 
                'tambah', 
                'penjualan', 
                f'id_{penjualan_id}', 
                f'Menambahkan penjualan: {nama_produk}',
                None,
                data_baru
            )
        
        # Hitung dan simpan data laba rugi
        app_funcs['calculate_and_save_laba_rugi'](proposal_id)
        
        # Update arus kas otomatis untuk bulan penjualan
        if tanggal_penjualan:
            bulan_penjualan = datetime.strptime(tanggal_penjualan, '%Y-%m-%d').strftime('%Y-%m')
            app_funcs['update_arus_kas_otomatis'](proposal_id, bulan_penjualan)
        
        cursor.close()
        return jsonify({'success': True, 'message': 'Data penjualan berhasil ditambahkan!'})
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

# Hapus penjualan: hapus data pencatatan
@mahasiswa_bp.route('/delete_penjualan', methods=['POST'])
def delete_penjualan():
    app_funcs = get_app_functions()
    if 'user_type' not in session or session['user_type'] != 'mahasiswa':
        return jsonify({'success': False, 'message': 'Anda harus login sebagai mahasiswa!'})
    if not hasattr(app_funcs["mysql"], 'connection') or app_funcs["mysql"].connection is None:
        return jsonify({'success': False, 'message': 'Koneksi ke database gagal!'})
    try:
        penjualan_id = request.form.get('penjualan_id')
        cursor = app_funcs["mysql"].connection.cursor(MySQLdb.cursors.DictCursor)
        # Ambil data penjualan dengan validasi kepemilikan
        cursor.execute('''
            SELECT p.proposal_id, p.nama_produk, p.quantity 
            FROM penjualan p
            JOIN proposal pr ON p.proposal_id = pr.id
            WHERE p.id = %s AND pr.nim = %s
        ''', (penjualan_id, session['nim']))
        penjualan = cursor.fetchone()
        if not penjualan:
            cursor.close()
            return jsonify({'success': False, 'message': 'Data penjualan tidak ditemukan atau tidak memiliki akses!'})
        # Tidak perlu mengembalikan stok produksi karena tidak pernah dikurangi
        # Hapus data penjualan
        cursor.execute('''DELETE FROM penjualan WHERE id = %s''', (penjualan_id,))
        app_funcs["mysql"].connection.commit()
        
        # Log aktivitas
        mahasiswa_info = app_funcs['get_mahasiswa_info_from_session']()
        if mahasiswa_info and penjualan:
            data_lama = dict(penjualan)
            app_funcs['log_mahasiswa_activity'](
                mahasiswa_info['id'], 
                mahasiswa_info['nim'], 
                mahasiswa_info['nama_ketua'], 
                'hapus', 
                'penjualan', 
                f'id_{penjualan_id}', 
                f'Menghapus penjualan: {data_lama.get("nama_produk", "Unknown")}',
                data_lama,
                None
            )
        
        # Hitung dan simpan data laba rugi
        app_funcs['calculate_and_save_laba_rugi'](penjualan['proposal_id'])
        
        # Update arus kas otomatis untuk semua bulan yang terpengaruh
        # Ambil semua bulan yang memiliki data penjualan untuk proposal ini
        cursor.execute('''
            SELECT DISTINCT DATE_FORMAT(tanggal_penjualan, '%%Y-%%m') as bulan
            FROM penjualan 
            WHERE proposal_id = %s
        ''', (penjualan['proposal_id'],))
        bulan_list = cursor.fetchall()
        
        for bulan_data in bulan_list:
            app_funcs['update_arus_kas_otomatis'](penjualan['proposal_id'], bulan_data['bulan'])
        
        cursor.close()
        return jsonify({'success': True, 'message': 'Data penjualan berhasil dihapus!'})
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

# Edit penjualan: update data pencatatan
@mahasiswa_bp.route('/update_penjualan', methods=['POST'])
def update_penjualan():
    app_funcs = get_app_functions()
    if 'user_type' not in session or session['user_type'] != 'mahasiswa':
        return jsonify({'success': False, 'message': 'Anda harus login sebagai mahasiswa!'})
    if not hasattr(app_funcs["mysql"], 'connection') or app_funcs["mysql"].connection is None:
        return jsonify({'success': False, 'message': 'Koneksi ke database gagal!'})
    try:
        penjualan_id = request.form.get('penjualan_id')
        proposal_id = request.form.get('proposal_id')
        tanggal_penjualan = request.form.get('tanggal_penjualan')
        nama_produk = request.form.get('nama_produk')
        quantity = request.form.get('quantity')
        harga_jual = request.form.get('harga_jual')
        if not all([penjualan_id, proposal_id, tanggal_penjualan, nama_produk, quantity, harga_jual]):
            return jsonify({'success': False, 'message': 'Semua field harus diisi!'})
        try:
            quantity = int(quantity) if quantity else 0
            harga_jual = float(harga_jual) if harga_jual else 0
        except ValueError:
            return jsonify({'success': False, 'message': 'Quantity harus berupa angka bulat dan harga jual harus berupa angka!'})
        if quantity <= 0 or harga_jual <= 0:
            return jsonify({'success': False, 'message': 'Quantity dan harga jual harus lebih dari 0!'})
        cursor = app_funcs["mysql"].connection.cursor(MySQLdb.cursors.DictCursor)
        # Ambil data penjualan lama dengan validasi kepemilikan
        cursor.execute('''
            SELECT p.quantity, p.proposal_id 
            FROM penjualan p
            JOIN proposal pr ON p.proposal_id = pr.id
            WHERE p.id = %s AND pr.nim = %s
        ''', (penjualan_id, session['nim']))
        penjualan_lama = cursor.fetchone()
        if not penjualan_lama:
            cursor.close()
            return jsonify({'success': False, 'message': 'Data penjualan tidak ditemukan atau tidak memiliki akses!'})
        quantity_lama = int(penjualan_lama['quantity'])
        selisih = quantity - quantity_lama
        # Tidak perlu validasi stok karena laporan penjualan tidak mengurangi stok produksi
        # Ambil harga jual terbaru dari produksi
        cursor.execute('''SELECT harga_jual FROM produksi WHERE proposal_id = %s AND nama_produk = %s''', (penjualan_lama['proposal_id'], nama_produk))
        prod = cursor.fetchone()
        if not prod:
            cursor.close()
            return jsonify({'success': False, 'message': 'Produk tidak ditemukan di produksi!'})
        
        # Gunakan harga jual terbaru dari produksi
        harga_jual_terbaru = float(prod['harga_jual'])
        total_terbaru = quantity * harga_jual_terbaru
        
        # Update data penjualan dengan harga jual terbaru
        cursor.execute('''UPDATE penjualan SET tanggal_penjualan = %s, nama_produk = %s, quantity = %s, harga_jual = %s, total = %s WHERE id = %s''', (tanggal_penjualan, nama_produk, quantity, harga_jual_terbaru, total_terbaru, penjualan_id))
        # Tidak mengupdate stok produksi - laporan penjualan hanya untuk pencatatan
        app_funcs["mysql"].connection.commit()
        
        # Log aktivitas
        mahasiswa_info = app_funcs['get_mahasiswa_info_from_session']()
        if mahasiswa_info and penjualan_id:
            data_baru = {
                'id': int(penjualan_id),
                'nama_produk': nama_produk,
                'quantity': quantity,
                'harga_jual': harga_jual_terbaru,
                'total': total_terbaru,
                'tanggal_penjualan': tanggal_penjualan
            }
            app_funcs['log_mahasiswa_activity'](
                mahasiswa_info['id'], 
                mahasiswa_info['nim'], 
                mahasiswa_info['nama_ketua'], 
                'edit', 
                'penjualan', 
                f'id_{penjualan_id}', 
                f'Mengubah penjualan: {nama_produk}',
                None,  # Data lama tidak diambil untuk performa
                data_baru
            )
        
        # Hitung dan simpan data laba rugi
        app_funcs['calculate_and_save_laba_rugi'](penjualan_lama['proposal_id'])
        
        # Update arus kas otomatis untuk bulan penjualan
        if tanggal_penjualan:
            bulan_penjualan = datetime.strptime(tanggal_penjualan, '%Y-%m-%d').strftime('%Y-%m')
            app_funcs['update_arus_kas_otomatis'](penjualan_lama['proposal_id'], bulan_penjualan)
        
        cursor.close()
        return jsonify({'success': True, 'message': 'Data penjualan berhasil diupdate!'})
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})


@mahasiswa_bp.route('/laporan_laba_rugi_mahasiswa')
def laporan_laba_rugi_mahasiswa():
    app_funcs = get_app_functions()
    if 'user_type' not in session or session['user_type'] != 'mahasiswa':
        flash('Anda harus login sebagai mahasiswa!', 'danger')
        return redirect(url_for('index'))
    
    if not hasattr(app_funcs["mysql"], 'connection') or app_funcs["mysql"].connection is None:
        flash('Koneksi ke database gagal. Cek konfigurasi database!', 'danger')
        return render_template('mahasiswa/laporan_laba_rugi.html', proposal_data={})
    
    try:
        cursor = app_funcs["mysql"].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Ambil data proposal yang lolos untuk mahasiswa ini
        cursor.execute('''
            SELECT p.id, p.judul_usaha, p.kategori, p.tahun_nib, p.status_admin
            FROM proposal p
            WHERE p.nim = %s AND p.status_admin = 'lolos'
            ORDER BY p.tanggal_buat DESC
        ''', (session['nim'],))
        
        proposals = cursor.fetchall()
        proposal_data = {}
        
        for proposal in proposals:
            proposal_id = proposal['id']
            
            # Ambil data laba rugi dari tabel laba_rugi
            cursor.execute('''
                SELECT tanggal_produksi, nama_produk, pendapatan, total_biaya_produksi, 
                       laba_rugi_kotor, biaya_operasional, laba_rugi_bersih
                FROM laba_rugi
                WHERE proposal_id = %s
                ORDER BY tanggal_produksi DESC
            ''', (proposal_id,))
            
            laba_rugi_data = cursor.fetchall()
            
            # DEBUG: Print data untuk memastikan data yang benar
            print(f"=== DEBUG LABA RUGI MAHASISWA ===")
            print(f"Proposal ID: {proposal_id}")
            print(f"Data count: {len(laba_rugi_data)}")
            for i, item in enumerate(laba_rugi_data):
                print(f"Row {i+1}:")
                print(f"  Pendapatan: {item['pendapatan']}")
                print(f"  Total Biaya Produksi: {item['total_biaya_produksi']}")
                print(f"  Laba Rugi Kotor: {item['laba_rugi_kotor']}")
                print(f"  Biaya Operasional: {item['biaya_operasional']}")
                print(f"  Laba Rugi Bersih: {item['laba_rugi_bersih']}")
            print("==================================")
            
            # Hitung total dari data laba rugi
            total_pendapatan = sum(item['pendapatan'] for item in laba_rugi_data)
            total_biaya_produksi = sum(item['total_biaya_produksi'] for item in laba_rugi_data)
            total_biaya_operasional = sum(item['biaya_operasional'] for item in laba_rugi_data)
            total_laba_kotor = sum(item['laba_rugi_kotor'] for item in laba_rugi_data)
            total_laba_bersih = sum(item['laba_rugi_bersih'] for item in laba_rugi_data)
            
            # DEBUG: Print total yang dihitung
            print(f"=== TOTALS CALCULATED ===")
            print(f"Total Pendapatan: {total_pendapatan}")
            print(f"Total Biaya Produksi: {total_biaya_produksi}")
            print(f"Total Biaya Operasional: {total_biaya_operasional}")
            print(f"Total Laba Kotor: {total_laba_kotor}")
            print(f"Total Laba Bersih: {total_laba_bersih}")
            print("=========================")
            
            proposal_data[proposal_id] = {
                'proposal': proposal,
                'laba_rugi_data': laba_rugi_data,
                'total_pendapatan': total_pendapatan,
                'total_biaya_produksi': total_biaya_produksi,
                'total_biaya_operasional': total_biaya_operasional,
                'total_laba_kotor': total_laba_kotor,
                'total_laba_bersih': total_laba_bersih
            }
        
        cursor.close()
        return render_template('mahasiswa/laporan_laba_rugi.html', proposal_data=proposal_data)
        
    except Exception as e:
        flash(f'Error: {str(e)}', 'danger')
        return render_template('mahasiswa/laporan_laba_rugi.html', proposal_data={})

@mahasiswa_bp.route('/download_laba_rugi', methods=['POST'])
def download_laba_rugi():
    logger.info("Download laba rugi mahasiswa dipanggil")
    app_funcs = get_app_functions()
    if 'user_type' not in session or session['user_type'] != 'mahasiswa':
        logger.warning(f"Akses ditolak untuk download laba rugi: {session.get('user_type', 'None')}")
        return jsonify({'success': False, 'message': 'Unauthorized'})
    
    try:
        proposal_id = request.form.get('proposal_id')
        format_type = request.form.get('format')
        start_date = request.form.get('start_date')
        end_date = request.form.get('end_date')
        
        if not all([proposal_id, format_type, start_date, end_date]):
            return jsonify({'success': False, 'message': 'Parameter tidak lengkap'})
        
        if not hasattr(app_funcs["mysql"], 'connection') or app_funcs["mysql"].connection is None:
            return jsonify({'success': False, 'message': 'Koneksi database tidak tersedia'})
        
        cursor = app_funcs["mysql"].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Ambil data proposal dan informasi mahasiswa
        cursor.execute('''
            SELECT p.judul_usaha, p.kategori, p.tahun_nib, m.id as mahasiswa_id, m.nama_ketua, m.nim
            FROM proposal p
            JOIN mahasiswa m ON p.nim = m.nim
            WHERE p.id = %s AND p.nim = %s
        ''', (proposal_id, session['nim']))
        
        proposal = cursor.fetchone()
        if not proposal:
            cursor.close()
            return jsonify({'success': False, 'message': 'Proposal tidak ditemukan'})
        
        # Ambil data laba rugi sesuai rentang tanggal
        cursor.execute('''
            SELECT tanggal_produksi, nama_produk, pendapatan, total_biaya_produksi, 
                   laba_rugi_kotor, biaya_operasional, laba_rugi_bersih
            FROM laba_rugi
            WHERE proposal_id = %s AND tanggal_produksi BETWEEN %s AND %s
            ORDER BY tanggal_produksi DESC
        ''', (proposal_id, start_date, end_date))
        
        laba_rugi_data = cursor.fetchall()
        
        # DEBUG: Log data untuk memastikan tidak kosong
        logger.debug(f"=== DEBUG DOWNLOAD LABA RUGI ===")
        logger.debug(f"Proposal ID: {proposal_id}")
        logger.debug(f"Start Date: {start_date}")
        logger.debug(f"End Date: {end_date}")
        logger.debug(f"Data count: {len(laba_rugi_data)}")
        logger.debug(f"Data sample: {laba_rugi_data[:2] if laba_rugi_data else 'KOSONG'}")
        logger.debug(f"Proposal data: {proposal}")
        logger.debug("==================================")
        
        if not laba_rugi_data:
            cursor.close()
            return jsonify({'success': False, 'message': 'Tidak ada data laba rugi untuk periode yang dipilih'})
        
        # Hitung total
        total_pendapatan = sum(item['pendapatan'] for item in laba_rugi_data)
        total_biaya_produksi = sum(item['total_biaya_produksi'] for item in laba_rugi_data)
        total_biaya_operasional = sum(item['biaya_operasional'] for item in laba_rugi_data)
        total_laba_kotor = sum(item['laba_rugi_kotor'] for item in laba_rugi_data)
        total_laba_bersih = sum(item['laba_rugi_bersih'] for item in laba_rugi_data)
        
        # DEBUG: Log totals
        logger.debug(f"=== TOTALS ===")
        logger.debug(f"Total Pendapatan: {total_pendapatan}")
        logger.debug(f"Total Biaya Produksi: {total_biaya_produksi}")
        logger.debug(f"Total Biaya Operasional: {total_biaya_operasional}")
        logger.debug(f"Total Laba Kotor: {total_laba_kotor}")
        logger.debug(f"Total Laba Bersih: {total_laba_bersih}")
        logger.debug("===============")
        
        cursor.close()
        
        # Log aktivitas download
        mahasiswa_info = {
            'id': proposal['mahasiswa_id'],
            'nim': proposal['nim'],
            'nama_ketua': proposal['nama_ketua']
        }
        periode = f'periode {start_date} sampai {end_date}'
        log_mahasiswa_download_activity(app_funcs, mahasiswa_info, 'laporan_laba_rugi', proposal_id, format_type, periode)
        
        # Generate filename dengan pembersihan karakter
        import re
        safe_judul = re.sub(r'[<>:"/\\|?*]', '_', proposal['judul_usaha'])
        filename = f"Laporan_Laba_Rugi_{safe_judul.replace(' ', '_')}_{start_date}_to_{end_date}"
        
        if format_type == 'excel':
            return app_funcs['generate_excel_laba_rugi'](laba_rugi_data, proposal, total_pendapatan, total_biaya_produksi, 
                                           total_biaya_operasional, total_laba_kotor, total_laba_bersih, 
                                           start_date, end_date, filename)
        elif format_type == 'pdf':
            return app_funcs['generate_pdf_laba_rugi'](laba_rugi_data, proposal, total_pendapatan, total_biaya_produksi, 
                                         total_biaya_operasional, total_laba_kotor, total_laba_bersih, 
                                         start_date, end_date, filename)
        elif format_type == 'word':
            return app_funcs['generate_word_laba_rugi'](laba_rugi_data, proposal, total_pendapatan, total_biaya_produksi, 
                                          total_biaya_operasional, total_laba_kotor, total_laba_bersih, 
                                          start_date, end_date, filename)
        else:
            return jsonify({'success': False, 'message': 'Format tidak didukung'})
            
    except Exception as e:
        logger.error('--- ERROR DOWNLOAD LABA RUGI ---')
        logger.error(traceback.format_exc())
        return jsonify({'success': False, 'message': f'Error: {str(e)}', 'trace': traceback.format_exc()})
@mahasiswa_bp.route('/download_arus_kas', methods=['POST'])
def download_arus_kas():
    app_funcs = get_app_functions()
    import traceback
    if 'user_type' not in session or session['user_type'] != 'mahasiswa':
        return jsonify({'success': False, 'message': 'Unauthorized'})
    
    try:
        proposal_id = request.form.get('proposal_id')
        format_type = request.form.get('format')
        bulan_tahun = request.form.get('bulan_tahun')
        
        if not all([proposal_id, format_type, bulan_tahun]):
            return jsonify({'success': False, 'message': 'Parameter tidak lengkap'})
        
        if not hasattr(app_funcs["mysql"], 'connection') or app_funcs["mysql"].connection is None:
            return jsonify({'success': False, 'message': 'Koneksi database tidak tersedia'})
        
        cursor = app_funcs["mysql"].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Ambil data proposal dan informasi mahasiswa
        cursor.execute('''
            SELECT p.judul_usaha, p.kategori, p.tahun_nib, m.id as mahasiswa_id, m.nama_ketua, m.nim
            FROM proposal p
            JOIN mahasiswa m ON p.nim = m.nim
            WHERE p.id = %s AND p.nim = %s
        ''', (proposal_id, session['nim']))
        
        proposal = cursor.fetchone()
        if not proposal:
            cursor.close()
            return jsonify({'success': False, 'message': 'Proposal tidak ditemukan'})
        
        # Parse bulan yang dipilih
        if not bulan_tahun:
            return jsonify({'success': False, 'message': 'Bulan tahun tidak boleh kosong'})
        
        try:
            selected_date = datetime.strptime(bulan_tahun, '%Y-%m')
        except ValueError:
            return jsonify({'success': False, 'message': 'Format bulan tahun tidak valid. Gunakan format YYYY-MM'})
        start_date = selected_date.replace(day=1)
        end_date = (selected_date.replace(day=1) + timedelta(days=32)).replace(day=1) - timedelta(days=1)
        
        # Ambil data arus kas untuk bulan yang dipilih
        cursor.execute('''
            SELECT bulan_tahun, total_penjualan, total_biaya_produksi, total_biaya_operasional,
                   total_biaya_non_operasional, kas_bersih_operasional, total_harga_jual_alat,
                   total_harga_alat, kas_bersih_investasi, kas_bersih_pembiayaan, total_kas_bersih
            FROM arus_kas
            WHERE proposal_id = %s AND bulan_tahun = %s
        ''', (proposal_id, bulan_tahun))
        
        arus_kas_data = cursor.fetchone()
        
        # DEBUG: Print data untuk memastikan tidak kosong
        print(f"=== DEBUG DOWNLOAD ARUS KAS ===")
        print(f"Proposal ID: {proposal_id}")
        print(f"Bulan Tahun: {bulan_tahun}")
        print(f"Data: {arus_kas_data}")
        print(f"Proposal data: {proposal}")
        print("==================================")
        
        if not arus_kas_data:
            cursor.close()
            return jsonify({'success': False, 'message': 'Tidak ada data arus kas untuk bulan yang dipilih'})
        
        cursor.close()
        
        # Log aktivitas download
        mahasiswa_info = {
            'id': proposal['mahasiswa_id'],
            'nim': proposal['nim'],
            'nama_ketua': proposal['nama_ketua']
        }
        periode = f'bulan {bulan_tahun}'
        log_mahasiswa_download_activity(app_funcs, mahasiswa_info, 'laporan_arus_kas', proposal_id, format_type, periode)
        
        # Generate filename dengan pembersihan karakter
        import re
        nama_bulan = selected_date.strftime('%B %Y')
        safe_judul = re.sub(r'[<>:"/\\|?*]', '_', proposal['judul_usaha'])
        filename = f"Laporan_Arus_Kas_{safe_judul.replace(' ', '_')}_{bulan_tahun}"
        
        if format_type == 'excel':
            return app_funcs['generate_excel_arus_kas'](arus_kas_data, proposal, bulan_tahun, nama_bulan, filename)
        elif format_type == 'pdf':
            return app_funcs['generate_pdf_arus_kas'](arus_kas_data, proposal, bulan_tahun, nama_bulan, filename)
        elif format_type == 'word':
            return app_funcs['generate_word_arus_kas'](arus_kas_data, proposal, bulan_tahun, nama_bulan, filename)
        else:
            return jsonify({'success': False, 'message': 'Format tidak didukung'})
            
    except Exception as e:
        print('--- ERROR DOWNLOAD ARUS KAS ---')
        print(traceback.format_exc())
        return jsonify({'success': False, 'message': f'Error: {str(e)}', 'trace': traceback.format_exc()})

@mahasiswa_bp.route('/laporan_arus_kas_mahasiswa')
def laporan_arus_kas_mahasiswa():
    app_funcs = get_app_functions()
    if 'user_type' not in session or session['user_type'] != 'mahasiswa':
        flash('Anda harus login sebagai mahasiswa!', 'danger')
        return redirect(url_for('index'))
    
    if not hasattr(app_funcs["mysql"], 'connection') or app_funcs["mysql"].connection is None:
        flash('Koneksi ke database gagal. Cek konfigurasi database!', 'danger')
        return render_template('mahasiswa/arus_kas.html', proposal_data={})
    
    try:
        cursor = app_funcs["mysql"].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Ambil data proposal yang lolos untuk mahasiswa ini
        cursor.execute('''
            SELECT p.id, p.judul_usaha, p.kategori, p.tahun_nib, p.status_admin
            FROM proposal p
            WHERE p.nim = %s AND p.status_admin = 'lolos'
            ORDER BY p.tanggal_buat DESC
        ''', (session['nim'],))
        
        proposals = cursor.fetchall()
        
        # Ambil bulan yang dipilih dari parameter URL
        selected_month = request.args.get('month', datetime.now().strftime('%Y-%m'))
        
        proposal_data = {}
        
        for proposal in proposals:
            proposal_id = proposal['id']
            
            # Parse bulan yang dipilih - Data arus kas per bulan (bukan kumulatif)
            try:
                selected_date = datetime.strptime(selected_month, '%Y-%m')
                # Hitung periode bulan yang dipilih (1-31, 1-28, dst)
                start_date = selected_date.replace(day=1)  # Tanggal 1 bulan yang dipilih
                end_date = (selected_date.replace(day=1) + timedelta(days=32)).replace(day=1) - timedelta(days=1)  # Tanggal terakhir bulan yang dipilih
            except:
                selected_date = datetime.now()
                start_date = selected_date.replace(day=1)
                end_date = (selected_date.replace(day=1) + timedelta(days=32)).replace(day=1) - timedelta(days=1)
            
            # 1. Total Penerimaan Kas dari Penjualan (PER BULAN YANG DIPILIH)
            cursor.execute("""
                SELECT COALESCE(SUM(total), 0) as total_penjualan
                FROM penjualan 
                WHERE proposal_id = %s 
                AND DATE(tanggal_penjualan) BETWEEN %s AND %s
            """, (proposal_id, start_date.date(), end_date.date()))
            total_penjualan = cursor.fetchone()['total_penjualan']
            
            # 2. Total Biaya Produksi (Supplier) - PER BULAN YANG DIPILIH
            cursor.execute("""
                SELECT COALESCE(SUM(total_biaya), 0) as total_biaya_produksi
                FROM produksi 
                WHERE proposal_id = %s 
                AND DATE(tanggal_produksi) BETWEEN %s AND %s
            """, (proposal_id, start_date.date(), end_date.date()))
            total_biaya_produksi = cursor.fetchone()['total_biaya_produksi']
            
            # 3. Total Biaya Operasional - PER BULAN YANG DIPILIH
            cursor.execute("""
                SELECT COALESCE(SUM(total_harga), 0) as total_biaya_operasional
                FROM biaya_operasional 
                WHERE proposal_id = %s 
                AND DATE(tanggal_beli) BETWEEN %s AND %s
            """, (proposal_id, start_date.date(), end_date.date()))
            total_biaya_operasional = cursor.fetchone()['total_biaya_operasional']
            
            # 4. Total Biaya Lain-lain - PER BULAN YANG DIPILIH
            cursor.execute("""
                SELECT COALESCE(SUM(total_harga), 0) as total_biaya_non_operasional
                FROM biaya_non_operasional 
                WHERE proposal_id = %s 
                AND DATE(tanggal_transaksi) BETWEEN %s AND %s
            """, (proposal_id, start_date.date(), end_date.date()))
            total_biaya_non_operasional = cursor.fetchone()['total_biaya_non_operasional']
            
            # 5. Total Harga Jual Alat Produksi - PER BULAN YANG DIPILIH
            cursor.execute("""
                SELECT COALESCE(SUM(harga_jual), 0) as total_harga_jual_alat
                FROM alat_produksi 
                WHERE proposal_id = %s 
                AND DATE(tanggal_beli) BETWEEN %s AND %s
                AND harga_jual > 0
            """, (proposal_id, start_date.date(), end_date.date()))
            total_harga_jual_alat = cursor.fetchone()['total_harga_jual_alat']
            
            # 6. Total Harga Alat Produksi (Pembelian) - PER BULAN YANG DIPILIH
            cursor.execute("""
                SELECT COALESCE(SUM(total_alat_produksi), 0) as total_harga_alat
                FROM alat_produksi 
                WHERE proposal_id = %s 
                AND DATE(tanggal_beli) BETWEEN %s AND %s
            """, (proposal_id, start_date.date(), end_date.date()))
            total_harga_alat = cursor.fetchone()['total_harga_alat']
            
            # Perhitungan Kas Bersih - PER BULAN YANG DIPILIH (BUKAN KUMULATIF)
            kas_bersih_operasional = total_penjualan - total_biaya_produksi - total_biaya_operasional - total_biaya_non_operasional
            kas_bersih_investasi = total_harga_jual_alat - total_harga_alat
            kas_bersih_pembiayaan = 0  # Selalu 0 sesuai permintaan
            total_kas_bersih = kas_bersih_operasional + kas_bersih_investasi + kas_bersih_pembiayaan
            
            # Simpan atau update data arus kas ke database - PER BULAN TERPISAH (OTOMATIS UPDATE)
            cursor.execute("""
                INSERT INTO arus_kas (
                    proposal_id, 
                    bulan_tahun,
                    total_penjualan,
                    total_biaya_produksi,
                    total_biaya_operasional,
                    total_biaya_non_operasional,
                    kas_bersih_operasional,
                    total_harga_jual_alat,
                    total_harga_alat,
                    kas_bersih_investasi,
                    kas_bersih_pembiayaan,
                    total_kas_bersih,
                    created_at,
                    updated_at
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, NOW(), NOW())
                ON DUPLICATE KEY UPDATE
                    total_penjualan = VALUES(total_penjualan),
                    total_biaya_produksi = VALUES(total_biaya_produksi),
                    total_biaya_operasional = VALUES(total_biaya_operasional),
                    total_biaya_non_operasional = VALUES(total_biaya_non_operasional),
                    kas_bersih_operasional = VALUES(kas_bersih_operasional),
                    total_harga_jual_alat = VALUES(total_harga_jual_alat),
                    total_harga_alat = VALUES(total_harga_alat),
                    kas_bersih_investasi = VALUES(kas_bersih_investasi),
                    kas_bersih_pembiayaan = VALUES(kas_bersih_pembiayaan),
                    total_kas_bersih = VALUES(total_kas_bersih),
                    updated_at = NOW()
            """, (
                proposal_id,
                selected_month,
                total_penjualan,
                total_biaya_produksi,
                total_biaya_operasional,
                total_biaya_non_operasional,
                kas_bersih_operasional,
                total_harga_jual_alat,
                total_harga_alat,
                kas_bersih_investasi,
                kas_bersih_pembiayaan,
                total_kas_bersih
            ))
            app_funcs["mysql"].connection.commit()
            
            proposal_data[proposal_id] = {
                'proposal': proposal,
                'total_penjualan': total_penjualan,
                'total_biaya_produksi': total_biaya_produksi,
                'total_biaya_operasional': total_biaya_operasional,
                'total_biaya_non_operasional': total_biaya_non_operasional,
                'kas_bersih_operasional': kas_bersih_operasional,
                'total_harga_jual_alat': total_harga_jual_alat,
                'total_harga_alat': total_harga_alat,
                'kas_bersih_investasi': kas_bersih_investasi,
                'kas_bersih_pembiayaan': kas_bersih_pembiayaan,
                'total_kas_bersih': total_kas_bersih
            }
        
        # Generate available months - Data per bulan terpisah (bukan kumulatif)
        available_months = []
        current_date = datetime.now()
        
        # Generate 12 bulan terakhir - Setiap bulan memiliki data arus kas yang terpisah
        for i in range(12):
            # Hitung mundur dari bulan saat ini
            month_date = current_date - timedelta(days=30*i)
            month_value = month_date.strftime('%Y-%m')  # Format: 2025-01, 2025-02, dst
            month_label = month_date.strftime('%B %Y')  # Format: January 2025, February 2025, dst
            available_months.append({
                'value': month_value,
                'label': month_label
            })
        
        # Format selected month year untuk display
        if not selected_month:
            selected_month_year = datetime.now().strftime('%B %Y')
        else:
            try:
                selected_date = datetime.strptime(selected_month, '%Y-%m')
                selected_month_year = selected_date.strftime('%B %Y')
            except:
                selected_month_year = datetime.now().strftime('%B %Y')
        
        cursor.close()
        
        return render_template('mahasiswa/arus_kas.html', 
                             proposal_data=proposal_data,
                             has_lolos_proposal=len(proposals) > 0,
                             available_months=available_months,
                             selected_month=selected_month,
                             selected_month_year=selected_month_year)
        
    except Exception as e:
        flash(f'Error: {str(e)}', 'danger')
        return render_template('mahasiswa/arus_kas.html', proposal_data={})


@mahasiswa_bp.route('/profile_mahasiswa')
def profile_mahasiswa():
    app_funcs = get_app_functions()
    if 'user_type' not in session or session['user_type'] != 'mahasiswa':
        flash('Anda harus login sebagai mahasiswa!', 'danger')
        return redirect(url_for('index'))
    
    if not hasattr(app_funcs["mysql"], 'connection') or app_funcs["mysql"].connection is None:
        flash('Koneksi ke database gagal. Cek konfigurasi database!', 'danger')
        return render_template('mahasiswa/profile.html', mahasiswa_data=None, anggota_tim_list=[], dosen_pembimbing_list=[])
    
    try:
        if not app_funcs["mysql"].connection:
            flash('Koneksi ke database gagal. Cek konfigurasi database!', 'danger')
            return render_template('mahasiswa/profile.html', mahasiswa_data=None, anggota_tim_list=[], dosen_pembimbing_list=[])
        
        cursor = app_funcs["mysql"].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Ambil data mahasiswa yang sedang login
        cursor.execute('''
            SELECT id, perguruan_tinggi, program_studi, nim, nama_ketua, no_telp, email, status, tanggal_daftar
            FROM mahasiswa 
            WHERE nim = %s
        ''', (session['nim'],))
        
        mahasiswa_data = cursor.fetchone()
        
        if not mahasiswa_data:
            flash('Data mahasiswa tidak ditemukan!', 'danger')
            cursor.close()
            return redirect(url_for('index'))
        
        # Ambil data anggota tim dari semua proposal mahasiswa
        cursor.execute('''
            SELECT at.*, p.judul_usaha, p.status as status_proposal
            FROM anggota_tim at
            INNER JOIN proposal p ON at.id_proposal = p.id
            WHERE p.nim = %s
            ORDER BY at.tanggal_tambah DESC
        ''', (session['nim'],))
        
        anggota_tim_list = cursor.fetchall()
        
        # Ambil data dosen pembimbing dari semua proposal mahasiswa
        cursor.execute('''
            SELECT DISTINCT p.dosen_pembimbing, p.nid_dosen, p.program_studi_dosen, p.judul_usaha, p.status as status_proposal
            FROM proposal p
            WHERE p.nim = %s AND p.dosen_pembimbing IS NOT NULL AND p.dosen_pembimbing != ''
            ORDER BY p.tanggal_buat DESC
        ''', (session['nim'],))
        
        dosen_pembimbing_list = cursor.fetchall()
        
        cursor.close()
        
        return render_template('mahasiswa/profile.html', 
                             mahasiswa_data=mahasiswa_data,
                             anggota_tim_list=anggota_tim_list,
                             dosen_pembimbing_list=dosen_pembimbing_list)
        
    except Exception as e:
        flash(f'Error saat mengambil data profile: {str(e)}', 'danger')
        return render_template('mahasiswa/profile.html', 
                             mahasiswa_data=None,
                             anggota_tim_list=[],
                             dosen_pembimbing_list=[])

# Route untuk CRUD Proposal
@mahasiswa_bp.route('/tambah_proposal', methods=['POST'])
def tambah_proposal():
    app_funcs = get_app_functions()
    if 'user_type' not in session or session['user_type'] != 'mahasiswa':
        flash('Anda harus login sebagai mahasiswa!', 'danger')
        return redirect(url_for('index'))
    
    # Cek status mahasiswa
    if not hasattr(app_funcs["mysql"], 'connection') or app_funcs["mysql"].connection is None:
        flash('Koneksi ke database gagal. Cek konfigurasi database!', 'danger')
        return redirect(url_for('mahasiswa.proposal'))
    
    try:
        cursor = app_funcs["mysql"].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Cek status mahasiswa
        cursor.execute('SELECT status FROM mahasiswa WHERE nim = %s', (session['nim'],))
        mahasiswa_data = cursor.fetchone()
        
        if not mahasiswa_data:
            flash('Data mahasiswa tidak ditemukan!', 'danger')
            cursor.close()
            return redirect(url_for('mahasiswa.proposal'))
        
        if mahasiswa_data['status'] == 'proses':
            flash('Akun Anda belum diverifikasi. Anda tidak dapat menambahkan proposal!', 'danger')
            cursor.close()
            return redirect(url_for('mahasiswa.proposal'))
        elif mahasiswa_data['status'] == 'tolak':
            flash('Akun Anda ditolak. Anda tidak dapat menambahkan proposal!', 'danger')
            cursor.close()
            return redirect(url_for('mahasiswa.proposal'))
        elif mahasiswa_data['status'] == 'selesai':
            flash('Akun Anda sudah selesai. Anda tidak dapat membuat proposal baru lagi!', 'danger')
            cursor.close()
            return redirect(url_for('mahasiswa.proposal'))
        
        # Ambil data dari form
        judul_usaha = request.form['judul_usaha'].strip().title()
        kategori = request.form['kategori'].strip().title()
        tahapan_usaha = request.form['tahapan_usaha'].strip().title()
        merk_produk = request.form['merk_produk'].strip().title()
        
        # Handle field opsional - ubah string kosong menjadi None untuk database
        nib = request.form['nib'].strip() or None
        tahun_nib = request.form['tahun_nib'].strip() or None
        platform_penjualan = request.form['platform_penjualan'].strip().title() or None
        
        # Jika tahun_nib ada, konversi ke integer
        if tahun_nib:
            try:
                tahun_nib = int(tahun_nib)
            except ValueError:
                tahun_nib = None
        
        dosen_pembimbing = request.form['dosen_pembimbing'].strip().title()
        nid_dosen = request.form['nid_dosen'].strip()
        program_studi_dosen = request.form['program_studi_dosen'].strip().title()
        
        # Validasi kuota dosen pembimbing
        if dosen_pembimbing:
            cursor.execute('''
                SELECT p.kuota_mahasiswa, COUNT(DISTINCT CASE WHEN pr.status != 'selesai' THEN pr.nim END) as jumlah_mahasiswa_bimbing
                FROM pembimbing p
                LEFT JOIN proposal pr ON p.nama = pr.dosen_pembimbing
                WHERE p.nama = %s AND p.status = 'aktif'
                GROUP BY p.id, p.kuota_mahasiswa
            ''', (dosen_pembimbing,))
            
            dosen_info = cursor.fetchone()
            if dosen_info:
                sisa_kuota = dosen_info['kuota_mahasiswa'] - dosen_info['jumlah_mahasiswa_bimbing']
                if sisa_kuota <= 0:
                    flash(f'Dosen pembimbing {dosen_pembimbing} sudah mencapai kuota maksimal!', 'danger')
                    cursor.close()
                    return redirect(url_for('mahasiswa.proposal'))
        
        # Ambil file proposal
        file = request.files.get('file_proposal')
        if file and file.filename and allowed_file(file.filename):
            file_extension = file.filename.rsplit('.', 1)[1].lower()
        else:
            if file is None or file.filename == '':
                flash('File proposal tidak dipilih!', 'danger')
            else:
                flash('Format file tidak diizinkan! Hanya PDF, DOC, atau DOCX.', 'danger')
            return redirect(url_for('mahasiswa.proposal'))
        
        # Ambil data mahasiswa dari session
        nim = session['nim']
        nama_ketua = session['nama']
        
        # Ambil data perguruan tinggi dari database
        cursor.execute('SELECT perguruan_tinggi FROM mahasiswa WHERE nim = %s', (nim,))
        mahasiswa_data = cursor.fetchone()
        
        if not mahasiswa_data:
            flash('Data mahasiswa tidak ditemukan!', 'danger')
            cursor.close()
            return redirect(url_for('mahasiswa.proposal'))
        
        perguruan_tinggi = mahasiswa_data['perguruan_tinggi']
        
        # Buat path upload
        safe_judul = re.sub(r'[^\w\s-]', '', judul_usaha).strip().replace(' ', '_')
        safe_nama_ketua = re.sub(r'[^\w\s-]', '', nama_ketua).strip().replace(' ', '_')
        upload_dir = os.path.join('static', 'uploads', 'Proposal', safe_judul)
        os.makedirs(upload_dir, exist_ok=True)
        filename = f"Proposal_{safe_judul}_{safe_nama_ketua}.{file_extension}"
        file_path = os.path.join(upload_dir, filename)
        
        # Simpan file
        file.save(file_path)
        
        # Simpan data ke database dengan tanggal yang eksplisit
        cursor.execute('''
            INSERT INTO proposal (
                nim, judul_usaha, kategori, tahapan_usaha, merk_produk, 
                nib, tahun_nib, platform_penjualan, dosen_pembimbing, 
                nid_dosen, program_studi_dosen, file_path, status,
                tanggal_buat
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, NOW())
        ''', (
            nim, judul_usaha, kategori, tahapan_usaha, merk_produk,
            nib, tahun_nib, platform_penjualan, dosen_pembimbing,
            nid_dosen, program_studi_dosen, file_path, 'draf'
        ))
        
        proposal_id = cursor.lastrowid
        
        if hasattr(app_funcs["mysql"], 'connection') and app_funcs["mysql"].connection is not None:
            app_funcs["mysql"].connection.commit()
        else:
            flash('Koneksi ke database gagal saat commit.', 'danger')
            return redirect(url_for('mahasiswa.proposal'))
        
        # Log aktivitas tambah proposal
        mahasiswa_info = app_funcs['get_mahasiswa_info_from_session']()
        if mahasiswa_info:
            data_baru = {
                'id': proposal_id,
                'judul_usaha': judul_usaha,
                'kategori': kategori,
                'tahapan_usaha': tahapan_usaha,
                'merk_produk': merk_produk,
                'dosen_pembimbing': dosen_pembimbing,
                'file_path': file_path,
                'status': 'draf'
            }
            app_funcs['log_mahasiswa_activity'](
                mahasiswa_info['id'], 
                mahasiswa_info['nim'], 
                mahasiswa_info['nama_ketua'], 
                'tambah', 
                'proposal', 
                f'id_{proposal_id}', 
                f'Menambahkan proposal: {judul_usaha}',
                None,
                data_baru
            )
        
        cursor.close()
        
        flash('Proposal berhasil disimpan!', 'success')
        return redirect(url_for('mahasiswa.proposal'))
        
    except Exception as e:
        flash(f'Error saat menyimpan proposal: {str(e)}', 'danger')
        return redirect(url_for('mahasiswa.proposal'))
@mahasiswa_bp.route('/tambah_anggota', methods=['POST'])
def tambah_anggota():
    app_funcs = get_app_functions()
    from flask import jsonify, request
    is_ajax = request.headers.get('X-Requested-With') == 'XMLHttpRequest' or request.accept_mimetypes['application/json'] > 0
    if 'user_type' not in session or session['user_type'] != 'mahasiswa':
        msg = 'Anda harus login sebagai mahasiswa!'
        if is_ajax:
            return jsonify(success=False, message=msg)
        flash(msg, 'danger')
        return redirect(url_for('index'))
    if not hasattr(app_funcs["mysql"], 'connection') or app_funcs["mysql"].connection is None:
        msg = 'Koneksi ke database gagal. Cek konfigurasi database!'
        if is_ajax:
            return jsonify(success=False, message=msg)
        flash(msg, 'danger')
        return redirect(url_for('mahasiswa.proposal'))
    cursor = app_funcs["mysql"].connection.cursor(MySQLdb.cursors.DictCursor)
    try:
        cursor.execute('SELECT status FROM mahasiswa WHERE nim = %s', (session['nim'],))
        mahasiswa_data = cursor.fetchone()
        if not mahasiswa_data:
            msg = 'Data mahasiswa tidak ditemukan!'
            if is_ajax:
                return jsonify(success=False, message=msg)
            flash(msg, 'danger')
            cursor.close()
            return redirect(url_for('mahasiswa.proposal'))
        if mahasiswa_data['status'] == 'proses':
            msg = 'Akun Anda belum diverifikasi. Anda tidak dapat menambahkan anggota tim!'
            if is_ajax:
                return jsonify(success=False, message=msg)
            flash(msg, 'danger')
            cursor.close()
            return redirect(url_for('mahasiswa.proposal'))
        elif mahasiswa_data['status'] == 'tolak':
            msg = 'Akun Anda ditolak. Anda tidak dapat menambahkan anggota tim!'
            if is_ajax:
                return jsonify(success=False, message=msg)
            flash(msg, 'danger')
            cursor.close()
            return redirect(url_for('mahasiswa.proposal'))
        elif mahasiswa_data['status'] == 'selesai':
            msg = 'Akun Anda sudah selesai. Anda tidak dapat menambahkan anggota tim!'
            if is_ajax:
                return jsonify(success=False, message=msg)
            flash(msg, 'danger')
            cursor.close()
            return redirect(url_for('mahasiswa.proposal'))
        id_proposal = request.form['id_proposal']
        cursor.execute('SELECT COUNT(*) as jumlah FROM anggota_tim WHERE id_proposal = %s', (id_proposal,))
        jumlah_anggota_existing = cursor.fetchone()['jumlah']
        anggota_data = []
        i = 1
        while f'nim_{i}' in request.form:
            anggota = {
                'perguruan_tinggi': request.form[f'perguruan_tinggi_{i}'].strip().title(),
                'program_studi': request.form[f'program_studi_{i}'].strip().title(),
                'nim': request.form[f'nim_{i}'].strip(),
                'nama': request.form[f'nama_{i}'].strip().title(),
                'email': request.form[f'email_{i}'].strip().lower(),
                'no_telp': request.form[f'no_telp_{i}'].strip()
            }
            anggota_data.append(anggota)
            i += 1
        total_anggota = jumlah_anggota_existing + len(anggota_data)
        
        # Validasi minimal 3 anggota
        if total_anggota < 3:
            msg = f'Tim harus memiliki minimal 3 anggota. Saat ini hanya ada {total_anggota} anggota ({jumlah_anggota_existing} di database + {len(anggota_data)} yang akan ditambahkan).'
            if is_ajax:
                return jsonify(success=False, message=msg)
            flash(msg, 'danger')
            return redirect(url_for('mahasiswa.proposal'))
        
        # Validasi maksimal 4 anggota
        if total_anggota > 4:
            msg = f'Total anggota tim maksimal 4 orang. Saat ini sudah ada {jumlah_anggota_existing} anggota, Anda hanya bisa menambah {4 - jumlah_anggota_existing} anggota lagi.'
            if is_ajax:
                return jsonify(success=False, message=msg)
            flash(msg, 'danger')
            return redirect(url_for('mahasiswa.proposal'))
        if not anggota_data:
            msg = 'Tidak ada data anggota yang dikirim!'
            if is_ajax:
                return jsonify(success=False, message=msg)
            flash(msg, 'danger')
            return redirect(url_for('mahasiswa.proposal'))
        for anggota in anggota_data:
            if not all(anggota.values()):
                msg = 'Semua field anggota harus diisi!'
                if is_ajax:
                    return jsonify(success=False, message=msg)
                flash(msg, 'danger')
                return redirect(url_for('mahasiswa.proposal'))
        for anggota in anggota_data:
            cursor.execute('SELECT * FROM anggota_tim WHERE nim = %s OR email = %s', (anggota['nim'], anggota['email']))
            existing = cursor.fetchone()
            if existing:
                msg = f'NIM {anggota["nim"]} atau Email {anggota["email"]} sudah terdaftar!'
                if is_ajax:
                    return jsonify(success=False, message=msg)
                flash(msg, 'danger')
                return redirect(url_for('mahasiswa.proposal'))
        success_count = 0
        for anggota in anggota_data:
            cursor.execute('''
                INSERT INTO anggota_tim (
                    id_proposal, perguruan_tinggi, program_studi, nim, nama, email, no_telp
                ) VALUES (%s, %s, %s, %s, %s, %s, %s)
            ''', (
                id_proposal, 
                anggota['perguruan_tinggi'], 
                anggota['program_studi'], 
                anggota['nim'], 
                anggota['nama'], 
                anggota['email'], 
                anggota['no_telp']
            ))
            success_count += 1
        if hasattr(app_funcs["mysql"], 'connection') and app_funcs["mysql"].connection is not None:
            app_funcs["mysql"].connection.commit()
        else:
            msg = 'Koneksi ke database gagal saat commit.'
            if is_ajax:
                return jsonify(success=False, message=msg)
            flash(msg, 'danger')
            return redirect(url_for('mahasiswa.proposal'))
        
        # Log aktivitas tambah anggota tim
        mahasiswa_info = app_funcs['get_mahasiswa_info_from_session']()
        if mahasiswa_info:
            data_baru = {
                'proposal_id': id_proposal,
                'anggota_count': success_count,
                'anggota_data': anggota_data
            }
            app_funcs['log_mahasiswa_activity'](
                mahasiswa_info['id'], 
                mahasiswa_info['nim'], 
                mahasiswa_info['nama_ketua'], 
                'tambah', 
                'anggota_tim', 
                f'proposal_id_{id_proposal}', 
                f'Menambahkan {success_count} anggota tim',
                None,
                data_baru
            )
        
        cursor.close()
        if success_count == 1:
            msg = 'Anggota tim berhasil ditambahkan!'
        else:
            msg = f'{success_count} anggota tim berhasil ditambahkan!'
        if is_ajax:
            return jsonify(success=True, message=msg)
        flash(msg, 'success')
        return redirect(url_for('mahasiswa.proposal'))
    except Exception as e:
        msg = f'Error saat menambah anggota: {str(e)}'
        if is_ajax:
            return jsonify(success=False, message=msg)
        flash(msg, 'danger')
        return redirect(url_for('mahasiswa.proposal'))
@mahasiswa_bp.route('/hapus_proposal/<int:id_proposal>', methods=['POST'])
def hapus_proposal(id_proposal):
    logger.info(f"Hapus proposal dipanggil untuk proposal_id: {id_proposal}")
    app_funcs = get_app_functions()
    if 'user_type' not in session or session['user_type'] != 'mahasiswa':
        logger.warning(f"Akses ditolak untuk hapus proposal: {session.get('user_type', 'None')}")
        flash('Anda harus login sebagai mahasiswa!', 'danger')
        return redirect(url_for('index'))
    
    # Cek status mahasiswa
    if not hasattr(app_funcs["mysql"], 'connection') or app_funcs["mysql"].connection is None:
        flash('Koneksi ke database gagal. Cek konfigurasi database!', 'danger')
        return redirect(url_for('mahasiswa.proposal'))
    
    cursor = app_funcs["mysql"].connection.cursor(MySQLdb.cursors.DictCursor)
    
    try:
        # Cek status mahasiswa
        cursor.execute('SELECT status FROM mahasiswa WHERE nim = %s', (session['nim'],))
        mahasiswa_data = cursor.fetchone()
        
        if not mahasiswa_data:
            flash('Data mahasiswa tidak ditemukan!', 'danger')
            cursor.close()
            return redirect(url_for('mahasiswa.proposal'))
        
        if mahasiswa_data['status'] == 'proses':
            flash('Akun Anda belum diverifikasi. Anda tidak dapat menghapus proposal!', 'danger')
            cursor.close()
            return redirect(url_for('mahasiswa.proposal'))
        elif mahasiswa_data['status'] == 'tolak':
            flash('Akun Anda ditolak. Anda tidak dapat menghapus proposal!', 'danger')
            cursor.close()
            return redirect(url_for('mahasiswa.proposal'))
        elif mahasiswa_data['status'] == 'selesai':
            flash('Akun Anda sudah selesai. Anda tidak dapat menghapus proposal!', 'danger')
            cursor.close()
            return redirect(url_for('mahasiswa.proposal'))
        
        # Ambil data proposal untuk mendapatkan file_path
        cursor.execute('''
            SELECT p.file_path, p.nim, p.judul_usaha
            FROM proposal p 
            WHERE p.id = %s AND p.nim = %s
        ''', (id_proposal, session['nim']))
        proposal_data = cursor.fetchone()
        
        if not proposal_data:
            flash('Proposal tidak ditemukan!', 'danger')
            cursor.close()
            return redirect(url_for('mahasiswa.proposal'))
        
        # Log aktivitas hapus proposal
        mahasiswa_info = app_funcs['get_mahasiswa_info_from_session']()
        if mahasiswa_info:
            data_lama = {
                'id': id_proposal,
                'judul_usaha': proposal_data.get('judul_usaha'),
                'nim': proposal_data.get('nim'),
                'file_path': proposal_data.get('file_path')
            }
            app_funcs['log_mahasiswa_activity'](
                mahasiswa_info['id'], 
                mahasiswa_info['nim'], 
                mahasiswa_info['nama_ketua'], 
                'hapus', 
                'proposal', 
                f'id_{id_proposal}', 
                f'Menghapus proposal: {proposal_data.get("judul_usaha")}',
                data_lama,
                None
            )
        
        file_path = proposal_data['file_path']
        
        # Fungsi helper untuk menghapus file dan folder
        def hapus_file_dan_folder(file_path, file_type):
            if file_path and os.path.exists(file_path):
                try:
                    # Hapus file terlebih dahulu
                    os.remove(file_path)
                    logger.info(f"✅ File {file_type} berhasil dihapus: {file_path}")
                    
                    # Cek folder setelah file dihapus
                    folder_path = os.path.dirname(file_path)
                    if os.path.exists(folder_path):
                        # Cek apakah folder kosong setelah file dihapus
                        remaining_files = [f for f in os.listdir(folder_path) if os.path.isfile(os.path.join(folder_path, f))]
                        
                        if not remaining_files:
                            # Jika folder kosong, hapus folder tersebut
                            os.rmdir(folder_path)
                            logger.info(f"✅ Folder kosong berhasil dihapus: {folder_path}")
                            
                            # Cek juga parent folder jika kosong
                            parent_folder = os.path.dirname(folder_path)
                            if os.path.exists(parent_folder):
                                parent_files = [f for f in os.listdir(parent_folder) if os.path.isfile(os.path.join(parent_folder, f))]
                                if not parent_files:
                                    os.rmdir(parent_folder)
                                    print(f"✅ Parent folder kosong berhasil dihapus: {parent_folder}")
                        else:
                            print(f"ℹ️ Folder tidak dihapus karena masih ada {len(remaining_files)} file lain: {folder_path}")
                        
                except PermissionError as e:
                    print(f"❌ Error permission saat menghapus file {file_path}: {e}")
                    flash(f'Warning: Tidak dapat menghapus file {file_type} karena masalah permission. File mungkin sedang digunakan.', 'warning')
                except FileNotFoundError as e:
                    print(f"⚠️ File {file_type} tidak ditemukan: {file_path}")
                    flash(f'Info: File {file_type} tidak ditemukan di server.', 'info')
                except Exception as e:
                    print(f"❌ Error tidak terduga saat menghapus file {file_path}: {e}")
                    flash(f'Warning: Gagal menghapus file {file_type}: {str(e)}', 'warning')
            else:
                if file_path:
                        print(f"⚠️ File {file_type} tidak ditemukan: {file_path}")
                else:
                        print(f"⚠️ Tidak ada path file {file_type} yang tersimpan")
        
        # Hapus file proposal
        hapus_file_dan_folder(file_path, "proposal")
        
        # Hapus file laporan kemajuan tidak lagi menggunakan kolom `laporan_kemajuan_path`.
        # Penghapusan file kemajuan/akhir dilakukan oleh helper hapus_semua_file_proposal di bawah.
        
        # Hapus semua file terkait proposal (komprehensif)
        app_funcs['hapus_semua_file_proposal'](id_proposal, session['nim'])
        
        # Hapus folder proposal jika kosong
        try:
            # Ambil data mahasiswa untuk nama ketua
            cursor.execute('SELECT nama_ketua FROM mahasiswa WHERE nim = %s', (session['nim'],))
            mahasiswa_info = cursor.fetchone()
            
            if mahasiswa_info and proposal_data['judul_usaha']:
                safe_judul = re.sub(r'[^\w\s-]', '', proposal_data['judul_usaha']).strip().replace(' ', '_')
                safe_nama_ketua = re.sub(r'[^\w\s-]', '', mahasiswa_info['nama_ketua']).strip().replace(' ', '_')
                proposal_folder = os.path.join('static', 'uploads', 'Proposal', safe_judul)
                
                # Hapus folder proposal jika kosong
                if os.path.exists(proposal_folder):
                    remaining_files = [f for f in os.listdir(proposal_folder) if os.path.isfile(os.path.join(proposal_folder, f))]
                    if not remaining_files:
                        os.rmdir(proposal_folder)
                        print(f"✅ Folder proposal kosong berhasil dihapus: {proposal_folder}")
                        
                        # Cek parent folder (Proposal) jika kosong
                        parent_proposal_folder = os.path.dirname(proposal_folder)
                        if os.path.exists(parent_proposal_folder):
                            parent_files = [f for f in os.listdir(parent_proposal_folder) if os.path.isfile(os.path.join(parent_proposal_folder, f))]
                            if not parent_files:
                                os.rmdir(parent_proposal_folder)
                                print(f"✅ Parent folder Proposal kosong berhasil dihapus: {parent_proposal_folder}")
        except Exception as e:
            print(f"⚠️ Error saat menghapus folder proposal: {e}")
        
        
        # Hapus anggaran awal terkait
        cursor.execute('DELETE FROM anggaran_awal WHERE id_proposal = %s', (id_proposal,))
        
        # Hapus anggaran bertumbuh terkait
        cursor.execute('DELETE FROM anggaran_bertumbuh WHERE id_proposal = %s', (id_proposal,))
        
        # Hapus laporan kemajuan awal terkait
        cursor.execute('DELETE FROM laporan_kemajuan_awal WHERE id_proposal = %s', (id_proposal,))
        
        # Hapus laporan kemajuan bertumbuh terkait
        cursor.execute('DELETE FROM laporan_kemajuan_bertumbuh WHERE id_proposal = %s', (id_proposal,))
        
        # Hapus laporan akhir awal terkait
        cursor.execute('DELETE FROM laporan_akhir_awal WHERE id_proposal = %s', (id_proposal,))
        
        # Hapus laporan akhir bertumbuh terkait
        cursor.execute('DELETE FROM laporan_akhir_bertumbuh WHERE id_proposal = %s', (id_proposal,))
        
        # Hapus data dari tabel-tabel terkait (dalam urutan yang benar untuk menghindari foreign key constraint)
        cursor.execute('DELETE FROM alat_produksi WHERE proposal_id = %s', (id_proposal,))
        cursor.execute('DELETE FROM bahan_baku WHERE proposal_id = %s', (id_proposal,))
        cursor.execute('DELETE FROM biaya_operasional WHERE proposal_id = %s', (id_proposal,))
        cursor.execute('DELETE FROM biaya_non_operasional WHERE proposal_id = %s', (id_proposal,))
        cursor.execute('DELETE FROM penjualan WHERE proposal_id = %s', (id_proposal,))
        cursor.execute('DELETE FROM laba_rugi WHERE proposal_id = %s', (id_proposal,))
        cursor.execute('DELETE FROM arus_kas WHERE proposal_id = %s', (id_proposal,))

        # Hapus data bimbingan terkait proposal
        cursor.execute('DELETE FROM bimbingan WHERE proposal_id = %s', (id_proposal,))
        
        # Hapus data penilaian terkait proposal (dalam urutan yang benar untuk menghindari foreign key constraint)
        
        # 1. Hapus detail penilaian proposal (pakai id_proposal langsung sesuai skema)
        cursor.execute('DELETE FROM detail_penilaian_proposal WHERE id_proposal = %s', (id_proposal,))
        
        # 2. Hapus catatan penilaian proposal (opsional jika tabel ada)
        def table_exists(table_name: str) -> bool:
            try:
                cur = app_funcs["mysql"].connection.cursor(MySQLdb.cursors.DictCursor)
                cur.execute(
                    """
                    SELECT 1 
                    FROM information_schema.tables 
                    WHERE table_schema = DATABASE() AND table_name = %s
                    LIMIT 1
                    """,
                    (table_name,)
                )
                exists = cur.fetchone() is not None
                cur.close()
                return exists
            except Exception:
                return False
        
        if table_exists('catatan_penilaian_proposal'):
            cursor.execute('''
                DELETE cpp FROM catatan_penilaian_proposal cpp 
                INNER JOIN penilaian_proposal pp ON cpp.id_penilaian_proposal = pp.id 
                WHERE pp.id_proposal = %s
            ''', (id_proposal,))
        
        # 3. Hapus rekomendasi penilaian proposal (opsional jika tabel ada)
        if table_exists('rekomendasi_penilaian_proposal'):
            cursor.execute('''
                DELETE rpp FROM rekomendasi_penilaian_proposal rpp 
                INNER JOIN penilaian_proposal pp ON rpp.id_penilaian_proposal = pp.id 
                WHERE pp.id_proposal = %s
            ''', (id_proposal,))
        
        # 4. Hapus penilaian proposal
        cursor.execute('DELETE FROM penilaian_proposal WHERE id_proposal = %s', (id_proposal,))
        
        # 5. Hapus detail penilaian laporan kemajuan, lalu header
        cursor.execute('''
            DELETE dplk FROM detail_penilaian_laporan_kemajuan dplk
            INNER JOIN penilaian_laporan_kemajuan plk ON dplk.id_penilaian_laporan_kemajuan = plk.id
            WHERE plk.id_proposal = %s
        ''', (id_proposal,))
        cursor.execute('DELETE FROM penilaian_laporan_kemajuan WHERE id_proposal = %s', (id_proposal,))

        # 6. Hapus detail penilaian laporan akhir, lalu header
        cursor.execute('''
            DELETE dpla FROM detail_penilaian_laporan_akhir dpla
            INNER JOIN penilaian_laporan_akhir pla ON dpla.id_penilaian_laporan_akhir = pla.id
            WHERE pla.id_proposal = %s
        ''', (id_proposal,))
        cursor.execute('DELETE FROM penilaian_laporan_akhir WHERE id_proposal = %s', (id_proposal,))

        # 7. Hapus detail penilaian mahasiswa (oleh pembimbing), lalu header
        cursor.execute('''
            DELETE dpm FROM detail_penilaian_mahasiswa dpm
            INNER JOIN penilaian_mahasiswa pm ON dpm.id_penilaian_mahasiswa = pm.id
            WHERE pm.id_proposal = %s
        ''', (id_proposal,))
        cursor.execute('DELETE FROM penilaian_mahasiswa WHERE id_proposal = %s', (id_proposal,))
        
        # Hapus produk_bahan_baku yang terkait dengan produksi dari proposal ini
        cursor.execute('''
            DELETE pbb FROM produk_bahan_baku pbb 
            INNER JOIN produksi p ON pbb.produksi_id = p.id 
            WHERE p.proposal_id = %s
        ''', (id_proposal,))
        
        # Hapus produksi (setelah produk_bahan_baku dihapus)
        cursor.execute('DELETE FROM produksi WHERE proposal_id = %s', (id_proposal,))
        
        # Hapus relasi reviewer dan data file_laporan_* terkait proposal
        cursor.execute('DELETE FROM proposal_reviewer WHERE id_proposal = %s', (id_proposal,))
        cursor.execute('DELETE FROM file_laporan_kemajuan WHERE id_proposal = %s', (id_proposal,))
        cursor.execute('DELETE FROM file_laporan_akhir WHERE id_proposal = %s', (id_proposal,))

        # Hapus anggota tim terkait
        cursor.execute('DELETE FROM anggota_tim WHERE id_proposal = %s', (id_proposal,))
        
        # Hapus proposal
        cursor.execute('DELETE FROM proposal WHERE id = %s AND nim = %s', (id_proposal, session['nim']))
        
        if hasattr(app_funcs["mysql"], 'connection') and app_funcs["mysql"].connection is not None:
            app_funcs["mysql"].connection.commit()
        else:
            flash('Koneksi ke database gagal saat commit.', 'danger')
            return redirect(url_for('mahasiswa.proposal'))
        cursor.close()
        
        flash('Proposal berhasil dihapus! Semua data terkait (file proposal, laporan kemajuan, file laporan kemajuan dari folder kemajuan, file laporan akhir dari folder proposal, data laporan kemajuan/akhir, anggaran, alat produksi, bahan baku, biaya operasional/non-operasional, produk bahan baku, produksi, penjualan, laba rugi, arus kas, bimbingan, penilaian proposal, penilaian laporan kemajuan, penilaian laporan akhir, dan anggota tim) juga telah dihapus secara permanen.', 'success')
        return redirect(url_for('mahasiswa.proposal'))
        
    except Exception as e:
        flash(f'Error saat menghapus proposal: {str(e)}', 'danger')
        return redirect(url_for('mahasiswa.proposal'))

@mahasiswa_bp.route('/kirim_proposal/<int:id_proposal>', methods=['POST'])
def kirim_proposal(id_proposal):
    app_funcs = get_app_functions()
    # Deteksi permintaan AJAX/Fetch agar tidak mengembalikan redirect 302
    wants_json = (
        (request.headers.get('X-Requested-With') == 'XMLHttpRequest')
        or (request.accept_mimetypes.accept_json and not request.accept_mimetypes.accept_html)
    )

    def respond_ok(message: str):
        if wants_json:
            return jsonify({
                'success': True,
                'message': message
            })
        flash(message, 'success')
        return redirect(url_for('mahasiswa.proposal'))

    def respond_fail(message: str, category: str = 'danger', code: int = 400):
        if wants_json:
            return jsonify({
                'success': False,
                'message': message
            }), code
        flash(message, category)
        return redirect(url_for('mahasiswa.proposal'))
    if 'user_type' not in session or session['user_type'] != 'mahasiswa':
        if wants_json:
            return jsonify({'success': False, 'message': 'Anda harus login sebagai mahasiswa!'}), 401
        flash('Anda harus login sebagai mahasiswa!', 'danger')
        return redirect(url_for('index'))
    
    # Cek status mahasiswa
    if not hasattr(app_funcs["mysql"], 'connection') or app_funcs["mysql"].connection is None:
        return respond_fail('Koneksi ke database gagal. Cek konfigurasi database!', 'danger', 500)
    
    cursor = app_funcs["mysql"].connection.cursor(MySQLdb.cursors.DictCursor)
    
    try:
        # Cek status mahasiswa
        cursor.execute('SELECT status FROM mahasiswa WHERE nim = %s', (session['nim'],))
        mahasiswa_data = cursor.fetchone()
        
        if not mahasiswa_data:
            cursor.close()
            return respond_fail('Data mahasiswa tidak ditemukan!', 'danger', 404)
        
        if mahasiswa_data['status'] == 'proses':
            cursor.close()
            return respond_fail('Akun Anda belum diverifikasi. Anda tidak dapat mengirim proposal!', 'danger', 403)
        elif mahasiswa_data['status'] == 'tolak':
            cursor.close()
            return respond_fail('Akun Anda ditolak. Anda tidak dapat mengirim proposal!', 'danger', 403)
        elif mahasiswa_data['status'] == 'selesai':
            cursor.close()
            return respond_fail('Akun Anda sudah selesai. Anda tidak dapat mengirim proposal!', 'danger', 403)
        
        # Cek jumlah anggota tim minimal 3
        cursor.execute('SELECT COUNT(*) as jumlah FROM anggota_tim WHERE id_proposal = %s', (id_proposal,))
        jumlah_anggota = cursor.fetchone()['jumlah']
        
        if jumlah_anggota < 3:
            cursor.close()
            return respond_fail(
                f'Tim harus memiliki minimal 3 anggota sebelum mengirim proposal. Saat ini hanya ada {jumlah_anggota} anggota. Silakan tambahkan anggota tim terlebih dahulu.',
                'warning',
                400
            )
        
        # Validasi total anggaran terhadap pengaturan min/max sebelum ubah status
        try:
            cursor2 = app_funcs["mysql"].connection.cursor(MySQLdb.cursors.DictCursor)
            # Pastikan tabel pengaturan_anggaran ada
            cursor2.execute(
                """
                CREATE TABLE IF NOT EXISTS pengaturan_anggaran (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    min_total_awal BIGINT DEFAULT 0,
                    max_total_awal BIGINT DEFAULT 0,
                    min_total_bertumbuh BIGINT DEFAULT 0,
                    max_total_bertumbuh BIGINT DEFAULT 0,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP
                )
                """
            )
            app_funcs["mysql"].connection.commit()
            cursor2.execute('SELECT * FROM pengaturan_anggaran ORDER BY id DESC LIMIT 1')
            batas = cursor2.fetchone() or {
                'min_total_awal': 0, 'max_total_awal': 0,
                'min_total_bertumbuh': 0, 'max_total_bertumbuh': 0
            }

            # Hitung total dan jumlah baris anggaran awal dan bertumbuh untuk proposal ini
            cursor2.execute('SELECT COALESCE(SUM(jumlah),0) AS total, COUNT(*) AS cnt FROM anggaran_awal WHERE id_proposal=%s', (id_proposal,))
            row_awal = cursor2.fetchone()
            total_awal = int(row_awal['total'] or 0)
            count_awal = int(row_awal['cnt'] or 0)

            cursor2.execute('SELECT COALESCE(SUM(jumlah),0) AS total, COUNT(*) AS cnt FROM anggaran_bertumbuh WHERE id_proposal=%s', (id_proposal,))
            row_bert = cursor2.fetchone()
            total_bert = int(row_bert['total'] or 0)
            count_bert = int(row_bert['cnt'] or 0)
            cursor2.close()

            # Validasi helper
            def check_range(total, min_v, max_v, label):
                if min_v and total < min_v:
                    return f'Total {label} (Rp {total:,.0f}) kurang dari minimal (Rp {min_v:,.0f}).'
                if max_v and total > max_v:
                    return f'Total {label} (Rp {total:,.0f}) melebihi maksimal (Rp {max_v:,.0f}).'
                return None

            # Terapkan batas minimal hanya jika ada data di tabel terkait
            min_awal_used = int(batas.get('min_total_awal') or 0) if count_awal > 0 else 0
            max_awal_used = int(batas.get('max_total_awal') or 0)
            min_bert_used = int(batas.get('min_total_bertumbuh') or 0) if count_bert > 0 else 0
            max_bert_used = int(batas.get('max_total_bertumbuh') or 0)

            err_awal = check_range(total_awal, min_awal_used, max_awal_used, 'Anggaran Awal')
            err_bert = check_range(total_bert, min_bert_used, max_bert_used, 'Anggaran Bertumbuh')

            if err_awal or err_bert:
                msg = err_awal or err_bert
                cursor.close()
                return respond_fail(msg, 'warning', 400)
        except Exception as _:
            # Jika gagal mengambil pengaturan, lanjut tanpa validasi agar tidak memblokir kirim (fail-open)
            pass

        # Update status proposal menjadi 'diajukan'
        cursor.execute('''
            UPDATE proposal 
            SET status = 'diajukan', 
                tanggal_kirim = NOW(),
                komentar_pembimbing = NULL,
                tanggal_komentar_pembimbing = NULL
            WHERE id = %s AND nim = %s
        ''', (id_proposal, session['nim']))
        
        # Update status anggaran awal menjadi 'diajukan' (jika ada)
        cursor.execute('''
            UPDATE anggaran_awal 
            SET status = 'diajukan' 
            WHERE id_proposal = %s
        ''', (id_proposal,))
        
        # Update status anggaran bertumbuh menjadi 'diajukan' (jika ada)
        cursor.execute('''
            UPDATE anggaran_bertumbuh 
            SET status = 'diajukan' 
            WHERE id_proposal = %s
        ''', (id_proposal,))
        
        if hasattr(app_funcs["mysql"], 'connection') and app_funcs["mysql"].connection is not None:
            app_funcs["mysql"].connection.commit()
        else:
            return respond_fail('Koneksi ke database gagal saat commit.', 'danger', 500)
        
        # Log aktivitas kirim proposal
        mahasiswa_info = app_funcs['get_mahasiswa_info_from_session']()
        if mahasiswa_info:
            # Ambil data proposal untuk logging
            cursor.execute('SELECT judul_usaha FROM proposal WHERE id = %s', (id_proposal,))
            proposal_info = cursor.fetchone()
            
            data_baru = {
                'id': id_proposal,
                'status': 'diajukan',
                'tanggal_kirim': datetime.now().isoformat(),
                'judul_usaha': proposal_info.get('judul_usaha') if proposal_info else 'Unknown'
            }
            app_funcs['log_mahasiswa_activity'](
                mahasiswa_info['id'], 
                mahasiswa_info['nim'], 
                mahasiswa_info['nama_ketua'], 
                'edit', 
                'proposal', 
                f'id_{id_proposal}', 
                f'Mengirim proposal untuk review: {proposal_info.get("judul_usaha") if proposal_info else "Unknown"}',
                {'status': 'draf'},
                data_baru
            )
        
        cursor.close()
        
        return respond_ok('Proposal berhasil dikirim! Status anggaran juga telah diubah menjadi diajukan.')
        
    except Exception as e:
        return respond_fail(f'Error saat mengirim proposal: {str(e)}', 'danger', 500)

@mahasiswa_bp.route('/cek_anggaran/<int:id_proposal>')
def cek_anggaran(id_proposal):
    app_funcs = get_app_functions()
    if 'user_type' not in session or session['user_type'] != 'mahasiswa':
        return jsonify({'success': False, 'message': 'Anda harus login sebagai mahasiswa!'})
    
    if not hasattr(app_funcs["mysql"], 'connection') or app_funcs["mysql"].connection is None:
        return jsonify({'success': False, 'message': 'Koneksi ke database gagal!'})
    
    try:
        tabel = request.args.get('tabel')
        if not tabel or tabel not in ['anggaran_awal', 'anggaran_bertumbuh']:
            return jsonify({'success': False, 'message': 'Tabel anggaran tidak valid!'})
        
        cursor = app_funcs["mysql"].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Cek apakah ada data anggaran untuk proposal ini
        cursor.execute(f'SELECT COUNT(*) FROM {tabel} WHERE id_proposal = %s', (id_proposal,))
        count = cursor.fetchone()['COUNT(*)']
        
        cursor.close()
        
        return jsonify({
            'success': True, 
            'has_anggaran': count > 0,
            'count': count,
            'message': f'Ditemukan {count} data anggaran' if count > 0 else 'Tidak ada data anggaran'
        })
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error saat mengecek anggaran: {str(e)}'})


@mahasiswa_bp.route('/tambah_anggaran_awal', methods=['POST'])
def tambah_anggaran_awal():
    app_funcs = get_app_functions()
    if 'user_type' not in session or session['user_type'] != 'mahasiswa':
        return jsonify({'success': False, 'message': 'Anda harus login sebagai mahasiswa!'})
    if not hasattr(app_funcs["mysql"], 'connection') or app_funcs["mysql"].connection is None:
        return jsonify({'success': False, 'message': 'Koneksi ke database gagal. Cek konfigurasi database!'})
    try:
        data = request.get_json()
        id_proposal = data['id_proposal']
        daftar_kegiatan_utama = data['daftar_kegiatan_utama']
        cursor = app_funcs["mysql"].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Check if there's already anggaran with status diajukan, disetujui, or ditolak
        cursor.execute('''
            SELECT COUNT(*) as count 
            FROM anggaran_awal 
            WHERE id_proposal = %s AND status IN ('diajukan', 'disetujui', 'ditolak')
        ''', (id_proposal,))
        
        existing_anggaran = cursor.fetchone()
        if existing_anggaran['count'] > 0:
            cursor.close()
            return jsonify({'success': False, 'message': 'Tidak dapat menambah anggaran karena sudah ada anggaran dengan status diajukan, disetujui, atau ditolak!'})
        
        for keg_utama in daftar_kegiatan_utama:
            kegiatan_utama = keg_utama['kegiatan_utama'].strip().title()
            for kegiatan in keg_utama['daftar_kegiatan']:
                nama_kegiatan = kegiatan['kegiatan'].strip().title()
                penanggung_jawab = kegiatan['penanggung_jawab'].strip().title()
                target_capaian = kegiatan['target_capaian'].strip().title()
                for barang in kegiatan['daftar_barang']:
                    nama_barang = barang['nama_barang'].strip().title()
                    satuan = barang['satuan'].strip().title()
                    keterangan = barang['keterangan'].strip().title() if 'keterangan' in barang else ''
                    cursor.execute('''
                        INSERT INTO anggaran_awal (
                            id_proposal, kegiatan_utama, kegiatan, penanggung_jawab, target_capaian,
                            nama_barang, kuantitas, satuan, harga_satuan, jumlah, keterangan, status
                        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    ''', (
                        id_proposal, kegiatan_utama, nama_kegiatan, penanggung_jawab, target_capaian,
                        nama_barang, barang['kuantitas'], satuan,
                        barang['harga_satuan'], barang['jumlah'], keterangan, 'draf'
                    ))
        if hasattr(app_funcs["mysql"], 'connection') and app_funcs["mysql"].connection is not None:
            app_funcs["mysql"].connection.commit()
        cursor.close()
        return jsonify({'success': True, 'message': 'Data anggaran berhasil disimpan!'})
    except Exception as e:
        return jsonify({'success': False, 'message': f'Gagal menyimpan data: {str(e)}'})

@mahasiswa_bp.route('/tambah_anggaran_bertumbuh', methods=['POST'])
def tambah_anggaran_bertumbuh():
    app_funcs = get_app_functions()
    if 'user_type' not in session or session['user_type'] != 'mahasiswa':
        return jsonify({'success': False, 'message': 'Anda harus login sebagai mahasiswa!'})
    if not hasattr(app_funcs["mysql"], 'connection') or app_funcs["mysql"].connection is None:
        return jsonify({'success': False, 'message': 'Koneksi ke database gagal. Cek konfigurasi database!'})
    try:
        data = request.get_json()
        id_proposal = data['id_proposal']
        daftar_kegiatan_utama = data['daftar_kegiatan_utama']
        cursor = app_funcs["mysql"].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Check if there's already anggaran with status diajukan, disetujui, or ditolak
        cursor.execute('''
            SELECT COUNT(*) as count 
            FROM anggaran_bertumbuh 
            WHERE id_proposal = %s AND status IN ('diajukan', 'disetujui', 'ditolak')
        ''', (id_proposal,))
        
        existing_anggaran = cursor.fetchone()
        if existing_anggaran['count'] > 0:
            cursor.close()
            return jsonify({'success': False, 'message': 'Tidak dapat menambah anggaran karena sudah ada anggaran dengan status diajukan, disetujui, atau ditolak!'})
        
        for keg_utama in daftar_kegiatan_utama:
            kegiatan_utama = keg_utama['kegiatan_utama'].strip().title()
            for kegiatan in keg_utama['daftar_kegiatan']:
                nama_kegiatan = kegiatan['kegiatan'].strip().title()
                penanggung_jawab = kegiatan['penanggung_jawab'].strip().title()
                target_capaian = kegiatan['target_capaian'].strip().title()
                for barang in kegiatan['daftar_barang']:
                    nama_barang = barang['nama_barang'].strip().title()
                    satuan = barang['satuan'].strip().title()
                    keterangan = barang['keterangan'].strip().title() if 'keterangan' in barang else ''
                    cursor.execute('''
                        INSERT INTO anggaran_bertumbuh (
                            id_proposal, kegiatan_utama, kegiatan, penanggung_jawab, target_capaian,
                            nama_barang, kuantitas, satuan, harga_satuan, jumlah, keterangan, status
                        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    ''', (
                        id_proposal, kegiatan_utama, nama_kegiatan, penanggung_jawab, target_capaian,
                        nama_barang, barang['kuantitas'], satuan,
                        barang['harga_satuan'], barang['jumlah'], keterangan, 'draf'
                    ))
        if hasattr(app_funcs["mysql"], 'connection') and app_funcs["mysql"].connection is not None:
            app_funcs["mysql"].connection.commit()
        cursor.close()
        return jsonify({'success': True, 'message': 'Data anggaran berhasil disimpan!'})
    except Exception as e:
        return jsonify({'success': False, 'message': f'Gagal menyimpan data: {str(e)}'})

@mahasiswa_bp.route('/get_anggaran/<int:anggaran_id>')
def get_anggaran(anggaran_id):
    app_funcs = get_app_functions()
    if 'user_type' not in session or session['user_type'] != 'mahasiswa':
        return jsonify({'success': False, 'message': 'Anda harus login sebagai mahasiswa!'})
    
    if not hasattr(app_funcs["mysql"], 'connection') or app_funcs["mysql"].connection is None:
        return jsonify({'success': False, 'message': 'Koneksi ke database gagal!'})
    
    try:
        # Ambil parameter tabel dari query string
        tabel = request.args.get('tabel', 'anggaran_awal')
        if tabel not in ['anggaran_awal', 'anggaran_bertumbuh']:
            return jsonify({'success': False, 'message': 'Tabel anggaran tidak valid!'})
        
        cursor = app_funcs["mysql"].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Get anggaran data by ID and ensure it belongs to the logged-in student
        cursor.execute(f'''
            SELECT aa.*, COALESCE(aa.nilai_bantuan, 0) as nilai_bantuan
            FROM {tabel} aa
            JOIN proposal p ON aa.id_proposal = p.id
            WHERE aa.id = %s AND p.nim = %s
        ''', (anggaran_id, session['nim']))
        
        anggaran = cursor.fetchone()
        cursor.close()
        
        if anggaran:
            return jsonify({'success': True, 'anggaran': anggaran})
        else:
            return jsonify({'success': False, 'message': 'Data anggaran tidak ditemukan!'})
            
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error saat mengambil data: {str(e)}'})

@mahasiswa_bp.route('/update_anggaran', methods=['POST'])
def update_anggaran():
    app_funcs = get_app_functions()
    if 'user_type' not in session or session['user_type'] != 'mahasiswa':
        return jsonify({'success': False, 'message': 'Anda harus login sebagai mahasiswa!'})
    
    if not hasattr(app_funcs["mysql"], 'connection') or app_funcs["mysql"].connection is None:
        return jsonify({'success': False, 'message': 'Koneksi ke database gagal!'})
    
    try:
        data = request.get_json()
        anggaran_id = data.get('id')
        tabel = data.get('tabel', 'anggaran_awal')  # Default ke anggaran_awal
        
        if not anggaran_id:
            return jsonify({'success': False, 'message': 'ID anggaran tidak ditemukan!'})
        
        if tabel not in ['anggaran_awal', 'anggaran_bertumbuh']:
            return jsonify({'success': False, 'message': 'Tabel anggaran tidak valid!'})
        
        cursor = app_funcs["mysql"].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Verify the anggaran belongs to the logged-in student
        cursor.execute(f'''
            SELECT aa.id 
            FROM {tabel} aa
            JOIN proposal p ON aa.id_proposal = p.id
            WHERE aa.id = %s AND p.nim = %s
        ''', (anggaran_id, session['nim']))
        
        anggaran = cursor.fetchone()
        if not anggaran:
            cursor.close()
            return jsonify({'success': False, 'message': 'Data anggaran tidak ditemukan!'})
        
        # Check if this is a full update (revision/draft) or partial update
        if 'kegiatan' in data:
            # Full update for revision/draft - validate required fields (except kegiatan_utama)
            kegiatan = data.get('kegiatan', '').strip().title()
            nama_barang = data.get('nama_barang', '').strip().title()
            penanggung_jawab = data.get('penanggung_jawab', '').strip().title()
            satuan = data.get('satuan', '').strip().title()
            target_capaian = data.get('target_capaian', '').strip().title()
            
            if not kegiatan:
                return jsonify({'success': False, 'message': 'Kegiatan harus diisi!'})
            if not nama_barang:
                return jsonify({'success': False, 'message': 'Nama barang harus diisi!'})
            if not penanggung_jawab:
                return jsonify({'success': False, 'message': 'Penanggung jawab harus diisi!'})
            if not satuan:
                return jsonify({'success': False, 'message': 'Satuan harus diisi!'})
            if not target_capaian:
                return jsonify({'success': False, 'message': 'Target capaian harus diisi!'})
            
            # Validasi satuan tidak boleh kosong
            if not satuan or satuan.strip() == '':
                return jsonify({'success': False, 'message': 'Satuan harus diisi!'})
            
            # Ambil data anggaran untuk validasi nilai bantuan
            cursor.execute(f'SELECT status, nilai_bantuan FROM {tabel} WHERE id = %s', (anggaran_id,))
            anggaran_data = cursor.fetchone()
            if not anggaran_data:
                return jsonify({'success': False, 'message': 'Data anggaran tidak ditemukan!'})
            
            old_status = anggaran_data['status']
            nilai_bantuan = anggaran_data.get('nilai_bantuan', 0) or 0
            
            # Validasi jumlah tidak boleh melebihi nilai bantuan jika ada nilai bantuan
            jumlah_baru = float(data.get('jumlah', 0))
            nilai_bantuan_float = float(nilai_bantuan) if nilai_bantuan else 0
            if nilai_bantuan_float > 0 and jumlah_baru > nilai_bantuan_float:
                return jsonify({
                    'success': False, 
                    'message': f'Jumlah anggaran (Rp {jumlah_baru:,.0f}) tidak boleh melebihi nilai bantuan (Rp {nilai_bantuan_float:,.0f})!'
                })
            
            # Full update for revision/draft - update all fields except kegiatan_utama
            cursor.execute(f'''
                UPDATE {tabel} 
                SET kegiatan = %s, nama_barang = %s, penanggung_jawab = %s,
                    satuan = %s, harga_satuan = %s, target_capaian = %s, keterangan = %s,
                    kuantitas = %s, jumlah = %s
                WHERE id = %s
            ''', (
                kegiatan,
                nama_barang,
                penanggung_jawab,
                satuan,
                data.get('harga_satuan', 0),
                target_capaian,
                data.get('keterangan', '').strip(),
                data.get('kuantitas', 0),
                data.get('jumlah', 0),
                anggaran_id
            ))
            
            # Jika status sebelumnya revisi, ubah ke diajukan
            if old_status == 'revisi':
                cursor.execute(f'UPDATE {tabel} SET status = %s WHERE id = %s', ('diajukan', anggaran_id))
        else:
            # Partial update - only kuantitas and jumlah
            cursor.execute(f'''
                UPDATE {tabel} 
                SET kuantitas = %s, jumlah = %s
                WHERE id = %s
            ''', (data.get('kuantitas', 0), data.get('jumlah', 0), anggaran_id))
        
        app_funcs["mysql"].connection.commit()
        cursor.close()
        
        return jsonify({'success': True, 'message': 'Data anggaran berhasil diperbarui!'})
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error saat memperbarui data: {str(e)}'})
@mahasiswa_bp.route('/delete_anggaran', methods=['POST'])
def delete_anggaran():
    app_funcs = get_app_functions()
    if 'user_type' not in session or session['user_type'] != 'mahasiswa':
        return jsonify({'success': False, 'message': 'Anda harus login sebagai mahasiswa!'})
    
    if not hasattr(app_funcs["mysql"], 'connection') or app_funcs["mysql"].connection is None:
        return jsonify({'success': False, 'message': 'Koneksi ke database gagal!'})
    
    try:
        data = request.get_json()
        anggaran_id = data.get('id')
        tabel = data.get('tabel')
        
        if not anggaran_id or not tabel:
            return jsonify({'success': False, 'message': 'Data tidak lengkap!'})
        
        if tabel not in ['anggaran_awal', 'anggaran_bertumbuh']:
            return jsonify({'success': False, 'message': 'Tabel tidak valid!'})
        
        cursor = app_funcs["mysql"].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Cek apakah anggaran milik mahasiswa ini
        cursor.execute(f'''
            SELECT a.*, p.nim, p.id as proposal_id, p.judul_usaha, p.tahapan_usaha
            FROM {tabel} a
            JOIN proposal p ON a.id_proposal = p.id
            WHERE a.id = %s AND p.nim = %s
        ''', (anggaran_id, session['nim']))
        
        anggaran = cursor.fetchone()
        if not anggaran:
            cursor.close()
            return jsonify({'success': False, 'message': 'Anggaran tidak ditemukan atau Anda tidak memiliki akses!'})
        
        proposal_id = anggaran['proposal_id']
        tahapan_usaha = anggaran['tahapan_usaha']
        
        # Tentukan tabel laporan berdasarkan tahapan usaha
        if 'bertumbuh' in tahapan_usaha.lower():
            tabel_laporan_kemajuan = 'laporan_kemajuan_bertumbuh'
            tabel_laporan_akhir = 'laporan_akhir_bertumbuh'
        else:
            tabel_laporan_kemajuan = 'laporan_kemajuan_awal'
            tabel_laporan_akhir = 'laporan_akhir_awal'
        
        # Hapus HANYA data laporan kemajuan yang sesuai dengan anggaran yang dihapus (berdasarkan nama_barang)
        # Tidak menghapus semua data laporan, hanya yang terkait dengan anggaran ini
        cursor.execute(f'''
            DELETE FROM {tabel_laporan_kemajuan} 
            WHERE id_proposal = %s 
            AND nama_barang = %s 
            AND kegiatan_utama = %s 
            AND kegiatan = %s
        ''', (proposal_id, anggaran['nama_barang'], anggaran['kegiatan_utama'], anggaran['kegiatan']))
        laporan_kemajuan_deleted = cursor.rowcount
        print(f"✅ Berhasil menghapus {laporan_kemajuan_deleted} data laporan kemajuan yang sesuai untuk: {anggaran['nama_barang']}")
        
        # Hapus HANYA data laporan akhir yang sesuai dengan anggaran yang dihapus (berdasarkan nama_barang)
        # Tidak menghapus semua data laporan, hanya yang terkait dengan anggaran ini
        cursor.execute(f'''
            DELETE FROM {tabel_laporan_akhir} 
            WHERE id_proposal = %s 
            AND nama_barang = %s 
            AND kegiatan_utama = %s 
            AND kegiatan = %s
        ''', (proposal_id, anggaran['nama_barang'], anggaran['kegiatan_utama'], anggaran['kegiatan']))
        laporan_akhir_deleted = cursor.rowcount
        print(f"✅ Berhasil menghapus {laporan_akhir_deleted} data laporan akhir yang sesuai untuk: {anggaran['nama_barang']}")
        
        # TIDAK menghapus file laporan kemajuan dan akhir secara otomatis
        # karena hanya 1 item anggaran yang dihapus, file laporan masih diperlukan untuk item lainnya
        # File hanya dihapus jika mahasiswa secara manual menghapusnya dari halaman laporan
        print(f"ℹ️  File laporan kemajuan dan akhir TIDAK dihapus karena masih ada item anggaran lainnya")
        
        # Hapus anggaran
        cursor.execute(f'DELETE FROM {tabel} WHERE id = %s', (anggaran_id,))
        anggaran_deleted = cursor.rowcount
        
        app_funcs["mysql"].connection.commit()
        cursor.close()
        
        message = f'Anggaran "{anggaran["nama_barang"]}" berhasil dihapus!'
        if laporan_kemajuan_deleted > 0:
            message += f' Data laporan kemajuan terkait "{anggaran["nama_barang"]}" juga dihapus.'
        if laporan_akhir_deleted > 0:
            message += f' Data laporan akhir terkait "{anggaran["nama_barang"]}" juga dihapus.'
        if laporan_kemajuan_deleted > 0 or laporan_akhir_deleted > 0:
            message += ' (Data laporan lainnya tetap aman!)'
        
        return jsonify({'success': True, 'message': message})
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@mahasiswa_bp.route('/update_laporan_kemajuan', methods=['POST'])
def update_laporan_kemajuan():
    app_funcs = get_app_functions()
    if 'user_type' not in session or session['user_type'] != 'mahasiswa':
        return jsonify({'success': False, 'message': 'Anda harus login sebagai mahasiswa!'})
    if not hasattr(app_funcs["mysql"], 'connection') or app_funcs["mysql"].connection is None:
        return jsonify({'success': False, 'message': 'Koneksi ke database gagal!'})
    
    # Validasi jadwal kemajuan sebelum mengizinkan update
    jadwal_status = get_jadwal_kemajuan_status()
    if not jadwal_status['bisa_upload']:
        return jsonify({'success': False, 'message': f"Tidak dapat mengupdate laporan kemajuan. {jadwal_status['pesan']}"})
    
    try:
        if request.content_type and request.content_type.startswith('multipart/form-data'):
            # Jika upload file
            laporan_id = request.form.get('id')
            tabel_laporan = request.form.get('tabel_laporan')
            kuantitas = request.form.get('kuantitas')
            harga_satuan = request.form.get('harga_satuan')
            jumlah = request.form.get('jumlah')
            status = request.form.get('status', 'draf')
            file = request.files.get('nota_file')
        else:
            data = request.get_json()
            laporan_id = data.get('id')
            tabel_laporan = data.get('tabel_laporan')
            kuantitas = data.get('kuantitas')
            harga_satuan = data.get('harga_satuan')
            jumlah = data.get('jumlah')
            status = data.get('status', 'draf')
            file = None
        if not laporan_id or not tabel_laporan or kuantitas is None or harga_satuan is None or jumlah is None:
            return jsonify({'success': False, 'message': 'Data tidak lengkap!'})
        if tabel_laporan not in ['laporan_kemajuan_awal', 'laporan_kemajuan_bertumbuh']:
            return jsonify({'success': False, 'message': 'Tabel laporan tidak valid!'})
        try:
            kuantitas = int(kuantitas)
            harga_satuan = int(harga_satuan)
            jumlah = int(jumlah)
        except ValueError:
            return jsonify({'success': False, 'message': 'Kuantitas, harga satuan, dan jumlah harus berupa angka!'})
        if kuantitas <= 0 or harga_satuan <= 0 or jumlah <= 0:
            return jsonify({'success': False, 'message': 'Kuantitas, harga satuan, dan jumlah harus lebih dari 0!'})
        cursor = app_funcs["mysql"].connection.cursor(MySQLdb.cursors.DictCursor)
        # Cek apakah laporan kemajuan milik mahasiswa ini
        cursor.execute(f'''
            SELECT lk.*, p.nim, p.judul_usaha, m.nama_ketua, m.perguruan_tinggi
            FROM {tabel_laporan} lk
            JOIN proposal p ON lk.id_proposal = p.id
            JOIN mahasiswa m ON p.nim = m.nim
            WHERE lk.id = %s AND p.nim = %s
        ''', (laporan_id, session['nim']))
        laporan = cursor.fetchone()
        if not laporan:
            cursor.close()
            return jsonify({'success': False, 'message': 'Laporan kemajuan tidak ditemukan atau Anda tidak memiliki akses!'})
        keterangan_path = None
        if file:
            if not file.filename or not isinstance(file.filename, str) or not file.filename.lower().endswith('.pdf'):
                cursor.close()
                return jsonify({'success': False, 'message': 'File nota harus berupa PDF!'})
            # Buat path upload: static/uploads/kemajuan/judul_usaha/nama_mahasiswa/
            nama_ketua = laporan['nama_ketua']
            id_laporan_kemajuan = laporan['id']
            upload_dir = os.path.join('static', 'uploads', 'Laporan Kemajuan', 'Nota')
            os.makedirs(upload_dir, exist_ok=True)
            safe_nama_ketua = re.sub(r'[^\w\s-]', '', nama_ketua).strip().replace(' ', '_')
            filename = f"Nota Transaksi_{safe_nama_ketua}_{id_laporan_kemajuan}.pdf"
            file_path = os.path.join(upload_dir, filename)
            # Hapus file lama jika ada dan berbeda path
            file_lama = laporan.get('keterangan')
            if file_lama and file_lama.endswith('.pdf'):
                file_lama_path = os.path.join('static', file_lama.replace('static/', '').replace('/', os.sep))
                if os.path.exists(file_lama_path) and os.path.abspath(file_lama_path) != os.path.abspath(file_path):
                    try:
                        os.remove(file_lama_path)
                    except Exception:
                        pass
            file.save(file_path)
            keterangan_path = os.path.relpath(file_path, 'static')
            keterangan_path = keterangan_path.replace("\\", "/")
            keterangan_path = 'static/' + keterangan_path
        # Update laporan kemajuan
        if keterangan_path:
            cursor.execute(f'''
                UPDATE {tabel_laporan} 
                SET kuantitas = %s, harga_satuan = %s, jumlah = %s, status = %s, keterangan = %s
                WHERE id = %s
            ''', (kuantitas, harga_satuan, jumlah, status, keterangan_path, laporan_id))
        else:
            cursor.execute(f'''
                UPDATE {tabel_laporan} 
                SET kuantitas = %s, harga_satuan = %s, jumlah = %s, status = %s
                WHERE id = %s
            ''', (kuantitas, harga_satuan, jumlah, status, laporan_id))
        if cursor.rowcount == 0:
            cursor.close()
            return jsonify({'success': False, 'message': 'Gagal mengupdate laporan kemajuan!'})
        app_funcs["mysql"].connection.commit()
        cursor.close()
        return jsonify({'success': True, 'message': 'Laporan kemajuan berhasil diperbarui!'})
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@mahasiswa_bp.route('/get_proposal_detail/<int:proposal_id>')
def get_proposal_detail(proposal_id):
    app_funcs = get_app_functions()
    if 'user_type' not in session or session['user_type'] != 'mahasiswa':
        return jsonify({'success': False, 'message': 'Anda harus login sebagai mahasiswa!'})
    
    if not hasattr(app_funcs["mysql"], 'connection') or app_funcs["mysql"].connection is None:
        return jsonify({'success': False, 'message': 'Koneksi ke database gagal!'})
    
    try:
        cursor = app_funcs["mysql"].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Get proposal details
        cursor.execute('''
            SELECT * FROM proposal 
            WHERE id = %s AND nim = %s
        ''', (proposal_id, session['nim']))
        
        proposal = cursor.fetchone()
        if not proposal:
            cursor.close()
            return jsonify({'success': False, 'message': 'Proposal tidak ditemukan!'})
        
        # Get team members
        cursor.execute('''
            SELECT * FROM anggota_tim 
            WHERE id_proposal = %s
            ORDER BY id
        ''', (proposal_id,))
        
        anggota = cursor.fetchall()
        cursor.close()
        
        return jsonify({
            'success': True,
            'proposal': proposal,
            'anggota': anggota
        })
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error saat mengambil data: {str(e)}'})

@mahasiswa_bp.route('/get_anggaran_data/<int:proposal_id>')
def get_anggaran_data(proposal_id):
    app_funcs = get_app_functions()
    """
    Endpoint untuk mendapatkan data anggaran yang sudah direview oleh reviewer
    """
    if 'user_type' not in session or session['user_type'] != 'mahasiswa':
        return jsonify({'success': False, 'message': 'Anda harus login sebagai mahasiswa!'})
    
    if not hasattr(app_funcs["mysql"], 'connection') or app_funcs["mysql"].connection is None:
        return jsonify({'success': False, 'message': 'Koneksi ke database gagal!'})
    
    try:
        cursor = app_funcs["mysql"].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Ambil data mahasiswa untuk menentukan tabel anggaran
        cursor.execute('''
    SELECT p.tahapan_usaha, p.id as proposal_id
    FROM mahasiswa m 
    LEFT JOIN proposal p ON m.nim = p.nim
    WHERE m.nim = %s AND p.id = %s
        ''', (session['nim'], proposal_id))
        
        mahasiswa_info = cursor.fetchone()
        if not mahasiswa_info:
            return jsonify({'success': False, 'message': 'Data tidak ditemukan!'})
        
        # Tentukan tabel anggaran berdasarkan tahapan usaha
        tahapan_usaha = mahasiswa_info.get('tahapan_usaha', '').lower()
        if 'bertumbuh' in tahapan_usaha:
            tabel_anggaran = 'anggaran_bertumbuh'
        else:
            tabel_anggaran = 'anggaran_awal'
        
        # Ambil data anggaran yang sudah direview oleh reviewer (status_reviewer = 'sudah_direview')
        cursor.execute(f'''
            SELECT id, kegiatan_utama, kegiatan, penanggung_jawab, target_capaian, 
                   nama_barang, satuan, status, nilai_bantuan
            FROM {tabel_anggaran} 
            WHERE id_proposal = %s AND status_reviewer = 'sudah_direview'
            ORDER BY kegiatan_utama, kegiatan, nama_barang
        ''', (proposal_id,))
        
        anggaran_data = cursor.fetchall()
        cursor.close()
        
        return jsonify({'success': True, 'data': anggaran_data})
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@mahasiswa_bp.route('/get_jenis_revisi/<int:proposal_id>')
def get_jenis_revisi(proposal_id):
    app_funcs = get_app_functions()
    """Mengambil informasi jenis revisi proposal (file atau anggaran)"""
    if 'user_type' not in session or session['user_type'] != 'mahasiswa':
        return jsonify({'success': False, 'message': 'Anda harus login sebagai mahasiswa!'})
    
    if not hasattr(app_funcs["mysql"], 'connection') or app_funcs["mysql"].connection is None:
        return jsonify({'success': False, 'message': 'Koneksi ke database gagal!'})
    
    try:
        cursor = app_funcs["mysql"].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Cek apakah ada anggaran yang status revisi
        cursor.execute('''
            SELECT COUNT(*) as total_revisi
            FROM (
                SELECT status FROM anggaran_awal WHERE id_proposal = %s AND status = 'revisi'
                UNION ALL
                SELECT status FROM anggaran_bertumbuh WHERE id_proposal = %s AND status = 'revisi'
            ) as combined_anggaran
        ''', (proposal_id, proposal_id))
        anggaran_revisi_count = cursor.fetchone()['total_revisi']
        
        # Cek apakah ada komentar revisi dari pembimbing
        cursor.execute('''
            SELECT komentar_revisi 
            FROM proposal 
            WHERE id = %s AND nim = %s
        ''', (proposal_id, session['nim']))
        proposal_data = cursor.fetchone()
        
        komentar_revisi = proposal_data.get('komentar_revisi') if proposal_data else None
        
        cursor.close()
        
        # Tentukan jenis revisi
        jenis_revisi = []
        if anggaran_revisi_count > 0:
            jenis_revisi.append('anggaran')
        if komentar_revisi and komentar_revisi.strip():
            jenis_revisi.append('file')
        
        return jsonify({
            'success': True,
            'jenis_revisi': jenis_revisi,
            'anggaran_revisi_count': anggaran_revisi_count,
            'ada_komentar_revisi': bool(komentar_revisi and komentar_revisi.strip())
        })
        
    except Exception as e:
        print(f"Error getting jenis revisi: {e}")
        return jsonify({'success': False, 'message': 'Terjadi kesalahan saat mengambil data jenis revisi'})

@mahasiswa_bp.route('/get_anggaran_revisi/<int:proposal_id>')
def get_anggaran_revisi(proposal_id):
    app_funcs = get_app_functions()
    """Mengambil data anggaran yang perlu direvisi untuk ditampilkan di SweetAlert"""
    if 'user_type' not in session or session['user_type'] != 'mahasiswa':
        return jsonify({'success': False, 'message': 'Anda harus login sebagai mahasiswa!'})
    
    if not hasattr(app_funcs["mysql"], 'connection') or app_funcs["mysql"].connection is None:
        return jsonify({'success': False, 'message': 'Koneksi ke database gagal!'})
    
    try:
        cursor = app_funcs["mysql"].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Ambil anggaran awal yang status revisi
        cursor.execute('''
            SELECT 
                nama_barang,
                jumlah,
                COALESCE(nilai_bantuan, 0) as nilai_bantuan,
                'anggaran_awal' as jenis
            FROM anggaran_awal 
            WHERE id_proposal = %s AND status = 'revisi'
        ''', (proposal_id,))
        anggaran_awal_revisi = cursor.fetchall()
        
        # Ambil anggaran bertumbuh yang status revisi
        cursor.execute('''
            SELECT 
                nama_barang,
                jumlah,
                COALESCE(nilai_bantuan, 0) as nilai_bantuan,
                'anggaran_bertumbuh' as jenis
            FROM anggaran_bertumbuh 
            WHERE id_proposal = %s AND status = 'revisi'
        ''', (proposal_id,))
        anggaran_bertumbuh_revisi = cursor.fetchall()
        
        # Gabungkan semua anggaran revisi
        anggaran_revisi = anggaran_awal_revisi + anggaran_bertumbuh_revisi
        
        cursor.close()
        
        return jsonify({
            'success': True, 
            'anggaran_revisi': anggaran_revisi,
            'total_revisi': len(anggaran_revisi)
        })
        
    except Exception as e:
        print(f"Error getting anggaran revisi: {e}")
        return jsonify({'success': False, 'message': 'Terjadi kesalahan saat mengambil data anggaran revisi'})

@mahasiswa_bp.route('/update_laporan_akhir', methods=['POST'])
def update_laporan_akhir():
    app_funcs = get_app_functions()
    if 'user_type' not in session or session['user_type'] != 'mahasiswa':
        return jsonify({'success': False, 'message': 'Anda harus login sebagai mahasiswa!'})
    
    if not hasattr(app_funcs["mysql"], 'connection') or app_funcs["mysql"].connection is None:
        return jsonify({'success': False, 'message': 'Koneksi ke database gagal!'})
    
    try:
        # Guard jadwal server-side: lebih fleksibel
        try:
            auto_reject_expired_akhir_for_current_student()
            cursor_guard = app_funcs["mysql"].connection.cursor(MySQLdb.cursors.DictCursor)
            cursor_guard.execute('SELECT akhir_mulai, akhir_selesai FROM pengaturan_jadwal ORDER BY id DESC LIMIT 1')
            row_guard = cursor_guard.fetchone()
            cursor_guard.close()
            
            from datetime import datetime
            now_dt = datetime.now()
            
            # Jika ada jadwal yang ditentukan
            if row_guard and (row_guard.get('akhir_mulai') or row_guard.get('akhir_selesai')):
                # Cek apakah sudah melewati jadwal selesai
                if row_guard.get('akhir_selesai') and now_dt > row_guard['akhir_selesai']:
                    # Jika sudah melewati jadwal selesai, cek status laporan yang ada
                    # Status selain 'disetujui' akan otomatis ditolak
                    pass  # Lanjutkan ke proses update, auto-reject akan dihandle di bawah
                elif row_guard.get('akhir_mulai') and now_dt < row_guard['akhir_mulai']:
                    # Jika belum masuk jadwal mulai
                    return jsonify({'success': False, 'message': 'Belum masuk jadwal laporan akhir. Silakan tunggu jadwal dimulai.'})
            # Jika tidak ada jadwal, izinkan penginputan kapan saja
            # Jika ada jadwal dan dalam rentang waktu, izinkan penginputan
            
        except Exception as e:
            print(f"=== DEBUG: Guard jadwal error: {str(e)} ===")
            # Jika guard gagal, tetap lanjutkan validasi biasa agar tidak mematikan alur
            pass

        # Debug: Cek database yang diakses
        cursor = app_funcs["mysql"].connection.cursor()
        cursor.execute('SELECT DATABASE()')
        current_db = cursor.fetchone()[0]
        print(f"=== DEBUG: Flask mengakses database: {current_db} ===")
        cursor.close()
        # Debug: Log semua data yang diterima
        print("=== DEBUG: Data yang diterima ===")
        print("Form data:", dict(request.form))
        print("Files:", [f.filename for f in request.files.values()] if request.files else "No files")
        
        laporan_id = request.form.get('id')
        tabel_laporan = request.form.get('tabel_laporan')
        kuantitas = request.form.get('kuantitas')
        harga_satuan = request.form.get('harga_satuan')
        jumlah = request.form.get('jumlah')
        keterangan = request.form.get('keterangan', '')
        status = request.form.get('status', 'diajukan')
        nota_file = request.files.get('nota_file')
        
        print("=== DEBUG: Data yang diproses ===")
        print(f"laporan_id: {laporan_id}")
        print(f"tabel_laporan: {tabel_laporan}")
        print(f"kuantitas: {kuantitas}")
        print(f"harga_satuan: {harga_satuan}")
        print(f"jumlah: {jumlah}")
        print(f"status: {status}")
        print(f"nota_file: {nota_file.filename if nota_file else 'None'}")
        
        # Validasi data yang diterima
        print("=== DEBUG: Validasi data ===")
        print(f"laporan_id valid: {laporan_id is not None and laporan_id != ''}")
        print(f"tabel_laporan valid: {tabel_laporan is not None and tabel_laporan != ''}")
        print(f"kuantitas valid: {kuantitas is not None and kuantitas != ''}")
        print(f"harga_satuan valid: {harga_satuan is not None and harga_satuan != ''}")
        print(f"jumlah valid: {jumlah is not None and jumlah != ''}")
        print(f"nota_file valid: {nota_file is not None}")
        
        if not laporan_id or not tabel_laporan or kuantitas is None or harga_satuan is None or jumlah is None:
            print("=== DEBUG: Data tidak lengkap ===")
            print(f"laporan_id: '{laporan_id}'")
            print(f"tabel_laporan: '{tabel_laporan}'")
            print(f"kuantitas: '{kuantitas}'")
            print(f"harga_satuan: '{harga_satuan}'")
            print(f"jumlah: '{jumlah}'")
            return jsonify({'success': False, 'message': 'Data tidak lengkap!'})
        
        if tabel_laporan not in ['laporan_akhir_awal', 'laporan_akhir_bertumbuh']:
            return jsonify({'success': False, 'message': 'Tabel laporan tidak valid!'})
        
        # Validasi input
        try:
            kuantitas = int(kuantitas)
            harga_satuan = int(harga_satuan)
            jumlah = int(jumlah)
        except ValueError:
            return jsonify({'success': False, 'message': 'Kuantitas, harga satuan, dan jumlah harus berupa angka!'})
        
        if kuantitas <= 0 or harga_satuan <= 0 or jumlah <= 0:
            return jsonify({'success': False, 'message': 'Kuantitas, harga satuan, dan jumlah harus lebih dari 0!'})
        
        if not nota_file:
            return jsonify({'success': False, 'message': 'File nota harus diupload!'})
        
        if not allowed_file(nota_file.filename):
            return jsonify({'success': False, 'message': 'Hanya file PDF yang diperbolehkan!'})
        
        cursor = app_funcs["mysql"].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Debug: Cek data laporan sebelum update
        print(f"=== DEBUG: Mencari laporan dengan id={laporan_id} dan nim={session['nim']} ===")
        
        # Cek apakah laporan akhir milik mahasiswa ini
        cursor.execute(f'''
            SELECT la.*, p.nim, p.id as proposal_id, p.judul_usaha, m.perguruan_tinggi, m.nama_ketua
            FROM {tabel_laporan} la
            JOIN proposal p ON la.id_proposal = p.id
            JOIN mahasiswa m ON p.nim = m.nim
            WHERE la.id = %s AND p.nim = %s
        ''', (laporan_id, session['nim']))
        
        laporan = cursor.fetchone()
        if not laporan:
            print(f"=== DEBUG: Laporan tidak ditemukan untuk id={laporan_id} ===")
            # Cek apakah baris dengan id tersebut ada di database
            cursor.execute(f'SELECT * FROM {tabel_laporan} WHERE id = %s', (laporan_id,))
            row_exists = cursor.fetchone()
            if row_exists:
                print(f"=== DEBUG: Baris dengan id={laporan_id} ada di database, tapi bukan milik mahasiswa ini ===")
                print(f"Baris yang ada: {row_exists}")
            else:
                print(f"=== DEBUG: Baris dengan id={laporan_id} tidak ada di database sama sekali ===")
            cursor.close()
            return jsonify({'success': False, 'message': 'Laporan akhir tidak ditemukan atau Anda tidak memiliki akses!'})
        
        print(f"=== DEBUG: Laporan ditemukan ===")
        print(f"ID: {laporan['id']}")
        print(f"Nama Barang: {laporan['nama_barang']}")
        print(f"Proposal ID: {laporan['proposal_id']}")
        print(f"Judul Usaha: {laporan['judul_usaha']}")
        print(f"Perguruan Tinggi: {laporan['perguruan_tinggi']}")
        print(f"Nama Ketua: {laporan['nama_ketua']}")
        
        # Cek jadwal lagi untuk menentukan status yang akan disimpan
        cursor_jadwal = app_funcs["mysql"].connection.cursor(MySQLdb.cursors.DictCursor)
        cursor_jadwal.execute('SELECT akhir_mulai, akhir_selesai FROM pengaturan_jadwal ORDER BY id DESC LIMIT 1')
        jadwal_info = cursor_jadwal.fetchone()
        cursor_jadwal.close()
        
        # Tentukan status berdasarkan jadwal
        final_status = status
        if jadwal_info and jadwal_info.get('akhir_selesai'):
            from datetime import datetime
            now_dt = datetime.now()
            if now_dt > jadwal_info['akhir_selesai']:
                # Jika sudah melewati jadwal selesai, status otomatis menjadi 'ditolak'
                final_status = 'ditolak'
                print(f"=== DEBUG: Sudah melewati jadwal selesai, status otomatis menjadi: {final_status} ===")
        
        # Upload file dengan penamaan yang lebih baik
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        judul_usaha_clean = laporan['judul_usaha'].replace(' ', '_').replace('/', '_').replace('\\', '_').replace(':', '_')
        laporan_id_clean = str(laporan_id)
        
        # Format: Bukti_judul_usaha_id_laporan_akhir.pdf
        filename = f"Bukti_{judul_usaha_clean}_{laporan_id_clean}_laporan_akhir.pdf"
        
        # Buat direktori untuk Laporan Akhir
        base_upload_dir = os.path.join('static', 'uploads', 'Laporan Akhir')
        judul_usaha_dir = os.path.join(base_upload_dir, judul_usaha_clean)
        
        # Buat direktori jika belum ada
        os.makedirs(judul_usaha_dir, exist_ok=True)
        
        # Path fisik untuk menyimpan file
        physical_path = os.path.join(judul_usaha_dir, filename)
        
        # Path untuk database (tanpa 'static/' di awal)
        db_path = os.path.join('uploads', 'Laporan Akhir', judul_usaha_clean, filename)
        
        print(f"=== DEBUG: File path info ===")
        print(f"Physical path: {physical_path}")
        print(f"Database path: {db_path}")
        print(f"Filename: {filename}")
        
        # Simpan file
        nota_file.save(physical_path)
        
        # Cek data sebelum update
        cursor.execute(f'SELECT kuantitas, harga_satuan, jumlah, keterangan, status FROM {tabel_laporan} WHERE id = %s', (laporan_id,))
        data_sebelum = cursor.fetchone()
        print(f"=== DEBUG: Data sebelum update ===")
        print(f"Kuantitas: {data_sebelum['kuantitas'] if data_sebelum else 'None'}")
        print(f"Harga Satuan: {data_sebelum['harga_satuan'] if data_sebelum else 'None'}")
        print(f"Jumlah: {data_sebelum['jumlah'] if data_sebelum else 'None'}")
        print(f"Keterangan: {data_sebelum['keterangan'] if data_sebelum else 'None'}")
        print(f"Status: {data_sebelum['status'] if data_sebelum else 'None'}")
        
        # Lakukan update dengan query yang lebih sederhana
        print(f"=== DEBUG: Melakukan update dengan query ===")
        update_query = f"UPDATE {tabel_laporan} SET kuantitas = %s, harga_satuan = %s, jumlah = %s, keterangan = %s, status = %s WHERE id = %s"
        update_params = (kuantitas, harga_satuan, jumlah, db_path, final_status, laporan_id)
        print(f"Query: {update_query}")
        print(f"Parameters: {update_params}")
        
        cursor.execute(update_query, update_params)
        
        rows_affected = cursor.rowcount
        print(f"=== DEBUG: Rows affected: {rows_affected} ===")
        
        # Cek data setelah update
        cursor.execute(f'SELECT kuantitas, harga_satuan, jumlah, keterangan, status FROM {tabel_laporan} WHERE id = %s', (laporan_id,))
        data_sesudah = cursor.fetchone()
        print(f"=== DEBUG: Data setelah update ===")
        print(f"Kuantitas: {data_sesudah['kuantitas'] if data_sesudah else 'None'}")
        print(f"Harga Satuan: {data_sesudah['harga_satuan'] if data_sesudah else 'None'}")
        print(f"Jumlah: {data_sesudah['jumlah'] if data_sesudah else 'None'}")
        print(f"Keterangan: {data_sesudah['keterangan'] if data_sesudah else 'None'}")
        print(f"Status: {data_sesudah['status'] if data_sesudah else 'None'}")
        
        print(f"=== DEBUG: Melakukan commit ===")
        app_funcs["mysql"].connection.commit()
        print(f"=== DEBUG: Commit berhasil ===")
        
        cursor.close()
        
        if rows_affected > 0:
            print("=== DEBUG: Update berhasil ===")
            # Verifikasi final bahwa data benar-benar tersimpan
            cursor = app_funcs["mysql"].connection.cursor(MySQLdb.cursors.DictCursor)
            cursor.execute(f'SELECT kuantitas, harga_satuan, jumlah, keterangan, status FROM {tabel_laporan} WHERE id = %s', (laporan_id,))
            final_verification = cursor.fetchone()
            cursor.close()
            
            print(f"=== DEBUG: Verifikasi final ===")
            print(f"Kuantitas: {final_verification['kuantitas'] if final_verification else 'None'}")
            print(f"Harga Satuan: {final_verification['harga_satuan'] if final_verification else 'None'}")
            print(f"Jumlah: {final_verification['jumlah'] if final_verification else 'None'}")
            print(f"Keterangan: {final_verification['keterangan'] if final_verification else 'None'}")
            print(f"Status: {final_verification['status'] if final_verification else 'None'}")
            
            # Cek apakah data benar-benar berubah
            if (final_verification and 
                final_verification['kuantitas'] == kuantitas and 
                final_verification['harga_satuan'] == harga_satuan and 
                final_verification['jumlah'] == jumlah and 
                final_verification['status'] == final_status):
                print("=== DEBUG: Data berhasil tersimpan dengan benar ===")
                # Pesan yang berbeda berdasarkan status
                if final_status == 'ditolak':
                    message = 'Data laporan akhir berhasil disimpan, namun status otomatis menjadi ditolak karena sudah melewati jadwal yang ditentukan.'
                else:
                    message = 'Data laporan akhir berhasil disimpan dan status diubah menjadi diajukan!'
                
                return jsonify({'success': True, 'message': message})
            else:
                print("=== DEBUG: Data tidak tersimpan dengan benar ===")
                return jsonify({'success': False, 'message': 'Data tidak tersimpan dengan benar!'})
        else:
            print("=== DEBUG: Update gagal - tidak ada baris yang terpengaruh ===")
            # Cek lagi apakah data benar-benar tidak berubah
            cursor = app_funcs["mysql"].connection.cursor(MySQLdb.cursors.DictCursor)
            cursor.execute(f'SELECT kuantitas, harga_satuan, jumlah, keterangan, status FROM {tabel_laporan} WHERE id = %s', (laporan_id,))
            final_check = cursor.fetchone()
            cursor.close()
            print(f"=== DEBUG: Data final check: {final_check} ===")
            return jsonify({'success': False, 'message': 'Gagal memperbarui data laporan akhir!'})
        
    except Exception as e:
        print(f"=== DEBUG: Exception terjadi: {str(e)} ===")
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@mahasiswa_bp.route('/download_laporan_kemajuan/<int:proposal_id>')
def mahasiswa_download_laporan_kemajuan(proposal_id):
    app_funcs = get_app_functions()
    if 'user_type' not in session or session['user_type'] != 'mahasiswa':
        flash('Anda harus login sebagai mahasiswa!', 'danger')
        return redirect(url_for('mahasiswa.proposal'))
    
    if not hasattr(app_funcs["mysql"], 'connection') or app_funcs["mysql"].connection is None:
        flash('Koneksi ke database gagal!', 'danger')
        return redirect(url_for('mahasiswa.proposal'))
    
    try:
        cursor = app_funcs["mysql"].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Ambil path file laporan kemajuan dan informasi mahasiswa
        cursor.execute('''
            SELECT p.laporan_kemajuan_path, m.id as mahasiswa_id, m.nama_ketua, m.nim
            FROM proposal p
            JOIN mahasiswa m ON p.nim = m.nim
            WHERE p.id = %s AND p.nim = %s
        ''', (proposal_id, session['nim']))
        
        proposal = cursor.fetchone()
        cursor.close()
        
        if not proposal or not proposal['laporan_kemajuan_path']:
            flash('File laporan kemajuan tidak ditemukan!', 'danger')
            return redirect(url_for('mahasiswa.proposal'))
        
        file_path = proposal['laporan_kemajuan_path']
        # Normalisasi path: ganti backslash ke slash dan awali dengan static/ jika belum ada
        file_path = file_path.replace('\\', '/').replace('\\', '/')
        if not file_path.startswith('static/'):
            file_path = 'static/' + file_path.lstrip('/')
        # Cek apakah file ada
        if not os.path.exists(file_path):
            flash(f'File laporan kemajuan tidak ditemukan di server! ({file_path})', 'danger')
            return redirect(url_for('mahasiswa.proposal'))
        
        # Log aktivitas download
        mahasiswa_info = {
            'id': proposal['mahasiswa_id'],
            'nim': proposal['nim'],
            'nama_ketua': proposal['nama_ketua']
        }
        log_mahasiswa_download_activity(app_funcs, mahasiswa_info, 'laporan_kemajuan', proposal_id)
        
        # Deteksi ekstensi file dan mimetype
        ext = os.path.splitext(file_path)[1].lower()
        if ext == '.pdf':
            return send_file(file_path, as_attachment=False, mimetype='application/pdf')
        else:
            # Deteksi mimetype otomatis untuk file selain PDF
            import mimetypes
            mimetype = mimetypes.guess_type(file_path)[0] or 'application/octet-stream'
            return send_file(file_path, as_attachment=False, mimetype=mimetype)
        
    except Exception as e:
        flash(f'Error saat mengakses file: {str(e)}', 'danger')
        return redirect(url_for('mahasiswa.proposal'))
@mahasiswa_bp.route('/download_proposal/<int:proposal_id>')
def mahasiswa_download_proposal(proposal_id):
    app_funcs = get_app_functions()
    if 'user_type' not in session or session['user_type'] != 'mahasiswa':
        flash('Anda harus login sebagai mahasiswa!', 'danger')
        return redirect(url_for('mahasiswa.proposal'))
    
    if not hasattr(app_funcs["mysql"], 'connection') or app_funcs["mysql"].connection is None:
        flash('Koneksi ke database gagal!', 'danger')
        return redirect(url_for('mahasiswa.proposal'))
    
    try:
        cursor = app_funcs["mysql"].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Ambil path file proposal dan informasi mahasiswa
        cursor.execute('''
            SELECT p.file_path, m.id as mahasiswa_id, m.nama_ketua, m.nim
            FROM proposal p
            JOIN mahasiswa m ON p.nim = m.nim
            WHERE p.id = %s AND p.nim = %s
        ''', (proposal_id, session['nim']))
        
        proposal_data = cursor.fetchone()
        cursor.close()
        
        if not proposal_data or not proposal_data['file_path']:
            flash('File proposal tidak ditemukan!', 'danger')
            return redirect(url_for('mahasiswa.proposal'))
        
        file_path = proposal_data['file_path']
        
        # Cek apakah file ada
        if not os.path.exists(file_path):
            flash('File proposal tidak ditemukan di server!', 'danger')
            return redirect(url_for('mahasiswa.proposal'))
        
        # Log aktivitas download
        mahasiswa_info = {
            'id': proposal_data['mahasiswa_id'],
            'nim': proposal_data['nim'],
            'nama_ketua': proposal_data['nama_ketua']
        }
        log_mahasiswa_download_activity(app_funcs, mahasiswa_info, 'proposal', proposal_id)
        
        # Ambil nama file dari path
        filename = os.path.basename(file_path)
        
        # Deteksi ekstensi file dan mimetype
        ext = os.path.splitext(file_path)[1].lower()
        if ext == '.pdf':
            return send_file(file_path, as_attachment=False, mimetype='application/pdf')
        else:
            # Deteksi mimetype otomatis untuk file selain PDF
            import mimetypes
            mimetype = mimetypes.guess_type(file_path)[0] or 'application/octet-stream'
            return send_file(file_path, as_attachment=False, mimetype=mimetype)
        
    except Exception as e:
        flash(f'Error saat mengunduh file proposal: {str(e)}', 'danger')
        return redirect(url_for('mahasiswa.proposal'))

@mahasiswa_bp.route('/submit_laporan_akhir', methods=['POST'])
def submit_laporan_akhir():
    app_funcs = get_app_functions()
    if 'user_type' not in session or session['user_type'] != 'mahasiswa':
        return jsonify({'success': False, 'message': 'Anda harus login sebagai mahasiswa!'})
    
    if not hasattr(app_funcs["mysql"], 'connection') or app_funcs["mysql"].connection is None:
        return jsonify({'success': False, 'message': 'Koneksi ke database gagal!'})
    
    try:
        data = request.get_json()
        proposal_id = data.get('proposal_id')
        tabel_laporan = data.get('tabel_laporan')
        
        if not proposal_id or not tabel_laporan:
            return jsonify({'success': False, 'message': 'Data tidak lengkap!'})
        
        if tabel_laporan not in ['laporan_akhir_awal', 'laporan_akhir_bertumbuh']:
            return jsonify({'success': False, 'message': 'Tabel laporan tidak valid!'})
        
        cursor = app_funcs["mysql"].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Cek apakah proposal milik mahasiswa ini
        cursor.execute('''
            SELECT p.*, m.nim 
            FROM proposal p
            JOIN mahasiswa m ON p.nim = m.nim
            WHERE p.id = %s AND m.nim = %s
        ''', (proposal_id, session['nim']))
        
        proposal = cursor.fetchone()
        if not proposal:
            cursor.close()
            return jsonify({'success': False, 'message': 'Proposal tidak ditemukan atau Anda tidak memiliki akses!'})
        
        # Update status semua laporan akhir menjadi 'diajukan'
        cursor.execute(f'''
            UPDATE {tabel_laporan} 
            SET status = 'diajukan'
            WHERE id_proposal = %s
        ''', (proposal_id,))
        
        app_funcs["mysql"].connection.commit()
        cursor.close()
        
        return jsonify({'success': True, 'message': 'Laporan akhir berhasil diajukan!'})
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})


@mahasiswa_bp.route('/get_program_studi_list')
def mahasiswa_get_program_studi_list():
    logger.info("get_program_studi_list dipanggil")
    app_funcs = get_app_functions()
    
    if 'user_type' not in session or session['user_type'] != 'mahasiswa':
        logger.warning(f"Akses ditolak untuk get_program_studi_list: {session.get('user_type', 'None')}")
        return jsonify({'success': False, 'message': 'Anda harus login sebagai mahasiswa!'})
    
    if not hasattr(app_funcs["mysql"], 'connection') or app_funcs["mysql"].connection is None:
        logger.error("Koneksi database gagal untuk get_program_studi_list")
        return jsonify({'success': False, 'message': 'Koneksi ke database gagal!'})
    
    try:
        cursor = app_funcs["mysql"].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Ambil data program studi untuk dropdown
        cursor.execute('SELECT id, nama_program_studi FROM program_studi ORDER BY nama_program_studi')
        program_studi_list = cursor.fetchall()
        
        logger.info(f"Berhasil mengambil {len(program_studi_list)} program studi")
        
        cursor.close()
        
        return jsonify({
            'success': True, 
            'program_studi_list': program_studi_list
        })
        
    except Exception as e:
        logger.error(f"Error dalam get_program_studi_list: {str(e)}")
        logger.error(f"Traceback: {traceback.format_exc()}")
        print(f"ERROR dalam get_program_studi_list: {str(e)}")
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'Error saat mengambil data: {str(e)}'})


@mahasiswa_bp.route('/submit_laporan_kemajuan', methods=['POST'])
def submit_laporan_kemajuan():
    app_funcs = get_app_functions()
    if 'user_type' not in session or session['user_type'] != 'mahasiswa':
        return jsonify({'success': False, 'message': 'Anda harus login sebagai mahasiswa!'})
    
    if not hasattr(app_funcs["mysql"], 'connection') or app_funcs["mysql"].connection is None:
        return jsonify({'success': False, 'message': 'Koneksi ke database gagal!'})
    
    # Validasi jadwal kemajuan sebelum mengizinkan submit
    jadwal_status = get_jadwal_kemajuan_status()
    if not jadwal_status['bisa_upload']:
        return jsonify({'success': False, 'message': f"Tidak dapat mengajukan laporan kemajuan. {jadwal_status['pesan']}"})
    
    try:
        data = request.get_json()
        proposal_id = data.get('proposal_id')
        tabel_laporan = data.get('tabel_laporan')
        
        if not proposal_id or not tabel_laporan:
            return jsonify({'success': False, 'message': 'Data tidak lengkap!'})
        
        if tabel_laporan not in ['laporan_kemajuan_awal', 'laporan_kemajuan_bertumbuh']:
            return jsonify({'success': False, 'message': 'Tabel laporan tidak valid!'})
        
        cursor = app_funcs["mysql"].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Cek apakah proposal milik mahasiswa ini
        cursor.execute('''
            SELECT p.*, m.nim 
            FROM proposal p
            JOIN mahasiswa m ON p.nim = m.nim
            WHERE p.id = %s AND m.nim = %s
        ''', (proposal_id, session['nim']))
        
        proposal = cursor.fetchone()
        if not proposal:
            cursor.close()
            return jsonify({'success': False, 'message': 'Proposal tidak ditemukan atau Anda tidak memiliki akses!'})
        
        # Update status semua laporan kemajuan menjadi 'diajukan'
        cursor.execute(f'''
            UPDATE {tabel_laporan} 
            SET status = 'diajukan'
            WHERE id_proposal = %s
        ''', (proposal_id,))
        
        app_funcs["mysql"].connection.commit()
        cursor.close()
        
        return jsonify({'success': True, 'message': 'Laporan kemajuan berhasil diajukan!'})
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})
@mahasiswa_bp.route('/tambah_biaya_non_operasional', methods=['POST'])
def tambah_biaya_non_operasional():
    app_funcs = get_app_functions()
    from decimal import Decimal, InvalidOperation
    
    if 'user_type' not in session or session['user_type'] != 'mahasiswa':
        return jsonify({'success': False, 'message': 'Anda harus login sebagai mahasiswa'})
    
    if not hasattr(app_funcs["mysql"], 'connection') or app_funcs["mysql"].connection is None:
        return jsonify({'success': False, 'message': 'Koneksi ke database gagal'})
    
    try:
        cursor = app_funcs["mysql"].connection.cursor(MySQLdb.cursors.DictCursor)
        
        proposal_id = request.form.get('proposal_id')
        nama_biaya = request.form.get('nama_biaya')
        quantity = request.form.get('quantity')
        harga_satuan = request.form.get('harga_satuan')
        tanggal_transaksi = request.form.get('tanggal_transaksi')
        
        # Validasi input
        if not all([proposal_id, nama_biaya, quantity, harga_satuan, tanggal_transaksi]):
            return jsonify({'success': False, 'message': 'Semua field wajib diisi'})
        
        # Validasi proposal_id
        if not proposal_id:
            return jsonify({'success': False, 'message': 'ID proposal tidak boleh kosong'})
        try:
            proposal_id = int(proposal_id)
        except (ValueError, TypeError):
            return jsonify({'success': False, 'message': 'ID proposal tidak valid'})
        
        # Cek apakah proposal ada dan milik mahasiswa yang login
        cursor.execute('SELECT id FROM proposal WHERE id = %s AND nim = %s AND status_admin = "lolos"', (proposal_id, session['nim']))
        proposal_exists = cursor.fetchone()
        if not proposal_exists:
            return jsonify({'success': False, 'message': 'Proposal tidak ditemukan atau belum disetujui'})
        
        # Validasi dan konversi nilai numerik
        if not quantity or not harga_satuan:
            return jsonify({'success': False, 'message': 'Quantity dan harga satuan tidak boleh kosong'})
        try:
            quantity_int = int(quantity)
            harga_satuan_dec = Decimal(str(harga_satuan).replace('.', '').replace(',', '.'))
        except (ValueError, TypeError, InvalidOperation):
            return jsonify({'success': False, 'message': 'Quantity dan harga satuan harus berupa angka'})
        
        if quantity_int <= 0 or harga_satuan_dec <= 0:
            return jsonify({'success': False, 'message': 'Quantity dan harga satuan harus lebih dari 0'})
        
        # Normalisasi nama_biaya
        if not nama_biaya:
            return jsonify({'success': False, 'message': 'Nama biaya tidak boleh kosong'})
        nama_biaya = nama_biaya.strip()
        nama_biaya = ' '.join([w.capitalize() for w in nama_biaya.split()])
        
        # Hitung total harga
        total_harga = quantity_int * float(harga_satuan_dec)
        
        # Insert ke database
        cursor.execute('''
            INSERT INTO biaya_non_operasional (proposal_id, nama_biaya, quantity, harga_satuan, total_harga, tanggal_transaksi)
            VALUES (%s, %s, %s, %s, %s, %s)
        ''', (proposal_id, nama_biaya, quantity_int, float(harga_satuan_dec), total_harga, tanggal_transaksi))
        
        biaya_id = cursor.lastrowid
        app_funcs["mysql"].connection.commit()
        
        # Log aktivitas
        mahasiswa_info = app_funcs['get_mahasiswa_info_from_session']()
        if mahasiswa_info:
            data_baru = {
                'id': biaya_id,
                'nama_biaya': nama_biaya,
                'quantity': quantity_int,
                'harga_satuan': float(harga_satuan_dec),
                'total_harga': total_harga,
                'tanggal_transaksi': tanggal_transaksi
            }
            app_funcs['log_mahasiswa_activity'](
                mahasiswa_info['id'], 
                mahasiswa_info['nim'], 
                mahasiswa_info['nama_ketua'], 
                'tambah', 
                'biaya_non_operasional', 
                f'id_{biaya_id}', 
                f'Menambahkan Biaya Lain-lain: {nama_biaya}',
                None,
                data_baru
            )
        
        # Update arus kas otomatis untuk bulan transaksi
        if tanggal_transaksi:
            bulan_transaksi = datetime.strptime(tanggal_transaksi, '%Y-%m-%d').strftime('%Y-%m')
            app_funcs['update_arus_kas_otomatis'](proposal_id, bulan_transaksi)
        
        cursor.close()
        
        return jsonify({'success': True, 'message': 'Biaya Lain-lain berhasil ditambahkan'})
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@mahasiswa_bp.route('/get_biaya_non_operasional/<int:biaya_id>')
def get_biaya_non_operasional(biaya_id):
    app_funcs = get_app_functions()
    if 'user_type' not in session or session['user_type'] != 'mahasiswa':
        return jsonify({'success': False, 'message': 'Anda harus login sebagai mahasiswa'})
    
    if not hasattr(app_funcs["mysql"], 'connection') or app_funcs["mysql"].connection is None:
        return jsonify({'success': False, 'message': 'Koneksi ke database gagal'})
    
    try:
        cursor = app_funcs["mysql"].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Cek apakah biaya ada dan milik mahasiswa yang login
        cursor.execute('''
            SELECT bno.* FROM biaya_non_operasional bno 
            JOIN proposal p ON bno.proposal_id = p.id 
            WHERE bno.id = %s AND p.nim = %s
        ''', (biaya_id, session['nim']))
        
        biaya = cursor.fetchone()
        if not biaya:
            return jsonify({'success': False, 'message': 'Biaya Lain-lain tidak ditemukan atau bukan milik Anda'})
        
        cursor.close()
        
        return jsonify({'success': True, 'data': biaya})
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@mahasiswa_bp.route('/update_biaya_non_operasional', methods=['POST'])
def update_biaya_non_operasional():
    app_funcs = get_app_functions()
    from decimal import Decimal, InvalidOperation
    
    if 'user_type' not in session or session['user_type'] != 'mahasiswa':
        return jsonify({'success': False, 'message': 'Anda harus login sebagai mahasiswa'})
    
    if not hasattr(app_funcs["mysql"], 'connection') or app_funcs["mysql"].connection is None:
        return jsonify({'success': False, 'message': 'Koneksi ke database gagal'})
    
    try:
        cursor = app_funcs["mysql"].connection.cursor(MySQLdb.cursors.DictCursor)
        
        biaya_id = request.form.get('biaya_non_operasional_id')
        nama_biaya = request.form.get('nama_biaya')
        quantity = request.form.get('quantity')
        harga_satuan = request.form.get('harga_satuan')
        tanggal_transaksi = request.form.get('tanggal_transaksi')
        
        # Validasi input
        if not all([biaya_id, nama_biaya, quantity, harga_satuan, tanggal_transaksi]):
            return jsonify({'success': False, 'message': 'Semua field wajib diisi'})
        
        # Validasi biaya_id
        if not biaya_id:
            return jsonify({'success': False, 'message': 'ID biaya tidak boleh kosong'})
        try:
            biaya_id = int(biaya_id)
        except (ValueError, TypeError):
            return jsonify({'success': False, 'message': 'ID biaya tidak valid'})
        
        # Cek apakah biaya ada dan milik mahasiswa yang login
        cursor.execute('''
            SELECT bno.* FROM biaya_non_operasional bno 
            JOIN proposal p ON bno.proposal_id = p.id 
            WHERE bno.id = %s AND p.nim = %s
        ''', (biaya_id, session['nim']))
        
        biaya_exists = cursor.fetchone()
        if not biaya_exists:
            return jsonify({'success': False, 'message': 'Biaya Lain-lain tidak ditemukan atau bukan milik Anda'})
        
        # Validasi dan konversi nilai numerik
        if not quantity or not harga_satuan:
            return jsonify({'success': False, 'message': 'Quantity dan harga satuan tidak boleh kosong'})
        try:
            quantity_int = int(quantity)
            harga_satuan_dec = Decimal(str(harga_satuan).replace('.', '').replace(',', '.'))
        except (ValueError, TypeError, InvalidOperation):
            return jsonify({'success': False, 'message': 'Quantity dan harga satuan harus berupa angka'})
        
        if quantity_int <= 0 or harga_satuan_dec <= 0:
            return jsonify({'success': False, 'message': 'Quantity dan harga satuan harus lebih dari 0'})
        
        # Normalisasi nama_biaya
        if not nama_biaya:
            return jsonify({'success': False, 'message': 'Nama biaya tidak boleh kosong'})
        nama_biaya = nama_biaya.strip()
        nama_biaya = ' '.join([w.capitalize() for w in nama_biaya.split()])
        
        # Hitung total harga
        total_harga = quantity_int * float(harga_satuan_dec)
        
        # Update database
        cursor.execute('''
            UPDATE biaya_non_operasional 
            SET nama_biaya = %s, quantity = %s, harga_satuan = %s, total_harga = %s, tanggal_transaksi = %s
            WHERE id = %s
        ''', (nama_biaya, quantity_int, float(harga_satuan_dec), total_harga, tanggal_transaksi, biaya_id))
        
        app_funcs["mysql"].connection.commit()
        
        # Log aktivitas
        mahasiswa_info = app_funcs['get_mahasiswa_info_from_session']()
        if mahasiswa_info:
            data_baru = {
                'id': biaya_id,
                'nama_biaya': nama_biaya,
                'quantity': quantity_int,
                'harga_satuan': float(harga_satuan_dec),
                'total_harga': total_harga,
                'tanggal_transaksi': tanggal_transaksi
            }
            app_funcs['log_mahasiswa_activity'](
                mahasiswa_info['id'], 
                mahasiswa_info['nim'], 
                mahasiswa_info['nama_ketua'], 
                'update', 
                'biaya_non_operasional', 
                f'id_{biaya_id}', 
                f'Mengupdate Biaya Lain-lain: {nama_biaya}',
                biaya_exists,
                data_baru
            )
        
        # Update arus kas otomatis untuk bulan transaksi
        if tanggal_transaksi:
            bulan_transaksi = datetime.strptime(tanggal_transaksi, '%Y-%m-%d').strftime('%Y-%m')
            app_funcs['update_arus_kas_otomatis'](biaya_exists['proposal_id'], bulan_transaksi)
        
        cursor.close()
        
        return jsonify({'success': True, 'message': 'Biaya Lain-lain berhasil diupdate'})
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@mahasiswa_bp.route('/delete_biaya_non_operasional', methods=['POST'])
def delete_biaya_non_operasional():
    app_funcs = get_app_functions()
    if 'user_type' not in session or session['user_type'] != 'mahasiswa':
        return jsonify({'success': False, 'message': 'Anda harus login sebagai mahasiswa'})
    
    if not hasattr(app_funcs["mysql"], 'connection') or app_funcs["mysql"].connection is None:
        return jsonify({'success': False, 'message': 'Koneksi ke database gagal'})
    
    try:
        cursor = app_funcs["mysql"].connection.cursor(MySQLdb.cursors.DictCursor)
        
        biaya_id = request.form.get('biaya_non_operasional_id')
        
        if not biaya_id:
            return jsonify({'success': False, 'message': 'ID biaya tidak boleh kosong'})
        try:
            biaya_id = int(biaya_id)
        except (ValueError, TypeError):
            return jsonify({'success': False, 'message': 'ID biaya tidak valid'})
        
        # Cek apakah biaya ada dan milik mahasiswa yang login
        cursor.execute('''
            SELECT bno.* FROM biaya_non_operasional bno 
            JOIN proposal p ON bno.proposal_id = p.id 
            WHERE bno.id = %s AND p.nim = %s
        ''', (biaya_id, session['nim']))
        
        biaya_exists = cursor.fetchone()
        if not biaya_exists:
            return jsonify({'success': False, 'message': 'Biaya Lain-lain tidak ditemukan atau bukan milik Anda'})
        
        # Hapus dari database
        cursor.execute('DELETE FROM biaya_non_operasional WHERE id = %s', (biaya_id,))
        app_funcs["mysql"].connection.commit()
        
        # Log aktivitas
        mahasiswa_info = app_funcs['get_mahasiswa_info_from_session']()
        if mahasiswa_info:
            app_funcs['log_mahasiswa_activity'](
                mahasiswa_info['id'], 
                mahasiswa_info['nim'], 
                mahasiswa_info['nama_ketua'], 
                'hapus', 
                'biaya_non_operasional', 
                f'id_{biaya_id}', 
                f'Menghapus Biaya Lain-lain: {biaya_exists["nama_biaya"]}',
                biaya_exists,
                None
            )
        
        # Update arus kas otomatis untuk bulan transaksi yang dihapus
        if biaya_exists.get('tanggal_transaksi'):
            bulan_transaksi = biaya_exists['tanggal_transaksi'].strftime('%Y-%m')
            app_funcs['update_arus_kas_otomatis'](biaya_exists['proposal_id'], bulan_transaksi)
        
        cursor.close()
        
        return jsonify({'success': True, 'message': 'Biaya Lain-lain berhasil dihapus'})
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@mahasiswa_bp.route('/laporan_neraca_mahasiswa')
def laporan_neraca_mahasiswa():
    """
    Route untuk menampilkan laporan neraca mahasiswa
    """
    if 'user_type' not in session or session['user_type'] != 'mahasiswa':
        flash('Anda harus login sebagai mahasiswa!', 'danger')
        return redirect(url_for('index'))
    
    app_funcs = get_app_functions()
    if not hasattr(app_funcs['mysql'], 'connection') or app_funcs['mysql'].connection is None:
        flash('Koneksi ke database gagal. Cek konfigurasi database!', 'danger')
        return render_template('mahasiswa/laporan_neraca.html', mahasiswa_info=None, neraca_data={})
    
    try:
        cursor = app_funcs['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Ambil informasi mahasiswa dan proposal
        cursor.execute('''
            SELECT m.id, m.nama_ketua, m.nim, m.program_studi, p.judul_usaha, p.id as proposal_id, p.dosen_pembimbing, p.kategori, p.tahapan_usaha
            FROM mahasiswa m
            JOIN proposal p ON m.nim = p.nim
            WHERE m.nim = %s
        ''', (session['nim'],))
        mahasiswa_info = cursor.fetchone()
        
        if not mahasiswa_info:
            flash('Data mahasiswa tidak ditemukan!', 'danger')
            return render_template('mahasiswa/laporan_neraca.html', mahasiswa_info=None, neraca_data={})
        
        # Hitung laporan neraca secara real-time dari data transaksi
        neraca_data = app_funcs['hitung_neraca_real_time'](mahasiswa_info['proposal_id'], cursor)
        
        cursor.close()
        
        return render_template('mahasiswa/laporan_neraca.html', 
                             mahasiswa_info=mahasiswa_info,
                             neraca_data=neraca_data)
        
    except Exception as e:
        print(f"Error in laporan_neraca_mahasiswa: {e}")
        flash('Terjadi kesalahan saat mengambil data laporan neraca!', 'danger')
        return render_template('mahasiswa/laporan_neraca.html', mahasiswa_info=None, neraca_data={})
@mahasiswa_bp.route('/download_neraca', methods=['POST'])
def download_neraca():
    """
    Route untuk download laporan neraca dalam format Excel, PDF, atau Word
    """
    if 'user_type' not in session or session['user_type'] != 'mahasiswa':
        flash('Anda harus login sebagai mahasiswa!', 'danger')
        return redirect(url_for('mahasiswa.laporan_neraca_mahasiswa'))
    
    app_funcs = get_app_functions()
    if not hasattr(app_funcs['mysql'], 'connection') or app_funcs['mysql'].connection is None:
        flash('Koneksi ke database gagal. Cek konfigurasi database!', 'danger')
        return redirect(url_for('mahasiswa.laporan_neraca_mahasiswa'))
    
    try:
        format_type = request.form.get('format', 'excel')
        proposal_id = request.form.get('proposal_id')
        
        if not proposal_id:
            flash('ID proposal tidak ditemukan!', 'danger')
            return redirect(url_for('mahasiswa.laporan_neraca_mahasiswa'))
        
        cursor = app_funcs['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Ambil informasi mahasiswa dan proposal
        cursor.execute('''
            SELECT m.id, m.nama_ketua, m.nim, m.program_studi, p.judul_usaha, p.id as proposal_id, p.dosen_pembimbing, p.kategori, p.tahapan_usaha
            FROM mahasiswa m
            JOIN proposal p ON m.nim = p.nim
            WHERE p.id = %s AND m.nim = %s
        ''', (proposal_id, session['nim']))
        
        mahasiswa_info = cursor.fetchone()
        if not mahasiswa_info:
            flash('Data mahasiswa tidak ditemukan!', 'danger')
            cursor.close()
            return redirect(url_for('mahasiswa.laporan_neraca_mahasiswa'))
        
        # Hitung laporan neraca secara real-time
        neraca_data = app_funcs['hitung_neraca_real_time'](proposal_id, cursor)
        cursor.close()
        
        # Generate file berdasarkan format
        if format_type == 'excel':
            file_buffer = app_funcs['generate_excel_neraca'](mahasiswa_info, neraca_data)
            if file_buffer:
                # Log aktivitas
                mahasiswa_info_session = app_funcs['get_mahasiswa_info_from_session']()
                if mahasiswa_info_session:
                    log_mahasiswa_download_activity(app_funcs, mahasiswa_info_session, 'laporan_neraca', proposal_id, format_type)
                
                filename = f"Laporan_Neraca_{mahasiswa_info['nama_ketua']}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                return send_file(
                    file_buffer,
                    as_attachment=True,
                    download_name=filename,
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
            else:
                flash('Gagal membuat file Excel!', 'danger')
                return redirect(url_for('mahasiswa.laporan_neraca_mahasiswa'))
        elif format_type == 'pdf':
            file_buffer = app_funcs['generate_pdf_neraca'](mahasiswa_info, neraca_data)
            if file_buffer:
                # Log aktivitas
                mahasiswa_info_session = app_funcs['get_mahasiswa_info_from_session']()
                if mahasiswa_info_session:
                    log_mahasiswa_download_activity(app_funcs, mahasiswa_info_session, 'laporan_neraca', proposal_id, format_type)
                
                filename = f"Laporan_Neraca_{mahasiswa_info['nama_ketua']}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
                return send_file(
                    file_buffer,
                    as_attachment=True,
                    download_name=filename,
                    mimetype='application/pdf'
                )
            else:
                flash('Gagal membuat file PDF!', 'danger')
                return redirect(url_for('mahasiswa.laporan_neraca_mahasiswa'))
        elif format_type == 'word':
            file_buffer = app_funcs['generate_word_neraca'](mahasiswa_info, neraca_data)
            if file_buffer:
                # Log aktivitas
                mahasiswa_info_session = app_funcs['get_mahasiswa_info_from_session']()
                if mahasiswa_info_session:
                    log_mahasiswa_download_activity(app_funcs, mahasiswa_info_session, 'laporan_neraca', proposal_id, format_type)
                
                filename = f"Laporan_Neraca_{mahasiswa_info['nama_ketua']}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.docx"
                return send_file(
                    file_buffer,
                    as_attachment=True,
                    download_name=filename,
                    mimetype='application/vnd.openxmlformats-officedocument.wordprocessingml.document'
                )
            else:
                flash('Gagal membuat file Word!', 'danger')
                return redirect(url_for('mahasiswa.laporan_neraca_mahasiswa'))
        else:
            flash('Format file tidak valid!', 'danger')
            return redirect(url_for('mahasiswa.laporan_neraca_mahasiswa'))
        
    except Exception as e:
        print(f"Error in download_neraca: {e}")
        flash(f'Error saat mengunduh file: {str(e)}', 'danger')
        return redirect(url_for('mahasiswa.laporan_neraca_mahasiswa'))

@mahasiswa_bp.route('/upload_file_laporan_kemajuan', methods=['POST'])
def upload_file_laporan_kemajuan():
    app_funcs = get_app_functions()
    if 'user_type' not in session or session['user_type'] != 'mahasiswa':
        return jsonify({'success': False, 'message': 'Anda harus login sebagai mahasiswa!'})
    
    try:
        # Guard jadwal: cek apakah bisa upload berdasarkan jadwal
        jadwal_status = get_jadwal_kemajuan_status()
        if not jadwal_status['bisa_upload']:
            return jsonify({'success': False, 'message': jadwal_status['pesan']})

        proposal_id = request.form.get('proposal_id')
        if not proposal_id:
            return jsonify({'success': False, 'message': 'ID Proposal tidak ditemukan!'})
        
        # Cek apakah file ada
        if 'file_laporan' not in request.files:
            return jsonify({'success': False, 'message': 'File tidak ditemukan!'})
        
        file = request.files['file_laporan']
        if file.filename == '':
            return jsonify({'success': False, 'message': 'Pilih file terlebih dahulu!'})
        
        # Validasi ekstensi file
        if not file.filename.lower().endswith('.pdf'):
            return jsonify({'success': False, 'message': 'Hanya file PDF yang diizinkan!'})
        
        # Validasi ukuran file (5MB)
        file.seek(0, 2)  # Pindah ke akhir file
        file_size = file.tell()
        file.seek(0)  # Kembali ke awal file
        
        if file_size > 5 * 1024 * 1024:  # 5MB
            return jsonify({'success': False, 'message': 'Ukuran file maksimal 5MB!'})
        
        # Ambil data mahasiswa untuk nama file
        cursor = app_funcs["mysql"].connection.cursor(MySQLdb.cursors.DictCursor)
        cursor.execute('''
            SELECT m.nama_ketua, p.judul_usaha
            FROM mahasiswa m 
            LEFT JOIN proposal p ON m.nim = p.nim
            WHERE p.id = %s AND m.nim = %s
        ''', (proposal_id, session['nim']))
        
        mahasiswa_data = cursor.fetchone()
        if not mahasiswa_data:
            cursor.close()
            return jsonify({'success': False, 'message': 'Data mahasiswa tidak ditemukan!'})
        
        # Buat direktori upload jika belum ada
        upload_dir = os.path.join('static', 'uploads', 'Laporan Kemajuan', 'Laporan')
        os.makedirs(upload_dir, exist_ok=True)
        
        # Generate nama file
        nama_ketua = mahasiswa_data['nama_ketua'].replace(' ', '_')
        judul_usaha = mahasiswa_data['judul_usaha'].replace(' ', '_')[:50]  # Batasi panjang judul
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"Laporan_Kemajuan_{nama_ketua}_{proposal_id}_{timestamp}.pdf"
        file_path = os.path.join(upload_dir, filename)
        
        # Cek apakah sudah ada file laporan untuk proposal ini
        cursor.execute('''
            SELECT id, nama_file, file_path FROM file_laporan_kemajuan 
            WHERE id_proposal = %s
        ''', (proposal_id,))
        
        existing_file = cursor.fetchone()
        
        if existing_file:
            # Hapus file lama jika ada
            old_file_path = existing_file['file_path']
            if os.path.exists(old_file_path):
                os.remove(old_file_path)
            
            # Update data di database - hapus komentar pembimbing saat status berubah ke diajukan atau disetujui
            cursor.execute('''
                UPDATE file_laporan_kemajuan 
                SET nama_file = %s, file_path = %s, status = 'diajukan', komentar_pembimbing = NULL, tanggal_update = CURRENT_TIMESTAMP
                WHERE id_proposal = %s
            ''', (filename, file_path, proposal_id))
        else:
            # Insert data baru
            cursor.execute('''
                INSERT INTO file_laporan_kemajuan 
                (id_proposal, nama_file, file_path, status) 
                VALUES (%s, %s, %s, 'diajukan')
            ''', (proposal_id, filename, file_path))
        
        # Simpan file
        file.save(file_path)
        app_funcs["mysql"].connection.commit()
        cursor.close()
        
        return jsonify({
            'success': True, 
            'message': 'File laporan kemajuan berhasil diupload!',
            'filename': filename,
            'status': 'diajukan'
        })
        
    except Exception as e:
        print(f"Error upload file laporan kemajuan: {str(e)}")
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@mahasiswa_bp.route('/edit_file_laporan_kemajuan', methods=['POST'])
def edit_file_laporan_kemajuan():
    app_funcs = get_app_functions()
    if 'user_type' not in session or session['user_type'] != 'mahasiswa':
        return jsonify({'success': False, 'message': 'Anda harus login sebagai mahasiswa!'})
    
    try:
        # Validasi jadwal kemajuan sebelum mengizinkan edit
        jadwal_status = get_jadwal_kemajuan_status()
        if not jadwal_status['bisa_upload']:
            return jsonify({'success': False, 'message': f"Tidak dapat mengedit file laporan kemajuan. {jadwal_status['pesan']}"})

        proposal_id = request.form.get('proposal_id')
        if not proposal_id:
            return jsonify({'success': False, 'message': 'ID Proposal tidak ditemukan!'})
        
        # Cek apakah file ada
        if 'file_laporan' not in request.files:
            return jsonify({'success': False, 'message': 'File tidak ditemukan!'})
        
        file = request.files['file_laporan']
        if file.filename == '':
            return jsonify({'success': False, 'message': 'Pilih file terlebih dahulu!'})
        
        # Validasi ekstensi file
        if not file.filename.lower().endswith('.pdf'):
            return jsonify({'success': False, 'message': 'Hanya file PDF yang diizinkan!'})
        
        # Validasi ukuran file (5MB)
        file.seek(0, 2)  # Pindah ke akhir file
        file_size = file.tell()
        file.seek(0)  # Kembali ke awal file
        
        if file_size > 5 * 1024 * 1024:  # 5MB
            return jsonify({'success': False, 'message': 'Ukuran file maksimal 5MB!'})
        
        # Ambil data mahasiswa untuk nama file
        cursor = app_funcs["mysql"].connection.cursor(MySQLdb.cursors.DictCursor)
        cursor.execute('''
            SELECT m.nama_ketua, p.judul_usaha
            FROM mahasiswa m 
            LEFT JOIN proposal p ON m.nim = p.nim
            WHERE p.id = %s AND m.nim = %s
        ''', (proposal_id, session['nim']))
        
        mahasiswa_data = cursor.fetchone()
        if not mahasiswa_data:
            cursor.close()
            return jsonify({'success': False, 'message': 'Data mahasiswa tidak ditemukan!'})
        
        # Cek apakah sudah ada file laporan untuk proposal ini
        cursor.execute('''
            SELECT id, nama_file, file_path, status FROM file_laporan_kemajuan 
            WHERE id_proposal = %s
        ''', (proposal_id,))
        
        existing_file = cursor.fetchone()
        if not existing_file:
            cursor.close()
            return jsonify({'success': False, 'message': 'File laporan kemajuan tidak ditemukan!'})
        
        # Hapus file lama
        old_file_path = existing_file['file_path']
        if os.path.exists(old_file_path):
            os.remove(old_file_path)
        
        # Buat direktori upload jika belum ada
        upload_dir = os.path.join('static', 'uploads', 'Laporan Kemajuan', 'Laporan')
        os.makedirs(upload_dir, exist_ok=True)
        
        # Generate nama file baru
        nama_ketua = mahasiswa_data['nama_ketua'].replace(' ', '_')
        judul_usaha = mahasiswa_data['judul_usaha'].replace(' ', '_')[:50]  # Batasi panjang judul
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"Laporan_Kemajuan_{nama_ketua}_{proposal_id}_{timestamp}.pdf"
        file_path = os.path.join(upload_dir, filename)
        
        # Update data di database - hapus komentar pembimbing saat status berubah ke diajukan atau disetujui
        cursor.execute('''
            UPDATE file_laporan_kemajuan 
            SET nama_file = %s, file_path = %s, status = 'diajukan', komentar_pembimbing = NULL, tanggal_update = CURRENT_TIMESTAMP
            WHERE id_proposal = %s
        ''', (filename, file_path, proposal_id))
        
        # Simpan file baru
        file.save(file_path)
        app_funcs["mysql"].connection.commit()
        cursor.close()
        
        return jsonify({
            'success': True, 
            'message': 'File laporan kemajuan berhasil diperbarui!',
            'filename': filename,
            'status': 'diajukan'
        })
        
    except Exception as e:
        print(f"Error edit file laporan kemajuan: {str(e)}")
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@mahasiswa_bp.route('/view_file_laporan_kemajuan/<int:proposal_id>')
def view_file_laporan_kemajuan(proposal_id):
    app_funcs = get_app_functions()
    if 'user_type' not in session or session['user_type'] != 'mahasiswa':
        flash('Anda harus login sebagai mahasiswa!', 'danger')
        return redirect(url_for('index'))
    
    try:
        cursor = app_funcs["mysql"].connection.cursor(MySQLdb.cursors.DictCursor)
        cursor.execute('''
            SELECT nama_file, file_path FROM file_laporan_kemajuan 
            WHERE id_proposal = %s
        ''', (proposal_id,))
        
        file_data = cursor.fetchone()
        cursor.close()
        
        if not file_data:
            flash('File laporan kemajuan tidak ditemukan!', 'danger')
            return redirect(url_for('mahasiswa.laporan_kemajuan'))
        
        file_path = file_data['file_path']
        if not os.path.exists(file_path):
            flash('File tidak ditemukan di server!', 'danger')
            return redirect(url_for('mahasiswa.laporan_kemajuan'))
        
        # Tampilkan file PDF di browser
        return send_file(file_path, as_attachment=False, mimetype='application/pdf')
        
    except Exception as e:
        flash(f'Error saat menampilkan file: {str(e)}', 'danger')
        return redirect(url_for('mahasiswa.laporan_kemajuan'))

@mahasiswa_bp.route('/upload_file_laporan_akhir', methods=['POST'])
def upload_file_laporan_akhir():
    app_funcs = get_app_functions()
    if 'user_type' not in session or session['user_type'] != 'mahasiswa':
        return jsonify({'success': False, 'message': 'Anda harus login sebagai mahasiswa!'})
    
    try:
        print(f"DEBUG: Upload file laporan akhir dimulai")
        print(f"DEBUG: Session NIM: {session.get('nim')}")
        print(f"DEBUG: Form data: {request.form}")
        print(f"DEBUG: Files: {request.files}")
        
        # Guard jadwal: cek apakah bisa upload berdasarkan jadwal
        jadwal_status = get_jadwal_akhir_status()
        print(f"DEBUG: Jadwal status: {jadwal_status}")
        if not jadwal_status['bisa_upload']:
            return jsonify({'success': False, 'message': jadwal_status['pesan']})

        proposal_id = request.form.get('proposal_id')
        print(f"DEBUG: Proposal ID: {proposal_id}")
        if not proposal_id:
            return jsonify({'success': False, 'message': 'ID Proposal tidak ditemukan!'})
        
        # Cek apakah file ada
        if 'file_laporan_akhir' not in request.files:
            print(f"DEBUG: File tidak ditemukan dalam request.files")
            return jsonify({'success': False, 'message': 'File tidak ditemukan!'})
        
        file = request.files['file_laporan_akhir']
        print(f"DEBUG: File yang diupload: {file.filename}")
        if file.filename == '':
            return jsonify({'success': False, 'message': 'Pilih file terlebih dahulu!'})
        
        # Validasi ekstensi file
        if not file.filename.lower().endswith('.pdf'):
            return jsonify({'success': False, 'message': 'Hanya file PDF yang diizinkan!'})
        
        # Validasi ukuran file (5MB)
        file.seek(0, 2)  # Pindah ke akhir file
        file_size = file.tell()
        file.seek(0)  # Kembali ke awal file
        print(f"DEBUG: Ukuran file: {file_size} bytes")
        
        if file_size > 5 * 1024 * 1024:  # 5MB
            return jsonify({'success': False, 'message': 'Ukuran file maksimal 5MB!'})
        
        # Ambil data mahasiswa untuk nama file
        cursor = app_funcs["mysql"].connection.cursor(MySQLdb.cursors.DictCursor)
        cursor.execute('''
            SELECT m.nama_ketua, p.judul_usaha
            FROM mahasiswa m 
            LEFT JOIN proposal p ON m.nim = p.nim
            WHERE p.id = %s AND m.nim = %s
        ''', (proposal_id, session['nim']))
        
        mahasiswa_data = cursor.fetchone()
        print(f"DEBUG: Data mahasiswa: {mahasiswa_data}")
        if not mahasiswa_data:
            cursor.close()
            return jsonify({'success': False, 'message': 'Data mahasiswa tidak ditemukan!'})
        
        # Buat direktori upload jika belum ada
        upload_dir = os.path.join('static', 'uploads', 'Laporan Akhir', 'Laporan')
        print(f"DEBUG: Upload directory: {upload_dir}")
        os.makedirs(upload_dir, exist_ok=True)
        
        # Generate nama file
        nama_ketua = mahasiswa_data['nama_ketua'].replace(' ', '_')
        judul_usaha = mahasiswa_data['judul_usaha'].replace(' ', '_')[:50]  # Batasi panjang judul
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"Laporan_Akhir_{nama_ketua}_{proposal_id}_{timestamp}.pdf"
        file_path = os.path.join(upload_dir, filename)
        print(f"DEBUG: Generated filename: {filename}")
        print(f"DEBUG: Full file path: {file_path}")
        
        # Cek apakah sudah ada file laporan untuk proposal ini
        cursor.execute('''
            SELECT id, nama_file, file_path FROM file_laporan_akhir 
            WHERE id_proposal = %s
        ''', (proposal_id,))
        
        existing_file = cursor.fetchone()
        print(f"DEBUG: Existing file: {existing_file}")
        
        if existing_file:
            # Hapus file lama jika ada
            old_file_path = existing_file['file_path']
            if os.path.exists(old_file_path):
                os.remove(old_file_path)
                print(f"DEBUG: File lama dihapus: {old_file_path}")
            
            # Update data di database - hapus komentar pembimbing saat status berubah ke diajukan atau disetujui
            cursor.execute('''
                UPDATE file_laporan_akhir 
                SET nama_file = %s, file_path = %s, status = 'diajukan', komentar_pembimbing = NULL, tanggal_update = CURRENT_TIMESTAMP
                WHERE id_proposal = %s
            ''', (filename, file_path, proposal_id))
            print(f"DEBUG: Database updated untuk existing file")
        else:
            # Insert data baru
            cursor.execute('''
                INSERT INTO file_laporan_akhir 
                (id_proposal, nama_file, file_path, status) 
                VALUES (%s, %s, %s, 'diajukan')
            ''', (proposal_id, filename, file_path))
            print(f"DEBUG: Database insert untuk file baru")
        
        # Simpan file
        file.save(file_path)
        print(f"DEBUG: File berhasil disimpan ke: {file_path}")
        app_funcs["mysql"].connection.commit()
        cursor.close()
        
        return jsonify({
            'success': True, 
            'message': 'File laporan akhir berhasil diupload!',
            'filename': filename,
            'status': 'diajukan'
        })
        
    except Exception as e:
        print(f"Error upload file laporan akhir: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@mahasiswa_bp.route('/view_file_laporan_akhir/<int:proposal_id>')
def view_file_laporan_akhir(proposal_id):
    if 'user_type' not in session or session['user_type'] != 'mahasiswa':
        flash('Anda harus login sebagai mahasiswa!', 'danger')
        return redirect(url_for('index'))
    
    try:
        app_funcs = get_app_functions()
        cursor = app_funcs['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Cek apakah file ada di database
        cursor.execute('''
            SELECT nama_file, file_path FROM file_laporan_akhir 
            WHERE id_proposal = %s
        ''', (proposal_id,))
        
        file_data = cursor.fetchone()
        cursor.close()
        
        if not file_data:
            flash('File laporan akhir tidak ditemukan!', 'danger')
            return redirect(url_for('mahasiswa.laporan_akhir'))
        
        file_path = file_data['file_path']
        if not os.path.exists(file_path):
            flash('File tidak ditemukan di server!', 'danger')
            return redirect(url_for('mahasiswa.laporan_akhir'))
        
        # Tampilkan file PDF di browser
        return send_file(file_path, as_attachment=False, mimetype='application/pdf')
        
    except Exception as e:
        flash(f'Error saat menampilkan file: {str(e)}', 'danger')
        return redirect(url_for('mahasiswa.laporan_akhir'))

# Route untuk halaman nilai mahasiswa
@mahasiswa_bp.route('/nilai')
def nilai_mahasiswa():
    if 'user_type' not in session or session['user_type'] != 'mahasiswa':
        flash('Anda harus login sebagai mahasiswa!', 'danger')
        return redirect(url_for('index'))
    
    return render_template('mahasiswa/nilai.html')

# Route untuk mendapatkan data nilai mahasiswa
@mahasiswa_bp.route('/get_nilai_mahasiswa')
def get_nilai_mahasiswa():
    if 'user_type' not in session or session['user_type'] != 'mahasiswa':
        return jsonify({'success': False, 'message': 'Akses ditolak!'})
    
    try:
        app_funcs = get_app_functions()
        cursor = app_funcs['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Ambil data nilai mahasiswa berdasarkan NIM yang login
        nim = session['nim']
        
        # Query untuk mendapatkan semua nilai
        cursor.execute('''
            SELECT 
                p.judul_usaha,
                p.id as proposal_id,
                COALESCE(pm.nilai_akhir, 0) as nilai_mahasiswa,
                COALESCE(pp.nilai_akhir, 0) as nilai_proposal,
                COALESCE(pk.nilai_akhir, 0) as nilai_kemajuan,
                COALESCE(pa.nilai_akhir, 0) as nilai_akhir,
                pa.rekomendasi_kelulusan as rekomendasi_akhir,
                CASE 
                    WHEN pm.updated_at IS NOT NULL THEN DATE_FORMAT(pm.updated_at, '%%d/%%m/%%Y')
                    ELSE NULL 
                END as tanggal_mahasiswa,
                CASE 
                    WHEN pp.tanggal_penilaian IS NOT NULL THEN DATE_FORMAT(pp.tanggal_penilaian, '%%d/%%m/%%Y')
                    ELSE NULL 
                END as tanggal_proposal,
                CASE 
                    WHEN pk.tanggal_penilaian IS NOT NULL THEN DATE_FORMAT(pk.tanggal_penilaian, '%%d/%%m/%%Y')
                    ELSE NULL 
                END as tanggal_kemajuan,
                CASE 
                    WHEN pa.tanggal_penilaian IS NOT NULL THEN DATE_FORMAT(pa.tanggal_penilaian, '%%d/%%m/%%Y')
                    ELSE NULL 
                END as tanggal_akhir
            FROM proposal p
            LEFT JOIN penilaian_mahasiswa pm ON p.id = pm.id_proposal
            LEFT JOIN penilaian_proposal pp ON p.id = pp.id_proposal
            LEFT JOIN penilaian_laporan_kemajuan pk ON p.id = pk.id_proposal
            LEFT JOIN penilaian_laporan_akhir pa ON p.id = pa.id_proposal
            WHERE p.nim = %s
            AND p.status IN ('disetujui', 'revisi', 'selesai')
            AND p.status_admin = 'lolos'
            ORDER BY p.tanggal_buat DESC
            LIMIT 1
        ''', (nim,))
        
        result = cursor.fetchone()
        cursor.close()
        
        if result:
            # Konversi nilai ke float untuk memastikan format yang benar
            data = {
                'judul_usaha': result.get('judul_usaha', '-'),
                'proposal_id': int(result.get('proposal_id', 0)),
                'nilai_mahasiswa': float(result.get('nilai_mahasiswa', 0)),
                'nilai_proposal': float(result.get('nilai_proposal', 0)),
                'nilai_kemajuan': float(result.get('nilai_kemajuan', 0)),
                'nilai_akhir': float(result.get('nilai_akhir', 0)),
                'rekomendasi_akhir': result.get('rekomendasi_akhir'),
                'tanggal_mahasiswa': result.get('tanggal_mahasiswa'),
                'tanggal_proposal': result.get('tanggal_proposal'),
                'tanggal_kemajuan': result.get('tanggal_kemajuan'),
                'tanggal_akhir': result.get('tanggal_akhir')
            }
            return jsonify({
                'success': True,
                'data': data
            })
        else:
            return jsonify({
                'success': True,
                'data': {
                    'judul_usaha': '-',
                    'proposal_id': 0,
                    'nilai_mahasiswa': 0.0,
                    'nilai_proposal': 0.0,
                    'nilai_kemajuan': 0.0,
                    'nilai_akhir': 0.0,
                    'rekomendasi_akhir': None,
                    'tanggal_mahasiswa': None,
                    'tanggal_proposal': None,
                    'tanggal_kemajuan': None,
                    'tanggal_akhir': None
                }
            })
            
    except Exception as e:
        print(f"Error get_nilai_mahasiswa: {str(e)}")
        # Log error untuk debugging
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'Terjadi kesalahan saat memuat data nilai: {str(e)}'})

# Route untuk mendapatkan detail penilaian mahasiswa
@mahasiswa_bp.route('/get_detail_penilaian_mahasiswa')
def get_detail_penilaian_mahasiswa():
    if 'user_type' not in session or session['user_type'] != 'mahasiswa':
        return jsonify({'success': False, 'message': 'Akses ditolak!'})
    
    try:
        app_funcs = get_app_functions()
        cursor = app_funcs['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        nim = session['nim']
        print(f"DEBUG: NIM: {nim}")
        
        # Helper function untuk nama kategori yang user-friendly dengan urutan yang benar
        def get_kategori_display_name(kategori):
            kategori_map = {
                'pondasi-bisnis': 'A. Pondasi Bisnis',
                'pelanggan': 'B. Pelanggan',
                'produk-jasa': 'C. Produk/Jasa',
                'pemasaran': 'D. Pemasaran',
                'penjualan': 'E. Penjualan',
                'keuntungan': 'F. Keuntungan',
                'sdm': 'G. Sumber Daya Manusia',
                'sistem-bisnis': 'H. Sistem Bisnis'
            }
            return kategori_map.get(kategori, kategori.title())
        
        # Helper function untuk urutan kategori yang benar
        def get_kategori_order(kategori):
            kategori_order = {
                'pondasi-bisnis': 1,
                'pelanggan': 2,
                'produk-jasa': 3,
                'pemasaran': 4,
                'penjualan': 5,
                'keuntungan': 6,
                'sdm': 7,
                'sistem-bisnis': 8
            }
            return kategori_order.get(kategori, 999)
        
        # Ambil data mahasiswa
        cursor.execute('''
            SELECT m.nama_ketua AS nama_mahasiswa, m.nim AS nim_mahasiswa, p.judul_usaha
            FROM mahasiswa m
            LEFT JOIN proposal p ON m.nim = p.nim
            WHERE m.nim = %s
            AND p.status IN ('disetujui', 'revisi', 'selesai')
            AND p.status_admin = 'lolos'
            ORDER BY p.tanggal_buat DESC
            LIMIT 1
        ''', (nim,))
        
        mahasiswa = cursor.fetchone()
        print(f"DEBUG: Mahasiswa data: {mahasiswa}")
        
        # Ambil data penilaian mahasiswa
        cursor.execute('''
            SELECT pm.*
            FROM penilaian_mahasiswa pm
            LEFT JOIN proposal p ON pm.id_proposal = p.id
            WHERE p.nim = %s
            AND p.status IN ('disetujui', 'revisi', 'selesai')
            AND p.status_admin = 'lolos'
            ORDER BY p.tanggal_buat DESC
            LIMIT 1
        ''', (nim,))
        
        penilaian = cursor.fetchone()
        print(f"DEBUG: Penilaian data: {penilaian}")
        
        # Ambil detail penilaian dengan kategori dan sesi
        detail_penilaian = []
        total_nilai = 0
        total_bobot = 0
        
        if penilaian:
            print(f"DEBUG: Penilaian ID: {penilaian['id']}")
            
            # PERBAIKAN: Ambil pertanyaan berdasarkan data historis yang ada
            # Ini memastikan pertanyaan nonaktif yang punya data tetap tampil
            cursor.execute('''
                SELECT DISTINCT p.id, p.pertanyaan, p.bobot, p.skor_maksimal, p.kategori, p.status
                FROM pertanyaan_penilaian_mahasiswa p
                INNER JOIN detail_penilaian_mahasiswa d ON p.id = d.id_pertanyaan
                WHERE d.id_penilaian_mahasiswa = %s
                ORDER BY p.kategori, p.created_at ASC
            ''', (penilaian['id'],))
            
            pertanyaan_list = cursor.fetchall()
            print(f"DEBUG: Pertanyaan count: {len(pertanyaan_list)}")
            
            # Ambil skor per sesi dengan metadata kolom - PERBAIKAN: Handle data lama dan baru
            try:
                cursor.execute('''
                    SELECT d.id_pertanyaan, d.sesi_penilaian, d.skor, d.nilai, d.tanggal_input,
                           m.nama_kolom, m.urutan_kolom, m.interval_tipe,
                           COALESCE(m.nama_kolom, 
                               CASE 
                                   WHEN d.sesi_penilaian <= 4 THEN CONCAT('Jam ', LPAD(0 + d.sesi_penilaian, 2, '0'), ':00')
                                   ELSE CONCAT('Sesi ', d.sesi_penilaian)
                               END
                           ) as nama_kolom_display
                    FROM detail_penilaian_mahasiswa d
                    LEFT JOIN metadata_kolom_penilaian m ON d.metadata_kolom_id = m.id
                    WHERE d.id_penilaian_mahasiswa = %s
                    ORDER BY d.id_pertanyaan, d.sesi_penilaian
                ''', (penilaian['id'],))
            except:
                # Jika tabel metadata_kolom_penilaian tidak ada, gunakan query sederhana
                cursor.execute('''
                    SELECT d.id_pertanyaan, d.sesi_penilaian, d.skor, d.nilai, d.tanggal_input,
                           NULL as nama_kolom, NULL as urutan_kolom, 'setiap_jam' as interval_tipe,
                           CASE 
                               WHEN d.sesi_penilaian <= 4 THEN CONCAT('Jam ', LPAD(0 + d.sesi_penilaian, 2, '0'), ':00')
                               ELSE CONCAT('Sesi ', d.sesi_penilaian)
                           END as nama_kolom_display
                    FROM detail_penilaian_mahasiswa d
                    WHERE d.id_penilaian_mahasiswa = %s
                    ORDER BY d.id_pertanyaan, d.sesi_penilaian
                ''', (penilaian['id'],))
            
            skor_data = cursor.fetchall()
            print(f"DEBUG: Skor data count: {len(skor_data)}")
            
            # Organize skor per sesi dengan metadata kolom
            skor_per_sesi = {}
            metadata_per_sesi = {}  # Untuk menyimpan metadata kolom per sesi
            
            for skor in skor_data:
                pertanyaan_id = str(skor['id_pertanyaan'])
                sesi = skor['sesi_penilaian']
                
                if pertanyaan_id not in skor_per_sesi:
                    skor_per_sesi[pertanyaan_id] = {}
                
                skor_per_sesi[pertanyaan_id][sesi] = {
                    'skor': skor['skor'],
                    'nilai': skor['nilai'],
                    'tanggal_input': skor['tanggal_input'].isoformat() if skor['tanggal_input'] else None
                }
                
                # Simpan metadata kolom per sesi - PERBAIKAN: Gunakan nama_kolom_display
                if sesi not in metadata_per_sesi:
                    metadata_per_sesi[sesi] = {
                        'nama_kolom': skor.get('nama_kolom_display', f'Sesi {sesi}'),
                        'urutan_kolom': skor.get('urutan_kolom', sesi),
                        'interval_tipe': skor.get('interval_tipe', 'unknown')
                    }
            
                    # PERBAIKAN: Ambil data jadwal untuk informasi sesi dari database (sama seperti admin)
        cursor.execute('''
            SELECT 
                pembimbing_interval_tipe as interval_tipe,
                pembimbing_interval_nilai as interval_nilai,
                pembimbing_nilai_mulai as jam_mulai,
                pembimbing_nilai_selesai as jam_selesai
            FROM pengaturan_jadwal
            ORDER BY created_at DESC LIMIT 1
        ''')
        jadwal_info = cursor.fetchone()
        
        # Generate jam_list berdasarkan interval dari database (sama seperti admin)
        if jadwal_info and jadwal_info['interval_tipe'] and jadwal_info['jam_mulai'] and jadwal_info['jam_selesai']:
            jam_mulai = jadwal_info['jam_mulai']
            jam_selesai = jadwal_info['jam_selesai']
            interval = jadwal_info['interval_nilai'] or 1
            
            # Generate list jam dari jam_mulai sampai jam_selesai dengan interval
            jam_list = []
            
            try:
                # Pastikan jam_mulai dan jam_selesai adalah time object
                if isinstance(jam_mulai, str):
                    try:
                        jam_mulai = datetime.strptime(jam_mulai, '%Y-%m-%d %H:%M:%S')
                        jam_mulai = jam_mulai.time()
                    except:
                        jam_mulai = datetime.strptime(jam_mulai, '%H:%M:%S').time()
                elif isinstance(jam_mulai, datetime):
                    jam_mulai = jam_mulai.time()
                elif isinstance(jam_mulai, timedelta):
                    total_seconds = int(jam_mulai.total_seconds())
                    hours = total_seconds // 3600
                    minutes = (total_seconds % 3600) // 60
                    jam_mulai = datetime.strptime(f"{hours:02d}:{minutes:02d}:00", '%H:%M:%S').time()
                
                if isinstance(jam_selesai, str):
                    try:
                        jam_selesai = datetime.strptime(jam_selesai, '%Y-%m-%d %H:%M:%S')
                        jam_selesai = jam_selesai.time()
                    except:
                        jam_selesai = datetime.strptime(jam_selesai, '%H:%M:%S').time()
                elif isinstance(jam_selesai, datetime):
                    jam_selesai = jam_selesai.time()
                elif isinstance(jam_selesai, timedelta):
                    total_seconds = int(jam_selesai.total_seconds())
                    hours = total_seconds // 3600
                    minutes = total_seconds % 3600 // 60
                    jam_selesai = datetime.strptime(f"{hours:02d}:{minutes:02d}:00", '%H:%M:%S').time()
                
                # Konversi ke datetime untuk perhitungan
                today = datetime.now().date()
                current_dt = datetime.combine(today, jam_mulai)
                end_dt = datetime.combine(today, jam_selesai)
                
                # Generate jam_list berdasarkan interval yang sebenarnya dari database
                # PERBAIKAN: Handle semua jenis interval dengan benar (sama seperti admin)
                if jadwal_info['interval_tipe'] == 'setiap_jam':
                    # PERBAIKAN: Handle kasus jam yang melewati tengah malam
                    while current_dt <= end_dt:
                        jam_list.append(current_dt.strftime('%H:%M'))
                        current_dt += timedelta(hours=interval)
                    
                    jadwal_info['jam_list'] = jam_list
                    
                    # PERBAIKAN: Jika jam_list kosong karena masalah tengah malam, generate manual
                    if not jam_list:
                        print("WARNING: jam_list kosong, generate manual berdasarkan metadata")
                        # Generate jam_list manual berdasarkan interval
                        current_time = jam_mulai
                        while current_time <= jam_selesai:
                            jam_list.append(current_time.strftime('%H:%M'))
                            current_time += timedelta(hours=interval)
                        jadwal_info['jam_list'] = jam_list
                
                elif jadwal_info['interval_tipe'] == 'harian':
                    # Generate label harian berdasarkan tanggal yang sebenarnya
                    hari_names = ['Senin', 'Selasa', 'Rabu', 'Kamis', 'Jumat', 'Sabtu', 'Minggu']
                    current_date = jam_mulai
                    urutan = 1
                    
                    while current_date <= jam_selesai:
                        hari_index = current_date.weekday()  # 0=Senin, 1=Selasa, dst
                        nama_hari = hari_names[hari_index]
                        jam_list.append(nama_hari)
                        current_date += timedelta(days=1)
                        urutan += 1
                    
                    jadwal_info['jam_list'] = jam_list
                
                elif jadwal_info['interval_tipe'] == 'mingguan':
                    # Generate label mingguan berdasarkan minggu dalam tahun
                    # Gunakan tanggal mulai dari jadwal sebagai starting point
                    if isinstance(jadwal_info['pembimbing_nilai_mulai'], str):
                        try:
                            start_date = datetime.strptime(jadwal_info['pembimbing_nilai_mulai'], '%Y-%m-%d %H:%M:%S').date()
                        except:
                            start_date = datetime.now().date()
                    elif isinstance(jadwal_info['pembimbing_nilai_mulai'], datetime):
                        start_date = jadwal_info['pembimbing_nilai_mulai'].date()
                    else:
                        start_date = datetime.now().date()
                    
                    current_date = start_date
                    urutan = 1
                    
                    # Generate minggu berdasarkan rentang jadwal yang sebenarnya
                    if isinstance(jadwal_info['pembimbing_nilai_selesai'], str):
                        try:
                            end_date = datetime.strptime(jadwal_info['pembimbing_nilai_selesai'], '%Y-%m-%d %H:%M:%S').date()
                        except:
                            end_date = start_date + timedelta(weeks=4)
                    elif isinstance(jadwal_info['pembimbing_nilai_selesai'], datetime):
                        end_date = jadwal_info['pembimbing_nilai_selesai'].date()
                    else:
                        end_date = start_date + timedelta(weeks=4)
                    
                    while current_date <= end_date:
                        # Hitung minggu ke berapa dalam tahun (1-52/53)
                        week_num = current_date.isocalendar()[1]  # ISO week number
                        jam_list.append(f"Minggu {week_num}")
                        current_date += timedelta(weeks=1)
                        urutan += 1
                    
                    jadwal_info['jam_list'] = jam_list
                
                elif jadwal_info['interval_tipe'] == 'bulanan':
                    # Generate label bulanan berdasarkan bulan yang sebenarnya
                    bulan_names = ['Januari', 'Februari', 'Maret', 'April', 'Mei', 'Juni', 
                                  'Juli', 'Agustus', 'September', 'Oktober', 'November', 'Desember']
                    
                    # Gunakan tanggal mulai dari jadwal sebagai starting point
                    if isinstance(jadwal_info['pembimbing_nilai_mulai'], str):
                        try:
                            start_date = datetime.strptime(jadwal_info['pembimbing_nilai_mulai'], '%Y-%m-%d %H:%M:%S').date()
                        except:
                            start_date = datetime.now().date()
                    elif isinstance(jadwal_info['pembimbing_nilai_mulai'], datetime):
                        start_date = jadwal_info['pembimbing_nilai_mulai'].date()
                    else:
                        start_date = datetime.now().date()
                    
                    current_date = start_date
                    urutan = 1
                    
                    # Generate bulan berdasarkan rentang jadwal yang sebenarnya
                    if isinstance(jadwal_info['pembimbing_nilai_selesai'], str):
                        try:
                            end_date = datetime.strptime(jadwal_info['pembimbing_nilai_selesai'], '%Y-%m-%d %H:%M:%S').date()
                        except:
                            end_date = start_date + timedelta(days=365)  # 1 tahun
                    elif isinstance(jadwal_info['pembimbing_nilai_selesai'], datetime):
                        end_date = jadwal_info['pembimbing_nilai_selesai'].date()
                    else:
                        end_date = start_date + timedelta(days=365)  # 1 tahun
                    
                    while current_date <= end_date:
                        bulan_index = current_date.month - 1  # 0-based index
                        nama_bulan = bulan_names[bulan_index]
                        jam_list.append(nama_bulan)
                        
                        # Pindah ke bulan berikutnya dengan aman
                        if current_date.month == 12:
                            current_date = current_date.replace(year=current_date.year + 1, month=1, day=1)
                        else:
                            # Gunakan tanggal 1 untuk menghindari masalah tanggal yang tidak valid
                            current_date = current_date.replace(month=current_date.month + 1, day=1)
                        urutan += 1
                    
                    jadwal_info['jam_list'] = jam_list
                
                else:
                    # Jika interval_tipe tidak dikenal, return error
                    cursor.close()
                    return jsonify({'success': False, 'message': f'Interval tipe tidak valid: {jadwal_info["interval_tipe"]}'})
                
            except Exception as e:
                print(f"Error generating jam_list: {e}")
                # Fallback ke default jika error
                jadwal_info['jam_list'] = ['08:00', '09:00', '10:00', '11:00']
        else:
            # Jika jam_mulai atau jam_selesai None, gunakan jam_list kosong
            jadwal_info['jam_list'] = []
        
        # Pastikan semua data di jadwal_info bisa di-serialize ke JSON
        if jadwal_info:
            # Konversi ke tipe yang aman untuk JSON
            jadwal_info['interval_tipe'] = str(jadwal_info['interval_tipe'])
            jadwal_info['interval_nilai'] = int(jadwal_info['interval_nilai']) if jadwal_info['interval_nilai'] else 1
            jadwal_info['jam_mulai'] = str(jadwal_info['jam_mulai'])
            jadwal_info['jam_selesai'] = str(jadwal_info['jam_selesai'])
            jadwal_info['jam_list'] = jadwal_info.get('jam_list', [])
            
            # ✅ PERBAIKAN: Ambil metadata kolom penilaian berdasarkan metadata_kolom_id yang tersimpan di detail_penilaian_mahasiswa (sama seperti admin)
            try:
                # ✅ PERBAIKAN: Ambil metadata_kolom_id yang digunakan dalam penilaian ini
                cursor.execute("""
                    SELECT DISTINCT metadata_kolom_id 
                    FROM detail_penilaian_mahasiswa 
                    WHERE id_penilaian_mahasiswa = %s 
                    AND metadata_kolom_id IS NOT NULL
                    ORDER BY metadata_kolom_id
                """, (penilaian['id'],))
                
                metadata_ids = cursor.fetchall()
                print(f"=== DEBUG METADATA KOLOM ID MAHASISWA ===")
                print(f"Found metadata_kolom_id: {[m['metadata_kolom_id'] for m in metadata_ids]}")
                
                if metadata_ids:
                    # Ambil metadata kolom berdasarkan metadata_kolom_id yang tersimpan
                    metadata_ids_list = [m['metadata_kolom_id'] for m in metadata_ids]
                    cursor.execute("""
                        SELECT id, nama_kolom, urutan_kolom, interval_tipe 
                        FROM metadata_kolom_penilaian 
                        WHERE id IN ({})
                        ORDER BY urutan_kolom
                    """.format(','.join(['%s'] * len(metadata_ids_list))), metadata_ids_list)
                    
                    metadata_kolom = cursor.fetchall()
                    
                    # Jika ada metadata kolom, gunakan sebagai jam_list
                    if metadata_kolom:
                        jadwal_info['metadata_kolom'] = metadata_kolom
                        # Override jam_list dengan metadata kolom yang sebenarnya
                        jadwal_info['jam_list'] = [meta['nama_kolom'] for meta in metadata_kolom]
                        print(f"=== DEBUG METADATA KOLOM MAHASISWA ===")
                        print(f"Found {len(metadata_kolom)} metadata columns for penilaian {penilaian['id']}")
                        print(f"Metadata kolom: {metadata_kolom}")
                        print(f"Updated jam_list: {jadwal_info['jam_list']}")
                        print(f"===============================")
                    else:
                        jadwal_info['metadata_kolom'] = []
                        print(f"No metadata columns found for penilaian {penilaian['id']}, using generated jam_list")
                else:
                    jadwal_info['metadata_kolom'] = []
                    print(f"No metadata columns found for penilaian {penilaian['id']}, using generated jam_list")
            
            except Exception as e:
                print(f"Error getting metadata kolom: {e}")
                # Jika error, tetap lanjutkan dengan jam_list yang sudah digenerate
                jadwal_info['metadata_kolom'] = []
            
            # Organize pertanyaan berdasarkan kategori
            kategori_organized = {}
            for pertanyaan in pertanyaan_list:
                kategori = pertanyaan['kategori']
                if kategori not in kategori_organized:
                    kategori_organized[kategori] = {
                        'nama_kategori': get_kategori_display_name(kategori),
                        'pertanyaan_list': [],
                        'total_bobot_kategori': 0,
                        'total_nilai_kategori': 0
                    }
                
                # Hitung nilai untuk pertanyaan ini
                pertanyaan_id = str(pertanyaan['id'])
                sesi_data = skor_per_sesi.get(pertanyaan_id, {})
                
                # ✅ PERBAIKAN: Hitung rata-rata skor dari semua sesi yang ada
                total_skor = 0
                sesi_count = 0
                for sesi, data in sesi_data.items():
                    total_skor += data['skor']  # Hitung semua sesi (termasuk skor 0)
                    sesi_count += 1
                
                # ✅ PERBAIKAN: Rata-rata = total skor / jumlah sesi (semua sesi)
                rata_rata_skor = (total_skor / sesi_count) if sesi_count > 0 else 0
                bobot = float(pertanyaan['bobot'])
                skor_max = int(pertanyaan['skor_maksimal'])
                nilai_akhir_pertanyaan = (rata_rata_skor / skor_max) * bobot
                
                # Tambahkan ke kategori
                kategori_organized[kategori]['pertanyaan_list'].append({
                    'id': pertanyaan['id'],
                    'pertanyaan': pertanyaan['pertanyaan'],
                    'bobot': bobot,
                    'skor_maksimal': skor_max,
                    'sesi_data': sesi_data,
                    'rata_rata_skor': rata_rata_skor,
                    'nilai_akhir': nilai_akhir_pertanyaan
                })
                
                kategori_organized[kategori]['total_bobot_kategori'] += bobot
                kategori_organized[kategori]['total_nilai_kategori'] += nilai_akhir_pertanyaan
                
                total_bobot += bobot
                total_nilai += nilai_akhir_pertanyaan
            
            # Convert ke list dan urutkan
            for kategori, kategori_info in kategori_organized.items():
                detail_penilaian.append({
                    'kategori': kategori,
                    'nama_kategori': kategori_info['nama_kategori'],
                    'total_bobot_kategori': kategori_info['total_bobot_kategori'],
                    'total_nilai_kategori': kategori_info['total_nilai_kategori'],
                    'pertanyaan_list': kategori_info['pertanyaan_list']
                })
            
            # Urutkan berdasarkan urutan kategori yang benar
            detail_penilaian.sort(key=lambda x: get_kategori_order(x['kategori']))
        
        cursor.close()
        
        # Pastikan jadwal_info selalu ada dengan fallback
        if 'jadwal_info' not in locals() or not jadwal_info:
            jadwal_info = {
                'interval_tipe': 'setiap_jam',
                'jam_mulai': '08:00',
                'jam_selesai': '17:00',
                'interval_nilai': 1,
                'jam_list': ['08:00', '09:00', '10:00', '11:00', '12:00', '13:00', '14:00', '15:00', '16:00', '17:00'],
                'metadata_kolom': []
            }
        
        # Pastikan metadata_kolom selalu ada jika belum diset
        if 'metadata_kolom' not in jadwal_info:
            jadwal_info['metadata_kolom'] = []
        
        # Pastikan metadata_per_sesi selalu ada
        if 'metadata_per_sesi' not in locals():
            metadata_per_sesi = {}
        
        response_data = {
            'success': True,
            'data': {
                'mahasiswa': mahasiswa or {},
                'penilaian': {
                    'id': penilaian['id'] if penilaian else None,
                    'nilai_akhir': penilaian['nilai_akhir'] if penilaian else 0,
                    'total_bobot': total_bobot,
                    'total_nilai': total_nilai
                },
                'detail_penilaian': detail_penilaian,
                'jadwal_info': jadwal_info,
                'metadata_per_sesi': metadata_per_sesi
            }
        }
        
        print(f"DEBUG: Response data: {response_data}")
        
        return jsonify(response_data)
        
    except Exception as e:
        print(f"Error get_detail_penilaian_mahasiswa: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'Terjadi kesalahan saat memuat detail penilaian mahasiswa: {str(e)}'})

def get_jadwal_info_for_ui_mahasiswa(nim):
    """Helper function untuk mendapatkan info jadwal untuk UI mahasiswa"""
    try:
        app_funcs = get_app_functions()
        cursor = app_funcs['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Ambil data pembimbing dari mahasiswa
        cursor.execute('''
            SELECT p.dosen_pembimbing
            FROM proposal p
            WHERE p.nim = %s AND p.status_admin = 'lolos'
        ''', (nim,))
        
        proposal_data = cursor.fetchone()
        if not proposal_data:
            return {}
        
        dosen_pembimbing = proposal_data['dosen_pembimbing']
        
        # Ambil ID pembimbing
        cursor.execute('SELECT id FROM pembimbing WHERE nama = %s', (dosen_pembimbing,))
        pembimbing_data = cursor.fetchone()
        if not pembimbing_data:
            return {}
        
        id_pembimbing = pembimbing_data['id']
        
        # Ambil pengaturan jadwal
        cursor.execute('''
            SELECT pj.*
            FROM pengaturan_jadwal pj
            WHERE pj.id_pembimbing = %s
            ORDER BY pj.created_at DESC
            LIMIT 1
        ''', (id_pembimbing,))
        
        jadwal_data = cursor.fetchone()
        if not jadwal_data:
            return {}
        
        # Ambil metadata kolom aktif (jika ada)
        metadata_kolom = []
        try:
            cursor.execute('''
                SELECT nama_kolom, urutan_kolom, interval_tipe
                FROM metadata_kolom_penilaian
                WHERE periode_metadata_id = %s
                ORDER BY urutan_kolom
            ''', (jadwal_data.get('periode_metadata_aktif', 0),))
            metadata_kolom = cursor.fetchall()
        except Exception as e:
            print(f"Warning: Tidak bisa ambil metadata kolom: {str(e)}")
            # Jika tabel tidak ada atau error, gunakan data default
            metadata_kolom = []
        
        cursor.close()
        
        # Generate jam_list untuk interval setiap_jam
        jam_list = []
        if jadwal_data['interval_tipe'] == 'setiap_jam' and jadwal_data['jam_mulai'] and jadwal_data['jam_selesai']:
            try:
                jam_mulai = int(jadwal_data['jam_mulai'].split(':')[0])
                jam_selesai = int(jadwal_data['jam_selesai'].split(':')[0])
                interval = int(jadwal_data['interval_nilai'] or 1)
                
                current_hour = jam_mulai
                while current_hour <= jam_selesai:
                    jam_list.append(f"{current_hour:02d}:00")
                    current_hour += interval
            except:
                pass
        
        return {
            'interval_tipe': jadwal_data['interval_tipe'],
            'jam_mulai': jadwal_data['jam_mulai'],
            'jam_selesai': jadwal_data['jam_selesai'],
            'interval_nilai': jadwal_data['interval_nilai'],
            'jam_list': jam_list,
            'metadata_kolom': metadata_kolom
        }
        
    except Exception as e:
        print(f"Error in get_jadwal_info_for_ui_mahasiswa: {str(e)}")
        return {}

# Route untuk mendapatkan detail penilaian proposal
@mahasiswa_bp.route('/get_detail_penilaian_proposal')
def get_detail_penilaian_proposal():
    if 'user_type' not in session or session['user_type'] != 'mahasiswa':
        return jsonify({'success': False, 'message': 'Akses ditolak!'})
    
    try:
        app_funcs = get_app_functions()
        cursor = app_funcs['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        nim = session['nim']
        
        # Ambil data mahasiswa
        cursor.execute('''
            SELECT m.nama_ketua AS nama_mahasiswa, m.nim, p.judul_usaha
            FROM mahasiswa m
            LEFT JOIN proposal p ON m.nim = p.nim
            WHERE m.nim = %s
            AND p.status IN ('disetujui', 'revisi', 'selesai')
            AND p.status_admin = 'lolos'
            ORDER BY p.tanggal_buat DESC
            LIMIT 1
        ''', (nim,))
        
        mahasiswa = cursor.fetchone()
        
        # Ambil data penilaian proposal
        cursor.execute('''
            SELECT pp.*
            FROM penilaian_proposal pp
            LEFT JOIN proposal p ON pp.id_proposal = p.id
            WHERE p.nim = %s
            AND p.status IN ('disetujui', 'revisi', 'selesai')
            AND p.status_admin = 'lolos'
            ORDER BY p.tanggal_buat DESC
            LIMIT 1
        ''', (nim,))
        
        penilaian = cursor.fetchone()
        
        # Ambil detail penilaian
        detail_penilaian = []
        if penilaian:
            cursor.execute('''
                SELECT dp.*, p.pertanyaan
                FROM detail_penilaian_proposal dp
                LEFT JOIN pertanyaan_penilaian_proposal p ON dp.id_pertanyaan = p.id
                WHERE dp.id_proposal = %s AND dp.id_reviewer = %s
                ORDER BY p.id
            ''', (penilaian['id_proposal'], penilaian['id_reviewer']))
            
            detail_penilaian = cursor.fetchall()
        
        # Ambil catatan dan rekomendasi langsung dari penilaian_proposal
        catatan = []
        rekomendasi = []
        if penilaian:
            # Catatan langsung dari kolom catatan di penilaian_proposal
            if penilaian.get('catatan'):
                catatan = [penilaian['catatan']]
            
            # Rekomendasi dari nilai_bantuan
            if penilaian.get('nilai_bantuan'):
                rekomendasi = [{'rekomendasi': f"Rp {penilaian['nilai_bantuan']:,.0f}"}]
        
        cursor.close()
        
        return jsonify({
            'success': True,
            'data': {
                'mahasiswa': mahasiswa or {},
                'penilaian': penilaian or {},
                'detail_penilaian': detail_penilaian,
                'catatan': catatan,
                'rekomendasi': rekomendasi
            }
        })
        
    except Exception as e:
        print(f"Error get_detail_penilaian_proposal: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'Terjadi kesalahan saat memuat detail penilaian proposal: {str(e)}'})

# Route untuk mendapatkan detail penilaian laporan kemajuan
@mahasiswa_bp.route('/get_detail_penilaian_kemajuan')
def get_detail_penilaian_kemajuan():
    if 'user_type' not in session or session['user_type'] != 'mahasiswa':
        return jsonify({'success': False, 'message': 'Akses ditolak!'})
    
    try:
        app_funcs = get_app_functions()
        cursor = app_funcs['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        nim = session['nim']
        
        # Ambil data mahasiswa
        cursor.execute('''
            SELECT m.nama_ketua AS nama_mahasiswa, m.nim, p.judul_usaha
            FROM mahasiswa m
            LEFT JOIN proposal p ON m.nim = p.nim
            WHERE m.nim = %s
            AND p.status IN ('disetujui', 'revisi', 'selesai')
            AND p.status_admin = 'lolos'
            ORDER BY p.tanggal_buat DESC
            LIMIT 1
        ''', (nim,))
        
        mahasiswa = cursor.fetchone()
        
        # Ambil data penilaian laporan kemajuan
        cursor.execute('''
            SELECT pk.*
            FROM penilaian_laporan_kemajuan pk
            LEFT JOIN proposal p ON pk.id_proposal = p.id
            WHERE p.nim = %s
            AND p.status IN ('disetujui', 'revisi', 'selesai')
            AND p.status_admin = 'lolos'
            ORDER BY p.tanggal_buat DESC
            LIMIT 1
        ''', (nim,))
        
        penilaian = cursor.fetchone()
        
        # Ambil detail penilaian
        detail_penilaian = []
        if penilaian:
            cursor.execute('''
                SELECT dp.skor_diberikan, dp.bobot_pertanyaan, dp.skor_maksimal, dp.nilai_terbobot, p.pertanyaan
                FROM detail_penilaian_laporan_kemajuan dp
                LEFT JOIN pertanyaan_penilaian_laporan_kemajuan p ON dp.id_pertanyaan = p.id
                WHERE dp.id_penilaian_laporan_kemajuan = %s
                ORDER BY p.id
            ''', (penilaian['id'],))
            
            detail_penilaian = cursor.fetchall()
        
        # Ambil catatan dan rekomendasi langsung dari penilaian_laporan_kemajuan
        catatan = []
        rekomendasi = []
        if penilaian:
            # Catatan langsung dari kolom komentar_reviewer di penilaian_laporan_kemajuan
            if penilaian.get('komentar_reviewer'):
                catatan = [penilaian['komentar_reviewer']]
            
            # Rekomendasi dari kolom yang ada
            if penilaian.get('rekomendasi_pendanaan') or penilaian.get('alasan_rekomendasi'):
                rekomendasi = [{
                    'rekomendasi_pendanaan': penilaian.get('rekomendasi_pendanaan', '-'),
                    'alasan_rekomendasi': penilaian.get('alasan_rekomendasi', '-')
                }]
        
        cursor.close()
        
        return jsonify({
            'success': True,
            'data': {
                'mahasiswa': mahasiswa or {},
                'penilaian': penilaian or {},
                'detail_penilaian': detail_penilaian,
                'catatan': catatan,
                'rekomendasi': rekomendasi
            }
        })
        
    except Exception as e:
        print(f"Error get_detail_penilaian_kemajuan: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'Terjadi kesalahan saat memuat detail penilaian kemajuan: {str(e)}'})

# Route untuk mendapatkan detail penilaian laporan akhir
@mahasiswa_bp.route('/get_detail_penilaian_akhir')
def get_detail_penilaian_akhir():
    if 'user_type' not in session or session['user_type'] != 'mahasiswa':
        return jsonify({'success': False, 'message': 'Akses ditolak!'})
    
    try:
        app_funcs = get_app_functions()
        cursor = app_funcs['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        nim = session['nim']
        
        # Ambil data mahasiswa
        cursor.execute('''
            SELECT m.nama_ketua AS nama_mahasiswa, m.nim, p.judul_usaha
            FROM mahasiswa m
            LEFT JOIN proposal p ON m.nim = p.nim
            WHERE m.nim = %s
            AND p.status IN ('disetujui', 'revisi', 'selesai')
            AND p.status_admin = 'lolos'
            ORDER BY p.tanggal_buat DESC
            LIMIT 1
        ''', (nim,))
        
        mahasiswa = cursor.fetchone()
        
        # Ambil data penilaian laporan akhir
        cursor.execute('''
            SELECT pa.*
            FROM penilaian_laporan_akhir pa
            LEFT JOIN proposal p ON pa.id_proposal = p.id
            WHERE p.nim = %s
            AND p.status IN ('disetujui', 'revisi', 'selesai')
            AND p.status_admin = 'lolos'
            ORDER BY p.tanggal_buat DESC
            LIMIT 1
        ''', (nim,))
        
        penilaian = cursor.fetchone()
        
        # Ambil detail penilaian
        detail_penilaian = []
        if penilaian:
            cursor.execute('''
                SELECT dp.id_pertanyaan, dp.skor_diberikan, dp.bobot_pertanyaan, 
                       dp.skor_maksimal, dp.nilai_terbobot, p.pertanyaan
                FROM detail_penilaian_laporan_akhir dp
                LEFT JOIN pertanyaan_penilaian_laporan_akhir p ON dp.id_pertanyaan = p.id
                WHERE dp.id_penilaian_laporan_akhir = %s
                ORDER BY dp.id_pertanyaan
            ''', (penilaian['id'],))
            
            detail_penilaian = cursor.fetchall()
        
        # Ambil catatan dan rekomendasi langsung dari penilaian_laporan_akhir
        catatan = []
        rekomendasi = []
        if penilaian:
            # Catatan langsung dari kolom komentar_reviewer di penilaian_laporan_akhir
            if penilaian.get('komentar_reviewer'):
                catatan = [penilaian['komentar_reviewer']]
            
            # Rekomendasi dari kolom yang ada
            if penilaian.get('rekomendasi_kelulusan') or penilaian.get('alasan_rekomendasi'):
                rekomendasi = [{
                    'rekomendasi_kelulusan': penilaian.get('rekomendasi_kelulusan', '-'),
                    'alasan_rekomendasi': penilaian.get('alasan_rekomendasi', '-')
                }]
        
        cursor.close()
        
        return jsonify({
            'success': True,
            'data': {
                'mahasiswa': mahasiswa or {},
                'penilaian': penilaian or {},
                'detail_penilaian': detail_penilaian,
                'catatan': catatan,
                'rekomendasi': rekomendasi
            }
        })
        
    except Exception as e:
        print(f"Error get_detail_penilaian_akhir: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': f'Terjadi kesalahan saat memuat detail penilaian akhir: {str(e)}'})

# ========================================
# HELPER FUNCTIONS UNTUK LOGGING
# ========================================

def log_mahasiswa_download_activity(app_funcs, mahasiswa_info, jenis_file, proposal_id, format_type=None, periode=None):
    """Helper function untuk logging aktivitas download mahasiswa"""
    try:
        if mahasiswa_info:
            deskripsi = f'Download {jenis_file} proposal {proposal_id}'
            if format_type:
                deskripsi += f' format {format_type.upper()}'
            if periode:
                deskripsi += f' {periode}'
            
            data_lama = {
                'proposal_id': proposal_id,
                'jenis_file': jenis_file,
                'format': format_type
            }
            if periode:
                data_lama['periode'] = periode
            
            app_funcs['log_mahasiswa_activity'](
                mahasiswa_info['id'],
                mahasiswa_info['nim'],
                mahasiswa_info['nama_ketua'],
                'download',
                jenis_file,
                f'proposal_{proposal_id}',
                deskripsi,
                data_lama,
                None
            )
    except Exception as log_error:
        print(f"Error logging download activity: {str(log_error)}")

def log_mahasiswa_bimbingan_activity(app_funcs, mahasiswa_info, jenis_aktivitas, detail_modul, deskripsi, data_lama=None, data_baru=None):
    """Helper function untuk logging aktivitas bimbingan mahasiswa"""
    try:
        if mahasiswa_info:
            app_funcs['log_mahasiswa_activity'](
                mahasiswa_info['id'],
                mahasiswa_info['nim'],
                mahasiswa_info['nama_ketua'],
                jenis_aktivitas,
                'bimbingan',
                detail_modul,
                deskripsi,
                data_lama,
                data_baru
            )
    except Exception as log_error:
        print(f"Error logging bimbingan activity: {str(log_error)}")

# ========================================
# ROUTE UNTUK SISTEM BIMBINGAN MAHASISWA
# ========================================

@mahasiswa_bp.route('/bimbingan')
def bimbingan():
    """Halaman bimbingan mahasiswa"""
    if 'user_type' not in session or session['user_type'] != 'mahasiswa':
        flash('Anda harus login sebagai mahasiswa!', 'danger')
        return redirect(url_for('index'))
    
    app_funcs = get_app_functions()
    if not hasattr(app_funcs['mysql'], 'connection') or app_funcs['mysql'].connection is None:
        flash('Koneksi ke database gagal. Cek konfigurasi database!', 'danger')
        return render_template('mahasiswa/bimbingan.html', riwayat_bimbingan=[], is_status_selesai=False)
    
    try:
        cursor = app_funcs['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # ✅ PERBAIKAN: Cek status mahasiswa
        cursor.execute('SELECT status FROM mahasiswa WHERE nim = %s', (session['nim'],))
        mahasiswa_status = cursor.fetchone()
        
        if not mahasiswa_status:
            flash('Data mahasiswa tidak ditemukan!', 'danger')
            cursor.close()
            return redirect(url_for('index'))
        
        # ✅ PERBAIKAN: Mahasiswa status selesai hanya bisa melihat bimbingan
        is_status_selesai = mahasiswa_status['status'] == 'selesai'
        
        # Cek apakah mahasiswa memiliki proposal yang lolos
        cursor.execute('''
            SELECT COUNT(*) as count FROM proposal 
            WHERE nim = %s AND status_admin = 'lolos'
        ''', (session['nim'],))
        
        has_lolos_proposal = cursor.fetchone()['count'] > 0
        
        # Ambil riwayat bimbingan mahasiswa
        cursor.execute('''
            SELECT b.*, p.judul_usaha, p.dosen_pembimbing
            FROM bimbingan b
            JOIN proposal p ON b.proposal_id = p.id
            WHERE b.nim = %s
            ORDER BY b.tanggal_buat DESC
        ''', (session['nim'],))
        
        riwayat_bimbingan = cursor.fetchall()
        
        cursor.close()
        
        return render_template('mahasiswa/bimbingan.html', 
                             riwayat_bimbingan=riwayat_bimbingan,
                             has_lolos_proposal=has_lolos_proposal,
                             is_status_selesai=is_status_selesai)
        
    except Exception as e:
        logger.error(f"Error in bimbingan: {str(e)}")
        flash('Terjadi kesalahan saat memuat halaman bimbingan!', 'danger')
        return render_template('mahasiswa/bimbingan.html', riwayat_bimbingan=[], is_status_selesai=False)

@mahasiswa_bp.route('/simpan_pertanyaan_bimbingan', methods=['POST'])
def simpan_pertanyaan_bimbingan():
    """API untuk menyimpan pertanyaan bimbingan"""
    if 'user_type' not in session or session['user_type'] != 'mahasiswa':
        return jsonify({'success': False, 'message': 'Anda harus login sebagai mahasiswa!'})
    
    app_funcs = get_app_functions()
    if not hasattr(app_funcs['mysql'], 'connection') or app_funcs['mysql'].connection is None:
        return jsonify({'success': False, 'message': 'Koneksi ke database gagal!'})
    
    try:
        cursor = app_funcs['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Ambil data dari form
        judul_bimbingan = request.form.get('judul_bimbingan', '').strip()
        hasil_bimbingan = request.form.get('hasil_bimbingan', '').strip()
        
        # ✅ PERBAIKAN: Cek status mahasiswa
        cursor.execute('SELECT status FROM mahasiswa WHERE nim = %s', (session['nim'],))
        mahasiswa_status = cursor.fetchone()
        
        if not mahasiswa_status:
            return jsonify({'success': False, 'message': 'Data mahasiswa tidak ditemukan!'})
        
        # ✅ PERBAIKAN: Mahasiswa status selesai tidak bisa melakukan bimbingan
        if mahasiswa_status['status'] == 'selesai':
            return jsonify({'success': False, 'message': 'Mahasiswa dengan status selesai tidak dapat melakukan bimbingan!'})
        
        # Validasi data
        if not judul_bimbingan:
            return jsonify({'success': False, 'message': 'Judul bimbingan harus diisi!'})
        
        if not hasil_bimbingan or hasil_bimbingan == '<p><br></p>':
            return jsonify({'success': False, 'message': 'Hasil bimbingan harus diisi!'})
        
        # Ambil proposal mahasiswa yang aktif
        cursor.execute('''
            SELECT id, judul_usaha, dosen_pembimbing 
            FROM proposal 
            WHERE nim = %s AND status_admin = 'lolos'
            ORDER BY tanggal_buat DESC 
            LIMIT 1
        ''', (session['nim'],))
        
        proposal = cursor.fetchone()
        if not proposal:
            return jsonify({'success': False, 'message': 'Anda harus memiliki proposal yang sudah disetujui!'})
        
        # Cek pengaturan bimbingan
        cursor.execute("""
            SELECT COUNT(*) as count 
            FROM information_schema.tables 
            WHERE table_schema = DATABASE() 
            AND table_name = 'pengaturan_bimbingan'
        """)
        pengaturan_table_exists = cursor.fetchone()['count'] > 0
        
        if pengaturan_table_exists:
            # Ambil pengaturan bimbingan
            cursor.execute('SELECT * FROM pengaturan_bimbingan ORDER BY id DESC LIMIT 1')
            pengaturan = cursor.fetchone()
            
            if pengaturan and pengaturan['status_aktif'] == 'aktif':
                # Cek jumlah bimbingan hari ini
                cursor.execute('''
                    SELECT COUNT(*) as count 
                    FROM bimbingan 
                    WHERE nim = %s AND DATE(tanggal_buat) = CURDATE()
                ''', (session['nim'],))
                
                bimbingan_hari_ini = cursor.fetchone()['count']
                
                if bimbingan_hari_ini >= pengaturan['maksimal_bimbingan_per_hari']:
                    auto_pesan = f"Anda telah mencapai batas maksimal bimbingan hari ini. Maksimal {pengaturan['maksimal_bimbingan_per_hari']} kali per hari. Silakan coba lagi besok."
                    cursor.close()
                    return jsonify({
                        'success': False, 
                        'message': auto_pesan
                    })
        
        # Buat tabel bimbingan jika belum ada
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS bimbingan (
                id INT AUTO_INCREMENT PRIMARY KEY,
                nim VARCHAR(20) NOT NULL,
                proposal_id INT NOT NULL,
                judul_bimbingan VARCHAR(255) NOT NULL,
                hasil_bimbingan TEXT,
                tanggal_buat DATETIME DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (nim) REFERENCES mahasiswa(nim) ON DELETE CASCADE,
                FOREIGN KEY (proposal_id) REFERENCES proposal(id) ON DELETE CASCADE
            )
        ''')
        
        # Simpan hasil bimbingan
        cursor.execute('''
            INSERT INTO bimbingan (nim, proposal_id, judul_bimbingan, hasil_bimbingan)
            VALUES (%s, %s, %s, %s)
        ''', (session['nim'], proposal['id'], judul_bimbingan, hasil_bimbingan))
        
        app_funcs['mysql'].connection.commit()
        
        # Log aktivitas mahasiswa
        mahasiswa_info = app_funcs['get_mahasiswa_info_from_session']()
        if mahasiswa_info:
            data_lama = {
                'proposal_id': proposal['id'],
                'judul_usaha': proposal['judul_usaha'],
                'jenis_aktivitas': 'bimbingan'
            }
            data_baru = {
                'judul_bimbingan': judul_bimbingan,
                'hasil_bimbingan': hasil_bimbingan
            }
            log_mahasiswa_bimbingan_activity(
                app_funcs, 
                mahasiswa_info, 
                'tambah', 
                'simpan_hasil_bimbingan',
                f'Menyimpan hasil bimbingan: {judul_bimbingan}',
                data_lama,
                data_baru
            )
        
        cursor.close()
        
        return jsonify({
            'success': True,
            'message': 'Hasil bimbingan berhasil disimpan!'
        })
        
    except Exception as e:
        logger.error(f"Error in simpan_pertanyaan_bimbingan: {str(e)}")
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@mahasiswa_bp.route('/cek_batasan_bimbingan')
def cek_batasan_bimbingan():
    """API untuk mengecek batasan bimbingan mahasiswa"""
    if 'user_type' not in session or session['user_type'] != 'mahasiswa':
        return jsonify({'success': False, 'message': 'Anda harus login sebagai mahasiswa!'})
    
    app_funcs = get_app_functions()
    if not hasattr(app_funcs['mysql'], 'connection') or app_funcs['mysql'].connection is None:
        return jsonify({'success': False, 'message': 'Koneksi ke database gagal!'})
    
    try:
        cursor = app_funcs['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Cek apakah tabel pengaturan_bimbingan ada
        cursor.execute("""
            SELECT COUNT(*) as count 
            FROM information_schema.tables 
            WHERE table_schema = DATABASE() 
            AND table_name = 'pengaturan_bimbingan'
        """)
        pengaturan_table_exists = cursor.fetchone()['count'] > 0
        
        if not pengaturan_table_exists:
            cursor.close()
            return jsonify({
                'success': True,
                'data': {
                    'status_aktif': 'nonaktif',
                    'maksimal_bimbingan_per_hari': 3,
                    'bimbingan_hari_ini': 0,
                    'pesan_batasan': 'Pengaturan bimbingan belum diatur'
                }
            })
        
        # Ambil pengaturan bimbingan
        cursor.execute('SELECT * FROM pengaturan_bimbingan ORDER BY id DESC LIMIT 1')
        pengaturan = cursor.fetchone()
        
        if not pengaturan:
            cursor.close()
            return jsonify({
                'success': True,
                'data': {
                    'status_aktif': 'nonaktif',
                    'maksimal_bimbingan_per_hari': 3,
                    'bimbingan_hari_ini': 0,
                    'pesan_batasan': 'Pengaturan bimbingan belum diatur'
                }
            })
        
        # Cek jumlah bimbingan hari ini
        cursor.execute('''
            SELECT COUNT(*) as count 
            FROM bimbingan 
            WHERE nim = %s AND DATE(tanggal_buat) = CURDATE()
        ''', (session['nim'],))
        
        bimbingan_hari_ini = cursor.fetchone()['count']
        
        cursor.close()
        
        auto_pesan = f"Anda telah mencapai batas maksimal bimbingan hari ini. Maksimal {pengaturan['maksimal_bimbingan_per_hari']} kali per hari. Silakan coba lagi besok."
        return jsonify({
            'success': True,
            'data': {
                'status_aktif': pengaturan['status_aktif'],
                'maksimal_bimbingan_per_hari': pengaturan['maksimal_bimbingan_per_hari'],
                'bimbingan_hari_ini': bimbingan_hari_ini,
                'pesan_batasan': auto_pesan
            }
        })
        
    except Exception as e:
        logger.error(f"Error in cek_batasan_bimbingan: {str(e)}")
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@mahasiswa_bp.route('/detail_bimbingan')
def detail_bimbingan():
    """API untuk mendapatkan detail bimbingan"""
    if 'user_type' not in session or session['user_type'] != 'mahasiswa':
        return jsonify({'success': False, 'message': 'Anda harus login sebagai mahasiswa!'})
    
    app_funcs = get_app_functions()
    if not hasattr(app_funcs['mysql'], 'connection') or app_funcs['mysql'].connection is None:
        return jsonify({'success': False, 'message': 'Koneksi ke database gagal!'})
    
    try:
        bimbingan_id = request.args.get('id')
        if not bimbingan_id:
            return jsonify({'success': False, 'message': 'ID bimbingan tidak ditemukan!'})
        
        cursor = app_funcs['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Ambil detail bimbingan
        cursor.execute('''
            SELECT b.*, p.judul_usaha, p.dosen_pembimbing
            FROM bimbingan b
            JOIN proposal p ON b.proposal_id = p.id
            WHERE b.id = %s AND b.nim = %s
        ''', (bimbingan_id, session['nim']))
        
        bimbingan = cursor.fetchone()
        if not bimbingan:
            cursor.close()
            return jsonify({'success': False, 'message': 'Data bimbingan tidak ditemukan!'})
        
        cursor.close()
        

        
        # Generate HTML untuk modal
        html = f'''
        <div class="row">
            <div class="col-12">
                <div class="mb-3">
                    <h6 class="text-primary">
                        <i class="bi bi-info-circle"></i>
                        Informasi Bimbingan:
                    </h6>
                    <div class="row">
                        <div class="col-md-6">
                            <p><strong>Judul Usaha:</strong> {bimbingan['judul_usaha']}</p>
                            <p><strong>Dosen Pembimbing:</strong> {bimbingan['dosen_pembimbing']}</p>
                        </div>
                        <div class="col-md-6">
                            <p><strong>Tanggal Buat:</strong> {bimbingan['tanggal_buat'].strftime('%d/%m/%Y %H:%M') if bimbingan['tanggal_buat'] else '-'}</p>
                        </div>
                    </div>
                </div>
                
                <div class="mb-3">
                    <h6 class="text-primary">
                        <i class="bi bi-type-bold"></i>
                        Judul/Topik Bimbingan:
                    </h6>
                    <div class="hasil-bimbingan-content">
                        <strong>{bimbingan['judul_bimbingan']}</strong>
                    </div>
                </div>
                
                <div class="mb-3">
                    <h6 class="text-success">
                        <i class="bi bi-pencil-square"></i>
                        Hasil Bimbingan:
                    </h6>
                    <div class="hasil-bimbingan-content">
                        {bimbingan['hasil_bimbingan']}
                    </div>
                </div>
            </div>
        </div>
        '''
        
        return jsonify({
            'success': True,
            'html': html
        })
        
    except Exception as e:
        logger.error(f"Error in detail_bimbingan: {str(e)}")
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@mahasiswa_bp.route('/download_excel_anggaran')
def mahasiswa_download_excel_anggaran():
    """Download Excel anggaran awal/bertumbuh dengan format kop surat untuk Mahasiswa"""
    try:
        # Cek login
        if 'user_type' not in session or session.get('user_type') != 'mahasiswa':
            return jsonify({'success': False, 'message': 'Unauthorized'}), 401
        
        id_proposal = request.args.get('id_proposal')
        jenis = request.args.get('jenis')  # 'awal' atau 'bertumbuh'
        
        if not id_proposal or not jenis:
            return jsonify({'success': False, 'message': 'Parameter tidak lengkap'}), 400
        
        # Koneksi database
        app_funcs = get_app_functions()
        if not hasattr(app_funcs['mysql'], 'connection') or app_funcs['mysql'].connection is None:
            return jsonify({'success': False, 'message': 'Koneksi database tidak tersedia'}), 500
        
        cursor = app_funcs['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Ambil data proposal dan pastikan mahasiswa berhak mengakses
        cursor.execute("""
            SELECT p.id, p.judul_usaha, m.nama_ketua as nama, m.nim 
            FROM proposal p
            JOIN mahasiswa m ON p.nim = m.nim
            WHERE p.id = %s AND m.nim = %s
        """, (id_proposal, session.get('nim')))
        proposal_data = cursor.fetchone()
        
        if not proposal_data:
            return jsonify({'success': False, 'message': 'Proposal tidak ditemukan atau Anda tidak berhak mengakses'}), 404
        
        # Tentukan tabel berdasarkan jenis
        tabel_anggaran = f'anggaran_{jenis}'
        
        # Ambil data anggaran
        cursor.execute(f"""
            SELECT kegiatan_utama, kegiatan, nama_barang, kuantitas, satuan, 
                   harga_satuan, jumlah, keterangan, target_capaian, penanggung_jawab
            FROM {tabel_anggaran}
            WHERE id_proposal = %s
            ORDER BY kegiatan_utama, kegiatan, nama_barang
        """, (id_proposal,))
        anggaran_data = cursor.fetchall()
        
        # Ambil tahapan usaha untuk menentukan urutan kegiatan utama
        cursor.execute("""
            SELECT tahapan_usaha FROM proposal WHERE id = %s
        """, (id_proposal,))
        proposal_info = cursor.fetchone()
        tahapan_usaha = proposal_info.get('tahapan_usaha', '').lower() if proposal_info else ''
        
        # Urutkan data sesuai urutan kegiatan utama yang standardisasi
        if 'bertumbuh' in tahapan_usaha:
            urutan_kegiatan = [
                "Pengembangan Pasar dan Saluran Distribusi",
                "Pengembangan Produk/Riset",
                "Produksi",
                "Legalitas, perizinan, sertifikasi, dan standarisasi",
                "Belanja ATK dan Penunjang"
            ]
        else:
            urutan_kegiatan = [
                "Pengembangan Produk/Riset",
                "Produksi",
                "Legalitas, perizinan, sertifikasi, dan standarisasi",
                "Belanja ATK dan Penunjang"
            ]
        
        # Buat mapping untuk case-insensitive matching dan normalisasi teks
        urutan_kegiatan_lower = [k.strip().lower() for k in urutan_kegiatan]
        
        # Fungsi helper untuk normalisasi teks kegiatan utama
        def normalize_kegiatan_utama(text):
            if not text:
                return ""
            # Normalisasi: lowercase, strip whitespace, replace multiple spaces
            normalized = text.strip().lower()
            # Handle berbagai variasi teks
            replacements = {
                "legalitas, perijinan, sertifikasi, pengujian produk, dan standarisasi": "legalitas, perizinan, sertifikasi, dan standarisasi",
                "legalitas, perijinan, sertifikasi, dan standarisasi": "legalitas, perizinan, sertifikasi, dan standarisasi",
                "legalitas perijinan sertifikasi pengujian produk dan standarisasi": "legalitas, perizinan, sertifikasi, dan standarisasi",
                "legalitas perijinan sertifikasi dan standarisasi": "legalitas, perizinan, sertifikasi, dan standarisasi"
            }
            for old, new in replacements.items():
                if old in normalized:
                    normalized = normalized.replace(old, new)
            return normalized
        
        # Sort data berdasarkan urutan kegiatan utama dengan normalisasi
        anggaran_data = sorted(
            anggaran_data,
            key=lambda x: urutan_kegiatan_lower.index(normalize_kegiatan_utama(x['kegiatan_utama'])) if x['kegiatan_utama'] and normalize_kegiatan_utama(x['kegiatan_utama']) in urutan_kegiatan_lower else 99
        )
        
        if not anggaran_data:
            return jsonify({'success': False, 'message': 'Data anggaran tidak ditemukan'}), 404
        
        # Buat workbook Excel
        wb = Workbook()
        ws = wb.active
        ws.title = f"Anggaran {jenis.title()}"
        
        # Styling
        header_font = Font(bold=True, color="000000", size=12)
        title_font = Font(bold=True, color="000000", size=14)
        border_thin = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        # Header kop surat (sama dengan laporan laba rugi)
        current_row = 1
        
        # Menggunakan format kop surat yang sama dengan laporan laba rugi
        from utils import create_pdf_header_from_kop_surat_pdf
        try:
            header_template = create_pdf_header_from_kop_surat_pdf()
            header_data = header_template['header_data']
            
            # Gunakan data dari kop surat yang sama
            for i, header_row in enumerate(header_data[:6]):  # Ambil 6 baris pertama
                if len(header_row) > 1 and header_row[1].strip():  # Jika ada teks
                    ws.merge_cells(f'A{current_row}:K{current_row}')
                    ws[f'A{current_row}'] = header_row[1].strip()
                    if i < 2:  # Baris 1-2 bold dan lebih besar
                        ws[f'A{current_row}'].font = Font(bold=True, size=12, name='Times New Roman')
                    else:  # Baris lainnya normal
                        ws[f'A{current_row}'].font = Font(bold=False, size=10, name='Times New Roman')
                    ws[f'A{current_row}'].alignment = center_alignment
                    current_row += 1
                elif not header_row[1].strip():  # Baris kosong
                    current_row += 1
        except Exception as e:
            # Fallback ke format default jika gagal
            print(f"Warning: Menggunakan fallback kop surat: {e}")
            ws.merge_cells(f'A{current_row}:K{current_row}')
            ws[f'A{current_row}'] = "YAYASAN KARTIKA EKA PAKSI"
            ws[f'A{current_row}'].font = Font(bold=True, size=12, name='Times New Roman')
            ws[f'A{current_row}'].alignment = center_alignment
            current_row += 1
            
            ws.merge_cells(f'A{current_row}:K{current_row}')
            ws[f'A{current_row}'] = "UNIVERSITAS JENDERAL ACHMAD YANI YOGYAKARTA"
            ws[f'A{current_row}'].font = Font(bold=True, size=12, name='Times New Roman')
            ws[f'A{current_row}'].alignment = center_alignment
            current_row += 1
            
            ws.merge_cells(f'A{current_row}:K{current_row}')
            ws[f'A{current_row}'] = "Jl Siliwangi Ringroad Barat, Banyuraden, Gamping, Sleman, Yogyakarta 55293"
            ws[f'A{current_row}'].font = Font(bold=False, size=10, name='Times New Roman')
            ws[f'A{current_row}'].alignment = center_alignment
            current_row += 1
            
            ws.merge_cells(f'A{current_row}:K{current_row}')
            ws[f'A{current_row}'] = "Telp. (0274) 552489, 552851 Fax. (0274) 557228 Website: www.unjaya.ac.id"
            ws[f'A{current_row}'].font = Font(bold=False, size=10, name='Times New Roman')
            ws[f'A{current_row}'].alignment = center_alignment
            current_row += 1
            
            ws.merge_cells(f'A{current_row}:K{current_row}')
            ws[f'A{current_row}'] = "Email: info@unjaya.ac.id"
            ws[f'A{current_row}'].font = Font(bold=False, size=10, name='Times New Roman')
            ws[f'A{current_row}'].alignment = center_alignment
            current_row += 1
        
        # Tambahkan spasi setelah kop surat
        current_row += 1
        
        # Judul laporan
        ws.merge_cells(f'A{current_row}:K{current_row}')
        ws[f'A{current_row}'] = f"RENCANA ANGGARAN {jenis.upper()}"
        ws[f'A{current_row}'].font = title_font
        ws[f'A{current_row}'].alignment = center_alignment
        current_row += 2
        
        # Info mahasiswa
        ws[f'A{current_row}'] = "Nama Mahasiswa:"
        ws[f'A{current_row}'].font = header_font
        ws[f'B{current_row}'] = proposal_data['nama']
        current_row += 1
        
        ws[f'A{current_row}'] = "NIM:"
        ws[f'A{current_row}'].font = header_font
        ws[f'B{current_row}'] = proposal_data['nim']
        current_row += 1
        
        ws[f'A{current_row}'] = "Nama Usaha:"
        ws[f'A{current_row}'].font = header_font
        ws[f'B{current_row}'] = proposal_data['judul_usaha']
        current_row += 1
        
        ws[f'A{current_row}'] = "Judul Usaha:"
        ws[f'A{current_row}'].font = header_font
        ws[f'B{current_row}'] = proposal_data['judul_usaha']
        current_row += 2
        
        # Header tabel dengan merged cells
        # Row 1: Header utama
        ws.cell(row=current_row, column=1, value="Kegiatan Utama").font = header_font
        ws.cell(row=current_row, column=1).fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
        ws.cell(row=current_row, column=1).border = border_thin
        ws.cell(row=current_row, column=1).alignment = center_alignment
        
        # Merge cells untuk "Rencana" (kolom B-G)
        ws.merge_cells(f'B{current_row}:G{current_row}')
        ws[f'B{current_row}'] = "Rencana"
        ws[f'B{current_row}'].font = header_font
        ws[f'B{current_row}'].fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
        ws[f'B{current_row}'].border = border_thin
        ws[f'B{current_row}'].alignment = center_alignment
        
        ws.cell(row=current_row, column=8, value="Keterangan/Ref. Harga").font = header_font
        ws.cell(row=current_row, column=8).fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
        ws.cell(row=current_row, column=8).border = border_thin
        ws.cell(row=current_row, column=8).alignment = center_alignment
        
        ws.cell(row=current_row, column=9, value="Target Capaian H=Output A").font = header_font
        ws.cell(row=current_row, column=9).fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
        ws.cell(row=current_row, column=9).border = border_thin
        ws.cell(row=current_row, column=9).alignment = center_alignment
        
        ws.cell(row=current_row, column=10, value="Penanggung Jawab").font = header_font
        ws.cell(row=current_row, column=10).fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
        ws.cell(row=current_row, column=10).border = border_thin
        ws.cell(row=current_row, column=10).alignment = center_alignment
        
        current_row += 1
        
        # Row 2: Sub-header untuk Rencana
        ws.cell(row=current_row, column=1, value="").border = border_thin  # Empty cell untuk alignment
        
        # Sub-header untuk kolom Rencana
        sub_headers = ["Kegiatan A", "Nama Barang B", "Kuantitas C", "Satuan D", "Harga Satuan E", "Jumlah (Rp) F=C×E"]
        for col, sub_header in enumerate(sub_headers, 2):
            cell = ws.cell(row=current_row, column=col, value=sub_header)
            cell.font = header_font
            cell.fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
            cell.border = border_thin
            cell.alignment = center_alignment
        
        # Empty cells untuk kolom lainnya
        for col in range(8, 11):
            ws.cell(row=current_row, column=col, value="").border = border_thin
        
        current_row += 1
        
        # Tulis data anggaran dengan grouping berdasarkan kegiatan utama
        total_jumlah = 0
        current_kegiatan_utama = None
        kegiatan_utama_start_row = None
        kegiatan_utama_row_count = 0
        
        current_kegiatan = None
        kegiatan_start_row = None
        kegiatan_row_count = 0
        
        current_target_capaian = None
        target_capaian_start_row = None
        target_capaian_row_count = 0
        
        current_penanggung_jawab = None
        penanggung_jawab_start_row = None
        penanggung_jawab_row_count = 0
        
        for i, row_data in enumerate(anggaran_data):
            # Cek jika kegiatan utama berubah
            if current_kegiatan_utama != row_data['kegiatan_utama']:
                # Merge cells untuk kegiatan utama sebelumnya jika ada
                if current_kegiatan_utama is not None and kegiatan_utama_row_count > 1:
                    ws.merge_cells(f'A{kegiatan_utama_start_row}:A{current_row-1}')
                    ws[f'A{kegiatan_utama_start_row}'].alignment = Alignment(horizontal='center', vertical='center')
                
                # Set kegiatan utama baru
                current_kegiatan_utama = row_data['kegiatan_utama']
                kegiatan_utama_start_row = current_row
                kegiatan_utama_row_count = 0
            
            # Cek jika kegiatan berubah
            if current_kegiatan != row_data['kegiatan']:
                # Merge cells untuk kegiatan sebelumnya jika ada
                if current_kegiatan is not None and kegiatan_row_count > 1:
                    ws.merge_cells(f'B{kegiatan_start_row}:B{current_row-1}')
                    ws[f'B{kegiatan_start_row}'].alignment = Alignment(horizontal='center', vertical='center')
                
                # Set kegiatan baru
                current_kegiatan = row_data['kegiatan']
                kegiatan_start_row = current_row
                kegiatan_row_count = 0
            
            # Cek jika target capaian berubah
            if current_target_capaian != row_data['target_capaian']:
                # Merge cells untuk target capaian sebelumnya jika ada
                if current_target_capaian is not None and target_capaian_row_count > 1:
                    ws.merge_cells(f'I{target_capaian_start_row}:I{current_row-1}')
                    ws[f'I{target_capaian_start_row}'].alignment = Alignment(horizontal='center', vertical='center')
                
                # Set target capaian baru
                current_target_capaian = row_data['target_capaian']
                target_capaian_start_row = current_row
                target_capaian_row_count = 0
            
            # Cek jika penanggung jawab berubah
            if current_penanggung_jawab != row_data['penanggung_jawab']:
                # Merge cells untuk penanggung jawab sebelumnya jika ada
                if current_penanggung_jawab is not None and penanggung_jawab_row_count > 1:
                    ws.merge_cells(f'J{penanggung_jawab_start_row}:J{current_row-1}')
                    ws[f'J{penanggung_jawab_start_row}'].alignment = Alignment(horizontal='center', vertical='center')
                
                # Set penanggung jawab baru
                current_penanggung_jawab = row_data['penanggung_jawab']
                penanggung_jawab_start_row = current_row
                penanggung_jawab_row_count = 0
            
            # Tulis data
            ws.cell(row=current_row, column=1, value=row_data['kegiatan_utama']).alignment = left_alignment
            ws.cell(row=current_row, column=2, value=row_data['kegiatan']).alignment = left_alignment
            ws.cell(row=current_row, column=3, value=row_data['nama_barang']).alignment = left_alignment
            ws.cell(row=current_row, column=4, value=row_data['kuantitas']).alignment = center_alignment
            ws.cell(row=current_row, column=5, value=row_data['satuan']).alignment = center_alignment
            ws.cell(row=current_row, column=6, value=f"Rp {row_data['harga_satuan']:,}").alignment = center_alignment
            ws.cell(row=current_row, column=7, value=f"Rp {row_data['jumlah']:,}").alignment = center_alignment
            ws.cell(row=current_row, column=8, value=row_data['keterangan'] or '-').alignment = left_alignment
            ws.cell(row=current_row, column=9, value=row_data['target_capaian']).alignment = left_alignment
            ws.cell(row=current_row, column=10, value=row_data['penanggung_jawab']).alignment = left_alignment
            
            # Apply borders
            for col in range(1, 11):
                ws.cell(row=current_row, column=col).border = border_thin
            
            total_jumlah += row_data['jumlah']
            current_row += 1
            kegiatan_utama_row_count += 1
            kegiatan_row_count += 1
            target_capaian_row_count += 1
            penanggung_jawab_row_count += 1
        
        # Merge cells untuk kegiatan utama terakhir jika ada lebih dari 1 row
        if kegiatan_utama_row_count > 1:
            ws.merge_cells(f'A{kegiatan_utama_start_row}:A{current_row-1}')
            ws[f'A{kegiatan_utama_start_row}'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Merge cells untuk kegiatan terakhir jika ada lebih dari 1 row
        if kegiatan_row_count > 1:
            ws.merge_cells(f'B{kegiatan_start_row}:B{current_row-1}')
            ws[f'B{kegiatan_start_row}'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Merge cells untuk target capaian terakhir jika ada lebih dari 1 row
        if target_capaian_row_count > 1:
            ws.merge_cells(f'I{target_capaian_start_row}:I{current_row-1}')
            ws[f'I{target_capaian_start_row}'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Merge cells untuk penanggung jawab terakhir jika ada lebih dari 1 row
        if penanggung_jawab_row_count > 1:
            ws.merge_cells(f'J{penanggung_jawab_start_row}:J{current_row-1}')
            ws[f'J{penanggung_jawab_start_row}'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Total row
        ws.merge_cells(f'A{current_row}:F{current_row}')
        ws[f'A{current_row}'] = "TOTAL ANGGARAN:"
        ws[f'A{current_row}'].font = Font(bold=True)
        ws[f'A{current_row}'].alignment = Alignment(horizontal='right', vertical='center')
        ws[f'A{current_row}'].border = border_thin
        
        ws[f'G{current_row}'] = f"Rp {total_jumlah:,}"
        ws[f'G{current_row}'].font = Font(bold=True)
        ws[f'G{current_row}'].fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
        ws[f'G{current_row}'].alignment = center_alignment
        ws[f'G{current_row}'].border = border_thin
        
        # Merge remaining cells in total row
        for col in range(8, 11):
            ws.cell(row=current_row, column=col).border = border_thin
        
        current_row += 2
        
        # Footer dengan tanggal
        ws[f'H{current_row}'] = f"Yogyakarta, {datetime.now().strftime('%d %B %Y')}"
        ws[f'H{current_row}'].alignment = center_alignment
        current_row += 3
        
        ws[f'H{current_row}'] = "Mahasiswa,"
        ws[f'H{current_row}'].alignment = center_alignment
        current_row += 4
        
        ws[f'H{current_row}'] = proposal_data['nama']
        ws[f'H{current_row}'].font = Font(bold=True)
        ws[f'H{current_row}'].alignment = center_alignment
        current_row += 1
        
        ws[f'H{current_row}'] = f"NIM: {proposal_data['nim']}"
        ws[f'H{current_row}'].alignment = center_alignment
        
        # Auto-adjust column widths
        for col in range(1, 11):
            column_letter = get_column_letter(col)
            max_length = 0
            for row in range(1, current_row + 1):
                try:
                    cell_value = str(ws[f'{column_letter}{row}'].value)
                    if len(cell_value) > max_length:
                        max_length = len(cell_value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Max width 50
            ws.column_dimensions[column_letter].width = max(adjusted_width, 15)  # Min width 15
        
        # Simpan ke temporary file
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        wb.save(temp_file.name)
        temp_file.close()
        
        # Buat response untuk download
        filename = f"Anggaran_{jenis.title()}_{proposal_data['judul_usaha']}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            temp_file.name,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        print(f"Error in mahasiswa_download_excel_anggaran: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'message': 'Terjadi kesalahan saat membuat file Excel'}), 500
    finally:
        if 'cursor' in locals():
            cursor.close()