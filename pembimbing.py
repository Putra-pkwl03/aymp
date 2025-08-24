from flask import Blueprint, render_template, request, jsonify, session, flash, redirect, url_for, send_file, abort, make_response
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import tempfile
import os
import os
import MySQLdb
import traceback
import logging
from datetime import datetime, timedelta
import re
from werkzeug.security import generate_password_hash, check_password_hash

# Import dari utils
from utils import group_anggaran_data, flatten_anggaran_data, format_tanggal_indonesia

# Setup logging untuk pembimbing blueprint (tanpa file handler)
logging.basicConfig(
    level=logging.DEBUG,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler()  # Hanya console handler, tidak ada file handler
    ]
)
logger = logging.getLogger('pembimbing')

def ensure_penilaian_tables(cursor, connection):
    """Pastikan semua tabel penilaian mahasiswa sudah dibuat"""
    try:
        # Tabel pertanyaan penilaian
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS pertanyaan_penilaian_mahasiswa (
                id INT AUTO_INCREMENT PRIMARY KEY,
                kategori VARCHAR(100) NOT NULL COMMENT 'Kategori penilaian mahasiswa',
                pertanyaan TEXT NOT NULL,
                bobot DECIMAL(5,2) NOT NULL COMMENT 'Bobot yang diinput admin (1-100)',
                skor_maksimal INT NOT NULL COMMENT 'Skor maksimal yang bisa diberikan (1-100)',
                status ENUM('Aktif', 'Nonaktif') DEFAULT 'Aktif',
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        
        # Tabel header penilaian
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS penilaian_mahasiswa (
                id INT AUTO_INCREMENT PRIMARY KEY,
                id_proposal INT NOT NULL,
                id_pembimbing INT NOT NULL,
                nilai_akhir DECIMAL(5,2) DEFAULT 0 COMMENT 'Total nilai akhir',
                komentar_pembimbing TEXT,
                status ENUM('Draft', 'Selesai') DEFAULT 'Draft',
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                FOREIGN KEY (id_proposal) REFERENCES proposal(id) ON DELETE CASCADE,
                FOREIGN KEY (id_pembimbing) REFERENCES pembimbing(id) ON DELETE CASCADE,
                UNIQUE KEY unique_penilaian_mahasiswa (id_proposal, id_pembimbing)
            )
        ''')
        
        # Tabel detail penilaian dengan fitur sesi
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS detail_penilaian_mahasiswa (
                id INT AUTO_INCREMENT PRIMARY KEY,
                id_penilaian_mahasiswa INT NOT NULL,
                id_pertanyaan INT NOT NULL,
                skor INT NOT NULL COMMENT 'Skor yang diberikan pembimbing',
                nilai DECIMAL(5,2) NOT NULL COMMENT 'Nilai = (skor/skor_maksimal) * bobot',
                sesi_penilaian INT NOT NULL DEFAULT 1 COMMENT 'Sesi penilaian ke-berapa (1,2,3,dst)',
                is_locked TINYINT(1) NOT NULL DEFAULT 0 COMMENT 'Apakah skor sudah dikunci',
                tanggal_input DATETIME NULL COMMENT 'Tanggal skor diinput',
                metadata_kolom_id INT NULL COMMENT 'Referensi ke metadata kolom dari admin',
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (id_penilaian_mahasiswa) REFERENCES penilaian_mahasiswa(id) ON DELETE CASCADE,
                FOREIGN KEY (id_pertanyaan) REFERENCES pertanyaan_penilaian_mahasiswa(id) ON DELETE CASCADE,
                UNIQUE KEY unique_detail_penilaian_mahasiswa (id_penilaian_mahasiswa, id_pertanyaan, sesi_penilaian)
            )
        ''')
        
        # Tambahkan kolom metadata_kolom_id jika belum ada (untuk backward compatibility)
        try:
            cursor.execute("ALTER TABLE detail_penilaian_mahasiswa ADD COLUMN metadata_kolom_id INT NULL")
        except:
            pass  # Kolom sudah ada
        
        # Tabel riwayat sudah tidak diperlukan karena data tersimpan di detail_penilaian_mahasiswa
        
        connection.commit()
        return True
    except Exception as e:
        logger.error(f"Error creating penilaian tables: {str(e)}")
        return False

# Lazy import untuk menghindari circular import
def get_app_functions():
    """Lazy import untuk mendapatkan fungsi dari app.py"""
    from app import (
        mysql, get_pembimbing_info_from_session, log_pembimbing_activity,
        generate_excel_laba_rugi, generate_pdf_laba_rugi, generate_word_laba_rugi,
        generate_excel_arus_kas, generate_pdf_arus_kas, generate_word_arus_kas,
        generate_excel_neraca, generate_pdf_neraca, generate_word_neraca,
        hitung_neraca_real_time,
        hapus_file_laporan_kemajuan, hapus_file_laporan_akhir, update_laporan_kemajuan_from_anggaran,
        hapus_file_laporan_akhir_terkait
    )
    return {
        'mysql': mysql,
        'get_pembimbing_info_from_session': get_pembimbing_info_from_session,
        'log_pembimbing_activity': log_pembimbing_activity,
        'generate_excel_laba_rugi': generate_excel_laba_rugi,
        'generate_pdf_laba_rugi': generate_pdf_laba_rugi,
        'generate_word_laba_rugi': generate_word_laba_rugi,
        'generate_excel_arus_kas': generate_excel_arus_kas,
        'generate_pdf_arus_kas': generate_pdf_arus_kas,
        'generate_word_arus_kas': generate_word_arus_kas,
        'generate_excel_neraca': generate_excel_neraca,
        'generate_pdf_neraca': generate_pdf_neraca,
        'generate_word_neraca': generate_word_neraca,
        'hitung_neraca_real_time': hitung_neraca_real_time,
        'hapus_file_laporan_kemajuan': hapus_file_laporan_kemajuan,
        'hapus_file_laporan_akhir': hapus_file_laporan_akhir,
        'update_laporan_kemajuan_from_anggaran': update_laporan_kemajuan_from_anggaran,
        'hapus_file_laporan_akhir_terkait': hapus_file_laporan_akhir_terkait
    }

# Import fungsi hitung_neraca_real_time akan dilakukan secara lazy di dalam fungsi

# Create blueprint
pembimbing_bp = Blueprint('pembimbing', __name__, url_prefix='/pembimbing')


@pembimbing_bp.route('/das_pembimbing')
def das_pembimbing():
    logger.info("Dashboard pembimbing dipanggil")
    if 'user_type' not in session or session['user_type'] != 'pembimbing':
        logger.warning(f"Akses ditolak untuk dashboard pembimbing: {session.get('user_type', 'None')}")
        flash('Anda harus login sebagai pembimbing!', 'danger')
        return redirect(url_for('index'))
    
    app_funcs = get_app_functions()
    if not hasattr(app_funcs['mysql'], 'connection') or app_funcs['mysql'].connection is None:
        logger.error("Koneksi database tidak tersedia untuk dashboard pembimbing")
        flash('Koneksi ke database gagal. Cek konfigurasi database!', 'danger')
        return render_template('pembimbing/index pembimbing.html', mahasiswa_bimbingan=[])
    
    try:
        cursor = app_funcs['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Ambil data mahasiswa yang dibimbing dengan informasi proposal dan anggota tim
        cursor.execute('''
            SELECT DISTINCT 
                m.id as mahasiswa_id,
                m.nama_ketua,
                m.nim,
                m.program_studi,
                m.perguruan_tinggi,
                m.status as status_mahasiswa,
                p.id as proposal_id,
                p.judul_usaha,
                p.status as status_proposal,
                p.status_admin,
                p.tanggal_buat,
                COUNT(at.id) as jumlah_anggota
            FROM mahasiswa m
            INNER JOIN proposal p ON m.nim = p.nim
            LEFT JOIN anggota_tim at ON p.id = at.id_proposal
            WHERE p.dosen_pembimbing = %s 
            AND p.status IN ('diajukan', 'disetujui', 'ditolak', 'revisi', 'selesai')
            GROUP BY m.id, m.nama_ketua, m.nim, m.program_studi, m.perguruan_tinggi, 
                     m.status, p.id, p.judul_usaha, p.status, p.status_admin, p.tanggal_buat
            ORDER BY p.tanggal_buat DESC
        ''', (session['nama'],))
        
        mahasiswa_bimbingan = cursor.fetchall()
        
        # Hitung statistik untuk dashboard
        total_mahasiswa = len(mahasiswa_bimbingan)
        proposal_lolos = len([m for m in mahasiswa_bimbingan if m['status_admin'] == 'lolos'])
        menunggu_review = len([m for m in mahasiswa_bimbingan if m['status_proposal'] == 'diajukan'])
        proposal_disetujui = len([m for m in mahasiswa_bimbingan if m['status_proposal'] == 'disetujui'])
        
        cursor.close()
        
        # Log aktivitas pembimbing
        pembimbing_info = app_funcs['get_pembimbing_info_from_session']()
        if pembimbing_info:
            app_funcs['log_pembimbing_activity'](
                pembimbing_info['id'],
                pembimbing_info['nip'],
                pembimbing_info['nama'],
                'view',
                'dashboard',
                'das_pembimbing',
                'Mengakses dashboard pembimbing'
            )
        
        return render_template('pembimbing/index pembimbing.html', 
                             mahasiswa_bimbingan=mahasiswa_bimbingan,
                             total_mahasiswa=total_mahasiswa,
                             proposal_lolos=proposal_lolos,
                             menunggu_review=menunggu_review,
                             proposal_disetujui=proposal_disetujui)
        
    except Exception as e:
        logger.error(f"Error saat mengambil data dashboard pembimbing: {str(e)}")
        logger.error(traceback.format_exc())
        flash(f'Error saat mengambil data: {str(e)}', 'danger')
        return render_template('pembimbing/index pembimbing.html', 
                             mahasiswa_bimbingan=[],
                             total_mahasiswa=0,
                             proposal_lolos=0,
                             menunggu_review=0,
                             proposal_disetujui=0)

@pembimbing_bp.route('/pembimbing_proposal')
def pembimbing_proposal():
    if 'user_type' not in session or session['user_type'] != 'pembimbing':
        flash('Anda harus login sebagai pembimbing!', 'danger')
        return redirect(url_for('index'))
    
    app_funcs = get_app_functions()
    if not hasattr(app_funcs['mysql'], 'connection') or app_funcs['mysql'].connection is None:
        flash('Koneksi ke database gagal. Cek konfigurasi database!', 'danger')
        return render_template('pembimbing/proposal.html', proposals=[])
    
    try:
        cursor = app_funcs['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Ambil data proposal mahasiswa yang dibimbing oleh pembimbing ini
        cursor.execute('''
            SELECT p.*, 
                   m.nama_ketua, m.perguruan_tinggi, m.program_studi, m.email, m.no_telp,
                   COUNT(at.id) as jumlah_anggota
            FROM proposal p 
            LEFT JOIN mahasiswa m ON p.nim = m.nim
            LEFT JOIN anggota_tim at ON p.id = at.id_proposal
            WHERE p.dosen_pembimbing = %s AND p.status IN ('diajukan', 'disetujui', 'ditolak', 'revisi')
            GROUP BY p.id 
            ORDER BY p.tanggal_buat DESC
        ''', (session['nama'],))
        
        proposals = cursor.fetchall()
        cursor.close()
        
        # Log aktivitas pembimbing
        pembimbing_info = app_funcs['get_pembimbing_info_from_session']()
        if pembimbing_info:
            app_funcs['log_pembimbing_activity'](
                pembimbing_info['id'],
                pembimbing_info['nip'],
                pembimbing_info['nama'],
                'view',
                'proposal',
                'daftar_proposal',
                'Melihat daftar proposal mahasiswa bimbingan'
            )
        
        return render_template('pembimbing/proposal.html', proposals=proposals)
        
    except Exception as e:
        flash(f'Error saat mengambil data proposal: {str(e)}', 'danger')
        return render_template('pembimbing/proposal.html', proposals=[])

@pembimbing_bp.route('/get_proposal_detail/<int:proposal_id>')
def pembimbing_get_proposal_detail(proposal_id):
    if 'user_type' not in session or session['user_type'] != 'pembimbing':
        return jsonify({'success': False, 'message': 'Anda harus login sebagai pembimbing!'})
    
    app_funcs = get_app_functions()
    if not hasattr(app_funcs['mysql'], 'connection') or app_funcs['mysql'].connection is None:
        return jsonify({'success': False, 'message': 'Koneksi ke database gagal!'})
    
    try:
        cursor = app_funcs['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Get proposal details with mahasiswa info, ensure it belongs to this pembimbing
        cursor.execute('''
            SELECT p.*, m.nama_ketua, m.perguruan_tinggi, m.program_studi, m.email, m.no_telp
            FROM proposal p 
            LEFT JOIN mahasiswa m ON p.nim = m.nim
            WHERE p.id = %s AND p.dosen_pembimbing = %s
        ''', (proposal_id, session['nama']))
        
        proposal = cursor.fetchone()
        if not proposal:
            cursor.close()
            return jsonify({'success': False, 'message': 'Proposal tidak ditemukan atau tidak ada akses!'})
        
        # Get team members
        cursor.execute('''
            SELECT * FROM anggota_tim 
            WHERE id_proposal = %s
            ORDER BY id
        ''', (proposal_id,))
        
        anggota = cursor.fetchall()
        
        # Get anggaran data if exists
        cursor.execute('''
            SELECT COUNT(*) as count FROM anggaran_awal WHERE id_proposal = %s
        ''', (proposal_id,))
        anggaran_awal_count = cursor.fetchone()['count']
        
        cursor.execute('''
            SELECT COUNT(*) as count FROM anggaran_bertumbuh WHERE id_proposal = %s
        ''', (proposal_id,))
        anggaran_bertumbuh_count = cursor.fetchone()['count']
        
        cursor.close()
        
        # Log aktivitas pembimbing
        pembimbing_info = app_funcs['get_pembimbing_info_from_session']()
        if pembimbing_info and proposal:
            app_funcs['log_pembimbing_activity'](
                pembimbing_info['id'],
                pembimbing_info['nip'],
                pembimbing_info['nama'],
                'view',
                'proposal',
                f'proposal_detail_id_{proposal_id}',
                f'Melihat detail proposal "{proposal["judul_usaha"]}"',
                None,
                None,
                None,  # mahasiswa_id tidak tersedia
                proposal_id
            )
        
        return jsonify({
            'success': True, 
            'proposal': proposal,
            'anggota': anggota,
            'anggaran_awal_count': anggaran_awal_count,
            'anggaran_bertumbuh_count': anggaran_bertumbuh_count
        })
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error saat mengambil data: {str(e)}'})
@pembimbing_bp.route('/update_proposal_status', methods=['POST'])
def pembimbing_update_proposal_status():
    if 'user_type' not in session or session['user_type'] != 'pembimbing':
        return jsonify({'success': False, 'message': 'Anda harus login sebagai pembimbing!'})
    
    app_funcs = get_app_functions()
    if not hasattr(app_funcs['mysql'], 'connection') or app_funcs['mysql'].connection is None:
        return jsonify({'success': False, 'message': 'Koneksi ke database gagal!'})
    
    try:
        data = request.get_json()
        proposal_id = data.get('proposal_id')
        new_status = data.get('status')  # 'disetujui', 'ditolak', 'revisi'
        catatan = data.get('catatan', '').strip()  # Ambil catatan dari request
        
        if not proposal_id or not new_status:
            return jsonify({'success': False, 'message': 'Data tidak lengkap!'})
        
        # Validasi status yang diizinkan
        if new_status not in ['disetujui', 'ditolak', 'revisi']:
            return jsonify({'success': False, 'message': 'Status tidak valid!'})
        
        cursor = app_funcs['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Verify the proposal belongs to this pembimbing
        cursor.execute('''
            SELECT id FROM proposal 
            WHERE id = %s AND dosen_pembimbing = %s
        ''', (proposal_id, session['nama']))
        
        proposal = cursor.fetchone()
        if not proposal:
            cursor.close()
            return jsonify({'success': False, 'message': 'Proposal tidak ditemukan atau tidak ada akses!'})
        
        # Ambil data proposal sebelum update untuk logging
        cursor.execute('''
            SELECT id, judul_usaha, status as status_lama
            FROM proposal 
            WHERE id = %s AND dosen_pembimbing = %s
        ''', (proposal_id, session['nama']))
        proposal_data = cursor.fetchone()
        
        # Update status proposal dan komentar pembimbing
        if new_status == 'revisi':
            cursor.execute('''
                UPDATE proposal 
                SET status = %s, 
                    komentar_revisi = %s,
                    tanggal_konfirmasi_pembimbing = NOW()
                WHERE id = %s AND dosen_pembimbing = %s
            ''', (new_status, catatan, proposal_id, session['nama']))
        else:
            cursor.execute('''
                UPDATE proposal 
                SET status = %s, 
                    komentar_pembimbing = %s,
                    tanggal_komentar_pembimbing = NOW(),
                    tanggal_konfirmasi_pembimbing = NOW()
                WHERE id = %s AND dosen_pembimbing = %s
            ''', (new_status, catatan, proposal_id, session['nama']))
        
        app_funcs['mysql'].connection.commit()
        
        # Log aktivitas pembimbing
        pembimbing_info = app_funcs['get_pembimbing_info_from_session']()
        if pembimbing_info and proposal_data:
            jenis_aktivitas = 'setuju' if new_status == 'disetujui' else 'tolak' if new_status == 'ditolak' else 'revisi'
            deskripsi = f'Mengubah status proposal "{proposal_data["judul_usaha"]}" dari {proposal_data["status_lama"]} menjadi {new_status}'
            if catatan:
                deskripsi += f' dengan catatan: {catatan[:100]}...' if len(catatan) > 100 else f' dengan catatan: {catatan}'
            
            # Data untuk tracking perubahan
            data_lama = {'status': proposal_data['status_lama']}
            data_baru = {'status': new_status, 'komentar_pembimbing': catatan}
            
            app_funcs['log_pembimbing_activity'](
                pembimbing_info['id'],
                pembimbing_info['nip'],
                pembimbing_info['nama'],
                jenis_aktivitas,
                'proposal',
                f'proposal_id_{proposal_id}',
                deskripsi,
                data_lama,
                data_baru,
                None,  # mahasiswa_id tidak tersedia
                proposal_id  # target_proposal_id
            )
        
        cursor.close()
        
        # Pesan sukses berdasarkan status dan komentar
        if catatan:
            message = f'Status proposal berhasil diubah menjadi {new_status} dengan komentar'
        else:
            message = f'Status proposal berhasil diubah menjadi {new_status}'
        
        return jsonify({'success': True, 'message': message})
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error saat update status: {str(e)}'})

@pembimbing_bp.route('/download_proposal/<int:proposal_id>')
def pembimbing_download_proposal(proposal_id):
    if 'user_type' not in session or session['user_type'] != 'pembimbing':
        flash('Anda harus login sebagai pembimbing!', 'danger')
        return redirect(url_for('index'))
    
    app_funcs = get_app_functions()
    if not hasattr(app_funcs['mysql'], 'connection') or app_funcs['mysql'].connection is None:
        flash('Koneksi ke database gagal!', 'danger')
        return redirect(url_for('pembimbing.pembimbing_proposal'))
    
    try:
        cursor = app_funcs['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Log aktivitas pembimbing
        pembimbing_info = app_funcs['get_pembimbing_info_from_session']()
        if pembimbing_info:
            app_funcs['log_pembimbing_activity'](
                pembimbing_info['id'],
                pembimbing_info['nip'],
                pembimbing_info['nama'],
                'view',
                'proposal',
                f'download_proposal_id_{proposal_id}',
                f'Mendownload file proposal ID {proposal_id}',
                None,
                None,
                None,
                proposal_id
            )
        
        # Get proposal file path, ensure it belongs to this pembimbing
        cursor.execute('''
            SELECT file_path, judul_usaha 
            FROM proposal 
            WHERE id = %s AND dosen_pembimbing = %s
        ''', (proposal_id, session['nama']))
        proposal = cursor.fetchone()
        cursor.close()
        
        if not proposal or not proposal['file_path']:
            flash('File proposal tidak ditemukan!', 'danger')
            return redirect(url_for('pembimbing.pembimbing_proposal'))
        
        file_path = proposal['file_path']
        
        if not os.path.exists(file_path):
            flash('File proposal tidak ditemukan di server!', 'danger')
            return redirect(url_for('pembimbing.pembimbing_proposal'))
        
        # Get file extension and create download filename
        file_extension = os.path.splitext(file_path)[1]
        
        # Ambil nama file asli dari path
        original_filename = os.path.basename(file_path)
        
        # Jika nama file sudah sesuai format yang diinginkan, gunakan itu
        if '_proposal.' in original_filename:
            download_filename = original_filename
        else:
            download_filename = f"{proposal['judul_usaha'].replace(' ', '_')}_proposal{file_extension}"
        
        from flask import send_file
        return send_file(file_path, as_attachment=False)
        
    except Exception as e:
        flash(f'Error saat download file: {str(e)}', 'danger')
        return redirect(url_for('pembimbing.pembimbing_proposal'))


@pembimbing_bp.route('/anggaran_awal/<int:proposal_id>')
def pembimbing_anggaran_awal(proposal_id):
    if 'user_type' not in session or session['user_type'] != 'pembimbing':
        flash('Anda harus login sebagai pembimbing!', 'danger')
        return redirect(url_for('index'))
    app_funcs = get_app_functions()
    if not hasattr(app_funcs['mysql'], 'connection') or app_funcs['mysql'].connection is None:
        flash('Koneksi ke database gagal. Cek konfigurasi database!', 'danger')
        return render_template('pembimbing/pengajuan anggaran awal.html', anggaran_data=[], proposal_id=proposal_id)
    try:
        cursor = app_funcs['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Verify the proposal belongs to this pembimbing
        cursor.execute('''
            SELECT id FROM proposal 
            WHERE id = %s AND dosen_pembimbing = %s
        ''', (proposal_id, session['nama']))
        
        proposal = cursor.fetchone()
        if not proposal:
            cursor.close()
            flash('Proposal tidak ditemukan atau tidak ada akses!', 'danger')
            return redirect(url_for('pembimbing.pembimbing_proposal'))
        
        cursor.execute('''
            SELECT id, kegiatan_utama, kegiatan, penanggung_jawab, target_capaian, 
                   nama_barang, kuantitas, satuan, harga_satuan, jumlah, keterangan, status,
                   COALESCE(nilai_bantuan, 0) as nilai_bantuan
            FROM anggaran_awal 
            WHERE id_proposal = %s 
            ORDER BY kegiatan_utama, kegiatan, nama_barang
        ''', (proposal_id,))
        anggaran_data = cursor.fetchall()
        
        cursor.execute('SELECT judul_usaha, tahapan_usaha FROM proposal WHERE id = %s', (proposal_id,))
        proposal_info = cursor.fetchone()
        
        # Urutkan data sesuai urutan kegiatan utama (logika dinamis berdasarkan tahapan usaha)
        tahapan_usaha = proposal_info.get('tahapan_usaha', '').lower() if proposal_info else ''
        if 'bertumbuh' in tahapan_usaha:
            urutan_kegiatan = [
                "Pengembangan Pasar dan Saluran Distribusi",
                "Pengembangan Produk/Riset",
                "Produksi",
                "Legalitas, perizinan, sertifikasi, dan standarisasi",
                "Belanja ATK dan Penunjang"
            ]
        else:
            urutan_kegiatan = [
                "Pengembangan Produk/Riset",
                "Produksi",
                "Legalitas, perizinan, sertifikasi, dan standarisasi",
                "Belanja ATK dan Penunjang"
            ]
        
        urutan_kegiatan_lower = [k.strip().lower() for k in urutan_kegiatan]
        anggaran_data = sorted(
            anggaran_data,
            key=lambda x: urutan_kegiatan_lower.index(x['kegiatan_utama'].strip().lower()) if x['kegiatan_utama'] and x['kegiatan_utama'].strip().lower() in urutan_kegiatan_lower else 99
        )
        cursor.close()
        
        # Log aktivitas pembimbing
        pembimbing_info = app_funcs['get_pembimbing_info_from_session']()
        if pembimbing_info and proposal_info:
            app_funcs['log_pembimbing_activity'](
                pembimbing_info['id'],
                pembimbing_info['nip'],
                pembimbing_info['nama'],
                'view',
                'anggaran_awal',
                f'proposal_id_{proposal_id}',
                f'Melihat anggaran awal proposal "{proposal_info["judul_usaha"]}"',
                None,
                None,
                None,
                proposal_id
            )
        
        grouped_data = group_anggaran_data(anggaran_data)
        anggaran_data_flat = flatten_anggaran_data(anggaran_data)
        # Total nilai bantuan pembimbing (kolom sudah COALESCE di SELECT)
        total_nilai_bantuan = sum((row.get('nilai_bantuan', 0) or 0) for row in anggaran_data)
        # Ambil batas min/maks untuk badge seperti mahasiswa
        try:
            cur2 = app_funcs['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
            cur2.execute('SELECT * FROM pengaturan_anggaran ORDER BY id DESC LIMIT 1')
            batas_row = cur2.fetchone() or {}
            batas_min_awal = int(batas_row.get('min_total_awal') or 0)
            batas_max_awal = int(batas_row.get('max_total_awal') or 0)
            cur2.close()
        except Exception:
            batas_min_awal = 0
            batas_max_awal = 0
        return render_template('pembimbing/pengajuan anggaran awal.html', 
                             anggaran_data=anggaran_data, 
                             grouped_data=grouped_data,
                             anggaran_data_flat=anggaran_data_flat,
                             proposal_id=proposal_id,
                              proposal_info=proposal_info,
                              total_nilai_bantuan=total_nilai_bantuan,
                              batas_min_awal=batas_min_awal,
                              batas_max_awal=batas_max_awal)
    except Exception as e:
        flash(f'Error saat mengambil data anggaran: {str(e)}', 'danger')
        return render_template('pembimbing/pengajuan anggaran awal.html', anggaran_data=[], proposal_id=proposal_id, total_nilai_bantuan=0)

@pembimbing_bp.route('/anggaran_bertumbuh/<int:proposal_id>')
def pembimbing_anggaran_bertumbuh(proposal_id):
    if 'user_type' not in session or session['user_type'] != 'pembimbing':
        flash('Anda harus login sebagai pembimbing!', 'danger')
        return redirect(url_for('index'))
    app_funcs = get_app_functions()
    if not hasattr(app_funcs['mysql'], 'connection') or app_funcs['mysql'].connection is None:
        flash('Koneksi ke database gagal. Cek konfigurasi database!', 'danger')
        return render_template('pembimbing/pengajuan anggaran bertumbuh.html', anggaran_data=[], proposal_id=proposal_id)
    try:
        cursor = app_funcs['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Verify the proposal belongs to this pembimbing
        cursor.execute('''
            SELECT id FROM proposal 
            WHERE id = %s AND dosen_pembimbing = %s
        ''', (proposal_id, session['nama']))
        
        proposal = cursor.fetchone()
        if not proposal:
            cursor.close()
            flash('Proposal tidak ditemukan atau tidak ada akses!', 'danger')
            return redirect(url_for('pembimbing.pembimbing_proposal'))
        
        cursor.execute('''
            SELECT id, kegiatan_utama, kegiatan, penanggung_jawab, target_capaian, 
                   nama_barang, kuantitas, satuan, harga_satuan, jumlah, keterangan, status,
                   COALESCE(nilai_bantuan, 0) as nilai_bantuan
            FROM anggaran_bertumbuh 
            WHERE id_proposal = %s 
            ORDER BY kegiatan_utama, kegiatan, nama_barang
        ''', (proposal_id,))
        anggaran_data = cursor.fetchall()
        
        cursor.execute('SELECT judul_usaha, tahapan_usaha FROM proposal WHERE id = %s', (proposal_id,))
        proposal_info = cursor.fetchone()
        
        # Urutkan data sesuai urutan kegiatan utama (logika dinamis berdasarkan tahapan usaha)
        tahapan_usaha = proposal_info.get('tahapan_usaha', '').lower() if proposal_info else ''
        if 'bertumbuh' in tahapan_usaha:
            urutan_kegiatan = [
                "Pengembangan Pasar dan Saluran Distribusi",
                "Pengembangan Produk/Riset",
                "Produksi",
                "Legalitas, perizinan, sertifikasi, dan standarisasi",
                "Belanja ATK dan Penunjang"
            ]
        else:
            urutan_kegiatan = [
                "Pengembangan Produk/Riset",
                "Produksi",
                "Legalitas, perizinan, sertifikasi, dan standarisasi",
                "Belanja ATK dan Penunjang"
            ]
        
        urutan_kegiatan_lower = [k.strip().lower() for k in urutan_kegiatan]
        anggaran_data = sorted(
            anggaran_data,
            key=lambda x: urutan_kegiatan_lower.index(x['kegiatan_utama'].strip().lower()) if x['kegiatan_utama'] and x['kegiatan_utama'].strip().lower() in urutan_kegiatan_lower else 99
        )
        cursor.close()
        
        # Log aktivitas pembimbing
        pembimbing_info = app_funcs['get_pembimbing_info_from_session']()
        if pembimbing_info and proposal_info:
            app_funcs['log_pembimbing_activity'](
                pembimbing_info['id'],
                pembimbing_info['nip'],
                pembimbing_info['nama'],
                'view',
                'anggaran_bertumbuh',
                f'proposal_id_{proposal_id}',
                f'Melihat anggaran bertumbuh proposal "{proposal_info["judul_usaha"]}"',
                None,
                None,
                None,
                proposal_id
            )
        
        grouped_data = group_anggaran_data(anggaran_data)
        anggaran_data_flat = flatten_anggaran_data(anggaran_data)
        total_nilai_bantuan = sum((row.get('nilai_bantuan', 0) or 0) for row in anggaran_data)
        # Ambil batas min/maks untuk badge seperti mahasiswa
        try:
            cur2 = app_funcs['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
            cur2.execute('SELECT * FROM pengaturan_anggaran ORDER BY id DESC LIMIT 1')
            batas_row = cur2.fetchone() or {}
            batas_min_bertumbuh = int(batas_row.get('min_total_bertumbuh') or 0)
            batas_max_bertumbuh = int(batas_row.get('max_total_bertumbuh') or 0)
            cur2.close()
        except Exception:
            batas_min_bertumbuh = 0
            batas_max_bertumbuh = 0
        return render_template('pembimbing/pengajuan anggaran bertumbuh.html', 
                             anggaran_data=anggaran_data, 
                             grouped_data=grouped_data,
                             anggaran_data_flat=anggaran_data_flat,
                             proposal_id=proposal_id,
                              proposal_info=proposal_info,
                              total_nilai_bantuan=total_nilai_bantuan,
                              batas_min_bertumbuh=batas_min_bertumbuh,
                              batas_max_bertumbuh=batas_max_bertumbuh)
    except Exception as e:
        flash(f'Error saat mengambil data anggaran: {str(e)}', 'danger')
        return render_template('pembimbing/pengajuan anggaran bertumbuh.html', anggaran_data=[], proposal_id=proposal_id, total_nilai_bantuan=0)

@pembimbing_bp.route('/get_anggaran/<int:anggaran_id>')
def pembimbing_get_anggaran(anggaran_id):
    if 'user_type' not in session or session['user_type'] != 'pembimbing':
        return jsonify({'success': False, 'message': 'Anda harus login sebagai pembimbing!'})
    app_funcs = get_app_functions()
    if not hasattr(app_funcs['mysql'], 'connection') or app_funcs['mysql'].connection is None:
        return jsonify({'success': False, 'message': 'Koneksi ke database gagal!'})
    try:
        tabel = request.args.get('tabel', 'anggaran_awal')
        if tabel not in ['anggaran_awal', 'anggaran_bertumbuh']:
            return jsonify({'success': False, 'message': 'Tabel anggaran tidak valid!'})
        
        # Log aktivitas pembimbing
        pembimbing_info = app_funcs['get_pembimbing_info_from_session']()
        if pembimbing_info:
            app_funcs['log_pembimbing_activity'](
                pembimbing_info['id'],
                pembimbing_info['nip'],
                pembimbing_info['nama'],
                'view',
                tabel,
                f'{tabel}_detail_id_{anggaran_id}',
                f'Mengambil detail {tabel} ID {anggaran_id}'
            )
        cursor = app_funcs['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Verify the anggaran belongs to a proposal supervised by this pembimbing
        cursor.execute(f'''
            SELECT aa.* 
            FROM {tabel} aa
            JOIN proposal p ON aa.id_proposal = p.id
            WHERE aa.id = %s AND p.dosen_pembimbing = %s
        ''', (anggaran_id, session['nama']))
        
        anggaran = cursor.fetchone()
        cursor.close()
        
        if anggaran:
            return jsonify({'success': True, 'anggaran': anggaran})
        else:
            return jsonify({'success': False, 'message': 'Data anggaran tidak ditemukan atau tidak ada akses!'})
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error saat mengambil data: {str(e)}'})


@pembimbing_bp.route('/monitoring_mahasiswa/produksi')
def pembimbing_monitoring_mahasiswa_produksi():
    if 'user_type' not in session or session['user_type'] != 'pembimbing':
        flash('Anda harus login sebagai pembimbing!', 'danger')
        return redirect(url_for('index'))
    
    app_funcs = get_app_functions()
    
    # Log aktivitas pembimbing
    pembimbing_info = app_funcs['get_pembimbing_info_from_session']()
    if pembimbing_info:
        app_funcs['log_pembimbing_activity'](
            pembimbing_info['id'],
            pembimbing_info['nip'],
            pembimbing_info['nama'],
            'view',
            'monitoring',
            'produksi_mahasiswa',
            'Melihat halaman monitoring produksi mahasiswa'
        )
    if not hasattr(app_funcs['mysql'], 'connection') or app_funcs['mysql'].connection is None:
        flash('Koneksi ke database gagal. Cek konfigurasi database!', 'danger')
        return render_template('pembimbing/produksi_bahan_baku.html', mahasiswa_list=[])
    
    try:
        cursor = app_funcs['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Ambil semua mahasiswa yang dibimbing oleh pembimbing ini
        cursor.execute('''
            SELECT m.id, m.nama_ketua, m.nim, p.judul_usaha, p.tahapan_usaha, m.program_studi, p.id as proposal_id
            FROM mahasiswa m
            JOIN proposal p ON m.nim = p.nim
            WHERE p.dosen_pembimbing = %s
            ORDER BY m.nama_ketua
        ''', (session['nama'],))
        
        mahasiswa_all = cursor.fetchall()
        mahasiswa_list = []
        
        for mhs in mahasiswa_all:
            proposal_id = mhs['proposal_id']
            
            # Cek apakah ada data bahan baku
            cursor.execute('SELECT COUNT(*) as cnt FROM bahan_baku WHERE proposal_id = %s', (proposal_id,))
            bahan_baku_count = cursor.fetchone()['cnt']
            
            # Cek apakah ada data produksi
            cursor.execute('SELECT COUNT(*) as cnt FROM produksi WHERE proposal_id = %s', (proposal_id,))
            produksi_count = cursor.fetchone()['cnt']
            
            # Tambahkan informasi data yang tersedia
            mhs['has_bahan_baku'] = bahan_baku_count > 0
            mhs['has_produksi'] = produksi_count > 0
            
            mahasiswa_list.append(mhs)
        
        cursor.close()
        
        return render_template('pembimbing/produksi_bahan_baku.html', mahasiswa_list=mahasiswa_list)
        
    except Exception as e:
        flash(f'Error: {str(e)}', 'danger')
        return render_template('pembimbing/produksi_bahan_baku.html', mahasiswa_list=[])

@pembimbing_bp.route('/monitoring_mahasiswa/penjualan_produk')
def pembimbing_monitoring_mahasiswa_penjualan_produk():
    if 'user_type' not in session or session['user_type'] != 'pembimbing':
        flash('Anda harus login sebagai pembimbing!', 'danger')
        return redirect(url_for('index'))
    
    app_funcs = get_app_functions()
    
    # Log aktivitas pembimbing
    pembimbing_info = app_funcs['get_pembimbing_info_from_session']()
    if pembimbing_info:
        app_funcs['log_pembimbing_activity'](
            pembimbing_info['id'],
            pembimbing_info['nip'],
            pembimbing_info['nama'],
            'view',
            'monitoring',
            'penjualan_produk_mahasiswa',
            'Melihat halaman monitoring penjualan produk mahasiswa'
        )
    if not hasattr(app_funcs['mysql'], 'connection') or app_funcs['mysql'].connection is None:
        flash('Koneksi ke database gagal. Cek konfigurasi database!', 'danger')
        return render_template('pembimbing/penjualan_produk.html', mahasiswa_list=[])
    
    try:
        cursor = app_funcs['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Ambil semua mahasiswa yang dibimbing oleh pembimbing ini
        cursor.execute('''
            SELECT m.id, m.nama_ketua, m.nim, p.judul_usaha, p.tahapan_usaha, m.program_studi, p.id as proposal_id
            FROM mahasiswa m
            JOIN proposal p ON m.nim = p.nim
            WHERE p.dosen_pembimbing = %s
            ORDER BY m.nama_ketua
        ''', (session['nama'],))
        
        mahasiswa_all = cursor.fetchall()
        mahasiswa_list = []
        
        for mhs in mahasiswa_all:
            proposal_id = mhs['proposal_id']
            
            # Cek apakah ada data penjualan
            cursor.execute('SELECT COUNT(*) as cnt FROM penjualan WHERE proposal_id = %s', (proposal_id,))
            penjualan_count = cursor.fetchone()['cnt']
            
            # Tambahkan informasi data yang tersedia
            mhs['has_penjualan'] = penjualan_count > 0
            
            mahasiswa_list.append(mhs)
        
        cursor.close()
        return render_template('pembimbing/penjualan_produk.html', mahasiswa_list=mahasiswa_list)
    except Exception as e:
        flash(f'Error: {str(e)}', 'danger')
        return render_template('pembimbing/penjualan_produk.html', mahasiswa_list=[])

@pembimbing_bp.route('/monitoring_mahasiswa/laporan_laba_rugi')
def pembimbing_monitoring_mahasiswa_laporan_laba_rugi():
    if 'user_type' not in session or session['user_type'] != 'pembimbing':
        flash('Anda harus login sebagai pembimbing!', 'danger')
        return redirect(url_for('index'))
    
    app_funcs = get_app_functions()
    
    # Log aktivitas pembimbing
    pembimbing_info = app_funcs['get_pembimbing_info_from_session']()
    if pembimbing_info:
        app_funcs['log_pembimbing_activity'](
            pembimbing_info['id'],
            pembimbing_info['nip'],
            pembimbing_info['nama'],
            'view',
            'monitoring',
            'laporan_laba_rugi_mahasiswa',
            'Melihat halaman monitoring laporan laba rugi mahasiswa'
        )
    if not hasattr(app_funcs['mysql'], 'connection') or app_funcs['mysql'].connection is None:
        flash('Koneksi ke database gagal. Cek konfigurasi database!', 'danger')
        return render_template('pembimbing/laba_rugi_produk.html', mahasiswa_list=[])
    
    try:
        cursor = app_funcs['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Ambil semua mahasiswa yang dibimbing oleh pembimbing ini
        cursor.execute('''
            SELECT m.id, m.nama_ketua, m.nim, p.judul_usaha, p.tahapan_usaha, m.program_studi, p.id as proposal_id
            FROM mahasiswa m
            JOIN proposal p ON m.nim = p.nim
            WHERE p.dosen_pembimbing = %s
            ORDER BY m.nama_ketua
        ''', (session['nama'],))
        
        mahasiswa_all = cursor.fetchall()
        
        mahasiswa_list = []
        
        for mhs in mahasiswa_all:
            proposal_id = mhs['proposal_id']
            
            # Cek apakah ada data laba rugi
            cursor.execute('SELECT COUNT(*) as cnt FROM laba_rugi WHERE proposal_id = %s', (proposal_id,))
            laba_rugi_count = cursor.fetchone()['cnt']
            
            # Tambahkan informasi data yang tersedia
            mhs['has_laba_rugi'] = laba_rugi_count > 0
            
            mahasiswa_list.append(mhs)
        
        cursor.close()
        
        return render_template('pembimbing/laba_rugi_produk.html', mahasiswa_list=mahasiswa_list)
        
    except Exception as e:
        flash(f'Error: {str(e)}', 'danger')
        return render_template('pembimbing/laba_rugi_produk.html', mahasiswa_list=[])

@pembimbing_bp.route('/laporan_laba_rugi/<int:mahasiswa_id>')
def pembimbing_laporan_laba_rugi(mahasiswa_id):
    if 'user_type' not in session or session['user_type'] != 'pembimbing':
        flash('Anda harus login sebagai pembimbing!', 'danger')
        return redirect(url_for('index'))
    
    app_funcs = get_app_functions()
    if not hasattr(app_funcs['mysql'], 'connection') or app_funcs['mysql'].connection is None:
        flash('Koneksi ke database gagal. Cek konfigurasi database!', 'danger')
        return render_template('pembimbing/laporan_laba_rugi.html', proposal_data={})
    
    try:
        cursor = app_funcs['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Ambil data mahasiswa yang dibimbing oleh pembimbing ini
        cursor.execute('''
            SELECT m.*, p.judul_usaha, p.kategori, p.tahun_nib, p.id as proposal_id, m.nama_ketua, m.nim
            FROM mahasiswa m
            JOIN proposal p ON m.nim = p.nim
            WHERE m.id = %s AND p.dosen_pembimbing = %s
        ''', (mahasiswa_id, session['nama']))
        
        mahasiswa_info = cursor.fetchone()
        
        if not mahasiswa_info:
            flash('Data mahasiswa tidak ditemukan atau Anda tidak memiliki akses!', 'danger')
            cursor.close()
            return redirect(url_for('pembimbing.pembimbing_monitoring_mahasiswa_laporan_laba_rugi'))
        
        proposal_id = mahasiswa_info['proposal_id']
        
        # Ambil data laba rugi dari tabel laba_rugi
        cursor.execute('''
            SELECT tanggal_produksi, nama_produk, pendapatan, total_biaya_produksi, 
                   laba_rugi_kotor, biaya_operasional, laba_rugi_bersih
            FROM laba_rugi
            WHERE proposal_id = %s
            ORDER BY tanggal_produksi DESC
        ''', (proposal_id,))
        
        laba_rugi_data = cursor.fetchall()
        
        # Hitung total dari data laba rugi
        total_pendapatan = sum(item['pendapatan'] for item in laba_rugi_data)
        total_biaya_produksi = sum(item['total_biaya_produksi'] for item in laba_rugi_data)
        total_biaya_operasional = sum(item['biaya_operasional'] for item in laba_rugi_data)
        total_laba_kotor = sum(item['laba_rugi_kotor'] for item in laba_rugi_data)
        total_laba_bersih = sum(item['laba_rugi_bersih'] for item in laba_rugi_data)
        
        proposal_data = {
            proposal_id: {
                'proposal': {
                    'id': proposal_id,
                    'judul_usaha': mahasiswa_info['judul_usaha'],
                    'kategori': mahasiswa_info['kategori'],
                    'tahun_nib': mahasiswa_info['tahun_nib'],
                    'nama_ketua': mahasiswa_info['nama_ketua'],
                    'nim': mahasiswa_info['nim']
                },
                'laba_rugi_data': laba_rugi_data,
                'total_pendapatan': total_pendapatan,
                'total_biaya_produksi': total_biaya_produksi,
                'total_biaya_operasional': total_biaya_operasional,
                'total_laba_kotor': total_laba_kotor,
                'total_laba_bersih': total_laba_bersih
            }
        }
        
        cursor.close()
        return render_template('pembimbing/laporan_laba_rugi.html', proposal_data=proposal_data)
        
    except Exception as e:
        flash(f'Error: {str(e)}', 'danger')
        return render_template('pembimbing/laporan_laba_rugi.html', proposal_data={})

@pembimbing_bp.route('/download_laba_rugi', methods=['POST'])
def pembimbing_download_laba_rugi():
    logger.info("Download laba rugi pembimbing dipanggil")
    if 'user_type' not in session or session['user_type'] != 'pembimbing':
        logger.warning(f"Akses ditolak untuk download laba rugi: {session.get('user_type', 'None')}")
        return jsonify({'success': False, 'message': 'Unauthorized'})
    
    try:
        proposal_id = request.form.get('proposal_id')
        format_type = request.form.get('format')
        start_date = request.form.get('start_date')
        end_date = request.form.get('end_date')
        
        if not all([proposal_id, format_type, start_date, end_date]):
            return jsonify({'success': False, 'message': 'Parameter tidak lengkap'})
        
        app_funcs = get_app_functions()
        if not hasattr(app_funcs['mysql'], 'connection') or app_funcs['mysql'].connection is None:
            return jsonify({'success': False, 'message': 'Koneksi database tidak tersedia'})
        
        cursor = app_funcs['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Ambil data proposal dan pastikan pembimbing memiliki akses
        cursor.execute('''
            SELECT p.judul_usaha, p.kategori, p.tahun_nib
            FROM proposal p
            WHERE p.id = %s AND p.dosen_pembimbing = %s
        ''', (proposal_id, session['nama']))
        
        proposal = cursor.fetchone()
        if not proposal:
            cursor.close()
            return jsonify({'success': False, 'message': 'Proposal tidak ditemukan atau tidak ada akses'})
        
        # Ambil data laba rugi sesuai rentang tanggal
        cursor.execute('''
            SELECT tanggal_produksi, nama_produk, pendapatan, total_biaya_produksi, 
                   laba_rugi_kotor, biaya_operasional, laba_rugi_bersih
            FROM laba_rugi
            WHERE proposal_id = %s AND tanggal_produksi BETWEEN %s AND %s
            ORDER BY tanggal_produksi DESC
        ''', (proposal_id, start_date, end_date))
        
        laba_rugi_data = cursor.fetchall()
        
        # DEBUG: Log data untuk memastikan tidak kosong
        logger.debug(f"=== DEBUG DOWNLOAD LABA RUGI PEMBIMBING ===")
        logger.debug(f"Proposal ID: {proposal_id}")
        logger.debug(f"Start Date: {start_date}")
        logger.debug(f"End Date: {end_date}")
        logger.debug(f"Data count: {len(laba_rugi_data)}")
        logger.debug(f"Data sample: {laba_rugi_data[:2] if laba_rugi_data else 'KOSONG'}")
        logger.debug(f"Proposal data: {proposal}")
        logger.debug("=============================================")
        
        if not laba_rugi_data:
            cursor.close()
            return jsonify({'success': False, 'message': 'Tidak ada data laba rugi untuk periode yang dipilih'})
        
        # Hitung total
        total_pendapatan = sum(item['pendapatan'] for item in laba_rugi_data)
        total_biaya_produksi = sum(item['total_biaya_produksi'] for item in laba_rugi_data)
        total_biaya_operasional = sum(item['biaya_operasional'] for item in laba_rugi_data)
        total_laba_kotor = sum(item['laba_rugi_kotor'] for item in laba_rugi_data)
        total_laba_bersih = sum(item['laba_rugi_bersih'] for item in laba_rugi_data)
        
        # DEBUG: Log totals
        logger.debug(f"=== TOTALS PEMBIMBING ===")
        logger.debug(f"Total Pendapatan: {total_pendapatan}")
        logger.debug(f"Total Biaya Produksi: {total_biaya_produksi}")
        logger.debug(f"Total Biaya Operasional: {total_biaya_operasional}")
        logger.debug(f"Total Laba Kotor: {total_laba_kotor}")
        logger.debug(f"Total Laba Bersih: {total_laba_bersih}")
        logger.debug("==========================")
        
        cursor.close()
        
        # Generate file berdasarkan format dengan pembersihan karakter
        import re
        safe_judul = re.sub(r'[<>:"/\\|?*]', '_', proposal['judul_usaha'])
        filename = f"Laporan_Laba_Rugi_{safe_judul.replace(' ', '_')}_{start_date}_to_{end_date}"
        
        if format_type == 'excel':
            return app_funcs['generate_excel_laba_rugi'](laba_rugi_data, proposal, total_pendapatan, 
                                           total_biaya_produksi, total_biaya_operasional, 
                                           total_laba_kotor, total_laba_bersih, 
                                           start_date, end_date, filename)
        elif format_type == 'pdf':
            return app_funcs['generate_pdf_laba_rugi'](laba_rugi_data, proposal, total_pendapatan, 
                                         total_biaya_produksi, total_biaya_operasional, 
                                         total_laba_kotor, total_laba_bersih, 
                                         start_date, end_date, filename)
        elif format_type == 'word':
            return app_funcs['generate_word_laba_rugi'](laba_rugi_data, proposal, total_pendapatan, 
                                          total_biaya_produksi, total_biaya_operasional, 
                                          total_laba_kotor, total_laba_bersih, 
                                          start_date, end_date, filename)
        else:
            return jsonify({'success': False, 'message': 'Format tidak didukung'})
        
    except Exception as e:
        logger.error('--- ERROR DOWNLOAD LABA RUGI PEMBIMBING ---')
        logger.error(traceback.format_exc())
        return jsonify({'success': False, 'message': f'Error: {str(e)}', 'trace': traceback.format_exc()})

@pembimbing_bp.route('/monitoring_mahasiswa/alat_produksi')
def pembimbing_monitoring_mahasiswa_alat_produksi():
    logger.info("Monitoring alat produksi mahasiswa dipanggil")
    if 'user_type' not in session or session['user_type'] != 'pembimbing':
        logger.warning(f"Akses ditolak untuk monitoring alat produksi: {session.get('user_type', 'None')}")
        flash('Anda harus login sebagai pembimbing!', 'danger')
        return redirect(url_for('index'))
    
    app_funcs = get_app_functions()
    # Log aktivitas pembimbing
    pembimbing_info = app_funcs['get_pembimbing_info_from_session']()
    if pembimbing_info:
        app_funcs['log_pembimbing_activity'](
            pembimbing_info['id'],
            pembimbing_info['nip'],
            pembimbing_info['nama'],
            'view',
            'monitoring',
            'alat_produksi_mahasiswa',
            'Melihat halaman monitoring alat produksi mahasiswa'
        )
    
    if not hasattr(app_funcs['mysql'], 'connection') or app_funcs['mysql'].connection is None:
        logger.error("Koneksi database tidak tersedia untuk monitoring alat produksi")
        flash('Koneksi ke database gagal. Cek konfigurasi database!', 'danger')
        return render_template('pembimbing/daftar_mahsiswa_alat_produksi.html', mahasiswa_list=[])
    
    try:
        cursor = app_funcs['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Ambil semua mahasiswa yang dibimbing oleh pembimbing ini
        cursor.execute('''
            SELECT m.id, m.nama_ketua, m.nim, p.judul_usaha, p.tahapan_usaha, m.program_studi, p.id as proposal_id
            FROM mahasiswa m
            JOIN proposal p ON m.nim = p.nim
            WHERE p.dosen_pembimbing = %s
            ORDER BY m.nama_ketua
        ''', (session['nama'],))
        
        mahasiswa_all = cursor.fetchall()
        mahasiswa_list = []
        
        for mhs in mahasiswa_all:
            proposal_id = mhs['proposal_id']
            
            # Cek apakah ada data alat produksi
            cursor.execute('SELECT COUNT(*) as cnt FROM alat_produksi WHERE proposal_id = %s', (proposal_id,))
            alat_produksi_count = cursor.fetchone()['cnt']
            
            # Tambahkan informasi data yang tersedia
            mhs['has_alat_produksi'] = alat_produksi_count > 0
            
            mahasiswa_list.append(mhs)
        
        cursor.close()
        
        return render_template('pembimbing/daftar_mahsiswa_alat_produksi.html', mahasiswa_list=mahasiswa_list)
        
    except Exception as e:
        flash(f'Error: {str(e)}', 'danger')
        return render_template('pembimbing/daftar_mahsiswa_alat_produksi.html', mahasiswa_list=[])

@pembimbing_bp.route('/alat_produksi/<int:mahasiswa_id>')
def pembimbing_alat_produksi(mahasiswa_id):
    if 'user_type' not in session or session['user_type'] != 'pembimbing':
        flash('Anda harus login sebagai pembimbing!', 'danger')
        return redirect(url_for('index'))
    
    app_funcs = get_app_functions()
    if not hasattr(app_funcs['mysql'], 'connection') or app_funcs['mysql'].connection is None:
        flash('Koneksi ke database gagal. Cek konfigurasi database!', 'danger')
        return render_template('pembimbing/alat_produksi.html', alat_produksi_list=[], mahasiswa_info={})
    
    try:
        cursor = app_funcs['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Ambil data mahasiswa yang dibimbing oleh pembimbing ini
        cursor.execute('''
            SELECT m.*, p.judul_usaha, p.id as proposal_id
            FROM mahasiswa m
            JOIN proposal p ON m.nim = p.nim
            WHERE m.id = %s AND p.dosen_pembimbing = %s
        ''', (mahasiswa_id, session['nama']))
        
        mahasiswa_info = cursor.fetchone()
        
        if not mahasiswa_info:
            flash('Data mahasiswa tidak ditemukan atau Anda tidak memiliki akses!', 'danger')
            cursor.close()
            return redirect(url_for('pembimbing.pembimbing_monitoring_mahasiswa_alat_produksi'))
        
        proposal_id = mahasiswa_info['proposal_id']
        
        # Ambil data alat produksi
        cursor.execute('''
            SELECT * FROM alat_produksi 
            WHERE proposal_id = %s 
            ORDER BY tanggal_beli DESC
        ''', (proposal_id,))
        
        alat_produksi_list = cursor.fetchall()
        cursor.close()
        
        return render_template('pembimbing/alat_produksi.html', 
                             alat_produksi_list=alat_produksi_list, 
                             mahasiswa_info=mahasiswa_info)
        
    except Exception as e:
        flash(f'Error: {str(e)}', 'danger')
        return render_template('pembimbing/alat_produksi.html', alat_produksi_list=[], mahasiswa_info={})

@pembimbing_bp.route('/monitoring_mahasiswa/biaya_operasional')
def pembimbing_monitoring_mahasiswa_biaya_operasional():
    if 'user_type' not in session or session['user_type'] != 'pembimbing':
        flash('Anda harus login sebagai pembimbing!', 'danger')
        return redirect(url_for('index'))
    
    app_funcs = get_app_functions()
    
    # Log aktivitas pembimbing
    pembimbing_info = app_funcs['get_pembimbing_info_from_session']()
    if pembimbing_info:
        app_funcs['log_pembimbing_activity'](
            pembimbing_info['id'],
            pembimbing_info['nip'],
            pembimbing_info['nama'],
            'view',
            'monitoring',
            'biaya_operasional_mahasiswa',
            'Melihat halaman monitoring biaya operasional mahasiswa'
        )
    if not hasattr(app_funcs['mysql'], 'connection') or app_funcs['mysql'].connection is None:
        flash('Koneksi ke database gagal. Cek konfigurasi database!', 'danger')
        return render_template('pembimbing/daftar_mahsiswa_biaya_operasional.html', mahasiswa_list=[])
    
    try:
        cursor = app_funcs['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Ambil semua mahasiswa yang dibimbing oleh pembimbing ini
        cursor.execute('''
            SELECT m.id, m.nama_ketua, m.nim, p.judul_usaha, p.tahapan_usaha, m.program_studi, p.id as proposal_id
            FROM mahasiswa m
            JOIN proposal p ON m.nim = p.nim
            WHERE p.dosen_pembimbing = %s
            ORDER BY m.nama_ketua
        ''', (session['nama'],))
        
        mahasiswa_all = cursor.fetchall()
        mahasiswa_list = []
        
        for mhs in mahasiswa_all:
            proposal_id = mhs['proposal_id']
            
            # Cek apakah ada data biaya operasional
            cursor.execute('SELECT COUNT(*) as cnt FROM biaya_operasional WHERE proposal_id = %s', (proposal_id,))
            biaya_operasional_count = cursor.fetchone()['cnt']
            
            # Tambahkan informasi data yang tersedia
            mhs['has_biaya_operasional'] = biaya_operasional_count > 0
            
            mahasiswa_list.append(mhs)
        
        cursor.close()
        
        return render_template('pembimbing/daftar_mahsiswa_biaya_operasional.html', mahasiswa_list=mahasiswa_list)
        
    except Exception as e:
        flash(f'Error: {str(e)}', 'danger')
        return render_template('pembimbing/daftar_mahsiswa_biaya_operasional.html', mahasiswa_list=[])

@pembimbing_bp.route('/monitoring_mahasiswa/biaya_non_operasional')
def pembimbing_monitoring_mahasiswa_biaya_non_operasional():
    if 'user_type' not in session or session['user_type'] != 'pembimbing':
        flash('Anda harus login sebagai pembimbing!', 'danger')
        return redirect(url_for('index'))
    
    app_funcs = get_app_functions()
    
    # Log aktivitas pembimbing
    pembimbing_info = app_funcs['get_pembimbing_info_from_session']()
    if pembimbing_info:
        app_funcs['log_pembimbing_activity'](
            pembimbing_info['id'],
            pembimbing_info['nip'],
            pembimbing_info['nama'],
            'view',
            'monitoring',
            'biaya_non_operasional_mahasiswa',
            'Melihat halaman monitoring Biaya Lain-lain mahasiswa'
        )
    if not hasattr(app_funcs['mysql'], 'connection') or app_funcs['mysql'].connection is None:
        flash('Koneksi ke database gagal. Cek konfigurasi database!', 'danger')
        return render_template('pembimbing/daftar_biaya_non_operasional.html', mahasiswa_list=[])
    
    try:
        cursor = app_funcs['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Ambil semua mahasiswa yang dibimbing oleh pembimbing ini
        cursor.execute('''
            SELECT m.id, m.nama_ketua, m.nim, p.judul_usaha, p.tahapan_usaha, m.program_studi, p.id as proposal_id
            FROM mahasiswa m
            JOIN proposal p ON m.nim = p.nim
            WHERE p.dosen_pembimbing = %s
            ORDER BY m.nama_ketua
        ''', (session['nama'],))
        
        mahasiswa_all = cursor.fetchall()
        mahasiswa_list = []
        
        for mhs in mahasiswa_all:
            proposal_id = mhs['proposal_id']
            
            # Cek apakah ada data Biaya Lain-lain
            cursor.execute('SELECT COUNT(*) as cnt FROM biaya_non_operasional WHERE proposal_id = %s', (proposal_id,))
            biaya_non_operasional_count = cursor.fetchone()['cnt']
            
            # Tambahkan informasi data yang tersedia
            mhs['has_biaya_non_operasional'] = biaya_non_operasional_count > 0
            
            mahasiswa_list.append(mhs)
        
        cursor.close()
        
        return render_template('pembimbing/daftar_biaya_non_operasional.html', mahasiswa_list=mahasiswa_list)
        
    except Exception as e:
        flash(f'Error: {str(e)}', 'danger')
        return render_template('pembimbing/daftar_biaya_non_operasional.html', mahasiswa_list=[])

@pembimbing_bp.route('/biaya_non_operasional/<int:mahasiswa_id>')
def pembimbing_biaya_non_operasional(mahasiswa_id):
    if 'user_type' not in session or session['user_type'] != 'pembimbing':
        flash('Anda harus login sebagai pembimbing!', 'danger')
        return redirect(url_for('index'))
    
    app_funcs = get_app_functions()
    if not hasattr(app_funcs['mysql'], 'connection') or app_funcs['mysql'].connection is None:
        flash('Koneksi ke database gagal. Cek konfigurasi database!', 'danger')
        return render_template('pembimbing/biaya_non_operasional.html', biaya_non_operasional_list=[], mahasiswa_info={})
    
    try:
        cursor = app_funcs['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Ambil data mahasiswa yang dibimbing oleh pembimbing ini
        cursor.execute('''
            SELECT m.*, p.judul_usaha, p.id as proposal_id
            FROM mahasiswa m
            JOIN proposal p ON m.nim = p.nim
            WHERE m.id = %s AND p.dosen_pembimbing = %s
        ''', (mahasiswa_id, session['nama']))
        
        mahasiswa_info = cursor.fetchone()
        
        if not mahasiswa_info:
            flash('Data mahasiswa tidak ditemukan atau Anda tidak memiliki akses!', 'danger')
            cursor.close()
            return redirect(url_for('pembimbing.pembimbing_monitoring_mahasiswa_biaya_non_operasional'))
        
        proposal_id = mahasiswa_info['proposal_id']
        
        # Ambil data Biaya Lain-lain
        cursor.execute('''
            SELECT * FROM biaya_non_operasional 
            WHERE proposal_id = %s 
            ORDER BY created_at DESC
        ''', (proposal_id,))
        
        biaya_non_operasional_list = cursor.fetchall()
        cursor.close()
        
        return render_template('pembimbing/biaya_non_operasional.html', 
                             biaya_non_operasional_list=biaya_non_operasional_list, 
                             mahasiswa_info=mahasiswa_info)
        
    except Exception as e:
        flash(f'Error: {str(e)}', 'danger')
        return render_template('pembimbing/biaya_non_operasional.html', biaya_non_operasional_list=[], mahasiswa_info={})

@pembimbing_bp.route('/biaya_operasional/<int:mahasiswa_id>')
def pembimbing_biaya_operasional(mahasiswa_id):
    if 'user_type' not in session or session['user_type'] != 'pembimbing':
        flash('Anda harus login sebagai pembimbing!', 'danger')
        return redirect(url_for('index'))
    
    app_funcs = get_app_functions()
    if not hasattr(app_funcs['mysql'], 'connection') or app_funcs['mysql'].connection is None:
        flash('Koneksi ke database gagal. Cek konfigurasi database!', 'danger')
        return render_template('pembimbing/biaya_operasional.html', biaya_operasional_list=[], mahasiswa_info={})
    
    try:
        cursor = app_funcs['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Ambil data mahasiswa yang dibimbing oleh pembimbing ini
        cursor.execute('''
            SELECT m.*, p.judul_usaha, p.id as proposal_id
            FROM mahasiswa m
            JOIN proposal p ON m.nim = p.nim
            WHERE m.id = %s AND p.dosen_pembimbing = %s
        ''', (mahasiswa_id, session['nama']))
        
        mahasiswa_info = cursor.fetchone()
        
        if not mahasiswa_info:
            flash('Data mahasiswa tidak ditemukan atau Anda tidak memiliki akses!', 'danger')
            cursor.close()
            return redirect(url_for('pembimbing.pembimbing_monitoring_mahasiswa_biaya_operasional'))
        
        proposal_id = mahasiswa_info['proposal_id']
        
        # Ambil data biaya operasional
        cursor.execute('''
            SELECT * FROM biaya_operasional 
            WHERE proposal_id = %s 
            ORDER BY created_at DESC
        ''', (proposal_id,))
        
        biaya_operasional_list = cursor.fetchall()
        cursor.close()
        
        return render_template('pembimbing/biaya_operasional.html', 
                             biaya_operasional_list=biaya_operasional_list, 
                             mahasiswa_info=mahasiswa_info)
        
    except Exception as e:
        flash(f'Error: {str(e)}', 'danger')
        return render_template('pembimbing/biaya_operasional.html', biaya_operasional_list=[], mahasiswa_info={})

@pembimbing_bp.route('/monitoring_mahasiswa/laporan_arus_kas')
def pembimbing_monitoring_mahasiswa_laporan_arus_kas():
    if 'user_type' not in session or session['user_type'] != 'pembimbing':
        flash('Anda harus login sebagai pembimbing!', 'danger')
        return redirect(url_for('index'))
    
    app_funcs = get_app_functions()
    if not hasattr(app_funcs['mysql'], 'connection') or app_funcs['mysql'].connection is None:
        flash('Koneksi ke database gagal. Cek konfigurasi database!', 'danger')
        return render_template('pembimbing/daftar_mahasiswa_arus_kas.html', mahasiswa_list=[])
    
    try:
        cursor = app_funcs['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Ambil semua mahasiswa yang dibimbing oleh pembimbing ini dengan status proposal
        cursor.execute('''
            SELECT m.id, m.nama_ketua, m.nim, p.judul_usaha, p.tahapan_usaha, m.program_studi, 
                   p.id as proposal_id, p.status_admin, p.status as status_proposal
            FROM mahasiswa m
            JOIN proposal p ON m.nim = p.nim
            WHERE p.dosen_pembimbing = %s
            ORDER BY m.nama_ketua
        ''', (session['nama'],))
        
        mahasiswa_all = cursor.fetchall()
        mahasiswa_list = []
        
        for mhs in mahasiswa_all:
            proposal_id = mhs['proposal_id']
            
            # Cek apakah proposal sudah lolos
            has_lolos_proposal = mhs['status_admin'] == 'lolos'
            
            if has_lolos_proposal:
                # Cek apakah ada data arus kas dari tabel arus_kas
                cursor.execute('SELECT COUNT(*) as cnt FROM arus_kas WHERE proposal_id = %s', (proposal_id,))
                arus_kas_count = cursor.fetchone()['cnt']
                
                # Jika tidak ada di tabel arus_kas, cek dari tabel individual
                if arus_kas_count == 0:
                    cursor.execute('SELECT COUNT(*) as cnt FROM penjualan WHERE proposal_id = %s', (proposal_id,))
                    penjualan_count = cursor.fetchone()['cnt']
                    
                    cursor.execute('SELECT COUNT(*) as cnt FROM biaya_operasional WHERE proposal_id = %s', (proposal_id,))
                    biaya_operasional_count = cursor.fetchone()['cnt']
                    
                    cursor.execute('SELECT COUNT(*) as cnt FROM biaya_non_operasional WHERE proposal_id = %s', (proposal_id,))
                    biaya_non_operasional_count = cursor.fetchone()['cnt']
                    
                    cursor.execute('SELECT COUNT(*) as cnt FROM alat_produksi WHERE proposal_id = %s', (proposal_id,))
                    alat_produksi_count = cursor.fetchone()['cnt']
                    
                    # Tambahkan informasi data yang tersedia
                    mhs['has_arus_kas'] = (penjualan_count > 0 or biaya_operasional_count > 0 or 
                                          biaya_non_operasional_count > 0 or alat_produksi_count > 0)
                else:
                    mhs['has_arus_kas'] = True
            else:
                mhs['has_arus_kas'] = False
            
            mahasiswa_list.append(mhs)
        
        cursor.close()
        
        return render_template('pembimbing/daftar_mahasiswa_arus_kas.html', mahasiswa_list=mahasiswa_list)
        
    except Exception as e:
        logger.error(f"Error in pembimbing_monitoring_mahasiswa_laporan_arus_kas: {str(e)}")
        flash(f'Error: {str(e)}', 'danger')
        return render_template('pembimbing/daftar_mahasiswa_arus_kas.html', mahasiswa_list=[])

@pembimbing_bp.route('/arus_kas/<int:mahasiswa_id>')
def pembimbing_arus_kas(mahasiswa_id):
    if 'user_type' not in session or session['user_type'] != 'pembimbing':
        flash('Anda harus login sebagai pembimbing!', 'danger')
        return redirect(url_for('index'))
    
    app_funcs = get_app_functions()
    if not hasattr(app_funcs['mysql'], 'connection') or app_funcs['mysql'].connection is None:
        flash('Koneksi ke database gagal. Cek konfigurasi database!', 'danger')
        return render_template('pembimbing/arus_kas.html', mahasiswa_info=None, proposal_data={})
    
    try:
        cursor = app_funcs['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Ambil data mahasiswa yang dibimbing oleh pembimbing ini
        cursor.execute('''
            SELECT m.*, p.judul_usaha, p.id as proposal_id, p.status as status_proposal, p.status_admin
            FROM mahasiswa m
            JOIN proposal p ON m.nim = p.nim
            WHERE m.id = %s AND p.dosen_pembimbing = %s
        ''', (mahasiswa_id, session['nama']))
        
        mahasiswa_info = cursor.fetchone()
        if not mahasiswa_info:
            flash('Data mahasiswa tidak ditemukan atau Anda tidak memiliki akses!', 'danger')
            cursor.close()
            return redirect(url_for('pembimbing.pembimbing_monitoring_mahasiswa_laporan_arus_kas'))
        
        # Cek apakah proposal sudah lolos
        has_lolos_proposal = mahasiswa_info['status_admin'] == 'lolos'
        
        if not has_lolos_proposal:
            flash('Proposal mahasiswa belum disetujui oleh admin!', 'warning')
            cursor.close()
            return render_template('pembimbing/arus_kas.html', 
                                 mahasiswa_info=mahasiswa_info,
                                 proposal_data={},
                                 has_lolos_proposal=False)
        
        # Ambil data arus kas dari tabel arus_kas
        proposal_id = mahasiswa_info['proposal_id']
        logger.debug(f"DEBUG - proposal_id dari mahasiswa_info: {proposal_id}")
        logger.debug(f"DEBUG - mahasiswa_info: {mahasiswa_info}")
        
        # Ambil bulan yang dipilih (default: bulan ini)
        selected_month = request.args.get('month', datetime.now().strftime('%Y-%m'))
        
        # Parse bulan dan tahun
        if not selected_month:
            selected_month = datetime.now().strftime('%Y-%m')
        
        try:
            selected_date = datetime.strptime(selected_month, '%Y-%m')
            selected_month_year = selected_date.strftime('%B %Y')
        except ValueError:
            selected_date = datetime.now()
            selected_month_year = selected_date.strftime('%B %Y')
        
        # Ambil data arus kas dari tabel arus_kas
        cursor.execute('''
            SELECT total_penjualan, total_biaya_produksi, total_biaya_operasional,
                   total_biaya_non_operasional, kas_bersih_operasional, total_harga_jual_alat,
                   total_harga_alat, kas_bersih_investasi, kas_bersih_pembiayaan, total_kas_bersih
            FROM arus_kas 
            WHERE proposal_id = %s AND bulan_tahun = %s
        ''', (proposal_id, selected_month))
        
        arus_kas_data = cursor.fetchone()
        
        if arus_kas_data:
            # Jika data ada di tabel arus_kas, gunakan data tersebut
            total_penjualan = float(arus_kas_data['total_penjualan'] or 0)
            total_biaya_produksi = float(arus_kas_data['total_biaya_produksi'] or 0)
            total_biaya_operasional = float(arus_kas_data['total_biaya_operasional'] or 0)
            total_biaya_non_operasional = float(arus_kas_data['total_biaya_non_operasional'] or 0)
            kas_bersih_operasional = float(arus_kas_data['kas_bersih_operasional'] or 0)
            total_harga_jual_alat = float(arus_kas_data['total_harga_jual_alat'] or 0)
            total_harga_alat = float(arus_kas_data['total_harga_alat'] or 0)
            kas_bersih_investasi = float(arus_kas_data['kas_bersih_investasi'] or 0)
            kas_bersih_pembiayaan = float(arus_kas_data['kas_bersih_pembiayaan'] or 0)
            total_kas_bersih = float(arus_kas_data['total_kas_bersih'] or 0)
        else:
            # Jika data tidak ada di tabel arus_kas, hitung dari tabel individual
            cursor.execute('''
                SELECT COALESCE(SUM(total), 0) as total_penjualan
                FROM penjualan 
                WHERE proposal_id = %s 
                AND DATE_FORMAT(tanggal_penjualan, '%%Y-%%m') = %s
            ''', (proposal_id, selected_month))
            total_penjualan = float(cursor.fetchone()['total_penjualan'] or 0)
            
            cursor.execute('''
                SELECT COALESCE(SUM(total_harga), 0) as total_biaya_produksi
                FROM bahan_baku 
                WHERE proposal_id = %s 
                AND DATE_FORMAT(tanggal_beli, '%%Y-%%m') = %s
            ''', (proposal_id, selected_month))
            total_biaya_produksi = float(cursor.fetchone()['total_biaya_produksi'] or 0)
            
            cursor.execute('''
                SELECT COALESCE(SUM(total_harga), 0) as total_biaya_operasional
                FROM biaya_operasional 
                WHERE proposal_id = %s 
                AND DATE_FORMAT(tanggal_beli, '%%Y-%%m') = %s
            ''', (proposal_id, selected_month))
            total_biaya_operasional = float(cursor.fetchone()['total_biaya_operasional'] or 0)
            
            cursor.execute('''
                SELECT COALESCE(SUM(total_harga), 0) as total_biaya_non_operasional
                FROM biaya_non_operasional 
                WHERE proposal_id = %s 
                AND DATE_FORMAT(tanggal_transaksi, '%%Y-%%m') = %s
            ''', (proposal_id, selected_month))
            total_biaya_non_operasional = float(cursor.fetchone()['total_biaya_non_operasional'] or 0)
            
            cursor.execute('''
                SELECT 
                    COALESCE(SUM(harga), 0) as total_harga_alat,
                    COALESCE(SUM(harga_jual), 0) as total_harga_jual_alat
                FROM alat_produksi 
                WHERE proposal_id = %s 
                AND DATE_FORMAT(tanggal_beli, '%%Y-%%m') = %s
            ''', (proposal_id, selected_month))
            
            alat_data = cursor.fetchone()
            total_harga_alat = float(alat_data['total_harga_alat'] or 0)
            total_harga_jual_alat = float(alat_data['total_harga_jual_alat'] or 0)
            
            # Hitung kas bersih dari kegiatan operasional
            kas_bersih_operasional = total_penjualan - total_biaya_produksi - total_biaya_operasional - total_biaya_non_operasional
            
            # Hitung kas bersih dari kegiatan investasi
            kas_bersih_investasi = total_harga_jual_alat - total_harga_alat
            
            # Hitung kas bersih dari kegiatan pembiayaan (selalu 0)
            kas_bersih_pembiayaan = 0
            
            # Hitung total kas bersih
            total_kas_bersih = kas_bersih_operasional + kas_bersih_investasi
        
        # Siapkan data untuk template
        proposal_data = {
            proposal_id: {
                'proposal': mahasiswa_info,
                'total_penjualan': total_penjualan,
                'total_biaya_produksi': total_biaya_produksi,
                'total_biaya_operasional': total_biaya_operasional,
                'total_biaya_non_operasional': total_biaya_non_operasional,
                'total_harga_alat': total_harga_alat,
                'total_harga_jual_alat': total_harga_jual_alat,
                'kas_bersih_operasional': kas_bersih_operasional,
                'kas_bersih_investasi': kas_bersih_investasi,
                'kas_bersih_pembiayaan': kas_bersih_pembiayaan,
                'total_kas_bersih': total_kas_bersih
            }
        }
        
        # Ambil daftar bulan yang tersedia dari tabel arus_kas
        cursor.execute('''
            SELECT DISTINCT bulan_tahun as month_value
            FROM arus_kas 
            WHERE proposal_id = %s
            ORDER BY bulan_tahun DESC
        ''', (proposal_id,))
        
        available_months = []
        for row in cursor.fetchall():
            month_value = row['month_value']
            try:
                month_date = datetime.strptime(month_value, '%Y-%m')
                month_label = month_date.strftime('%B %Y')
                available_months.append({
                    'value': month_value,
                    'label': month_label
                })
            except ValueError:
                continue
        
        # Jika tidak ada bulan di tabel arus_kas, tambahkan bulan saat ini
        if not available_months:
            current_month = datetime.now().strftime('%Y-%m')
            current_month_label = datetime.now().strftime('%B %Y')
            available_months.append({
                'value': current_month,
                'label': current_month_label
            })
        
        cursor.close()
        
        logger.debug(f"DEBUG - proposal_data yang dikirim ke template: {proposal_data}")
        
        return render_template('pembimbing/arus_kas.html', 
                             mahasiswa_info=mahasiswa_info,
                             proposal_data=proposal_data,
                             proposal_id=proposal_id,
                             selected_month=selected_month,
                             selected_month_year=selected_month_year,
                             available_months=available_months,
                             has_lolos_proposal=True)
        
    except Exception as e:
        logger.error(f"Error in pembimbing_arus_kas: {str(e)}")
        flash(f'Error: {str(e)}', 'danger')
        return render_template('pembimbing/arus_kas.html', 
                             mahasiswa_info=None, 
                             proposal_data={},
                             has_lolos_proposal=False)

@pembimbing_bp.route('/download_arus_kas', methods=['POST'])
def pembimbing_download_arus_kas():
    import traceback
    if 'user_type' not in session or session['user_type'] != 'pembimbing':
        return jsonify({'success': False, 'message': 'Unauthorized'})
    
    try:
        proposal_id = request.form.get('proposal_id')
        format_type = request.form.get('format')
        bulan_tahun = request.form.get('bulan_tahun')
        
        if not all([proposal_id, format_type, bulan_tahun]):
            return jsonify({'success': False, 'message': 'Parameter tidak lengkap'})
        
        app_funcs = get_app_functions()
        if not hasattr(app_funcs['mysql'], 'connection') or app_funcs['mysql'].connection is None:
            return jsonify({'success': False, 'message': 'Koneksi database tidak tersedia'})
        
        cursor = app_funcs['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Debug: Print parameter yang diterima
        logger.debug(f"DEBUG - proposal_id: {proposal_id}")
        logger.debug(f"DEBUG - session['nama']: {session.get('nama')}")
        
        # Verifikasi bahwa mahasiswa dibimbing oleh pembimbing ini
        cursor.execute('''
            SELECT m.nama_ketua, m.nim, m.program_studi, p.judul_usaha, p.kategori, p.tahun_nib
            FROM mahasiswa m
            JOIN proposal p ON m.nim = p.nim
            WHERE p.id = %s AND LOWER(p.dosen_pembimbing) = LOWER(%s)
        ''', (proposal_id, session['nama']))
        mahasiswa_info = cursor.fetchone()
        logger.debug(f"DEBUG - mahasiswa_info: {mahasiswa_info}")
        
        # Debug: Cek data proposal dan pembimbing
        if not mahasiswa_info:
            logger.debug("DEBUG - Query gagal, cek data proposal...")
            cursor.execute('SELECT id, nim, dosen_pembimbing, judul_usaha FROM proposal WHERE id = %s', (proposal_id,))
            proposal_data = cursor.fetchone()
            logger.debug(f"DEBUG - proposal_data: {proposal_data}")
            
            if proposal_data:
                logger.debug(f"DEBUG - dosen_pembimbing di DB: '{proposal_data['dosen_pembimbing']}'")
                logger.debug(f"DEBUG - session['nama']: '{session['nama']}'")
                logger.debug(f"DEBUG - Apakah sama? {proposal_data['dosen_pembimbing'] == session['nama']}")
            
            cursor.close()
            return jsonify({'success': False, 'message': 'Data mahasiswa tidak ditemukan atau tidak ada akses'})
        
        # Parse bulan yang dipilih
        if not bulan_tahun:
            return jsonify({'success': False, 'message': 'Bulan tahun tidak boleh kosong'})
        
        try:
            selected_date = datetime.strptime(bulan_tahun, '%Y-%m')
            selected_month_year = selected_date.strftime('%B %Y')
        except ValueError:
            return jsonify({'success': False, 'message': 'Format bulan tahun tidak valid. Gunakan format YYYY-MM'})
        
        # Ambil data arus kas dari tabel arus_kas
        cursor.execute('''
            SELECT total_penjualan, total_biaya_produksi, total_biaya_operasional,
                   total_biaya_non_operasional, kas_bersih_operasional, total_harga_jual_alat,
                   total_harga_alat, kas_bersih_investasi, kas_bersih_pembiayaan, total_kas_bersih
            FROM arus_kas 
            WHERE proposal_id = %s AND bulan_tahun = %s
        ''', (proposal_id, bulan_tahun))
        
        arus_kas_data = cursor.fetchone()
        
        if arus_kas_data:
            # Jika data ada di tabel arus_kas, gunakan data tersebut
            total_penjualan = arus_kas_data['total_penjualan'] or 0
            total_biaya_produksi = arus_kas_data['total_biaya_produksi'] or 0
            total_biaya_operasional = arus_kas_data['total_biaya_operasional'] or 0
            total_biaya_non_operasional = arus_kas_data['total_biaya_non_operasional'] or 0
            kas_bersih_operasional = arus_kas_data['kas_bersih_operasional'] or 0
            total_harga_jual_alat = arus_kas_data['total_harga_jual_alat'] or 0
            total_harga_alat = arus_kas_data['total_harga_alat'] or 0
            kas_bersih_investasi = arus_kas_data['kas_bersih_investasi'] or 0
            kas_bersih_pembiayaan = arus_kas_data['kas_bersih_pembiayaan'] or 0
            total_kas_bersih = arus_kas_data['total_kas_bersih'] or 0
        else:
            # Jika data tidak ada di tabel arus_kas, hitung dari tabel individual
            cursor.execute('''
                SELECT COALESCE(SUM(total), 0) as total_penjualan
                FROM penjualan 
                WHERE proposal_id = %s 
                AND DATE_FORMAT(tanggal_penjualan, '%%Y-%%m') = %s
            ''', (proposal_id, bulan_tahun))
            total_penjualan = cursor.fetchone()['total_penjualan'] or 0
            
            cursor.execute('''
                SELECT COALESCE(SUM(total_harga), 0) as total_biaya_produksi
                FROM bahan_baku 
                WHERE proposal_id = %s 
                AND DATE_FORMAT(tanggal_beli, '%%Y-%%m') = %s
            ''', (proposal_id, bulan_tahun))
            total_biaya_produksi = cursor.fetchone()['total_biaya_produksi'] or 0
            
            cursor.execute('''
                SELECT COALESCE(SUM(total_harga), 0) as total_biaya_operasional
                FROM biaya_operasional 
                WHERE proposal_id = %s 
                AND DATE_FORMAT(tanggal_beli, '%%Y-%%m') = %s
            ''', (proposal_id, bulan_tahun))
            total_biaya_operasional = cursor.fetchone()['total_biaya_operasional'] or 0
            
            cursor.execute('''
                SELECT COALESCE(SUM(total_harga), 0) as total_biaya_non_operasional
                FROM biaya_non_operasional 
                WHERE proposal_id = %s 
                AND DATE_FORMAT(tanggal_transaksi, '%%Y-%%m') = %s
            ''', (proposal_id, bulan_tahun))
            total_biaya_non_operasional = cursor.fetchone()['total_biaya_non_operasional'] or 0
            
            cursor.execute('''
                SELECT 
                    COALESCE(SUM(harga), 0) as total_harga_alat,
                    COALESCE(SUM(harga_jual), 0) as total_harga_jual_alat
                FROM alat_produksi 
                WHERE proposal_id = %s 
                AND DATE_FORMAT(tanggal_beli, '%%Y-%%m') = %s
            ''', (proposal_id, bulan_tahun))
            alat_data = cursor.fetchone()
            total_harga_alat = alat_data['total_harga_alat'] or 0
            total_harga_jual_alat = alat_data['total_harga_jual_alat'] or 0
            
            # Hitung kas bersih
            kas_bersih_operasional = total_penjualan - total_biaya_produksi - total_biaya_operasional - total_biaya_non_operasional
            kas_bersih_investasi = total_harga_jual_alat - total_harga_alat
            kas_bersih_pembiayaan = 0  # Selalu 0 sesuai permintaan
            total_kas_bersih = kas_bersih_operasional + kas_bersih_investasi
        
        # Siapkan data untuk download
        arus_kas_data = {
            'total_penjualan': total_penjualan,
            'total_biaya_produksi': total_biaya_produksi,
            'total_biaya_operasional': total_biaya_operasional,
            'total_biaya_non_operasional': total_biaya_non_operasional,
            'kas_bersih_operasional': kas_bersih_operasional,
            'total_harga_jual_alat': total_harga_jual_alat,
            'total_harga_alat': total_harga_alat,
            'kas_bersih_investasi': kas_bersih_investasi,
            'kas_bersih_pembiayaan': 0,  # Selalu 0 sesuai permintaan
            'total_kas_bersih': total_kas_bersih
        }
        
        cursor.close()
        
        # Generate filename
        filename = f"Laporan_Arus_Kas_{mahasiswa_info['nama_ketua'].replace(' ', '_')}_{mahasiswa_info['judul_usaha'].replace(' ', '_')}_{bulan_tahun}"
        
        if format_type == 'excel':
            return app_funcs['generate_excel_arus_kas'](arus_kas_data, mahasiswa_info, bulan_tahun, selected_month_year, filename)
        elif format_type == 'pdf':
            return app_funcs['generate_pdf_arus_kas'](arus_kas_data, mahasiswa_info, bulan_tahun, selected_month_year, filename)
        elif format_type == 'word':
            return app_funcs['generate_word_arus_kas'](arus_kas_data, mahasiswa_info, bulan_tahun, selected_month_year, filename)
        else:
            return jsonify({'success': False, 'message': 'Format tidak didukung'})
            
    except Exception as e:
        logger.error('--- ERROR DOWNLOAD ARUS KAS PEMBIMBING ---')
        logger.error(traceback.format_exc())
        return jsonify({'success': False, 'message': f'Error: {str(e)}', 'trace': traceback.format_exc()})



@pembimbing_bp.route('/laporan_penjualan/<int:mahasiswa_id>')
def pembimbing_laporan_penjualan(mahasiswa_id):
    if 'user_type' not in session or session['user_type'] != 'pembimbing':
        flash('Anda harus login sebagai pembimbing!', 'danger')
        return redirect(url_for('index'))
    
    app_funcs = get_app_functions()
    if not hasattr(app_funcs['mysql'], 'connection') or app_funcs['mysql'].connection is None:
        flash('Koneksi ke database gagal. Cek konfigurasi database!', 'danger')
        return render_template('pembimbing/laporan_penjualan.html', mahasiswa_info=None, penjualan_list=[])
    
    try:
        cursor = app_funcs['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Ambil data mahasiswa yang dibimbing oleh pembimbing ini
        cursor.execute('''
            SELECT m.*, p.judul_usaha, p.id as proposal_id
            FROM mahasiswa m
            JOIN proposal p ON m.nim = p.nim
            WHERE m.id = %s AND p.dosen_pembimbing = %s
        ''', (mahasiswa_id, session['nama']))
        
        mahasiswa_info = cursor.fetchone()
        if not mahasiswa_info:
            flash('Data mahasiswa tidak ditemukan atau Anda tidak memiliki akses!', 'danger')
            cursor.close()
            return redirect(url_for('pembimbing.pembimbing_monitoring_mahasiswa_penjualan_produk'))
        
        # Ambil data penjualan
        cursor.execute('''
            SELECT * FROM penjualan 
            WHERE proposal_id = %s
            ORDER BY tanggal_penjualan DESC
        ''', (mahasiswa_info['proposal_id'],))
        
        penjualan_list = cursor.fetchall()
        for penjualan in penjualan_list:
            if penjualan['tanggal_penjualan']:
                try:
                    penjualan['tanggal_penjualan_indo'] = format_tanggal_indonesia(penjualan['tanggal_penjualan'])
                except Exception:
                    penjualan['tanggal_penjualan_indo'] = penjualan['tanggal_penjualan']
            else:
                penjualan['tanggal_penjualan_indo'] = '-'
        cursor.close()
        
        return render_template('pembimbing/laporan_penjualan.html', mahasiswa_info=mahasiswa_info, penjualan_list=penjualan_list)
        
    except Exception as e:
        flash(f'Error: {str(e)}', 'danger')
        return render_template('pembimbing/laporan_penjualan.html', mahasiswa_info=None, penjualan_list=[])

@pembimbing_bp.route('/bahan_baku/<int:mahasiswa_id>')
def pembimbing_bahan_baku(mahasiswa_id):
    if 'user_type' not in session or session['user_type'] != 'pembimbing':
        flash('Anda harus login sebagai pembimbing!', 'danger')
        return redirect(url_for('index'))
    
    app_funcs = get_app_functions()
    if not hasattr(app_funcs['mysql'], 'connection') or app_funcs['mysql'].connection is None:
        flash('Koneksi ke database gagal. Cek konfigurasi database!', 'danger')
        return render_template('pembimbing/bahan_baku.html', mahasiswa_info=None, bahan_baku_list=[])
    
    try:
        cursor = app_funcs['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Log aktivitas pembimbing
        pembimbing_info = app_funcs['get_pembimbing_info_from_session']()
        if pembimbing_info:
            app_funcs['log_pembimbing_activity'](
                pembimbing_info['id'],
                pembimbing_info['nip'],
                pembimbing_info['nama'],
                'view',
                'monitoring',
                f'bahan_baku_mahasiswa_id_{mahasiswa_id}',
                f'Melihat detail bahan baku mahasiswa ID {mahasiswa_id}',
                None,
                None,
                mahasiswa_id
            )
        
        # Ambil data mahasiswa yang dibimbing oleh pembimbing ini
        cursor.execute('''
            SELECT m.*, p.judul_usaha, p.id as proposal_id
            FROM mahasiswa m
            JOIN proposal p ON m.nim = p.nim
            WHERE m.id = %s AND p.dosen_pembimbing = %s
        ''', (mahasiswa_id, session['nama']))
        
        mahasiswa_info = cursor.fetchone()
        if not mahasiswa_info:
            flash('Data mahasiswa tidak ditemukan atau Anda tidak memiliki akses!', 'danger')
            cursor.close()
            return redirect(url_for('pembimbing.pembimbing_monitoring_mahasiswa_produksi'))
        
        # Ambil data bahan baku
        cursor.execute('''
            SELECT * FROM bahan_baku 
            WHERE proposal_id = %s 
            ORDER BY tanggal_beli DESC
        ''', (mahasiswa_info['proposal_id'],))
        
        bahan_baku_list = cursor.fetchall()
        cursor.close()
        
        return render_template('pembimbing/bahan_baku.html', mahasiswa_info=mahasiswa_info, bahan_baku_list=bahan_baku_list)
        
    except Exception as e:
        flash(f'Error: {str(e)}', 'danger')
        return render_template('pembimbing/bahan_baku.html', mahasiswa_info=None, bahan_baku_list=[])

@pembimbing_bp.route('/produksi/<int:mahasiswa_id>')
def pembimbing_produksi(mahasiswa_id):
    if 'user_type' not in session or session['user_type'] != 'pembimbing':
        flash('Anda harus login sebagai pembimbing!', 'danger')
        return redirect(url_for('index'))
    
    app_funcs = get_app_functions()
    if not hasattr(app_funcs['mysql'], 'connection') or app_funcs['mysql'].connection is None:
        flash('Koneksi ke database gagal. Cek konfigurasi database!', 'danger')
        return render_template('pembimbing/produksi.html', mahasiswa_info=None, produksi_list=[])
    
    try:
        cursor = app_funcs['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Ambil data mahasiswa yang dibimbing oleh pembimbing ini
        cursor.execute('''
            SELECT m.*, p.judul_usaha, p.id as proposal_id
            FROM mahasiswa m
            JOIN proposal p ON m.nim = p.nim
            WHERE m.id = %s AND p.dosen_pembimbing = %s
        ''', (mahasiswa_id, session['nama']))
        
        mahasiswa_info = cursor.fetchone()
        if not mahasiswa_info:
            flash('Data mahasiswa tidak ditemukan atau Anda tidak memiliki akses!', 'danger')
            cursor.close()
            return redirect(url_for('pembimbing.pembimbing_monitoring_mahasiswa_produksi'))
        
        # Ambil data produksi
        cursor.execute('''
            SELECT * FROM produksi 
            WHERE proposal_id = %s 
            ORDER BY tanggal_produksi DESC
        ''', (mahasiswa_info['proposal_id'],))
        
        produksi_list = cursor.fetchall()
        
        # Tambahkan data bahan baku untuk setiap produksi
        for produksi in produksi_list:
            cursor.execute('''
                SELECT pbb.*, bb.nama_bahan, bb.satuan
                FROM produk_bahan_baku pbb
                JOIN bahan_baku bb ON pbb.bahan_baku_id = bb.id
                WHERE pbb.produksi_id = %s
            ''', (produksi['id'],))
            bahan_baku_list = cursor.fetchall()
            produksi['bahan_baku'] = bahan_baku_list
        
        cursor.close()
        
        return render_template('pembimbing/produksi.html', mahasiswa_info=mahasiswa_info, produksi_list=produksi_list)
        
    except Exception as e:
        flash(f'Error: {str(e)}', 'danger')
        return render_template('pembimbing/produksi.html', mahasiswa_info=None, produksi_list=[])


@pembimbing_bp.route('/update_anggaran', methods=['POST'])
def pembimbing_update_anggaran():
    if 'user_type' not in session or session['user_type'] != 'pembimbing':
        return jsonify({'success': False, 'message': 'Anda harus login sebagai pembimbing!'})
    app_funcs = get_app_functions()
    if not hasattr(app_funcs['mysql'], 'connection') or app_funcs['mysql'].connection is None:
        return jsonify({'success': False, 'message': 'Koneksi ke database gagal!'})
    try:
        # Handle both JSON and form data
        if request.is_json:
            data = request.get_json()
        else:
            data = request.form.to_dict()
        
        anggaran_id = data.get('id')
        tabel = data.get('tabel', 'anggaran_awal')
        if not anggaran_id:
            return jsonify({'success': False, 'message': 'ID anggaran tidak ditemukan!'})
        if tabel not in ['anggaran_awal', 'anggaran_bertumbuh']:
            return jsonify({'success': False, 'message': 'Tabel anggaran tidak valid!'})
        
        cursor = app_funcs['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Verify the anggaran belongs to a proposal supervised by this pembimbing
        cursor.execute(f'''
            SELECT aa.id 
            FROM {tabel} aa
            JOIN proposal p ON aa.id_proposal = p.id
            WHERE aa.id = %s AND p.dosen_pembimbing = %s
        ''', (anggaran_id, session['nama']))
        
        anggaran = cursor.fetchone()
        if not anggaran:
            cursor.close()
            return jsonify({'success': False, 'message': 'Data anggaran tidak ditemukan atau tidak ada akses!'})
        
        # Ambil data lama untuk logging
        cursor.execute(f'''
            SELECT aa.*, p.judul_usaha
            FROM {tabel} aa
            JOIN proposal p ON aa.id_proposal = p.id
            WHERE aa.id = %s
        ''', (anggaran_id,))
        data_lama = cursor.fetchone()
        
        # Ambil dan validasi data dari request
        kegiatan_utama = data.get('kegiatan_utama', '').strip()
        kegiatan = data.get('kegiatan', '').strip()
        nama_barang = data.get('nama_barang', '').strip()
        penanggung_jawab = data.get('penanggung_jawab', '').strip()
        satuan = data.get('satuan', '').strip()
        target_capaian = data.get('target_capaian', '').strip()
        keterangan = data.get('keterangan', '').strip()
        
        # Validasi dan konversi tipe data numerik
        try:
            kuantitas = int(data.get('kuantitas', 0))
            if kuantitas < 0:
                return jsonify({'success': False, 'message': 'Kuantitas tidak boleh negatif!'})
        except (ValueError, TypeError):
            return jsonify({'success': False, 'message': 'Kuantitas harus berupa angka!'})
        
        try:
            harga_satuan = float(data.get('harga_satuan', 0))
            if harga_satuan < 0:
                return jsonify({'success': False, 'message': 'Harga satuan tidak boleh negatif!'})
        except (ValueError, TypeError):
            return jsonify({'success': False, 'message': 'Harga satuan harus berupa angka!'})
        
        try:
            jumlah = float(data.get('jumlah', 0))
            if jumlah < 0:
                return jsonify({'success': False, 'message': 'Jumlah tidak boleh negatif!'})
        except (ValueError, TypeError):
            return jsonify({'success': False, 'message': 'Jumlah harus berupa angka!'})
        
        # Validasi field yang required
        if not kegiatan:
            return jsonify({'success': False, 'message': 'Kegiatan harus diisi!'})
        if not penanggung_jawab:
            return jsonify({'success': False, 'message': 'Penanggung jawab harus diisi!'})
        if not target_capaian:
            return jsonify({'success': False, 'message': 'Target capaian harus diisi!'})
        if not nama_barang:
            return jsonify({'success': False, 'message': 'Nama barang harus diisi!'})
        if not satuan:
            return jsonify({'success': False, 'message': 'Satuan harus diisi!'})
        
        # Validasi satuan tidak boleh kosong
        if not satuan or satuan.strip() == '':
            return jsonify({'success': False, 'message': 'Satuan harus diisi!'})
        
        # Update data anggaran
        cursor.execute(f'''
            UPDATE {tabel} SET 
                kegiatan_utama = %s, kegiatan = %s, nama_barang = %s, penanggung_jawab = %s,
                satuan = %s, harga_satuan = %s, target_capaian = %s, keterangan = %s,
                kuantitas = %s, jumlah = %s
            WHERE id = %s
        ''', (
            kegiatan_utama, kegiatan, nama_barang, penanggung_jawab,
            satuan, harga_satuan, target_capaian, keterangan,
            kuantitas, jumlah, anggaran_id
        ))
        
        app_funcs['mysql'].connection.commit()
        
        # Log aktivitas pembimbing
        pembimbing_info = app_funcs['get_pembimbing_info_from_session']()
        if pembimbing_info and data_lama:
            data_lama_log = {
                'kegiatan_utama': data_lama.get('kegiatan_utama'),
                'kegiatan': data_lama.get('kegiatan'),
                'nama_barang': data_lama.get('nama_barang'),
                'kuantitas': data_lama.get('kuantitas'),
                'harga_satuan': data_lama.get('harga_satuan'),
                'jumlah': data_lama.get('jumlah')
            }
            data_baru_log = {
                'kegiatan_utama': kegiatan_utama,
                'kegiatan': kegiatan,
                'nama_barang': nama_barang,
                'kuantitas': kuantitas,
                'harga_satuan': harga_satuan,
                'jumlah': jumlah
            }
            
            app_funcs['log_pembimbing_activity'](
                pembimbing_info['id'],
                pembimbing_info['nip'],
                pembimbing_info['nama'],
                'edit',
                tabel,
                f'{tabel}_id_{anggaran_id}',
                f'Mengedit {tabel} ID {anggaran_id} untuk proposal "{data_lama["judul_usaha"]}"',
                data_lama_log,
                data_baru_log,
                None,  # mahasiswa_id tidak tersedia
                data_lama.get('id_proposal')
            )
        
        cursor.close()
        return jsonify({'success': True, 'message': 'Data anggaran berhasil diperbarui!'})
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error saat memperbarui data: {str(e)}'})

@pembimbing_bp.route('/delete_anggaran', methods=['POST'])
def pembimbing_delete_anggaran():
    if 'user_type' not in session or session['user_type'] != 'pembimbing':
        return jsonify({'success': False, 'message': 'Anda harus login sebagai pembimbing!'})
    app_funcs = get_app_functions()
    if not hasattr(app_funcs['mysql'], 'connection') or app_funcs['mysql'].connection is None:
        return jsonify({'success': False, 'message': 'Koneksi ke database gagal!'})
    try:
        # Handle both JSON and form data
        if request.is_json:
            data = request.get_json()
        else:
            data = request.form.to_dict()
        
        anggaran_id = data.get('id')
        tabel = data.get('tabel', 'anggaran_awal')
        
        if not anggaran_id:
            return jsonify({'success': False, 'message': 'ID anggaran tidak ditemukan!'})
        
        if tabel not in ['anggaran_awal', 'anggaran_bertumbuh']:
            return jsonify({'success': False, 'message': 'Tabel anggaran tidak valid!'})
        
        cursor = app_funcs['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Verify the anggaran belongs to a proposal supervised by this pembimbing
        cursor.execute(f'''
            SELECT aa.*, p.id as proposal_id, p.judul_usaha, p.tahapan_usaha, p.nim
            FROM {tabel} aa
            JOIN proposal p ON aa.id_proposal = p.id
            WHERE aa.id = %s AND p.dosen_pembimbing = %s
        ''', (anggaran_id, session['nama']))
        
        anggaran = cursor.fetchone()
        if not anggaran:
            cursor.close()
            return jsonify({'success': False, 'message': 'Data anggaran tidak ditemukan atau tidak ada akses!'})
        
        proposal_id = anggaran['proposal_id']
        tahapan_usaha = anggaran['tahapan_usaha']
        nim = anggaran['nim']
        
        # Tentukan tabel laporan berdasarkan tahapan usaha
        if 'bertumbuh' in tahapan_usaha.lower():
            tabel_laporan_kemajuan = 'laporan_kemajuan_bertumbuh'
            tabel_laporan_akhir = 'laporan_akhir_bertumbuh'
        else:
            tabel_laporan_kemajuan = 'laporan_kemajuan_awal'
            tabel_laporan_akhir = 'laporan_akhir_awal'
        
        # Hapus data laporan kemajuan terkait
        cursor.execute(f'DELETE FROM {tabel_laporan_kemajuan} WHERE id_proposal = %s', (proposal_id,))
        laporan_kemajuan_deleted = cursor.rowcount
        logger.info(f"✅ Berhasil menghapus {laporan_kemajuan_deleted} data laporan kemajuan untuk proposal {proposal_id}")
        
        # Hapus data laporan akhir terkait
        cursor.execute(f'DELETE FROM {tabel_laporan_akhir} WHERE id_proposal = %s', (proposal_id,))
        laporan_akhir_deleted = cursor.rowcount
        logger.info(f"✅ Berhasil menghapus {laporan_akhir_deleted} data laporan akhir untuk proposal {proposal_id}")
        
        # Hapus file laporan kemajuan
        app_funcs['hapus_file_laporan_kemajuan'](proposal_id, nim)
        
        # Hapus file laporan akhir
        app_funcs['hapus_file_laporan_akhir'](proposal_id, nim)
        
        # Hapus data anggaran
        cursor.execute(f'DELETE FROM {tabel} WHERE id = %s', (anggaran_id,))
        anggaran_deleted = cursor.rowcount
        
        # Cek apakah ada baris yang terhapus
        if anggaran_deleted == 0:
            cursor.close()
            return jsonify({'success': False, 'message': 'Gagal menghapus data anggaran!'})
        
        app_funcs['mysql'].connection.commit()
        
        # Log aktivitas pembimbing
        pembimbing_info = app_funcs['get_pembimbing_info_from_session']()
        if pembimbing_info and anggaran:
            app_funcs['log_pembimbing_activity'](
                pembimbing_info['id'],
                pembimbing_info['nip'],
                pembimbing_info['nama'],
                'hapus',
                tabel,
                f'{tabel}_id_{anggaran_id}',
                f'Menghapus {tabel} ID {anggaran_id} "{anggaran["nama_barang"]}" untuk proposal "{anggaran["judul_usaha"]}" beserta laporan kemajuan dan akhir',
                {
                    'kegiatan': anggaran.get('kegiatan'),
                    'nama_barang': anggaran.get('nama_barang'),
                    'kuantitas': anggaran.get('kuantitas'),
                    'harga_satuan': anggaran.get('harga_satuan'),
                    'jumlah': anggaran.get('jumlah'),
                    'laporan_kemajuan_deleted': laporan_kemajuan_deleted,
                    'laporan_akhir_deleted': laporan_akhir_deleted
                },
                None,
                None,  # mahasiswa_id tidak tersedia
                proposal_id
            )
        
        cursor.close()
        
        message = f'Data anggaran berhasil dihapus!'
        if laporan_kemajuan_deleted > 0:
            message += f' {laporan_kemajuan_deleted} data laporan kemajuan juga dihapus.'
        if laporan_akhir_deleted > 0:
            message += f' {laporan_akhir_deleted} data laporan akhir juga dihapus.'
        if laporan_kemajuan_deleted > 0 or laporan_akhir_deleted > 0:
            message += ' File-file terkait juga dihapus.'
        
        return jsonify({'success': True, 'message': message})
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error saat menghapus data: {str(e)}'})

@pembimbing_bp.route('/get_anggaran_by_id', methods=['GET'])
def pembimbing_get_anggaran_by_id():
    if 'user_type' not in session or session['user_type'] != 'pembimbing':
        return jsonify({'success': False, 'message': 'Anda harus login sebagai pembimbing!'})
    
    app_funcs = get_app_functions()
    if not hasattr(app_funcs['mysql'], 'connection') or app_funcs['mysql'].connection is None:
        return jsonify({'success': False, 'message': 'Koneksi ke database gagal!'})
    
    try:
        anggaran_id = request.args.get('id')
        tabel = request.args.get('tabel', 'anggaran_awal')
        
        if not anggaran_id:
            return jsonify({'success': False, 'message': 'ID anggaran tidak ditemukan!'})
        
        if tabel not in ['anggaran_awal', 'anggaran_bertumbuh']:
            return jsonify({'success': False, 'message': 'Tabel anggaran tidak valid!'})
        
        cursor = app_funcs['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Verify the anggaran belongs to a proposal supervised by this pembimbing
        cursor.execute(f'''
            SELECT aa.* 
            FROM {tabel} aa
            JOIN proposal p ON aa.id_proposal = p.id
            WHERE aa.id = %s AND p.dosen_pembimbing = %s
        ''', (anggaran_id, session['nama']))
        
        anggaran = cursor.fetchone()
        cursor.close()
        
        if not anggaran:
            return jsonify({'success': False, 'message': 'Data anggaran tidak ditemukan atau tidak ada akses!'})
        
        return jsonify({'success': True, 'data': anggaran})
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error saat mengambil data: {str(e)}'})

@pembimbing_bp.route('/konfirmasi_anggaran', methods=['POST'])
def pembimbing_konfirmasi_anggaran():
    if 'user_type' not in session or session['user_type'] != 'pembimbing':
        return jsonify({'success': False, 'message': 'Anda harus login sebagai pembimbing!'})
    
    app_funcs = get_app_functions()
    if not hasattr(app_funcs['mysql'], 'connection') or app_funcs['mysql'].connection is None:
        return jsonify({'success': False, 'message': 'Koneksi ke database gagal!'})
    
    try:
        data = request.get_json()
        anggaran_id = data.get('id')
        status = data.get('status')  # 'disetujui', 'ditolak', 'revisi'
        jenis = data.get('jenis')  # 'anggaran_awal', 'anggaran_bertumbuh'
        
        if not anggaran_id or not status:
            return jsonify({'success': False, 'message': 'Data tidak lengkap!'})
        
        # Tentukan tabel berdasarkan jenis
        if jenis == 'anggaran_bertumbuh':
            tabel = 'anggaran_bertumbuh'
        else:
            tabel = 'anggaran_awal'
        
        if tabel not in ['anggaran_awal', 'anggaran_bertumbuh']:
            return jsonify({'success': False, 'message': 'Tabel anggaran tidak valid!'})
        
        cursor = app_funcs['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Verify the anggaran belongs to a proposal supervised by this pembimbing
        cursor.execute(f'''
            SELECT aa.id 
            FROM {tabel} aa
            JOIN proposal p ON aa.id_proposal = p.id
            WHERE aa.id = %s AND p.dosen_pembimbing = %s
        ''', (anggaran_id, session['nama']))
        
        anggaran = cursor.fetchone()
        if not anggaran:
            cursor.close()
            return jsonify({'success': False, 'message': 'Data anggaran tidak ditemukan atau tidak ada akses!'})
        
        # Get proposal ID from anggaran first
        cursor.execute(f'''
            SELECT id_proposal FROM {tabel} WHERE id = %s
        ''', (anggaran_id,))
        
        anggaran = cursor.fetchone()
        if not anggaran:
            cursor.close()
            return jsonify({'success': False, 'message': 'Data anggaran tidak ditemukan!'})
        
        proposal_id = anggaran['id_proposal']
        
        # Ambil data anggaran sebelum update untuk logging
        cursor.execute(f'''
            SELECT aa.*, p.judul_usaha 
            FROM {tabel} aa
            JOIN proposal p ON aa.id_proposal = p.id
            WHERE aa.id = %s
        ''', (anggaran_id,))
        anggaran_data = cursor.fetchone()
        
        # Update status anggaran
        cursor.execute(f'''
            UPDATE {tabel} 
            SET status = %s 
            WHERE id = %s
        ''', (status, anggaran_id))
        
        # Log aktivitas pembimbing
        pembimbing_info = app_funcs['get_pembimbing_info_from_session']()
        if pembimbing_info and anggaran_data:
            jenis_aktivitas = 'setuju' if status == 'disetujui' else 'tolak' if status == 'ditolak' else 'revisi'
            modul = 'anggaran_awal' if tabel == 'anggaran_awal' else 'anggaran_bertumbuh'
            deskripsi = f'Mengubah status {modul} ID {anggaran_id} menjadi {status} untuk proposal "{anggaran_data["judul_usaha"]}"'
            
            app_funcs['log_pembimbing_activity'](
                pembimbing_info['id'],
                pembimbing_info['nip'],
                pembimbing_info['nama'],
                jenis_aktivitas,
                modul,
                f'{modul}_id_{anggaran_id}',
                deskripsi,
                {'status': anggaran_data.get('status', 'pending')},
                {'status': status},
                None,  # mahasiswa_id tidak tersedia
                proposal_id
            )
        
        # Update laporan kemajuan setiap kali status berubah (tidak hanya disetujui)
        # Tentukan tabel laporan kemajuan berdasarkan tabel anggaran
        if tabel == 'anggaran_awal':
            tabel_laporan = 'laporan_kemajuan_awal'
        else:
            tabel_laporan = 'laporan_kemajuan_bertumbuh'
        
        # Update laporan kemajuan dengan data anggaran terbaru
        logger.debug(f"Debug: Calling update_laporan_kemajuan_from_anggaran for proposal {proposal_id}")
        app_funcs['update_laporan_kemajuan_from_anggaran'](cursor, proposal_id, tabel_laporan, tabel)
        
        # Check if all anggaran for this proposal are approved
        cursor.execute(f'''
            SELECT COUNT(*) as total, 
                   SUM(CASE WHEN status = 'disetujui' THEN 1 ELSE 0 END) as approved
            FROM {tabel} 
            WHERE id_proposal = %s
        ''', (proposal_id,))
        
        result = cursor.fetchone()
        if result['total'] > 0 and result['total'] == result['approved']:
                # All anggaran approved, update proposal status
                cursor.execute('''
                    UPDATE proposal 
                    SET status = 'disetujui' 
                    WHERE id = %s AND dosen_pembimbing = %s
                ''', (proposal_id, session['nama']))
        
        app_funcs['mysql'].connection.commit()
        cursor.close()
        
        return jsonify({'success': True, 'message': f'Status anggaran berhasil diubah menjadi {status}!'})
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@pembimbing_bp.route('/update_anggaran_status/<int:anggaran_id>', methods=['POST'])
def pembimbing_update_anggaran_status(anggaran_id):
    if 'user_type' not in session or session['user_type'] != 'pembimbing':
        return jsonify({'success': False, 'message': 'Anda harus login sebagai pembimbing!'})
    app_funcs = get_app_functions()
    if not hasattr(app_funcs['mysql'], 'connection') or app_funcs['mysql'].connection is None:
        return jsonify({'success': False, 'message': 'Koneksi ke database gagal!'})
    try:
        data = request.get_json()
        status = data.get('status')
        tabel = data.get('tabel', 'anggaran_awal')
        
        if status not in ['disetujui', 'ditolak', 'revisi']:
            return jsonify({'success': False, 'message': 'Status tidak valid!'})
        
        if tabel not in ['anggaran_awal', 'anggaran_bertumbuh']:
            return jsonify({'success': False, 'message': 'Tabel anggaran tidak valid!'})
        
        cursor = app_funcs['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Validasi bahwa anggaran milik mahasiswa yang dibimbing oleh pembimbing ini
        query = f"""
            SELECT aa.id FROM {tabel} aa
            JOIN proposal p ON aa.id_proposal = p.id
            WHERE aa.id = %s AND p.dosen_pembimbing = %s
        """
        cursor.execute(query, (anggaran_id, session['nama']))
        anggaran = cursor.fetchone()
        
        if not anggaran:
            return jsonify({'success': False, 'message': 'Anggaran tidak ditemukan atau tidak memiliki akses!'})
        
        # Ambil proposal_id dari anggaran
        cursor.execute(f"SELECT id_proposal FROM {tabel} WHERE id = %s", (anggaran_id,))
        anggaran_data = cursor.fetchone()
        if not anggaran_data:
            return jsonify({'success': False, 'message': 'Data anggaran tidak ditemukan!'})
        
        proposal_id = anggaran_data['id_proposal']
        
        # Update status anggaran
        update_query = f"UPDATE {tabel} SET status = %s WHERE id = %s"
        cursor.execute(update_query, (status, anggaran_id))
        
        # Update laporan kemajuan setiap kali status berubah (tidak hanya disetujui)
        # Tentukan tabel laporan kemajuan berdasarkan tabel anggaran
        if tabel == 'anggaran_awal':
            tabel_laporan = 'laporan_kemajuan_awal'
        else:
            tabel_laporan = 'laporan_kemajuan_bertumbuh'
        
        # Update laporan kemajuan dengan data anggaran terbaru
        logger.debug(f"Debug: Calling update_laporan_kemajuan_from_anggaran for proposal {proposal_id}")
        app_funcs['update_laporan_kemajuan_from_anggaran'](cursor, proposal_id, tabel_laporan, tabel)
        
        app_funcs['mysql'].connection.commit()
        
        return jsonify({'success': True, 'message': f'Status anggaran berhasil diupdate menjadi {status}!'})
        
    except Exception as e:
        logger.error(f"Error updating anggaran status: {e}")
        return jsonify({'success': False, 'message': 'Terjadi kesalahan saat mengupdate status anggaran!'})
    finally:
        if 'cursor' in locals():
            cursor.close()

@pembimbing_bp.route('/pembimbing_laporan_kemajuan')
def pembimbing_laporan_kemajuan():
    if 'user_type' not in session or session['user_type'] != 'pembimbing':
        flash('Anda harus login sebagai pembimbing!', 'danger')
        return redirect(url_for('index'))
    
    app_funcs = get_app_functions()
    
    # Log aktivitas pembimbing
    pembimbing_info = app_funcs['get_pembimbing_info_from_session']()
    if pembimbing_info:
        app_funcs['log_pembimbing_activity'](
            pembimbing_info['id'],
            pembimbing_info['nip'],
            pembimbing_info['nama'],
            'view',
            'laporan_kemajuan',
            'daftar_laporan_kemajuan',
            'Melihat halaman daftar laporan kemajuan mahasiswa'
        )
    if not hasattr(app_funcs['mysql'], 'connection') or app_funcs['mysql'].connection is None:
        flash('Koneksi ke database gagal. Cek konfigurasi database!', 'danger')
        return render_template('pembimbing/laporan kemajuan.html', mahasiswa_list=[])
    try:
        cursor = app_funcs['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        # Ambil semua mahasiswa yang dibimbing oleh pembimbing ini beserta proposal dan tahapan_usaha
        cursor.execute('''
            SELECT m.id, m.nama_ketua, m.nim, p.judul_usaha, m.program_studi, p.id as proposal_id, p.tahapan_usaha
            FROM mahasiswa m
            JOIN proposal p ON m.nim = p.nim
            WHERE p.dosen_pembimbing = %s
            ORDER BY m.nama_ketua
        ''', (session['nama'],))
        mahasiswa_all = cursor.fetchall()
        print(f"DEBUG LAPORAN KEMAJUAN: Total mahasiswa yang dibimbing: {len(mahasiswa_all)}")
        mahasiswa_list = []
        for mhs in mahasiswa_all:
            proposal_id = mhs['proposal_id']
            tahapan = (mhs.get('tahapan_usaha') or '').lower()
            
            if 'bertumbuh' in tahapan:
                tabel_laporan = 'laporan_kemajuan_bertumbuh'
                tabel_anggaran = 'anggaran_bertumbuh'
            else:
                tabel_laporan = 'laporan_kemajuan_awal'
                tabel_anggaran = 'anggaran_awal'
            
            # TRIGGER: Cek apakah sudah ada data laporan kemajuan untuk proposal ini
            cursor.execute(f"SELECT COUNT(*) as cnt FROM {tabel_laporan} WHERE id_proposal = %s", (proposal_id,))
            cnt = cursor.fetchone()['cnt']
            
            # Jika belum ada data laporan kemajuan, buat dari anggaran yang sudah direview
            if cnt == 0:
                print(f"DEBUG: Membuat laporan kemajuan untuk proposal {proposal_id}")
                print(f"DEBUG: Menggunakan tabel anggaran: {tabel_anggaran}")
                
                # Tambahkan kolom nilai_bantuan jika belum ada
                try:
                    cursor.execute(f'ALTER TABLE {tabel_laporan} ADD COLUMN nilai_bantuan DECIMAL(15,2) DEFAULT 0.00')
                    app_funcs['mysql'].connection.commit()
                    print(f"DEBUG: Berhasil menambahkan kolom nilai_bantuan ke tabel {tabel_laporan}")
                except Exception as e:
                    if "Duplicate column name" in str(e):
                        print(f"DEBUG: Kolom nilai_bantuan sudah ada di tabel {tabel_laporan}")
                    else:
                        print(f"DEBUG: Error menambahkan kolom nilai_bantuan: {str(e)}")
                
                # Cek apakah ada data anggaran dengan status_reviewer sudah_direview
                cursor.execute(f'''
                    SELECT COUNT(*) as cnt_anggaran 
                    FROM {tabel_anggaran} 
                    WHERE id_proposal = %s AND status_reviewer = 'sudah_direview'
                ''', (proposal_id,))
                cnt_anggaran = cursor.fetchone()['cnt_anggaran']
                print(f"DEBUG: Ditemukan {cnt_anggaran} data anggaran dengan status_reviewer 'sudah_direview'")
                
                if cnt_anggaran > 0:
                    try:
                        cursor.execute(f'''
                            SELECT id, kegiatan_utama, kegiatan, penanggung_jawab, target_capaian, 
                                   nama_barang, satuan, status, status_reviewer, nilai_bantuan
                            FROM {tabel_anggaran} 
                            WHERE id_proposal = %s AND status_reviewer = 'sudah_direview'
                            ORDER BY kegiatan_utama, kegiatan, nama_barang
                        ''', (proposal_id,))
                        anggaran_data = cursor.fetchall()
                        print(f"DEBUG: Berhasil mengambil {len(anggaran_data)} data anggaran")
                        
                        if anggaran_data:
                            print(f"DEBUG: Sample data - kegiatan_utama: {anggaran_data[0].get('kegiatan_utama', 'N/A')}")
                            print(f"DEBUG: Sample data - nama_barang: {anggaran_data[0].get('nama_barang', 'N/A')}")
                            
                            for anggaran in anggaran_data:
                                try:
                                    cursor.execute(f'''
                                        INSERT INTO {tabel_laporan} (
                                            id_proposal, kegiatan_utama, kegiatan, penanggung_jawab, target_capaian,
                                            nama_barang, kuantitas, satuan, harga_satuan, jumlah, keterangan, status, nilai_bantuan
                                        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                                    ''', (
                                        proposal_id, anggaran['kegiatan_utama'], anggaran['kegiatan'], 
                                        anggaran['penanggung_jawab'], anggaran['target_capaian'], anggaran['nama_barang'],
                                        0, anggaran['satuan'], 0, 0, '', 'draf', anggaran.get('nilai_bantuan', 0)
                                    ))
                                    print(f"DEBUG: Berhasil INSERT data untuk kegiatan: {anggaran.get('kegiatan_utama', 'N/A')}")
                                except Exception as insert_error:
                                    print(f"DEBUG: Error saat INSERT: {str(insert_error)}")
                                    print(f"DEBUG: Data yang gagal INSERT: {anggaran}")
                                    raise insert_error
                            
                            app_funcs['mysql'].connection.commit()
                            print(f"DEBUG: Berhasil membuat {len(anggaran_data)} data laporan kemajuan untuk proposal {proposal_id}")
                        else:
                            print(f"DEBUG: Tidak ada data anggaran yang sudah direview untuk proposal {proposal_id}")
                    except Exception as e:
                        print(f"DEBUG: Error dalam proses pembuatan laporan kemajuan: {str(e)}")
                        app_funcs['mysql'].connection.rollback()
                else:
                    print(f"DEBUG: Tidak ada anggaran dengan status_reviewer 'sudah_direview' untuk proposal {proposal_id}")
            else:
                print(f"DEBUG: Data laporan kemajuan sudah ada untuk proposal {proposal_id}")
            
            # Cek ulang apakah sekarang ada data laporan kemajuan
            cursor.execute(f"SELECT COUNT(*) as cnt FROM {tabel_laporan} WHERE id_proposal = %s", (proposal_id,))
            cnt = cursor.fetchone()['cnt']
            if cnt > 0:
                mahasiswa_list.append(mhs)
        cursor.close()
        return render_template('pembimbing/laporan kemajuan.html', mahasiswa_list=mahasiswa_list)
    except Exception as e:
        flash(f'Error: {str(e)}', 'danger')
        return render_template('pembimbing/laporan kemajuan.html', mahasiswa_list=[])

@pembimbing_bp.route('/konfirmasi_laporan_kemajuan', methods=['POST'])
def pembimbing_konfirmasi_laporan_kemajuan():
    if 'user_type' not in session or session['user_type'] != 'pembimbing':
        return jsonify({'success': False, 'message': 'Anda harus login sebagai pembimbing!'})
    
    app_funcs = get_app_functions()
    if not hasattr(app_funcs['mysql'], 'connection') or app_funcs['mysql'].connection is None:
        return jsonify({'success': False, 'message': 'Koneksi ke database gagal!'})
    
    try:
        data = request.get_json()
        laporan_id = data.get('id')
        tabel_laporan = data.get('tabel_laporan')
        action = data.get('action')  # 'setuju', 'tolak', 'revisi'
        
        if not laporan_id or not tabel_laporan or not action:
            return jsonify({'success': False, 'message': 'Data tidak lengkap!'})
        
        if tabel_laporan not in ['laporan_kemajuan_awal', 'laporan_kemajuan_bertumbuh']:
            return jsonify({'success': False, 'message': 'Tabel laporan tidak valid!'})
        
        if action not in ['setuju', 'tolak', 'revisi']:
            return jsonify({'success': False, 'message': 'Aksi tidak valid!'})
        
        cursor = app_funcs['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Cek apakah laporan kemajuan milik mahasiswa yang dibimbing oleh pembimbing ini
        cursor.execute(f'''
            SELECT lk.*, p.dosen_pembimbing 
            FROM {tabel_laporan} lk
            JOIN proposal p ON lk.id_proposal = p.id
            WHERE lk.id = %s AND p.dosen_pembimbing = %s
        ''', (laporan_id, session['nama']))
        
        laporan = cursor.fetchone()
        if not laporan:
            cursor.close()
            return jsonify({'success': False, 'message': 'Laporan kemajuan tidak ditemukan atau Anda tidak memiliki akses!'})
        
        # Update status laporan kemajuan
        if action == 'setuju':
            new_status = 'disetujui'
        elif action == 'tolak':
            new_status = 'ditolak'
        elif action == 'revisi':
            new_status = 'revisi'
        
        # Ambil data lengkap untuk logging
        cursor.execute(f'''
            SELECT lk.*, p.judul_usaha, p.nim
            FROM {tabel_laporan} lk
            JOIN proposal p ON lk.id_proposal = p.id
            WHERE lk.id = %s
        ''', (laporan_id,))
        laporan_data = cursor.fetchone()
        
        cursor.execute(f'''
            UPDATE {tabel_laporan} 
            SET status = %s, tanggal_review = NOW()
            WHERE id = %s
        ''', (new_status, laporan_id))
        
        # Jika status diubah menjadi 'ditolak' atau 'revisi', hapus laporan akhir yang bersangkutan
        laporan_akhir_deleted = 0
        if action in ['tolak', 'revisi']:
            proposal_id = laporan['id_proposal']
            
            # Tentukan tabel laporan akhir berdasarkan tabel laporan kemajuan
            if tabel_laporan == 'laporan_kemajuan_awal':
                tabel_laporan_akhir = 'laporan_akhir_awal'
            else:
                tabel_laporan_akhir = 'laporan_akhir_bertumbuh'
            
            # Hapus data laporan akhir yang bersangkutan
            cursor.execute(f'''
                DELETE FROM {tabel_laporan_akhir} 
                WHERE id_proposal = %s AND kegiatan_utama = %s AND kegiatan = %s AND nama_barang = %s
            ''', (proposal_id, laporan['kegiatan_utama'], laporan['kegiatan'], laporan['nama_barang']))
            
            laporan_akhir_deleted = cursor.rowcount
            logger.debug(f"DEBUG: Menghapus {laporan_akhir_deleted} data laporan akhir yang bersangkutan")
            
            # Hapus file laporan akhir yang terkait (jika ada)
            if laporan_akhir_deleted > 0:
                try:
                    # Ambil data mahasiswa untuk path file
                    cursor.execute('SELECT nim FROM proposal WHERE id = %s', (proposal_id,))
                    proposal_data = cursor.fetchone()
                    if proposal_data:
                        app_funcs['hapus_file_laporan_akhir_terkait'](proposal_id, proposal_data['nim'], laporan['kegiatan_utama'], laporan['kegiatan'], laporan['nama_barang'])
                except Exception as e:
                    logger.error(f"DEBUG: Error saat menghapus file laporan akhir: {e}")
        
        app_funcs['mysql'].connection.commit()
        
        # Log aktivitas pembimbing
        pembimbing_info = app_funcs['get_pembimbing_info_from_session']()
        if pembimbing_info and laporan_data:
            jenis_aktivitas = 'setuju' if action == 'setuju' else 'tolak' if action == 'tolak' else 'revisi'
            modul = 'laporan_kemajuan_awal' if 'awal' in tabel_laporan else 'laporan_kemajuan_bertumbuh'
            deskripsi = f'Mengubah status {modul} ID {laporan_id} menjadi {new_status} untuk proposal "{laporan_data["judul_usaha"]}"'
            
            # Ambil mahasiswa_id berdasarkan nim
            mahasiswa_id = None
            if laporan_data.get('nim'):
                cursor.execute('SELECT id FROM mahasiswa WHERE nim = %s', (laporan_data['nim'],))
                mahasiswa_result = cursor.fetchone()
                if mahasiswa_result:
                    mahasiswa_id = mahasiswa_result['id']
            
            app_funcs['log_pembimbing_activity'](
                pembimbing_info['id'],
                pembimbing_info['nip'],
                pembimbing_info['nama'],
                jenis_aktivitas,
                modul,
                f'{modul}_id_{laporan_id}',
                deskripsi,
                {'status': laporan_data.get('status', 'pending')},
                {'status': new_status},
                mahasiswa_id,
                laporan_data.get('id_proposal')
            )
        
        cursor.close()
        
        if action == 'setuju':
            action_text = 'disetujui'
            message = f'Laporan kemajuan berhasil {action_text}!'
        elif action == 'tolak':
            action_text = 'ditolak'
            message = f'Laporan kemajuan berhasil {action_text}!'
            if laporan_akhir_deleted > 0:
                message += f' {laporan_akhir_deleted} data laporan akhir yang bersangkutan juga dihapus beserta file-file terkait.'
        elif action == 'revisi':
            action_text = 'diminta revisi'
            message = f'Laporan kemajuan berhasil {action_text}!'
            if laporan_akhir_deleted > 0:
                message += f' {laporan_akhir_deleted} data laporan akhir yang bersangkutan juga dihapus beserta file-file terkait.'
        
        return jsonify({'success': True, 'message': message})
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@pembimbing_bp.route('/laporan_kemajuan_awal_bertumbuh/<int:mahasiswa_id>')
def pembimbing_laporan_kemajuan_awal_bertumbuh(mahasiswa_id):
    if 'user_type' not in session or session['user_type'] != 'pembimbing':
        flash('Anda harus login sebagai pembimbing!', 'danger')
        
    app_funcs = get_app_functions()
    if not hasattr(app_funcs['mysql'], 'connection') or app_funcs['mysql'].connection is None:
        flash('Koneksi ke database gagal. Cek konfigurasi database!', 'danger')
        return render_template('pembimbing/laporan_kemajuan_awal_bertumbuh.html', anggaran_data=[], mahasiswa_info=None, proposal_id=None, total_nilai_bantuan=0, total_laporan_kemajuan=0)
        
    try:
        cursor = app_funcs['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Ambil data mahasiswa yang dibimbing oleh pembimbing ini
        cursor.execute('''
            SELECT m.*, p.judul_usaha, p.tahapan_usaha, p.status as status_proposal, p.id as proposal_id
            FROM mahasiswa m 
            LEFT JOIN proposal p ON m.nim = p.nim
            WHERE m.id = %s AND p.dosen_pembimbing = %s
        ''', (mahasiswa_id, session['nama']))
        
        mahasiswa_info = cursor.fetchone()
        if not mahasiswa_info:
            flash('Data mahasiswa tidak ditemukan atau Anda tidak memiliki akses!', 'danger')
            cursor.close()
            return redirect(url_for('pembimbing.pembimbing_laporan_kemajuan'))
        
        # Tentukan jenis anggaran berdasarkan tahapan usaha
        tahapan_usaha = mahasiswa_info.get('tahapan_usaha', '').lower()
        if 'bertumbuh' in tahapan_usaha:
            tabel_laporan = 'laporan_kemajuan_bertumbuh'
        else:
            tabel_laporan = 'laporan_kemajuan_awal'
        
        proposal_id = mahasiswa_info.get('proposal_id')
        
        if not proposal_id:
            flash('Proposal tidak ditemukan!', 'danger')
            cursor.close()
            return render_template('pembimbing/laporan_kemajuan_awal_bertumbuh.html', 
                                 anggaran_data=[], mahasiswa_info=mahasiswa_info, proposal_id=None, total_nilai_bantuan=0, total_laporan_kemajuan=0)
        
        # Tentukan tabel anggaran berdasarkan tahapan usaha
        if 'bertumbuh' in tahapan_usaha:
            tabel_anggaran = 'anggaran_bertumbuh'
        else:
            tabel_anggaran = 'anggaran_awal'
        
        # Buat tabel laporan kemajuan jika belum ada
        if tabel_laporan == 'laporan_kemajuan_awal':
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS laporan_kemajuan_awal (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    id_proposal INT NOT NULL,
                    kegiatan_utama VARCHAR(255),
                    kegiatan VARCHAR(255) NOT NULL,
                    penanggung_jawab VARCHAR(100) NOT NULL,
                    target_capaian TEXT NOT NULL,
                    nama_barang VARCHAR(255) NOT NULL,
                    kuantitas INT NOT NULL,
                    satuan VARCHAR(50) NOT NULL,
                    harga_satuan DECIMAL(15,2) NOT NULL,
                    jumlah DECIMAL(15,2) NOT NULL,
                    keterangan TEXT,
                    nilai_bantuan DECIMAL(15,2) DEFAULT 0.00,
                    tanggal_buat TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    status ENUM('draf','diajukan','disetujui','ditolak','revisi') DEFAULT 'draf',
                    tanggal_review TIMESTAMP NULL,
                    FOREIGN KEY (id_proposal) REFERENCES proposal(id) ON DELETE CASCADE
                )
            ''')
        else:
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS laporan_kemajuan_bertumbuh (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    id_proposal INT NOT NULL,
                    kegiatan_utama VARCHAR(255) NOT NULL,
                    kegiatan VARCHAR(255) NOT NULL,
                    penanggung_jawab VARCHAR(100) NOT NULL,
                    target_capaian TEXT NOT NULL,
                    nama_barang VARCHAR(255) NOT NULL,
                    kuantitas INT NOT NULL,
                    satuan VARCHAR(50) NOT NULL,
                    harga_satuan DECIMAL(15,2) NOT NULL,
                    jumlah DECIMAL(15,2) NOT NULL,
                    keterangan TEXT,
                    nilai_bantuan DECIMAL(15,2) DEFAULT 0.00,
                    tanggal_buat TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    status ENUM('draf','diajukan','disetujui','ditolak','revisi') DEFAULT 'draf',
                    tanggal_review TIMESTAMP NULL,
                    FOREIGN KEY (id_proposal) REFERENCES proposal(id) ON DELETE CASCADE
                )
            ''')
        
        # Cek apakah sudah ada data laporan kemajuan untuk proposal ini
        cursor.execute(f'''
            SELECT COUNT(*) as count FROM {tabel_laporan} WHERE id_proposal = %s
        ''', (proposal_id,))
        existing_count = cursor.fetchone()['count']
        
        # Jika belum ada data laporan kemajuan, salin data dari anggaran yang disetujui
        if existing_count == 0:
            cursor.execute(f'''
                SELECT id, kegiatan_utama, kegiatan, penanggung_jawab, target_capaian, 
                       nama_barang, satuan, status, nilai_bantuan
                FROM {tabel_anggaran} 
                WHERE id_proposal = %s AND status = 'disetujui'
                ORDER BY kegiatan_utama, kegiatan, nama_barang
            ''', (proposal_id,))
            
            anggaran_data = cursor.fetchall()
            
            if anggaran_data:
                for anggaran in anggaran_data:
                    cursor.execute(f'''
                        INSERT INTO {tabel_laporan} (
                            id_proposal, kegiatan_utama, kegiatan, penanggung_jawab, target_capaian,
                            nama_barang, kuantitas, satuan, harga_satuan, jumlah, keterangan, status, nilai_bantuan
                        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    ''', (
                        proposal_id, anggaran['kegiatan_utama'], anggaran['kegiatan'], 
                        anggaran['penanggung_jawab'], anggaran['target_capaian'], anggaran['nama_barang'],
                        0, anggaran['satuan'], 0, 
                        0, '', 'draf', anggaran.get('nilai_bantuan', 0)
                    ))
                
                app_funcs['mysql'].connection.commit()
        
        # Debug: Log informasi untuk troubleshooting
        logger.debug(f"Debug: Proposal ID: {proposal_id}")
        logger.debug(f"Debug: Tabel laporan: {tabel_laporan}")
        logger.debug(f"Debug: Tabel anggaran: {tabel_anggaran}")
        logger.debug(f"Debug: Existing count: {existing_count}")
        
        # Ambil data laporan kemajuan
        cursor.execute(f'''
            SELECT id, kegiatan_utama, kegiatan, penanggung_jawab, target_capaian, 
                   nama_barang, kuantitas, satuan, harga_satuan, jumlah, keterangan, status, nilai_bantuan
            FROM {tabel_laporan} 
            WHERE id_proposal = %s
            ORDER BY kegiatan_utama, kegiatan, nama_barang
        ''', (proposal_id,))
        
        laporan_data = cursor.fetchall()
        cursor.close()
        
        # Urutkan data sesuai urutan kegiatan utama dengan normalisasi teks
        if 'bertumbuh' in tahapan_usaha:
            urutan_kegiatan = [
                "Pengembangan Pasar dan Saluran Distribusi",
                "Pengembangan Produk/Riset",
                "Produksi",
                "Legalitas, perizinan, sertifikasi, dan standarisasi",
                "Belanja ATK dan Penunjang"
            ]
        else:
            urutan_kegiatan = [
                "Pengembangan Produk/Riset",
                "Produksi",
                "Legalitas, perizinan, sertifikasi, dan standarisasi",
                "Belanja ATK dan Penunjang"
            ]
        
        # Buat mapping untuk case-insensitive matching dan normalisasi teks
        urutan_kegiatan_lower = [k.strip().lower() for k in urutan_kegiatan]
        
        # Fungsi helper untuk normalisasi teks kegiatan utama
        def normalize_kegiatan_utama(text):
            if not text:
                return ""
            # Normalisasi: lowercase, strip whitespace, replace multiple spaces
            normalized = text.strip().lower()
            # Handle berbagai variasi teks
            replacements = {
                "legalitas, perijinan, sertifikasi, pengujian produk, dan standarisasi": "legalitas, perizinan, sertifikasi, dan standarisasi",
                "legalitas, perijinan, sertifikasi, dan standarisasi": "legalitas, perizinan, sertifikasi, dan standarisasi",
                "legalitas perijinan sertifikasi pengujian produk dan standarisasi": "legalitas, perizinan, sertifikasi, dan standarisasi",
                "legalitas perijinan sertifikasi dan standarisasi": "legalitas, perizinan, sertifikasi, dan standarisasi"
            }
            for old, new in replacements.items():
                if old in normalized:
                    normalized = normalized.replace(old, new)
            return normalized
        
        # Sort data berdasarkan urutan kegiatan utama dengan normalisasi
        laporan_data = sorted(
            laporan_data,
            key=lambda x: urutan_kegiatan_lower.index(normalize_kegiatan_utama(x['kegiatan_utama'])) if x['kegiatan_utama'] and normalize_kegiatan_utama(x['kegiatan_utama']) in urutan_kegiatan_lower else 99
        )
        
        grouped_data = group_anggaran_data(laporan_data)
        laporan_data_flat = flatten_anggaran_data(laporan_data)
        
        # Hitung total nilai bantuan dari anggaran yang sudah disetujui
        cursor = app_funcs['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        cursor.execute(f'''
            SELECT SUM(nilai_bantuan) as total_nilai_bantuan
            FROM {tabel_anggaran} 
            WHERE id_proposal = %s AND status = 'disetujui'
        ''', (proposal_id,))
        
        nilai_bantuan_result = cursor.fetchone()
        total_nilai_bantuan = nilai_bantuan_result['total_nilai_bantuan'] or 0
        cursor.close()
        
        # Debug: Log total nilai bantuan
        logger.debug(f"Debug: Total nilai bantuan dari {tabel_anggaran}: {total_nilai_bantuan}")
        
        # Hitung total laporan kemajuan yang sudah disetujui
        cursor = app_funcs['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        cursor.execute(f'''
            SELECT SUM(jumlah) as total_laporan_kemajuan
            FROM {tabel_laporan} 
            WHERE id_proposal = %s AND status = 'disetujui'
        ''', (proposal_id,))
        
        laporan_kemajuan_result = cursor.fetchone()
        total_laporan_kemajuan = laporan_kemajuan_result['total_laporan_kemajuan'] or 0
        cursor.close()
        
        # Jika tidak ada laporan kemajuan yang disetujui, hitung dari semua data
        if total_laporan_kemajuan == 0:
            cursor = app_funcs['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
            cursor.execute(f'''
                SELECT SUM(jumlah) as total_laporan_kemajuan
                FROM {tabel_laporan} 
                WHERE id_proposal = %s
            ''', (proposal_id,))
            
            laporan_kemajuan_result = cursor.fetchone()
            total_laporan_kemajuan = laporan_kemajuan_result['total_laporan_kemajuan'] or 0
            cursor.close()
        
        # Debug: Log total laporan kemajuan
        logger.debug(f"Debug: Total laporan kemajuan dari {tabel_laporan}: {total_laporan_kemajuan}")
        
        # Debug: Print jumlah data yang ditemukan
        logger.debug(f"Debug: Found {len(laporan_data)} laporan kemajuan records for proposal {proposal_id}")
        
        # Ambil data file laporan kemajuan
        cursor = app_funcs['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        cursor.execute('''
            SELECT * FROM file_laporan_kemajuan 
            WHERE id_proposal = %s 
            ORDER BY tanggal_upload DESC 
            LIMIT 1
        ''', (proposal_id,))
        file_laporan = cursor.fetchone()
        cursor.close()
        
        if laporan_data:
            logger.debug(f"Debug: Sample data - Status: {laporan_data[0].get('status', 'N/A')}")
            logger.debug(f"Debug: Sample data - Kegiatan: {laporan_data[0].get('kegiatan_utama', 'N/A')}")
        
        return render_template('pembimbing/laporan_kemajuan_awal_bertumbuh.html', 
                             anggaran_data=laporan_data, 
                             grouped_data=grouped_data,
                             anggaran_data_flat=laporan_data_flat,
                             mahasiswa_info=mahasiswa_info,
                             tabel_laporan=tabel_laporan,
                             proposal_id=proposal_id,
                             total_nilai_bantuan=total_nilai_bantuan,
                             total_laporan_kemajuan=total_laporan_kemajuan,
                             file_laporan=file_laporan)
        
    except Exception as e:
        flash(f'Error saat mengambil data laporan kemajuan: {str(e)}', 'danger')
        return render_template('pembimbing/laporan_kemajuan_awal_bertumbuh.html', 
                             anggaran_data=[], 
                             grouped_data=[],
                             anggaran_data_flat=[],
                             mahasiswa_info=None,
                             tabel_laporan='',
                             proposal_id=None,
                             total_nilai_bantuan=0,
                             total_laporan_kemajuan=0,
                             file_laporan=None)


@pembimbing_bp.route('/pembimbing_laporan_akhir')
def pembimbing_laporan_akhir():
    if 'user_type' not in session or session['user_type'] != 'pembimbing':
        flash('Anda harus login sebagai pembimbing!', 'danger')
        return redirect(url_for('index'))
    
    if 'nama' not in session:
        flash('Data pembimbing tidak lengkap! Silakan login ulang.', 'danger')
        return redirect(url_for('index'))
    
    app_funcs = get_app_functions()
    
    # Log aktivitas pembimbing
    pembimbing_info = app_funcs['get_pembimbing_info_from_session']()
    if pembimbing_info:
        app_funcs['log_pembimbing_activity'](
            pembimbing_info['id'],
            pembimbing_info['nip'],
            pembimbing_info['nama'],
            'view',
            'laporan_akhir',
            'daftar_laporan_akhir',
            'Melihat halaman daftar laporan akhir mahasiswa'
        )
    if not hasattr(app_funcs['mysql'], 'connection') or app_funcs['mysql'].connection is None:
        flash('Koneksi ke database gagal. Cek konfigurasi database!', 'danger')
        return render_template('pembimbing/laporan_akhir.html', mahasiswa_list=[])
    
    try:
        cursor = app_funcs['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Ambil semua mahasiswa yang dibimbing oleh pembimbing ini beserta proposal dan tahapan_usaha
        cursor.execute('''
            SELECT m.id, m.nama_ketua, m.nim, p.judul_usaha, m.program_studi, p.id as proposal_id, p.tahapan_usaha
            FROM mahasiswa m
            INNER JOIN proposal p ON m.nim = p.nim
            WHERE p.dosen_pembimbing = %s AND p.status IN ('disetujui', 'selesai')
            ORDER BY m.nama_ketua
        ''', (session['nama'],))
        
        mahasiswa_all = cursor.fetchall()
        print(f"DEBUG LAPORAN AKHIR: Total mahasiswa yang dibimbing: {len(mahasiswa_all)}")
        print(f"DEBUG LAPORAN AKHIR: Session nama: {session.get('nama', 'N/A')}")
        print(f"DEBUG LAPORAN AKHIR: Session user_type: {session.get('user_type', 'N/A')}")
        
        if not mahasiswa_all:
            print("DEBUG LAPORAN AKHIR: Tidak ada mahasiswa yang dibimbing atau tidak ada proposal yang sesuai")
            
            # Coba query alternatif untuk debugging
            cursor.execute('''
                SELECT m.id, m.nama_ketua, m.nim, p.judul_usaha, m.program_studi, p.id as proposal_id, p.tahapan_usaha, p.status as proposal_status
                FROM mahasiswa m
                LEFT JOIN proposal p ON m.nim = p.nim
                WHERE p.dosen_pembimbing = %s
                ORDER BY m.nama_ketua
            ''', (session['nama'],))
            
            all_proposals = cursor.fetchall()
            print(f"DEBUG LAPORAN AKHIR: Total proposal dengan dosen_pembimbing = {session['nama']}: {len(all_proposals)}")
            for prop in all_proposals:
                print(f"DEBUG LAPORAN AKHIR: Proposal {prop['proposal_id']}: {prop['judul_usaha']} - Status: {prop.get('proposal_status', 'N/A')}")
            
            cursor.close()
            return render_template('pembimbing/laporan_akhir.html', mahasiswa_list=[])
        
        mahasiswa_list = []
        for mhs in mahasiswa_all:
            proposal_id = mhs['proposal_id']
            tahapan = (mhs.get('tahapan_usaha') or '').lower()
            
            print(f"DEBUG LAPORAN AKHIR: Processing mahasiswa {mhs['nama_ketua']} (NIM: {mhs['nim']})")
            print(f"DEBUG LAPORAN AKHIR: Proposal ID: {proposal_id}, Tahapan: {tahapan}")
            
            if 'bertumbuh' in tahapan:
                tabel_laporan_akhir = 'laporan_akhir_bertumbuh'
                tabel_laporan_kemajuan = 'laporan_kemajuan_bertumbuh'
            else:
                tabel_laporan_akhir = 'laporan_akhir_awal'
                tabel_laporan_kemajuan = 'laporan_kemajuan_awal'
            
            # TRIGGER: Cek apakah sudah ada data laporan akhir untuk proposal ini
            cursor.execute(f"SELECT COUNT(*) as cnt FROM {tabel_laporan_akhir} WHERE id_proposal = %s", (proposal_id,))
            cnt = cursor.fetchone()['cnt']
            
            # Jika sudah ada data laporan akhir, tambahkan ke list
            if cnt > 0:
                print(f"DEBUG: Data laporan akhir sudah ada untuk proposal {proposal_id}")
                mahasiswa_list.append(mhs)
                continue
            
            # Jika belum ada data laporan akhir, buat dari laporan kemajuan yang disetujui
                print(f"DEBUG: Membuat laporan akhir untuk proposal {proposal_id}")
                print(f"DEBUG: Menggunakan tabel laporan kemajuan: {tabel_laporan_kemajuan}")
                
                # Tambahkan kolom nilai_bantuan jika belum ada
                try:
                    cursor.execute(f'ALTER TABLE {tabel_laporan_akhir} ADD COLUMN nilai_bantuan DECIMAL(15,2) DEFAULT 0.00')
                    app_funcs['mysql'].connection.commit()
                    print(f"DEBUG: Berhasil menambahkan kolom nilai_bantuan ke tabel {tabel_laporan_akhir}")
                except Exception as e:
                    if "Duplicate column name" in str(e):
                        print(f"DEBUG: Kolom nilai_bantuan sudah ada di tabel {tabel_laporan_akhir}")
                    else:
                        print(f"DEBUG: Error menambahkan kolom nilai_bantuan: {str(e)}")
                
                # Tambahkan kolom rekomendasi_pendanaan jika belum ada
                try:
                    cursor.execute('ALTER TABLE penilaian_laporan_kemajuan ADD COLUMN rekomendasi_pendanaan VARCHAR(50) DEFAULT NULL')
                    app_funcs['mysql'].connection.commit()
                    print("DEBUG: Berhasil menambahkan kolom rekomendasi_pendanaan ke tabel penilaian_laporan_kemajuan")
                except Exception as e:
                    if "Duplicate column name" in str(e):
                        print("DEBUG: Kolom rekomendasi_pendanaan sudah ada di tabel penilaian_laporan_kemajuan")
                    else:
                        print(f"DEBUG: Error menambahkan kolom rekomendasi_pendanaan: {str(e)}")
                
                # Cek status rekomendasi pendanaan dari penilaian laporan kemajuan
                cursor.execute('''
                    SELECT rekomendasi_pendanaan 
                    FROM penilaian_laporan_kemajuan 
                    WHERE id_proposal = %s AND rekomendasi_pendanaan IS NOT NULL
                    ORDER BY id DESC 
                    LIMIT 1
                ''', (proposal_id,))
                
                penilaian_result = cursor.fetchone()
                rekomendasi_pendanaan = penilaian_result['rekomendasi_pendanaan'] if penilaian_result else None
                
                print(f"DEBUG: Rekomendasi pendanaan untuk proposal {proposal_id}: {rekomendasi_pendanaan}")
                
                # Jika rekomendasi adalah 'berhenti pendanaan', tidak buat laporan akhir
                if rekomendasi_pendanaan == 'berhenti pendanaan':
                    print(f"DEBUG: Rekomendasi pendanaan 'berhenti pendanaan' untuk proposal {proposal_id}. Laporan akhir tidak dibuat.")
                    continue
                
                # Jika rekomendasi adalah 'lanjutkan pendanaan' atau belum ada penilaian, cek laporan kemajuan yang disetujui
                if rekomendasi_pendanaan == 'lanjutkan pendanaan' or rekomendasi_pendanaan is None:
                    cursor.execute(f'''
                        SELECT COUNT(*) as cnt_kemajuan 
                        FROM {tabel_laporan_kemajuan} 
                        WHERE id_proposal = %s AND status = 'disetujui'
                    ''', (proposal_id,))
                    cnt_kemajuan = cursor.fetchone()['cnt_kemajuan']
                    print(f"DEBUG: Ditemukan {cnt_kemajuan} data laporan kemajuan dengan status 'disetujui'")
                    
                    if cnt_kemajuan > 0:
                        print(f"DEBUG: Menyalin data dari laporan kemajuan ke laporan akhir dengan reset field tertentu")
                        try:
                            cursor.execute(f'''
                                SELECT id, kegiatan_utama, kegiatan, penanggung_jawab, target_capaian, 
                                       nama_barang, satuan, status, nilai_bantuan
                                FROM {tabel_laporan_kemajuan} 
                                WHERE id_proposal = %s AND status = 'disetujui'
                                ORDER BY kegiatan_utama, kegiatan, nama_barang
                            ''', (proposal_id,))
                            kemajuan_data = cursor.fetchall()
                            print(f"DEBUG: Berhasil mengambil {len(kemajuan_data)} data laporan kemajuan")
                            
                            if kemajuan_data:
                                print(f"DEBUG: Sample data - kegiatan_utama: {kemajuan_data[0].get('kegiatan_utama', 'N/A')}")
                                print(f"DEBUG: Sample data - nama_barang: {kemajuan_data[0].get('nama_barang', 'N/A')}")
                                
                                for kemajuan in kemajuan_data:
                                    try:
                                        cursor.execute(f'''
                                            INSERT INTO {tabel_laporan_akhir} (
                                                id_proposal, kegiatan_utama, kegiatan, penanggung_jawab, target_capaian,
                                                nama_barang, kuantitas, satuan, harga_satuan, jumlah, keterangan, status, nilai_bantuan
                                            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                                        ''', (
                                            proposal_id, kemajuan['kegiatan_utama'], kemajuan['kegiatan'], 
                                            kemajuan['penanggung_jawab'], kemajuan['target_capaian'], kemajuan['nama_barang'],
                                            0, kemajuan['satuan'], 0, 0, '', 'draf', kemajuan.get('nilai_bantuan', 0)
                                        ))
                                        print(f"DEBUG: Berhasil INSERT data untuk kegiatan: {kemajuan.get('kegiatan_utama', 'N/A')}")
                                    except Exception as insert_error:
                                        print(f"DEBUG: Error saat INSERT: {str(insert_error)}")
                                        print(f"DEBUG: Data yang gagal INSERT: {kemajuan}")
                                        raise insert_error
                                
                                app_funcs['mysql'].connection.commit()
                                print(f"DEBUG: Berhasil membuat {len(kemajuan_data)} data laporan akhir untuk proposal {proposal_id}")
                                print(f"DEBUG: Field yang disalin: kegiatan_utama, kegiatan, penanggung_jawab, target_capaian, nama_barang, satuan, nilai_bantuan")
                                print(f"DEBUG: Field yang di-reset: kuantitas=0, harga_satuan=0, jumlah=0, keterangan=''")
                                
                                # Tambahkan mahasiswa ke list setelah berhasil membuat laporan akhir
                                mahasiswa_list.append(mhs)
                            else:
                                print(f"DEBUG: Tidak ada data laporan kemajuan yang disetujui untuk proposal {proposal_id}")
                        except Exception as e:
                            print(f"DEBUG: Error dalam proses pembuatan laporan akhir: {str(e)}")
                            app_funcs['mysql'].connection.rollback()
                else:
                    print(f"DEBUG: Tidak ada laporan kemajuan dengan status 'disetujui' untuk proposal {proposal_id}")
            else:
                print(f"DEBUG: Rekomendasi pendanaan bukan 'lanjutkan', tidak membuat laporan akhir")
        
        # Tambahkan mahasiswa ke list jika belum ada (untuk memastikan semua mahasiswa yang dibimbing ditampilkan)
        if mhs not in mahasiswa_list:
            print(f"DEBUG: Menambahkan mahasiswa {mhs['nama_ketua']} ke list meskipun belum ada laporan akhir")
            mahasiswa_list.append(mhs)
        
        cursor.close()
        
        print(f"DEBUG LAPORAN AKHIR: Final mahasiswa_list length: {len(mahasiswa_list)}")
        print(f"DEBUG LAPORAN AKHIR: Mahasiswa yang akan ditampilkan: {[m['nama_ketua'] for m in mahasiswa_list]}")
        
        return render_template('pembimbing/laporan_akhir.html', mahasiswa_list=mahasiswa_list)
        
    except Exception as e:
        flash(f'Error: {str(e)}', 'danger')
        return render_template('pembimbing/laporan_akhir.html', mahasiswa_list=[])

@pembimbing_bp.route('/laporan_akhir_awal_bertumbuh/<int:mahasiswa_id>')
def pembimbing_laporan_akhir_awal_bertumbuh(mahasiswa_id):
    if 'user_type' not in session or session['user_type'] != 'pembimbing':
        flash('Anda harus login sebagai pembimbing!', 'danger')
        return redirect(url_for('index'))
    
    app_funcs = get_app_functions()
    if not hasattr(app_funcs['mysql'], 'connection') or app_funcs['mysql'].connection is None:
        flash('Koneksi ke database gagal. Cek konfigurasi database!', 'danger')
        return render_template('pembimbing/laporan_akhir_awal_bertumbuh.html', anggaran_data=[], mahasiswa_info=None, proposal_id=None, total_anggaran_disetujui=0, total_nilai_bantuan=0, total_laporan_kemajuan_disetujui=0)
    
    try:
        cursor = app_funcs['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Ambil data mahasiswa dan proposal
        cursor.execute('''
            SELECT m.*, p.judul_usaha, p.tahapan_usaha, p.status as status_proposal, p.id as proposal_id
            FROM mahasiswa m 
            LEFT JOIN proposal p ON m.nim = p.nim
            WHERE m.id = %s AND p.dosen_pembimbing = %s
        ''', (mahasiswa_id, session['nama']))
        
        mahasiswa_info = cursor.fetchone()
        if not mahasiswa_info:
            flash('Data mahasiswa tidak ditemukan atau Anda tidak memiliki akses!', 'danger')
            cursor.close()
            return redirect(url_for('pembimbing.pembimbing_laporan_akhir'))
        
        # Tentukan jenis laporan berdasarkan tahapan usaha
        tahapan_usaha = mahasiswa_info.get('tahapan_usaha', '').lower()
        if 'bertumbuh' in tahapan_usaha:
            tabel_laporan_akhir = 'laporan_akhir_bertumbuh'
        else:
            tabel_laporan_akhir = 'laporan_akhir_awal'
        
        proposal_id = mahasiswa_info.get('proposal_id')
        
        if not proposal_id:
            flash('Proposal tidak ditemukan!', 'danger')
            cursor.close()
            return render_template('pembimbing/laporan_akhir_awal_bertumbuh.html', 
                                 anggaran_data=[], mahasiswa_info=mahasiswa_info, proposal_id=None, total_anggaran_disetujui=0, total_nilai_bantuan=0, total_laporan_kemajuan_disetujui=0)
        
        
        
        # Tentukan tabel laporan kemajuan
        if 'bertumbuh' in tahapan_usaha:
            tabel_laporan_kemajuan = 'laporan_kemajuan_bertumbuh'
        else:
            tabel_laporan_kemajuan = 'laporan_kemajuan_awal'
        
        # Buat tabel laporan akhir jika belum ada
        if tabel_laporan_akhir == 'laporan_akhir_awal':
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS laporan_akhir_awal (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    id_proposal INT NOT NULL,
                    kegiatan_utama VARCHAR(255),
                    kegiatan VARCHAR(255) NOT NULL,
                    penanggung_jawab VARCHAR(100) NOT NULL,
                    target_capaian TEXT NOT NULL,
                    nama_barang VARCHAR(255) NOT NULL,
                    kuantitas INT NOT NULL,
                    satuan VARCHAR(50) NOT NULL,
                    harga_satuan DECIMAL(15,2) NOT NULL,
                    jumlah DECIMAL(15,2) NOT NULL,
                    keterangan TEXT,
                    nilai_bantuan DECIMAL(15,2) DEFAULT 0.00,
                    tanggal_buat TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    status ENUM('draf','diajukan','disetujui','ditolak','revisi') DEFAULT 'draf',
                    tanggal_review TIMESTAMP NULL,
                    FOREIGN KEY (id_proposal) REFERENCES proposal(id) ON DELETE CASCADE
                )
            ''')
        else:
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS laporan_akhir_bertumbuh (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    id_proposal INT NOT NULL,
                    kegiatan_utama VARCHAR(255) NOT NULL,
                    kegiatan VARCHAR(255) NOT NULL,
                    penanggung_jawab VARCHAR(100) NOT NULL,
                    target_capaian TEXT NOT NULL,
                    nama_barang VARCHAR(255) NOT NULL,
                    kuantitas INT NOT NULL,
                    satuan VARCHAR(50) NOT NULL,
                    harga_satuan DECIMAL(15,2) NOT NULL,
                    jumlah DECIMAL(15,2) NOT NULL,
                    keterangan TEXT,
                    nilai_bantuan DECIMAL(15,2) DEFAULT 0.00,
                    tanggal_buat TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    status ENUM('draf','diajukan','disetujui','ditolak','revisi') DEFAULT 'draf',
                    tanggal_review TIMESTAMP NULL,
                    FOREIGN KEY (id_proposal) REFERENCES proposal(id) ON DELETE CASCADE
                )
            ''')
       
        # Ambil data laporan akhir
        cursor.execute(f'''
            SELECT id, kegiatan_utama, kegiatan, penanggung_jawab, target_capaian, 
                   nama_barang, kuantitas, satuan, harga_satuan, jumlah, keterangan, status, nilai_bantuan
            FROM {tabel_laporan_akhir} 
            WHERE id_proposal = %s
            ORDER BY kegiatan_utama, kegiatan, nama_barang
        ''', (proposal_id,))
        
        laporan_data = cursor.fetchall()
        print(f"DEBUG: Data laporan akhir untuk proposal {proposal_id}: {len(laporan_data)} rows")
        
        # Debug: Log detail data laporan akhir
        for i, row in enumerate(laporan_data):
            print(f"DEBUG: Row {i+1}: kegiatan_utama='{row.get('kegiatan_utama', 'N/A')}', kegiatan='{row.get('kegiatan', 'N/A')}', nama_barang='{row.get('nama_barang', 'N/A')}', jumlah={row.get('jumlah', 'N/A')}, status='{row.get('status', 'N/A')}")
        
        # Urutkan data sesuai urutan kegiatan utama dengan normalisasi teks
        if 'bertumbuh' in tahapan_usaha:
            urutan_kegiatan = [
                "Pengembangan Pasar dan Saluran Distribusi",
                "Pengembangan Produk/Riset",
                "Produksi",
                "Legalitas, perizinan, sertifikasi, dan standarisasi",
                "Belanja ATK dan Penunjang"
            ]
        else:
            urutan_kegiatan = [
                "Pengembangan Produk/Riset",
                "Produksi",
                "Legalitas, perizinan, sertifikasi, dan standarisasi",
                "Belanja ATK dan Penunjang"
            ]
        
        # Buat mapping untuk case-insensitive matching dan normalisasi teks
        urutan_kegiatan_lower = [k.strip().lower() for k in urutan_kegiatan]
        
        # Fungsi helper untuk normalisasi teks kegiatan utama
        def normalize_kegiatan_utama(text):
            if not text:
                return ""
            # Normalisasi: lowercase, strip whitespace, replace multiple spaces
            normalized = text.strip().lower()
            # Handle berbagai variasi teks
            replacements = {
                "legalitas, perijinan, sertifikasi, pengujian produk, dan standarisasi": "legalitas, perizinan, sertifikasi, dan standarisasi",
                "legalitas, perijinan, sertifikasi, dan standarisasi": "legalitas, perizinan, sertifikasi, dan standarisasi",
                "legalitas perijinan sertifikasi pengujian produk dan standarisasi": "legalitas, perizinan, sertifikasi, dan standarisasi",
                "legalitas perijinan sertifikasi dan standarisasi": "legalitas, perizinan, sertifikasi, dan standarisasi"
            }
            for old, new in replacements.items():
                if old in normalized:
                    normalized = normalized.replace(old, new)
            return normalized
        
        # Sort data berdasarkan urutan kegiatan utama dengan normalisasi
        laporan_data = sorted(
            laporan_data,
            key=lambda x: urutan_kegiatan_lower.index(normalize_kegiatan_utama(x['kegiatan_utama'])) if x['kegiatan_utama'] and normalize_kegiatan_utama(x['kegiatan_utama']) in urutan_kegiatan_lower else 99
        )
        
        grouped_data = group_anggaran_data(laporan_data)
        laporan_data_flat = flatten_anggaran_data(laporan_data)
        
        print(f"DEBUG: Data berhasil di-flatten menjadi {len(laporan_data_flat)} baris")
        
        # Debug: Log detail data yang akan dikirim ke template
        print("=== DEBUG DETAIL DATA LAPORAN AKHIR FLAT ===")
        for i, row in enumerate(laporan_data_flat):
            print(f"Row {i+1}: kegiatan_utama='{row.get('kegiatan_utama', 'N/A')}', kegiatan='{row.get('kegiatan', 'N/A')}', nama_barang='{row.get('nama_barang', 'N/A')}', jumlah={row.get('jumlah', 'N/A')}, status='{row.get('status', 'N/A')}")
        print("============================================")
        
        # Hitung total nilai bantuan dari anggaran yang sudah direview
        if 'bertumbuh' in tahapan_usaha:
            tabel_anggaran = 'anggaran_bertumbuh'
        else:
            tabel_anggaran = 'anggaran_awal'
        
        print(f"DEBUG: Menggunakan tabel anggaran: {tabel_anggaran}")
        
        cursor.execute(f'''
            SELECT SUM(nilai_bantuan) as total_nilai_bantuan
            FROM {tabel_anggaran} 
            WHERE id_proposal = %s AND status_reviewer = 'sudah_direview'
        ''', (proposal_id,))
        
        nilai_bantuan_result = cursor.fetchone()
        total_nilai_bantuan = nilai_bantuan_result['total_nilai_bantuan'] or 0
        print(f"DEBUG: Total nilai bantuan dari {tabel_anggaran}: {total_nilai_bantuan}")
        
        # Hitung total anggaran yang disetujui
        cursor.execute(f'''
            SELECT SUM(jumlah) as total_anggaran
            FROM {tabel_anggaran} 
            WHERE id_proposal = %s AND status = 'disetujui'
        ''', (proposal_id,))
        
        anggaran_result = cursor.fetchone()
        total_anggaran_disetujui = anggaran_result['total_anggaran'] or 0
        print(f"DEBUG: Total anggaran disetujui dari {tabel_anggaran}: {total_anggaran_disetujui}")
        
        # Hitung total laporan kemajuan yang disetujui
        cursor.execute(f'''
            SELECT SUM(jumlah) as total_laporan_kemajuan
            FROM {tabel_laporan_kemajuan} 
            WHERE id_proposal = %s AND status = 'disetujui'
        ''', (proposal_id,))
        
        laporan_kemajuan_result = cursor.fetchone()
        total_laporan_kemajuan_disetujui = laporan_kemajuan_result['total_laporan_kemajuan'] or 0
        print(f"DEBUG: Total laporan kemajuan disetujui dari {tabel_laporan_kemajuan}: {total_laporan_kemajuan_disetujui}")
        
        # Hitung total laporan akhir (jumlah dari laporan_data)
        total_laporan_akhir = sum(row.get('jumlah', 0) for row in laporan_data) if laporan_data else 0
        print(f"DEBUG: Total laporan akhir (dihitung dari laporan_data): {total_laporan_akhir}")
        
        # Summary log
        print(f"=== SUMMARY LAPORAN AKHIR PROPOSAL {proposal_id} ===")
        print(f"Total Nilai Bantuan: {total_nilai_bantuan}")
        print(f"Total Anggaran Disetujui: {total_anggaran_disetujui}")
        print(f"Total Laporan Kemajuan Disetujui: {total_laporan_kemajuan_disetujui}")
        print(f"Total Laporan Akhir: {total_laporan_akhir}")
        print(f"Jumlah Data Laporan: {len(laporan_data)}")
        print(f"Jumlah Data Flat: {len(laporan_data_flat)}")
        print(f"================================================")
        
        # Ambil data file laporan akhir
        cursor.execute('''
            SELECT * FROM file_laporan_akhir 
            WHERE id_proposal = %s
        ''', (proposal_id,))
        
        file_laporan_akhir = cursor.fetchone()
        
        cursor.close()
        
        return render_template('pembimbing/laporan_akhir_awal_bertumbuh.html', 
                             anggaran_data=laporan_data, 
                             grouped_data=grouped_data,
                             anggaran_data_flat=laporan_data_flat,
                             mahasiswa_info=mahasiswa_info,
                             tabel_laporan=tabel_laporan_akhir,
                             proposal_id=proposal_id,
                             total_anggaran_disetujui=total_anggaran_disetujui,
                             total_nilai_bantuan=total_nilai_bantuan,
                             total_laporan_kemajuan_disetujui=total_laporan_kemajuan_disetujui,
                             total_laporan_akhir=total_laporan_akhir,
                             file_laporan_akhir=file_laporan_akhir)
        
    except Exception as e:
        flash(f'Error saat mengambil data laporan akhir: {str(e)}', 'danger')
        return render_template('pembimbing/laporan_akhir_awal_bertumbuh.html', 
                             anggaran_data=[], 
                             mahasiswa_info=None,
                             proposal_id=None,
                             total_anggaran_disetujui=0,
                             total_nilai_bantuan=0,
                             total_laporan_kemajuan_disetujui=0,
                             file_laporan_akhir=None)

@pembimbing_bp.route('/download_laporan_kemajuan/<int:proposal_id>')
def pembimbing_download_laporan_kemajuan(proposal_id):
    if 'user_type' not in session or session['user_type'] != 'pembimbing':
        flash('Anda harus login sebagai pembimbing!', 'danger')
        return redirect(url_for('pembimbing.pembimbing_proposal'))
    
    app_funcs = get_app_functions()
    if not hasattr(app_funcs['mysql'], 'connection') or app_funcs['mysql'].connection is None:
        flash('Koneksi ke database gagal. Cek konfigurasi database!', 'danger')
        return redirect(url_for('pembimbing.pembimbing_proposal'))
    
    try:
        cursor = app_funcs['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Ambil data proposal dan laporan kemajuan
        cursor.execute('''
            SELECT p.*, m.nama_ketua 
            FROM proposal p 
            LEFT JOIN mahasiswa m ON p.nim = m.nim
            WHERE p.id = %s AND p.dosen_pembimbing = %s
        ''', (proposal_id, session['nama']))
        
        proposal = cursor.fetchone()
        cursor.close()
        
        if not proposal:
            flash('Proposal tidak ditemukan atau Anda tidak memiliki akses!', 'danger')
            return redirect(url_for('pembimbing.pembimbing_proposal'))
        
        # Ambil path file laporan kemajuan dari struktur file, bukan dari kolom DB yang tidak ada
        # Gunakan helper yang sudah ada untuk menentukan lokasi file berdasarkan proposal
        # Jika tidak ditemukan, tampilkan pesan yang sama
        file_path = proposal.get('laporan_kemajuan_path')
        if not file_path:
            flash('Laporan kemajuan belum diupload oleh mahasiswa!', 'danger')
            return redirect(url_for('pembimbing.pembimbing_proposal'))
        file_path = file_path.replace('\\', '/').replace('\\', '/')
        if not file_path.startswith('static/'):
            file_path = 'static/' + file_path.lstrip('/')
        if not os.path.exists(file_path):
            flash(f'File laporan kemajuan tidak ditemukan! ({file_path})', 'danger')
            return redirect(url_for('pembimbing.pembimbing_proposal'))
        ext = os.path.splitext(file_path)[1].lower()
        if ext == '.pdf':
            return send_file(file_path, as_attachment=False, mimetype='application/pdf')
        else:
            import mimetypes
            mimetype = mimetypes.guess_type(file_path)[0] or 'application/octet-stream'
            return send_file(file_path, as_attachment=False, mimetype=mimetype)
    except Exception as e:
        flash(f'Error saat download laporan kemajuan: {str(e)}', 'danger')
        return redirect(url_for('pembimbing.pembimbing_proposal'))


@pembimbing_bp.route('/update_file_laporan_kemajuan_status', methods=['POST'])
def pembimbing_update_file_laporan_kemajuan_status():
    if 'user_type' not in session or session['user_type'] != 'pembimbing':
        return jsonify({'success': False, 'message': 'Anda harus login sebagai pembimbing!'})
    
    try:
        app_funcs = get_app_functions()
        cursor = app_funcs['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        proposal_id = request.form.get('proposal_id')
        status = request.form.get('status')
        komentar = request.form.get('komentar', '').strip()
        
        if not proposal_id or not status:
            return jsonify({'success': False, 'message': 'Data tidak lengkap!'})
        
        # Validasi status
        if status not in ['draf', 'diajukan', 'disetujui', 'revisi']:
            return jsonify({'success': False, 'message': 'Status tidak valid!'})
        
        # Cek apakah file laporan kemajuan ada
        cursor.execute('''
            SELECT * FROM file_laporan_kemajuan 
            WHERE id_proposal = %s
        ''', (proposal_id,))
        
        file_data = cursor.fetchone()
        if not file_data:
            cursor.close()
            return jsonify({'success': False, 'message': 'File laporan kemajuan tidak ditemukan!'})
        
        # Update status dan komentar - hapus komentar jika status disetujui
        if status == 'disetujui':
            cursor.execute('''
                UPDATE file_laporan_kemajuan 
                SET status = %s, komentar_pembimbing = NULL, tanggal_update = CURRENT_TIMESTAMP
                WHERE id_proposal = %s
            ''', (status, proposal_id))
        else:
            cursor.execute('''
                UPDATE file_laporan_kemajuan 
                SET status = %s, komentar_pembimbing = %s, tanggal_update = CURRENT_TIMESTAMP
                WHERE id_proposal = %s
            ''', (status, komentar, proposal_id))
        
        app_funcs['mysql'].connection.commit()
        cursor.close()
        
        return jsonify({
            'success': True, 
            'message': f'Status file laporan kemajuan berhasil diupdate menjadi {status}!',
            'status': status
        })
        
    except Exception as e:
        print(f"Error update file laporan kemajuan status: {str(e)}")
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})


@pembimbing_bp.route('/view_file_laporan_kemajuan/<int:proposal_id>')
def pembimbing_view_file_laporan_kemajuan(proposal_id):
    if 'user_type' not in session or session['user_type'] != 'pembimbing':
        flash('Anda harus login sebagai pembimbing!', 'danger')
        return redirect(url_for('pembimbing.pembimbing_laporan_kemajuan'))
    
    try:
        app_funcs = get_app_functions()
        cursor = app_funcs['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        cursor.execute('''
            SELECT nama_file, file_path FROM file_laporan_kemajuan 
            WHERE id_proposal = %s
        ''', (proposal_id,))
        
        file_data = cursor.fetchone()
        cursor.close()
        
        if not file_data:
            flash('File laporan kemajuan tidak ditemukan!', 'danger')
            return redirect(url_for('pembimbing.pembimbing_laporan_kemajuan'))
        
        file_path = file_data['file_path']
        if not os.path.exists(file_path):
            flash('File tidak ditemukan di server!', 'danger')
            return redirect(url_for('pembimbing.pembimbing_laporan_kemajuan'))
        
        # Tampilkan file PDF di browser
        return send_file(file_path, as_attachment=False, mimetype='application/pdf')
        
    except Exception as e:
        flash(f'Error saat menampilkan file: {str(e)}', 'danger')
        return redirect(url_for('pembimbing.pembimbing_laporan_kemajuan'))

@pembimbing_bp.route('/update_file_laporan_akhir_status', methods=['POST'])
def pembimbing_update_file_laporan_akhir_status():
    if 'user_type' not in session or session['user_type'] != 'pembimbing':
        return jsonify({'success': False, 'message': 'Anda harus login sebagai pembimbing!'})
    
    try:
        app_funcs = get_app_functions()
        cursor = app_funcs['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        data = request.get_json()
        proposal_id = data.get('proposal_id')
        status = data.get('status')
        komentar = data.get('komentar', '').strip()
        
        if not proposal_id or not status:
            return jsonify({'success': False, 'message': 'Data tidak lengkap!'})
        
        # Validasi status
        if status not in ['draf', 'diajukan', 'disetujui', 'revisi']:
            return jsonify({'success': False, 'message': 'Status tidak valid!'})
        
        # Cek apakah file laporan akhir ada
        cursor.execute('''
            SELECT * FROM file_laporan_akhir 
            WHERE id_proposal = %s
        ''', (proposal_id,))
        
        file_data = cursor.fetchone()
        if not file_data:
            cursor.close()
            return jsonify({'success': False, 'message': 'File laporan akhir tidak ditemukan!'})
        
        # Update status dan komentar - hapus komentar jika status disetujui
        if status == 'disetujui':
            cursor.execute('''
                UPDATE file_laporan_akhir 
                SET status = %s, komentar_pembimbing = NULL, tanggal_update = CURRENT_TIMESTAMP
                WHERE id_proposal = %s
            ''', (status, proposal_id))
        else:
            cursor.execute('''
                UPDATE file_laporan_akhir 
                SET status = %s, komentar_pembimbing = %s, tanggal_update = CURRENT_TIMESTAMP
                WHERE id_proposal = %s
            ''', (status, komentar, proposal_id))
        
        app_funcs['mysql'].connection.commit()
        cursor.close()
        
        return jsonify({
            'success': True, 
            'message': f'Status file laporan akhir berhasil diupdate menjadi {status}!',
            'status': status
        })
        
    except Exception as e:
        print(f"Error update file laporan akhir status: {str(e)}")
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@pembimbing_bp.route('/pembimbing_view_file_laporan_akhir/<int:proposal_id>')
def pembimbing_view_file_laporan_akhir(proposal_id):
    if 'user_type' not in session or session['user_type'] != 'pembimbing':
        flash('Anda harus login sebagai pembimbing!', 'danger')
        return redirect(url_for('pembimbing.pembimbing_laporan_akhir'))
    
    try:
        app_funcs = get_app_functions()
        cursor = app_funcs['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        cursor.execute('''
            SELECT nama_file, file_path FROM file_laporan_akhir 
            WHERE id_proposal = %s
        ''', (proposal_id,))
        
        file_data = cursor.fetchone()
        cursor.close()
        
        if not file_data:
            flash('File laporan akhir tidak ditemukan!', 'danger')
            return redirect(url_for('pembimbing.pembimbing_laporan_akhir'))
        
        file_path = file_data['file_path']
        if not os.path.exists(file_path):
            flash('File tidak ditemukan di server!', 'danger')
            return redirect(url_for('pembimbing.pembimbing_laporan_akhir'))
        
        # Tampilkan file PDF di browser
        return send_file(file_path, as_attachment=False, mimetype='application/pdf')
        
    except Exception as e:
        flash(f'Error saat menampilkan file: {str(e)}', 'danger')
        return redirect(url_for('pembimbing.pembimbing_laporan_akhir'))


@pembimbing_bp.route('/konfirmasi_proposal', methods=['POST'])
def pembimbing_konfirmasi_proposal():
    if 'user_type' not in session or session['user_type'] != 'pembimbing':
        flash('Anda harus login sebagai pembimbing!', 'danger')
        return redirect(url_for('pembimbing.pembimbing_proposal'))
    
    app_funcs = get_app_functions()
    if not hasattr(app_funcs['mysql'], 'connection') or app_funcs['mysql'].connection is None:
        flash('Koneksi ke database gagal. Cek konfigurasi database!', 'danger')
        return redirect(url_for('pembimbing.pembimbing_proposal'))
    
    try:
        cursor = app_funcs['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Ambil data dari form
        proposal_id = request.form.get('proposal_id')
        action = request.form.get('action')
        catatan = request.form.get('catatan', '').strip()
        
        if not proposal_id or not action:
            flash('Data tidak lengkap!', 'danger')
            return redirect(url_for('pembimbing.pembimbing_proposal'))
        
        # Validasi aksi
        if action not in ['setuju', 'tolak', 'revisi']:
            flash('Aksi tidak valid!', 'danger')
            return redirect(url_for('pembimbing.pembimbing_proposal'))
        
        # Cek apakah proposal ada dan dibimbing oleh pembimbing ini
        cursor.execute('''
            SELECT p.*, m.nama_ketua 
            FROM proposal p 
            LEFT JOIN mahasiswa m ON p.nim = m.nim
            WHERE p.id = %s AND p.dosen_pembimbing = %s
        ''', (proposal_id, session['nama']))
        
        proposal = cursor.fetchone()
        
        if not proposal:
            flash('Proposal tidak ditemukan atau Anda tidak memiliki akses!', 'danger')
            cursor.close()
            return redirect(url_for('pembimbing.pembimbing_proposal'))
        
        # Update status proposal berdasarkan aksi
        if action == 'setuju':
            status = 'disetujui'
        elif action == 'tolak':
            status = 'ditolak'
        elif action == 'revisi':
            status = 'revisi'
        
        # Update database
        if action == 'revisi':
            cursor.execute('''
                UPDATE proposal 
                SET status = %s, komentar_revisi = %s, 
                    tanggal_konfirmasi_pembimbing = CURRENT_TIMESTAMP
                WHERE id = %s
            ''', (status, catatan, proposal_id))
        else:
            cursor.execute('''
                UPDATE proposal 
                SET status = %s, catatan_pembimbing = %s, 
                    tanggal_konfirmasi_pembimbing = CURRENT_TIMESTAMP
                WHERE id = %s
            ''', (status, catatan, proposal_id))
        
        app_funcs['mysql'].connection.commit()
        
        # Log aktivitas pembimbing
        pembimbing_info = app_funcs['get_pembimbing_info_from_session']()
        if pembimbing_info and proposal:
            jenis_aktivitas = 'setuju' if action == 'setuju' else 'tolak' if action == 'tolak' else 'revisi'
            deskripsi = f'Mengkonfirmasi proposal "{proposal["judul"]}" dengan status {status}'
            if catatan:
                deskripsi += f' dengan catatan: {catatan[:100]}...' if len(catatan) > 100 else f' dengan catatan: {catatan}'
            
            app_funcs['log_pembimbing_activity'](
                pembimbing_info['id'],
                pembimbing_info['nip'],
                pembimbing_info['nama'],
                jenis_aktivitas,
                'proposal',
                f'proposal_konfirmasi_id_{proposal_id}',
                deskripsi,
                {'status': proposal.get('status', 'pending')},
                {'status': status, 'catatan_pembimbing': catatan},
                proposal.get('mahasiswa_id'),
                proposal_id
            )
        
        cursor.close()
        
        flash(f'Proposal berhasil dikonfirmasi! Status: {status}', 'success')
        return redirect(url_for('pembimbing.pembimbing_proposal'))
        
    except Exception as e:
        flash(f'Error saat konfirmasi proposal: {str(e)}', 'danger')
        return redirect(url_for('pembimbing.pembimbing_proposal'))


@pembimbing_bp.route('/konfirmasi_laporan_akhir', methods=['POST'])
def pembimbing_konfirmasi_laporan_akhir():
    if 'user_type' not in session or session['user_type'] != 'pembimbing':
        return jsonify({'success': False, 'message': 'Anda harus login sebagai pembimbing!'})
    
    app_funcs = get_app_functions()
    if not hasattr(app_funcs['mysql'], 'connection') or app_funcs['mysql'].connection is None:
        return jsonify({'success': False, 'message': 'Koneksi ke database gagal!'})
    
    try:
        data = request.get_json()
        laporan_id = data.get('id')
        tabel_laporan = data.get('tabel_laporan')
        action = data.get('action')  # 'setuju', 'tolak', 'revisi'
        
        if not laporan_id or not tabel_laporan or not action:
            return jsonify({'success': False, 'message': 'Data tidak lengkap!'})
        
        if tabel_laporan not in ['laporan_akhir_awal', 'laporan_akhir_bertumbuh']:
            return jsonify({'success': False, 'message': 'Tabel laporan tidak valid!'})
        
        if action not in ['setuju', 'tolak', 'revisi']:
            return jsonify({'success': False, 'message': 'Aksi tidak valid!'})
        
        # Tentukan status berdasarkan action
        status_map = {
            'setuju': 'disetujui',
            'tolak': 'ditolak',
            'revisi': 'revisi'
        }
        new_status = status_map[action]
        
        cursor = app_funcs['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Cek apakah laporan akhir milik mahasiswa yang dibimbing oleh pembimbing ini
        cursor.execute(f'''
            SELECT la.*, p.dosen_pembimbing
            FROM {tabel_laporan} la
            JOIN proposal p ON la.id_proposal = p.id
            WHERE la.id = %s AND p.dosen_pembimbing = %s
        ''', (laporan_id, session['nama']))
        
        laporan = cursor.fetchone()
        if not laporan:
            cursor.close()
            return jsonify({'success': False, 'message': 'Laporan akhir tidak ditemukan atau Anda tidak memiliki akses!'})
        
        # Ambil data lengkap untuk logging
        cursor.execute(f'''
            SELECT la.*, p.judul_usaha, m.id as mahasiswa_id
            FROM {tabel_laporan} la
            JOIN proposal p ON la.id_proposal = p.id
            JOIN mahasiswa m ON p.nim = m.nim
            WHERE la.id = %s
        ''', (laporan_id,))
        laporan_data = cursor.fetchone()
        
        # Update status laporan akhir
        cursor.execute(f'''
            UPDATE {tabel_laporan} 
            SET status = %s, tanggal_review = NOW()
            WHERE id = %s
        ''', (new_status, laporan_id))
        
        app_funcs['mysql'].connection.commit()
        
        # Log aktivitas pembimbing
        pembimbing_info = app_funcs['get_pembimbing_info_from_session']()
        if pembimbing_info and laporan_data:
            jenis_aktivitas = 'setuju' if action == 'setuju' else 'tolak' if action == 'tolak' else 'revisi'
            modul = 'laporan_akhir_awal' if 'awal' in tabel_laporan else 'laporan_akhir_bertumbuh'
            deskripsi = f'Mengubah status {modul} ID {laporan_id} menjadi {new_status} untuk proposal "{laporan_data["judul_usaha"]}"'
            
            app_funcs['log_pembimbing_activity'](
                pembimbing_info['id'],
                pembimbing_info['nip'],
                pembimbing_info['nama'],
                jenis_aktivitas,
                modul,
                f'{modul}_id_{laporan_id}',
                deskripsi,
                {'status': laporan_data.get('status', 'pending')},
                {'status': new_status},
                laporan_data.get('mahasiswa_id'),
                laporan_data.get('id_proposal')
            )
        
        cursor.close()
        
        return jsonify({'success': True, 'message': f'Status laporan akhir berhasil diubah menjadi {new_status}'})
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@pembimbing_bp.route('/get_pembimbing_list')
def pembimbing_get_pembimbing_list():
    if 'user_type' not in session or session['user_type'] != 'pembimbing':
        return jsonify({'success': False, 'message': 'Anda harus login sebagai pembimbing!'})
    app_funcs = get_app_functions()
    if not hasattr(app_funcs['mysql'], 'connection') or app_funcs['mysql'].connection is None:
        return jsonify({'success': False, 'message': 'Koneksi ke database gagal!'})
    try:
        cursor = app_funcs['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        cursor.execute('''
            SELECT p.nama, p.nip, p.program_studi, p.kuota_mahasiswa,
                   COUNT(DISTINCT CASE WHEN pr.status != 'selesai' THEN pr.nim END) as jumlah_mahasiswa_bimbing
            FROM pembimbing p
            LEFT JOIN proposal pr ON p.nama = pr.dosen_pembimbing
            WHERE p.status = 'aktif'
            GROUP BY p.id, p.nama, p.nip, p.program_studi, p.kuota_mahasiswa
            ORDER BY p.nama
        ''')
        pembimbing_list = cursor.fetchall()
        for pembimbing in pembimbing_list:
            pembimbing['sisa_kuota'] = pembimbing['kuota_mahasiswa'] - pembimbing['jumlah_mahasiswa_bimbing']
            pembimbing['status_kuota'] = 'tersedia' if pembimbing['sisa_kuota'] > 0 else 'penuh'
        cursor.close()
        return jsonify({'success': True, 'pembimbing': pembimbing_list})
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error saat mengambil data pembimbing: {str(e)}'})


    
    try:
        cursor = app_funcs['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Ambil data mahasiswa yang dibimbing oleh pembimbing ini
        cursor.execute('''
            SELECT m.*, p.judul_usaha, p.id as proposal_id, p.tahapan_usaha
            FROM mahasiswa m
            JOIN proposal p ON m.nim = p.nim
            WHERE m.id = %s AND p.dosen_pembimbing = %s
        ''', (mahasiswa_id, session['nama']))
        
        mahasiswa_info = cursor.fetchone()
        if not mahasiswa_info:
            flash('Data mahasiswa tidak ditemukan atau Anda tidak memiliki akses!', 'danger')
            cursor.close()
            return redirect(url_for('pembimbing.pembimbing_monitoring_mahasiswa_laporan_neraca'))
        
        proposal_id = mahasiswa_info['proposal_id']
        
        # Ambil data neraca dari tabel neraca
        cursor.execute('''
            SELECT * FROM neraca 
            WHERE proposal_id = %s 
            ORDER BY tanggal_neraca DESC 
            LIMIT 1
        ''', (proposal_id,))
        
        neraca_data = cursor.fetchone()
        
        if not neraca_data:
            # Jika tidak ada data neraca, buat data kosong dengan format yang benar
            neraca_data = {
                'kas_dan_setara_kas': 0,
                'piutang_usaha': 0,
                'persediaan': 0,
                'beban_dibayar_dimuka': 0,
                'total_aset_lancar': 0,
                'tanah': 0,
                'bangunan': 0,
                'kendaraan': 0,
                'peralatan_dan_mesin': 0,
                'akumulasi_penyusutan': 0,
                'total_aset_tetap_bersih': 0,
                'investasi_jangka_panjang': 0,
                'hak_paten_merek_dagang': 0,
                'total_aset_lain_lain': 0,
                'total_aset': 0,
                'utang_usaha': 0,
                'beban_harus_dibayar': 0,
                'utang_pajak': 0,
                'utang_jangka_pendek_lainnya': 0,
                'total_liabilitas_jangka_pendek': 0,
                'utang_bank': 0,
                'total_liabilitas_jangka_panjang': 0,
                'total_liabilitas': 0,
                'modal_disetor': 0,
                'laba_ditahan': 0,
                'total_ekuitas': 0,
                'total_liabilitas_dan_ekuitas': 0,
                'tanggal_neraca': datetime.now().strftime('%Y-%m-%d')
            }
        
        cursor.close()
        
        return render_template('pembimbing/laporan_neraca.html', 
                             mahasiswa_info=mahasiswa_info,
                             neraca_data=neraca_data)
        
    except Exception as e:
        flash(f'Error: {str(e)}', 'danger')
        return render_template('pembimbing/laporan_neraca.html', 
                             mahasiswa_info=None,
                             neraca_data={})

@pembimbing_bp.route('/monitoring_mahasiswa/laporan_neraca')
def pembimbing_monitoring_mahasiswa_laporan_neraca():
    if 'user_type' not in session or session['user_type'] != 'pembimbing':
        flash('Anda harus login sebagai pembimbing!', 'danger')
        return redirect(url_for('index'))
    
    app_funcs = get_app_functions()
    if not hasattr(app_funcs['mysql'], 'connection') or app_funcs['mysql'].connection is None:
        flash('Koneksi ke database gagal. Cek konfigurasi database!', 'danger')
        return render_template('pembimbing/daftar_laporan_neraca.html', mahasiswa_list=[])
    
    try:
        cursor = app_funcs['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Ambil semua mahasiswa yang dibimbing oleh pembimbing ini
        # Coba dengan exact match terlebih dahulu
        cursor.execute('''
            SELECT m.id as mahasiswa_id, m.nama_ketua, m.nim, p.judul_usaha, p.tahapan_usaha, m.program_studi, p.id as proposal_id
            FROM mahasiswa m
            JOIN proposal p ON m.nim = p.nim
            WHERE p.dosen_pembimbing = %s
            ORDER BY m.nama_ketua
        ''', (session['nama'],))
        
        mahasiswa_all = cursor.fetchall()
        
        # Jika tidak ada hasil, coba dengan LIKE match
        if not mahasiswa_all:
            cursor.execute('''
                SELECT m.id as mahasiswa_id, m.nama_ketua, m.nim, p.judul_usaha, p.tahapan_usaha, m.program_studi, p.id as proposal_id
                FROM mahasiswa m
                JOIN proposal p ON m.nim = p.nim
                WHERE p.dosen_pembimbing LIKE %s
                ORDER BY m.nama_ketua
            ''', (f"%{session['nama']}%",))
            mahasiswa_all = cursor.fetchall()
        
        # Jika masih tidak ada, ambil semua mahasiswa yang memiliki dosen pembimbing
        if not mahasiswa_all:
            cursor.execute('''
                SELECT m.id as mahasiswa_id, m.nama_ketua, m.nim, p.judul_usaha, p.tahapan_usaha, m.program_studi, p.id as proposal_id
                FROM mahasiswa m
                JOIN proposal p ON m.nim = p.nim
                WHERE p.dosen_pembimbing IS NOT NULL AND p.dosen_pembimbing != ''
                ORDER BY m.nama_ketua
            ''')
            mahasiswa_all = cursor.fetchall()
        mahasiswa_list = []
        
        for mhs in mahasiswa_all:
            proposal_id = mhs['proposal_id']
            
            # Cek apakah ada data neraca
            try:
                cursor.execute('SELECT COUNT(*) as cnt FROM neraca WHERE proposal_id = %s', (proposal_id,))
                neraca_count = cursor.fetchone()['cnt']
            except:
                # Jika tabel tidak ada, cek dari data lain
                cursor.execute('SELECT COUNT(*) as cnt FROM penjualan WHERE proposal_id = %s', (proposal_id,))
                penjualan_count = cursor.fetchone()['cnt']
                cursor.execute('SELECT COUNT(*) as cnt FROM biaya_operasional WHERE proposal_id = %s', (proposal_id,))
                biaya_operasional_count = cursor.fetchone()['cnt']
                neraca_count = 1 if (penjualan_count > 0 or biaya_operasional_count > 0) else 0
            
            mhs['has_neraca'] = neraca_count > 0
            mahasiswa_list.append(mhs)
        
        cursor.close()
        
        return render_template('pembimbing/daftar_laporan_neraca.html', mahasiswa_list=mahasiswa_list)
        
    except Exception as e:
        flash(f'Error: {str(e)}', 'danger')
        return render_template('pembimbing/daftar_laporan_neraca.html', mahasiswa_list=[])


    
    try:
        cursor = app_funcs['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Ambil data mahasiswa yang dibimbing oleh pembimbing ini
        cursor.execute('''
            SELECT m.*, p.judul_usaha, p.id as proposal_id, p.tahapan_usaha
            FROM mahasiswa m
            JOIN proposal p ON m.nim = p.nim
            WHERE m.id = %s AND p.dosen_pembimbing = %s
        ''', (mahasiswa_id, session['nama']))
        
        mahasiswa_info = cursor.fetchone()
        if not mahasiswa_info:
            flash('Data mahasiswa tidak ditemukan atau Anda tidak memiliki akses!', 'danger')
            cursor.close()
            return redirect(url_for('pembimbing.pembimbing_monitoring_mahasiswa_laporan_neraca'))
        
        proposal_id = mahasiswa_info['proposal_id']
        
        # Hitung laporan neraca secara real-time dari data transaksi
        neraca_data = app_funcs['hitung_neraca_real_time'](proposal_id, cursor)
        
        cursor.close()
        
        return render_template('pembimbing/laporan_neraca.html', 
                             mahasiswa_info=mahasiswa_info,
                             neraca_data=neraca_data)
        
    except Exception as e:
        flash(f'Error: {str(e)}', 'danger')
        return render_template('pembimbing/laporan_neraca.html', 
                             mahasiswa_info=None,
                             neraca_data={})

@pembimbing_bp.route('/laporan_neraca/<int:mahasiswa_id>')
def pembimbing_laporan_neraca(mahasiswa_id):
    if 'user_type' not in session or session['user_type'] != 'pembimbing':
        flash('Anda harus login sebagai pembimbing!', 'danger')
        return redirect(url_for('index'))
    
    app_funcs = get_app_functions()
    if not hasattr(app_funcs['mysql'], 'connection') or app_funcs['mysql'].connection is None:
        flash('Koneksi ke database gagal. Cek konfigurasi database!', 'danger')
        return render_template('pembimbing/laporan_neraca.html', mahasiswa_info=None, neraca_data={})
    
    try:
        cursor = app_funcs['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Ambil data mahasiswa yang dibimbing oleh pembimbing ini
        cursor.execute('''
            SELECT m.*, p.judul_usaha, p.id as proposal_id, p.tahapan_usaha
            FROM mahasiswa m
            JOIN proposal p ON m.nim = p.nim
            WHERE m.id = %s AND p.dosen_pembimbing = %s
        ''', (mahasiswa_id, session['nama']))
        
        mahasiswa_info = cursor.fetchone()
        if not mahasiswa_info:
            flash('Data mahasiswa tidak ditemukan atau Anda tidak memiliki akses!', 'danger')
            cursor.close()
            return redirect(url_for('pembimbing.pembimbing_monitoring_mahasiswa_laporan_neraca'))
        
        proposal_id = mahasiswa_info['proposal_id']
        
        # Hitung laporan neraca secara real-time dari data transaksi
        neraca_data = app_funcs['hitung_neraca_real_time'](proposal_id, cursor)
        
        cursor.close()
        
        return render_template('pembimbing/laporan_neraca.html', 
                             mahasiswa_info=mahasiswa_info,
                             neraca_data=neraca_data)
        
    except Exception as e:
        flash(f'Error: {str(e)}', 'danger')
        return render_template('pembimbing/laporan_neraca.html', 
                             mahasiswa_info=None,
                             neraca_data={})

@pembimbing_bp.route('/download_neraca', methods=['POST'])
def download_neraca():
    if 'user_type' not in session or session['user_type'] != 'pembimbing':
        flash('Anda harus login sebagai pembimbing!', 'danger')
        return redirect(url_for('index'))
    
    app_funcs = get_app_functions()
    if not hasattr(app_funcs['mysql'], 'connection') or app_funcs['mysql'].connection is None:
        flash('Koneksi ke database gagal. Cek konfigurasi database!', 'danger')
        return redirect(url_for('pembimbing.pembimbing_monitoring_mahasiswa_laporan_neraca'))
    
    try:
        format_type = request.form.get('format', 'excel')
        proposal_id = request.form.get('proposal_id')
        
        if not proposal_id:
            flash('ID proposal tidak ditemukan!', 'danger')
            return redirect(url_for('pembimbing.pembimbing_monitoring_mahasiswa_laporan_neraca'))
        
        cursor = app_funcs['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Ambil data mahasiswa
        cursor.execute('''
            SELECT m.*, p.judul_usaha, p.id as proposal_id, p.tahapan_usaha
            FROM mahasiswa m
            JOIN proposal p ON m.nim = p.nim
            WHERE p.id = %s AND p.dosen_pembimbing = %s
        ''', (proposal_id, session['nama']))
        
        mahasiswa_info = cursor.fetchone()
        if not mahasiswa_info:
            flash('Data mahasiswa tidak ditemukan atau Anda tidak memiliki akses!', 'danger')
            cursor.close()
            return redirect(url_for('pembimbing.pembimbing_monitoring_mahasiswa_laporan_neraca'))
        
        # Hitung laporan neraca secara real-time
        neraca_data = app_funcs['hitung_neraca_real_time'](proposal_id, cursor)
        
        cursor.close()
        
        # Generate file berdasarkan format
        if format_type == 'excel':
            file_buffer = app_funcs['generate_excel_neraca'](mahasiswa_info, neraca_data)
            if file_buffer:
                filename = f"Laporan_Neraca_{mahasiswa_info['nama_ketua']}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                return send_file(
                    file_buffer,
                    as_attachment=True,
                    download_name=filename,
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
            else:
                flash('Gagal membuat file Excel!', 'danger')
                return redirect(url_for('pembimbing.pembimbing_monitoring_mahasiswa_laporan_neraca'))
        elif format_type == 'pdf':
            file_buffer = app_funcs['generate_pdf_neraca'](mahasiswa_info, neraca_data)
            if file_buffer:
                filename = f"Laporan_Neraca_{mahasiswa_info['nama_ketua']}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
                return send_file(
                    file_buffer,
                    as_attachment=True,
                    download_name=filename,
                    mimetype='application/pdf'
                )
            else:
                flash('Gagal membuat file PDF!', 'danger')
                return redirect(url_for('pembimbing.pembimbing_monitoring_mahasiswa_laporan_neraca'))
        elif format_type == 'word':
            file_buffer = app_funcs['generate_word_neraca'](mahasiswa_info, neraca_data)
            if file_buffer:
                filename = f"Laporan_Neraca_{mahasiswa_info['nama_ketua']}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.docx"
                return send_file(
                    file_buffer,
                    as_attachment=True,
                    download_name=filename,
                    mimetype='application/vnd.openxmlformats-officedocument.wordprocessingml.document'
                )
            else:
                flash('Gagal membuat file Word!', 'danger')
                return redirect(url_for('pembimbing.pembimbing_monitoring_mahasiswa_laporan_neraca'))
        else:
            flash('Format file tidak valid!', 'danger')
            return redirect(url_for('pembimbing.pembimbing_monitoring_mahasiswa_laporan_neraca'))
        
    except Exception as e:
        flash(f'Error saat mengunduh file: {str(e)}', 'danger')
        return redirect(url_for('pembimbing.pembimbing_monitoring_mahasiswa_laporan_neraca'))

# ========================================
# ROUTE UNTUK PENILAIAN MAHASISWA
# ========================================

@pembimbing_bp.route('/daftar_penilaian_mahasiswa')
def daftar_penilaian_mahasiswa():
    """Halaman daftar mahasiswa yang bisa dinilai"""
    if 'user_type' not in session or session['user_type'] != 'pembimbing':
        flash('Anda harus login sebagai pembimbing!', 'danger')
        return redirect(url_for('index'))
    
    app_funcs = get_app_functions()
    if not hasattr(app_funcs['mysql'], 'connection') or app_funcs['mysql'].connection is None:
        flash('Koneksi ke database gagal. Cek konfigurasi database!', 'danger')
        return render_template('pembimbing/daftar_penilaian_mahasiswa.html')
    
    try:
        # Log aktivitas pembimbing
        pembimbing_info = app_funcs['get_pembimbing_info_from_session']()
        if pembimbing_info:
            app_funcs['log_pembimbing_activity'](
                pembimbing_info['id'],
                pembimbing_info['nip'],
                pembimbing_info['nama'],
                'view',
                'penilaian',
                'daftar_penilaian_mahasiswa',
                'Melihat daftar mahasiswa untuk penilaian'
            )
        
        # Tambahkan status jadwal penilaian pembimbing
        jadwal_bimbingan = get_jadwal_pembimbing_status()
        
        return render_template('pembimbing/daftar_penilaian_mahasiswa.html',
                             jadwal_bimbingan=jadwal_bimbingan)
        
    except Exception as e:
        logger.error(f"Error in daftar_penilaian_mahasiswa: {str(e)}")
        flash('Terjadi kesalahan saat memuat halaman!', 'danger')
        return render_template('pembimbing/daftar_penilaian_mahasiswa.html')

@pembimbing_bp.route('/penilaian_mahasiswa')
def penilaian_mahasiswa():
    """Halaman form penilaian mahasiswa"""
    if 'user_type' not in session or session['user_type'] != 'pembimbing':
        flash('Anda harus login sebagai pembimbing!', 'danger')
        return redirect(url_for('index'))
    
    app_funcs = get_app_functions()
    if not hasattr(app_funcs['mysql'], 'connection') or app_funcs['mysql'].connection is None:
        flash('Koneksi ke database gagal. Cek konfigurasi database!', 'danger')
        # Auto-simpan nilai 0 untuk pembimbing jika melewati jadwal selesai pembimbing_nilai_selesai dan belum ada penilaian
        try:
            cur = app_funcs['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
            cur.execute('SELECT pembimbing_nilai_mulai FROM pengaturan_jadwal ORDER BY id DESC LIMIT 1')
            row = cur.fetchone()
            if row and row.get('pembimbing_nilai_selesai'):
                from datetime import datetime
                if datetime.now() > row['pembimbing_nilai_selesai']:
                    # Ambil daftar proposal yang dibimbing oleh pembimbing ini dan belum ada penilaian
                    cur.execute('SELECT id FROM pembimbing WHERE nama=%s', (session['nama'],))
                    pemb = cur.fetchone()
                    if pemb:
                        id_pembimbing = pemb['id']
                        cur.execute('''
                            SELECT p.id as id_proposal
                            FROM proposal p
                            WHERE p.dosen_pembimbing=%s AND NOT EXISTS (
                                SELECT 1 FROM penilaian_mahasiswa pm WHERE pm.id_proposal=p.id AND pm.id_pembimbing=%s
                            )
                        ''', (session['nama'], id_pembimbing))
                        rows = cur.fetchall()
                        # Pertanyaan aktif
                        cur.execute('SELECT id, bobot, skor_maksimal FROM pertanyaan_penilaian_mahasiswa WHERE status="Aktif" ORDER BY created_at ASC')
                        pts = cur.fetchall()
                        for r in rows:
                            # insert header 0
                            cur.execute('''
                                INSERT INTO penilaian_mahasiswa (id_proposal, id_pembimbing, nilai_akhir, komentar_pembimbing, status)
                                VALUES (%s,%s,%s,%s,'Selesai')
                            ''', (r['id_proposal'], id_pembimbing, 0, ''))
                            pen_id = cur.lastrowid
                            for p in pts:
                                cur.execute('''
                                    INSERT INTO detail_penilaian_mahasiswa (id_penilaian_mahasiswa, id_pertanyaan, skor, nilai)
                                    VALUES (%s,%s,%s,%s)
                                ''', (pen_id, p['id'], 0, 0))
                        app_funcs['mysql'].connection.commit()
            cur.close()
        except Exception:
            try:
                cur.close()
            except Exception:
                pass

        # Cek status jadwal penilaian pembimbing
        jadwal_bimbingan = get_jadwal_pembimbing_status()
        if not jadwal_bimbingan['bisa_nilai']:
            if jadwal_bimbingan['status'] == 'sudah_selesai':
                flash(f"Tidak bisa melakukan penilaian: {jadwal_bimbingan['pesan']}", 'warning')
            else:
                flash(f"Tidak bisa melakukan penilaian: {jadwal_bimbingan['pesan']}", 'info')
            return redirect(url_for('pembimbing.daftar_penilaian_mahasiswa'))
        
        return render_template('pembimbing/penilaian_mahasiswa.html')
    
    try:
        # Log aktivitas pembimbing
        pembimbing_info = app_funcs['get_pembimbing_info_from_session']()
        if pembimbing_info:
            app_funcs['log_pembimbing_activity'](
                pembimbing_info['id'],
                pembimbing_info['nip'],
                pembimbing_info['nama'],
                'view',
                'penilaian',
                'form_penilaian_mahasiswa',
                'Melihat form penilaian mahasiswa'
            )
        
        # Cek status jadwal penilaian pembimbing
        jadwal_bimbingan = get_jadwal_pembimbing_status()
        if not jadwal_bimbingan['bisa_nilai']:
            if jadwal_bimbingan['status'] == 'sudah_selesai':
                flash(f"Tidak bisa melakukan penilaian: {jadwal_bimbingan['pesan']}", 'warning')
            else:
                flash(f"Tidak bisa melakukan penilaian: {jadwal_bimbingan['pesan']}", 'info')
            return redirect(url_for('pembimbing.daftar_penilaian_mahasiswa'))
        
        return render_template('pembimbing/penilaian_mahasiswa.html')
        
    except Exception as e:
        logger.error(f"Error in penilaian_mahasiswa: {str(e)}")
        flash('Terjadi kesalahan saat memuat halaman!', 'danger')
        return render_template('pembimbing/penilaian_mahasiswa.html')

def update_nilai_akhir_penilaian(cursor, id_penilaian_mahasiswa):
    """
    Helper function untuk mengupdate nilai_akhir di tabel penilaian_mahasiswa
    """
    try:
        # PERBAIKAN: Ambil pertanyaan berdasarkan data yang ada (aktif + nonaktif)
        # Ini memastikan perhitungan nilai konsisten meskipun pertanyaan dinonaktifkan
        cursor.execute('''
            SELECT DISTINCT p.id, p.bobot, p.skor_maksimal
            FROM pertanyaan_penilaian_mahasiswa p
            INNER JOIN detail_penilaian_mahasiswa d ON p.id = d.id_pertanyaan
            WHERE d.id_penilaian_mahasiswa = %s
            ORDER BY p.created_at ASC
        ''', (id_penilaian_mahasiswa,))
        pertanyaan_list = cursor.fetchall()
        
        if not pertanyaan_list:
            return
        
        # Ambil skor per sesi
        cursor.execute('''
            SELECT id_pertanyaan, sesi_penilaian, skor
            FROM detail_penilaian_mahasiswa 
            WHERE id_penilaian_mahasiswa = %s
            ORDER BY id_pertanyaan, sesi_penilaian
        ''', (id_penilaian_mahasiswa,))
        
        skor_data = cursor.fetchall()
        
        if not skor_data:
            # Jika tidak ada skor, set nilai_akhir = 0
            cursor.execute('''
                UPDATE penilaian_mahasiswa 
                SET nilai_akhir = 0.00 
                WHERE id = %s
            ''', (id_penilaian_mahasiswa,))
            return
        
        # Organize skor per sesi
        skor_per_sesi = {}
        for skor in skor_data:
            pertanyaan_id = str(skor['id_pertanyaan'])
            sesi = skor['sesi_penilaian']
            
            if pertanyaan_id not in skor_per_sesi:
                skor_per_sesi[pertanyaan_id] = {}
            
            skor_per_sesi[pertanyaan_id][sesi] = skor['skor']
        
        # Hitung nilai akhir
        nilai_akhir = 0
        logger.info(f"DEBUG PERHITUNGAN: Mulai hitung nilai akhir untuk penilaian_id {id_penilaian_mahasiswa}")
        logger.info(f"DEBUG PERHITUNGAN: Total pertanyaan = {len(pertanyaan_list)}")
        logger.info(f"DEBUG PERHITUNGAN: Skor per sesi = {skor_per_sesi}")
        
        for pertanyaan in pertanyaan_list:
            pertanyaan_id = str(pertanyaan['id'])
            if pertanyaan_id in skor_per_sesi:
                # Hitung rata-rata skor untuk pertanyaan ini
                sesi_keys = list(skor_per_sesi[pertanyaan_id].keys())
                if sesi_keys:
                    total_skor = sum(skor_per_sesi[pertanyaan_id][sesi] for sesi in sesi_keys)
                    rata_rata_skor = total_skor / len(sesi_keys)
                    skor_maksimal = float(pertanyaan['skor_maksimal'])
                    bobot = float(pertanyaan['bobot'])
                    
                    # Hitung nilai untuk pertanyaan ini
                    nilai_pertanyaan = (rata_rata_skor / skor_maksimal) * bobot
                    nilai_akhir += nilai_pertanyaan
                    
                    # Debug log untuk setiap pertanyaan
                    logger.info(f"DEBUG PERHITUNGAN: Pertanyaan {pertanyaan_id} - Skor per sesi: {[skor_per_sesi[pertanyaan_id][sesi] for sesi in sesi_keys]}")
                    logger.info(f"DEBUG PERHITUNGAN: Pertanyaan {pertanyaan_id} - Total skor: {total_skor}, Jumlah sesi: {len(sesi_keys)}, Rata-rata: {rata_rata_skor}")
                    logger.info(f"DEBUG PERHITUNGAN: Pertanyaan {pertanyaan_id} - Skor maks: {skor_maksimal}, Bobot: {bobot}, Nilai: {nilai_pertanyaan}")
            else:
                logger.warning(f"DEBUG PERHITUNGAN: Pertanyaan {pertanyaan_id} tidak ada di skor_per_sesi")
        
        logger.info(f"DEBUG PERHITUNGAN: Total nilai akhir = {nilai_akhir}")
        
        # Update nilai_akhir di database
        cursor.execute('''
            UPDATE penilaian_mahasiswa 
            SET nilai_akhir = %s 
            WHERE id = %s
        ''', (round(nilai_akhir, 2), id_penilaian_mahasiswa))
        
        logger.info(f"SUCCESS: Updated nilai_akhir to {round(nilai_akhir, 2)} for penilaian_id {id_penilaian_mahasiswa}")
        
    except Exception as e:
        logger.error(f"Error in update_nilai_akhir_penilaian: {str(e)}")

def get_jadwal_penilaian_mahasiswa(cursor, id_penilaian_mahasiswa):
    """
    ✅ PERBAIKAN: Mendapatkan jadwal yang sesuai dengan penilaian mahasiswa
    - Jika ada jadwal_id, gunakan jadwal tersebut
    - Jika tidak ada jadwal_id, gunakan jadwal terbaru
    """
    try:
        if not id_penilaian_mahasiswa:
            # Jika belum ada penilaian, gunakan jadwal terbaru
            return get_jadwal_terbaru_status(cursor)
        
        # Ambil jadwal_id dari penilaian mahasiswa
        cursor.execute('''
            SELECT jadwal_id 
            FROM penilaian_mahasiswa 
            WHERE id = %s
        ''', (id_penilaian_mahasiswa,))
        
        result = cursor.fetchone()
        if not result or not result['jadwal_id']:
            # Jika tidak ada jadwal_id, gunakan jadwal terbaru
            return get_jadwal_terbaru_status(cursor)
        
        jadwal_id = result['jadwal_id']
        
        # Ambil jadwal berdasarkan jadwal_id
        cursor.execute('''
            SELECT 
                pembimbing_nilai_mulai, pembimbing_nilai_selesai,
                pembimbing_interval_tipe, pembimbing_interval_nilai,
                pembimbing_jam_mulai, pembimbing_jam_selesai
            FROM pengaturan_jadwal 
            WHERE id = %s
        ''', (jadwal_id,))
        
        jadwal = cursor.fetchone()
        if not jadwal:
            # Jika jadwal tidak ditemukan, gunakan jadwal terbaru
            return get_jadwal_terbaru_status(cursor)
        
        # Hitung status jadwal
        now = datetime.now()
        mulai = jadwal['pembimbing_nilai_mulai']
        selesai = jadwal['pembimbing_nilai_selesai']
        
        if not mulai or not selesai:
            return {
                'status': 'tidak ada jadwal',
                'pesan': 'Jadwal penilaian tidak valid',
                'bisa_nilai': False,
                'jadwal_id': jadwal_id
            }
        
        # Konversi ke datetime jika string
        if isinstance(mulai, str):
            mulai = datetime.strptime(mulai, '%Y-%m-%d %H:%M:%S')
        if isinstance(selesai, str):
            selesai = datetime.strptime(selesai, '%Y-%m-%d %H:%M:%S')
        
        if now < mulai:
            return {
                'status': 'belum_mulai',
                'pesan': 'Jadwal penilaian belum dimulai',
                'bisa_nilai': False,
                'jadwal_mulai': mulai,
                'jadwal_selesai': selesai,
                'jadwal_id': jadwal_id,
                'interval_tipe': jadwal['pembimbing_interval_tipe'],
                'interval_nilai': jadwal['pembimbing_interval_nilai'],
                'jam_mulai': jadwal['pembimbing_jam_mulai'],
                'jam_selesai': jadwal['pembimbing_jam_selesai']
            }
        elif now > selesai:
            return {
                'status': 'sudah_selesai',
                'pesan': 'Jadwal penilaian sudah berakhir',
                'bisa_nilai': False,
                'jadwal_mulai': mulai,
                'jadwal_selesai': selesai,
                'jadwal_id': jadwal_id,
                'interval_tipe': jadwal['pembimbing_interval_tipe'],
                'interval_nilai': jadwal['pembimbing_interval_nilai'],
                'jam_mulai': jadwal['pembimbing_jam_mulai'],
                'jam_selesai': jadwal['pembimbing_jam_selesai']
            }
        else:
            return {
                'status': 'aktif',
                'pesan': 'Jadwal penilaian sedang berlangsung',
                'bisa_nilai': True,
                'jadwal_mulai': mulai,
                'jadwal_selesai': selesai,
                'jadwal_id': jadwal_id,
                'interval_tipe': jadwal['pembimbing_interval_tipe'],
                'interval_nilai': jadwal['pembimbing_interval_nilai'],
                'jam_mulai': jadwal['pembimbing_jam_mulai'],
                'jam_selesai': jadwal['pembimbing_jam_selesai']
            }
            
    except Exception as e:
        logger.error(f"Error in get_jadwal_penilaian_mahasiswa: {str(e)}")
        return get_jadwal_terbaru_status(cursor)

def get_jadwal_terbaru_status(cursor):
    """Mendapatkan status jadwal terbaru (untuk mahasiswa baru)"""
    try:
        cursor.execute('''
            SELECT 
                pembimbing_nilai_mulai, pembimbing_nilai_selesai,
                pembimbing_interval_tipe, pembimbing_interval_nilai,
                pembimbing_jam_mulai, pembimbing_jam_selesai
            FROM pengaturan_jadwal 
            ORDER BY id DESC 
            LIMIT 1
        ''')
        
        jadwal = cursor.fetchone()
        if not jadwal:
            return {
                'status': 'tidak ada jadwal',
                'pesan': 'Belum ada jadwal penilaian yang diatur',
                'bisa_nilai': False
            }
        
        now = datetime.now()
        mulai = jadwal['pembimbing_nilai_mulai']
        selesai = jadwal['pembimbing_nilai_selesai']
        
        if not mulai or not selesai:
            return {
                'status': 'tidak ada jadwal',
                'pesan': 'Jadwal penilaian belum diatur',
                'bisa_nilai': False
            }
        
        # Konversi ke datetime jika string
        if isinstance(mulai, str):
            mulai = datetime.strptime(mulai, '%Y-%m-%d %H:%M:%S')
        if isinstance(selesai, str):
            selesai = datetime.strptime(selesai, '%Y-%m-%d %H:%M:%S')
        
        if now < mulai:
            return {
                'status': 'belum_mulai',
                'pesan': 'Jadwal penilaian belum dimulai',
                'bisa_nilai': False,
                'jadwal_mulai': mulai,
                'jadwal_selesai': selesai,
                'interval_tipe': jadwal['pembimbing_interval_tipe'],
                'interval_nilai': jadwal['pembimbing_interval_nilai'],
                'jam_mulai': jadwal['pembimbing_jam_mulai'],
                'jam_selesai': jadwal['pembimbing_jam_selesai']
            }
        elif now > selesai:
            return {
                'status': 'sudah_selesai',
                'pesan': 'Jadwal penilaian sudah berakhir',
                'bisa_nilai': False,
                'jadwal_mulai': mulai,
                'jadwal_selesai': selesai,
                'interval_tipe': jadwal['pembimbing_interval_tipe'],
                'interval_nilai': jadwal['pembimbing_interval_nilai'],
                'jam_mulai': jadwal['pembimbing_jam_mulai'],
                'jam_selesai': jadwal['pembimbing_jam_selesai']
            }
        else:
            return {
                'status': 'aktif',
                'pesan': 'Jadwal penilaian sedang berlangsung',
                'bisa_nilai': True,
                'jadwal_mulai': mulai,
                'jadwal_selesai': selesai,
                'interval_tipe': jadwal['pembimbing_interval_tipe'],
                'interval_nilai': jadwal['pembimbing_interval_nilai'],
                'jam_mulai': jadwal['pembimbing_jam_mulai'],
                'jam_selesai': jadwal['pembimbing_jam_selesai']
            }
            
    except Exception as e:
        logger.error(f"Error in get_jadwal_terbaru_status: {str(e)}")
        return {'status': 'error', 'pesan': str(e), 'bisa_nilai': False}

def hitung_nilai_penilaian_mahasiswa(cursor, id_penilaian_mahasiswa):
    """
    Helper function untuk menghitung nilai akhir penilaian mahasiswa
    PERBAIKAN: Ambil pertanyaan berdasarkan data historis per mahasiswa
    """
    try:
        if not id_penilaian_mahasiswa:
            return 0, 0, 'Belum Dinilai'  # nilai, total_sesi, status
        
        # Ambil pertanyaan berdasarkan data yang ada (aktif + nonaktif)
        cursor.execute('''
            SELECT DISTINCT p.id, p.bobot, p.skor_maksimal
            FROM pertanyaan_penilaian_mahasiswa p
            INNER JOIN detail_penilaian_mahasiswa d ON p.id = d.id_pertanyaan
            WHERE d.id_penilaian_mahasiswa = %s
            ORDER BY p.created_at ASC
        ''', (id_penilaian_mahasiswa,))
        
        pertanyaan_list = cursor.fetchall()
        
        if not pertanyaan_list:
            return 0, 0, 'Belum Dinilai'  # nilai, total_sesi, status
        
        # Ambil skor per sesi
        cursor.execute('''
            SELECT id_pertanyaan, sesi_penilaian, skor, nilai, tanggal_input, is_locked
            FROM detail_penilaian_mahasiswa 
            WHERE id_penilaian_mahasiswa = %s
            ORDER BY id_pertanyaan, sesi_penilaian
        ''', (id_penilaian_mahasiswa,))
        
        skor_data = cursor.fetchall()
        
        if not skor_data:
            return 0, 0, 'Belum Dinilai'
        
        # Organize skor per sesi
        skor_per_sesi = {}
        for skor in skor_data:
            pertanyaan_id = str(skor['id_pertanyaan'])
            sesi = skor['sesi_penilaian']
            
            if pertanyaan_id not in skor_per_sesi:
                skor_per_sesi[pertanyaan_id] = {}
            
            skor_per_sesi[pertanyaan_id][sesi] = {
                'skor': skor['skor'],
                'nilai': skor['nilai'],
                'tanggal_input': skor['tanggal_input'],
                'is_locked': skor['is_locked']
            }
        
        # Hitung total sesi
        total_sesi = 0
        if skor_per_sesi:
            for pertanyaan_id in skor_per_sesi:
                sesi_keys = list(skor_per_sesi[pertanyaan_id].keys())
                if sesi_keys:
                    total_sesi = max(total_sesi, max(sesi_keys))
        
        # PERBAIKAN: Hitung nilai akhir dengan membedakan skor 0 real vs auto-filled
        nilai_akhir = 0
        sesi_dengan_nilai_real = 0  # Hitung sesi yang benar-benar dinilai (bukan auto-filled)
        
        if pertanyaan_list and skor_per_sesi:
            total_bobot = sum(float(p['bobot']) for p in pertanyaan_list)
            if total_bobot > 0:
                for pertanyaan in pertanyaan_list:
                    pertanyaan_id = str(pertanyaan['id'])
                    if pertanyaan_id in skor_per_sesi:
                        # ✅ PERBAIKAN: Hitung rata-rata skor untuk pertanyaan ini dari semua sesi yang ada
                        sesi_keys = list(skor_per_sesi[pertanyaan_id].keys())
                        if sesi_keys:
                            # ✅ PERBAIKAN: Hitung rata-rata dari semua sesi (termasuk skor 0)
                            total_skor = 0
                            for sesi in sesi_keys:
                                skor = skor_per_sesi[pertanyaan_id][sesi]['skor']
                                total_skor += skor
                                if skor > 0:  # Track sesi dengan nilai real untuk status
                                    sesi_dengan_nilai_real = max(sesi_dengan_nilai_real, sesi)
                            
                            # ✅ PERBAIKAN: Rata-rata = total skor / jumlah sesi (semua sesi)
                            rata_rata_skor = total_skor / len(sesi_keys)
                            skor_maksimal = float(pertanyaan['skor_maksimal'])
                            bobot = float(pertanyaan['bobot'])
                            
                            # Hitung nilai untuk pertanyaan ini
                            nilai_pertanyaan = (rata_rata_skor / skor_maksimal) * bobot
                            nilai_akhir += nilai_pertanyaan
        
        # PERBAIKAN: Tentukan status penilaian berdasarkan keberadaan data skor dan nilai
        # ✅ PERBAIKAN: Selalu ambil status dari database terlebih dahulu
        cursor.execute("""
            SELECT pm.status as status_penilaian_db, pm.jadwal_id
            FROM penilaian_mahasiswa pm
            WHERE pm.id = %s
        """, (id_penilaian_mahasiswa,))
        
        status_result = cursor.fetchone()
        status_penilaian = 'Belum Dinilai'  # Default
        
        if status_result:
            # ✅ PERBAIKAN: Jika status di database sudah Selesai, gunakan status tersebut
            if status_result['status_penilaian_db'] == 'Selesai':
                status_penilaian = 'Selesai'
            else:
                # Untuk status lain, cek berdasarkan skor dan nilai
                if skor_per_sesi:
                    # Jika ada data skor, cek apakah ada nilai yang lebih dari 0
                    ada_nilai_real = False
                    for pertanyaan_id in skor_per_sesi:
                        for sesi in skor_per_sesi[pertanyaan_id]:
                            if skor_per_sesi[pertanyaan_id][sesi]['skor'] > 0:
                                ada_nilai_real = True
                                break
                        if ada_nilai_real:
                            break
                    
                    if ada_nilai_real:
                        # ✅ PERBAIKAN: Jika ada nilai real dan jadwal sudah selesai, status otomatis Selesai
                        if status_result['jadwal_id']:
                            cursor.execute("""
                                SELECT pembimbing_nilai_selesai
                                FROM pengaturan_jadwal 
                                WHERE id = %s
                            """, (status_result['jadwal_id'],))
                            
                            jadwal_result = cursor.fetchone()
                            if jadwal_result and jadwal_result['pembimbing_nilai_selesai']:
                                from datetime import datetime
                                jadwal_selesai = jadwal_result['pembimbing_nilai_selesai']
                                if isinstance(jadwal_selesai, str):
                                    jadwal_selesai = datetime.strptime(jadwal_selesai, '%Y-%m-%d %H:%M:%S')
                                
                                # Jika jadwal sudah selesai dan ada nilai real, status otomatis menjadi Selesai
                                if datetime.now() > jadwal_selesai:
                                    status_penilaian = 'Selesai'
                                else:
                                    # Jadwal masih aktif, gunakan status dari database
                                    status_penilaian = status_result['status_penilaian_db']
                            else:
                                # Tidak ada jadwal, gunakan status dari database
                                status_penilaian = status_result['status_penilaian_db']
                        else:
                            # Tidak ada jadwal_id, gunakan status dari database
                            status_penilaian = status_result['status_penilaian_db']
                    else:
                        # Jika semua skor 0 (auto-fill), gunakan status dari database
                        status_penilaian = status_result['status_penilaian_db']
                else:
                    # Tidak ada data skor, gunakan status dari database
                    status_penilaian = status_result['status_penilaian_db']
        
        # Debug log untuk troubleshooting
        logger.info(f"DEBUG NILAI: penilaian_id={id_penilaian_mahasiswa}, nilai_akhir={nilai_akhir}, status={status_penilaian}")
        
        return nilai_akhir, sesi_dengan_nilai_real, status_penilaian
        
    except Exception as e:
        logger.error(f"Error in hitung_nilai_penilaian_mahasiswa: {str(e)}")
        return 0, 0, 'Error'

@pembimbing_bp.route('/get_daftar_mahasiswa_penilaian')
def get_daftar_mahasiswa_penilaian():
    """API untuk mendapatkan daftar mahasiswa yang bisa dinilai"""
    if 'user_type' not in session or session['user_type'] != 'pembimbing':
        return jsonify({'success': False, 'message': 'Anda harus login sebagai pembimbing!'})
    
    app_funcs = get_app_functions()
    if not hasattr(app_funcs['mysql'], 'connection') or app_funcs['mysql'].connection is None:
        return jsonify({'success': False, 'message': 'Koneksi ke database gagal!'})
    
    try:
        cursor = app_funcs['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Pastikan semua tabel penilaian sudah ada
        if not ensure_penilaian_tables(cursor, app_funcs['mysql'].connection):
            logger.error("Gagal membuat tabel penilaian")
            return jsonify({'success': False, 'message': 'Gagal membuat tabel penilaian'})
        
        # Ambil data mahasiswa yang dibimbing
        cursor.execute('''
            SELECT DISTINCT 
                m.nim as nim_mahasiswa,
                m.nama_ketua as nama_mahasiswa,
                m.status as status_mahasiswa,
                p.judul_usaha,
                p.id as proposal_id,
                p.tanggal_buat,
                p.tanggal_kirim,
                pm.id as penilaian_id,
                pm.status as status_penilaian
            FROM mahasiswa m
            INNER JOIN proposal p ON m.nim = p.nim
            LEFT JOIN penilaian_mahasiswa pm ON p.id = pm.id_proposal 
                AND pm.id_pembimbing = (SELECT id FROM pembimbing WHERE nama = %s)
            WHERE p.dosen_pembimbing = %s 
            AND p.status IN ('disetujui', 'revisi', 'selesai')
            AND p.status_admin = 'lolos'
            ORDER BY p.tanggal_buat DESC
        ''', (session['nama'], session['nama']))
        
        mahasiswa_list = cursor.fetchall()
        
        # Debug log
        logger.info(f"Raw data from database: {mahasiswa_list}")
        
        # Validasi data tidak kosong
        if not mahasiswa_list:
            logger.warning("Tidak ada data mahasiswa yang ditemukan")
            return jsonify({
                'success': True,
                'data': [],
                'message': 'Tidak ada data mahasiswa yang ditemukan'
            })
        
        # PERBAIKAN: Tidak perlu ambil pertanyaan global karena setiap mahasiswa
        # bisa punya set pertanyaan yang berbeda (historis)
        
        # Format data untuk response dengan perhitungan nilai yang konsisten
        formatted_data = []
        for item in mahasiswa_list:
            nilai = 0
            status_penilaian = item['status_penilaian'] or 'Belum Dinilai'
            total_sesi = 0
            
            if item['penilaian_id']:
                # Hitung nilai menggunakan helper function (tanpa pertanyaan_list global)
                nilai, total_sesi, status_penilaian = hitung_nilai_penilaian_mahasiswa(
                    cursor, item['penilaian_id']
                )
            
            # ✅ PERBAIKAN: Ambil jadwal yang sesuai dengan penilaian mahasiswa ini
            jadwal_info = get_jadwal_penilaian_mahasiswa(cursor, item['penilaian_id'])
            
            # ✅ PERBAIKAN: Konversi datetime dan timedelta ke string untuk JSON serialization
            jadwal_info_serializable = {}
            if jadwal_info:
                for key, value in jadwal_info.items():
                    if isinstance(value, datetime):
                        jadwal_info_serializable[key] = value.strftime('%Y-%m-%d %H:%M:%S')
                    elif isinstance(value, timedelta):
                        jadwal_info_serializable[key] = str(value)
                    else:
                        jadwal_info_serializable[key] = value
            
            formatted_data.append({
                'nim_mahasiswa': item['nim_mahasiswa'],
                'nama_mahasiswa': item['nama_mahasiswa'],
                'status_mahasiswa': item['status_mahasiswa'],  # ✅ PERBAIKAN: Tambahkan status mahasiswa
                'judul_usaha': item['judul_usaha'],
                'proposal_id': item['proposal_id'],
                'nilai': nilai,
                'status_penilaian': status_penilaian,
                'total_sesi': total_sesi,
                'tanggal_buat': item['tanggal_buat'].strftime('%Y-%m-%d %H:%M:%S') if item['tanggal_buat'] else None,
                'tanggal_kirim': item['tanggal_kirim'].strftime('%Y-%m-%d %H:%M:%S') if item['tanggal_kirim'] else None,
                'jadwal_info': jadwal_info_serializable  # ✅ PERBAIKAN: Gunakan versi yang bisa di-serialize
            })
        
        # Debug log untuk data yang sudah diformat
        logger.info(f"Formatted data: {formatted_data}")
        
        cursor.close()
        
        return jsonify({
            'success': True,
            'data': formatted_data
        })
        
    except Exception as e:
        logger.error(f"Error in get_daftar_mahasiswa_penilaian: {str(e)}")
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@pembimbing_bp.route('/get_info_mahasiswa')
def get_info_mahasiswa():
    """API untuk mendapatkan informasi mahasiswa"""
    if 'user_type' not in session or session['user_type'] != 'pembimbing':
        return jsonify({'success': False, 'message': 'Anda harus login sebagai pembimbing!'})
    
    app_funcs = get_app_functions()
    if not hasattr(app_funcs['mysql'], 'connection') or app_funcs['mysql'].connection is None:
        return jsonify({'success': False, 'message': 'Koneksi ke database gagal!'})
    
    try:
        nim = request.args.get('nim')
        if not nim:
            return jsonify({'success': False, 'message': 'NIM mahasiswa tidak ditemukan!'})
        
        cursor = app_funcs['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Ambil informasi mahasiswa
        cursor.execute('''
            SELECT m.nim, m.nama_ketua, m.program_studi, p.judul_usaha
            FROM mahasiswa m
            INNER JOIN proposal p ON m.nim = p.nim
            WHERE m.nim = %s AND p.dosen_pembimbing = %s
        ''', (nim, session['nama']))
        
        mahasiswa = cursor.fetchone()
        if not mahasiswa:
            cursor.close()
            return jsonify({'success': False, 'message': 'Data mahasiswa tidak ditemukan!'})
        
        cursor.close()
        
        return jsonify({
            'success': True,
            'data': {
                'nim': mahasiswa['nim'],
                'nama': mahasiswa['nama_ketua'],
                'program_studi': mahasiswa['program_studi'],
                'judul_usaha': mahasiswa['judul_usaha']
            }
        })
        
    except Exception as e:
        logger.error(f"Error in get_info_mahasiswa: {str(e)}")
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

def get_pertanyaan_for_evaluation(cursor):
    """
    Helper function untuk mendapatkan pertanyaan berdasarkan logic snapshot jadwal
    
    Logic:
    1. Jika ada jadwal aktif yang sedang berlangsung, gunakan pertanyaan yang aktif pada saat jadwal mulai
    2. Jika jadwal sudah selesai, gunakan pertanyaan yang aktif pada saat jadwal mulai  
    3. Jika tidak ada jadwal, gunakan pertanyaan yang aktif saat ini
    4. Pertanyaan baru atau yang diedit setelah jadwal mulai tidak akan muncul sampai ada jadwal baru
    """
    try:
        # Cek apakah ada jadwal pembimbing
        cursor.execute('''
            SELECT pembimbing_nilai_mulai, pembimbing_nilai_selesai
            FROM pengaturan_jadwal 
            ORDER BY id DESC 
            LIMIT 1
        ''')
        
        jadwal = cursor.fetchone()
        
        if jadwal and jadwal['pembimbing_nilai_mulai']:
            jadwal_mulai = jadwal['pembimbing_nilai_mulai']
            

            # Cek apakah ada pertanyaan yang dibuat sebelum jadwal mulai
            cursor.execute('''
                SELECT COUNT(*) as count
                FROM pertanyaan_penilaian_mahasiswa
                WHERE status = 'Aktif' AND created_at <= %s
            ''', (jadwal_mulai,))
            
            count_result = cursor.fetchone()
            pertanyaan_sebelum_jadwal = count_result['count'] if count_result else 0
            
            if pertanyaan_sebelum_jadwal > 0:
                # Gunakan pertanyaan yang aktif pada saat jadwal mulai
                cursor.execute('''
                    SELECT id, kategori, pertanyaan, bobot, skor_maksimal, status
                    FROM pertanyaan_penilaian_mahasiswa
                    WHERE status = 'Aktif' AND created_at <= %s
                    ORDER BY kategori ASC, created_at ASC
                ''', (jadwal_mulai,))
                
                logger.info(f"Menggunakan snapshot pertanyaan pada jadwal mulai: {jadwal_mulai}")
            else:
                # Jika tidak ada pertanyaan sebelum jadwal mulai, gunakan pertanyaan aktif saat ini
                cursor.execute('''
                    SELECT id, kategori, pertanyaan, bobot, skor_maksimal, status
                    FROM pertanyaan_penilaian_mahasiswa
                    WHERE status = 'Aktif'
                    ORDER BY kategori ASC, created_at ASC
                ''')
                
                logger.info(f"Tidak ada pertanyaan sebelum jadwal mulai ({jadwal_mulai}), menggunakan pertanyaan aktif saat ini")
        else:
            # Jika tidak ada jadwal, gunakan pertanyaan aktif saat ini
            cursor.execute('''
                SELECT id, kategori, pertanyaan, bobot, skor_maksimal, status
                FROM pertanyaan_penilaian_mahasiswa
                WHERE status = 'Aktif'
                ORDER BY kategori ASC, created_at ASC
            ''')
            
            logger.info("Tidak ada jadwal, menggunakan pertanyaan aktif saat ini")
        
        return cursor.fetchall()
        
    except Exception as e:
        logger.error(f"Error in get_pertanyaan_for_evaluation: {str(e)}")
        # Fallback ke pertanyaan aktif
        cursor.execute('''
            SELECT id, kategori, pertanyaan, bobot, skor_maksimal, status
            FROM pertanyaan_penilaian_mahasiswa
            WHERE status = 'Aktif'
            ORDER BY kategori ASC, created_at ASC
        ''')
        return cursor.fetchall()

@pembimbing_bp.route('/get_pertanyaan_penilaian')
def get_pertanyaan_penilaian():
    """API untuk mendapatkan pertanyaan penilaian mahasiswa"""
    if 'user_type' not in session or session['user_type'] != 'pembimbing':
        return jsonify({'success': False, 'message': 'Anda harus login sebagai pembimbing!'})
    
    app_funcs = get_app_functions()
    if not hasattr(app_funcs['mysql'], 'connection') or app_funcs['mysql'].connection is None:
        return jsonify({'success': False, 'message': 'Koneksi ke database gagal!'})
    
    try:
        cursor = app_funcs['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Buat tabel jika belum ada
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS pertanyaan_penilaian_mahasiswa (
                id INT AUTO_INCREMENT PRIMARY KEY,
                kategori VARCHAR(100) NOT NULL COMMENT 'Kategori penilaian mahasiswa',
                pertanyaan TEXT NOT NULL,
                bobot DECIMAL(5,2) NOT NULL COMMENT 'Bobot yang diinput admin (1-100)',
                skor_maksimal INT NOT NULL COMMENT 'Skor maksimal yang bisa diberikan (1-100)',
                status ENUM('Aktif', 'Nonaktif') DEFAULT 'Aktif',
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        
        # PERBAIKAN: Ambil pertanyaan berdasarkan snapshot jadwal mulai
        # Jika ada jadwal aktif, gunakan pertanyaan yang ada pada saat jadwal mulai
        # Jika tidak ada jadwal, gunakan pertanyaan aktif saat ini
        pertanyaan_list = get_pertanyaan_for_evaluation(cursor)
        cursor.close()
        
        # Log untuk debugging
        logger.info(f"Pertanyaan aktif yang dikirim: {len(pertanyaan_list)} pertanyaan")
        for pertanyaan in pertanyaan_list:
            logger.info(f"ID: {pertanyaan['id']}, Status: {pertanyaan['status']}, Pertanyaan: {pertanyaan['pertanyaan'][:50]}...")
        
        return jsonify({
            'success': True,
            'data': pertanyaan_list
        })
        
    except Exception as e:
        logger.error(f"Error in get_pertanyaan_penilaian: {str(e)}")
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})


@pembimbing_bp.route('/get_daftar_nilai_mahasiswa')
def get_daftar_nilai_mahasiswa():
    """API untuk mendapatkan daftar nilai mahasiswa"""
    if 'user_type' not in session or session['user_type'] != 'pembimbing':
        return jsonify({'success': False, 'message': 'Anda harus login sebagai pembimbing!'})
    
    app_funcs = get_app_functions()
    if not hasattr(app_funcs['mysql'], 'connection') or app_funcs['mysql'].connection is None:
        return jsonify({'success': False, 'message': 'Koneksi ke database gagal!'})
    
    try:
        cursor = app_funcs['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Pastikan semua tabel penilaian sudah ada
        ensure_penilaian_tables(cursor, app_funcs['mysql'].connection)
        
        # Ambil data mahasiswa yang dibimbing dengan nilai terbaru
        cursor.execute('''
            SELECT DISTINCT 
                m.nim as nim_mahasiswa,
                m.nama_ketua as nama_mahasiswa,
                p.judul_usaha,
                p.id as proposal_id,
                COALESCE(pm.nilai_akhir, 0) as nilai,
                COALESCE(pm.status, 'Belum Dinilai') as status_penilaian,
                p.tanggal_buat,
                p.tanggal_kirim
            FROM mahasiswa m
            INNER JOIN proposal p ON m.nim = p.nim
            LEFT JOIN penilaian_mahasiswa pm ON p.id = pm.id_proposal 
                AND pm.id_pembimbing = (SELECT id FROM pembimbing WHERE nama = %s)
            WHERE p.dosen_pembimbing = %s 
            AND p.status IN ('disetujui', 'revisi', 'selesai')
            AND p.status_admin = 'lolos'
            ORDER BY p.tanggal_buat DESC
        ''', (session['nama'], session['nama']))
        
        mahasiswa_list = cursor.fetchall()
        cursor.close()
        
        return jsonify({
            'success': True,
            'data': mahasiswa_list
        })
        
    except Exception as e:
        print(f"Error get_daftar_nilai_mahasiswa: {str(e)}")
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@pembimbing_bp.route('/get_skor_tersimpan')
def get_skor_tersimpan():
    """API untuk mengambil skor yang sudah tersimpan"""
    if 'user_type' not in session or session['user_type'] != 'pembimbing':
        return jsonify({'success': False, 'message': 'Anda harus login sebagai pembimbing!'})
    
    app_funcs = get_app_functions()
    if not hasattr(app_funcs['mysql'], 'connection') or app_funcs['mysql'].connection is None:
        return jsonify({'success': False, 'message': 'Koneksi ke database gagal!'})
    
    try:
        cursor = app_funcs['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        nim = request.args.get('nim')
        
        if not nim:
            return jsonify({'success': False, 'message': 'NIM mahasiswa tidak ditemukan!'})
        
        # Ambil ID pembimbing
        cursor.execute('SELECT id FROM pembimbing WHERE nama = %s', (session['nama'],))
        pembimbing = cursor.fetchone()
        if not pembimbing:
            cursor.close()
            return jsonify({'success': False, 'message': 'Data pembimbing tidak ditemukan!'})
        
        id_pembimbing = pembimbing['id']
        
        # Ambil ID proposal
        cursor.execute('''
            SELECT id FROM proposal 
            WHERE nim = %s AND dosen_pembimbing = %s
        ''', (nim, session['nama']))
        
        proposal = cursor.fetchone()
        if not proposal:
            cursor.close()
            return jsonify({'success': False, 'message': 'Data proposal tidak ditemukan!'})
        
        id_proposal = proposal['id']
        
        # Ambil skor yang sudah tersimpan
        cursor.execute('''
            SELECT dp.id_pertanyaan, dp.skor, dp.nilai
            FROM detail_penilaian_mahasiswa dp
            INNER JOIN penilaian_mahasiswa pm ON dp.id_penilaian_mahasiswa = pm.id
            WHERE pm.id_proposal = %s AND pm.id_pembimbing = %s
        ''', (id_proposal, id_pembimbing))
        
        skor_data = cursor.fetchall()
        cursor.close()
        
        return jsonify({'success': True, 'data': skor_data})
        
    except Exception as e:
        logger.error(f"Error in get_skor_tersimpan: {str(e)}")
        if 'cursor' in locals():
            cursor.close()
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

# ========================================
# ROUTE UNTUK SISTEM BIMBINGAN PEMBIMBING
# ========================================

@pembimbing_bp.route('/daftar_bimbingan_mahasiswa')
def daftar_bimbingan_mahasiswa():
    """Halaman daftar mahasiswa bimbingan"""
    if 'user_type' not in session or session['user_type'] != 'pembimbing':
        flash('Anda harus login sebagai pembimbing!', 'danger')
        return redirect(url_for('index'))
    
    app_funcs = get_app_functions()
    if not hasattr(app_funcs['mysql'], 'connection') or app_funcs['mysql'].connection is None:
        flash('Koneksi ke database gagal. Cek konfigurasi database!', 'danger')
        return render_template('pembimbing/daftar_bimbingan_mahasiswa.html', mahasiswa_bimbingan=[])
    
    try:
        cursor = app_funcs['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Ambil daftar mahasiswa yang dibimbing oleh pembimbing ini
        cursor.execute('''
            SELECT DISTINCT m.id, m.nim, m.nama_ketua, m.program_studi, p.judul_usaha
            FROM mahasiswa m
            JOIN proposal p ON m.nim = p.nim
            WHERE p.dosen_pembimbing = %s AND p.status_admin = 'lolos'
            ORDER BY m.nama_ketua ASC
        ''', (session['nama'],))
        
        mahasiswa_bimbingan = cursor.fetchall()
        cursor.close()
        
        return render_template('pembimbing/daftar_bimbingan_mahasiswa.html', 
                             mahasiswa_bimbingan=mahasiswa_bimbingan)
        
    except Exception as e:
        logger.error(f"Error in daftar_bimbingan_mahasiswa: {str(e)}")
        flash('Terjadi kesalahan saat memuat halaman!', 'danger')
        return render_template('pembimbing/daftar_bimbingan_mahasiswa.html', mahasiswa_bimbingan=[])

@pembimbing_bp.route('/bimbingan_mahasiswa/<int:mahasiswa_id>')
def bimbingan_mahasiswa(mahasiswa_id):
    """Halaman hasil bimbingan mahasiswa"""
    if 'user_type' not in session or session['user_type'] != 'pembimbing':
        flash('Anda harus login sebagai pembimbing!', 'danger')
        return redirect(url_for('index'))
    
    app_funcs = get_app_functions()
    if not hasattr(app_funcs['mysql'], 'connection') or app_funcs['mysql'].connection is None:
        flash('Koneksi ke database gagal. Cek konfigurasi database!', 'danger')
        return render_template('pembimbing/bimbingan_mahasiswa.html', mahasiswa_info=None, riwayat_bimbingan=[])
    
    try:
        cursor = app_funcs['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Ambil data mahasiswa yang dibimbing oleh pembimbing ini
        cursor.execute('''
            SELECT m.*, p.judul_usaha, p.id as proposal_id
            FROM mahasiswa m
            JOIN proposal p ON m.nim = p.nim
            WHERE m.id = %s AND p.dosen_pembimbing = %s AND p.status_admin = 'lolos'
        ''', (mahasiswa_id, session['nama']))
        
        mahasiswa_info = cursor.fetchone()
        if not mahasiswa_info:
            flash('Data mahasiswa tidak ditemukan atau Anda tidak memiliki akses!', 'danger')
            cursor.close()
            return redirect(url_for('pembimbing.daftar_bimbingan_mahasiswa'))
        
        # Ambil riwayat bimbingan mahasiswa
        cursor.execute('''
            SELECT b.*, p.judul_usaha, p.dosen_pembimbing
            FROM bimbingan b
            JOIN proposal p ON b.proposal_id = p.id
            WHERE b.nim = %s
            ORDER BY b.tanggal_buat DESC
        ''', (mahasiswa_info['nim'],))
        
        riwayat_bimbingan = cursor.fetchall()
        cursor.close()
        
        return render_template('pembimbing/bimbingan_mahasiswa.html', 
                             mahasiswa_info=mahasiswa_info,
                             riwayat_bimbingan=riwayat_bimbingan)
        
    except Exception as e:
        logger.error(f"Error in bimbingan_mahasiswa: {str(e)}")
        flash('Terjadi kesalahan saat memuat halaman!', 'danger')
        return render_template('pembimbing/bimbingan_mahasiswa.html', mahasiswa_info=None, riwayat_bimbingan=[])

@pembimbing_bp.route('/detail_bimbingan')
def detail_bimbingan():
    """API untuk mendapatkan detail bimbingan"""
    if 'user_type' not in session or session['user_type'] != 'pembimbing':
        return jsonify({'success': False, 'message': 'Anda harus login sebagai pembimbing!'})
    
    app_funcs = get_app_functions()
    if not hasattr(app_funcs['mysql'], 'connection') or app_funcs['mysql'].connection is None:
        return jsonify({'success': False, 'message': 'Koneksi ke database gagal!'})
    
    try:
        bimbingan_id = request.args.get('id')
        if not bimbingan_id:
            return jsonify({'success': False, 'message': 'ID bimbingan tidak ditemukan!'})
        
        cursor = app_funcs['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Ambil detail bimbingan dengan validasi pembimbing
        cursor.execute('''
            SELECT b.*, p.judul_usaha, p.dosen_pembimbing, m.nama_ketua, m.nim
            FROM bimbingan b
            JOIN proposal p ON b.proposal_id = p.id
            JOIN mahasiswa m ON b.nim = m.nim
            WHERE b.id = %s AND p.dosen_pembimbing = %s
        ''', (bimbingan_id, session['nama']))
        
        bimbingan = cursor.fetchone()
        if not bimbingan:
            cursor.close()
            return jsonify({'success': False, 'message': 'Data bimbingan tidak ditemukan!'})
        
        cursor.close()
        
        # Generate HTML untuk modal
        html = f'''
        <div class="row">
            <div class="col-12">
                <div class="mb-3">
                    <h6 class="text-primary">
                        <i class="bi bi-info-circle"></i>
                        Informasi Bimbingan:
                    </h6>
                    <div class="row">
                        <div class="col-md-6">
                            <p><strong>Mahasiswa:</strong> {bimbingan['nama_ketua']} ({bimbingan['nim']})</p>
                            <p><strong>Judul Usaha:</strong> {bimbingan['judul_usaha']}</p>
                        </div>
                        <div class="col-md-6">
                            <p><strong>Dosen Pembimbing:</strong> {bimbingan['dosen_pembimbing']}</p>
                            <p><strong>Tanggal Buat:</strong> {bimbingan['tanggal_buat'].strftime('%d/%m/%Y %H:%M') if bimbingan['tanggal_buat'] else '-'}</p>
                        </div>
                    </div>
                </div>
                
                <div class="mb-3">
                    <h6 class="text-primary">
                        <i class="bi bi-type-bold"></i>
                        Judul/Topik Bimbingan:
                    </h6>
                    <div class="hasil-bimbingan-content">
                        <strong>{bimbingan['judul_bimbingan']}</strong>
                    </div>
                </div>
                
                <div class="mb-3">
                    <h6 class="text-success">
                        <i class="bi bi-pencil-square"></i>
                        Hasil Bimbingan:
                    </h6>
                    <div class="hasil-bimbingan-content">
                        {bimbingan['hasil_bimbingan']}
                    </div>
                </div>
            </div>
        </div>
        '''
        
        return jsonify({
            'success': True,
            'html': html
        })
        
    except Exception as e:
        logger.error(f"Error in detail_bimbingan: {str(e)}")
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})


@pembimbing_bp.route('/pembimbing_hapus_bimbingan', methods=['POST'])
def pembimbing_hapus_bimbingan():
    """API untuk menghapus bimbingan"""
    if 'user_type' not in session or session['user_type'] != 'pembimbing':
        return jsonify({'success': False, 'message': 'Akses ditolak!'})
    
    app_funcs = get_app_functions()
    if not hasattr(app_funcs['mysql'], 'connection') or app_funcs['mysql'].connection is None:
        return jsonify({'success': False, 'message': 'Koneksi ke database gagal!'})
    
    try:
        bimbingan_id = request.form.get('id')
        
        if not bimbingan_id:
            return jsonify({'success': False, 'message': 'ID bimbingan tidak ditemukan!'})
        
        cursor = app_funcs['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Cek apakah bimbingan ada dan milik mahasiswa yang dibimbing oleh pembimbing ini
        cursor.execute('''
            SELECT b.*, p.dosen_pembimbing 
            FROM bimbingan b
            JOIN proposal p ON b.proposal_id = p.id
            WHERE b.id = %s AND p.dosen_pembimbing = %s
        ''', (bimbingan_id, session['nama']))
        
        bimbingan = cursor.fetchone()
        
        if not bimbingan:
            return jsonify({'success': False, 'message': 'Data bimbingan tidak ditemukan atau Anda tidak memiliki akses!'})
        
        # Hapus bimbingan
        cursor.execute('DELETE FROM bimbingan WHERE id = %s', (bimbingan_id,))
        app_funcs['mysql'].connection.commit()
        
        cursor.close()
        
        return jsonify({
            'success': True, 
            'message': 'Bimbingan berhasil dihapus!'
        })
        
    except Exception as e:
        logger.error(f"Error in pembimbing_hapus_bimbingan: {str(e)}")
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})


def check_pembimbing_penilaian_interval(id_pembimbing, nim_mahasiswa):
    """Cek apakah pembimbing bisa menilai berdasarkan interval yang diatur"""
    try:
        cursor = get_app_functions()['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Ambil pengaturan jadwal dan interval
        cursor.execute('''
            SELECT 
                pembimbing_nilai_mulai, pembimbing_nilai_selesai,
                pembimbing_interval_tipe, pembimbing_interval_nilai,
                pembimbing_hari_aktif,
                pembimbing_jam_mulai, pembimbing_jam_selesai
            FROM pengaturan_jadwal 
            ORDER BY id DESC 
            LIMIT 1
        ''')
        
        jadwal = cursor.fetchone()
        if not jadwal:
            return {'bisa_nilai': False, 'pesan': 'Belum ada pengaturan jadwal penilaian. Silakan hubungi admin untuk mengatur jadwal.'}
            
        now = datetime.now()
        mulai = jadwal.get('pembimbing_nilai_mulai')
        selesai = jadwal.get('pembimbing_nilai_selesai')
        
        # Cek periode utama (mulai-selesai) - WAJIB ADA
        if not mulai or not selesai:
            return {'bisa_nilai': False, 'pesan': 'Jadwal penilaian belum lengkap. Admin harus mengatur tanggal mulai dan selesai.'}
            
        if now < mulai:
            return {'bisa_nilai': False, 'pesan': f'Periode penilaian belum dimulai. Dimulai: {mulai.strftime("%d/%m/%Y %H:%M")}'}
        if now > selesai:
            return {'bisa_nilai': False, 'pesan': f'Periode penilaian sudah berakhir pada: {selesai.strftime("%d/%m/%Y %H:%M")}'}
            
        # ✅ PERBAIKAN: Cek jam aktif berdasarkan jam kerja pembimbing
        jam_kerja_mulai = jadwal.get('pembimbing_jam_mulai')
        jam_kerja_selesai = jadwal.get('pembimbing_jam_selesai')
        current_time = now.time()
        
        # ✅ PERBAIKAN: Cek apakah jam kerja valid (bukan 00:00:00)
        jam_kerja_valid = False
        if jam_kerja_mulai and jam_kerja_selesai:
            # Konversi timedelta ke time jika perlu (MySQL TIME field bisa return timedelta)
            if hasattr(jam_kerja_mulai, 'total_seconds'):  # timedelta
                jam_kerja_mulai = (datetime.min + jam_kerja_mulai).time()
            if hasattr(jam_kerja_selesai, 'total_seconds'):  # timedelta
                jam_kerja_selesai = (datetime.min + jam_kerja_selesai).time()
            
            # ✅ PERBAIKAN: Cek apakah jam kerja bukan 00:00:00 (tidak valid)
            if jam_kerja_mulai != datetime.strptime('00:00:00', '%H:%M:%S').time() or jam_kerja_selesai != datetime.strptime('00:00:00', '%H:%M:%S').time():
                jam_kerja_valid = True
                logger.info(f"Using jam kerja pembimbing: {jam_kerja_mulai.strftime('%H:%M')} - {jam_kerja_selesai.strftime('%H:%M')}")
            else:
                logger.info("Jam kerja pembimbing tidak valid (00:00:00), menggunakan fallback")
        
        if jam_kerja_valid:
            # Gunakan jam kerja pembimbing
            if current_time < jam_kerja_mulai or current_time > jam_kerja_selesai:
                return {'bisa_nilai': False, 'pesan': f'Penilaian hanya dapat dilakukan pada jam kerja pembimbing {jam_kerja_mulai.strftime("%H:%M")} - {jam_kerja_selesai.strftime("%H:%M")}. Sekarang: {current_time.strftime("%H:%M")}'}
        else:
            # ✅ PERBAIKAN: Fallback: gunakan jam dari periode penilaian
            jam_mulai_obj = mulai.time()
            jam_selesai_obj = selesai.time()
            logger.info(f"Using jam periode penilaian: {jam_mulai_obj.strftime('%H:%M')} - {jam_selesai_obj.strftime('%H:%M')}, current: {current_time.strftime('%H:%M')}")
            
            # ✅ PERBAIKAN: Handle kasus jam yang melewati tengah malam
            if jam_mulai_obj <= jam_selesai_obj:
                # Normal case: jam mulai < jam selesai (contoh: 08:00 - 17:00)
                if current_time < jam_mulai_obj or current_time > jam_selesai_obj:
                    return {'bisa_nilai': False, 'pesan': f'Penilaian hanya dapat dilakukan pada jam {jam_mulai_obj.strftime("%H:%M")} - {jam_selesai_obj.strftime("%H:%M")}. Sekarang: {current_time.strftime("%H:%M")}'}
            else:
                # Special case: jam mulai > jam selesai (contoh: 23:00 - 01:00, melewati tengah malam)
                if current_time < jam_mulai_obj and current_time > jam_selesai_obj:
                    return {'bisa_nilai': False, 'pesan': f'Penilaian hanya dapat dilakukan pada jam {jam_mulai_obj.strftime("%H:%M")} - {jam_selesai_obj.strftime("%H:%M")}. Sekarang: {current_time.strftime("%H:%M")}'}
        
        # ✅ PERBAIKAN: Cek hari aktif - HANYA JIKA RANGE LEBIH DARI 1 HARI
        range_days = (selesai.date() - mulai.date()).days
        
        if range_days > 0:  # Range lebih dari 1 hari, cek hari aktif
            hari_aktif = jadwal.get('pembimbing_hari_aktif')
            
            # ✅ PERBAIKAN: Jika hari_aktif NULL atau kosong, anggap semua hari aktif
            if not hari_aktif or hari_aktif.strip() == '':
                logger.info("Hari aktif tidak diatur, anggap semua hari aktif")
            else:
                hari_aktif_list = [int(x.strip()) for x in hari_aktif.split(',') if x.strip().isdigit()]
                if not hari_aktif_list:
                    logger.warning(f"Format hari aktif tidak valid: '{hari_aktif}', anggap semua hari aktif")
                else:
                    hari_sekarang = now.weekday() + 1  # Python: 0=Monday, kita: 1=Monday
                    nama_hari = ['', 'Senin', 'Selasa', 'Rabu', 'Kamis', 'Jumat', 'Sabtu', 'Minggu']
                    if hari_sekarang not in hari_aktif_list:
                        hari_aktif_nama = [nama_hari[i] for i in hari_aktif_list if 1 <= i <= 7]
                        return {'bisa_nilai': False, 'pesan': f'Hari {nama_hari[hari_sekarang]} bukan hari aktif penilaian. Hari aktif: {", ".join(hari_aktif_nama)}'}
        # Jika range 1 hari, skip pengecekan hari aktif (sudah pasti hari yang sama)
        
        # Cek interval penilaian - WAJIB ADA
        interval_tipe = jadwal.get('pembimbing_interval_tipe')
        interval_nilai = jadwal.get('pembimbing_interval_nilai')
        
        if not interval_tipe:
            return {'bisa_nilai': False, 'pesan': 'Tipe interval penilaian belum diatur. Admin harus memilih: setiap_jam, harian, mingguan, atau bulanan.'}
        
        # Auto-default untuk tipe tertentu
        if interval_tipe == 'setiap_jam' and (not interval_nilai or interval_nilai == '' or interval_nilai is None):
            interval_nilai = 1  # Auto default ke 1 jam
        elif not interval_nilai or interval_nilai is None:
            return {'bisa_nilai': False, 'pesan': f'Admin sudah memilih tipe "{interval_tipe}" tapi belum mengisi "Setiap Berapa". Contoh: jika tipe harian, isi "1" untuk setiap 1 hari.'}
            
        if interval_tipe not in ['harian', 'mingguan', 'bulanan', 'setiap_jam']:
            return {'bisa_nilai': False, 'pesan': 'Tipe interval penilaian tidak valid. Silakan hubungi admin.'}
            
        try:
            interval_nilai = int(interval_nilai)
            if interval_nilai < 1:
                return {'bisa_nilai': False, 'pesan': 'Nilai interval harus minimal 1. Silakan hubungi admin.'}
        except (ValueError, TypeError):
            return {'bisa_nilai': False, 'pesan': 'Nilai interval tidak valid. Silakan hubungi admin.'}
        
        # Ambil riwayat penilaian terakhir untuk mahasiswa ini dari detail_penilaian_mahasiswa
        cursor.execute('''
            SELECT dpm.tanggal_input as tanggal_penilaian, dpm.sesi_penilaian
            FROM detail_penilaian_mahasiswa dpm
            JOIN penilaian_mahasiswa pm ON dpm.id_penilaian_mahasiswa = pm.id
            WHERE pm.id_pembimbing = %s AND pm.id_proposal IN (
                SELECT id FROM proposal WHERE nim = %s
            )
            ORDER BY dpm.tanggal_input DESC 
            LIMIT 1
        ''', (id_pembimbing, nim_mahasiswa))
        
        riwayat = cursor.fetchone()
        
        if not riwayat:
            # Cek apakah benar-benar fresh start (tidak ada data penilaian sama sekali)
            cursor.execute('''
                SELECT COUNT(*) as total
                FROM detail_penilaian_mahasiswa dpm
                JOIN penilaian_mahasiswa pm ON dpm.id_penilaian_mahasiswa = pm.id
                JOIN proposal p ON pm.id_proposal = p.id
                WHERE p.nim = %s AND pm.id_pembimbing = %s
            ''', (nim_mahasiswa, id_pembimbing))
            
            existing_data = cursor.fetchone()
            has_existing_data = existing_data and existing_data['total'] > 0
            
            if not has_existing_data:
                # FIX: Cek apakah waktu sudah masuk sesi 2 atau lebih
                # Jika ya, maka sesi 1 sudah tidak bisa diakses
                sesi_berdasarkan_waktu = calculate_sesi_by_time()
                if sesi_berdasarkan_waktu > 1:
                    # Waktu sudah masuk sesi 2+, sesi 1 sudah tidak bisa diakses
                    # TAPI sesi 2 sudah bisa diakses!
                    # ✅ PERBAIKAN: Cek apakah ini data lama sebelum auto-fill
                    cursor.execute('''
                        SELECT p.tanggal_buat, p.tanggal_kirim 
                        FROM proposal p 
                        WHERE p.nim = %s
                    ''', (nim_mahasiswa,))
                    
                    proposal_info = cursor.fetchone()
                    if proposal_info:
                        tanggal_buat = proposal_info['tanggal_buat']
                        tanggal_kirim = proposal_info['tanggal_kirim']
                        
                        # ✅ PERBAIKAN: Ambil jadwal aktif untuk validasi
                        cursor.execute('''
                            SELECT pembimbing_nilai_mulai, pembimbing_nilai_selesai
                            FROM pengaturan_jadwal 
                            ORDER BY id DESC 
                            LIMIT 1
                        ''')
                        
                        jadwal_info = cursor.fetchone()
                        if jadwal_info and jadwal_info['pembimbing_nilai_mulai']:
                            waktu_mulai = jadwal_info['pembimbing_nilai_mulai']
                            waktu_selesai = jadwal_info['pembimbing_nilai_selesai']
                            
                            # ✅ PERBAIKAN: Jika proposal dibuat sebelum jadwal aktif, ini data lama - SKIP auto-fill
                            if tanggal_buat and tanggal_buat < waktu_mulai:
                                logger.info(f"Data lama terdeteksi - proposal dibuat {tanggal_buat} sebelum jadwal aktif {waktu_mulai}, SKIP auto-fill")
                                cursor.close()
                                return {'bisa_nilai': True, 'pesan': f'Waktu sudah masuk sesi {sesi_berdasarkan_waktu}. Data lama terdeteksi, tidak ada auto-fill.'}
                            
                            # ✅ PERBAIKAN: Jika proposal dibuat setelah jadwal selesai, ini data baru - SKIP auto-fill
                            if tanggal_buat and waktu_selesai and tanggal_buat > waktu_selesai:
                                logger.info(f"Data baru terdeteksi - proposal dibuat {tanggal_buat} setelah jadwal selesai {waktu_selesai}, SKIP auto-fill")
                                cursor.close()
                                return {'bisa_nilai': True, 'pesan': f'Waktu sudah masuk sesi {sesi_berdasarkan_waktu}. Data baru terdeteksi, tidak ada auto-fill.'}
                    
                    # ✅ PERBAIKAN: Biarkan worker yang menangani auto-fill
                    logger.info(f"Mahasiswa {nim_mahasiswa} - sesi {sesi_berdasarkan_waktu} memerlukan auto-fill, biarkan worker menangani")
                    cursor.close()
                    return {'bisa_nilai': True, 'pesan': f'Waktu sudah masuk sesi {sesi_berdasarkan_waktu}. Auto-fill akan diproses oleh sistem.'}
                else:
                    # Masih dalam sesi 1 - boleh nilai
                    cursor.close()
                    return {'bisa_nilai': True, 'pesan': 'Fresh start - dapat melakukan penilaian pertama tanpa menunggu interval'}
            else:
                # Ada data tapi tidak ada riwayat, mungkin data lama - cek waktu juga
                sesi_berdasarkan_waktu = calculate_sesi_by_time()
                if sesi_berdasarkan_waktu > 1:
                    # Waktu sudah masuk sesi 2+, sesi 1 sudah tidak bisa diakses
                    # TAPI sesi 2 sudah bisa diakses!
                    # ✅ PERBAIKAN: Cek apakah ini data lama sebelum auto-fill
                    cursor.execute('''
                        SELECT p.tanggal_buat, p.tanggal_kirim 
                        FROM proposal p 
                        WHERE p.nim = %s
                    ''', (nim_mahasiswa,))
                    
                    proposal_info = cursor.fetchone()
                    if proposal_info:
                        tanggal_buat = proposal_info['tanggal_buat']
                        tanggal_kirim = proposal_info['tanggal_kirim']
                        
                        # ✅ PERBAIKAN: Ambil jadwal aktif untuk validasi
                        cursor.execute('''
                            SELECT pembimbing_nilai_mulai, pembimbing_nilai_selesai
                            FROM pengaturan_jadwal 
                            ORDER BY id DESC 
                            LIMIT 1
                        ''')
                        
                        jadwal_info = cursor.fetchone()
                        if jadwal_info and jadwal_info['pembimbing_nilai_mulai']:
                            waktu_mulai = jadwal_info['pembimbing_nilai_mulai']
                            waktu_selesai = jadwal_info['pembimbing_nilai_selesai']
                            
                            # ✅ PERBAIKAN: Jika proposal dibuat sebelum jadwal aktif, ini data lama - SKIP auto-fill
                            if tanggal_buat and tanggal_buat < waktu_mulai:
                                logger.info(f"Data lama terdeteksi - proposal dibuat {tanggal_buat} sebelum jadwal aktif {waktu_mulai}, SKIP auto-fill")
                                cursor.close()
                                return {'bisa_nilai': True, 'pesan': f'Waktu sudah masuk sesi {sesi_berdasarkan_waktu}. Data lama terdeteksi, tidak ada auto-fill.'}
                            
                            # ✅ PERBAIKAN: Jika proposal dibuat setelah jadwal selesai, ini data baru - SKIP auto-fill
                            if tanggal_buat and waktu_selesai and tanggal_buat > waktu_selesai:
                                logger.info(f"Data baru terdeteksi - proposal dibuat {tanggal_buat} setelah jadwal selesai {waktu_selesai}, SKIP auto-fill")
                                cursor.close()
                                return {'bisa_nilai': True, 'pesan': f'Waktu sudah masuk sesi {sesi_berdasarkan_waktu}. Data baru terdeteksi, tidak ada auto-fill.'}
                    
                    # ✅ PERBAIKAN: Biarkan worker yang menangani auto-fill
                    logger.info(f"Mahasiswa {nim_mahasiswa} - sesi {sesi_berdasarkan_waktu} memerlukan auto-fill, biarkan worker menangani")
                    cursor.close()
                    return {'bisa_nilai': True, 'pesan': f'Waktu sudah masuk sesi {sesi_berdasarkan_waktu}. Auto-fill akan diproses oleh sistem.'}
                else:
                    cursor.close()
                    return {'bisa_nilai': True, 'pesan': 'Dapat melakukan penilaian (data penilaian ditemukan tapi tidak ada riwayat interval)'}
        
        tanggal_terakhir = riwayat['tanggal_penilaian']
        if isinstance(tanggal_terakhir, str):
            tanggal_terakhir = datetime.strptime(tanggal_terakhir, '%Y-%m-%d %H:%M:%S')
            
        # Hitung kapan boleh nilai lagi berdasarkan interval
        tanggal_terakhir_str = tanggal_terakhir.strftime("%d/%m/%Y %H:%M")
        
        if interval_tipe == 'setiap_jam':
            # FIX: Hitung berdasarkan jadwal mulai + (sesi - 1) * interval
            # Ambil jadwal mulai untuk perhitungan yang tepat
            cursor.execute('''
                SELECT pembimbing_nilai_mulai
                FROM pengaturan_jadwal 
                ORDER BY id DESC 
                LIMIT 1
            ''')
            
            jadwal_mulai_result = cursor.fetchone()
            if jadwal_mulai_result and jadwal_mulai_result['pembimbing_nilai_mulai']:
                jadwal_mulai = jadwal_mulai_result['pembimbing_nilai_mulai']
                
                # FIX: Hitung sesi yang seharusnya aktif berdasarkan waktu
                sesi_berdasarkan_waktu = calculate_sesi_by_time()
                
                # FIX: Gunakan sesi_berdasarkan_waktu sebagai sesi yang aktif sekarang
                sesi_sekarang = sesi_berdasarkan_waktu
                logger.info(f"DEBUG FORCE SESI: Waktu sudah masuk sesi {sesi_berdasarkan_waktu}, sesi_sekarang = {sesi_sekarang}")
                
                # ✅ PERBAIKAN: Jika ada riwayat lama yang lebih rendah, cek dulu apakah data lama
                if riwayat.get('sesi_penilaian') and riwayat.get('sesi_penilaian') < sesi_berdasarkan_waktu:
                    # ✅ PERBAIKAN: Cek apakah ini data lama sebelum auto-fill
                    cursor.execute('''
                        SELECT p.tanggal_buat, p.tanggal_kirim 
                        FROM proposal p 
                        WHERE p.nim = %s
                    ''', (nim_mahasiswa,))
                    
                    proposal_info = cursor.fetchone()
                    if proposal_info:
                        tanggal_buat = proposal_info['tanggal_buat']
                        tanggal_kirim = proposal_info['tanggal_kirim']
                        
                        # ✅ PERBAIKAN: Jika proposal dibuat sebelum jadwal aktif, ini data lama - SKIP auto-fill
                        if tanggal_buat and tanggal_buat < jadwal_mulai:
                            logger.info(f"Data lama terdeteksi - proposal dibuat {tanggal_buat} sebelum jadwal aktif {jadwal_mulai}, SKIP auto-fill")
                        else:
                            logger.info(f"DEBUG AUTO-FILL MISSED: Riwayat lama sesi {riwayat.get('sesi_penilaian')}, biarkan worker menangani auto-fill")
                            # Biarkan worker yang menangani auto-fill
                
                # Sesi 1: jadwal_mulai + 0 jam
                # Sesi 2: jadwal_mulai + 1 jam  
                # Sesi 3: jadwal_mulai + 2 jam, dst
                waktu_boleh_nilai = jadwal_mulai + timedelta(hours=(sesi_sekarang - 1) * interval_nilai)
                
                logger.info(f"DEBUG INTERVAL: sesi_sekarang={sesi_sekarang}, jadwal_mulai={jadwal_mulai}, waktu_boleh_nilai={waktu_boleh_nilai}, now={now}")
                
                # FIX: Tambah log detail untuk debug
                logger.info(f"DEBUG INTERVAL DETAIL: riwayat.sesi_penilaian={riwayat.get('sesi_penilaian')}, sesi_berdasarkan_waktu={sesi_berdasarkan_waktu}, interval_nilai={interval_nilai}")
                logger.info(f"DEBUG INTERVAL CALC: jadwal_mulai={jadwal_mulai}, (sesi_sekarang-1)*interval_nilai={(sesi_sekarang-1)*interval_nilai}, waktu_boleh_nilai={waktu_boleh_nilai}")
                logger.info(f"DEBUG INTERVAL COMPARE: now={now}, waktu_boleh_nilai={waktu_boleh_nilai}, now < waktu_boleh_nilai={now < waktu_boleh_nilai}")
                
                if now < waktu_boleh_nilai:
                    sisa_detik = (waktu_boleh_nilai - now).total_seconds()
                    sisa_jam = int(sisa_detik // 3600)
                    sisa_menit = int((sisa_detik % 3600) // 60)
                    logger.info(f"DEBUG INTERVAL BLOCKED: Sesi {sesi_sekarang} diblokir, sisa waktu: {sisa_jam} jam {sisa_menit} menit")
                    return {'bisa_nilai': False, 'pesan': f'Sesi {sesi_sekarang} dapat dimulai pada {waktu_boleh_nilai.strftime("%d/%m/%Y %H:%M")}. Sisa waktu: {sisa_jam} jam {sisa_menit} menit.'}
                else:
                    logger.info(f"DEBUG INTERVAL ALLOWED: Sesi {sesi_sekarang} diizinkan, waktu sudah cukup")
                    cursor.close()
                    return {'bisa_nilai': True, 'pesan': f'Sesi {sesi_sekarang} sudah dapat dimulai. Waktu mulai: {waktu_boleh_nilai.strftime("%d/%m/%Y %H:%M")}'}
            else:
                # Fallback ke method lama jika tidak ada jadwal mulai
                selisih_jam = (now - tanggal_terakhir).total_seconds() / 3600
                if selisih_jam < interval_nilai:
                    sisa_jam = interval_nilai - int(selisih_jam)
                    kapan_boleh = tanggal_terakhir + timedelta(hours=interval_nilai)
                    return {'bisa_nilai': False, 'pesan': f'Terakhir dinilai: {tanggal_terakhir_str}. Dapat menilai lagi dalam {sisa_jam} jam (pada {kapan_boleh.strftime("%d/%m/%Y %H:%M")})'}
                
        elif interval_tipe == 'harian':
            selisih_hari = (now.date() - tanggal_terakhir.date()).days
            if selisih_hari < interval_nilai:
                sisa_hari = interval_nilai - selisih_hari
                kapan_boleh = tanggal_terakhir + timedelta(days=interval_nilai)
                return {'bisa_nilai': False, 'pesan': f'Terakhir dinilai: {tanggal_terakhir_str}. Dapat menilai lagi dalam {sisa_hari} hari (pada {kapan_boleh.strftime("%d/%m/%Y")})'}
                
        elif interval_tipe == 'mingguan':
            # Hitung minggu berdasarkan hari mulai periode
            start_of_week_last = tanggal_terakhir - timedelta(days=tanggal_terakhir.weekday())
            start_of_week_now = now - timedelta(days=now.weekday())
            selisih_minggu = (start_of_week_now.date() - start_of_week_last.date()).days // 7
            if selisih_minggu < interval_nilai:
                sisa_minggu = interval_nilai - selisih_minggu
                kapan_boleh = tanggal_terakhir + timedelta(weeks=interval_nilai)
                return {'bisa_nilai': False, 'pesan': f'Terakhir dinilai: {tanggal_terakhir_str}. Dapat menilai lagi dalam {sisa_minggu} minggu (pada {kapan_boleh.strftime("%d/%m/%Y")})'}
                
        elif interval_tipe == 'bulanan':
            # Hitung bulan
            bulan_terakhir = tanggal_terakhir.year * 12 + tanggal_terakhir.month
            bulan_sekarang = now.year * 12 + now.month
            selisih_bulan = bulan_sekarang - bulan_terakhir
            if selisih_bulan < interval_nilai:
                sisa_bulan = interval_nilai - selisih_bulan
                # Approximate next month
                next_year = tanggal_terakhir.year
                next_month = tanggal_terakhir.month + interval_nilai
                if next_month > 12:
                    next_year += (next_month - 1) // 12
                    next_month = ((next_month - 1) % 12) + 1
                kapan_boleh = tanggal_terakhir.replace(year=next_year, month=next_month)
                return {'bisa_nilai': False, 'pesan': f'Terakhir dinilai: {tanggal_terakhir_str}. Dapat menilai lagi dalam {sisa_bulan} bulan (pada {kapan_boleh.strftime("%d/%m/%Y")})'}
        
        cursor.close()
        
        # Pastikan interval_nilai ada untuk display
        if interval_tipe == 'setiap_jam' and (not interval_nilai or interval_nilai == ''):
            interval_nilai = 1  # Auto default untuk display
        
        interval_text = {
            'setiap_jam': f'setiap {interval_nilai} jam',
            'harian': f'setiap {interval_nilai} hari',
            'mingguan': f'setiap {interval_nilai} minggu',
            'bulanan': f'setiap {interval_nilai} bulan'
        }.get(interval_tipe, 'sesuai interval')
        
        return {'bisa_nilai': True, 'pesan': f'Dapat melakukan penilaian sesi berikutnya (interval: {interval_text})'}
        
    except Exception as e:
        return {'bisa_nilai': False, 'pesan': f'Error: {str(e)}'}

def is_schedule_active():
    """Cek apakah jadwal penilaian pembimbing masih aktif"""
    try:
        cursor = get_app_functions()['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        cursor.execute('''
            SELECT pembimbing_nilai_mulai, pembimbing_nilai_selesai
            FROM pengaturan_jadwal 
            ORDER BY id DESC 
            LIMIT 1
        ''')
        
        jadwal = cursor.fetchone()
        cursor.close()
        
        if not jadwal:
            logger.warning("Tidak ada jadwal penilaian ditemukan, anggap tidak aktif")
            return False  # ✅ PERBAIKAN: Jika tidak ada jadwal, anggap tidak aktif
            
        mulai = jadwal.get('pembimbing_nilai_mulai')
        selesai = jadwal.get('pembimbing_nilai_selesai')
        
        if not mulai or not selesai:
            logger.warning("Jadwal tidak lengkap (mulai atau selesai kosong), anggap tidak aktif")
            return False  # ✅ PERBAIKAN: Jika tidak ada batas waktu, anggap tidak aktif
            
        now = datetime.now()
        
        # ✅ PERBAIKAN: Jadwal aktif jika dalam rentang waktu
        is_active = mulai <= now <= selesai
        logger.info(f"Jadwal status: {'AKTIF' if is_active else 'TIDAK AKTIF'} (mulai: {mulai}, selesai: {selesai}, now: {now})")
        return is_active
        
    except Exception as e:
        logger.error(f"Error checking schedule: {str(e)}")
        return False  # ✅ PERBAIKAN: Default tidak aktif jika error

def calculate_sesi_by_time():
    """Hitung sesi yang seharusnya aktif berdasarkan waktu yang sudah berlalu dari jadwal mulai"""
    try:
        cursor = get_app_functions()['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        cursor.execute('''
            SELECT pembimbing_nilai_mulai, pembimbing_interval_tipe, pembimbing_interval_nilai
            FROM pengaturan_jadwal 
            ORDER BY id DESC LIMIT 1
        ''')
        
        jadwal = cursor.fetchone()
        cursor.close()
        
        if not jadwal:
            return 1  # Default sesi 1 jika tidak ada jadwal
            
        mulai = jadwal.get('pembimbing_nilai_mulai')
        interval_tipe = jadwal.get('pembimbing_interval_tipe')
        interval_nilai = jadwal.get('pembimbing_interval_nilai')
        
        if not mulai or not interval_tipe or not interval_nilai:
            return 1  # Default sesi 1 jika data tidak lengkap
            
        now = datetime.now()
        
        # Jika waktu sekarang masih sebelum jadwal mulai
        if now < mulai:
            return 1
            
        # Hitung selisih waktu dari jadwal mulai
        selisih = now - mulai
        
        # Hitung sesi berdasarkan interval
        if interval_tipe == 'setiap_jam':
            # Setiap jam: sesi 1 pada jam 0, sesi 2 pada jam 1, dst
            sesi_berdasarkan_waktu = int(selisih.total_seconds() // 3600) + 1
        elif interval_tipe == 'harian':
            # Setiap hari: sesi 1 pada hari 0, sesi 2 pada hari 1, dst
            sesi_berdasarkan_waktu = int(selisih.days // interval_nilai) + 1
        elif interval_tipe == 'mingguan':
            # Setiap minggu: sesi 1 pada minggu 0, sesi 2 pada minggu 1, dst
            minggu_berlalu = int(selisih.days // 7 // interval_nilai)
            sesi_berdasarkan_waktu = minggu_berlalu + 1
        elif interval_tipe == 'bulanan':
            # ✅ PERBAIKAN: Untuk interval bulanan, hitung berdasarkan bulan yang sebenarnya
            # Ambil jadwal selesai untuk menentukan total bulan
            cursor = get_app_functions()['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
            cursor.execute('''
                SELECT pembimbing_nilai_selesai
                FROM pengaturan_jadwal 
                ORDER BY id DESC LIMIT 1
            ''')
            jadwal_selesai_result = cursor.fetchone()
            cursor.close()
            
            if jadwal_selesai_result and jadwal_selesai_result['pembimbing_nilai_selesai']:
                jadwal_selesai = jadwal_selesai_result['pembimbing_nilai_selesai']
                
                # Hitung total bulan antara jadwal mulai dan selesai
                total_bulan = (jadwal_selesai.year - mulai.year) * 12 + (jadwal_selesai.month - mulai.month) + 1
                
                # Jika waktu sekarang sudah melewati jadwal selesai, sesi = total_bulan
                if now > jadwal_selesai:
                    sesi_berdasarkan_waktu = total_bulan
                else:
                    # Jika masih dalam jadwal, sesi = 1 (karena hanya 1 periode bulanan)
                    sesi_berdasarkan_waktu = 1
                    
                logger.info(f"DEBUG BULANAN: mulai={mulai}, selesai={jadwal_selesai}, total_bulan={total_bulan}, now={now}, sesi={sesi_berdasarkan_waktu}")
            else:
                # Fallback: gunakan perhitungan lama jika tidak ada jadwal selesai
                bulan_berlalu = int(selisih.days // 30 // interval_nilai)
                sesi_berdasarkan_waktu = bulan_berlalu + 1
        else:
            sesi_berdasarkan_waktu = 1
            
        logger.info(f"DEBUG CALCULATE SESI BY TIME: mulai={mulai}, now={now}, selisih={selisih}, interval_tipe={interval_tipe}, interval_nilai={interval_nilai}, sesi_calculated={sesi_berdasarkan_waktu}")
        return max(1, sesi_berdasarkan_waktu)
        
    except Exception as e:
        logger.error(f"Error calculating sesi by time: {str(e)}")
        return 1  # Default sesi 1 jika error

def get_current_sesi_penilaian(id_pembimbing, nim_mahasiswa):
    """Menentukan sesi penilaian yang sedang aktif berdasarkan riwayat DAN waktu yang sudah berlalu"""
    try:
        cursor = get_app_functions()['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Cek apakah ada data penilaian aktual
        cursor.execute('''
            SELECT COUNT(*) as total_data
            FROM detail_penilaian_mahasiswa dpm
            JOIN penilaian_mahasiswa pm ON dpm.id_penilaian_mahasiswa = pm.id
            JOIN proposal p ON pm.id_proposal = p.id
            WHERE p.nim = %s AND pm.id_pembimbing = %s
        ''', (nim_mahasiswa, id_pembimbing))
        
        data_result = cursor.fetchone()
        total_data = data_result['total_data'] if data_result else 0
        
        # Hitung sesi tertinggi yang sudah dilakukan dari detail_penilaian_mahasiswa
        cursor.execute('''
            SELECT COALESCE(MAX(dpm.sesi_penilaian), 0) as max_sesi_penilaian
            FROM detail_penilaian_mahasiswa dpm
            JOIN penilaian_mahasiswa pm ON dpm.id_penilaian_mahasiswa = pm.id
            WHERE pm.id_pembimbing = %s AND pm.id_proposal IN (
                SELECT id FROM proposal WHERE nim = %s
            )
        ''', (id_pembimbing, nim_mahasiswa))
        
        result = cursor.fetchone()
        max_sesi_penilaian = result['max_sesi_penilaian'] if result else 0
        
        # Cek apakah jadwal masih aktif
        schedule_active = is_schedule_active()
        
        # NEW: Hitung sesi berdasarkan waktu yang sudah berlalu
        sesi_berdasarkan_waktu = calculate_sesi_by_time()
        
        # Logic untuk menentukan sesi
        if total_data == 0 and max_sesi_penilaian == 0:
            # Fresh start - gunakan sesi berdasarkan waktu atau minimal 1
            current_sesi = max(1, sesi_berdasarkan_waktu)
        elif total_data == 0 and max_sesi_penilaian > 0:
            # Ada riwayat tapi tidak ada data aktual (data dihapus) - reset ke sesi 1
            current_sesi = 1
            logger.info(f"DEBUG: Reset to sesi 1 - no actual data but has history")
        elif not schedule_active:
            # Jika jadwal sudah berakhir, hitung berapa sesi yang seharusnya ada
            # berdasarkan total durasi jadwal, lalu autofill yang kosong
            total_sesi_seharusnya = sesi_berdasarkan_waktu  # Berdasarkan waktu penuh
            current_sesi = max(1, max_sesi_penilaian, total_sesi_seharusnya)  # Ambil yang terbesar
        else:
            # FIX: Prioritaskan sesi berdasarkan waktu untuk memaksa pindah ke sesi berikutnya
            # Jika waktu sudah masuk sesi 2, maka current_sesi HARUS 2, bukan 1
            if sesi_berdasarkan_waktu > 1:
                current_sesi = sesi_berdasarkan_waktu
                logger.info(f"DEBUG FORCE UPDATE: Waktu sudah masuk sesi {sesi_berdasarkan_waktu}, force update current_sesi")
            else:
                # Gunakan yang lebih tinggi: sesi berdasarkan riwayat + 1 atau sesi berdasarkan waktu
                sesi_berdasarkan_riwayat = max_sesi_penilaian + 1
                current_sesi = max(1, sesi_berdasarkan_riwayat, sesi_berdasarkan_waktu)
            
            # Jika ada gap sesi (sesi terlewat), beri warning di log
            if sesi_berdasarkan_waktu > max_sesi_penilaian:
                logger.warning(f"WARNING: Sesi terlewat! Waktu sudah masuk sesi {sesi_berdasarkan_waktu}, tapi riwayat baru sampai sesi {max_sesi_penilaian}")
        
        cursor.close()
        logger.info(f"DEBUG CURRENT SESI: total_data={total_data}, max_sesi_penilaian={max_sesi_penilaian}, sesi_berdasarkan_waktu={sesi_berdasarkan_waktu}, schedule_active={schedule_active}, current_sesi={current_sesi}")
        return current_sesi
        
    except Exception as e:
        logger.error(f"Error getting current sesi: {str(e)}")
        return 1

def get_jadwal_info_for_ui(id_pembimbing, nim_mahasiswa):
    """Mendapatkan info jadwal untuk ditampilkan di UI (fungsi lama untuk kompatibilitas)"""
    try:
        cursor = get_app_functions()['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Ambil pengaturan jadwal
        cursor.execute('''
            SELECT 
                pembimbing_nilai_mulai, pembimbing_nilai_selesai,
                pembimbing_interval_tipe, pembimbing_interval_nilai,
                pembimbing_hari_aktif
            FROM pengaturan_jadwal 
            ORDER BY id DESC LIMIT 1
        ''')
        
        jadwal = cursor.fetchone()
        if not jadwal:
            return {'status': 'Tidak ada pengaturan jadwal'}
        
        # Format info untuk UI - ambil jam dari datetime utama
        mulai = jadwal.get('pembimbing_nilai_mulai')
        selesai = jadwal.get('pembimbing_nilai_selesai')
        
        if mulai and selesai:
            jam_mulai = mulai.strftime('%H:%M')
            jam_selesai = selesai.strftime('%H:%M')
        else:
            jam_mulai = '08:00'
            jam_selesai = '17:00'
        
        interval_tipe = jadwal.get('pembimbing_interval_tipe', 'harian')
        interval_nilai = jadwal.get('pembimbing_interval_nilai', 1)
        
        # Cek apakah range 1 hari atau multi-hari
        if mulai and selesai:
            range_days = (selesai.date() - mulai.date()).days
            if range_days > 0:  # Range lebih dari 1 hari
                hari_mapping = {1: 'Sen', 2: 'Sel', 3: 'Rab', 4: 'Kam', 5: 'Jum', 6: 'Sab', 7: 'Min'}
                hari_aktif = jadwal.get('pembimbing_hari_aktif', '1,2,3,4,5')
                hari_list = [hari_mapping.get(int(h.strip()), h.strip()) for h in hari_aktif.split(',') if h.strip().isdigit()]
                hari_aktif_text = ', '.join(hari_list) if hari_list else 'Belum diatur'
            else:  # Range 1 hari saja
                hari_aktif_text = f"Hanya 1 hari ({mulai.strftime('%d/%m/%Y')})"
        else:
            hari_aktif_text = 'Belum diatur'
        
        interval_text = f"Setiap {interval_nilai} {interval_tipe}"
        if interval_tipe == 'setiap_jam':
            interval_text = f"Setiap {interval_nilai} jam"
        
        cursor.close()
        
        # Hitung daftar jam untuk interval setiap jam
        jam_list = []
        if interval_tipe == 'setiap_jam' and mulai and selesai:
            current_time = mulai
            while current_time <= selesai:
                jam_list.append(current_time.strftime('%H:%M'))
                # Tambah interval jam
                current_time = current_time + timedelta(hours=interval_nilai)
        
        # Ambil metadata kolom dari admin - PERBAIKAN: Ambil berdasarkan periode aktif
        metadata_kolom = []
        if interval_tipe and interval_nilai:
            try:
                # Ambil periode aktif terbaru
                cursor.execute("""
                    SELECT periode_metadata_aktif 
                    FROM pengaturan_jadwal 
                    ORDER BY id DESC LIMIT 1
                """)
                periode_result = cursor.fetchone()
                periode_aktif = periode_result['periode_metadata_aktif'] if periode_result else 1
                
                # Ambil metadata berdasarkan periode aktif dan interval tipe yang sesuai
                cursor.execute("""
                    SELECT id, nama_kolom, urutan_kolom, interval_tipe, interval_nilai
                    FROM metadata_kolom_penilaian 
                    WHERE interval_tipe = %s AND periode_interval_id = %s
                    ORDER BY urutan_kolom
                """, (interval_tipe, periode_aktif))
                
                metadata_kolom = cursor.fetchall()
                logger.info(f"Metadata kolom untuk interval {interval_tipe}, periode {periode_aktif}: {len(metadata_kolom)} items")
            except Exception as e:
                logger.warning(f"Tidak bisa ambil metadata kolom: {str(e)}")
                metadata_kolom = []
        
        return {
            'jam_aktif': f"{jam_mulai} - {jam_selesai}",
            'hari_aktif': hari_aktif_text,
            'interval': interval_text,
            'periode': f"{jadwal.get('pembimbing_nilai_mulai', 'Tidak diatur')} s/d {jadwal.get('pembimbing_nilai_selesai', 'Tidak diatur')}",
            'jam_list': jam_list,
            'interval_tipe': interval_tipe,
            'interval_nilai': interval_nilai,
            'metadata_kolom': metadata_kolom
        }
        
    except Exception as e:
        logger.error(f"Error getting jadwal info: {str(e)}")
        return {'status': f'Error: {str(e)}'}

def get_session_info_for_ui(id_pembimbing, nim_mahasiswa):
    """
    Mendapatkan informasi sesi untuk UI dengan perhitungan otomatis
    """
    try:
        cursor = get_app_functions()['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Ambil pengaturan jadwal
        cursor.execute('''
            SELECT 
                pembimbing_nilai_mulai, pembimbing_nilai_selesai,
                pembimbing_interval_tipe, pembimbing_interval_nilai,
                pembimbing_hari_aktif
            FROM pengaturan_jadwal 
            ORDER BY id DESC LIMIT 1
        ''')
        
        jadwal = cursor.fetchone()
        if not jadwal:
            return None
        
        mulai = jadwal.get('pembimbing_nilai_mulai')
        selesai = jadwal.get('pembimbing_nilai_selesai')
        interval_tipe = jadwal.get('pembimbing_interval_tipe', 'harian')
        interval_nilai = jadwal.get('pembimbing_interval_nilai', 1)
        
        if not mulai or not selesai:
            return None
        
        # Hitung waktu sekarang
        now = datetime.now()
        
        # Hitung total sesi berdasarkan interval
        total_sesi = 1
        current_sesi = 1
        sessions_available = [1]
        
        if interval_tipe == 'setiap_jam':
            # Hitung sesi berdasarkan jam
            if mulai <= now <= selesai:
                # Waktu sekarang dalam range jadwal
                time_diff = now - mulai
                hours_passed = int(time_diff.total_seconds() / 3600)
                current_sesi = (hours_passed // interval_nilai) + 1
                
                # Hitung total sesi yang seharusnya ada
                total_hours = int((selesai - mulai).total_seconds() / 3600)
                total_sesi = (total_hours // interval_nilai) + 1
                
                # Sesi yang tersedia untuk input
                sessions_available = list(range(1, current_sesi + 1))
                
            elif now < mulai:
                # Belum waktunya jadwal
                current_sesi = 1
                total_hours = int((selesai - mulai).total_seconds() / 3600)
                total_sesi = (total_hours // interval_nilai) + 1
                sessions_available = []
                
            else:
                # Jadwal sudah selesai
                total_hours = int((selesai - mulai).total_seconds() / 3600)
                total_sesi = (total_hours // interval_nilai) + 1
                current_sesi = total_sesi
                sessions_available = list(range(1, total_sesi + 1))
                
        elif interval_tipe == 'harian':
            # Hitung sesi berdasarkan hari
            if mulai <= now <= selesai:
                days_passed = (now.date() - mulai.date()).days
                current_sesi = (days_passed // interval_nilai) + 1
                
                total_days = (selesai.date() - mulai.date()).days
                total_sesi = (total_days // interval_nilai) + 1
                
                sessions_available = list(range(1, current_sesi + 1))
                
            elif now < mulai:
                current_sesi = 1
                total_days = (selesai.date() - mulai.date()).days
                total_sesi = (total_days // interval_nilai) + 1
                sessions_available = []
                
            else:
                total_days = (selesai.date() - mulai.date()).days
                total_sesi = (total_days // interval_nilai) + 1
                current_sesi = total_sesi
                sessions_available = list(range(1, total_sesi + 1))
                
        elif interval_tipe == 'mingguan':
            # Hitung sesi berdasarkan minggu
            if mulai <= now <= selesai:
                weeks_passed = int((now.date() - mulai.date()).days / 7)
                current_sesi = (weeks_passed // interval_nilai) + 1
                
                total_weeks = int((selesai.date() - mulai.date()).days / 7)
                total_sesi = (total_weeks // interval_nilai) + 1
                
                sessions_available = list(range(1, current_sesi + 1))
                
            elif now < mulai:
                current_sesi = 1
                total_weeks = int((selesai.date() - mulai.date()).days / 7)
                total_sesi = (total_weeks // interval_nilai) + 1
                sessions_available = []
                
            else:
                total_weeks = int((selesai.date() - mulai.date()).days / 7)
                total_sesi = (total_weeks // interval_nilai) + 1
                current_sesi = total_sesi
                sessions_available = list(range(1, total_sesi + 1))
                
        elif interval_tipe == 'bulanan':
            # ✅ PERBAIKAN: Hitung sesi berdasarkan bulan yang sebenarnya
            # Untuk interval bulanan, gunakan metadata kolom untuk menentukan total sesi
            metadata_kolom_bulanan = []
            try:
                cursor.execute('''
                    SELECT urutan_kolom, nama_kolom
                    FROM metadata_kolom_penilaian 
                    WHERE jadwal_id = (SELECT id FROM pengaturan_jadwal ORDER BY id DESC LIMIT 1)
                    AND interval_tipe = 'bulanan'
                    ORDER BY urutan_kolom ASC
                ''')
                metadata_kolom_bulanan = cursor.fetchall()
            except Exception as e:
                logger.error(f"Error getting metadata kolom bulanan: {str(e)}")
            
            # Jika ada metadata kolom, gunakan jumlah metadata sebagai total sesi
            if metadata_kolom_bulanan:
                total_sesi = len(metadata_kolom_bulanan)
                logger.info(f"Using metadata kolom count for bulanan total_sesi: {total_sesi}")
            else:
                # Fallback ke perhitungan lama
                total_months = (selesai.year - mulai.year) * 12 + (selesai.month - mulai.month) + 1
                total_sesi = max(1, total_months // interval_nilai)
                logger.info(f"Using fallback calculation for bulanan total_sesi: {total_sesi}")
            
            # Hitung current sesi
            if mulai <= now <= selesai:
                months_passed = (now.year - mulai.year) * 12 + (now.month - mulai.month)
                current_sesi = (months_passed // interval_nilai) + 1
                current_sesi = min(current_sesi, total_sesi)  # Pastikan tidak melebihi total sesi
                sessions_available = list(range(1, current_sesi + 1))
                
            elif now < mulai:
                current_sesi = 1
                sessions_available = []
                
            else:
                current_sesi = total_sesi
                sessions_available = list(range(1, total_sesi + 1))
        
        # Hitung daftar jam untuk interval setiap jam
        jam_list = []
        if interval_tipe == 'setiap_jam':
            current_time = mulai
            while current_time <= selesai:
                jam_list.append(current_time.strftime('%H:%M'))
                current_time = current_time + timedelta(hours=interval_nilai)
        
        # ✅ PERBAIKAN: Ambil metadata kolom dari database yang sesuai dengan interval aktif dan periode aktif
        metadata_kolom = []
        try:
            # Ambil periode aktif terbaru
            cursor.execute("""
                SELECT periode_metadata_aktif 
                FROM pengaturan_jadwal 
                ORDER BY id DESC LIMIT 1
            """)
            periode_result = cursor.fetchone()
            periode_aktif = periode_result['periode_metadata_aktif'] if periode_result else 1
            
            # Untuk interval bulanan, gunakan metadata yang sudah diambil di atas
            if interval_tipe == 'bulanan' and 'metadata_kolom_bulanan' in locals():
                metadata_kolom = metadata_kolom_bulanan
            else:
                cursor.execute('''
                    SELECT urutan_kolom, nama_kolom, deskripsi, tanggal_mulai, tanggal_selesai
                    FROM metadata_kolom_penilaian 
                    WHERE jadwal_id = (SELECT id FROM pengaturan_jadwal ORDER BY id DESC LIMIT 1)
                    AND interval_tipe = %s AND periode_interval_id = %s
                    ORDER BY urutan_kolom ASC
                ''', (interval_tipe, periode_aktif))
                metadata_kolom = cursor.fetchall()
            
            logger.info(f"Metadata kolom untuk interval {interval_tipe}, periode {periode_aktif}: {len(metadata_kolom)} items")
            
            # Convert datetime objects to strings for JSON serialization
            for meta in metadata_kolom:
                if meta.get('tanggal_mulai'):
                    meta['tanggal_mulai'] = meta['tanggal_mulai'].strftime('%Y-%m-%d')
                if meta.get('tanggal_selesai'):
                    meta['tanggal_selesai'] = meta['tanggal_selesai'].strftime('%Y-%m-%d')
                    
        except Exception as e:
            logger.error(f"Error getting metadata kolom: {str(e)}")
            metadata_kolom = []
        
        cursor.close()
        
        return {
            'current_sesi': current_sesi,
            'total_sesi': total_sesi,
            'sessions_available': sessions_available,
            'jam_list': jam_list,
            'interval_tipe': interval_tipe,
            'interval_nilai': interval_nilai,
            'jadwal_mulai': mulai,
            'jadwal_selesai': selesai,
            'metadata_kolom': metadata_kolom  # PERBAIKAN: Tambahkan metadata kolom
        }
        
    except Exception as e:
        logger.error(f"Error getting session info: {str(e)}")
        return None

@pembimbing_bp.route('/get_penilaian_data_with_sesi')
def get_penilaian_data_with_sesi():
    """API untuk mengambil data penilaian dengan info sesi yang sudah diperbaiki"""
    if 'user_type' not in session or session['user_type'] != 'pembimbing':
        return jsonify({'success': False, 'message': 'Anda harus login sebagai pembimbing!'})
    
    try:
        nim = request.args.get('nim')
        if not nim:
            return jsonify({'success': False, 'message': 'NIM mahasiswa tidak ditemukan!'})
        
        cursor = get_app_functions()['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Ambil data proposal dengan join ke tabel mahasiswa untuk nama
        cursor.execute('''
            SELECT p.id as proposal_id, m.nama_ketua as nama_mahasiswa, p.nim, p.judul_usaha 
            FROM proposal p 
            JOIN mahasiswa m ON p.nim = m.nim
            WHERE p.nim = %s AND p.dosen_pembimbing = %s
        ''', (nim, session['nama']))
        
        proposal = cursor.fetchone()
        if not proposal:
            return jsonify({'success': False, 'message': 'Data proposal tidak ditemukan!'})
        
        # Ambil ID pembimbing
        cursor.execute('SELECT id FROM pembimbing WHERE nama = %s', (session['nama'],))
        pembimbing_data = cursor.fetchone()
        if not pembimbing_data:
            return jsonify({'success': False, 'message': 'Data pembimbing tidak ditemukan!'})
        
        id_pembimbing = pembimbing_data['id']
        
        # PERBAIKAN: Ambil pertanyaan berdasarkan konteks
        # Jika ada data penilaian, ambil semua pertanyaan yang punya data (aktif + nonaktif)
        # Jika belum ada data, ambil hanya pertanyaan aktif
        
        # Cek apakah sudah ada data penilaian
        cursor.execute('''
            SELECT pm.id as penilaian_id
            FROM penilaian_mahasiswa pm 
            WHERE pm.id_proposal = %s AND pm.id_pembimbing = %s
        ''', (proposal['proposal_id'], id_pembimbing))
        
        existing_penilaian = cursor.fetchone()
        pertanyaan_list = []  # Inisialisasi pertanyaan_list
        
        if existing_penilaian:
            # PERBAIKAN: Cek apakah ini data lama atau data baru berdasarkan waktu pembuatan penilaian
            # Ambil waktu pembuatan penilaian terlebih dahulu
            cursor.execute('''
                SELECT created_at FROM penilaian_mahasiswa 
                WHERE id = %s
            ''', (existing_penilaian['penilaian_id'],))
            penilaian_created = cursor.fetchone()
            
            if penilaian_created and penilaian_created['created_at']:
                # Cek apakah ini data lama (dibuat sebelum jadwal mulai) atau data baru (dibuat setelah jadwal mulai)
                cursor.execute('''
                    SELECT pembimbing_nilai_mulai 
                    FROM pengaturan_jadwal 
                    ORDER BY id DESC 
                    LIMIT 1
                ''')
                jadwal = cursor.fetchone()
                
                if jadwal and jadwal['pembimbing_nilai_mulai']:
                    jadwal_mulai = jadwal['pembimbing_nilai_mulai']
                    penilaian_created_time = penilaian_created['created_at']
                    
                    if penilaian_created_time < jadwal_mulai:
                        # DATA LAMA: Gunakan snapshot berdasarkan waktu pembuatan penilaian
                        cursor.execute('''
                            SELECT DISTINCT p.id, p.kategori, p.pertanyaan, p.bobot, p.skor_maksimal, p.status
                            FROM pertanyaan_penilaian_mahasiswa p
                            INNER JOIN detail_penilaian_mahasiswa d ON p.id = d.id_pertanyaan
                            WHERE d.id_penilaian_mahasiswa = %s 
                            AND p.created_at <= %s
                            ORDER BY p.kategori ASC, p.created_at ASC
                        ''', (existing_penilaian['penilaian_id'], penilaian_created['created_at']))
                        pertanyaan_list = cursor.fetchall()
                        logger.info(f"DATA LAMA: Menggunakan snapshot pertanyaan berdasarkan created_at: {penilaian_created['created_at']}")
                    else:
                        # DATA BARU: Gunakan logic snapshot jadwal (seperti penilaian baru)
                        pertanyaan_raw = get_pertanyaan_for_evaluation(cursor)
                        # Convert hasil ke format yang diharapkan
                        formatted_pertanyaan = []
                        for p in pertanyaan_raw:
                            formatted_pertanyaan.append({
                                'id': p['id'],
                                'kategori': p['kategori'], 
                                'pertanyaan': p['pertanyaan'],
                                'bobot': p['bobot'],
                                'skor_maksimal': p['skor_maksimal'],
                                'status': p['status']
                            })
                        pertanyaan_list = formatted_pertanyaan
                        logger.info(f"DATA BARU: Menggunakan logic snapshot jadwal untuk penilaian created: {penilaian_created['created_at']}")
                else:
                    # Tidak ada jadwal, gunakan semua pertanyaan yang punya data
                    cursor.execute('''
                        SELECT DISTINCT p.id, p.kategori, p.pertanyaan, p.bobot, p.skor_maksimal, p.status
                        FROM pertanyaan_penilaian_mahasiswa p
                        INNER JOIN detail_penilaian_mahasiswa d ON p.id = d.id_pertanyaan
                        WHERE d.id_penilaian_mahasiswa = %s
                        ORDER BY p.kategori ASC, p.created_at ASC
                    ''', (existing_penilaian['penilaian_id'],
                    ))
                pertanyaan_list = cursor.fetchall()
                        
                logger.warning(f"Tidak ada jadwal: Menggunakan semua pertanyaan untuk penilaian_id: {existing_penilaian['penilaian_id']}")
            else:
                 # Fallback ke semua pertanyaan yang punya data jika tidak ada created_at
                cursor.execute('''
                        SELECT DISTINCT p.id, p.kategori, p.pertanyaan, p.bobot, p.skor_maksimal, p.status
                        FROM pertanyaan_penilaian_mahasiswa p
                        INNER JOIN detail_penilaian_mahasiswa d ON p.id = d.id_pertanyaan
                        WHERE d.id_penilaian_mahasiswa = %s
                        ORDER BY p.kategori ASC, p.created_at ASC
                    ''', (existing_penilaian['penilaian_id'],))
                pertanyaan_list = cursor.fetchall()
                logger.warning(f"Fallback ke semua pertanyaan untuk penilaian_id: {existing_penilaian['penilaian_id']}")
        else:
            # Untuk penilaian baru - gunakan SEMUA pertanyaan aktif saat ini (termasuk yang baru)
            # Data baru dibuat setelah jadwal mulai, jadi berhak mendapatkan semua pertanyaan
            cursor.execute('''
                SELECT id, kategori, pertanyaan, bobot, skor_maksimal, status
                FROM pertanyaan_penilaian_mahasiswa 
                WHERE status = 'Aktif'
                ORDER BY FIELD(kategori, 'pondasi-bisnis', 'pelanggan', 'produk-jasa', 'pemasaran', 'penjualan', 'keuntungan', 'sdm', 'sistem-bisnis'), created_at ASC
            ''')
            pertanyaan_raw = cursor.fetchall()
            
            # Convert hasil ke format yang diharapkan
            formatted_pertanyaan = []
            for p in pertanyaan_raw:
                formatted_pertanyaan.append({
                    'id': p['id'],
                    'kategori': p['kategori'], 
                    'pertanyaan': p['pertanyaan'],
                    'bobot': p['bobot'],
                    'skor_maksimal': p['skor_maksimal'],
                    'status': p['status']
                })
            # Set pertanyaan_list untuk digunakan selanjutnya
            pertanyaan_list = formatted_pertanyaan
            logger.info(f"DATA BARU: Menggunakan semua pertanyaan aktif ({len(pertanyaan_list)} pertanyaan) untuk penilaian baru")
        
        # Helper function untuk nama kategori yang user-friendly
        def get_kategori_display_name(kategori):
            kategori_map = {
                'pondasi-bisnis': 'A. Pondasi Bisnis',
                'pelanggan': 'B. Pelanggan',
                'produk-jasa': 'C. Produk/Jasa',
                'pemasaran': 'D. Pemasaran',
                'penjualan': 'E. Penjualan',
                'keuntungan': 'F. Keuntungan',
                'sdm': 'G. Sumber Daya Manusia',
                'sistem-bisnis': 'H. Sistem Bisnis'
            }
            return kategori_map.get(kategori, kategori.title())
        
        # Helper function untuk urutan kategori yang benar
        def get_kategori_order(kategori):
            kategori_order = {
                'pondasi-bisnis': 1,
                'pelanggan': 2,
                'produk-jasa': 3,
                'pemasaran': 4,
                'penjualan': 5,
                'keuntungan': 6,
                'sdm': 7,
                'sistem-bisnis': 8
            }
            return kategori_order.get(kategori, 999)
        
        # Organize pertanyaan per kategori
        kategori_data = {}
        for pertanyaan in pertanyaan_list:
            kategori = pertanyaan['kategori']
            if kategori not in kategori_data:
                kategori_data[kategori] = {
                    'nama_kategori': get_kategori_display_name(kategori),
                    'pertanyaan_list': [],
                    'total_bobot_kategori': 0,
                    'total_nilai_kategori': 0
                }
            
            kategori_data[kategori]['pertanyaan_list'].append(pertanyaan)
            kategori_data[kategori]['total_bobot_kategori'] += float(pertanyaan['bobot'])
        
        # Convert ke list dan urutkan
        pertanyaan_organized = []
        for kategori, kategori_info in kategori_data.items():
            pertanyaan_organized.append({
                'kategori': kategori,
                'nama_kategori': kategori_info['nama_kategori'],
                'total_bobot_kategori': kategori_info['total_bobot_kategori'],
                'total_nilai_kategori': kategori_info['total_nilai_kategori'],
                'pertanyaan_list': kategori_info['pertanyaan_list']
            })
        
        # Urutkan berdasarkan urutan kategori yang benar
        pertanyaan_organized.sort(key=lambda x: get_kategori_order(x['kategori']))

        # Fallback ekstra: jika karena suatu sebab pertanyaan_list kosong, muat pertanyaan aktif saat ini
        if not pertanyaan_list:
            try:
                cursor.execute('''
                    SELECT id, kategori, pertanyaan, bobot, skor_maksimal, status
                    FROM pertanyaan_penilaian_mahasiswa 
                    WHERE status = 'Aktif'
                    ORDER BY FIELD(kategori, 'pondasi-bisnis', 'pelanggan', 'produk-jasa', 'pemasaran', 'penjualan', 'keuntungan', 'sdm', 'sistem-bisnis'), created_at ASC
                ''')
                backup_rows = cursor.fetchall()
                if backup_rows:
                    pertanyaan_list = [
                        {
                            'id': p['id'],
                            'kategori': p['kategori'],
                            'pertanyaan': p['pertanyaan'],
                            'bobot': p['bobot'],
                            'skor_maksimal': p['skor_maksimal'],
                            'status': p['status']
                        } for p in backup_rows
                    ]
                    # Susun kembali organized
                    kategori_data = {}
                    for pertanyaan in pertanyaan_list:
                        kategori = pertanyaan['kategori']
                        if kategori not in kategori_data:
                            kategori_data[kategori] = {
                                'nama_kategori': get_kategori_display_name(kategori),
                                'pertanyaan_list': [],
                                'total_bobot_kategori': 0,
                                'total_nilai_kategori': 0
                            }
                        kategori_data[kategori]['pertanyaan_list'].append(pertanyaan)
                        kategori_data[kategori]['total_bobot_kategori'] += float(pertanyaan['bobot'])
                    pertanyaan_organized = []
                    for kategori, kategori_info in kategori_data.items():
                        pertanyaan_organized.append({
                            'kategori': kategori,
                            'nama_kategori': kategori_info['nama_kategori'],
                            'total_bobot_kategori': kategori_info['total_bobot_kategori'],
                            'total_nilai_kategori': kategori_info['total_nilai_kategori'],
                            'pertanyaan_list': kategori_info['pertanyaan_list']
                        })
                    pertanyaan_organized.sort(key=lambda x: get_kategori_order(x['kategori']))
                    logger.warning("FALLBACK SNAPSHOT: pertanyaan_list kosong, menggunakan pertanyaan aktif saat ini")
            except Exception as e:
                logger.error(f"FALLBACK SNAPSHOT ERROR: {str(e)}")
        
        # Gunakan fungsi baru yang sudah diperbaiki
        jadwal_info = get_session_info_for_ui(id_pembimbing, nim)
        
        if not jadwal_info:
            return jsonify({'success': False, 'message': 'Gagal mendapatkan informasi jadwal!'})
        
        # Gunakan existing_penilaian yang sudah diambil di atas
        penilaian_data = existing_penilaian
        
        # Ambil skor per sesi jika ada
        skor_per_sesi = {}
        if penilaian_data:
            # Pastikan kolom baru ada
            try:
                cursor.execute("SHOW COLUMNS FROM detail_penilaian_mahasiswa LIKE 'sesi_penilaian'")
                if not cursor.fetchone():
                    cursor.execute("ALTER TABLE detail_penilaian_mahasiswa ADD COLUMN sesi_penilaian INT NOT NULL DEFAULT 1")
                    cursor.execute("ALTER TABLE detail_penilaian_mahasiswa ADD COLUMN is_locked TINYINT(1) NOT NULL DEFAULT 0")
                    cursor.execute("ALTER TABLE detail_penilaian_mahasiswa ADD COLUMN tanggal_input DATETIME NULL")
                    get_app_functions()['mysql'].connection.commit()
            except Exception:
                pass
            
            cursor.execute('''
                SELECT dpm.id_pertanyaan, dpm.skor, dpm.nilai, 
                       COALESCE(dpm.sesi_penilaian, 1) as sesi_penilaian,
                       COALESCE(dpm.is_locked, 0) as is_locked,
                       dpm.tanggal_input
                FROM detail_penilaian_mahasiswa dpm
                WHERE dpm.id_penilaian_mahasiswa = %s
                ORDER BY dpm.sesi_penilaian ASC
            ''', (penilaian_data['penilaian_id'],))
            
            detail_list = cursor.fetchall()
            
            for detail in detail_list:
                pertanyaan_id = detail['id_pertanyaan']
                sesi = detail['sesi_penilaian']
                
                if pertanyaan_id not in skor_per_sesi:
                    skor_per_sesi[pertanyaan_id] = {}
                
                skor_per_sesi[pertanyaan_id][sesi] = {
                    'skor': detail['skor'],
                    'nilai': float(detail['nilai']),
                    'is_locked': bool(detail['is_locked']),
                    'tanggal_input': detail['tanggal_input'].strftime('%Y-%m-%d %H:%M:%S') if detail['tanggal_input'] else None
                }
        
        # Gunakan informasi sesi dari jadwal_info yang sudah diperbaiki
        current_sesi = jadwal_info.get('current_sesi', 1)
        total_sesi = jadwal_info.get('total_sesi', 1)
        sessions_available = jadwal_info.get('sessions_available', [1])
        
        # Cek status jadwal untuk frontend
        schedule_active = is_schedule_active()
        
        # Cek interval check untuk sesi yang sedang aktif
        interval_check = check_pembimbing_penilaian_interval(id_pembimbing, nim)
        
        # Deteksi sesi terlewat dan beri pilihan auto-fill
        sesi_terlewat = []
        
        # ✅ PERBAIKAN: Auto-fill sesi terlewat dengan logika yang lebih jelas
        if current_sesi > 1:
            logger.info(f"CHECKING AUTO-FILL: current_sesi={current_sesi}, schedule_active={schedule_active}")
            
            # Tentukan range sesi yang perlu dicek
            if schedule_active:
                # ✅ PERBAIKAN: Jadwal masih aktif: cek sesi yang sudah lewat (1 sampai current_sesi - 1)
                range_sesi = range(1, current_sesi)
                logger.info(f"JADWAL AKTIF: Cek sesi terlewat 1-{current_sesi-1}, sesi {current_sesi} sedang berlangsung")
            else:
                # ✅ PERBAIKAN: Jadwal sudah selesai: cek SEMUA sesi yang seharusnya ada (1 sampai current_sesi)
                range_sesi = range(1, current_sesi + 1)
                logger.info(f"JADWAL SELESAI: Cek semua sesi yang seharusnya ada 1-{current_sesi}")
            
            for sesi_check in range_sesi:
                # Cek apakah sesi ini kosong untuk semua pertanyaan
                sesi_has_data = False
                for pertanyaan in pertanyaan_list:
                    if str(pertanyaan['id']) in skor_per_sesi and sesi_check in skor_per_sesi[str(pertanyaan['id'])]:
                        sesi_has_data = True
                        break
                
                if not sesi_has_data:
                    sesi_terlewat.append(sesi_check)
                    logger.info(f"Sesi {sesi_check} terdeteksi kosong untuk mahasiswa {nim}")
            
            if sesi_terlewat:
                logger.warning(f"DETEKSI SESI TERLEWAT: Sesi {sesi_terlewat} kosong. Schedule active: {schedule_active}")
                
                # ✅ PERBAIKAN: Auto-fill otomatis jika jadwal sudah selesai
                if not schedule_active and sesi_terlewat:
                    logger.info(f"AUTO-FILL OTOMATIS: Jadwal sudah selesai, mengisi sesi {sesi_terlewat} dengan skor 0")
                    success_count = 0
                    for sesi_kosong in sesi_terlewat:
                        if auto_fill_sesi_terlewat_helper(id_pembimbing, nim, sesi_kosong):
                            success_count += 1
                    
                    logger.info(f"AUTO-FILL COMPLETED: Berhasil mengisi {success_count} dari {len(sesi_terlewat)} sesi")
                    # Refresh data setelah autofill
                    sesi_terlewat = []  # Clear karena sudah diisi
                else:
                    logger.info(f"AUTO-FILL SKIPPED: Jadwal masih aktif ({schedule_active}), auto-fill tidak dijalankan")
        else:
            logger.info(f"AUTO-FILL SKIPPED: current_sesi={current_sesi} <= 1, tidak perlu auto-fill")
        
        # ✅ PERBAIKAN: Ambil metadata per sesi berdasarkan metadata_kolom_id yang tersimpan
        metadata_per_sesi = {}
        interval_tipe_penilaian = None  # Untuk menyimpan interval tipe dari penilaian yang ada
        
        if penilaian_data:
            # Ambil metadata_kolom_id yang digunakan dalam penilaian ini
            cursor.execute('''
                SELECT DISTINCT metadata_kolom_id, sesi_penilaian
                FROM detail_penilaian_mahasiswa 
                WHERE id_penilaian_mahasiswa = %s 
                AND metadata_kolom_id IS NOT NULL
                ORDER BY sesi_penilaian
            ''', (penilaian_data['penilaian_id'],))
            
            metadata_ids = cursor.fetchall()
            
            if metadata_ids:
                # Ambil periode aktif terbaru
                cursor.execute("""
                    SELECT periode_metadata_aktif 
                    FROM pengaturan_jadwal 
                    ORDER BY id DESC LIMIT 1
                """)
                periode_result = cursor.fetchone()
                periode_aktif = periode_result['periode_metadata_aktif'] if periode_result else 1
                
                # Ambil metadata kolom berdasarkan metadata_kolom_id yang tersimpan dan periode aktif
                metadata_ids_list = [m['metadata_kolom_id'] for m in metadata_ids]
                cursor.execute("""
                    SELECT id, nama_kolom, urutan_kolom, interval_tipe 
                    FROM metadata_kolom_penilaian 
                    WHERE id IN ({}) AND periode_interval_id = %s
                    ORDER BY urutan_kolom
                """.format(','.join(['%s'] * len(metadata_ids_list))), metadata_ids_list + [periode_aktif])
                
                metadata_kolom = cursor.fetchall()
                
                # ✅ PERBAIKAN: Map metadata ke sesi berdasarkan sesi_penilaian, bukan urutan_kolom
                for metadata in metadata_kolom:
                    # Cari sesi_penilaian yang menggunakan metadata_kolom_id ini
                    cursor.execute('''
                        SELECT DISTINCT sesi_penilaian 
                        FROM detail_penilaian_mahasiswa 
                        WHERE id_penilaian_mahasiswa = %s 
                        AND metadata_kolom_id = %s
                    ''', (penilaian_data['penilaian_id'], metadata['id']))
                    
                    sesi_data = cursor.fetchone()
                    if sesi_data:
                        sesi = sesi_data['sesi_penilaian']
                        metadata_per_sesi[sesi] = {
                            'nama_kolom': metadata['nama_kolom'],
                            'urutan_kolom': metadata['urutan_kolom'],
                            'interval_tipe': metadata['interval_tipe'],
                            'metadata_kolom_id': metadata['id']
                        }
                        
                        # ✅ PERBAIKAN: Simpan interval tipe dari penilaian yang ada
                        if interval_tipe_penilaian is None:
                            interval_tipe_penilaian = metadata['interval_tipe']
                        elif interval_tipe_penilaian != metadata['interval_tipe']:
                            # Jika ada perbedaan interval tipe, gunakan yang pertama
                            logger.warning(f"Perbedaan interval tipe dalam penilaian: {interval_tipe_penilaian} vs {metadata['interval_tipe']}, menggunakan {interval_tipe_penilaian}")
                
                logger.info(f"Metadata per sesi untuk penilaian {penilaian_data['penilaian_id']}: {metadata_per_sesi}")
                logger.info(f"Interval tipe penilaian: {interval_tipe_penilaian}")
            else:
                logger.warning(f"Tidak ada metadata_kolom_id untuk penilaian {penilaian_data['penilaian_id']}")
        
        # ✅ PERBAIKAN: Validasi interval tipe - hanya izinkan input jika sesuai dengan penilaian yang ada
        if interval_tipe_penilaian and jadwal_info.get('interval_tipe') != interval_tipe_penilaian:
            logger.warning(f"INTERVAL TIPE TIDAK SESUAI: Jadwal saat ini {jadwal_info.get('interval_tipe')} vs penilaian yang ada {interval_tipe_penilaian}")
            # Blok input jika interval tipe tidak sesuai
            interval_check = {
                'bisa_nilai': False,
                'pesan': f'Interval tipe tidak sesuai. Penilaian ini menggunakan {interval_tipe_penilaian}, sedangkan jadwal saat ini {jadwal_info.get("interval_tipe")}.'
            }
            schedule_active = False
        
        cursor.close()
        
        # Siapkan response data
        response_data = {
            'success': True,
            'proposal': proposal,
            'pertanyaan': pertanyaan_list,  # Keep for backward compatibility
            'pertanyaan_organized': pertanyaan_organized,  # New structured data
            'skor_per_sesi': skor_per_sesi,
            'current_sesi': current_sesi,
            'total_sesi': total_sesi,
            'sessions_available': sessions_available,
            'jadwal_info': jadwal_info,
            'schedule_active': schedule_active,
            'interval_check': interval_check,
            'penilaian_id': penilaian_data['penilaian_id'] if penilaian_data else None,
            'sesi_terlewat': sesi_terlewat,
            'metadata_per_sesi': metadata_per_sesi  # ✅ PERBAIKAN: Tambahkan metadata per sesi
        }
        
        logger.info(f"DEBUG RESPONSE: {response_data}")
        
        return jsonify(response_data)
        
    except Exception as e:
        logger.error(f"Error in get_penilaian_data_with_sesi: {str(e)}")
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@pembimbing_bp.route('/simpan_skor_sesi', methods=['POST'])
def simpan_skor_sesi():
    """API untuk menyimpan skor per sesi dan mengunci skor tersebut"""
    if 'user_type' not in session or session['user_type'] != 'pembimbing':
        return jsonify({'success': False, 'message': 'Anda harus login sebagai pembimbing!'})
    
    try:
        nim_mahasiswa = request.form.get('nim_mahasiswa')
        id_pertanyaan = request.form.get('id_pertanyaan')
        skor = request.form.get('skor')
        sesi_penilaian = request.form.get('sesi_penilaian')
        
        if not all([nim_mahasiswa, id_pertanyaan, skor, sesi_penilaian]):
            return jsonify({'success': False, 'message': 'Data tidak lengkap!'})
        
        skor = int(skor)
        sesi_penilaian = int(sesi_penilaian)
        
        cursor = get_app_functions()['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Ambil ID pembimbing
        cursor.execute('SELECT id FROM pembimbing WHERE nama = %s', (session['nama'],))
        pembimbing_data = cursor.fetchone()
        if not pembimbing_data:
            return jsonify({'success': False, 'message': 'Data pembimbing tidak ditemukan!'})
        
        id_pembimbing = pembimbing_data['id']
        
        # Validasi interval penilaian - hanya cek jika ini adalah sesi baru
        # Jika sesi yang sama, tidak perlu cek interval
        if sesi_penilaian > 1:  # Hanya cek interval untuk sesi 2 ke atas
            interval_check = check_pembimbing_penilaian_interval(id_pembimbing, nim_mahasiswa)
            if not interval_check['bisa_nilai']:
                return jsonify({'success': False, 'message': interval_check['pesan']})
        
        # Ambil data proposal
        cursor.execute('''
            SELECT p.id as proposal_id 
            FROM proposal p 
            WHERE p.nim = %s AND p.dosen_pembimbing = %s
        ''', (nim_mahasiswa, session['nama']))
        
        proposal = cursor.fetchone()
        if not proposal:
            return jsonify({'success': False, 'message': 'Data proposal tidak ditemukan!'})
        
        # Tambahan validasi: cek apakah sudah ada data untuk pertanyaan dan sesi ini
        # Ini untuk mencegah duplicate insert yang tidak perlu
        cursor.execute('''
            SELECT dp.id, dp.skor, dp.is_locked
            FROM detail_penilaian_mahasiswa dp
            JOIN penilaian_mahasiswa pm ON dp.id_penilaian_mahasiswa = pm.id
            WHERE pm.id_proposal = %s AND pm.id_pembimbing = %s 
            AND dp.id_pertanyaan = %s AND dp.sesi_penilaian = %s
        ''', (proposal['proposal_id'], id_pembimbing, id_pertanyaan, sesi_penilaian))
        
        existing_detail = cursor.fetchone()
        
        # Jika sudah ada dan sudah dikunci, return error
        if existing_detail and existing_detail['is_locked']:
            return jsonify({
                'success': False, 
                'message': f'Skor untuk pertanyaan ini pada sesi {sesi_penilaian} sudah dikunci dan tidak bisa diubah!'
            })
        
        # Log untuk debugging
        if existing_detail:
            logger.info(f"EXISTING DATA: Pertanyaan {id_pertanyaan}, sesi {sesi_penilaian} sudah ada dengan skor {existing_detail['skor']}, akan diupdate")
        else:
            logger.info(f"NEW DATA: Pertanyaan {id_pertanyaan}, sesi {sesi_penilaian} belum ada, akan diinsert")
        
        # Ambil data pertanyaan untuk validasi skor
        cursor.execute('''
            SELECT bobot, skor_maksimal 
            FROM pertanyaan_penilaian_mahasiswa 
            WHERE id = %s AND status = 'Aktif'
        ''', (id_pertanyaan,))
        
        pertanyaan = cursor.fetchone()
        if not pertanyaan:
            return jsonify({'success': False, 'message': 'Pertanyaan tidak ditemukan atau tidak aktif!'})
        
        if skor > pertanyaan['skor_maksimal']:
            return jsonify({'success': False, 'message': f'Skor tidak boleh melebihi {pertanyaan["skor_maksimal"]}!'})
        
        # Pastikan kolom jadwal_id ada di tabel penilaian_mahasiswa
        try:
            cursor.execute("SHOW COLUMNS FROM penilaian_mahasiswa LIKE 'jadwal_id'")
            if not cursor.fetchone():
                cursor.execute("ALTER TABLE penilaian_mahasiswa ADD COLUMN jadwal_id INT NULL")
                get_app_functions()['mysql'].connection.commit()
        except Exception:
            pass
        
        # Cek atau buat penilaian_mahasiswa
        cursor.execute('''
            SELECT id FROM penilaian_mahasiswa 
            WHERE id_proposal = %s AND id_pembimbing = %s
        ''', (proposal['proposal_id'], id_pembimbing))
        
        penilaian_data = cursor.fetchone()
        
        if not penilaian_data:
            # Dapatkan jadwal_id yang sedang aktif
            current_jadwal_id = None
            try:
                cursor.execute('SELECT id FROM pengaturan_jadwal ORDER BY id DESC LIMIT 1')
                jadwal_result = cursor.fetchone()
                if jadwal_result:
                    current_jadwal_id = jadwal_result['id']
            except:
                pass  # Jika tidak ada jadwal, jadwal_id akan NULL
            
            # Buat header penilaian baru dengan jadwal_id
            cursor.execute('''
                INSERT INTO penilaian_mahasiswa (id_proposal, id_pembimbing, jadwal_id, nilai_akhir, komentar_pembimbing, status)
                VALUES (%s, %s, %s, 0, '', 'Draft')
            ''', (proposal['proposal_id'], id_pembimbing, current_jadwal_id))
            
            penilaian_id = cursor.lastrowid
        else:
            penilaian_id = penilaian_data['id']
        
        # Pastikan kolom baru ada di detail_penilaian_mahasiswa
        try:
            cursor.execute("SHOW COLUMNS FROM detail_penilaian_mahasiswa LIKE 'sesi_penilaian'")
            if not cursor.fetchone():
                cursor.execute("ALTER TABLE detail_penilaian_mahasiswa ADD COLUMN sesi_penilaian INT NOT NULL DEFAULT 1")
                cursor.execute("ALTER TABLE detail_penilaian_mahasiswa ADD COLUMN is_locked TINYINT(1) NOT NULL DEFAULT 0")
                cursor.execute("ALTER TABLE detail_penilaian_mahasiswa ADD COLUMN tanggal_input DATETIME NULL")
                cursor.execute("ALTER TABLE detail_penilaian_mahasiswa ADD COLUMN metadata_kolom_id INT NULL")
                get_app_functions()['mysql'].connection.commit()
        except Exception:
            pass
        
        # Hitung nilai - konversi ke float untuk menghindari decimal error
        skor_float = float(skor)
        skor_maksimal_float = float(pertanyaan['skor_maksimal'])
        bobot_float = float(pertanyaan['bobot'])
        nilai = (skor_float / skor_maksimal_float) * bobot_float
        
        # ✅ PERBAIKAN: Ambil metadata kolom untuk sesi ini berdasarkan interval yang aktif
        metadata_kolom_id = None
        try:
            # Ambil interval tipe yang aktif dari pengaturan jadwal
            cursor.execute("""
                SELECT pembimbing_interval_tipe, periode_metadata_aktif
                FROM pengaturan_jadwal 
                ORDER BY id DESC LIMIT 1
            """)
            
            jadwal_result = cursor.fetchone()
            if jadwal_result:
                interval_tipe_aktif = jadwal_result['pembimbing_interval_tipe']
                periode_aktif = jadwal_result['periode_metadata_aktif']
                
                # ✅ PERBAIKAN: Ambil metadata yang sesuai dengan jadwal aktif (bukan yang terbaru)
                cursor.execute("""
                    SELECT mkp.id 
                    FROM metadata_kolom_penilaian mkp
                    JOIN pengaturan_jadwal pj ON mkp.jadwal_id = pj.id
                    WHERE mkp.urutan_kolom = %s 
                    AND mkp.interval_tipe = %s 
                    AND mkp.periode_interval_id = %s
                    AND pj.id = (
                        SELECT id FROM pengaturan_jadwal 
                        WHERE pembimbing_nilai_mulai IS NOT NULL 
                        AND pembimbing_nilai_selesai IS NOT NULL
                        ORDER BY id DESC LIMIT 1
                    )
                    ORDER BY mkp.created_at DESC LIMIT 1
                """, (sesi_penilaian, interval_tipe_aktif, periode_aktif))
                
                metadata_result = cursor.fetchone()
                if metadata_result:
                    metadata_kolom_id = metadata_result['id']
                    logger.info(f"Metadata kolom dipilih: ID {metadata_kolom_id} untuk sesi {sesi_penilaian}, interval {interval_tipe_aktif}")
                else:
                    logger.warning(f"Tidak ada metadata yang sesuai untuk sesi {sesi_penilaian}, interval {interval_tipe_aktif}, periode {periode_aktif}")
            else:
                logger.warning("Tidak ada pengaturan jadwal yang aktif")
                
        except Exception as e:
            logger.warning(f"Tidak bisa ambil metadata kolom untuk sesi {sesi_penilaian}: {str(e)}")
        
        # Gunakan INSERT ... ON DUPLICATE KEY UPDATE untuk mencegah duplicate entry error
        # Ini akan handle kasus race condition atau multiple request
        try:
            cursor.execute('''
                INSERT INTO detail_penilaian_mahasiswa 
                (id_penilaian_mahasiswa, id_pertanyaan, skor, nilai, sesi_penilaian, is_locked, tanggal_input, metadata_kolom_id)
                VALUES (%s, %s, %s, %s, %s, 1, NOW(), %s)
                ON DUPLICATE KEY UPDATE 
                    skor = VALUES(skor),
                    nilai = VALUES(nilai),
                    is_locked = VALUES(is_locked),
                    tanggal_input = VALUES(tanggal_input),
                    metadata_kolom_id = VALUES(metadata_kolom_id)
            ''', (penilaian_id, id_pertanyaan, skor, nilai, sesi_penilaian, metadata_kolom_id))
            
            # Log apakah ini insert baru atau update existing
            if cursor.rowcount == 1:
                logger.info(f"INSERT NEW: Skor baru berhasil disimpan untuk pertanyaan {id_pertanyaan}, sesi {sesi_penilaian}")
            else:
                logger.info(f"UPDATE EXISTING: Skor berhasil diupdate untuk pertanyaan {id_pertanyaan}, sesi {sesi_penilaian}")
                
        except Exception as insert_error:
            logger.error(f"Error saat INSERT/UPDATE detail penilaian: {str(insert_error)}")
            logger.error(f"Data yang dicoba insert: penilaian_id={penilaian_id}, pertanyaan_id={id_pertanyaan}, sesi={sesi_penilaian}")
            raise insert_error
        
        # Update nilai_akhir di tabel penilaian_mahasiswa
        update_nilai_akhir_penilaian(cursor, penilaian_id)
        
        # Data riwayat penilaian tersimpan di detail_penilaian_mahasiswa dengan sesi_penilaian dan tanggal_input
        
        get_app_functions()['mysql'].connection.commit()
        cursor.close()
        
        logger.info(f"SUCCESS: Skor berhasil disimpan - NIM: {nim_mahasiswa}, Pertanyaan: {id_pertanyaan}, Skor: {skor}, Sesi: {sesi_penilaian}, Nilai: {nilai}")
        
        return jsonify({
            'success': True,
            'message': 'Skor berhasil disimpan dan dikunci!',
            'nilai': float(nilai)
        })
        
    except Exception as e:
        logger.error(f"Error in simpan_skor_sesi: {str(e)}")
        logger.error(f"Error details - NIM: {nim_mahasiswa}, Pertanyaan: {id_pertanyaan}, Skor: {skor}, Sesi: {sesi_penilaian}")
        
        # Rollback jika ada error
        try:
            get_app_functions()['mysql'].connection.rollback()
            logger.info("Database rollback successful")
        except Exception as rollback_error:
            logger.error(f"Rollback failed: {str(rollback_error)}")
        
        # Berikan pesan error yang lebih user-friendly
        error_message = str(e)
        if "decimal" in error_message.lower():
            error_message = "Error perhitungan nilai. Silakan coba lagi atau hubungi admin."
        elif "connection" in error_message.lower():
            error_message = "Koneksi database error. Silakan coba lagi."
        else:
            error_message = f"Terjadi kesalahan: {error_message}"
            
        return jsonify({'success': False, 'message': error_message})

@pembimbing_bp.route('/auto_fill_sesi_terlewat', methods=['POST'])
def auto_fill_sesi_terlewat():
    """API untuk auto-fill sesi yang terlewat dengan skor 0 menggunakan logika baru"""
    if 'user_type' not in session or session['user_type'] != 'pembimbing':
        return jsonify({'success': False, 'message': 'Anda harus login sebagai pembimbing!'})
    
    try:
        nim_mahasiswa = request.form.get('nim_mahasiswa')
        sesi_list = request.form.get('sesi_list')  # Format: "1,2,3"
        
        if not nim_mahasiswa or not sesi_list:
            return jsonify({'success': False, 'message': 'Parameter tidak lengkap!'})
        
        # Parse sesi list
        try:
            sesi_to_fill = [int(s.strip()) for s in sesi_list.split(',') if s.strip().isdigit()]
        except:
            return jsonify({'success': False, 'message': 'Format sesi tidak valid!'})
        
        if not sesi_to_fill:
            return jsonify({'success': False, 'message': 'Tidak ada sesi untuk diisi!'})
        
        cursor = get_app_functions()['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Ambil ID pembimbing
        cursor.execute('SELECT id FROM pembimbing WHERE nama = %s', (session['nama'],))
        pembimbing_data = cursor.fetchone()
        if not pembimbing_data:
            return jsonify({'success': False, 'message': 'Data pembimbing tidak ditemukan!'})
        
        id_pembimbing = pembimbing_data['id']
        
        # Ambil atau buat data penilaian mahasiswa
        cursor.execute('''
            SELECT pm.id as penilaian_id
            FROM penilaian_mahasiswa pm 
            JOIN proposal p ON pm.id_proposal = p.id 
            WHERE p.nim = %s AND pm.id_pembimbing = %s
        ''', (nim_mahasiswa, id_pembimbing))
        
        penilaian_data = cursor.fetchone()
        
        if not penilaian_data:
            # Buat data penilaian baru
            cursor.execute('SELECT id FROM proposal WHERE nim = %s', (nim_mahasiswa,))
            proposal_data = cursor.fetchone()
            if not proposal_data:
                return jsonify({'success': False, 'message': 'Data proposal tidak ditemukan!'})
            
            cursor.execute('''
                INSERT INTO penilaian_mahasiswa (id_proposal, id_pembimbing, nilai_akhir, komentar_pembimbing, status)
                VALUES (%s, %s, 0, '', 'Draft')
            ''', (proposal_data['id'], id_pembimbing))
            
            penilaian_id = cursor.lastrowid
        else:
            penilaian_id = penilaian_data['penilaian_id']
        
        # Ambil pertanyaan aktif
        cursor.execute('''
            SELECT id, pertanyaan, bobot, skor_maksimal
            FROM pertanyaan_penilaian_mahasiswa 
            WHERE status = 'Aktif'
        ''')
        pertanyaan_list = cursor.fetchall()
        
        filled_count = 0
        
        # Auto-fill setiap sesi dengan skor 0
        for sesi in sesi_to_fill:
            for pertanyaan in pertanyaan_list:
                try:
                    # Cek apakah sudah ada data untuk sesi dan pertanyaan ini
                    cursor.execute('''
                        SELECT id FROM detail_penilaian_mahasiswa 
                        WHERE id_penilaian_mahasiswa = %s AND id_pertanyaan = %s AND sesi_penilaian = %s
                    ''', (penilaian_id, pertanyaan['id'], sesi))
                    
                    existing = cursor.fetchone()
                    
                    if not existing:
                        # ✅ PERBAIKAN: Ambil metadata kolom untuk sesi ini
                        metadata_kolom_id = None
                        try:
                            # Ambil interval tipe yang aktif dari pengaturan jadwal
                            cursor.execute('''
                                SELECT pembimbing_interval_tipe, periode_metadata_aktif
                                FROM pengaturan_jadwal 
                                ORDER BY id DESC LIMIT 1
                            ''')
                            
                            jadwal_result = cursor.fetchone()
                            if jadwal_result:
                                interval_tipe_aktif = jadwal_result['pembimbing_interval_tipe']
                                periode_aktif = jadwal_result['periode_metadata_aktif']
                                
                                # Ambil metadata yang sesuai dengan interval aktif
                                cursor.execute('''
                                    SELECT id FROM metadata_kolom_penilaian 
                                    WHERE urutan_kolom = %s AND interval_tipe = %s AND periode_interval_id = %s
                                    ORDER BY created_at DESC LIMIT 1
                                ''', (sesi, interval_tipe_aktif, periode_aktif))
                                
                                metadata_result = cursor.fetchone()
                                if metadata_result:
                                    metadata_kolom_id = metadata_result['id']
                                else:
                                    logger.warning(f"Tidak ada metadata yang sesuai untuk sesi {sesi}, interval {interval_tipe_aktif}, periode {periode_aktif}")
                            else:
                                logger.warning("Tidak ada pengaturan jadwal yang aktif")
                                
                        except Exception as e:
                            logger.error(f"Tidak bisa ambil metadata kolom untuk sesi {sesi}: {str(e)}")
                        
                        # Insert skor 0 dengan metadata_kolom_id
                        cursor.execute('''
                            INSERT INTO detail_penilaian_mahasiswa 
                            (id_penilaian_mahasiswa, id_pertanyaan, skor, nilai, sesi_penilaian, is_locked, tanggal_input, metadata_kolom_id)
                            VALUES (%s, %s, 0, 0, %s, 1, NOW(), %s)
                        ''', (penilaian_id, pertanyaan['id'], sesi, metadata_kolom_id))
                        
                        filled_count += 1
                        
                except Exception as e:
                    logger.error(f"Error auto-filling sesi {sesi} pertanyaan {pertanyaan['id']}: {str(e)}")
        
        # Update nilai_akhir di tabel penilaian_mahasiswa setelah auto-fill
        update_nilai_akhir_penilaian(cursor, penilaian_id)
        
        # Data riwayat penilaian tersimpan di detail_penilaian_mahasiswa dengan sesi_penilaian dan tanggal_input
        
        get_app_functions()['mysql'].connection.commit()
        cursor.close()
        
        logger.info(f"AUTO-FILL SUCCESS: {filled_count} data terisi untuk sesi {sesi_to_fill} mahasiswa {nim_mahasiswa}")
        
        return jsonify({
            'success': True,
            'message': f'Berhasil mengisi {filled_count} data dengan skor 0 untuk sesi {sesi_to_fill}',
            'filled_count': filled_count,
            'sesi_filled': sesi_to_fill
        })
        
    except Exception as e:
        logger.error(f"Error in auto_fill_sesi_terlewat: {str(e)}")
        try:
            get_app_functions()['mysql'].connection.rollback()
        except:
            pass
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

def get_jadwal_pembimbing_status():
    """Mendapatkan status jadwal penilaian pembimbing"""
    try:
        cursor = get_app_functions()['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Ambil jadwal dari tabel pengaturan_jadwal
        cursor.execute('''
            SELECT 
                pembimbing_nilai_mulai, pembimbing_nilai_selesai
            FROM pengaturan_jadwal 
            ORDER BY id DESC 
            LIMIT 1
        ''')
        
        jadwal = cursor.fetchone()
        if not jadwal:
            return {
                'status': 'tidak ada jadwal',
                'pesan': 'Belum ada jadwal penilaian pembimbing yang diatur',
                'bisa_nilai': False  # Jika tidak ada jadwal, tidak bisa nilai
            }
        
        now = datetime.now()
        mulai = jadwal['pembimbing_nilai_mulai']
        selesai = jadwal['pembimbing_nilai_selesai']
        
        if not mulai or not selesai:
            return {
                'status': 'tidak ada jadwal',
                'pesan': 'Jadwal penilaian pembimbing belum diatur',
                'bisa_nilai': False
            }
        
        # Konversi ke datetime jika string
        if isinstance(mulai, str):
            mulai = datetime.strptime(mulai, '%Y-%m-%d %H:%M:%S')
        if isinstance(selesai, str):
            selesai = datetime.strptime(selesai, '%Y-%m-%d %H:%M:%S')
        
        if now < mulai:
            return {
                'status': 'belum_mulai',
                'pesan': 'Jadwal penilaian pembimbing belum dimulai',
                'bisa_nilai': False,
                'jadwal_mulai': mulai,
                'jadwal_selesai': selesai
            }
        elif now > selesai:
            return {
                'status': 'sudah_selesai',
                'pesan': 'Jadwal penilaian pembimbing sudah berakhir',
                'bisa_nilai': False,
                'jadwal_mulai': mulai,
                'jadwal_selesai': selesai
            }
        else:
            return {
                'status': 'aktif',
                'pesan': 'Jadwal penilaian pembimbing sedang berlangsung',
                'bisa_nilai': True,
                'jadwal_mulai': mulai,
                'jadwal_selesai': selesai
            }
            
    except Exception as e:
        return {'status': 'error', 'pesan': str(e), 'bisa_nilai': False}
    finally:
        if 'cursor' in locals():
            cursor.close()


def auto_fill_sesi_terlewat_helper(id_pembimbing, nim_mahasiswa, sesi_terlewat):
    """Helper function untuk auto-fill sesi yang terlewat dengan skor 0"""
    try:
        cursor = get_app_functions()['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # ✅ PERBAIKAN: Validasi data lama sebelum auto-fill
        cursor.execute('''
            SELECT p.tanggal_buat, p.tanggal_kirim, pm.jadwal_id
            FROM proposal p
            JOIN penilaian_mahasiswa pm ON p.id = pm.id_proposal
            WHERE p.nim = %s
        ''', (nim_mahasiswa,))
        
        proposal_info = cursor.fetchone()
        if proposal_info:
            tanggal_buat = proposal_info['tanggal_buat']
            jadwal_id = proposal_info['jadwal_id']
            
            # ✅ PERBAIKAN: Jika jadwal_id NULL, ini data lama - SKIP auto-fill
            if jadwal_id is None:
                logger.info(f"Data lama terdeteksi - jadwal_id NULL untuk mahasiswa {nim_mahasiswa}, SKIP auto-fill")
                cursor.close()
                return False
            
            # ✅ PERBAIKAN: Ambil jadwal aktif untuk validasi
            cursor.execute('''
                SELECT pembimbing_nilai_mulai, pembimbing_nilai_selesai
                FROM pengaturan_jadwal 
                ORDER BY id DESC 
                LIMIT 1
            ''')
            
            jadwal_info = cursor.fetchone()
            if jadwal_info and jadwal_info['pembimbing_nilai_mulai']:
                waktu_mulai = jadwal_info['pembimbing_nilai_mulai']
                waktu_selesai = jadwal_info['pembimbing_nilai_selesai']
                
                # ✅ PERBAIKAN: Jika proposal dibuat sebelum jadwal aktif, ini data lama - SKIP auto-fill
                if tanggal_buat and tanggal_buat < waktu_mulai:
                    logger.info(f"Data lama terdeteksi - proposal dibuat {tanggal_buat} sebelum jadwal aktif {waktu_mulai}, SKIP auto-fill")
                    cursor.close()
                    return False
                
                # ✅ PERBAIKAN: Jika proposal dibuat setelah jadwal selesai, ini data baru - SKIP auto-fill
                if tanggal_buat and waktu_selesai and tanggal_buat > waktu_selesai:
                    logger.info(f"Data baru terdeteksi - proposal dibuat {tanggal_buat} setelah jadwal selesai {waktu_selesai}, SKIP auto-fill")
                    cursor.close()
                    return False
        
        # PERBAIKAN: Ambil pertanyaan berdasarkan snapshot jadwal mulai
        # Ambil jadwal mulai terlebih dahulu
        cursor.execute('''
            SELECT pembimbing_nilai_mulai
            FROM pengaturan_jadwal 
            ORDER BY id DESC 
            LIMIT 1
        ''')
        
        jadwal = cursor.fetchone()
        
        if jadwal and jadwal['pembimbing_nilai_mulai']:
            jadwal_mulai = jadwal['pembimbing_nilai_mulai']
            
            # Cek apakah ada pertanyaan yang dibuat sebelum jadwal mulai
            cursor.execute('''
                SELECT COUNT(*) as count
                FROM pertanyaan_penilaian_mahasiswa
                WHERE status = 'Aktif' AND created_at <= %s
            ''', (jadwal_mulai,))
            
            count_result = cursor.fetchone()
            pertanyaan_sebelum_jadwal = count_result['count'] if count_result else 0
            
            if pertanyaan_sebelum_jadwal > 0:
                # Gunakan pertanyaan yang aktif pada saat jadwal mulai
                cursor.execute('''
                    SELECT id, bobot, skor_maksimal
                    FROM pertanyaan_penilaian_mahasiswa
                    WHERE status = 'Aktif' AND created_at <= %s
                    ORDER BY id
                ''', (jadwal_mulai,))
                pertanyaan_list = cursor.fetchall()
                logger.info(f"Auto-fill menggunakan snapshot pertanyaan pada jadwal mulai: {jadwal_mulai}")
            else:
                # Fallback: gunakan pertanyaan aktif saat ini
                cursor.execute('''
                    SELECT id, bobot, skor_maksimal
                    FROM pertanyaan_penilaian_mahasiswa
                    WHERE status = 'Aktif'
                    ORDER BY id
                ''')
                pertanyaan_list = cursor.fetchall()
                logger.info(f"Auto-fill fallback: tidak ada pertanyaan sebelum jadwal mulai ({jadwal_mulai}), menggunakan pertanyaan aktif saat ini")
        else:
            # Fallback ke pertanyaan aktif saat ini
            cursor.execute('''
                SELECT id, bobot, skor_maksimal
                FROM pertanyaan_penilaian_mahasiswa
                WHERE status = 'Aktif'
                ORDER BY id
            ''')
            pertanyaan_list = cursor.fetchall()
            logger.info("Auto-fill fallback: tidak ada jadwal, menggunakan pertanyaan aktif saat ini")
        
        if not pertanyaan_list:
            logger.warning(f"Tidak ada pertanyaan penilaian aktif untuk auto-fill sesi {sesi_terlewat}")
            return False
        
        # Ambil atau buat penilaian_mahasiswa
        cursor.execute('''
            SELECT id FROM penilaian_mahasiswa 
            WHERE id_pembimbing = %s AND id_proposal = (
                SELECT id FROM proposal WHERE nim = %s
            )
        ''', (id_pembimbing, nim_mahasiswa))
        
        penilaian_result = cursor.fetchone()
        if penilaian_result:
            penilaian_id = penilaian_result['id']
        else:
            # Buat baru jika belum ada
            cursor.execute('''
                INSERT INTO penilaian_mahasiswa (id_pembimbing, id_proposal)
                SELECT %s, id FROM proposal WHERE nim = %s
            ''', (id_pembimbing, nim_mahasiswa))
            penilaian_id = cursor.lastrowid
        
        # ✅ PERBAIKAN: Ambil metadata kolom untuk sesi ini
        metadata_kolom_id = None
        try:
            # Ambil interval tipe yang aktif dari pengaturan jadwal
            cursor.execute('''
                SELECT pembimbing_interval_tipe, periode_metadata_aktif
                FROM pengaturan_jadwal 
                ORDER BY id DESC LIMIT 1
            ''')
            
            jadwal_result = cursor.fetchone()
            if jadwal_result:
                interval_tipe_aktif = jadwal_result['pembimbing_interval_tipe']
                periode_aktif = jadwal_result['periode_metadata_aktif']
                
                # ✅ PERBAIKAN: Ambil metadata yang sesuai dengan jadwal aktif (bukan yang terbaru)
                cursor.execute('''
                    SELECT mkp.id 
                    FROM metadata_kolom_penilaian mkp
                    JOIN pengaturan_jadwal pj ON mkp.jadwal_id = pj.id
                    WHERE mkp.urutan_kolom = %s 
                    AND mkp.interval_tipe = %s 
                    AND mkp.periode_interval_id = %s
                    AND pj.id = (
                        SELECT id FROM pengaturan_jadwal 
                        WHERE pembimbing_nilai_mulai IS NOT NULL 
                        AND pembimbing_nilai_selesai IS NOT NULL
                        ORDER BY id DESC LIMIT 1
                    )
                    ORDER BY mkp.created_at DESC LIMIT 1
                ''', (sesi_terlewat, interval_tipe_aktif, periode_aktif))
                
                metadata_result = cursor.fetchone()
                if metadata_result:
                    metadata_kolom_id = metadata_result['id']
                    logger.info(f"Metadata kolom dipilih: ID {metadata_kolom_id} untuk sesi {sesi_terlewat}, interval {interval_tipe_aktif}")
                else:
                    logger.warning(f"Tidak ada metadata yang sesuai untuk sesi {sesi_terlewat}, interval {interval_tipe_aktif}, periode {periode_aktif}")
            else:
                logger.warning("Tidak ada pengaturan jadwal yang aktif")
                
        except Exception as e:
            logger.error(f"Tidak bisa ambil metadata kolom untuk sesi {sesi_terlewat}: {str(e)}")
        
        # Auto-fill setiap pertanyaan dengan skor 0
        for pertanyaan in pertanyaan_list:
            cursor.execute('''
                INSERT INTO detail_penilaian_mahasiswa 
                (id_penilaian_mahasiswa, id_pertanyaan, skor, nilai, sesi_penilaian, is_locked, tanggal_input, metadata_kolom_id)
                VALUES (%s, %s, 0, 0, %s, 1, NOW(), %s)
                ON DUPLICATE KEY UPDATE
                skor = 0, nilai = 0, is_locked = 1, tanggal_input = NOW(), metadata_kolom_id = VALUES(metadata_kolom_id)
            ''', (penilaian_id, pertanyaan['id'], sesi_terlewat, metadata_kolom_id))
        
        # Data riwayat penilaian tersimpan di detail_penilaian_mahasiswa dengan sesi_penilaian dan tanggal_input
        
        get_app_functions()['mysql'].connection.commit()
        cursor.close()
        
        logger.info(f"Auto-fill sesi {sesi_terlewat} berhasil dengan skor 0 untuk {len(pertanyaan_list)} pertanyaan")
        return True
        
    except Exception as e:
        logger.error(f"Error saat auto-fill sesi {sesi_terlewat}: {str(e)}")
        if 'cursor' in locals():
            cursor.close()
        return False

def auto_fill_all_missed_sessions(id_pembimbing, nim_mahasiswa, current_sesi_waktu):
    """Auto-fill semua sesi yang terlewat dengan skor 0"""
    try:
        cursor = get_app_functions()['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Ambil sesi yang sudah ada data dari detail_penilaian_mahasiswa
        cursor.execute('''
            SELECT DISTINCT dpm.sesi_penilaian 
            FROM detail_penilaian_mahasiswa dpm
            JOIN penilaian_mahasiswa pm ON dpm.id_penilaian_mahasiswa = pm.id
            WHERE pm.id_pembimbing = %s AND pm.id_proposal IN (
                SELECT id FROM proposal WHERE nim = %s
            )
        ''', (id_pembimbing, nim_mahasiswa))
        
        existing_sessions = {row['sesi_penilaian'] for row in cursor.fetchall()}
        
        # Auto-fill semua sesi yang terlewat (dari 1 sampai current_sesi_waktu, tidak termasuk current_sesi)
        sessions_to_fill = []
        for sesi in range(1, current_sesi_waktu):  # Tidak termasuk current_sesi karena itu sesi yang sedang aktif
            if sesi not in existing_sessions:
                sessions_to_fill.append(sesi)
        
        if not sessions_to_fill:
            logger.info(f"Tidak ada sesi yang perlu di auto-fill")
            return True
        
        # Auto-fill setiap sesi yang terlewat
        success_count = 0
        for sesi in sessions_to_fill:
            if auto_fill_sesi_terlewat_helper(id_pembimbing, nim_mahasiswa, sesi):
                success_count += 1
                logger.info(f"Auto-fill sesi {sesi} berhasil")
            else:
                logger.error(f"Auto-fill sesi {sesi} gagal")
        
        cursor.close()
        
        if success_count > 0:
            logger.info(f"Auto-fill berhasil untuk {success_count} dari {len(sessions_to_fill)} sesi yang terlewat")
            return True
        else:
            logger.error(f"Semua auto-fill gagal")
            return False
            
    except Exception as e:
        logger.error(f"Error saat auto-fill semua sesi terlewat: {str(e)}")
        if 'cursor' in locals():
            cursor.close()
        return False




@pembimbing_bp.route('/get_detail_penilaian_mahasiswa')
def get_detail_penilaian_mahasiswa():
    """API untuk mengambil detail lengkap penilaian mahasiswa untuk modal"""
    if 'user_type' not in session or session['user_type'] != 'pembimbing':
        return jsonify({'success': False, 'message': 'Anda harus login sebagai pembimbing!'})
    
    try:
        nim_mahasiswa = request.args.get('nim')
        if not nim_mahasiswa:
            return jsonify({'success': False, 'message': 'NIM mahasiswa tidak ditemukan!'})
        
        cursor = get_app_functions()['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Ambil ID pembimbing
        cursor.execute('SELECT id FROM pembimbing WHERE nama = %s', (session['nama'],))
        pembimbing_data = cursor.fetchone()
        if not pembimbing_data:
            return jsonify({'success': False, 'message': 'Data pembimbing tidak ditemukan!'})
        
        id_pembimbing = pembimbing_data['id']
        
        # Ambil data proposal dan mahasiswa
        cursor.execute('''
            SELECT p.id as proposal_id, p.nim, p.judul_usaha, p.status, p.status_admin,
                   m.nama_ketua as nama_mahasiswa, m.program_studi
            FROM proposal p
            JOIN mahasiswa m ON p.nim = m.nim
            WHERE p.nim = %s AND p.dosen_pembimbing = %s
            AND p.status IN ('disetujui', 'revisi', 'selesai') AND p.status_admin = 'lolos'
        ''', (nim_mahasiswa, session['nama']))
        
        proposal = cursor.fetchone()
        if not proposal:
            return jsonify({'success': False, 'message': 'Data proposal tidak ditemukan!'})
        
        # Ambil data penilaian
        cursor.execute('''
            SELECT id FROM penilaian_mahasiswa 
            WHERE id_pembimbing = %s AND id_proposal = %s
        ''', (id_pembimbing, proposal['proposal_id']))
        
        penilaian_data = cursor.fetchone()
        if not penilaian_data:
            return jsonify({'success': False, 'message': 'Data penilaian tidak ditemukan!'})
        
        penilaian_id = penilaian_data['id']
        
        # PERBAIKAN: Cek apakah ini data lama atau data baru berdasarkan waktu pembuatan penilaian
        # Ambil waktu pembuatan penilaian terlebih dahulu
        cursor.execute('''
            SELECT created_at FROM penilaian_mahasiswa 
            WHERE id = %s
        ''', (penilaian_id,))
        penilaian_created = cursor.fetchone()
        
        if penilaian_created and penilaian_created['created_at']:
            # Cek apakah ini data lama (dibuat sebelum jadwal mulai) atau data baru (dibuat setelah jadwal mulai)
            cursor.execute('''
                SELECT pembimbing_nilai_mulai 
                FROM pengaturan_jadwal 
                ORDER BY id DESC 
                LIMIT 1
            ''')
            jadwal = cursor.fetchone()
            
            if jadwal and jadwal['pembimbing_nilai_mulai']:
                jadwal_mulai = jadwal['pembimbing_nilai_mulai']
                penilaian_created_time = penilaian_created['created_at']
                
                if penilaian_created_time < jadwal_mulai:
                    # DATA LAMA: Gunakan snapshot berdasarkan waktu pembuatan penilaian
                    cursor.execute('''
                        SELECT DISTINCT p.id, p.pertanyaan, p.bobot, p.skor_maksimal, p.kategori, p.status
                        FROM pertanyaan_penilaian_mahasiswa p
                        INNER JOIN detail_penilaian_mahasiswa d ON p.id = d.id_pertanyaan
                        WHERE d.id_penilaian_mahasiswa = %s 
                        AND p.created_at <= %s
                        ORDER BY p.kategori, p.created_at ASC
                    ''', (penilaian_id, penilaian_created['created_at']))
                    pertanyaan_list = cursor.fetchall()
                    logger.info(f"Detail modal DATA LAMA: Menggunakan snapshot pertanyaan berdasarkan created_at: {penilaian_created['created_at']}")
                else:
                    # DATA BARU: Gunakan SEMUA pertanyaan aktif saat ini (termasuk pertanyaan baru)
                    # Data baru dibuat setelah jadwal mulai, jadi berhak mendapatkan semua pertanyaan
                    cursor.execute('''
                        SELECT p.id, p.pertanyaan, p.bobot, p.skor_maksimal, p.kategori, p.status
                        FROM pertanyaan_penilaian_mahasiswa p
                        WHERE p.status = 'Aktif'
                        ORDER BY FIELD(p.kategori, 'pondasi-bisnis', 'pelanggan', 'produk-jasa', 'pemasaran', 'penjualan', 'keuntungan', 'sdm', 'sistem-bisnis'), p.created_at ASC
                    ''')
                    pertanyaan_list = cursor.fetchall()
                    logger.info(f"Detail modal DATA BARU: Menggunakan semua pertanyaan aktif ({len(pertanyaan_list)} pertanyaan) untuk penilaian created: {penilaian_created['created_at']}")
            else:
                # Tidak ada jadwal, gunakan semua pertanyaan yang punya data
                cursor.execute('''
                    SELECT DISTINCT p.id, p.pertanyaan, p.bobot, p.skor_maksimal, p.kategori, p.status
                    FROM pertanyaan_penilaian_mahasiswa p
                    INNER JOIN detail_penilaian_mahasiswa d ON p.id = d.id_pertanyaan
                    WHERE d.id_penilaian_mahasiswa = %s
                    ORDER BY p.kategori, p.created_at ASC
                ''', (penilaian_id,))
                pertanyaan_list = cursor.fetchall()
                logger.warning(f"Detail modal: Tidak ada jadwal, menggunakan semua pertanyaan untuk penilaian_id: {penilaian_id}")
        else:
            # Fallback ke semua pertanyaan yang punya data jika tidak ada created_at
            cursor.execute('''
                SELECT DISTINCT p.id, p.pertanyaan, p.bobot, p.skor_maksimal, p.kategori, p.status
                FROM pertanyaan_penilaian_mahasiswa p
                INNER JOIN detail_penilaian_mahasiswa d ON p.id = d.id_pertanyaan
                WHERE d.id_penilaian_mahasiswa = %s
                ORDER BY p.kategori, p.created_at ASC
            ''', (penilaian_id,))
            pertanyaan_list = cursor.fetchall()
            logger.warning(f"Detail modal: Fallback ke semua pertanyaan untuk penilaian_id: {penilaian_id}")
        
        # ✅ PERBAIKAN: Cek apakah penilaian menggunakan metadata_kolom_id (data lama) atau jadwal_id (data baru)
        penilaian_jadwal_id = None
        is_data_lama = False
        
        try:
            # Pastikan kolom jadwal_id ada sebelum query
            cursor.execute("SHOW COLUMNS FROM penilaian_mahasiswa LIKE 'jadwal_id'")
            if cursor.fetchone():
                cursor.execute('SELECT jadwal_id FROM penilaian_mahasiswa WHERE id = %s', (penilaian_id,))
                jadwal_result = cursor.fetchone()
                if jadwal_result and jadwal_result['jadwal_id']:
                    penilaian_jadwal_id = jadwal_result['jadwal_id']
                else:
                    # Jika jadwal_id NULL, kemungkinan ini data lama
                    is_data_lama = True
            else:
                # Kolom jadwal_id tidak ada, ini pasti data lama
                is_data_lama = True
        except Exception as e:
            logger.warning(f"Tidak bisa ambil jadwal_id untuk penilaian {penilaian_id}: {str(e)}")
            is_data_lama = True
        
        # ✅ PERBAIKAN: Cek apakah detail_penilaian_mahasiswa menggunakan metadata_kolom_id
        if not is_data_lama:
            cursor.execute('''
                SELECT metadata_kolom_id FROM detail_penilaian_mahasiswa 
                WHERE id_penilaian_mahasiswa = %s 
                LIMIT 1
            ''', (penilaian_id,))
            metadata_check = cursor.fetchone()
            if metadata_check and metadata_check['metadata_kolom_id']:
                # Jika ada metadata_kolom_id, ini data lama yang menggunakan metadata tersimpan
                is_data_lama = True
                penilaian_jadwal_id = None
                logger.info(f"Detail penilaian mahasiswa {nim_mahasiswa}: Detected as DATA LAMA (using metadata_kolom_id)")
            else:
                logger.info(f"Detail penilaian mahasiswa {nim_mahasiswa}: Detected as DATA BARU (using jadwal_id)")
        else:
            logger.info(f"Detail penilaian mahasiswa {nim_mahasiswa}: Detected as DATA LAMA (no jadwal_id)")
        
        # Ambil skor per sesi dengan metadata kolom berdasarkan jadwal_id yang tepat
        if not is_data_lama and penilaian_jadwal_id:
            # Data baru: gunakan metadata berdasarkan jadwal_id
            cursor.execute('''
                SELECT d.id_pertanyaan, d.sesi_penilaian, d.skor, d.nilai, d.tanggal_input,
                       m.nama_kolom, m.urutan_kolom, m.interval_tipe,
                       COALESCE(m.nama_kolom, 
                           CASE 
                               WHEN m.interval_tipe = 'setiap_jam' THEN CONCAT('Jam ', LPAD(20 + d.sesi_penilaian, 2, '0'), ':20')
                               WHEN m.interval_tipe = 'harian' THEN CONCAT('Hari ', d.sesi_penilaian)
                               WHEN m.interval_tipe = 'mingguan' THEN CONCAT('Minggu ', d.sesi_penilaian)
                               WHEN m.interval_tipe = 'bulanan' THEN CONCAT('Bulan ', d.sesi_penilaian)
                               WHEN m.interval_tipe IS NULL THEN (
                                   -- Jika tidak ada metadata, gunakan interval dari jadwal
                                   SELECT 
                                       CASE 
                                           WHEN j.pembimbing_interval_tipe = 'setiap_jam' THEN CONCAT('Jam ', LPAD(20 + d.sesi_penilaian, 2, '0'), ':20')
                                           WHEN j.pembimbing_interval_tipe = 'harian' THEN CONCAT('Hari ', d.sesi_penilaian)
                                           WHEN j.pembimbing_interval_tipe = 'mingguan' THEN CONCAT('Minggu ', d.sesi_penilaian)
                                           WHEN j.pembimbing_interval_tipe = 'bulanan' THEN CONCAT('Bulan ', d.sesi_penilaian)
                                           ELSE CONCAT('Sesi ', d.sesi_penilaian)
                                       END
                                   FROM pengaturan_jadwal j 
                                   WHERE j.id = %s
                               )
                               ELSE CONCAT('Sesi ', d.sesi_penilaian)
                           END
                       ) as nama_kolom_display
                FROM detail_penilaian_mahasiswa d
                LEFT JOIN metadata_kolom_penilaian m ON (m.jadwal_id = %s AND m.urutan_kolom = d.sesi_penilaian)
                WHERE d.id_penilaian_mahasiswa = %s
                ORDER BY d.id_pertanyaan, d.sesi_penilaian
            ''', (penilaian_jadwal_id, penilaian_jadwal_id, penilaian_id))
        else:
            # ✅ PERBAIKAN: Data lama - gunakan metadata yang sudah tersimpan di database (metadata_kolom_id)
            # Jangan menggunakan interval dominan, tapi hormati metadata yang sudah tersimpan
            cursor.execute('''
                SELECT d.id_pertanyaan, d.sesi_penilaian, d.skor, d.nilai, d.tanggal_input,
                       m.nama_kolom, m.urutan_kolom, m.interval_tipe,
                       COALESCE(m.nama_kolom, 
                           CASE 
                               WHEN m.interval_tipe = 'setiap_jam' THEN CONCAT('Jam ', LPAD(20 + d.sesi_penilaian, 2, '0'), ':20')
                               WHEN m.interval_tipe = 'harian' THEN CONCAT('Hari ', d.sesi_penilaian)
                               WHEN m.interval_tipe = 'mingguan' THEN CONCAT('Minggu ', d.sesi_penilaian)
                               WHEN m.interval_tipe = 'bulanan' THEN CONCAT('Bulan ', d.sesi_penilaian)
                               ELSE CONCAT('Sesi ', d.sesi_penilaian)
                           END
                       ) as nama_kolom_display
                FROM detail_penilaian_mahasiswa d
                LEFT JOIN metadata_kolom_penilaian m ON d.metadata_kolom_id = m.id
                WHERE d.id_penilaian_mahasiswa = %s
                ORDER BY d.id_pertanyaan, d.sesi_penilaian
            ''', (penilaian_id,))
        
        skor_data = cursor.fetchall()
        
        # Organize skor per sesi dengan metadata kolom
        skor_per_sesi = {}
        metadata_per_sesi = {}  # Untuk menyimpan metadata kolom per sesi
        
        for skor in skor_data:
            pertanyaan_id = str(skor['id_pertanyaan'])
            sesi = skor['sesi_penilaian']
            
            if pertanyaan_id not in skor_per_sesi:
                skor_per_sesi[pertanyaan_id] = {}
            
            skor_per_sesi[pertanyaan_id][sesi] = {
                'skor': skor['skor'],
                'nilai': skor['nilai'],
                'tanggal_input': skor['tanggal_input'].isoformat() if skor['tanggal_input'] else None
            }
            
            # Simpan metadata kolom per sesi - PERBAIKAN: Gunakan nama_kolom_display yang sudah diperbaiki
            if sesi not in metadata_per_sesi:
                metadata_per_sesi[sesi] = {
                    'nama_kolom': skor.get('nama_kolom_display', f'Sesi {sesi}'),
                    'urutan_kolom': skor.get('urutan_kolom', sesi),
                    'interval_tipe': skor.get('interval_tipe', 'unknown')
                }
            else:
                # Update jika sudah ada, pastikan konsisten
                existing_metadata = metadata_per_sesi[sesi]
                if skor.get('nama_kolom_display') and skor.get('nama_kolom_display') != existing_metadata['nama_kolom']:
                    metadata_per_sesi[sesi]['nama_kolom'] = skor.get('nama_kolom_display')
                if skor.get('interval_tipe') and skor.get('interval_tipe') != 'unknown':
                    metadata_per_sesi[sesi]['interval_tipe'] = skor.get('interval_tipe')
        
        # PERBAIKAN: Untuk data lama, pastikan semua sesi menggunakan interval yang konsisten
        if not penilaian_jadwal_id and metadata_per_sesi:
            print(f"Detail penilaian mahasiswa {nim_mahasiswa}: Processing data lama with jadwal_id = NULL")
        elif penilaian_jadwal_id:
            print(f"Detail penilaian mahasiswa {nim_mahasiswa}: Processing data baru with jadwal_id = {penilaian_jadwal_id}")
            
        # PERBAIKAN: Untuk data baru, pastikan sesi yang tidak ada metadata menggunakan interval yang benar
        if not is_data_lama and penilaian_jadwal_id and metadata_per_sesi:
            # Ambil interval dari jadwal untuk data baru
            cursor.execute('''
                SELECT pembimbing_interval_tipe as interval_tipe
                FROM pengaturan_jadwal 
                WHERE id = %s
            ''', (penilaian_jadwal_id,))
            jadwal_interval = cursor.fetchone()
            
            if jadwal_interval and jadwal_interval['interval_tipe']:
                interval_dominan = jadwal_interval['interval_tipe']
                print(f"Detail penilaian mahasiswa {nim_mahasiswa}: Using jadwal interval = {interval_dominan}")
                
                # Update metadata untuk sesi yang tidak memiliki nama_kolom
                for sesi in metadata_per_sesi:
                    if not metadata_per_sesi[sesi].get('nama_kolom') or metadata_per_sesi[sesi]['nama_kolom'].startswith('Sesi '):
                        if interval_dominan == 'setiap_jam':
                            metadata_per_sesi[sesi]['nama_kolom'] = f'Jam {20 + sesi:02d}:20'
                        elif interval_dominan == 'harian':
                            metadata_per_sesi[sesi]['nama_kolom'] = f'Hari {sesi}'
                        elif interval_dominan == 'mingguan':
                            metadata_per_sesi[sesi]['nama_kolom'] = f'Minggu {sesi}'
                        elif interval_dominan == 'bulanan':
                            # ✅ PERBAIKAN: Untuk interval bulanan, gunakan nama bulan yang sebenarnya
                            # Ambil jadwal mulai untuk menentukan nama bulan
                            cursor.execute('''
                                SELECT pembimbing_nilai_mulai, pembimbing_nilai_selesai
                                FROM pengaturan_jadwal 
                                WHERE id = %s
                            ''', (penilaian_jadwal_id,))
                            jadwal_info_result = cursor.fetchone()
                            
                            if jadwal_info_result and jadwal_info_result['pembimbing_nilai_mulai']:
                                jadwal_mulai = jadwal_info_result['pembimbing_nilai_mulai']
                                jadwal_selesai = jadwal_info_result['pembimbing_nilai_selesai']
                                
                                # Jika jadwal mulai dan selesai dalam bulan yang sama
                                if jadwal_mulai.month == jadwal_selesai.month and jadwal_mulai.year == jadwal_selesai.year:
                                    # Gunakan nama bulan (Agustus 2025)
                                    nama_bulan = jadwal_mulai.strftime('%B %Y')
                                    metadata_per_sesi[sesi]['nama_kolom'] = nama_bulan
                                else:
                                    # Jika berbeda bulan, gunakan format "Bulan X"
                                    metadata_per_sesi[sesi]['nama_kolom'] = f'Bulan {sesi}'
                            else:
                                # Fallback ke format lama
                                metadata_per_sesi[sesi]['nama_kolom'] = f'Bulan {sesi}'
                        metadata_per_sesi[sesi]['interval_tipe'] = interval_dominan
            
        elif is_data_lama and metadata_per_sesi:
            # ✅ PERBAIKAN: Untuk data lama, JANGAN tentukan interval dominan
            # Biarkan setiap sesi menggunakan metadata aslinya dari database
            print(f"Detail penilaian mahasiswa {nim_mahasiswa}: Data lama - menggunakan metadata asli dari database")
            
            # Debug: tampilkan metadata yang digunakan
            for sesi, metadata in metadata_per_sesi.items():
                print(f"  Sesi {sesi}: {metadata.get('nama_kolom', 'N/A')} (interval: {metadata.get('interval_tipe', 'unknown')})")
            
            # Tidak ada interval dominan untuk data lama - setiap sesi menggunakan metadata aslinya
            # Metadata sudah diambil dari database berdasarkan metadata_kolom_id yang tersimpan
        
        # Gunakan helper function untuk menghitung nilai yang konsisten
        nilai_akhir, total_sesi, status_penilaian = hitung_nilai_penilaian_mahasiswa(
            cursor, penilaian_id
        )
        
        # PERBAIKAN: Ambil data jadwal untuk informasi sesi dari database
        if not is_data_lama and penilaian_jadwal_id:
            # Untuk data baru, gunakan jadwal yang sesuai dengan jadwal_id
            cursor.execute('''
                SELECT 
                    pembimbing_interval_tipe as interval_tipe,
                    pembimbing_interval_nilai as interval_nilai,
                    pembimbing_nilai_mulai as jam_mulai,
                    pembimbing_nilai_selesai as jam_selesai
                FROM pengaturan_jadwal
                WHERE id = %s
            ''', (penilaian_jadwal_id,))
            jadwal_info = cursor.fetchone()
            print(f"Detail penilaian mahasiswa {nim_mahasiswa}: Using jadwal_id {penilaian_jadwal_id} with interval = {jadwal_info['interval_tipe'] if jadwal_info else 'None'}")
        else:
            # ✅ PERBAIKAN: Untuk data lama, JANGAN gunakan jadwal terbaru
            # Gunakan metadata yang sudah tersimpan di database
            jadwal_info = {
                'interval_tipe': 'data_lama',
                'interval_nilai': 1,
                'jam_mulai': None,
                'jam_selesai': None
            }
            print(f"Detail penilaian mahasiswa {nim_mahasiswa}: Data lama - menggunakan metadata asli dari database")
        
        # Generate jam_list berdasarkan interval dari database (sama seperti admin)
        if jadwal_info and jadwal_info['interval_tipe'] and jadwal_info['interval_tipe'] != 'data_lama':
            jam_mulai = jadwal_info['jam_mulai']
            jam_selesai = jadwal_info['jam_selesai']
            interval = jadwal_info['interval_nilai'] or 1
            
            # Generate list jam dari jam_mulai sampai jam_selesai dengan interval
            jam_list = []
            
            try:
                # Pastikan jam_mulai dan jam_selesai adalah time object
                if isinstance(jam_mulai, str):
                    try:
                        jam_mulai = datetime.strptime(jam_mulai, '%Y-%m-%d %H:%M:%S')
                        jam_mulai = jam_mulai.time()
                    except:
                        jam_mulai = datetime.strptime(jam_mulai, '%H:%M:%S').time()
                elif isinstance(jam_mulai, datetime):
                    jam_mulai = jam_mulai.time()
                elif isinstance(jam_mulai, timedelta):
                    total_seconds = int(jam_mulai.total_seconds())
                    hours = total_seconds // 3600
                    minutes = (total_seconds % 3600) // 60
                    jam_mulai = datetime.strptime(f"{hours:02d}:{minutes:02d}:00", '%H:%M:%S').time()
                
                if isinstance(jam_selesai, str):
                    try:
                        jam_selesai = datetime.strptime(jam_selesai, '%Y-%m-%d %H:%M:%S')
                        jam_selesai = jam_selesai.time()
                    except:
                        jam_selesai = datetime.strptime(jam_selesai, '%H:%M:%S').time()
                elif isinstance(jam_selesai, datetime):
                    jam_selesai = jam_selesai.time()
                elif isinstance(jam_selesai, timedelta):
                    total_seconds = int(jam_selesai.total_seconds())
                    hours = total_seconds // 3600
                    minutes = total_seconds % 3600 // 60
                    jam_selesai = datetime.strptime(f"{hours:02d}:{minutes:02d}:00", '%H:%M:%S').time()
                
                # Konversi ke datetime untuk perhitungan
                today = datetime.now().date()
                current_dt = datetime.combine(today, jam_mulai)
                end_dt = datetime.combine(today, jam_selesai)
                
                # Generate jam_list berdasarkan interval yang sebenarnya dari database
                while current_dt <= end_dt:
                    jam_list.append(current_dt.strftime('%H:%M'))
                    
                    # Tambah interval sesuai pengaturan database
                    if jadwal_info['interval_tipe'] == 'setiap_jam':
                        current_dt += timedelta(hours=interval)
                    elif jadwal_info['interval_tipe'] == 'harian':
                        current_dt += timedelta(days=interval)
                    elif jadwal_info['interval_tipe'] == 'mingguan':
                        current_dt += timedelta(weeks=interval)
                    elif jadwal_info['interval_tipe'] == 'bulanan':
                        current_dt += timedelta(days=interval * 30)
                
                jadwal_info['jam_list'] = jam_list
                
            except Exception as e:
                print(f"Error generating jam_list: {e}")
                # Fallback ke default jika error
                jadwal_info['jam_list'] = ['08:00', '09:00', '10:00', '11:00']
        else:
            # Untuk data lama, gunakan jam_list kosong atau default
            jadwal_info['jam_list'] = []
        
        # Pastikan semua data di jadwal_info bisa di-serialize ke JSON
        if jadwal_info:
            # Konversi ke tipe yang aman untuk JSON
            jadwal_info['interval_tipe'] = str(jadwal_info['interval_tipe'])
            jadwal_info['interval_nilai'] = int(jadwal_info['interval_nilai']) if jadwal_info['interval_nilai'] else 1
            jadwal_info['jam_mulai'] = str(jadwal_info['jam_mulai'])
            jadwal_info['jam_selesai'] = str(jadwal_info['jam_selesai'])
            jadwal_info['jam_list'] = jadwal_info.get('jam_list', [])
        
        # Ambil tanggal penilaian terbaru
        tanggal_penilaian = '-'
        if skor_data:
            latest_date = max(skor['tanggal_input'] for skor in skor_data if skor['tanggal_input'])
            if latest_date:
                tanggal_penilaian = latest_date.strftime('%d %B %Y, %H:%M')
        
        # Organize pertanyaan berdasarkan kategori
        kategori_organized = {}
        for pertanyaan in pertanyaan_list:
            kategori = pertanyaan['kategori']
            if kategori not in kategori_organized:
                kategori_organized[kategori] = {
                    'nama_kategori': kategori.replace('-', ' ').title(),
                    'pertanyaan_list': [],
                    'total_bobot_kategori': 0,
                    'total_nilai_kategori': 0
                }
            
            # Hitung nilai untuk pertanyaan ini
            pertanyaan_id = str(pertanyaan['id'])
            sesi_data = skor_per_sesi.get(pertanyaan_id, {})
            
            # ✅ PERBAIKAN: Hitung rata-rata skor dari semua sesi yang ada
            total_skor = 0
            sesi_count = 0
            for sesi, data in sesi_data.items():
                total_skor += data['skor']  # Hitung semua sesi (termasuk skor 0)
                sesi_count += 1
            
            # ✅ PERBAIKAN: Rata-rata = total skor / jumlah sesi (semua sesi)
            rata_rata_skor = (total_skor / sesi_count) if sesi_count > 0 else 0
            bobot = float(pertanyaan['bobot'])
            skor_max = int(pertanyaan['skor_maksimal'])
            nilai_akhir_pertanyaan = (rata_rata_skor / skor_max) * bobot
            
            # Tambahkan ke kategori
            kategori_organized[kategori]['pertanyaan_list'].append({
                'id': pertanyaan['id'],
                'pertanyaan': pertanyaan['pertanyaan'],
                'bobot': bobot,
                'skor_maksimal': skor_max,
                'sesi_data': sesi_data,
                'rata_rata_skor': rata_rata_skor,
                'nilai_akhir': nilai_akhir_pertanyaan
            })
            
            kategori_organized[kategori]['total_bobot_kategori'] += bobot
            kategori_organized[kategori]['total_nilai_kategori'] += nilai_akhir_pertanyaan
        
        # Convert ke list untuk response
        detail_penilaian = list(kategori_organized.values())
        
        cursor.close()
        
        # Siapkan response data
        response_data = {
            'success': True,
            'data': {
                'nim_mahasiswa': proposal['nim'],
                'nama_mahasiswa': proposal['nama_mahasiswa'],
                'judul_usaha': proposal['judul_usaha'],
                'program_studi': proposal['program_studi'],
                'status_penilaian': status_penilaian,
                'nilai': nilai_akhir,
                'total_sesi': total_sesi,
                'tanggal_penilaian': tanggal_penilaian,
                'pertanyaan': pertanyaan_list,
                'skor_per_sesi': skor_per_sesi,
                'metadata_per_sesi': metadata_per_sesi,
                'jadwal_info': jadwal_info,
                'detail_penilaian': detail_penilaian
            }
        }
        
        return jsonify(response_data)
        
    except Exception as e:
        logger.error(f"Error in get_detail_penilaian_mahasiswa: {str(e)}")
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@pembimbing_bp.route('/download_excel_anggaran')
def download_excel_anggaran():
    """Download Excel anggaran awal/bertumbuh dengan format kop surat"""
    try:
        # Cek login
        if 'user_type' not in session or session.get('user_type') != 'pembimbing':
            return jsonify({'success': False, 'message': 'Unauthorized'}), 401
        
        id_proposal = request.args.get('id_proposal')
        jenis = request.args.get('jenis')  # 'awal' atau 'bertumbuh'
        
        if not id_proposal or not jenis:
            return jsonify({'success': False, 'message': 'Parameter tidak lengkap'}), 400
        
        # Koneksi database
        app_funcs = get_app_functions()
        cursor = app_funcs['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Ambil data proposal
        cursor.execute("""
            SELECT p.id, p.judul_usaha, m.nama_ketua as nama, m.nim 
            FROM proposal p
            JOIN mahasiswa m ON p.nim = m.nim
            WHERE p.id = %s
        """, (id_proposal,))
        proposal_data = cursor.fetchone()
        
        if not proposal_data:
            return jsonify({'success': False, 'message': 'Proposal tidak ditemukan'}), 404
        
        # Tentukan tabel berdasarkan jenis
        tabel_anggaran = f'anggaran_{jenis}'
        
        # Ambil data anggaran
        cursor.execute(f"""
            SELECT kegiatan_utama, kegiatan, nama_barang, kuantitas, satuan, 
                   harga_satuan, jumlah, keterangan, target_capaian, penanggung_jawab
            FROM {tabel_anggaran}
            WHERE id_proposal = %s
            ORDER BY kegiatan_utama, kegiatan, nama_barang
        """, (id_proposal,))
        anggaran_data = cursor.fetchall()
        
        # Ambil tahapan usaha untuk menentukan urutan kegiatan utama
        cursor.execute("""
            SELECT tahapan_usaha FROM proposal WHERE id = %s
        """, (id_proposal,))
        proposal_info = cursor.fetchone()
        tahapan_usaha = proposal_info.get('tahapan_usaha', '').lower() if proposal_info else ''
        
        # Urutkan data sesuai urutan kegiatan utama yang standardisasi
        if 'bertumbuh' in tahapan_usaha:
            urutan_kegiatan = [
                "Pengembangan Pasar dan Saluran Distribusi",
                "Pengembangan Produk/Riset",
                "Produksi",
                "Legalitas, perizinan, sertifikasi, dan standarisasi",
                "Belanja ATK dan Penunjang"
            ]
        else:
            urutan_kegiatan = [
                "Pengembangan Produk/Riset",
                "Produksi",
                "Legalitas, perizinan, sertifikasi, dan standarisasi",
                "Belanja ATK dan Penunjang"
            ]
        
        # Buat mapping untuk case-insensitive matching dan normalisasi teks
        urutan_kegiatan_lower = [k.strip().lower() for k in urutan_kegiatan]
        
        # Fungsi helper untuk normalisasi teks kegiatan utama
        def normalize_kegiatan_utama(text):
            if not text:
                return ""
            # Normalisasi: lowercase, strip whitespace, replace multiple spaces
            normalized = text.strip().lower()
            # Handle berbagai variasi teks
            replacements = {
                "legalitas, perijinan, sertifikasi, pengujian produk, dan standarisasi": "legalitas, perizinan, sertifikasi, dan standarisasi",
                "legalitas, perijinan, sertifikasi, dan standarisasi": "legalitas, perizinan, sertifikasi, dan standarisasi",
                "legalitas perijinan sertifikasi pengujian produk dan standarisasi": "legalitas, perizinan, sertifikasi, dan standarisasi",
                "legalitas perijinan sertifikasi dan standarisasi": "legalitas, perizinan, sertifikasi, dan standarisasi"
            }
            for old, new in replacements.items():
                if old in normalized:
                    normalized = normalized.replace(old, new)
            return normalized
        
        # Sort data berdasarkan urutan kegiatan utama dengan normalisasi
        anggaran_data = sorted(
            anggaran_data,
            key=lambda x: urutan_kegiatan_lower.index(normalize_kegiatan_utama(x['kegiatan_utama'])) if x['kegiatan_utama'] and normalize_kegiatan_utama(x['kegiatan_utama']) in urutan_kegiatan_lower else 99
        )
        
        if not anggaran_data:
            return jsonify({'success': False, 'message': 'Data anggaran tidak ditemukan'}), 404
        
        # Buat workbook Excel
        wb = Workbook()
        ws = wb.active
        ws.title = f"Anggaran {jenis.title()}"
        
        # Styling
        header_font = Font(bold=True, color="000000", size=12)
        title_font = Font(bold=True, color="000000", size=14)
        border_thin = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        # Header kop surat (sama dengan laporan laba rugi)
        current_row = 1
        
        # Menggunakan format kop surat yang sama dengan laporan laba rugi
        from utils import create_pdf_header_from_kop_surat_pdf
        try:
            header_template = create_pdf_header_from_kop_surat_pdf()
            header_data = header_template['header_data']
            
            # Gunakan data dari kop surat yang sama
            for i, header_row in enumerate(header_data[:6]):  # Ambil 6 baris pertama
                if len(header_row) > 1 and header_row[1].strip():  # Jika ada teks
                    ws.merge_cells(f'A{current_row}:K{current_row}')
                    ws[f'A{current_row}'] = header_row[1].strip()
                    if i < 2:  # Baris 1-2 bold dan lebih besar
                        ws[f'A{current_row}'].font = Font(bold=True, size=12, name='Times New Roman')
                    else:  # Baris lainnya normal
                        ws[f'A{current_row}'].font = Font(bold=False, size=10, name='Times New Roman')
                    ws[f'A{current_row}'].alignment = center_alignment
                    current_row += 1
                elif not header_row[1].strip():  # Baris kosong
                    current_row += 1
        except Exception as e:
            # Fallback ke format default jika gagal
            print(f"Warning: Menggunakan fallback kop surat: {e}")
            ws.merge_cells(f'A{current_row}:K{current_row}')
            ws[f'A{current_row}'] = "YAYASAN KARTIKA EKA PAKSI"
            ws[f'A{current_row}'].font = Font(bold=True, size=12, name='Times New Roman')
            ws[f'A{current_row}'].alignment = center_alignment
            current_row += 1
            
            ws.merge_cells(f'A{current_row}:K{current_row}')
            ws[f'A{current_row}'] = "UNIVERSITAS JENDERAL ACHMAD YANI YOGYAKARTA"
            ws[f'A{current_row}'].font = Font(bold=True, size=12, name='Times New Roman')
            ws[f'A{current_row}'].alignment = center_alignment
            current_row += 1
            
            ws.merge_cells(f'A{current_row}:K{current_row}')
            ws[f'A{current_row}'] = "Jl Siliwangi Ringroad Barat, Banyuraden, Gamping, Sleman, Yogyakarta 55293"
            ws[f'A{current_row}'].font = Font(bold=False, size=10, name='Times New Roman')
            ws[f'A{current_row}'].alignment = center_alignment
            current_row += 1
            
            ws.merge_cells(f'A{current_row}:K{current_row}')
            ws[f'A{current_row}'] = "Telp. (0274) 552489, 552851 Fax. (0274) 557228 Website: www.unjaya.ac.id"
            ws[f'A{current_row}'].font = Font(bold=False, size=10, name='Times New Roman')
            ws[f'A{current_row}'].alignment = center_alignment
            current_row += 1
            
            ws.merge_cells(f'A{current_row}:K{current_row}')
            ws[f'A{current_row}'] = "Email: info@unjaya.ac.id"
            ws[f'A{current_row}'].font = Font(bold=False, size=10, name='Times New Roman')
            ws[f'A{current_row}'].alignment = center_alignment
            current_row += 1
        
        # Tambahkan spasi setelah kop surat
        current_row += 1
        
        # Judul laporan
        ws.merge_cells(f'A{current_row}:K{current_row}')
        ws[f'A{current_row}'] = f"RENCANA ANGGARAN {jenis.upper()}"
        ws[f'A{current_row}'].font = title_font
        ws[f'A{current_row}'].alignment = center_alignment
        current_row += 2
        
        # Info mahasiswa
        ws[f'A{current_row}'] = "Nama Mahasiswa:"
        ws[f'A{current_row}'].font = header_font
        ws[f'B{current_row}'] = proposal_data['nama']
        current_row += 1
        
        ws[f'A{current_row}'] = "NIM:"
        ws[f'A{current_row}'].font = header_font
        ws[f'B{current_row}'] = proposal_data['nim']
        current_row += 1
        
        ws[f'A{current_row}'] = "Nama Usaha:"
        ws[f'A{current_row}'].font = header_font
        ws[f'B{current_row}'] = proposal_data['judul_usaha']
        current_row += 1
        
        ws[f'A{current_row}'] = "Judul Usaha:"
        ws[f'A{current_row}'].font = header_font
        ws[f'B{current_row}'] = proposal_data['judul_usaha']
        current_row += 2
        
        # Header tabel dengan merged cells
        # Row 1: Header utama
        ws.cell(row=current_row, column=1, value="Kegiatan Utama").font = header_font
        ws.cell(row=current_row, column=1).fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
        ws.cell(row=current_row, column=1).border = border_thin
        ws.cell(row=current_row, column=1).alignment = center_alignment
        
        # Merge cells untuk "Rencana" (kolom B-G)
        ws.merge_cells(f'B{current_row}:G{current_row}')
        ws[f'B{current_row}'] = "Rencana"
        ws[f'B{current_row}'].font = header_font
        ws[f'B{current_row}'].fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
        ws[f'B{current_row}'].border = border_thin
        ws[f'B{current_row}'].alignment = center_alignment
        
        ws.cell(row=current_row, column=8, value="Keterangan/Ref. Harga").font = header_font
        ws.cell(row=current_row, column=8).fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
        ws.cell(row=current_row, column=8).border = border_thin
        ws.cell(row=current_row, column=8).alignment = center_alignment
        
        ws.cell(row=current_row, column=9, value="Target Capaian H=Output A").font = header_font
        ws.cell(row=current_row, column=9).fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
        ws.cell(row=current_row, column=9).border = border_thin
        ws.cell(row=current_row, column=9).alignment = center_alignment
        
        ws.cell(row=current_row, column=10, value="Penanggung Jawab").font = header_font
        ws.cell(row=current_row, column=10).fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
        ws.cell(row=current_row, column=10).border = border_thin
        ws.cell(row=current_row, column=10).alignment = center_alignment
        
        current_row += 1
        
        # Row 2: Sub-header untuk Rencana
        ws.cell(row=current_row, column=1, value="").border = border_thin  # Empty cell untuk alignment
        
        # Sub-header untuk kolom Rencana
        sub_headers = ["Kegiatan A", "Nama Barang B", "Kuantitas C", "Satuan D", "Harga Satuan E", "Jumlah (Rp) F=C×E"]
        for col, sub_header in enumerate(sub_headers, 2):
            cell = ws.cell(row=current_row, column=col, value=sub_header)
            cell.font = header_font
            cell.fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
            cell.border = border_thin
            cell.alignment = center_alignment
        
        # Empty cells untuk kolom lainnya
        for col in range(8, 11):
            ws.cell(row=current_row, column=col, value="").border = border_thin
        
        current_row += 1
        
        # Tulis data anggaran dengan grouping berdasarkan kegiatan utama
        total_jumlah = 0
        current_kegiatan_utama = None
        kegiatan_utama_start_row = None
        kegiatan_utama_row_count = 0
        
        current_kegiatan = None
        kegiatan_start_row = None
        kegiatan_row_count = 0
        
        current_target_capaian = None
        target_capaian_start_row = None
        target_capaian_row_count = 0
        
        current_penanggung_jawab = None
        penanggung_jawab_start_row = None
        penanggung_jawab_row_count = 0
        
        for i, row_data in enumerate(anggaran_data):
            # Cek jika kegiatan utama berubah
            if current_kegiatan_utama != row_data['kegiatan_utama']:
                # Merge cells untuk kegiatan utama sebelumnya jika ada
                if current_kegiatan_utama is not None and kegiatan_utama_row_count > 1:
                    ws.merge_cells(f'A{kegiatan_utama_start_row}:A{current_row-1}')
                    ws[f'A{kegiatan_utama_start_row}'].alignment = Alignment(horizontal='center', vertical='center')
                
                # Set kegiatan utama baru
                current_kegiatan_utama = row_data['kegiatan_utama']
                kegiatan_utama_start_row = current_row
                kegiatan_utama_row_count = 0
            
            # Cek jika kegiatan berubah
            if current_kegiatan != row_data['kegiatan']:
                # Merge cells untuk kegiatan sebelumnya jika ada
                if current_kegiatan is not None and kegiatan_row_count > 1:
                    ws.merge_cells(f'B{kegiatan_start_row}:B{current_row-1}')
                    ws[f'B{kegiatan_start_row}'].alignment = Alignment(horizontal='center', vertical='center')
                
                # Set kegiatan baru
                current_kegiatan = row_data['kegiatan']
                kegiatan_start_row = current_row
                kegiatan_row_count = 0
            
            # Cek jika target capaian berubah
            if current_target_capaian != row_data['target_capaian']:
                # Merge cells untuk target capaian sebelumnya jika ada
                if current_target_capaian is not None and target_capaian_row_count > 1:
                    ws.merge_cells(f'I{target_capaian_start_row}:I{current_row-1}')
                    ws[f'I{target_capaian_start_row}'].alignment = Alignment(horizontal='center', vertical='center')
                
                # Set target capaian baru
                current_target_capaian = row_data['target_capaian']
                target_capaian_start_row = current_row
                target_capaian_row_count = 0
            
            # Cek jika penanggung jawab berubah
            if current_penanggung_jawab != row_data['penanggung_jawab']:
                # Merge cells untuk penanggung jawab sebelumnya jika ada
                if current_penanggung_jawab is not None and penanggung_jawab_row_count > 1:
                    ws.merge_cells(f'J{penanggung_jawab_start_row}:J{current_row-1}')
                    ws[f'J{penanggung_jawab_start_row}'].alignment = Alignment(horizontal='center', vertical='center')
                
                # Set penanggung jawab baru
                current_penanggung_jawab = row_data['penanggung_jawab']
                penanggung_jawab_start_row = current_row
                penanggung_jawab_row_count = 0
            
            # Tulis data
            ws.cell(row=current_row, column=1, value=row_data['kegiatan_utama']).alignment = left_alignment
            ws.cell(row=current_row, column=2, value=row_data['kegiatan']).alignment = left_alignment
            ws.cell(row=current_row, column=3, value=row_data['nama_barang']).alignment = left_alignment
            ws.cell(row=current_row, column=4, value=row_data['kuantitas']).alignment = center_alignment
            ws.cell(row=current_row, column=5, value=row_data['satuan']).alignment = center_alignment
            ws.cell(row=current_row, column=6, value=f"Rp {row_data['harga_satuan']:,}").alignment = center_alignment
            ws.cell(row=current_row, column=7, value=f"Rp {row_data['jumlah']:,}").alignment = center_alignment
            ws.cell(row=current_row, column=8, value=row_data['keterangan'] or '-').alignment = left_alignment
            ws.cell(row=current_row, column=9, value=row_data['target_capaian']).alignment = left_alignment
            ws.cell(row=current_row, column=10, value=row_data['penanggung_jawab']).alignment = left_alignment
            
            # Apply borders
            for col in range(1, 11):
                ws.cell(row=current_row, column=col).border = border_thin
            
            total_jumlah += row_data['jumlah']
            current_row += 1
            kegiatan_utama_row_count += 1
            kegiatan_row_count += 1
            target_capaian_row_count += 1
            penanggung_jawab_row_count += 1
        
        # Merge cells untuk kegiatan utama terakhir jika ada lebih dari 1 row
        if kegiatan_utama_row_count > 1:
            ws.merge_cells(f'A{kegiatan_utama_start_row}:A{current_row-1}')
            ws[f'A{kegiatan_utama_start_row}'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Merge cells untuk kegiatan terakhir jika ada lebih dari 1 row
        if kegiatan_row_count > 1:
            ws.merge_cells(f'B{kegiatan_start_row}:B{current_row-1}')
            ws[f'B{kegiatan_start_row}'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Merge cells untuk target capaian terakhir jika ada lebih dari 1 row
        if target_capaian_row_count > 1:
            ws.merge_cells(f'I{target_capaian_start_row}:I{current_row-1}')
            ws[f'I{target_capaian_start_row}'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Merge cells untuk penanggung jawab terakhir jika ada lebih dari 1 row
        if penanggung_jawab_row_count > 1:
            ws.merge_cells(f'J{penanggung_jawab_start_row}:J{current_row-1}')
            ws[f'J{penanggung_jawab_start_row}'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Total row
        ws.merge_cells(f'A{current_row}:F{current_row}')
        ws[f'A{current_row}'] = "TOTAL ANGGARAN:"
        ws[f'A{current_row}'].font = Font(bold=True)
        ws[f'A{current_row}'].alignment = Alignment(horizontal='right', vertical='center')
        ws[f'A{current_row}'].border = border_thin
        
        ws[f'G{current_row}'] = f"Rp {total_jumlah:,}"
        ws[f'G{current_row}'].font = Font(bold=True)
        ws[f'G{current_row}'].fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
        ws[f'G{current_row}'].alignment = center_alignment
        ws[f'G{current_row}'].border = border_thin
        
        # Merge remaining cells in total row
        for col in range(8, 11):
            ws.cell(row=current_row, column=col).border = border_thin
        
        current_row += 2
        
        # Footer dengan tanggal
        ws[f'H{current_row}'] = f"Yogyakarta, {datetime.now().strftime('%d %B %Y')}"
        ws[f'H{current_row}'].alignment = center_alignment
        current_row += 3
        
        ws[f'H{current_row}'] = "Mahasiswa,"
        ws[f'H{current_row}'].alignment = center_alignment
        current_row += 4
        
        ws[f'H{current_row}'] = proposal_data['nama']
        ws[f'H{current_row}'].font = Font(bold=True)
        ws[f'H{current_row}'].alignment = center_alignment
        current_row += 1
        
        ws[f'H{current_row}'] = f"NIM: {proposal_data['nim']}"
        ws[f'H{current_row}'].alignment = center_alignment
        
        # Auto-adjust column widths
        for col in range(1, 11):
            column_letter = get_column_letter(col)
            max_length = 0
            for row in range(1, current_row + 1):
                try:
                    cell_value = str(ws[f'{column_letter}{row}'].value)
                    if len(cell_value) > max_length:
                        max_length = len(cell_value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Max width 50
            ws.column_dimensions[column_letter].width = max(adjusted_width, 15)  # Min width 15
        
        # Simpan ke temporary file
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        wb.save(temp_file.name)
        temp_file.close()
        
        # Buat response untuk download
        filename = f"Anggaran_{jenis.title()}_{proposal_data['judul_usaha']}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        def remove_file(response):
            try:
                os.unlink(temp_file.name)
            except Exception:
                pass
            return response
        
        return send_file(
            temp_file.name,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        print(f"Error in download_excel_anggaran: {e}")
        traceback.print_exc()
        return jsonify({'success': False, 'message': 'Terjadi kesalahan saat membuat file Excel'}), 500
    finally:
        if 'cursor' in locals():
            cursor.close()
