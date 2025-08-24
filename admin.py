from flask import Blueprint, render_template, request, jsonify, session, flash, redirect, url_for, send_file, make_response, abort

# Fungsi helper untuk pengecekan duplikasi komprehensif
def check_comprehensive_duplicates(cursor, nama=None, nim=None, nip=None, username=None, email=None, exclude_id=None, exclude_table=None):
    """
    Fungsi untuk mengecek duplikasi data di semua tabel
    Returns: (is_duplicate, message, table_name)
    """
    checks = []
    
    # Cek nama di semua tabel
    if nama:
        # Cek di mahasiswa
        if exclude_table != 'mahasiswa' or exclude_id is None:
            cursor.execute('SELECT id, nama_ketua FROM mahasiswa WHERE nama_ketua = %s', (nama,))
            mahasiswa = cursor.fetchone()
            if mahasiswa and (exclude_id is None or mahasiswa[0] != exclude_id):
                return True, f"Nama '{nama}' sudah digunakan oleh mahasiswa", "mahasiswa"
        
        # Cek di pembimbing
        if exclude_table != 'pembimbing' or exclude_id is None:
            cursor.execute('SELECT id, nama FROM pembimbing WHERE nama = %s', (nama,))
            pembimbing = cursor.fetchone()
            if pembimbing and (exclude_id is None or pembimbing[0] != exclude_id):
                return True, f"Nama '{nama}' sudah digunakan oleh pembimbing", "pembimbing"
        
        # Cek di reviewer
        if exclude_table != 'reviewer' or exclude_id is None:
            cursor.execute('SELECT id, nama FROM reviewer WHERE nama = %s', (nama,))
            reviewer = cursor.fetchone()
            if reviewer and (exclude_id is None or reviewer[0] != exclude_id):
                return True, f"Nama '{nama}' sudah digunakan oleh reviewer", "reviewer"
        
        # Cek di admin
        if exclude_table != 'admin' or exclude_id is None:
            cursor.execute('SELECT id, nama FROM admin WHERE nama = %s', (nama,))
            admin = cursor.fetchone()
            if admin and (exclude_id is None or admin[0] != exclude_id):
                return True, f"Nama '{nama}' sudah digunakan oleh admin", "admin"
    
    # Cek NIM/NIP di semua tabel
    if nim:
        # Cek NIM di mahasiswa
        if exclude_table != 'mahasiswa' or exclude_id is None:
            cursor.execute('SELECT id, nim FROM mahasiswa WHERE nim = %s', (nim,))
            mahasiswa = cursor.fetchone()
            if mahasiswa and (exclude_id is None or mahasiswa[0] != exclude_id):
                return True, f"NIM '{nim}' sudah digunakan oleh mahasiswa", "mahasiswa"
        
        # Cek NIM di pembimbing (sebagai NIP)
        if exclude_table != 'pembimbing' or exclude_id is None:
            cursor.execute('SELECT id, nip FROM pembimbing WHERE nip = %s', (nim,))
            pembimbing = cursor.fetchone()
            if pembimbing and (exclude_id is None or pembimbing[0] != exclude_id):
                return True, f"NIM '{nim}' sudah digunakan sebagai NIP pembimbing", "pembimbing"
        
        # Cek NIM di admin (sebagai NIP)
        if exclude_table != 'admin' or exclude_id is None:
            cursor.execute('SELECT id, nip FROM admin WHERE nip = %s', (nim,))
            admin = cursor.fetchone()
            if admin and (exclude_id is None or admin[0] != exclude_id):
                return True, f"NIM '{nim}' sudah digunakan sebagai NIP admin", "admin"
    
    # Cek NIP di semua tabel
    if nip:
        # Cek NIP di mahasiswa (sebagai NIM)
        if exclude_table != 'mahasiswa' or exclude_id is None:
            cursor.execute('SELECT id, nim FROM mahasiswa WHERE nim = %s', (nip,))
            mahasiswa = cursor.fetchone()
            if mahasiswa and (exclude_id is None or mahasiswa[0] != exclude_id):
                return True, f"NIP '{nip}' sudah digunakan sebagai NIM mahasiswa", "mahasiswa"
        
        # Cek NIP di pembimbing
        if exclude_table != 'pembimbing' or exclude_id is None:
            cursor.execute('SELECT id, nip FROM pembimbing WHERE nip = %s', (nip,))
            pembimbing = cursor.fetchone()
            if pembimbing and (exclude_id is None or pembimbing[0] != exclude_id):
                return True, f"NIP '{nip}' sudah digunakan oleh pembimbing", "pembimbing"
        
        # Cek NIP di admin
        if exclude_table != 'admin' or exclude_id is None:
            cursor.execute('SELECT id, nip FROM admin WHERE nip = %s', (nip,))
            admin = cursor.fetchone()
            if admin and (exclude_id is None or admin[0] != exclude_id):
                return True, f"NIP '{nip}' sudah digunakan oleh admin", "admin"
    
    # Cek username di reviewer
    if username:
        if exclude_table != 'reviewer' or exclude_id is None:
            cursor.execute('SELECT id, username FROM reviewer WHERE username = %s', (username,))
            reviewer = cursor.fetchone()
            if reviewer and (exclude_id is None or reviewer[0] != exclude_id):
                return True, f"Username '{username}' sudah digunakan oleh reviewer", "reviewer"
        
        # Cek username di tabel lain sebagai nama
        if exclude_table != 'mahasiswa' or exclude_id is None:
            cursor.execute('SELECT id, nama_ketua FROM mahasiswa WHERE nama_ketua = %s', (username,))
            mahasiswa = cursor.fetchone()
            if mahasiswa and (exclude_id is None or mahasiswa[0] != exclude_id):
                return True, f"Username '{username}' sudah digunakan sebagai nama mahasiswa", "mahasiswa"
        
        if exclude_table != 'pembimbing' or exclude_id is None:
            cursor.execute('SELECT id, nama FROM pembimbing WHERE nama = %s', (username,))
            pembimbing = cursor.fetchone()
            if pembimbing and (exclude_id is None or pembimbing[0] != exclude_id):
                return True, f"Username '{username}' sudah digunakan sebagai nama pembimbing", "pembimbing"
        
        if exclude_table != 'admin' or exclude_id is None:
            cursor.execute('SELECT id, nama FROM admin WHERE nama = %s', (username,))
            admin = cursor.fetchone()
            if admin and (exclude_id is None or admin[0] != exclude_id):
                return True, f"Username '{username}' sudah digunakan sebagai nama admin", "admin"
    
    # Cek email di mahasiswa
    if email:
        if exclude_table != 'mahasiswa' or exclude_id is None:
            cursor.execute('SELECT id, email FROM mahasiswa WHERE email = %s', (email,))
            mahasiswa = cursor.fetchone()
            if mahasiswa and (exclude_id is None or mahasiswa[0] != exclude_id):
                return True, f"Email '{email}' sudah digunakan oleh mahasiswa", "mahasiswa"
    
    return False, "", ""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import tempfile
from datetime import datetime, timedelta
import os
import re
import MySQLdb
import traceback
import logging
from datetime import datetime
from werkzeug.security import generate_password_hash, check_password_hash

# Lazy import untuk menghindari circular import
def get_app_functions():
    """Lazy import untuk mendapatkan fungsi dari app.py"""
    from app import (
        mysql, generate_excel_laba_rugi, generate_pdf_laba_rugi, generate_word_laba_rugi,
        generate_excel_arus_kas, generate_pdf_arus_kas, generate_word_arus_kas,
        generate_excel_neraca, generate_pdf_neraca, generate_word_neraca,
        hitung_neraca_real_time,
        hapus_semua_file_proposal, hapus_file_laporan_kemajuan, hapus_file_laporan_akhir
    )
    return {
        'mysql': mysql,
        'generate_excel_laba_rugi': generate_excel_laba_rugi,
        'generate_pdf_laba_rugi': generate_pdf_laba_rugi,
        'generate_word_laba_rugi': generate_word_laba_rugi,
        'generate_excel_arus_kas': generate_excel_arus_kas,
        'generate_pdf_arus_kas': generate_pdf_arus_kas,
        'generate_word_arus_kas': generate_word_arus_kas,
        'generate_excel_neraca': generate_excel_neraca,
        'generate_pdf_neraca': generate_pdf_neraca,
        'generate_word_neraca': generate_word_neraca,
        'hitung_neraca_real_time': hitung_neraca_real_time,
        'hapus_semua_file_proposal': hapus_semua_file_proposal,
        'hapus_file_laporan_kemajuan': hapus_file_laporan_kemajuan,
        'hapus_file_laporan_akhir': hapus_file_laporan_akhir
    }

# Import dari utils
from utils import group_anggaran_data, flatten_anggaran_data, format_tanggal_indonesia

# Setup logging untuk admin blueprint (tanpa file handler)
logging.basicConfig(
    level=logging.DEBUG,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.StreamHandler()  # Hanya console handler, tidak ada file handler
    ]
)
logger = logging.getLogger('admin')

# Import fungsi hitung_neraca_real_time akan dilakukan secara lazy di dalam fungsi

# Create blueprint
admin_bp = Blueprint('admin', __name__, url_prefix='/admin')

def generate_metadata_kolom_penilaian(interval_tipe, interval_nilai, tanggal_mulai, tanggal_selesai):
    """
    Generate metadata kolom berdasarkan interval yang dipilih admin
    """
    metadata = []
    
    if not interval_tipe or not interval_nilai or not tanggal_mulai or not tanggal_selesai:
        print(f"DEBUG: Missing required parameters - interval_tipe: {interval_tipe}, interval_nilai: {interval_nilai}, tanggal_mulai: {tanggal_mulai}, tanggal_selesai: {tanggal_selesai}")
        return metadata
    
    # Pastikan tanggal_mulai dan tanggal_selesai adalah datetime objects
    if isinstance(tanggal_mulai, str):
        try:
            if 'T' in tanggal_mulai:
                tanggal_mulai = datetime.strptime(tanggal_mulai, '%Y-%m-%dT%H:%M')
            else:
                tanggal_mulai = datetime.strptime(tanggal_mulai, '%Y-%m-%d %H:%M:%S')
        except Exception as e:
            print(f"DEBUG: Error parsing tanggal_mulai '{tanggal_mulai}': {str(e)}")
            return metadata
    
    if isinstance(tanggal_selesai, str):
        try:
            if 'T' in tanggal_selesai:
                tanggal_selesai = datetime.strptime(tanggal_selesai, '%Y-%m-%dT%H:%M')
            else:
                tanggal_selesai = datetime.strptime(tanggal_selesai, '%Y-%m-%d %H:%M:%S')
        except Exception as e:
            print(f"DEBUG: Error parsing tanggal_selesai '{tanggal_selesai}': {str(e)}")
            return metadata
    
    # Validasi tanggal
    if tanggal_mulai >= tanggal_selesai:
        print(f"DEBUG: Invalid date range - tanggal_mulai ({tanggal_mulai}) >= tanggal_selesai ({tanggal_selesai})")
        return metadata
    
    print(f"DEBUG: Generating metadata for {interval_tipe} with interval {interval_nilai}")
    print(f"DEBUG: Date range from {tanggal_mulai} to {tanggal_selesai}")
    
    if interval_tipe == 'setiap_jam':
        # PERBAIKAN: Generate kolom dengan validasi batas waktu
        current_time = tanggal_mulai
        urutan = 1
        
        while current_time <= tanggal_selesai:
            metadata.append({
                'nama_kolom': f"Jam {current_time.strftime('%H:%M')}",
                'urutan_kolom': urutan,
                'interval_tipe': interval_tipe,
                'interval_nilai': interval_nilai,
                'tanggal_mulai': current_time.date(),
                'tanggal_selesai': current_time.date()
            })
            
            current_time += timedelta(hours=interval_nilai)
            urutan += 1
            
            # Validasi batas waktu
            if current_time > tanggal_selesai:
                break
            
    elif interval_tipe == 'harian':
        # PERBAIKAN: Generate kolom berdasarkan tanggal yang sebenarnya dengan validasi batas waktu
        hari_names = ['Senin', 'Selasa', 'Rabu', 'Kamis', 'Jumat', 'Sabtu', 'Minggu']
        current_date = tanggal_mulai
        urutan = 1
        
        while current_date <= tanggal_selesai:
            hari_index = current_date.weekday()  # 0=Senin, 1=Selasa, dst
            nama_hari = hari_names[hari_index]
            metadata.append({
                'nama_kolom': f"{nama_hari} ({current_date.strftime('%d/%m')})",
                'urutan_kolom': urutan,
                'interval_tipe': interval_tipe,
                'interval_nilai': interval_nilai,
                'tanggal_mulai': current_date.date(),
                'tanggal_selesai': current_date.date()
            })
            current_date += timedelta(days=interval_nilai)
            urutan += 1
            
            # Validasi batas waktu
            if current_date > tanggal_selesai:
                break
            
    elif interval_tipe == 'mingguan':
        # PERBAIKAN: Generate kolom berdasarkan minggu yang sebenarnya dengan validasi batas waktu
        current_date = tanggal_mulai
        urutan = 1
        
        while current_date <= tanggal_selesai:
            metadata.append({
                'nama_kolom': f"Minggu {urutan} ({current_date.strftime('%d/%m')})",
                'urutan_kolom': urutan,
                'interval_tipe': interval_tipe,
                'interval_nilai': interval_nilai,
                'tanggal_mulai': current_date.date(),
                'tanggal_selesai': (current_date + timedelta(days=6)).date()
            })
            current_date += timedelta(weeks=interval_nilai)
            urutan += 1
            
            # Validasi batas waktu
            if current_date > tanggal_selesai:
                break
            
    elif interval_tipe == 'bulanan':
        # PERBAIKAN: Generate kolom berdasarkan bulan yang sebenarnya dengan validasi batas waktu
        bulan_names = ['Januari', 'Februari', 'Maret', 'April', 'Mei', 'Juni', 
                      'Juli', 'Agustus', 'September', 'Oktober', 'November', 'Desember']
        
        current_date = tanggal_mulai
        urutan = 1
        
        while current_date <= tanggal_selesai:
            bulan_index = current_date.month - 1  # 0-based index
            nama_bulan = bulan_names[bulan_index]
            
            # Hitung tanggal akhir bulan
            if current_date.month == 12:
                next_month = current_date.replace(year=current_date.year + 1, month=1)
            else:
                next_month = current_date.replace(month=current_date.month + 1)
            tanggal_selesai_bulan = (next_month - timedelta(days=1)).date()
            
            metadata.append({
                'nama_kolom': f"{nama_bulan} {current_date.year}",
                'urutan_kolom': urutan,
                'interval_tipe': interval_tipe,
                'interval_nilai': interval_nilai,
                'tanggal_mulai': current_date.date(),
                'tanggal_selesai': tanggal_selesai_bulan
            })
            
            # Pindah ke bulan berikutnya
            if current_date.month == 12:
                current_date = current_date.replace(year=current_date.year + 1, month=1)
            else:
                current_date = current_date.replace(month=current_date.month + 1)
            urutan += 1
            
            # Validasi batas waktu
            if current_date > tanggal_selesai:
                break
    
    print(f"DEBUG: Generated {len(metadata)} metadata columns")
    return metadata


@admin_bp.route('/das_admin')
def das_admin():
    logger.info("Dashboard admin dipanggil")
    if 'user_type' not in session or session['user_type'] != 'admin':
        logger.warning(f"Akses ditolak untuk dashboard admin: {session.get('user_type', 'None')}")
        flash('Anda harus login sebagai admin!', 'danger')
        return redirect(url_for('index'))
    
    if not hasattr(get_app_functions()['mysql'], 'connection') or get_app_functions()['mysql'].connection is None:
        logger.error("Koneksi database tidak tersedia untuk dashboard admin")
        flash('Koneksi ke database gagal. Cek konfigurasi database!', 'danger')
        return render_template('admin/index admin.html', 
                             total_mahasiswa=0,
                             proposal_diajukan=0,
                             proposal_disetujui=0,
                             proposal_lolos=0,
                             pembimbing_list=[],
                             mahasiswa_list=[])
    
    try:
        cursor = get_app_functions()['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Hitung total mahasiswa dengan status aktif
        cursor.execute('''
            SELECT COUNT(*) as total
            FROM mahasiswa 
            WHERE status = 'aktif'
        ''')
        total_mahasiswa = cursor.fetchone()['total']
        
        # Hitung proposal dengan status diajukan
        cursor.execute('''
            SELECT COUNT(*) as total
            FROM proposal 
            WHERE status = 'diajukan'
        ''')
        proposal_diajukan = cursor.fetchone()['total']
        
        # Hitung proposal dengan status disetujui
        cursor.execute('''
            SELECT COUNT(*) as total
            FROM proposal 
            WHERE status = 'disetujui'
        ''')
        proposal_disetujui = cursor.fetchone()['total']
        
        # Hitung proposal yang lolos pendanaan
        cursor.execute('''
            SELECT COUNT(*) as total
            FROM proposal 
            WHERE status_admin = 'lolos'
        ''')
        proposal_lolos = cursor.fetchone()['total']
        
        # Ambil data pembimbing dengan jumlah mahasiswa yang dibimbing
        cursor.execute('''
            SELECT p.*, 
                   COALESCE(COUNT(DISTINCT m.id), 0) as jumlah_mahasiswa
            FROM pembimbing p
            LEFT JOIN proposal pr ON p.nama = pr.dosen_pembimbing
            LEFT JOIN mahasiswa m ON pr.nim = m.nim AND m.status = 'aktif'
            WHERE p.status = 'aktif'
            GROUP BY p.id, p.nama, p.nip, p.program_studi, p.password, p.tanggal_dibuat, p.status, p.kuota_mahasiswa
            ORDER BY p.nama ASC
        ''')
        
        pembimbing_list = cursor.fetchall()
        
        # Ambil data mahasiswa aktif dengan informasi proposal
        try:
            cursor.execute('''
                SELECT m.id, m.nim, m.nama_ketua, m.program_studi, m.tanggal_daftar, m.status,
                       p.judul_usaha, p.status as status_proposal, p.status_admin, p.tanggal_buat, 
                       CASE 
                           WHEN p.dosen_pembimbing IS NOT NULL AND p.dosen_pembimbing != '' 
                           THEN p.dosen_pembimbing 
                           ELSE 'Belum ditentukan' 
                       END as dosen_pembimbing,
                       COALESCE(COUNT(at.id), 0) as jumlah_anggota
                FROM mahasiswa m
                LEFT JOIN proposal p ON m.nim = p.nim
                LEFT JOIN anggota_tim at ON p.id = at.id_proposal
                WHERE m.status = 'aktif'
                GROUP BY m.id, m.nim, m.nama_ketua, m.program_studi, m.tanggal_daftar, m.status, p.judul_usaha, p.status, p.status_admin, p.tanggal_buat, p.dosen_pembimbing
                ORDER BY m.nama_ketua ASC
            ''')
            
            mahasiswa_list = cursor.fetchall()
            
            # Debug: print data untuk memeriksa
            print("Debug - Mahasiswa data:")
            for mahasiswa in mahasiswa_list:
                print(f"NIM: {mahasiswa['nim']}, Pembimbing: {mahasiswa.get('dosen_pembimbing', 'NULL')}")
                
        except Exception as e:
            logger.error(f"Error fetching mahasiswa data: {str(e)}")
            logger.error(traceback.format_exc())
            # Fallback query jika ada masalah dengan GROUP BY
            cursor.execute('''
                SELECT m.id, m.nim, m.nama_ketua, m.program_studi, m.tanggal_daftar, m.status,
                       p.judul_usaha, p.status as status_proposal, p.status_admin, p.tanggal_buat, 
                       CASE 
                           WHEN p.dosen_pembimbing IS NOT NULL AND p.dosen_pembimbing != '' 
                           THEN p.dosen_pembimbing 
                           ELSE 'Belum ditentukan' 
                       END as dosen_pembimbing
                FROM mahasiswa m
                LEFT JOIN proposal p ON m.nim = p.nim
                WHERE m.status = 'aktif'
                ORDER BY m.nama_ketua ASC
            ''')
            
            mahasiswa_list = cursor.fetchall()
            # Tambahkan jumlah_anggota default dan status_admin default
            for mahasiswa in mahasiswa_list:
                mahasiswa['jumlah_anggota'] = 0
                if 'status_admin' not in mahasiswa:
                    mahasiswa['status_admin'] = None
        
        cursor.close()
        
        return render_template('admin/index admin.html', 
                             total_mahasiswa=total_mahasiswa,
                             proposal_diajukan=proposal_diajukan,
                             proposal_disetujui=proposal_disetujui,
                             proposal_lolos=proposal_lolos,
                             pembimbing_list=pembimbing_list,
                             mahasiswa_list=mahasiswa_list)
        
    except Exception as e:
        flash(f'Error saat mengambil data statistik: {str(e)}', 'danger')
        return render_template('admin/index admin.html', 
                             total_mahasiswa=0,
                             proposal_diajukan=0,
                             proposal_disetujui=0,
                             proposal_lolos=0,
                             pembimbing_list=[],
                             mahasiswa_list=[])

@admin_bp.route('/pengajuan_proposal')
def pengajuan_proposal():
    if 'user_type' not in session or session['user_type'] != 'admin':
        flash('Anda harus login sebagai admin!', 'danger')
        return redirect(url_for('index'))
    
    if not hasattr(get_app_functions()['mysql'], 'connection') or get_app_functions()['mysql'].connection is None:
        flash('Koneksi ke database gagal. Cek konfigurasi database!', 'danger')
        return render_template('admin/proposal.html', proposals=[])
    
    try:
        cursor = get_app_functions()['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Ambil data proposal yang sudah disetujui pembimbing dengan informasi mahasiswa dan anggota tim
        cursor.execute('''
            SELECT p.*, 
                   m.nama_ketua, m.perguruan_tinggi, m.program_studi, m.email, m.no_telp,
                   COUNT(at.id) as jumlah_anggota,
                   (SELECT COUNT(*) FROM anggaran_awal aa 
                    WHERE aa.id_proposal = p.id AND aa.status_reviewer = 'sudah_direview') as anggaran_awal_reviewed,
                   (SELECT COUNT(*) FROM anggaran_bertumbuh ab 
                    WHERE ab.id_proposal = p.id AND ab.status_reviewer = 'sudah_direview') as anggaran_bertumbuh_reviewed,
                   (SELECT COUNT(*) FROM anggaran_awal aa 
                    WHERE aa.id_proposal = p.id) as total_anggaran_awal,
                   (SELECT COUNT(*) FROM anggaran_bertumbuh ab 
                    WHERE ab.id_proposal = p.id) as total_anggaran_bertumbuh,
                   (SELECT COUNT(*) FROM anggaran_awal aa 
                    WHERE aa.id_proposal = p.id AND aa.status = 'revisi') as anggaran_awal_revisi,
                   (SELECT COUNT(*) FROM anggaran_bertumbuh ab 
                    WHERE ab.id_proposal = p.id AND ab.status = 'revisi') as anggaran_bertumbuh_revisi,
                   (SELECT COUNT(*) FROM anggaran_awal aa 
                    WHERE aa.id_proposal = p.id AND aa.status_reviewer = 'tolak_bantuan') as anggaran_awal_ditolak,
                   (SELECT COUNT(*) FROM anggaran_bertumbuh ab 
                    WHERE ab.id_proposal = p.id AND ab.status_reviewer = 'tolak_bantuan') as anggaran_bertumbuh_ditolak,
                   (SELECT pr.status_review FROM proposal_reviewer pr 
                    WHERE pr.id_proposal = p.id ORDER BY pr.tanggal_assign DESC LIMIT 1) as reviewer_status,
                   (SELECT r.nama FROM proposal_reviewer pr 
                    JOIN reviewer r ON pr.id_reviewer = r.id 
                    WHERE pr.id_proposal = p.id ORDER BY pr.tanggal_assign DESC LIMIT 1) as nama_reviewer
            FROM proposal p 
            LEFT JOIN mahasiswa m ON p.nim = m.nim
            LEFT JOIN anggota_tim at ON p.id = at.id_proposal
            WHERE p.status IN ('disetujui', 'revisi', 'selesai')
            GROUP BY p.id 
            ORDER BY p.tanggal_konfirmasi_pembimbing DESC
        ''')
        
        proposals = cursor.fetchall()
        cursor.close()
        
        return render_template('admin/proposal.html', proposals=proposals)
        
    except Exception as e:
        flash(f'Error saat mengambil data proposal: {str(e)}', 'danger')
        return render_template('admin/proposal.html', proposals=[])

@admin_bp.route('/get_proposal_detail/<int:proposal_id>')
def admin_get_proposal_detail(proposal_id):
    if 'user_type' not in session or session['user_type'] != 'admin':
        return jsonify({'success': False, 'message': 'Anda harus login sebagai admin!'})
    
    if not hasattr(get_app_functions()['mysql'], 'connection') or get_app_functions()['mysql'].connection is None:
        return jsonify({'success': False, 'message': 'Koneksi ke database gagal!'})
    
    try:
        cursor = get_app_functions()['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Get proposal details with mahasiswa info
        cursor.execute('''
            SELECT p.*, m.nama_ketua, m.perguruan_tinggi, m.program_studi, m.email, m.no_telp
            FROM proposal p 
            LEFT JOIN mahasiswa m ON p.nim = m.nim
            WHERE p.id = %s
        ''', (proposal_id,))
        
        proposal = cursor.fetchone()
        if not proposal:
            cursor.close()
            return jsonify({'success': False, 'message': 'Proposal tidak ditemukan!'})
        
        # Get team members
        cursor.execute('''
            SELECT * FROM anggota_tim 
            WHERE id_proposal = %s
            ORDER BY id
        ''', (proposal_id,))
        
        anggota = cursor.fetchall()
        
        # Get anggaran data if exists
        cursor.execute('''
            SELECT COUNT(*) as count FROM anggaran_awal WHERE id_proposal = %s
        ''', (proposal_id,))
        anggaran_awal_count = cursor.fetchone()['count']
        
        cursor.execute('''
            SELECT COUNT(*) as count FROM anggaran_bertumbuh WHERE id_proposal = %s
        ''', (proposal_id,))
        anggaran_bertumbuh_count = cursor.fetchone()['count']
        
        cursor.close()
        
        return jsonify({
            'success': True, 
            'proposal': proposal,
            'anggota': anggota,
            'anggaran_awal_count': anggaran_awal_count,
            'anggaran_bertumbuh_count': anggaran_bertumbuh_count
        })
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error saat mengambil data: {str(e)}'})

@admin_bp.route('/update_proposal_status', methods=['POST'])
def update_proposal_status():
    if 'user_type' not in session or session['user_type'] != 'admin':
        return jsonify({'success': False, 'message': 'Anda harus login sebagai admin!'})
    
    if not hasattr(get_app_functions()['mysql'], 'connection') or get_app_functions()['mysql'].connection is None:
        return jsonify({'success': False, 'message': 'Koneksi ke database gagal!'})
    
    try:
        data = request.get_json()
        proposal_id = data.get('proposal_id')
        new_status = data.get('status')  # 'disetujui', 'ditolak', 'revisi', 'draf', 'diajukan'
        
        if not proposal_id or not new_status:
            return jsonify({'success': False, 'message': 'Data tidak lengkap!'})
        
        # Validasi status yang diizinkan
        if new_status not in ['disetujui', 'ditolak', 'revisi', 'draf', 'diajukan']:
            return jsonify({'success': False, 'message': 'Status tidak valid!'})
        
        cursor = get_app_functions()['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Update status proposal
        cursor.execute('''
            UPDATE proposal 
            SET status = %s, tanggal_review = NOW()
            WHERE id = %s
        ''', (new_status, proposal_id))
        
        get_app_functions()['mysql'].connection.commit()
        cursor.close()
        
        return jsonify({'success': True, 'message': f'Status proposal berhasil diubah menjadi {new_status}'})
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error saat update status: {str(e)}'})

@admin_bp.route('/download_proposal/<int:proposal_id>')
def download_proposal(proposal_id):
    if 'user_type' not in session or session['user_type'] != 'admin':
        flash('Anda harus login sebagai admin!', 'danger')
        return redirect(url_for('index'))
    
    if not hasattr(get_app_functions()['mysql'], 'connection') or get_app_functions()['mysql'].connection is None:
        flash('Koneksi ke database gagal!', 'danger')
        return redirect(url_for('pengajuan_proposal'))
    
    try:
        cursor = get_app_functions()['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Get proposal file path
        cursor.execute('SELECT file_path, judul_usaha FROM proposal WHERE id = %s', (proposal_id,))
        proposal = cursor.fetchone()
        cursor.close()
        
        if not proposal or not proposal['file_path']:
            flash('File proposal tidak ditemukan!', 'danger')
            return redirect(url_for('pengajuan_proposal'))
        
        file_path = proposal['file_path']
        
        if not os.path.exists(file_path):
            flash('File proposal tidak ditemukan di server!', 'danger')
            return redirect(url_for('pengajuan_proposal'))
        
        # Get file extension and create download filename
        file_extension = os.path.splitext(file_path)[1]
        
        # Ambil nama file asli dari path
        original_filename = os.path.basename(file_path)
        
        # Jika nama file sudah sesuai format yang diinginkan, gunakan itu
        if '_proposal.' in original_filename:
            download_filename = original_filename
        else:
            # Jika tidak, buat nama file baru dengan format yang diinginkan
            download_filename = f"{proposal['judul_usaha'].replace(' ', '_')}_proposal{file_extension}"
        
        return send_file(file_path, as_attachment=True, download_name=download_filename)
        
    except Exception as e:
        flash(f'Error saat download file: {str(e)}', 'danger')
        return redirect(url_for('pengajuan_proposal'))

@admin_bp.route('/mahasiswa_belum_diverifikasi')
def mahasiswa_belum_diverifikasi():
    if 'user_type' not in session or session['user_type'] != 'admin':
        flash('Anda harus login sebagai admin!', 'danger')
        return redirect(url_for('index'))
    
    if not hasattr(get_app_functions()['mysql'], 'connection') or get_app_functions()['mysql'].connection is None:
        flash('Koneksi ke database gagal. Cek konfigurasi database!', 'danger')
        return render_template('admin/mahasiswa belum.html', mahasiswa_list=[])
    
    try:
        cursor = get_app_functions()['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Ambil data mahasiswa yang belum diverifikasi (status = 'proses')
        cursor.execute('''
            SELECT id, perguruan_tinggi, program_studi, nim, nama_ketua, no_telp, email, status, tanggal_daftar
            FROM mahasiswa 
            WHERE status = 'proses' 
            ORDER BY tanggal_daftar DESC
        ''')
        
        mahasiswa_list = cursor.fetchall()
        cursor.close()
        
        return render_template('admin/mahasiswa belum.html', mahasiswa_list=mahasiswa_list)
        
    except Exception as e:
        flash(f'Error saat mengambil data mahasiswa: {str(e)}', 'danger')
        return render_template('admin/mahasiswa belum.html', mahasiswa_list=[])

@admin_bp.route('/update_status_mahasiswa', methods=['POST'])
def update_status_mahasiswa():
    if 'user_type' not in session or session['user_type'] != 'admin':
        return {'success': False, 'message': 'Anda harus login sebagai admin!'}
    
    if not hasattr(get_app_functions()['mysql'], 'connection') or get_app_functions()['mysql'].connection is None:
        return {'success': False, 'message': 'Koneksi ke database gagal!'}
    
    try:
        data = request.get_json()
        mahasiswa_id = data.get('mahasiswa_id')
        new_status = data.get('status')  # 'aktif' atau 'tolak'
        
        if not mahasiswa_id or not new_status:
            return {'success': False, 'message': 'Data tidak lengkap!'}
        
        # Validasi status yang diizinkan
        if new_status not in ['aktif', 'tolak', 'proses', 'selesai']:
            return {'success': False, 'message': 'Status tidak valid!'}
        
        # Gunakan status langsung karena sudah sesuai dengan database
        db_status = new_status
        
        cursor = get_app_functions()['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Update status mahasiswa
        cursor.execute('''
            UPDATE mahasiswa 
            SET status = %s, tanggal_verifikasi = NOW()
            WHERE id = %s
        ''', (db_status, mahasiswa_id))
        
        if hasattr(get_app_functions()['mysql'], 'connection') and get_app_functions()['mysql'].connection is not None:
            get_app_functions()['mysql'].connection.commit()
        else:
            return {'success': False, 'message': 'Koneksi database gagal saat commit!'}
        
        cursor.close()
        
        # Tampilkan pesan yang sesuai dengan input user
        status_display = {
            'aktif': 'Aktif',
            'tolak': 'Ditolak',
            'proses': 'Proses',
            'selesai': 'Selesai'
        }
        display_status = status_display.get(new_status, new_status)
        return {'success': True, 'message': f'Status mahasiswa berhasil diubah menjadi {display_status}'}
        
    except Exception as e:
        return {'success': False, 'message': f'Error saat update status: {str(e)}'}

@admin_bp.route('/mahasiswa_sudah_diverifikasi')
def mahasiswa_sudah_diverifikasi():
    """
    Halaman untuk menampilkan mahasiswa yang sudah diverifikasi.
    Menampilkan mahasiswa dengan status 'aktif', 'selesai', dan 'tolak'.
    Data diurutkan berdasarkan status (aktif -> selesai -> tolak) kemudian berdasarkan tanggal_daftar DESC.
    """
    if 'user_type' not in session or session['user_type'] != 'admin':
        flash('Anda harus login sebagai admin!', 'danger')
        return redirect(url_for('index'))
    
    if not hasattr(get_app_functions()['mysql'], 'connection') or get_app_functions()['mysql'].connection is None:
        flash('Koneksi ke database gagal. Cek konfigurasi database!', 'danger')
        return render_template('admin/mahasiswa sudah.html', mahasiswa_list=[])
    
    try:
        cursor = get_app_functions()['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Ambil data mahasiswa yang sudah diverifikasi (status = 'aktif', 'selesai', atau 'tolak')
        cursor.execute('''
            SELECT id, perguruan_tinggi, program_studi, nim, nama_ketua, no_telp, email, status, tanggal_daftar
            FROM mahasiswa 
            WHERE status IN ('aktif', 'selesai', 'tolak') 
            ORDER BY FIELD(status, 'aktif', 'selesai', 'tolak'), tanggal_daftar DESC
        ''')
        
        mahasiswa_list = cursor.fetchall()
        cursor.close()
        
        return render_template('admin/mahasiswa sudah.html', mahasiswa_list=mahasiswa_list)
        
    except Exception as e:
        flash(f'Error saat mengambil data mahasiswa: {str(e)}', 'danger')
        return render_template('admin/mahasiswa sudah.html', mahasiswa_list=[])

@admin_bp.route('/laporan_akhir')
def admin_laporan_akhir():
    if 'user_type' not in session or session['user_type'] != 'admin':
        flash('Anda harus login sebagai admin!', 'danger')
        return redirect(url_for('index'))
    
    if not hasattr(get_app_functions()['mysql'], 'connection') or get_app_functions()['mysql'].connection is None:
        flash('Koneksi ke database gagal. Cek konfigurasi database!', 'danger')
        return render_template('admin/laporan_akhir.html', mahasiswa_list=[])
    
    try:
        cursor = get_app_functions()['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Ambil semua mahasiswa dengan proposal
        cursor.execute('''
            SELECT DISTINCT
                m.id,
                m.nama_ketua,
                m.nim,
                m.perguruan_tinggi,
                p.judul_usaha,
                p.tahapan_usaha,
                p.dosen_pembimbing,
                p.id as proposal_id
            FROM mahasiswa m
            INNER JOIN proposal p ON m.nim = p.nim
            ORDER BY m.nama_ketua
        ''')
        
        mahasiswa_all = cursor.fetchall()
        print(f"DEBUG ADMIN LAPORAN AKHIR: Total mahasiswa: {len(mahasiswa_all)}")
        mahasiswa_list = []
        
        for mhs in mahasiswa_all:
            proposal_id = mhs['proposal_id']
            tahapan = (mhs.get('tahapan_usaha') or '').lower()
            
            # Tambahkan kolom rekomendasi_pendanaan jika belum ada
            try:
                cursor.execute('ALTER TABLE penilaian_laporan_kemajuan ADD COLUMN rekomendasi_pendanaan VARCHAR(50) DEFAULT NULL')
                get_app_functions()['mysql'].connection.commit()
                print("DEBUG: Berhasil menambahkan kolom rekomendasi_pendanaan ke tabel penilaian_laporan_kemajuan")
            except Exception as e:
                if "Duplicate column name" in str(e):
                    print("DEBUG: Kolom rekomendasi_pendanaan sudah ada di tabel penilaian_laporan_kemajuan")
                else:
                    print(f"DEBUG: Error menambahkan kolom rekomendasi_pendanaan: {str(e)}")
            
            if 'bertumbuh' in tahapan:
                tabel_laporan_akhir = 'laporan_akhir_bertumbuh'
                tabel_laporan_kemajuan = 'laporan_kemajuan_bertumbuh'
            else:
                tabel_laporan_akhir = 'laporan_akhir_awal'
                tabel_laporan_kemajuan = 'laporan_kemajuan_awal'
            
            # TRIGGER: Cek apakah sudah ada data laporan akhir untuk proposal ini
            cursor.execute(f"SELECT COUNT(*) as cnt FROM {tabel_laporan_akhir} WHERE id_proposal = %s", (proposal_id,))
            cnt = cursor.fetchone()['cnt']
            
            # Jika belum ada data laporan akhir, buat dari laporan kemajuan yang disetujui
            if cnt == 0:
                print(f"DEBUG: Membuat laporan akhir untuk proposal {proposal_id}")
                print(f"DEBUG: Menggunakan tabel laporan kemajuan: {tabel_laporan_kemajuan}")
                
              # Tambahkan kolom nilai_bantuan jika belum ada
                try:
                    cursor.execute(f'ALTER TABLE {tabel_laporan_akhir} ADD COLUMN nilai_bantuan DECIMAL(15,2) DEFAULT 0.00')
                    get_app_functions()['mysql'].connection.commit()
                    print(f"DEBUG: Berhasil menambahkan kolom nilai_bantuan ke tabel {tabel_laporan_akhir}")
                except Exception as e:
                    if "Duplicate column name" in str(e):
                        print(f"DEBUG: Kolom nilai_bantuan sudah ada di tabel {tabel_laporan_akhir}")
                    else:
                        print(f"DEBUG: Error menambahkan kolom nilai_bantuan: {str(e)}")
                
                # Cek status rekomendasi pendanaan dari penilaian laporan kemajuan
                cursor.execute('''
                    SELECT rekomendasi_pendanaan 
                    FROM penilaian_laporan_kemajuan 
                    WHERE id_proposal = %s AND rekomendasi_pendanaan IS NOT NULL
                    ORDER BY id DESC 
                    LIMIT 1
                ''', (proposal_id,))
                
                penilaian_result = cursor.fetchone()
                rekomendasi_pendanaan = penilaian_result['rekomendasi_pendanaan'] if penilaian_result else None
                
                print(f"DEBUG: Rekomendasi pendanaan untuk proposal {proposal_id}: {rekomendasi_pendanaan}")
                
                # Jika rekomendasi adalah 'berhenti pendanaan', tidak buat laporan akhir
                if rekomendasi_pendanaan == 'berhenti pendanaan':
                    print(f"DEBUG: Rekomendasi pendanaan 'berhenti pendanaan' untuk proposal {proposal_id}. Laporan akhir tidak dibuat.")
                    continue
                
                # Jika rekomendasi adalah 'lanjutkan pendanaan' atau belum ada penilaian, cek laporan kemajuan yang disetujui
                if rekomendasi_pendanaan == 'lanjutkan pendanaan' or rekomendasi_pendanaan is None:
                    cursor.execute(f'''
                        SELECT COUNT(*) as cnt_kemajuan 
                        FROM {tabel_laporan_kemajuan} 
                        WHERE id_proposal = %s AND status = 'disetujui'
                    ''', (proposal_id,))
                    cnt_kemajuan = cursor.fetchone()['cnt_kemajuan']
                    print(f"DEBUG: Ditemukan {cnt_kemajuan} data laporan kemajuan dengan status 'disetujui'")
                    
                    if cnt_kemajuan > 0:
                        print(f"DEBUG: Menyalin data dari laporan kemajuan ke laporan akhir dengan reset field tertentu")
                        try:
                            cursor.execute(f'''
                                SELECT id, kegiatan_utama, kegiatan, penanggung_jawab, target_capaian, 
                                       nama_barang, satuan, status, nilai_bantuan
                                FROM {tabel_laporan_kemajuan} 
                                WHERE id_proposal = %s AND status = 'disetujui'
                                ORDER BY kegiatan_utama, kegiatan, nama_barang
                            ''', (proposal_id,))
                            kemajuan_data = cursor.fetchall()
                            print(f"DEBUG: Berhasil mengambil {len(kemajuan_data)} data laporan kemajuan")
                            
                            if kemajuan_data:
                                print(f"DEBUG: Sample data - kegiatan_utama: {kemajuan_data[0].get('kegiatan_utama', 'N/A')}")
                                print(f"DEBUG: Sample data - nama_barang: {kemajuan_data[0].get('nama_barang', 'N/A')}")
                                
                                for kemajuan in kemajuan_data:
                                    try:
                                        cursor.execute(f'''
                                            INSERT INTO {tabel_laporan_akhir} (
                                                id_proposal, kegiatan_utama, kegiatan, penanggung_jawab, target_capaian,
                                                nama_barang, kuantitas, satuan, harga_satuan, jumlah, keterangan, status, nilai_bantuan
                                            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                                        ''', (
                                            proposal_id, kemajuan['kegiatan_utama'], kemajuan['kegiatan'], 
                                            kemajuan['penanggung_jawab'], kemajuan['target_capaian'], kemajuan['nama_barang'],
                                            0, kemajuan['satuan'], 0, 0, '', 'draf', kemajuan.get('nilai_bantuan', 0)
                                        ))
                                        print(f"DEBUG: Berhasil INSERT data untuk kegiatan: {kemajuan.get('kegiatan_utama', 'N/A')}")
                                    except Exception as insert_error:
                                        print(f"DEBUG: Error saat INSERT: {str(insert_error)}")
                                        print(f"DEBUG: Data yang gagal INSERT: {kemajuan}")
                                        raise insert_error
                                
                                get_app_functions()['mysql'].connection.commit()
                                print(f"DEBUG: Berhasil membuat {len(kemajuan_data)} data laporan akhir untuk proposal {proposal_id}")
                                print(f"DEBUG: Field yang disalin: kegiatan_utama, kegiatan, penanggung_jawab, target_capaian, nama_barang, satuan, nilai_bantuan")
                                print(f"DEBUG: Field yang di-reset: kuantitas=0, harga_satuan=0, jumlah=0, keterangan=''")
                            else:
                                print(f"DEBUG: Tidak ada data laporan kemajuan yang disetujui untuk proposal {proposal_id}")
                        except Exception as e:
                            print(f"DEBUG: Error dalam proses pembuatan laporan akhir: {str(e)}")
                            get_app_functions()['mysql'].connection.rollback()
                    else:
                        print(f"DEBUG: Tidak ada laporan kemajuan dengan status 'disetujui' untuk proposal {proposal_id}")
            else:
                print(f"DEBUG: Data laporan akhir sudah ada untuk proposal {proposal_id}")
            
            # Cek ulang apakah sekarang ada data laporan akhir
            cursor.execute(f"SELECT COUNT(*) as cnt FROM {tabel_laporan_akhir} WHERE id_proposal = %s", (proposal_id,))
            cnt = cursor.fetchone()['cnt']
            if cnt > 0:
                # Tentukan jenis laporan akhir
                if 'bertumbuh' in tahapan:
                    jenis_laporan = 'Bertumbuh'
                else:
                    jenis_laporan = 'Awal'
                
                mhs['jenis_laporan_akhir'] = jenis_laporan
                mahasiswa_list.append(mhs)
        
        cursor.close()
        
        return render_template('admin/laporan_akhir.html', mahasiswa_list=mahasiswa_list)
        
    except Exception as e:
        flash(f'Error: {str(e)}', 'danger')
        return render_template('admin/laporan_akhir.html', mahasiswa_list=[])

@admin_bp.route('/laporan_kemajuan')
def admin_laporan_kemajuan():
    if 'user_type' not in session or session['user_type'] != 'admin':
        flash('Anda harus login sebagai admin!', 'danger')
        return redirect(url_for('index'))
    
    if not hasattr(get_app_functions()['mysql'], 'connection') or get_app_functions()['mysql'].connection is None:
        flash('Koneksi ke database gagal. Cek konfigurasi database!', 'danger')
        return render_template('admin/laporan_kemajuan.html', mahasiswa_list=[])
    
    try:
        cursor = get_app_functions()['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Ambil semua mahasiswa dengan proposal
        cursor.execute('''
            SELECT DISTINCT
                m.id,
                m.nama_ketua,
                m.nim,
                m.perguruan_tinggi,
                p.judul_usaha,
                p.tahapan_usaha,
                p.dosen_pembimbing,
                p.id as proposal_id
            FROM mahasiswa m
            INNER JOIN proposal p ON m.nim = p.nim
            ORDER BY m.nama_ketua
        ''')
        
        mahasiswa_all = cursor.fetchall()
        print(f"DEBUG ADMIN LAPORAN KEMAJUAN: Total mahasiswa: {len(mahasiswa_all)}")
        mahasiswa_list = []
        
        for mhs in mahasiswa_all:
            proposal_id = mhs['proposal_id']
            tahapan = (mhs.get('tahapan_usaha') or '').lower()
            
            if 'bertumbuh' in tahapan:
                tabel_laporan = 'laporan_kemajuan_bertumbuh'
                tabel_anggaran = 'anggaran_bertumbuh'
            else:
                tabel_laporan = 'laporan_kemajuan_awal'
                tabel_anggaran = 'anggaran_awal'
            
            # TRIGGER: Cek apakah sudah ada data laporan kemajuan untuk proposal ini
            cursor.execute(f"SELECT COUNT(*) as cnt FROM {tabel_laporan} WHERE id_proposal = %s", (proposal_id,))
            cnt = cursor.fetchone()['cnt']
            
            # Jika belum ada data laporan kemajuan, buat dari anggaran yang sudah direview
            if cnt == 0:
                print(f"DEBUG: Membuat laporan kemajuan untuk proposal {proposal_id}")
                print(f"DEBUG: Menggunakan tabel anggaran: {tabel_anggaran}")
                
                # Tambahkan kolom nilai_bantuan jika belum ada
                try:
                    cursor.execute(f'ALTER TABLE {tabel_laporan} ADD COLUMN nilai_bantuan DECIMAL(15,2) DEFAULT 0.00')
                    get_app_functions()['mysql'].connection.commit()
                    print(f"DEBUG: Berhasil menambahkan kolom nilai_bantuan ke tabel {tabel_laporan}")
                except Exception as e:
                    if "Duplicate column name" in str(e):
                        print(f"DEBUG: Kolom nilai_bantuan sudah ada di tabel {tabel_laporan}")
                    else:
                        print(f"DEBUG: Error menambahkan kolom nilai_bantuan: {str(e)}")
                
                # Cek apakah ada data anggaran dengan status_reviewer sudah_direview
                cursor.execute(f'''
                    SELECT COUNT(*) as cnt_anggaran 
                    FROM {tabel_anggaran} 
                    WHERE id_proposal = %s AND status_reviewer = 'sudah_direview'
                ''', (proposal_id,))
                cnt_anggaran = cursor.fetchone()['cnt_anggaran']
                print(f"DEBUG: Ditemukan {cnt_anggaran} data anggaran dengan status_reviewer 'sudah_direview'")
                
                if cnt_anggaran > 0:
                    print(f"DEBUG: Menyalin data dari anggaran ke laporan kemajuan dengan reset field tertentu")
                    try:
                        cursor.execute(f'''
                            SELECT id, kegiatan_utama, kegiatan, penanggung_jawab, target_capaian, 
                                   nama_barang, satuan, status, status_reviewer, nilai_bantuan
                            FROM {tabel_anggaran} 
                            WHERE id_proposal = %s AND status_reviewer = 'sudah_direview'
                            ORDER BY kegiatan_utama, kegiatan, nama_barang
                        ''', (proposal_id,))
                        anggaran_data = cursor.fetchall()
                        print(f"DEBUG: Berhasil mengambil {len(anggaran_data)} data anggaran")
                        
                        if anggaran_data:
                            print(f"DEBUG: Sample data - kegiatan_utama: {anggaran_data[0].get('kegiatan_utama', 'N/A')}")
                            print(f"DEBUG: Sample data - nama_barang: {anggaran_data[0].get('nama_barang', 'N/A')}")
                            
                            for anggaran in anggaran_data:
                                try:
                                    cursor.execute(f'''
                                        INSERT INTO {tabel_laporan} (
                                            id_proposal, kegiatan_utama, kegiatan, penanggung_jawab, target_capaian,
                                            nama_barang, kuantitas, satuan, harga_satuan, jumlah, keterangan, status, nilai_bantuan
                                        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                                    ''', (
                                        proposal_id, anggaran['kegiatan_utama'], anggaran['kegiatan'], 
                                        anggaran['penanggung_jawab'], anggaran['target_capaian'], anggaran['nama_barang'],
                                        0, anggaran['satuan'], 0, 0, '', 'draf', anggaran.get('nilai_bantuan', 0)
                                    ))
                                    print(f"DEBUG: Berhasil INSERT data untuk kegiatan: {anggaran.get('kegiatan_utama', 'N/A')}")
                                except Exception as insert_error:
                                    print(f"DEBUG: Error saat INSERT: {str(insert_error)}")
                                    print(f"DEBUG: Data yang gagal INSERT: {anggaran}")
                                    raise insert_error
                            
                            get_app_functions()['mysql'].connection.commit()
                            print(f"DEBUG: Berhasil membuat {len(anggaran_data)} data laporan kemajuan untuk proposal {proposal_id}")
                            print(f"DEBUG: Field yang disalin: kegiatan_utama, kegiatan, penanggung_jawab, target_capaian, nama_barang, satuan, nilai_bantuan")
                            print(f"DEBUG: Field yang di-reset: kuantitas=0, harga_satuan=0, jumlah=0, keterangan=''")
                        else:
                            print(f"DEBUG: Tidak ada data anggaran yang sudah direview untuk proposal {proposal_id}")
                    except Exception as e:
                        print(f"DEBUG: Error dalam proses pembuatan laporan kemajuan: {str(e)}")
                        get_app_functions()['mysql'].connection.rollback()
                else:
                    print(f"DEBUG: Tidak ada anggaran dengan status_reviewer 'sudah_direview' untuk proposal {proposal_id}")
            else:
                print(f"DEBUG: Data laporan kemajuan sudah ada untuk proposal {proposal_id}")
            
            # Cek ulang apakah sekarang ada data laporan kemajuan
            cursor.execute(f"SELECT COUNT(*) as cnt FROM {tabel_laporan} WHERE id_proposal = %s", (proposal_id,))
            cnt = cursor.fetchone()['cnt']
            if cnt > 0:
                # Tentukan jenis laporan kemajuan
                if 'bertumbuh' in tahapan:
                    jenis_laporan = 'Bertumbuh'
                else:
                    jenis_laporan = 'Awal'
                
                mhs['jenis_laporan_kemajuan'] = jenis_laporan
                mahasiswa_list.append(mhs)
        
        cursor.close()
        
        return render_template('admin/laporan_kemajuan.html', mahasiswa_list=mahasiswa_list)
        
    except Exception as e:
        flash(f'Error: {str(e)}', 'danger')
        return render_template('admin/laporan_kemajuan.html', mahasiswa_list=[])

@admin_bp.route('/update_laporan_kemajuan_status', methods=['POST'])
def admin_update_laporan_kemajuan_status():
    if 'user_type' not in session or session['user_type'] != 'admin':
        return jsonify({'success': False, 'message': 'Anda harus login sebagai admin!'})
    
    if not hasattr(get_app_functions()['mysql'], 'connection') or get_app_functions()['mysql'].connection is None:
        return jsonify({'success': False, 'message': 'Koneksi ke database gagal!'})
    
    try:
        data = request.get_json()
        laporan_id = data.get('id')
        tabel_laporan = data.get('tabel_laporan')
        new_status = data.get('status')  # 'disetujui', 'ditolak', 'revisi'
        
        if not laporan_id or not tabel_laporan or not new_status:
            return jsonify({'success': False, 'message': 'Data tidak lengkap!'})
        
        if tabel_laporan not in ['laporan_kemajuan_awal', 'laporan_kemajuan_bertumbuh']:
            return jsonify({'success': False, 'message': 'Tabel laporan tidak valid!'})
        
        if new_status not in ['disetujui', 'ditolak', 'revisi']:
            return jsonify({'success': False, 'message': 'Status tidak valid!'})
        
        cursor = get_app_functions()['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Update status laporan kemajuan
        cursor.execute(f'''
            UPDATE {tabel_laporan} 
            SET status = %s, tanggal_review = NOW()
            WHERE id = %s
        ''', (new_status, laporan_id))
        
        get_app_functions()['mysql'].connection.commit()
        cursor.close()
        
        return jsonify({'success': True, 'message': f'Status laporan kemajuan berhasil diubah menjadi {new_status}'})
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@admin_bp.route('/update_mahasiswa_data', methods=['POST'])
def update_mahasiswa_data():
    if 'user_type' not in session or session['user_type'] != 'admin':
        return jsonify({'success': False, 'message': 'Anda harus login sebagai admin!'})
    
    if not hasattr(get_app_functions()['mysql'], 'connection') or get_app_functions()['mysql'].connection is None:
        return jsonify({'success': False, 'message': 'Koneksi ke database gagal!'})
    
    try:
        data = request.get_json()
        student_id = data.get('student_id')
        nama_ketua = data.get('nama_ketua')
        nim = data.get('nim')
        email = data.get('email')
        no_telp = data.get('no_telp')
        perguruan_tinggi = data.get('perguruan_tinggi')
        program_studi = data.get('program_studi')
        status = data.get('status')
        password = data.get('password', '').strip()  # Password opsional
        
        if not student_id or not nama_ketua or not nim or not email:
            return jsonify({'success': False, 'message': 'Data tidak lengkap!'})
        
        # Validasi status yang diizinkan
        if status not in ['aktif', 'tolak', 'proses', 'selesai']:
            return jsonify({'success': False, 'message': 'Status tidak valid!'})
        
        # Validasi password jika diisi
        if password and len(password) < 1:
            return jsonify({'success': False, 'message': 'Password minimal 1 karakter!'})
        
        # Set perguruan tinggi default
        perguruan_tinggi = "Universitas Jenderal Achmad Yani Yogyakarta"
        
        cursor = get_app_functions()['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Cek apakah mahasiswa dengan ID tersebut ada
        cursor.execute('SELECT * FROM mahasiswa WHERE id = %s', (student_id,))
        existing_student = cursor.fetchone()
        
        if not existing_student:
            cursor.close()
            return jsonify({'success': False, 'message': 'Data mahasiswa tidak ditemukan!'})
        
        # Cek duplikasi komprehensif (exclude mahasiswa yang sedang diedit)
        is_duplicate, message, table_name = check_comprehensive_duplicates(
            cursor, 
            nama=nama_ketua, 
            nim=nim, 
            email=email,
            exclude_id=student_id,
            exclude_table='mahasiswa'
        )
        if is_duplicate:
            cursor.close()
            return jsonify({'success': False, 'message': message})
        
        # Update data mahasiswa
        if password:
            # Jika ada password baru, update dengan password yang di-hash
            from werkzeug.security import generate_password_hash
            password_hash = generate_password_hash(password)
            cursor.execute('''
                UPDATE mahasiswa 
                SET nama_ketua = %s, nim = %s, email = %s, no_telp = %s, 
                    perguruan_tinggi = %s, program_studi = %s, status = %s, password = %s
                WHERE id = %s
            ''', (nama_ketua, nim, email, no_telp, perguruan_tinggi, program_studi, status, password_hash, student_id))
        else:
            # Jika tidak ada password baru, update tanpa password
            cursor.execute('''
                UPDATE mahasiswa 
                SET nama_ketua = %s, nim = %s, email = %s, no_telp = %s, 
                    perguruan_tinggi = %s, program_studi = %s, status = %s
                WHERE id = %s
            ''', (nama_ketua, nim, email, no_telp, perguruan_tinggi, program_studi, status, student_id))
        
        get_app_functions()['mysql'].connection.commit()
        cursor.close()
        
        if password:
            return jsonify({'success': True, 'message': 'Data mahasiswa dan password berhasil diperbarui!'})
        else:
            return jsonify({'success': True, 'message': 'Data mahasiswa berhasil diperbarui!'})
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error saat update data: {str(e)}'})

@admin_bp.route('/get_mahasiswa_detail/<int:student_id>')
def get_mahasiswa_detail(student_id):
    if 'user_type' not in session or session['user_type'] != 'admin':
        return jsonify({'success': False, 'message': 'Anda harus login sebagai admin!'})
    
    if not hasattr(get_app_functions()['mysql'], 'connection') or get_app_functions()['mysql'].connection is None:
        return jsonify({'success': False, 'message': 'Koneksi ke database gagal!'})
    
    try:
        cursor = get_app_functions()['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Ambil detail mahasiswa
        cursor.execute('''
            SELECT id, perguruan_tinggi, program_studi, nim, nama_ketua, no_telp, email, status
            FROM mahasiswa 
            WHERE id = %s
        ''', (student_id,))
        
        mahasiswa = cursor.fetchone()
        
        # Ambil data program studi untuk dropdown
        cursor.execute('SELECT id, nama_program_studi FROM program_studi ORDER BY nama_program_studi')
        program_studi_list = cursor.fetchall()
        
        cursor.close()
        
        if mahasiswa:
            return jsonify({
                'success': True, 
                'mahasiswa': mahasiswa,
                'program_studi_list': program_studi_list
            })
        else:
            return jsonify({'success': False, 'message': 'Data mahasiswa tidak ditemukan!'})
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error saat mengambil data: {str(e)}'})

@admin_bp.route('/update_proposal', methods=['POST'])
def admin_update_proposal():
    if 'user_type' not in session or session['user_type'] != 'admin':
        return jsonify({'success': False, 'message': 'Anda harus login sebagai admin!'})
    if not hasattr(get_app_functions()['mysql'], 'connection') or get_app_functions()['mysql'].connection is None:
        return jsonify({'success': False, 'message': 'Koneksi ke database gagal!'})
    try:
        data = request.get_json()
        proposal_id = data.get('id')
        fields = [
            'judul_usaha', 'kategori', 'tahapan_usaha', 'merk_produk', 'nib', 'tahun_nib',
            'platform_penjualan', 'dosen_pembimbing', 'nid_dosen', 'program_studi_dosen', 'status'
        ]
        values = [data.get(f) for f in fields]
        cursor = get_app_functions()['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        cursor.execute(f'''
            UPDATE proposal SET
                judul_usaha=%s, kategori=%s, tahapan_usaha=%s, merk_produk=%s, nib=%s, tahun_nib=%s,
                platform_penjualan=%s, dosen_pembimbing=%s, nid_dosen=%s, program_studi_dosen=%s, status=%s
            WHERE id=%s
        ''', (*values, proposal_id))
        get_app_functions()['mysql'].connection.commit()
        cursor.close()
        return jsonify({'success': True, 'message': 'Proposal berhasil diperbarui!'})
    except Exception as e:
        return jsonify({'success': False, 'message': f'Gagal update proposal: {str(e)}'})

@admin_bp.route('/delete_proposal/<int:proposal_id>', methods=['POST'])
def admin_delete_proposal(proposal_id):
    logger.info(f"Delete proposal dipanggil untuk proposal_id: {proposal_id}")
    if 'user_type' not in session or session['user_type'] != 'admin':
        logger.warning(f"Akses ditolak untuk delete proposal: {session.get('user_type', 'None')}")
        return jsonify({'success': False, 'message': 'Anda harus login sebagai admin!'})
    if not hasattr(get_app_functions()['mysql'], 'connection') or get_app_functions()['mysql'].connection is None:
        logger.error("Koneksi database tidak tersedia untuk delete proposal")
        return jsonify({'success': False, 'message': 'Koneksi ke database gagal!'})
    try:
        cursor = get_app_functions()['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)

        # Helper: cek keberadaan tabel di schema aktif
        def table_exists(table_name: str) -> bool:
            try:
                cur = get_app_functions()['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
                cur.execute(
                    """
                    SELECT 1 
                    FROM information_schema.tables 
                    WHERE table_schema = DATABASE() AND table_name = %s
                    LIMIT 1
                    """,
                    (table_name,)
                )
                exists = cur.fetchone() is not None
                cur.close()
                return exists
            except Exception:
                return False
        cursor.execute('SELECT file_path FROM proposal WHERE id = %s', (proposal_id,))
        proposal = cursor.fetchone()
        
        # Fungsi helper untuk menghapus file dan folder
        def hapus_file_dan_folder(file_path, file_type):
            if file_path and os.path.exists(file_path):
                try:
                    # Cek jumlah file di folder sebelum menghapus
                    folder_path = os.path.dirname(file_path)
                    files_in_folder = []
                    if os.path.exists(folder_path):
                        files_in_folder = [f for f in os.listdir(folder_path) if os.path.isfile(os.path.join(folder_path, f))]
                    
                        # Hapus file
                        os.remove(file_path)
                        logger.info(f"✅ File {file_type} berhasil dihapus: {file_path}")
                    
                    # Hapus folder jika hanya tersisa 1 file (file yang baru saja dihapus)
                    if len(files_in_folder) <= 1:
                        if os.path.exists(folder_path) and not os.listdir(folder_path):
                                # Jika folder kosong, hapus folder tersebut
                            os.rmdir(folder_path)
                            logger.info(f"✅ Folder kosong berhasil dihapus: {folder_path}")
                                
                                # Cek juga parent folder (perguruan_tinggi) jika kosong
                            parent_folder = os.path.dirname(folder_path)
                            if os.path.exists(parent_folder) and not os.listdir(parent_folder):
                                os.rmdir(parent_folder)
                                logger.info(f"✅ Parent folder kosong berhasil dihapus: {parent_folder}")
                            else:
                                logger.info(f"ℹ️ Folder tidak dihapus karena masih ada file lain: {folder_path}")
                        else:
                            logger.info(f"ℹ️ Folder tidak dihapus karena masih ada {len(files_in_folder)-1} file lain: {folder_path}")
                            
                except PermissionError as e:
                    logger.error(f"❌ Error permission saat menghapus file {file_path}: {e}")
                except FileNotFoundError as e:
                    logger.warning(f"⚠️ File {file_type} tidak ditemukan: {file_path}")
                except Exception as e:
                    logger.error(f"❌ Error tidak terduga saat menghapus file {file_path}: {e}")
            else:
                if file_path:
                    logger.warning(f"⚠️ File {file_type} tidak ditemukan: {file_path}")
                else:
                    logger.info(f"⚠️ Tidak ada path file {file_type} yang tersimpan")
        
        # Hapus file proposal
        hapus_file_dan_folder(proposal['file_path'], "proposal")
        
        # Hapus file laporan kemajuan tidak lagi menggunakan kolom `laporan_kemajuan_path`.
        # Penghapusan file kemajuan/akhir dilakukan oleh helper hapus_semua_file_proposal di bawah.
        
        # Hapus semua file terkait proposal (komprehensif)
        cursor.execute('SELECT nim, judul_usaha FROM proposal WHERE id = %s', (proposal_id,))
        proposal_ident = cursor.fetchone()
        if proposal_ident:
            get_app_functions()['hapus_semua_file_proposal'](proposal_id, proposal_ident['nim'])
            # Selaraskan perilaku dengan hapus oleh mahasiswa: hapus folder static/uploads/Proposal/<Judul> jika kosong
            try:
                judul_usaha = proposal_ident.get('judul_usaha') if proposal_ident else None
                if judul_usaha:
                    safe_judul = re.sub(r'[^\w\s-]', '', judul_usaha).strip().replace(' ', '_')
                    proposal_folder = os.path.join('static', 'uploads', 'Proposal', safe_judul)
                    if os.path.exists(proposal_folder):
                        remaining_files = [f for f in os.listdir(proposal_folder) if os.path.isfile(os.path.join(proposal_folder, f))]
                        if not remaining_files and not os.listdir(proposal_folder):
                            os.rmdir(proposal_folder)
                            logger.info(f"✅ Folder proposal kosong berhasil dihapus: {proposal_folder}")
                            # Cek parent folder (Proposal) jika kosong
                            parent_proposal_folder = os.path.dirname(proposal_folder)
                            if os.path.exists(parent_proposal_folder) and not os.listdir(parent_proposal_folder):
                                os.rmdir(parent_proposal_folder)
                                logger.info(f"✅ Parent folder Proposal kosong berhasil dihapus: {parent_proposal_folder}")
            except Exception as e:
                logger.warning(f"⚠️ Error saat menghapus folder proposal: {e}")
        
        cursor.execute('DELETE FROM anggaran_awal WHERE id_proposal = %s', (proposal_id,))
        cursor.execute('DELETE FROM anggaran_bertumbuh WHERE id_proposal = %s', (proposal_id,))
        cursor.execute('DELETE FROM laporan_kemajuan_awal WHERE id_proposal = %s', (proposal_id,))
        cursor.execute('DELETE FROM laporan_kemajuan_bertumbuh WHERE id_proposal = %s', (proposal_id,))
        cursor.execute('DELETE FROM laporan_akhir_awal WHERE id_proposal = %s', (proposal_id,))
        cursor.execute('DELETE FROM laporan_akhir_bertumbuh WHERE id_proposal = %s', (proposal_id,))
        
        # Hapus data dari tabel-tabel terkait (dalam urutan yang benar untuk menghindari foreign key constraint)
        cursor.execute('DELETE FROM alat_produksi WHERE proposal_id = %s', (proposal_id,))
        cursor.execute('DELETE FROM bahan_baku WHERE proposal_id = %s', (proposal_id,))
        cursor.execute('DELETE FROM biaya_operasional WHERE proposal_id = %s', (proposal_id,))
        cursor.execute('DELETE FROM biaya_non_operasional WHERE proposal_id = %s', (proposal_id,))
        cursor.execute('DELETE FROM penjualan WHERE proposal_id = %s', (proposal_id,))
        cursor.execute('DELETE FROM laba_rugi WHERE proposal_id = %s', (proposal_id,))
        cursor.execute('DELETE FROM arus_kas WHERE proposal_id = %s', (proposal_id,))

        # Hapus data bimbingan terkait proposal
        cursor.execute('DELETE FROM bimbingan WHERE proposal_id = %s', (proposal_id,))
        
        # Hapus data penilaian terkait proposal (dalam urutan yang benar untuk menghindari foreign key constraint)
        
        # 1. Hapus detail penilaian proposal (berdasarkan id_proposal, karena kolom relasi adalah id_proposal/id_reviewer)
        cursor.execute('DELETE FROM detail_penilaian_proposal WHERE id_proposal = %s', (proposal_id,))
        
        # 2. Hapus catatan penilaian proposal (opsional jika tabel ada)
        if table_exists('catatan_penilaian_proposal'):
            cursor.execute('''
                DELETE cpp FROM catatan_penilaian_proposal cpp 
                INNER JOIN penilaian_proposal pp ON cpp.id_penilaian_proposal = pp.id 
                WHERE pp.id_proposal = %s
            ''', (proposal_id,))
        
        # 3. Hapus rekomendasi penilaian proposal (opsional jika tabel ada)
        if table_exists('rekomendasi_penilaian_proposal'):
            cursor.execute('''
                DELETE rpp FROM rekomendasi_penilaian_proposal rpp 
                INNER JOIN penilaian_proposal pp ON rpp.id_penilaian_proposal = pp.id 
                WHERE pp.id_proposal = %s
            ''', (proposal_id,))
        
        # 4. Hapus penilaian proposal
        cursor.execute('DELETE FROM penilaian_proposal WHERE id_proposal = %s', (proposal_id,))
        
        # 5. Hapus detail penilaian laporan kemajuan, lalu header
        cursor.execute('''
            DELETE dplk FROM detail_penilaian_laporan_kemajuan dplk
            INNER JOIN penilaian_laporan_kemajuan plk ON dplk.id_penilaian_laporan_kemajuan = plk.id
            WHERE plk.id_proposal = %s
        ''', (proposal_id,))
        cursor.execute('DELETE FROM penilaian_laporan_kemajuan WHERE id_proposal = %s', (proposal_id,))

        # 6. Hapus detail penilaian laporan akhir, lalu header
        cursor.execute('''
            DELETE dpla FROM detail_penilaian_laporan_akhir dpla
            INNER JOIN penilaian_laporan_akhir pla ON dpla.id_penilaian_laporan_akhir = pla.id
            WHERE pla.id_proposal = %s
        ''', (proposal_id,))
        cursor.execute('DELETE FROM penilaian_laporan_akhir WHERE id_proposal = %s', (proposal_id,))

        # 7. Hapus detail penilaian mahasiswa (oleh pembimbing), lalu header
        cursor.execute('''
            DELETE dpm FROM detail_penilaian_mahasiswa dpm
            INNER JOIN penilaian_mahasiswa pm ON dpm.id_penilaian_mahasiswa = pm.id
            WHERE pm.id_proposal = %s
        ''', (proposal_id,))
        cursor.execute('DELETE FROM penilaian_mahasiswa WHERE id_proposal = %s', (proposal_id,))
        
        # Hapus produk_bahan_baku yang terkait dengan produksi dari proposal ini
        cursor.execute('''
            DELETE pbb FROM produk_bahan_baku pbb 
            INNER JOIN produksi p ON pbb.produksi_id = p.id 
            WHERE p.proposal_id = %s
        ''', (proposal_id,))
        
        # Hapus produksi (setelah produk_bahan_baku dihapus)
        cursor.execute('DELETE FROM produksi WHERE proposal_id = %s', (proposal_id,))
        
        # Hapus relasi reviewer dan data file_laporan_* terkait proposal
        cursor.execute('DELETE FROM proposal_reviewer WHERE id_proposal = %s', (proposal_id,))
        cursor.execute('DELETE FROM file_laporan_kemajuan WHERE id_proposal = %s', (proposal_id,))
        cursor.execute('DELETE FROM file_laporan_akhir WHERE id_proposal = %s', (proposal_id,))

        cursor.execute('DELETE FROM anggota_tim WHERE id_proposal = %s', (proposal_id,))
        cursor.execute('DELETE FROM proposal WHERE id = %s', (proposal_id,))
        get_app_functions()['mysql'].connection.commit()
        cursor.close()
        logger.info(f"Proposal {proposal_id} berhasil dihapus")
        return jsonify({'success': True, 'message': 'Proposal berhasil dihapus!'})
    except Exception as e:
        logger.error(f"Gagal menghapus proposal {proposal_id}: {str(e)}")
        logger.error(traceback.format_exc())
        return jsonify({'success': False, 'message': f'Gagal menghapus proposal: {str(e)}'})

@admin_bp.route('/get_pembimbing_list')
def get_pembimbing_list():
    if 'user_type' not in session or session['user_type'] != 'admin':
        return jsonify({'success': False, 'message': 'Anda harus login sebagai admin!'})
    
    if not hasattr(get_app_functions()['mysql'], 'connection') or get_app_functions()['mysql'].connection is None:
        return jsonify({'success': False, 'message': 'Koneksi ke database gagal!'})
    
    try:
        cursor = get_app_functions()['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Ambil daftar pembimbing dengan informasi kuota
        cursor.execute('''
            SELECT p.nama, p.nip, p.program_studi, p.kuota_mahasiswa,
                   COUNT(DISTINCT CASE WHEN pr.status != 'selesai' THEN pr.nim END) as jumlah_mahasiswa_bimbing
            FROM pembimbing p
            LEFT JOIN proposal pr ON p.nama = pr.dosen_pembimbing
            WHERE p.status = 'aktif'
            GROUP BY p.id, p.nama, p.nip, p.program_studi, p.kuota_mahasiswa
            ORDER BY p.nama
        ''')
        
        pembimbing_list = cursor.fetchall()
        
        # Tambahkan informasi ketersediaan kuota
        for pembimbing in pembimbing_list:
            pembimbing['sisa_kuota'] = pembimbing['kuota_mahasiswa'] - pembimbing['jumlah_mahasiswa_bimbing']
            pembimbing['status_kuota'] = 'tersedia' if pembimbing['sisa_kuota'] > 0 else 'penuh'
        
        cursor.close()
        
        return jsonify({'success': True, 'pembimbing': pembimbing_list})
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error saat mengambil data pembimbing: {str(e)}'})

@admin_bp.route('/anggaran_awal/<int:proposal_id>')
def admin_anggaran_awal(proposal_id):
    if 'user_type' not in session or session['user_type'] != 'admin':
        flash('Anda harus login sebagai admin!', 'danger')
        return redirect(url_for('index'))
    if not hasattr(get_app_functions()['mysql'], 'connection') or get_app_functions()['mysql'].connection is None:
        flash('Koneksi ke database gagal. Cek konfigurasi database!', 'danger')
        return render_template('admin/pengajuan anggaran awal.html', anggaran_data=[], proposal_id=proposal_id)
    try:
        cursor = get_app_functions()['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        cursor.execute('''
            SELECT id, kegiatan_utama, kegiatan, penanggung_jawab, target_capaian, 
                   nama_barang, kuantitas, satuan, harga_satuan, jumlah, keterangan, status,
                   status_reviewer, nilai_bantuan
            FROM anggaran_awal 
            WHERE id_proposal = %s 
            ORDER BY kegiatan_utama, kegiatan, nama_barang
        ''', (proposal_id,))
        anggaran_data = cursor.fetchall()
        
        cursor.execute('SELECT judul_usaha, tahapan_usaha FROM proposal WHERE id = %s', (proposal_id,))
        proposal_info = cursor.fetchone()
        
        # Urutkan data sesuai urutan kegiatan utama (logika dinamis berdasarkan tahapan usaha)
        tahapan_usaha = proposal_info.get('tahapan_usaha', '').lower() if proposal_info else ''
        if 'bertumbuh' in tahapan_usaha:
            urutan_kegiatan = [
                "Pengembangan Pasar dan Saluran Distribusi",
                "Pengembangan Produk/Riset",
                "Produksi",
                "Legalitas, perizinan, sertifikasi, dan standarisasi",
                "Belanja ATK dan Penunjang"
            ]
        else:
            urutan_kegiatan = [
                "Pengembangan Produk/Riset",
                "Produksi",
                "Legalitas, perizinan, sertifikasi, dan standarisasi",
                "Belanja ATK dan Penunjang"
            ]
        
        anggaran_data = sorted(
            anggaran_data,
            key=lambda x: urutan_kegiatan.index(x['kegiatan_utama']) if x['kegiatan_utama'] in urutan_kegiatan else 99
        )
        cursor.close()
        grouped_data = group_anggaran_data(anggaran_data)
        anggaran_data_flat = flatten_anggaran_data(anggaran_data)
        # Ambil batas min/maks untuk badge seperti mahasiswa
        try:
            cur2 = get_app_functions()['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
            cur2.execute('SELECT * FROM pengaturan_anggaran ORDER BY id DESC LIMIT 1')
            batas_row = cur2.fetchone() or {}
            batas_min_awal = int(batas_row.get('min_total_awal') or 0)
            batas_max_awal = int(batas_row.get('max_total_awal') or 0)
            cur2.close()
        except Exception:
            batas_min_awal = 0
            batas_max_awal = 0
        return render_template('admin/pengajuan anggaran awal.html',
                             anggaran_data=anggaran_data, 
                             grouped_data=grouped_data,
                             anggaran_data_flat=anggaran_data_flat,
                             proposal_id=proposal_id,
                             proposal_info=proposal_info,
                             batas_min_awal=batas_min_awal,
                             batas_max_awal=batas_max_awal)
    except Exception as e:
        flash(f'Error saat mengambil data anggaran: {str(e)}', 'danger')
        return render_template('admin/pengajuan anggaran awal.html', anggaran_data=[], proposal_id=proposal_id)

@admin_bp.route('/anggaran_bertumbuh/<int:proposal_id>')
def admin_anggaran_bertumbuh(proposal_id):
    if 'user_type' not in session or session['user_type'] != 'admin':
        flash('Anda harus login sebagai admin!', 'danger')
        return redirect(url_for('index'))
    if not hasattr(get_app_functions()['mysql'], 'connection') or get_app_functions()['mysql'].connection is None:
        flash('Koneksi ke database gagal. Cek konfigurasi database!', 'danger')
        return render_template('admin/pengajuan anggaran bertumbuh.html', anggaran_data=[], proposal_id=proposal_id)
    try:
        cursor = get_app_functions()['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        cursor.execute('''
            SELECT id, kegiatan_utama, kegiatan, penanggung_jawab, target_capaian, 
                   nama_barang, kuantitas, satuan, harga_satuan, jumlah, keterangan, status,
                   status_reviewer, nilai_bantuan
            FROM anggaran_bertumbuh 
            WHERE id_proposal = %s 
            ORDER BY kegiatan_utama, kegiatan, nama_barang
        ''', (proposal_id,))
        anggaran_data = cursor.fetchall()
        
        cursor.execute('SELECT judul_usaha, tahapan_usaha FROM proposal WHERE id = %s', (proposal_id,))
        proposal_info = cursor.fetchone()
        
        # Urutkan data sesuai urutan kegiatan utama dengan normalisasi teks
        tahapan_usaha = proposal_info.get('tahapan_usaha', '').lower() if proposal_info else ''
        if 'bertumbuh' in tahapan_usaha:
            urutan_kegiatan = [
                "Pengembangan Pasar dan Saluran Distribusi",
                "Pengembangan Produk/Riset",
                "Produksi",
                "Legalitas, perizinan, sertifikasi, dan standarisasi",
                "Belanja ATK dan Penunjang"
            ]
        else:
            urutan_kegiatan = [
                "Pengembangan Produk/Riset",
                "Produksi",
                "Legalitas, perizinan, sertifikasi, dan standarisasi",
                "Belanja ATK dan Penunjang"
            ]
        
        # Buat mapping untuk case-insensitive matching dan normalisasi teks
        urutan_kegiatan_lower = [k.strip().lower() for k in urutan_kegiatan]
        
        # Fungsi helper untuk normalisasi teks kegiatan utama
        def normalize_kegiatan_utama(text):
            if not text:
                return ""
            # Normalisasi: lowercase, strip whitespace, replace multiple spaces
            normalized = text.strip().lower()
            # Handle berbagai variasi teks
            replacements = {
                "legalitas, perijinan, sertifikasi, pengujian produk, dan standarisasi": "legalitas, perizinan, sertifikasi, dan standarisasi",
                "legalitas, perijinan, sertifikasi, dan standarisasi": "legalitas, perizinan, sertifikasi, dan standarisasi",
                "legalitas perijinan sertifikasi pengujian produk dan standarisasi": "legalitas, perizinan, sertifikasi, dan standarisasi",
                "legalitas perijinan sertifikasi dan standarisasi": "legalitas, perizinan, sertifikasi, dan standarisasi"
            }
            for old, new in replacements.items():
                if old in normalized:
                    normalized = normalized.replace(old, new)
            return normalized
        
        # Sort data berdasarkan urutan kegiatan utama dengan normalisasi
        anggaran_data = sorted(
            anggaran_data,
            key=lambda x: urutan_kegiatan_lower.index(normalize_kegiatan_utama(x['kegiatan_utama'])) if x['kegiatan_utama'] and normalize_kegiatan_utama(x['kegiatan_utama']) in urutan_kegiatan_lower else 99
        )
        cursor.close()
        grouped_data = group_anggaran_data(anggaran_data)
        anggaran_data_flat = flatten_anggaran_data(anggaran_data)
        # Ambil batas min/maks untuk badge seperti mahasiswa
        try:
            cur2 = get_app_functions()['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
            cur2.execute('SELECT * FROM pengaturan_anggaran ORDER BY id DESC LIMIT 1')
            batas_row = cur2.fetchone() or {}
            batas_min_bertumbuh = int(batas_row.get('min_total_bertumbuh') or 0)
            batas_max_bertumbuh = int(batas_row.get('max_total_bertumbuh') or 0)
            cur2.close()
        except Exception:
            batas_min_bertumbuh = 0
            batas_max_bertumbuh = 0
        return render_template('admin/pengajuan anggaran bertumbuh.html', 
                             anggaran_data=anggaran_data, 
                             grouped_data=grouped_data,
                             anggaran_data_flat=anggaran_data_flat,
                             proposal_id=proposal_id,
                             proposal_info=proposal_info,
                             batas_min_bertumbuh=batas_min_bertumbuh,
                             batas_max_bertumbuh=batas_max_bertumbuh)
    except Exception as e:
        flash(f'Error saat mengambil data anggaran: {str(e)}', 'danger')
        return render_template('admin/pengajuan anggaran bertumbuh.html', anggaran_data=[], proposal_id=proposal_id)












@admin_bp.route('/pembimbing')
def admin_pembimbing():
    if 'user_type' not in session or session['user_type'] != 'admin':
        flash('Anda harus login sebagai admin!', 'danger')
        return redirect(url_for('index'))
    
    if not hasattr(get_app_functions()['mysql'], 'connection') or get_app_functions()['mysql'].connection is None:
        flash('Koneksi ke database gagal. Cek konfigurasi database!', 'danger')
        return render_template('admin/pembimbing.html', pembimbing_list=[])
    
    try:
        cursor = get_app_functions()['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Ambil data pembimbing beserta status dan informasi kuota
        cursor.execute('''
            SELECT p.id, p.nama, p.nip, p.program_studi, p.status, p.kuota_mahasiswa, p.tanggal_dibuat,
                   COUNT(DISTINCT CASE WHEN pr.status != 'selesai' THEN pr.nim END) as jumlah_mahasiswa_bimbing
            FROM pembimbing p
            LEFT JOIN proposal pr ON p.nama = pr.dosen_pembimbing
            GROUP BY p.id, p.nama, p.nip, p.program_studi, p.status, p.kuota_mahasiswa, p.tanggal_dibuat
            ORDER BY p.nama ASC
        ''')
        
        pembimbing_list = cursor.fetchall()
        
        # Tambahkan informasi ketersediaan kuota
        for pembimbing in pembimbing_list:
            pembimbing['sisa_kuota'] = pembimbing['kuota_mahasiswa'] - pembimbing['jumlah_mahasiswa_bimbing']
            pembimbing['status_kuota'] = 'tersedia' if pembimbing['sisa_kuota'] > 0 else 'penuh'
        
        # Ambil data program studi untuk dropdown
        cursor.execute('SELECT id, nama_program_studi FROM program_studi ORDER BY nama_program_studi')
        program_studi_list = cursor.fetchall()
        
        cursor.close()
        
        return render_template('admin/pembimbing.html', pembimbing_list=pembimbing_list, program_studi_list=program_studi_list)
        
    except Exception as e:
        flash(f'Error saat mengambil data pembimbing: {str(e)}', 'danger')
        return render_template('admin/pembimbing.html', pembimbing_list=[])

@admin_bp.route('/add_pembimbing', methods=['POST'])
def add_pembimbing():
    if 'user_type' not in session or session['user_type'] != 'admin':
        return jsonify({'success': False, 'message': 'Anda harus login sebagai admin!'})
    
    if not hasattr(get_app_functions()['mysql'], 'connection') or get_app_functions()['mysql'].connection is None:
        return jsonify({'success': False, 'message': 'Koneksi ke database gagal!'})
    
    try:
        data = request.get_json()
        nama = data.get('nama', '').strip()
        nip = data.get('nip', '').strip()
        program_studi = data.get('program_studi', '').strip()
        password = data.get('password', '').strip()
        # Gunakan kuota dari request atau default 5
        kuota_mahasiswa = data.get('kuota_mahasiswa', 5)
        
        if not nama or not nip or not program_studi or not password:
            return jsonify({'success': False, 'message': 'Semua field harus diisi!'})
        
        # Validasi password minimal 1 karakter
        if len(password) < 1:
            return jsonify({'success': False, 'message': 'Password minimal 1 karakter!'})
        
        # Validasi kuota
        try:
            kuota_mahasiswa = int(kuota_mahasiswa)
            if kuota_mahasiswa < 1 or kuota_mahasiswa > 50:
                return jsonify({'success': False, 'message': 'Kuota mahasiswa harus antara 1-50!'})
        except (ValueError, TypeError):
            return jsonify({'success': False, 'message': 'Kuota mahasiswa harus berupa angka!'})
        
        cursor = get_app_functions()['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Cek duplikasi komprehensif
        is_duplicate, message, table_name = check_comprehensive_duplicates(
            cursor, 
            nama=nama, 
            nip=nip
        )
        if is_duplicate:
            cursor.close()
            return jsonify({'success': False, 'message': message})
        
        # Hash password
        password_hash = generate_password_hash(password)
        
        # Insert pembimbing baru dengan kuota
        cursor.execute('''
            INSERT INTO pembimbing (nama, nip, program_studi, password, status, kuota_mahasiswa)
            VALUES (%s, %s, %s, %s, %s, %s)
        ''', (nama, nip, program_studi, password_hash, 'nonaktif', kuota_mahasiswa))
        
        get_app_functions()['mysql'].connection.commit()
        cursor.close()
        
        return jsonify({'success': True, 'message': 'Data pembimbing berhasil ditambahkan!'})
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error saat menambah data: {str(e)}'})

@admin_bp.route('/update_pembimbing_data', methods=['POST'])
def update_pembimbing_data():
    if 'user_type' not in session or session['user_type'] != 'admin':
        return jsonify({'success': False, 'message': 'Anda harus login sebagai admin!'})
    
    if not hasattr(get_app_functions()['mysql'], 'connection') or get_app_functions()['mysql'].connection is None:
        return jsonify({'success': False, 'message': 'Koneksi ke database gagal!'})
    
    try:
        data = request.get_json()
        pembimbing_id = data.get('pembimbing_id')
        nama = data.get('nama', '').strip()
        nip = data.get('nip', '').strip()
        program_studi = data.get('program_studi', '').strip()
        password = data.get('password', '').strip()
        kuota_mahasiswa = data.get('kuota_mahasiswa', 5)
        
        if not pembimbing_id or not nama or not nip or not program_studi:
            return jsonify({'success': False, 'message': 'Data tidak lengkap!'})
        
        # Validasi kuota
        try:
            kuota_mahasiswa = int(kuota_mahasiswa)
            if kuota_mahasiswa < 1 or kuota_mahasiswa > 50:
                return jsonify({'success': False, 'message': 'Kuota mahasiswa harus antara 1-50!'})
        except (ValueError, TypeError):
            return jsonify({'success': False, 'message': 'Kuota mahasiswa harus berupa angka!'})
        
        # Validasi password jika diisi
        if password and len(password) < 1:
            return jsonify({'success': False, 'message': 'Password minimal 1 karakter!'})
        
        cursor = get_app_functions()['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Cek apakah pembimbing dengan ID tersebut ada
        cursor.execute('SELECT * FROM pembimbing WHERE id = %s', (pembimbing_id,))
        existing_pembimbing = cursor.fetchone()
        
        if not existing_pembimbing:
            cursor.close()
            return jsonify({'success': False, 'message': 'Data pembimbing tidak ditemukan!'})
        
        # Cek duplikasi komprehensif (exclude pembimbing yang sedang diedit)
        is_duplicate, message, table_name = check_comprehensive_duplicates(
            cursor, 
            nama=nama, 
            nip=nip,
            exclude_id=pembimbing_id,
            exclude_table='pembimbing'
        )
        if is_duplicate:
            cursor.close()
            return jsonify({'success': False, 'message': message})
        
        # Update data pembimbing
        if password:
            # Jika ada password baru, update dengan password yang di-hash
            password_hash = generate_password_hash(password)
            cursor.execute('''
                UPDATE pembimbing 
                SET nama = %s, nip = %s, program_studi = %s, password = %s, kuota_mahasiswa = %s
                WHERE id = %s
            ''', (nama, nip, program_studi, password_hash, kuota_mahasiswa, pembimbing_id))
        else:
            # Jika tidak ada password baru, update tanpa password
            cursor.execute('''
                UPDATE pembimbing 
                SET nama = %s, nip = %s, program_studi = %s, kuota_mahasiswa = %s
                WHERE id = %s
            ''', (nama, nip, program_studi, kuota_mahasiswa, pembimbing_id))
        
        get_app_functions()['mysql'].connection.commit()
        cursor.close()
        
        if password:
            return jsonify({'success': True, 'message': 'Data pembimbing dan password berhasil diperbarui!'})
        else:
            return jsonify({'success': True, 'message': 'Data pembimbing berhasil diperbarui!'})
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error saat update data: {str(e)}'})

@admin_bp.route('/delete_pembimbing', methods=['POST'])
def delete_pembimbing():
    if 'user_type' not in session or session['user_type'] != 'admin':
        return jsonify({'success': False, 'message': 'Anda harus login sebagai admin!'})
    
    if not hasattr(get_app_functions()['mysql'], 'connection') or get_app_functions()['mysql'].connection is None:
        return jsonify({'success': False, 'message': 'Koneksi ke database gagal!'})
    
    try:
        data = request.get_json()
        pembimbing_id = data.get('pembimbing_id')
        
        if not pembimbing_id:
            return jsonify({'success': False, 'message': 'ID pembimbing tidak ditemukan!'})
        
        cursor = get_app_functions()['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Cek apakah pembimbing dengan ID tersebut ada
        cursor.execute('SELECT * FROM pembimbing WHERE id = %s', (pembimbing_id,))
        existing_pembimbing = cursor.fetchone()
        
        if not existing_pembimbing:
            cursor.close()
            return jsonify({'success': False, 'message': 'Data pembimbing tidak ditemukan!'})
        
        nama_pembimbing = existing_pembimbing['nama']
        
        # Update semua proposal yang menggunakan pembimbing ini
        cursor.execute('''
            UPDATE proposal 
            SET dosen_pembimbing = NULL, 
                nid_dosen = NULL, 
                program_studi_dosen = NULL, 
                status = 'draf', 
                status_admin = 'belum_ditinjau' 
            WHERE dosen_pembimbing = %s
        ''', (nama_pembimbing,))
        
        # Hapus semua session pembimbing terkait
        cursor.execute('DELETE FROM session_pembimbing WHERE pembimbing_id = %s', (pembimbing_id,))
        
        # Hapus semua log aktivitas pembimbing terkait
        cursor.execute('DELETE FROM log_aktivitas_pembimbing WHERE pembimbing_id = %s', (pembimbing_id,))
        
        # Hapus pembimbing
        cursor.execute('DELETE FROM pembimbing WHERE id = %s', (pembimbing_id,))
        
        get_app_functions()['mysql'].connection.commit()
        cursor.close()
        
        return jsonify({'success': True, 'message': 'Data pembimbing, session, log aktivitas, dan proposal terkait sudah dihapus/diupdate!'})
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error saat menghapus data: {str(e)}'})


@admin_bp.route('/set_status_pembimbing', methods=['POST'])
def set_status_pembimbing():
    if 'user_type' not in session or session['user_type'] != 'admin':
        return jsonify({'success': False, 'message': 'Anda harus login sebagai admin!'})
    if not hasattr(get_app_functions()['mysql'], 'connection') or get_app_functions()['mysql'].connection is None:
        return jsonify({'success': False, 'message': 'Koneksi ke database gagal!'})
    try:
        data = request.get_json()
        pembimbing_id = data.get('pembimbing_id')
        status = data.get('status')
        if not pembimbing_id or status not in ['aktif', 'nonaktif']:
            return jsonify({'success': False, 'message': 'Data tidak lengkap atau status tidak valid!'})
        cursor = get_app_functions()['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        cursor.execute('SELECT * FROM pembimbing WHERE id = %s', (pembimbing_id,))
        pembimbing = cursor.fetchone()
        if not pembimbing:
            cursor.close()
            return jsonify({'success': False, 'message': 'Data pembimbing tidak ditemukan!'})
        cursor.execute('UPDATE pembimbing SET status = %s WHERE id = %s', (status, pembimbing_id))
        # Jika dinonaktifkan, update semua proposal yang memakai pembimbing ini
        if status == 'nonaktif':
            nama_pembimbing = pembimbing['nama']
            cursor.execute('''
                UPDATE proposal
                SET dosen_pembimbing = NULL,
                    nid_dosen = NULL,
                    program_studi_dosen = NULL,
                    status = 'draf',
                    status_admin = 'belum_ditinjau'
                WHERE dosen_pembimbing = %s
            ''', (nama_pembimbing,))
        get_app_functions()['mysql'].connection.commit()
        cursor.close()
        return jsonify({'success': True, 'message': f'Status pembimbing berhasil diubah menjadi {status}.'})
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})




@admin_bp.route('/delete_mahasiswa', methods=['POST'])
def delete_mahasiswa():
    if 'user_type' not in session or session['user_type'] != 'admin':
        return jsonify({'success': False, 'message': 'Anda harus login sebagai admin!'})
    
    if not hasattr(get_app_functions()['mysql'], 'connection') or get_app_functions()['mysql'].connection is None:
        return jsonify({'success': False, 'message': 'Koneksi ke database gagal!'})
    
    try:
        data = request.get_json()
        mahasiswa_id = data.get('mahasiswa_id')
        
        if not mahasiswa_id:
            return jsonify({'success': False, 'message': 'ID mahasiswa tidak ditemukan!'})
        
        cursor = get_app_functions()['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Cek apakah mahasiswa dengan ID tersebut ada
        cursor.execute('SELECT * FROM mahasiswa WHERE id = %s', (mahasiswa_id,))
        existing_mahasiswa = cursor.fetchone()
        
        if not existing_mahasiswa:
            cursor.close()
            return jsonify({'success': False, 'message': 'Data mahasiswa tidak ditemukan!'})
        
        nim = existing_mahasiswa['nim']
        nama_ketua = existing_mahasiswa['nama_ketua']
        perguruan_tinggi = existing_mahasiswa['perguruan_tinggi']
        
        # 1. Ambil semua proposal mahasiswa untuk penghapusan file (tanpa kolom yang tidak ada)
        cursor.execute('SELECT id, file_path FROM proposal WHERE nim = %s', (nim,))
        proposals = cursor.fetchall()
        
        # 2. Hapus file proposal dan laporan kemajuan dari server
        for proposal in proposals:
            file_path = proposal['file_path']
            
            # Fungsi helper untuk menghapus file dan folder
            def hapus_file_dan_folder(file_path, file_type):
                if file_path and os.path.exists(file_path):
                    try:
                        # Cek jumlah file di folder sebelum menghapus
                        folder_path = os.path.dirname(file_path)
                        files_in_folder = []
                        if os.path.exists(folder_path):
                            files_in_folder = [f for f in os.listdir(folder_path) if os.path.isfile(os.path.join(folder_path, f))]
                        
                            # Hapus file
                        os.remove(file_path)
                        print(f"✅ File {file_type} berhasil dihapus: {file_path}")
                        
                        # Hapus folder jika hanya tersisa 1 file (file yang baru saja dihapus)
                        if len(files_in_folder) <= 1:
                            if os.path.exists(folder_path) and not os.listdir(folder_path):
                                # Jika folder kosong, hapus folder tersebut
                                os.rmdir(folder_path)
                                print(f"✅ Folder kosong berhasil dihapus: {folder_path}")
                                
                                # Cek juga parent folder (perguruan_tinggi) jika kosong
                                parent_folder = os.path.dirname(folder_path)
                                if os.path.exists(parent_folder) and not os.listdir(parent_folder):
                                    os.rmdir(parent_folder)
                                    print(f"✅ Parent folder kosong berhasil dihapus: {parent_folder}")
                            else:
                                print(f"ℹ️ Folder tidak dihapus karena masih ada file lain: {folder_path}")
                        else:
                            print(f"ℹ️ Folder tidak dihapus karena masih ada {len(files_in_folder)-1} file lain: {folder_path}")
                            
                    except PermissionError as e:
                        print(f"❌ Error permission saat menghapus file {file_path}: {e}")
                    except FileNotFoundError as e:
                            print(f"⚠️ File {file_type} tidak ditemukan: {file_path}")
                    except Exception as e:
                        print(f"❌ Error tidak terduga saat menghapus file {file_path}: {e}")
                else:
                    if file_path:
                            print(f"⚠️ File {file_type} tidak ditemukan: {file_path}")
                    else:
                            print(f"⚠️ Tidak ada path file {file_type} yang tersimpan")
            
            # Hapus file proposal
            hapus_file_dan_folder(file_path, "proposal")

            # Hapus file-file laporan kemajuan dan laporan akhir berdasarkan tabel file_*
            try:
                # Laporan kemajuan
                cursor2 = get_app_functions()['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
                cursor2.execute('SELECT nama_file, file_path FROM file_laporan_kemajuan WHERE id_proposal = %s', (proposal['id'],))
                kemajuan_files = cursor2.fetchall()
                for f in kemajuan_files:
                    hapus_file_dan_folder(f.get('file_path'), f"laporan kemajuan ({f.get('nama_file')})")
                cursor2.close()

                # Laporan akhir
                cursor3 = get_app_functions()['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
                cursor3.execute('SELECT nama_file, file_path FROM file_laporan_akhir WHERE id_proposal = %s', (proposal['id'],))
                akhir_files = cursor3.fetchall()
                for f in akhir_files:
                    hapus_file_dan_folder(f.get('file_path'), f"laporan akhir ({f.get('nama_file')})")
                cursor3.close()
            except Exception as e:
                print(f"⚠️ Gagal menghapus file dari tabel file_laporan_* untuk proposal {proposal['id']}: {e}")
        
        # Hapus semua file terkait proposal untuk semua proposal mahasiswa ini
        cursor.execute('SELECT id FROM proposal WHERE nim = %s', (nim,))
        proposals = cursor.fetchall()
        
        for proposal in proposals:
            get_app_functions()['hapus_semua_file_proposal'](proposal['id'], nim)
        
        # 3. Hapus data dari database (dalam urutan yang benar untuk menghindari foreign key constraint)
        
        # Hapus anggaran awal terkait
        cursor.execute('''
            DELETE aa FROM anggaran_awal aa 
            INNER JOIN proposal p ON aa.id_proposal = p.id 
            WHERE p.nim = %s
        ''', (nim,))
        anggaran_awal_deleted = cursor.rowcount
        
        # Hapus anggaran bertumbuh terkait
        cursor.execute('''
            DELETE ab FROM anggaran_bertumbuh ab 
            INNER JOIN proposal p ON ab.id_proposal = p.id 
            WHERE p.nim = %s
        ''', (nim,))
        anggaran_bertumbuh_deleted = cursor.rowcount
        
        # Hapus laporan kemajuan awal terkait
        cursor.execute('''
            DELETE lka FROM laporan_kemajuan_awal lka 
            INNER JOIN proposal p ON lka.id_proposal = p.id 
            WHERE p.nim = %s
        ''', (nim,))
        laporan_kemajuan_awal_deleted = cursor.rowcount
        
        # Hapus laporan kemajuan bertumbuh terkait
        cursor.execute('''
            DELETE lkb FROM laporan_kemajuan_bertumbuh lkb 
            INNER JOIN proposal p ON lkb.id_proposal = p.id 
            WHERE p.nim = %s
        ''', (nim,))
        laporan_kemajuan_bertumbuh_deleted = cursor.rowcount
        
        # Hapus laporan akhir awal terkait
        cursor.execute('''
            DELETE laa FROM laporan_akhir_awal laa 
            INNER JOIN proposal p ON laa.id_proposal = p.id 
            WHERE p.nim = %s
        ''', (nim,))
        laporan_akhir_awal_deleted = cursor.rowcount
        
        # Hapus laporan akhir bertumbuh terkait
        cursor.execute('''
            DELETE lab FROM laporan_akhir_bertumbuh lab 
            INNER JOIN proposal p ON lab.id_proposal = p.id 
            WHERE p.nim = %s
        ''', (nim,))
        laporan_akhir_bertumbuh_deleted = cursor.rowcount
        
        # Hapus data dari tabel-tabel terkait (dalam urutan yang benar untuk menghindari foreign key constraint)
        cursor.execute('''
            DELETE ap FROM alat_produksi ap 
            INNER JOIN proposal p ON ap.proposal_id = p.id 
            WHERE p.nim = %s
        ''', (nim,))
        alat_produksi_deleted = cursor.rowcount
        
        cursor.execute('''
            DELETE bb FROM bahan_baku bb 
            INNER JOIN proposal p ON bb.proposal_id = p.id 
            WHERE p.nim = %s
        ''', (nim,))
        bahan_baku_deleted = cursor.rowcount
        
        cursor.execute('''
            DELETE bo FROM biaya_operasional bo 
            INNER JOIN proposal p ON bo.proposal_id = p.id 
            WHERE p.nim = %s
        ''', (nim,))
        biaya_operasional_deleted = cursor.rowcount
        
        cursor.execute('''
            DELETE bno FROM biaya_non_operasional bno 
            INNER JOIN proposal p ON bno.proposal_id = p.id 
            WHERE p.nim = %s
        ''', (nim,))
        biaya_non_operasional_deleted = cursor.rowcount
        
        cursor.execute('''
            DELETE pj FROM penjualan pj 
            INNER JOIN proposal p ON pj.proposal_id = p.id 
            WHERE p.nim = %s
        ''', (nim,))
        penjualan_deleted = cursor.rowcount
        
        cursor.execute('''
            DELETE lr FROM laba_rugi lr 
            INNER JOIN proposal p ON lr.proposal_id = p.id 
            WHERE p.nim = %s
        ''', (nim,))
        laba_rugi_deleted = cursor.rowcount
        
        cursor.execute('''
            DELETE ak FROM arus_kas ak 
            INNER JOIN proposal p ON ak.proposal_id = p.id 
            WHERE p.nim = %s
        ''', (nim,))
        arus_kas_deleted = cursor.rowcount
        
        # Hapus data bimbingan terkait
        cursor.execute('DELETE FROM bimbingan WHERE nim = %s', (nim,))
        bimbingan_deleted = cursor.rowcount
        
        # Hapus data penilaian terkait (dalam urutan yang benar untuk menghindari foreign key constraint)
        
        # 1. Hapus detail penilaian proposal (berdasarkan id_proposal langsung)
        cursor.execute('''
            DELETE dpp FROM detail_penilaian_proposal dpp 
            INNER JOIN proposal p ON dpp.id_proposal = p.id 
            WHERE p.nim = %s
        ''', (nim,))
        detail_penilaian_proposal_deleted = cursor.rowcount
        
        # 2. Hapus catatan penilaian proposal (opsional, hanya jika tabel ada)
        def table_exists(table_name: str) -> bool:
            try:
                cur = get_app_functions()['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
                cur.execute(
                    """
                    SELECT 1 
                    FROM information_schema.tables 
                    WHERE table_schema = DATABASE() AND table_name = %s
                    LIMIT 1
                    """,
                    (table_name,)
                )
                exists = cur.fetchone() is not None
                cur.close()
                return exists
            except Exception:
                return False

        if table_exists('catatan_penilaian_proposal'):
            cursor.execute('''
                DELETE cpp FROM catatan_penilaian_proposal cpp 
                INNER JOIN penilaian_proposal pp ON cpp.id_penilaian_proposal = pp.id 
                INNER JOIN proposal p ON pp.id_proposal = p.id 
                WHERE p.nim = %s
            ''', (nim,))
        catatan_penilaian_proposal_deleted = cursor.rowcount
        
        # 3. Hapus rekomendasi penilaian proposal (opsional, hanya jika tabel ada)
        if table_exists('rekomendasi_penilaian_proposal'):
            cursor.execute('''
                DELETE rpp FROM rekomendasi_penilaian_proposal rpp 
                INNER JOIN penilaian_proposal pp ON rpp.id_penilaian_proposal = pp.id 
                INNER JOIN proposal p ON pp.id_proposal = p.id 
                WHERE p.nim = %s
            ''', (nim,))
        rekomendasi_penilaian_proposal_deleted = cursor.rowcount
        
        # 4. Hapus penilaian proposal
        cursor.execute('''
            DELETE pp FROM penilaian_proposal pp 
            INNER JOIN proposal p ON pp.id_proposal = p.id 
            WHERE p.nim = %s
        ''', (nim,))
        penilaian_proposal_deleted = cursor.rowcount
        
        # 5. Hapus detail penilaian laporan kemajuan, lalu header
        cursor.execute('''
            DELETE dplk FROM detail_penilaian_laporan_kemajuan dplk
            INNER JOIN penilaian_laporan_kemajuan plk ON dplk.id_penilaian_laporan_kemajuan = plk.id
            INNER JOIN proposal p ON plk.id_proposal = p.id
            WHERE p.nim = %s
        ''', (nim,))
        cursor.execute('''
            DELETE plk FROM penilaian_laporan_kemajuan plk 
            INNER JOIN proposal p ON plk.id_proposal = p.id 
            WHERE p.nim = %s
        ''', (nim,))
        penilaian_laporan_kemajuan_deleted = cursor.rowcount
        
        # 6. Hapus detail penilaian laporan akhir, lalu header
        cursor.execute('''
            DELETE dpla FROM detail_penilaian_laporan_akhir dpla
            INNER JOIN penilaian_laporan_akhir pla ON dpla.id_penilaian_laporan_akhir = pla.id
            INNER JOIN proposal p ON pla.id_proposal = p.id
            WHERE p.nim = %s
        ''', (nim,))
        cursor.execute('''
            DELETE pla FROM penilaian_laporan_akhir pla 
            INNER JOIN proposal p ON pla.id_proposal = p.id 
            WHERE p.nim = %s
        ''', (nim,))

        # 7. Hapus detail penilaian mahasiswa (oleh pembimbing), lalu header
        cursor.execute('''
            DELETE dpm FROM detail_penilaian_mahasiswa dpm
            INNER JOIN penilaian_mahasiswa pm ON dpm.id_penilaian_mahasiswa = pm.id
            INNER JOIN proposal p ON pm.id_proposal = p.id
            WHERE p.nim = %s
        ''', (nim,))
        cursor.execute('''
            DELETE pm FROM penilaian_mahasiswa pm
            INNER JOIN proposal p ON pm.id_proposal = p.id
            WHERE p.nim = %s
        ''', (nim,))
        penilaian_laporan_akhir_deleted = cursor.rowcount
        
        # Hapus log aktivitas mahasiswa
        cursor.execute('DELETE FROM log_aktivitas_mahasiswa WHERE nim = %s', (nim,))
        log_aktivitas_mahasiswa_deleted = cursor.rowcount
        
        # Hapus session mahasiswa
        cursor.execute('DELETE FROM session_mahasiswa WHERE nim = %s', (nim,))
        session_mahasiswa_deleted = cursor.rowcount
        
        # Hapus produk_bahan_baku yang terkait dengan produksi dari proposal mahasiswa ini
        cursor.execute('''
            DELETE pbb FROM produk_bahan_baku pbb 
            INNER JOIN produksi prd ON pbb.produksi_id = prd.id 
            INNER JOIN proposal p ON prd.proposal_id = p.id 
            WHERE p.nim = %s
        ''', (nim,))
        produk_bahan_baku_deleted = cursor.rowcount
        
        # Hapus produksi (setelah produk_bahan_baku dihapus)
        cursor.execute('''
            DELETE pr FROM produksi pr 
            INNER JOIN proposal p ON pr.proposal_id = p.id 
            WHERE p.nim = %s
        ''', (nim,))
        produksi_deleted = cursor.rowcount

        # Hapus relasi proposal_reviewer yang terkait
        cursor.execute('''
            DELETE pr FROM proposal_reviewer pr
            INNER JOIN proposal p ON pr.id_proposal = p.id
            WHERE p.nim = %s
        ''', (nim,))
        proposal_reviewer_deleted = cursor.rowcount

        # Hapus catatan file_laporan_kemajuan dan file_laporan_akhir dari database
        cursor.execute('''
            DELETE flk FROM file_laporan_kemajuan flk
            INNER JOIN proposal p ON flk.id_proposal = p.id
            WHERE p.nim = %s
        ''', (nim,))
        file_laporan_kemajuan_deleted = cursor.rowcount
        cursor.execute('''
            DELETE fla FROM file_laporan_akhir fla
            INNER JOIN proposal p ON fla.id_proposal = p.id
            WHERE p.nim = %s
        ''', (nim,))
        file_laporan_akhir_deleted = cursor.rowcount
        
        # Hapus anggota tim terkait
        cursor.execute('''
            DELETE at FROM anggota_tim at 
            INNER JOIN proposal p ON at.id_proposal = p.id 
            WHERE p.nim = %s
        ''', (nim,))
        anggota_tim_deleted = cursor.rowcount
        
        # Hapus proposal
        cursor.execute('DELETE FROM proposal WHERE nim = %s', (nim,))
        proposal_deleted = cursor.rowcount
        
        # Hapus mahasiswa
        cursor.execute('DELETE FROM mahasiswa WHERE id = %s', (mahasiswa_id,))
        mahasiswa_deleted = cursor.rowcount
        
        get_app_functions()['mysql'].connection.commit()
        cursor.close()
        
        return jsonify({
            'success': True, 
            'message': (
                'Mahasiswa berhasil dihapus! Data yang dihapus: '
                f"{mahasiswa_deleted} mahasiswa, {proposal_deleted} proposal, {anggota_tim_deleted} anggota tim, "
                f"{anggaran_awal_deleted} anggaran awal, {anggaran_bertumbuh_deleted} anggaran bertumbuh, {bimbingan_deleted} bimbingan, "
                f"{penilaian_proposal_deleted} penilaian proposal, {penilaian_laporan_kemajuan_deleted} penilaian kemajuan, {penilaian_laporan_akhir_deleted} penilaian akhir, "
                f"{log_aktivitas_mahasiswa_deleted} log aktivitas, {session_mahasiswa_deleted} session, {proposal_reviewer_deleted} relasi reviewer, "
                f"{file_laporan_kemajuan_deleted} file laporan kemajuan, {file_laporan_akhir_deleted} file laporan akhir, dan semua file terkait."
            )
        })
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error saat menghapus data: {str(e)}'})

@admin_bp.route('/get_pembimbing_detail/<int:pembimbing_id>')
def get_pembimbing_detail(pembimbing_id):
    if 'user_type' not in session or session['user_type'] != 'admin':
        return jsonify({'success': False, 'message': 'Anda harus login sebagai admin!'})
    
    if not hasattr(get_app_functions()['mysql'], 'connection') or get_app_functions()['mysql'].connection is None:
        return jsonify({'success': False, 'message': 'Koneksi ke database gagal!'})
    
    try:
        cursor = get_app_functions()['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Ambil detail pembimbing
        cursor.execute('''
            SELECT id, nama, nip, program_studi, status, kuota_mahasiswa
            FROM pembimbing 
            WHERE id = %s
        ''', (pembimbing_id,))
        
        pembimbing = cursor.fetchone()
        
        # Ambil data program studi untuk dropdown
        cursor.execute('SELECT id, nama_program_studi FROM program_studi ORDER BY nama_program_studi')
        program_studi_list = cursor.fetchall()
        
        cursor.close()
        
        if pembimbing:
            return jsonify({
                'success': True, 
                'pembimbing': pembimbing,
                'program_studi_list': program_studi_list
            })
        else:
            return jsonify({'success': False, 'message': 'Data pembimbing tidak ditemukan!'})
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error saat mengambil data: {str(e)}'})

@admin_bp.route('/get_program_studi_list')
def get_program_studi_list():
    if 'user_type' not in session or session['user_type'] != 'admin':
        return jsonify({'success': False, 'message': 'Anda harus login sebagai admin!'})
    
    if not hasattr(get_app_functions()['mysql'], 'connection') or get_app_functions()['mysql'].connection is None:
        return jsonify({'success': False, 'message': 'Koneksi ke database gagal!'})
    
    try:
        cursor = get_app_functions()['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Ambil data program studi untuk dropdown
        cursor.execute('SELECT id, nama_program_studi FROM program_studi ORDER BY nama_program_studi')
        program_studi_list = cursor.fetchall()
        
        cursor.close()
        
        return jsonify({
            'success': True, 
            'program_studi_list': program_studi_list
        })
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error saat mengambil data: {str(e)}'})


@admin_bp.route('/monitoring_mahasiswa/bahan_baku')
def admin_monitoring_mahasiswa_bahan_baku():
    if 'user_type' not in session or session['user_type'] != 'admin':
        flash('Anda harus login sebagai admin!', 'danger')
        return redirect(url_for('index'))
    
    if not hasattr(get_app_functions()['mysql'], 'connection') or get_app_functions()['mysql'].connection is None:
        flash('Koneksi ke database gagal. Cek konfigurasi database!', 'danger')
        return render_template('admin/produksi_bahan_baku.html', mahasiswa_list=[])
    
    try:
        cursor = get_app_functions()['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Ambil semua mahasiswa
        cursor.execute('''
            SELECT m.id, m.nama_ketua, m.nim, p.judul_usaha, m.program_studi, p.id as proposal_id, p.dosen_pembimbing, p.kategori, p.tahapan_usaha
            FROM mahasiswa m
            JOIN proposal p ON m.nim = p.nim
            ORDER BY m.nama_ketua
        ''')
        
        mahasiswa_all = cursor.fetchall()
        mahasiswa_list = []
        
        for mhs in mahasiswa_all:
            proposal_id = mhs['proposal_id']
            
            # Cek apakah ada data bahan baku
            cursor.execute('SELECT COUNT(*) as cnt FROM bahan_baku WHERE proposal_id = %s', (proposal_id,))
            bahan_baku_count = cursor.fetchone()['cnt']
            
            # Cek apakah ada data produksi
            cursor.execute('SELECT COUNT(*) as cnt FROM produksi WHERE proposal_id = %s', (proposal_id,))
            produksi_count = cursor.fetchone()['cnt']
            
            # Tambahkan informasi data yang tersedia
            mhs['has_bahan_baku'] = bahan_baku_count > 0
            mhs['has_produksi'] = produksi_count > 0
            
            mahasiswa_list.append(mhs)
        
        cursor.close()
        
        return render_template('admin/produksi_bahan_baku.html', mahasiswa_list=mahasiswa_list, mahasiswa_info=None, bahan_baku_list=[])
        
    except Exception as e:
        print(f"Error: {e}")
        flash('Terjadi kesalahan saat mengambil data!', 'danger')
        return render_template('admin/daftar_mahsiswa_bahan_baku.html', mahasiswa_list=[], mahasiswa_info=None, bahan_baku_list=[])

@admin_bp.route('/monitoring_mahasiswa/produksi')
def admin_monitoring_mahasiswa_produksi():
    if 'user_type' not in session or session['user_type'] != 'admin':
        flash('Anda harus login sebagai admin!', 'danger')
        return redirect(url_for('index'))
    
    if not hasattr(get_app_functions()['mysql'], 'connection') or get_app_functions()['mysql'].connection is None:
        flash('Koneksi ke database gagal. Cek konfigurasi database!', 'danger')
        return render_template('admin/produksi_bahan_baku.html', mahasiswa_list=[])
    
    try:
        cursor = get_app_functions()['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Ambil semua mahasiswa
        cursor.execute('''
            SELECT m.id, m.nama_ketua, m.nim, p.judul_usaha, m.program_studi, p.id as proposal_id, p.dosen_pembimbing, p.kategori, p.tahapan_usaha
            FROM mahasiswa m
            JOIN proposal p ON m.nim = p.nim
            ORDER BY m.nama_ketua
        ''')
        
        mahasiswa_all = cursor.fetchall()
        mahasiswa_list = []
        
        for mhs in mahasiswa_all:
            proposal_id = mhs['proposal_id']
            
            # Cek apakah ada data bahan baku
            cursor.execute('SELECT COUNT(*) as cnt FROM bahan_baku WHERE proposal_id = %s', (proposal_id,))
            bahan_baku_count = cursor.fetchone()['cnt']
            
            # Cek apakah ada data produksi
            cursor.execute('SELECT COUNT(*) as cnt FROM produksi WHERE proposal_id = %s', (proposal_id,))
            produksi_count = cursor.fetchone()['cnt']
            
            # Tambahkan informasi data yang tersedia
            mhs['has_bahan_baku'] = bahan_baku_count > 0
            mhs['has_produksi'] = produksi_count > 0
            
            mahasiswa_list.append(mhs)
        
        cursor.close()
        
        return render_template('admin/produksi_bahan_baku.html', mahasiswa_list=mahasiswa_list, mahasiswa_info=None, produksi_list=[])
        
    except Exception as e:
        print(f"Error: {e}")
        flash('Terjadi kesalahan saat mengambil data!', 'danger')
        return render_template('admin/produksi_bahan_baku.html', mahasiswa_list=[], mahasiswa_info=None, produksi_list=[])

@admin_bp.route('/monitoring_mahasiswa/bahan_baku/<int:mahasiswa_id>')
def admin_monitoring_mahasiswa_bahan_baku_detail(mahasiswa_id):
    if 'user_type' not in session or session['user_type'] != 'admin':
        flash('Anda harus login sebagai admin!', 'danger')
        return redirect(url_for('index'))
    
    if not hasattr(get_app_functions()['mysql'], 'connection') or get_app_functions()['mysql'].connection is None:
        flash('Koneksi ke database gagal. Cek konfigurasi database!', 'danger')
        return render_template('admin/bahan_baku.html', mahasiswa_info=None, bahan_baku_list=[])
    
    try:
        cursor = get_app_functions()['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Ambil data mahasiswa
        cursor.execute('''
            SELECT m.*, p.judul_usaha, p.id as proposal_id, p.dosen_pembimbing
            FROM mahasiswa m
            JOIN proposal p ON m.nim = p.nim
            WHERE m.id = %s
        ''', (mahasiswa_id,))
        
        mahasiswa_info = cursor.fetchone()
        if not mahasiswa_info:
            flash('Data mahasiswa tidak ditemukan!', 'danger')
            cursor.close()
            return redirect(url_for('admin_monitoring_mahasiswa_produksi'))
        
        # Cek apakah ada data produksi
        cursor.execute('SELECT COUNT(*) as cnt FROM produksi WHERE proposal_id = %s', (mahasiswa_info['proposal_id'],))
        produksi_count = cursor.fetchone()['cnt']
        mahasiswa_info['has_produksi'] = produksi_count > 0
        
        # Ambil data bahan baku
        cursor.execute('''
            SELECT * FROM bahan_baku 
            WHERE proposal_id = %s 
            ORDER BY tanggal_beli DESC
        ''', (mahasiswa_info['proposal_id'],))
        
        bahan_baku_list = cursor.fetchall()
        
        cursor.close()
        
        return render_template('admin/bahan_baku.html', mahasiswa_info=mahasiswa_info, bahan_baku_list=bahan_baku_list, mahasiswa_list=[])
        
    except Exception as e:
        print(f"Error: {e}")
        flash('Terjadi kesalahan saat mengambil data!', 'danger')
        return render_template('admin/bahan_baku.html', mahasiswa_info=None, bahan_baku_list=[], mahasiswa_list=[])

@admin_bp.route('/monitoring_mahasiswa/produksi/<int:mahasiswa_id>')
def admin_monitoring_mahasiswa_produksi_detail(mahasiswa_id):
    if 'user_type' not in session or session['user_type'] != 'admin':
        flash('Anda harus login sebagai admin!', 'danger')
        return redirect(url_for('index'))
    
    if not hasattr(get_app_functions()['mysql'], 'connection') or get_app_functions()['mysql'].connection is None:
        flash('Koneksi ke database gagal. Cek konfigurasi database!', 'danger')
        return render_template('admin/produksi.html', mahasiswa_info=None, produksi_list=[])
    
    try:
        cursor = get_app_functions()['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Ambil data mahasiswa
        cursor.execute('''
            SELECT m.*, p.judul_usaha, p.id as proposal_id, p.dosen_pembimbing
            FROM mahasiswa m
            JOIN proposal p ON m.nim = p.nim
            WHERE m.id = %s
        ''', (mahasiswa_id,))
        
        mahasiswa_info = cursor.fetchone()
        if not mahasiswa_info:
            flash('Data mahasiswa tidak ditemukan!', 'danger')
            cursor.close()
            return redirect(url_for('admin_monitoring_mahasiswa_produksi'))
        
        # Pastikan mahasiswa_info memiliki field id
        mahasiswa_info['id'] = mahasiswa_id
        
        # Cek apakah ada data bahan baku
        cursor.execute('SELECT COUNT(*) as cnt FROM bahan_baku WHERE proposal_id = %s', (mahasiswa_info['proposal_id'],))
        bahan_baku_count = cursor.fetchone()['cnt']
        mahasiswa_info['has_bahan_baku'] = bahan_baku_count > 0
        
        # Ambil data produksi
        cursor.execute('''
            SELECT * FROM produksi 
            WHERE proposal_id = %s 
            ORDER BY tanggal_produksi DESC
        ''', (mahasiswa_info['proposal_id'],))
        
        produksi_list = cursor.fetchall()
        
        # Tambahkan data bahan baku untuk setiap produksi
        for produksi in produksi_list:
            cursor.execute('''
                SELECT pbb.*, bb.nama_bahan, bb.satuan
                FROM produk_bahan_baku pbb
                JOIN bahan_baku bb ON pbb.bahan_baku_id = bb.id
                WHERE pbb.produksi_id = %s
            ''', (produksi['id'],))
            bahan_baku_list = cursor.fetchall()
            produksi['bahan_baku'] = bahan_baku_list
        
        cursor.close()
        

        return render_template('admin/produksi.html', mahasiswa_info=mahasiswa_info, produksi_list=produksi_list, mahasiswa_list=[])
        
    except Exception as e:
        print(f"Error: {e}")
        flash('Terjadi kesalahan saat mengambil data!', 'danger')
        return render_template('admin/produksi.html', mahasiswa_info=None, produksi_list=[], mahasiswa_list=[])

@admin_bp.route('/monitoring_mahasiswa/penjualan_produk')
def admin_monitoring_mahasiswa_penjualan_produk():
    if 'user_type' not in session or session['user_type'] != 'admin':
        flash('Anda harus login sebagai admin!', 'danger')
        return redirect(url_for('index'))
    
    if not hasattr(get_app_functions()['mysql'], 'connection') or get_app_functions()['mysql'].connection is None:
        flash('Koneksi ke database gagal. Cek konfigurasi database!', 'danger')
        return render_template('admin/penjualan.html', mahasiswa_list=[])
    
    try:
        cursor = get_app_functions()['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Ambil semua mahasiswa
        cursor.execute('''
            SELECT m.id, m.nama_ketua, m.nim, p.judul_usaha, m.program_studi, p.id as proposal_id, p.dosen_pembimbing, p.kategori, p.tahapan_usaha
            FROM mahasiswa m
            JOIN proposal p ON m.nim = p.nim
            ORDER BY m.nama_ketua
        ''')
        
        mahasiswa_all = cursor.fetchall()
        mahasiswa_list = []
        
        for mhs in mahasiswa_all:
            proposal_id = mhs['proposal_id']
            
            # Cek apakah ada data penjualan
            cursor.execute('SELECT COUNT(*) as cnt FROM penjualan WHERE proposal_id = %s', (proposal_id,))
            penjualan_count = cursor.fetchone()['cnt']
            
            # Tambahkan informasi data yang tersedia
            mhs['has_penjualan'] = penjualan_count > 0
            
            mahasiswa_list.append(mhs)
        
        cursor.close()
        
        return render_template('admin/penjualan.html', mahasiswa_list=mahasiswa_list)
        
    except Exception as e:
        print(f"Error: {e}")
        flash('Terjadi kesalahan saat mengambil data!', 'danger')
        return render_template('admin/penjualan.html', mahasiswa_list=[])

@admin_bp.route('/monitoring_mahasiswa/laporan_laba_rugi')
def admin_monitoring_mahasiswa_laporan_laba_rugi():
    if 'user_type' not in session or session['user_type'] != 'admin':
        flash('Anda harus login sebagai admin!', 'danger')
        return redirect(url_for('index'))
    
    if not hasattr(get_app_functions()['mysql'], 'connection') or get_app_functions()['mysql'].connection is None:
        flash('Koneksi ke database gagal. Cek konfigurasi database!', 'danger')
        return render_template('admin/laba_rugi_produk.html', mahasiswa_list=[])
    
    try:
        cursor = get_app_functions()['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Ambil semua mahasiswa
        cursor.execute('''
            SELECT m.id, m.nama_ketua, m.nim, p.judul_usaha, m.program_studi, p.id as proposal_id, p.dosen_pembimbing, p.kategori, p.tahapan_usaha
            FROM mahasiswa m
            JOIN proposal p ON m.nim = p.nim
            ORDER BY m.nama_ketua
        ''')
        mahasiswa_list = cursor.fetchall()
        
        # Cek apakah setiap mahasiswa memiliki data laba rugi
        for mahasiswa in mahasiswa_list:
            if mahasiswa['proposal_id']:
                cursor.execute('''
                    SELECT COUNT(*) as count
                    FROM produksi pr
                    WHERE pr.proposal_id = %s
                ''', (mahasiswa['proposal_id'],))
                result = cursor.fetchone()
                mahasiswa['has_laba_rugi'] = result['count'] > 0
            else:
                mahasiswa['has_laba_rugi'] = False
        
        cursor.close()
        return render_template('admin/laba_rugi_produk.html', mahasiswa_list=mahasiswa_list)
        
    except Exception as e:
        print(f"Error in admin_monitoring_mahasiswa_laporan_laba_rugi: {e}")
        flash('Terjadi kesalahan saat mengambil data mahasiswa!', 'danger')
        return render_template('admin/laba_rugi_produk.html', mahasiswa_list=[])

@admin_bp.route('/monitoring_mahasiswa/laporan_arus_kas')
def admin_monitoring_mahasiswa_laporan_arus_kas():
    if 'user_type' not in session or session['user_type'] != 'admin':
        flash('Anda harus login sebagai admin!', 'danger')
        return redirect(url_for('index'))
    
    if not hasattr(get_app_functions()['mysql'], 'connection') or get_app_functions()['mysql'].connection is None:
        flash('Koneksi ke database gagal. Cek konfigurasi database!', 'danger')
        return render_template('admin/daftar_mahasiswa_arus_kas.html', mahasiswa_list=[])
    
    try:
        cursor = get_app_functions()['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Ambil semua mahasiswa (tanpa filter pembimbing untuk admin)
        cursor.execute('''
            SELECT m.id, m.nama_ketua, m.nim, p.judul_usaha, m.program_studi, p.id as proposal_id, p.dosen_pembimbing, p.kategori, p.tahapan_usaha
            FROM mahasiswa m
            JOIN proposal p ON m.nim = p.nim
            ORDER BY m.nama_ketua
        ''')
        
        mahasiswa_all = cursor.fetchall()
        mahasiswa_list = []
        
        for mhs in mahasiswa_all:
            proposal_id = mhs['proposal_id']
            
            # Cek apakah ada data arus kas dari tabel arus_kas
            cursor.execute('SELECT COUNT(*) as cnt FROM arus_kas WHERE proposal_id = %s', (proposal_id,))
            arus_kas_count = cursor.fetchone()['cnt']
            
            # Jika tidak ada di tabel arus_kas, cek dari tabel individual
            if arus_kas_count == 0:
                cursor.execute('SELECT COUNT(*) as cnt FROM penjualan WHERE proposal_id = %s', (proposal_id,))
                penjualan_count = cursor.fetchone()['cnt']
                
                cursor.execute('SELECT COUNT(*) as cnt FROM biaya_operasional WHERE proposal_id = %s', (proposal_id,))
                biaya_operasional_count = cursor.fetchone()['cnt']
                
                cursor.execute('SELECT COUNT(*) as cnt FROM biaya_non_operasional WHERE proposal_id = %s', (proposal_id,))
                biaya_non_operasional_count = cursor.fetchone()['cnt']
                
                cursor.execute('SELECT COUNT(*) as cnt FROM alat_produksi WHERE proposal_id = %s', (proposal_id,))
                alat_produksi_count = cursor.fetchone()['cnt']
                
                # Tambahkan informasi data yang tersedia
                mhs['has_arus_kas'] = (penjualan_count > 0 or biaya_operasional_count > 0 or 
                                      biaya_non_operasional_count > 0 or alat_produksi_count > 0)
            else:
                mhs['has_arus_kas'] = True
            
            mahasiswa_list.append(mhs)
        
        cursor.close()
        
        return render_template('admin/daftar_mahasiswa_arus_kas.html', mahasiswa_list=mahasiswa_list)
        
    except Exception as e:
        flash(f'Error: {str(e)}', 'danger')
        return render_template('admin/daftar_mahasiswa_arus_kas.html', mahasiswa_list=[])

@admin_bp.route('/monitoring_mahasiswa/alat_produksi')
def admin_monitoring_mahasiswa_alat_produksi():
    if 'user_type' not in session or session['user_type'] != 'admin':
        flash('Anda harus login sebagai admin!', 'danger')
        return redirect(url_for('index'))
    
    if not hasattr(get_app_functions()['mysql'], 'connection') or get_app_functions()['mysql'].connection is None:
        flash('Koneksi ke database gagal. Cek konfigurasi database!', 'danger')
        return render_template('admin/daftar_mahsiswa_alat_produksi.html', mahasiswa_list=[])
    
    try:
        cursor = get_app_functions()['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Ambil semua mahasiswa (tanpa filter pembimbing untuk admin)
        cursor.execute('''
            SELECT m.id, m.nama_ketua, m.nim, p.judul_usaha, m.program_studi, p.id as proposal_id, p.dosen_pembimbing, p.kategori, p.tahapan_usaha
            FROM mahasiswa m
            JOIN proposal p ON m.nim = p.nim
            ORDER BY m.nama_ketua
        ''')
        
        mahasiswa_all = cursor.fetchall()
        mahasiswa_list = []
        
        for mhs in mahasiswa_all:
            proposal_id = mhs['proposal_id']
            
            # Cek apakah ada data alat produksi
            cursor.execute('SELECT COUNT(*) as cnt FROM alat_produksi WHERE proposal_id = %s', (proposal_id,))
            alat_produksi_count = cursor.fetchone()['cnt']
            
            # Tambahkan informasi data yang tersedia
            mhs['has_alat_produksi'] = alat_produksi_count > 0
            
            mahasiswa_list.append(mhs)
        
        cursor.close()
        
        return render_template('admin/daftar_mahsiswa_alat_produksi.html', mahasiswa_list=mahasiswa_list)
        
    except Exception as e:
        flash(f'Error: {str(e)}', 'danger')
        return render_template('admin/daftar_mahsiswa_alat_produksi.html', mahasiswa_list=[])

@admin_bp.route('/monitoring_mahasiswa/biaya_operasional')
def admin_monitoring_mahasiswa_biaya_operasional():
    if 'user_type' not in session or session['user_type'] != 'admin':
        flash('Anda harus login sebagai admin!', 'danger')
        return redirect(url_for('index'))
    
    if not hasattr(get_app_functions()['mysql'], 'connection') or get_app_functions()['mysql'].connection is None:
        flash('Koneksi ke database gagal. Cek konfigurasi database!', 'danger')
        return render_template('admin/daftar_mahsiswa_biaya_operasional.html', mahasiswa_list=[])
    
    try:
        cursor = get_app_functions()['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Ambil semua mahasiswa (tanpa filter pembimbing untuk admin)
        cursor.execute('''
            SELECT m.id, m.nama_ketua, m.nim, p.judul_usaha, m.program_studi, p.id as proposal_id, p.dosen_pembimbing, p.kategori, p.tahapan_usaha
            FROM mahasiswa m
            JOIN proposal p ON m.nim = p.nim
            ORDER BY m.nama_ketua
        ''')
        
        mahasiswa_all = cursor.fetchall()
        mahasiswa_list = []
        
        for mhs in mahasiswa_all:
            proposal_id = mhs['proposal_id']
            
            # Cek apakah ada data biaya operasional
            cursor.execute('SELECT COUNT(*) as cnt FROM biaya_operasional WHERE proposal_id = %s', (proposal_id,))
            biaya_operasional_count = cursor.fetchone()['cnt']
            
            # Tambahkan informasi data yang tersedia
            mhs['has_biaya_operasional'] = biaya_operasional_count > 0
            
            mahasiswa_list.append(mhs)
        
        cursor.close()
        
        return render_template('admin/daftar_mahsiswa_biaya_operasional.html', mahasiswa_list=mahasiswa_list)
        
    except Exception as e:
        flash(f'Error: {str(e)}', 'danger')
        return render_template('admin/daftar_mahsiswa_biaya_operasional.html', mahasiswa_list=[])

@admin_bp.route('/monitoring_mahasiswa/biaya_non_operasional')
def admin_monitoring_mahasiswa_biaya_non_operasional():
    if 'user_type' not in session or session['user_type'] != 'admin':
        flash('Anda harus login sebagai admin!', 'danger')
        return redirect(url_for('index'))
    
    if not hasattr(get_app_functions()['mysql'], 'connection') or get_app_functions()['mysql'].connection is None:
        flash('Koneksi ke database gagal. Cek konfigurasi database!', 'danger')
        return render_template('admin/daftar_non_biaya_operasional.html', mahasiswa_list=[])
    
    try:
        cursor = get_app_functions()['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Ambil semua mahasiswa (tanpa filter pembimbing untuk admin)
        cursor.execute('''
            SELECT m.id, m.nama_ketua, m.nim, p.judul_usaha, m.program_studi, p.id as proposal_id, p.dosen_pembimbing, p.kategori, p.tahapan_usaha
            FROM mahasiswa m
            JOIN proposal p ON m.nim = p.nim
            ORDER BY m.nama_ketua
        ''')
        
        mahasiswa_all = cursor.fetchall()
        mahasiswa_list = []
        
        for mhs in mahasiswa_all:
            proposal_id = mhs['proposal_id']
            
            # Cek apakah ada data Biaya Lain-lain
            cursor.execute('SELECT COUNT(*) as cnt FROM biaya_non_operasional WHERE proposal_id = %s', (proposal_id,))
            biaya_non_operasional_count = cursor.fetchone()['cnt']
            
            # Tambahkan informasi data yang tersedia
            mhs['has_biaya_non_operasional'] = biaya_non_operasional_count > 0
            
            mahasiswa_list.append(mhs)
        
        cursor.close()
        
        return render_template('admin/daftar_non_biaya_operasional.html', mahasiswa_list=mahasiswa_list)
        
    except Exception as e:
        flash(f'Error: {str(e)}', 'danger')
        return render_template('admin/daftar_non_biaya_operasional.html', mahasiswa_list=[])

@admin_bp.route('/biaya_non_operasional/<int:mahasiswa_id>')
def admin_biaya_non_operasional(mahasiswa_id):
    if 'user_type' not in session or session['user_type'] != 'admin':
        flash('Anda harus login sebagai admin!', 'danger')
        return redirect(url_for('index'))
    
    if not hasattr(get_app_functions()['mysql'], 'connection') or get_app_functions()['mysql'].connection is None:
        flash('Koneksi ke database gagal. Cek konfigurasi database!', 'danger')
        return render_template('admin/biaya_non_operasional.html', biaya_non_operasional_list=[], mahasiswa_info={})
    
    try:
        cursor = get_app_functions()['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Ambil data mahasiswa
        cursor.execute('''
            SELECT m.*, p.judul_usaha, p.id as proposal_id
            FROM mahasiswa m
            JOIN proposal p ON m.nim = p.nim
            WHERE m.id = %s
        ''', (mahasiswa_id,))
        
        mahasiswa_info = cursor.fetchone()
        
        if not mahasiswa_info:
            flash('Data mahasiswa tidak ditemukan!', 'danger')
            cursor.close()
            return redirect(url_for('admin_monitoring_mahasiswa_biaya_non_operasional'))
        
        proposal_id = mahasiswa_info['proposal_id']
        
        # Ambil data Biaya Lain-lain
        cursor.execute('''
            SELECT * FROM biaya_non_operasional 
            WHERE proposal_id = %s 
            ORDER BY created_at DESC
        ''', (proposal_id,))
        
        biaya_non_operasional_list = cursor.fetchall()
        cursor.close()
        
        return render_template('admin/biaya_non_operasional.html', 
                             biaya_non_operasional_list=biaya_non_operasional_list, 
                             mahasiswa_info=mahasiswa_info)
        
    except Exception as e:
        flash(f'Error: {str(e)}', 'danger')
        return render_template('admin/biaya_non_operasional.html', biaya_non_operasional_list=[], mahasiswa_info={})

@admin_bp.route('/arus_kas/<int:mahasiswa_id>')
def admin_arus_kas(mahasiswa_id):
    if 'user_type' not in session or session['user_type'] != 'admin':
        flash('Anda harus login sebagai admin!', 'danger')
        return redirect(url_for('index'))
    
    if not hasattr(get_app_functions()['mysql'], 'connection') or get_app_functions()['mysql'].connection is None:
        flash('Koneksi ke database gagal. Cek konfigurasi database!', 'danger')
        return render_template('admin/arus_kas.html', mahasiswa_info=None, proposal_data={}, has_lolos_proposal=False)
    
    try:
        cursor = get_app_functions()['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Ambil data mahasiswa
        cursor.execute('''
            SELECT m.*, p.judul_usaha, p.id as proposal_id
            FROM mahasiswa m
            JOIN proposal p ON m.nim = p.nim
            WHERE m.id = %s
        ''', (mahasiswa_id,))
        
        mahasiswa_info = cursor.fetchone()
        if not mahasiswa_info:
            flash('Data mahasiswa tidak ditemukan!', 'danger')
            cursor.close()
            return redirect(url_for('admin_monitoring_mahasiswa_laporan_arus_kas'))
        
        # Pastikan mahasiswa_info memiliki field id
        mahasiswa_info['id'] = mahasiswa_id
        
        # Ambil data arus kas dari tabel arus_kas
        proposal_id = mahasiswa_info['proposal_id']
        
        # Ambil bulan yang dipilih (default: cek data terbaru di database)
        selected_month = request.args.get('month')
        
        # Jika tidak ada bulan yang dipilih, ambil bulan terbaru dari database
        if not selected_month:
            cursor.execute('''
                SELECT bulan_tahun 
                FROM arus_kas 
                WHERE proposal_id = %s 
                ORDER BY bulan_tahun DESC 
                LIMIT 1
            ''', (proposal_id,))
            latest_month = cursor.fetchone()
            if latest_month:
                selected_month = latest_month['bulan_tahun']
            else:
                selected_month = datetime.now().strftime('%Y-%m')
        
        # Parse bulan dan tahun
        try:
            selected_date = datetime.strptime(selected_month, '%Y-%m')
            selected_month_year = selected_date.strftime('%B %Y')
        except:
            selected_date = datetime.now()
            selected_month_year = selected_date.strftime('%B %Y')
        
        # Ambil data arus kas dari tabel arus_kas
        cursor.execute('''
            SELECT total_penjualan, total_biaya_produksi, total_biaya_operasional,
                   total_biaya_non_operasional, kas_bersih_operasional, total_harga_jual_alat,
                   total_harga_alat, kas_bersih_investasi, kas_bersih_pembiayaan, total_kas_bersih
            FROM arus_kas 
            WHERE proposal_id = %s AND bulan_tahun = %s
        ''', (proposal_id, selected_month))
        
        arus_kas_data = cursor.fetchone()
        
        if arus_kas_data:
            # Jika data ada di tabel arus_kas, gunakan data tersebut
            total_penjualan = arus_kas_data['total_penjualan'] or 0
            total_biaya_produksi = arus_kas_data['total_biaya_produksi'] or 0
            total_biaya_operasional = arus_kas_data['total_biaya_operasional'] or 0
            total_biaya_non_operasional = arus_kas_data['total_biaya_non_operasional'] or 0
            kas_bersih_operasional = arus_kas_data['kas_bersih_operasional'] or 0
            total_harga_jual_alat = arus_kas_data['total_harga_jual_alat'] or 0
            total_harga_alat = arus_kas_data['total_harga_alat'] or 0
            kas_bersih_investasi = arus_kas_data['kas_bersih_investasi'] or 0
            kas_bersih_pembiayaan = arus_kas_data['kas_bersih_pembiayaan'] or 0
            total_kas_bersih = arus_kas_data['total_kas_bersih'] or 0
        else:
            # Jika data tidak ada di tabel arus_kas, hitung dari tabel individual
            cursor.execute('''
                SELECT COALESCE(SUM(total_harga), 0) as total_penjualan
                FROM penjualan 
                WHERE proposal_id = %s 
                AND DATE_FORMAT(tanggal_penjualan, '%%Y-%%m') = %s
            ''', (proposal_id, selected_month))
            total_penjualan = cursor.fetchone()['total_penjualan'] or 0
            
            cursor.execute('''
                SELECT COALESCE(SUM(total_harga), 0) as total_biaya_produksi
                FROM bahan_baku 
                WHERE proposal_id = %s 
                AND DATE_FORMAT(tanggal_beli, '%%Y-%%m') = %s
            ''', (proposal_id, selected_month))
            total_biaya_produksi = cursor.fetchone()['total_biaya_produksi'] or 0
            
            cursor.execute('''
                SELECT COALESCE(SUM(total_biaya), 0) as total_biaya_operasional
                FROM biaya_operasional 
                WHERE proposal_id = %s 
                AND DATE_FORMAT(tanggal, '%%Y-%%m') = %s
            ''', (proposal_id, selected_month))
            total_biaya_operasional = cursor.fetchone()['total_biaya_operasional'] or 0
            
            cursor.execute('''
                SELECT COALESCE(SUM(total_biaya), 0) as total_biaya_non_operasional
                FROM biaya_non_operasional 
                WHERE proposal_id = %s 
                AND DATE_FORMAT(tanggal, '%%Y-%%m') = %s
            ''', (proposal_id, selected_month))
            total_biaya_non_operasional = cursor.fetchone()['total_biaya_non_operasional'] or 0
            
            cursor.execute('''
                SELECT 
                    COALESCE(SUM(harga_alat), 0) as total_harga_alat,
                    COALESCE(SUM(harga_jual), 0) as total_harga_jual_alat
                FROM alat_produksi 
                WHERE proposal_id = %s 
                AND DATE_FORMAT(tanggal_beli, '%%Y-%%m') = %s
            ''', (proposal_id, selected_month))
            
            alat_data = cursor.fetchone()
            total_harga_alat = alat_data['total_harga_alat'] or 0
            total_harga_jual_alat = alat_data['total_harga_jual_alat'] or 0
            
            # Hitung kas bersih dari kegiatan operasional
            kas_bersih_operasional = total_penjualan - total_biaya_produksi - total_biaya_operasional - total_biaya_non_operasional
            
            # Hitung kas bersih dari kegiatan investasi
            kas_bersih_investasi = total_harga_jual_alat - total_harga_alat
            
            # Hitung kas bersih dari kegiatan pembiayaan (selalu 0)
            kas_bersih_pembiayaan = 0
            
            # Hitung total kas bersih
            total_kas_bersih = kas_bersih_operasional + kas_bersih_investasi
        
        # Siapkan data untuk template
        proposal_data = {
            proposal_id: {
                'proposal': {
                    'id': proposal_id,
                    'judul_usaha': mahasiswa_info['judul_usaha'],
                    'nama_ketua': mahasiswa_info['nama_ketua'],
                    'nim': mahasiswa_info['nim'],
                    'program_studi': mahasiswa_info['program_studi']
                },
                'total_penjualan': total_penjualan,
                'total_biaya_produksi': total_biaya_produksi,
                'total_biaya_operasional': total_biaya_operasional,
                'total_biaya_non_operasional': total_biaya_non_operasional,
                'total_harga_alat': total_harga_alat,
                'total_harga_jual_alat': total_harga_jual_alat,
                'kas_bersih_operasional': kas_bersih_operasional,
                'kas_bersih_investasi': kas_bersih_investasi,
                'kas_bersih_pembiayaan': kas_bersih_pembiayaan,
                'total_kas_bersih': total_kas_bersih
            }
        }
        
        # Ambil daftar bulan yang tersedia dari tabel arus_kas
        cursor.execute('''
            SELECT DISTINCT bulan_tahun as month_value
            FROM arus_kas 
            WHERE proposal_id = %s
            ORDER BY bulan_tahun DESC
        ''', (proposal_id,))
        
        available_months = []
        for row in cursor.fetchall():
            month_value = row['month_value']
            try:
                month_date = datetime.strptime(month_value, '%Y-%m')
                month_label = month_date.strftime('%B %Y')
                available_months.append({
                    'value': month_value,
                    'label': month_label
                })
            except:
                continue
        
        cursor.close()
        

        return render_template('admin/arus_kas.html', 
                             mahasiswa_info=mahasiswa_info,
                             proposal_data=proposal_data,
                             selected_month=selected_month,
                             selected_month_year=selected_month_year,
                             available_months=available_months,
                             has_lolos_proposal=True)
        
    except Exception as e:
        flash(f'Error: {str(e)}', 'danger')
        return render_template('admin/arus_kas.html', 
                             mahasiswa_info=None, 
                             proposal_data={},
                             has_lolos_proposal=False)

@admin_bp.route('/download_arus_kas', methods=['POST'])
def admin_download_arus_kas():
    import traceback
    if 'user_type' not in session or session['user_type'] != 'admin':
        return jsonify({'success': False, 'message': 'Unauthorized'})
    
    try:
        proposal_id = request.form.get('proposal_id')
        mahasiswa_id = request.form.get('mahasiswa_id')
        format_type = request.form.get('format')
        bulan_tahun = request.form.get('bulan_tahun')
        
        if not all([proposal_id, mahasiswa_id, format_type, bulan_tahun]):
            return jsonify({'success': False, 'message': 'Parameter tidak lengkap'})
        
        if not hasattr(get_app_functions()['mysql'], 'connection') or get_app_functions()['mysql'].connection is None:
            return jsonify({'success': False, 'message': 'Koneksi database tidak tersedia'})
        
        cursor = get_app_functions()['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Debug: Print parameter yang diterima
        print(f"=== DEBUG ADMIN DOWNLOAD ARUS KAS ===")
        print(f"Proposal ID: {proposal_id}")
        print(f"Mahasiswa ID: {mahasiswa_id}")
        print(f"Format: {format_type}")
        print(f"Bulan Tahun: {bulan_tahun}")
        print(f"Proposal ID type: {type(proposal_id)}")
        print(f"Mahasiswa ID type: {type(mahasiswa_id)}")
        print("=====================================")
        
        # Verifikasi bahwa mahasiswa ada - gunakan query yang lebih sederhana
        print(f"DEBUG - Proposal ID dan Mahasiswa ID sama? {proposal_id == mahasiswa_id}")
        
        cursor.execute('''
            SELECT m.nama_ketua, m.nim, m.program_studi, p.judul_usaha, p.kategori, p.tahun_nib
            FROM mahasiswa m
            JOIN proposal p ON m.nim = p.nim
            WHERE p.id = %s
        ''', (proposal_id,))
        
        mahasiswa_info = cursor.fetchone()
        
        print(f"Mahasiswa info found: {mahasiswa_info is not None}")
        if mahasiswa_info:
            print(f"Mahasiswa data: {mahasiswa_info}")
        
        if not mahasiswa_info:
            # Debug: Cek data proposal
            print("DEBUG - Query gagal, cek data proposal...")
            cursor.execute('SELECT id, nim, judul_usaha FROM proposal WHERE id = %s', (proposal_id,))
            proposal_data = cursor.fetchone()
            print(f"DEBUG - proposal_data: {proposal_data}")
            
            if proposal_data:
                print("DEBUG - Cek data mahasiswa berdasarkan NIM...")
                cursor.execute('SELECT id, nama_ketua, nim FROM mahasiswa WHERE nim = %s', (proposal_data['nim'],))
                mahasiswa_data = cursor.fetchone()
                print(f"DEBUG - mahasiswa_data: {mahasiswa_data}")
                
                # Jika mahasiswa ditemukan berdasarkan NIM, gunakan data tersebut
                if mahasiswa_data:
                    mahasiswa_info = {
                        'nama_ketua': mahasiswa_data['nama_ketua'],
                        'nim': mahasiswa_data['nim'],
                        'program_studi': '',  # Default value
                        'judul_usaha': proposal_data['judul_usaha'],
                        'kategori': '',  # Default value
                        'tahun_nib': ''  # Default value
                    }
                    print(f"Using fallback mahasiswa info: {mahasiswa_info}")
            else:
                # Jika proposal tidak ditemukan, coba cek semua proposal yang ada
                print("DEBUG - Proposal tidak ditemukan, cek semua proposal...")
                cursor.execute('SELECT id, nim, judul_usaha FROM proposal ORDER BY id')
                all_proposals = cursor.fetchall()
                print(f"DEBUG - All proposals: {all_proposals}")
                
                # Coba cek mahasiswa berdasarkan mahasiswa_id
                print("DEBUG - Cek mahasiswa berdasarkan mahasiswa_id...")
                cursor.execute('SELECT id, nama_ketua, nim FROM mahasiswa WHERE id = %s', (mahasiswa_id,))
                mahasiswa_by_id = cursor.fetchone()
                print(f"DEBUG - mahasiswa_by_id: {mahasiswa_by_id}")
                
                if mahasiswa_by_id:
                    # Jika mahasiswa ditemukan, cari proposal berdasarkan NIM
                    cursor.execute('SELECT id, nim, judul_usaha FROM proposal WHERE nim = %s', (mahasiswa_by_id['nim'],))
                    proposal_by_nim = cursor.fetchone()
                    print(f"DEBUG - proposal_by_nim: {proposal_by_nim}")
                    
                    if proposal_by_nim:
                        # Update proposal_id dengan ID yang benar
                        proposal_id = proposal_by_nim['id']
                        print(f"DEBUG - Updated proposal_id to: {proposal_id}")
                        
                        mahasiswa_info = {
                            'nama_ketua': mahasiswa_by_id['nama_ketua'],
                            'nim': mahasiswa_by_id['nim'],
                            'program_studi': '',  # Default value
                            'judul_usaha': proposal_by_nim['judul_usaha'],
                            'kategori': '',  # Default value
                            'tahun_nib': ''  # Default value
                        }
                        print(f"Using mahasiswa_id fallback: {mahasiswa_info}")
            
            if not mahasiswa_info:
                cursor.close()
                return jsonify({'success': False, 'message': 'Data mahasiswa tidak ditemukan'})
        
        # Parse bulan yang dipilih
        if not bulan_tahun:
            return jsonify({'success': False, 'message': 'Bulan tahun tidak boleh kosong'})
        
        try:
            selected_date = datetime.strptime(bulan_tahun, '%Y-%m')
            selected_month_year = selected_date.strftime('%B %Y')
        except ValueError:
            return jsonify({'success': False, 'message': 'Format bulan tahun tidak valid. Gunakan format YYYY-MM'})
        
        # Ambil data arus kas dari tabel arus_kas
        cursor.execute('''
            SELECT total_penjualan, total_biaya_produksi, total_biaya_operasional,
                   total_biaya_non_operasional, kas_bersih_operasional, total_harga_jual_alat,
                   total_harga_alat, kas_bersih_investasi, kas_bersih_pembiayaan, total_kas_bersih
            FROM arus_kas 
            WHERE proposal_id = %s AND bulan_tahun = %s
        ''', (proposal_id, bulan_tahun))
        
        arus_kas_data = cursor.fetchone()
        
        if arus_kas_data:
            # Jika data ada di tabel arus_kas, gunakan data tersebut
            total_penjualan = arus_kas_data['total_penjualan'] or 0
            total_biaya_produksi = arus_kas_data['total_biaya_produksi'] or 0
            total_biaya_operasional = arus_kas_data['total_biaya_operasional'] or 0
            total_biaya_non_operasional = arus_kas_data['total_biaya_non_operasional'] or 0
            kas_bersih_operasional = arus_kas_data['kas_bersih_operasional'] or 0
            total_harga_jual_alat = arus_kas_data['total_harga_jual_alat'] or 0
            total_harga_alat = arus_kas_data['total_harga_alat'] or 0
            kas_bersih_investasi = arus_kas_data['kas_bersih_investasi'] or 0
            kas_bersih_pembiayaan = arus_kas_data['kas_bersih_pembiayaan'] or 0
            total_kas_bersih = arus_kas_data['total_kas_bersih'] or 0
        else:
            # Jika data tidak ada di tabel arus_kas, hitung dari tabel individual
            cursor.execute('''
                SELECT COALESCE(SUM(total_harga), 0) as total_penjualan
                FROM penjualan 
                WHERE proposal_id = %s 
                AND DATE_FORMAT(tanggal_penjualan, '%%Y-%%m') = %s
            ''', (proposal_id, bulan_tahun))
            total_penjualan = cursor.fetchone()['total_penjualan'] or 0
            
            cursor.execute('''
                SELECT COALESCE(SUM(total_harga), 0) as total_biaya_produksi
                FROM bahan_baku 
                WHERE proposal_id = %s 
                AND DATE_FORMAT(tanggal_beli, '%%Y-%%m') = %s
            ''', (proposal_id, bulan_tahun))
            total_biaya_produksi = cursor.fetchone()['total_biaya_produksi'] or 0
            
            cursor.execute('''
                SELECT COALESCE(SUM(total_biaya), 0) as total_biaya_operasional
                FROM biaya_operasional 
                WHERE proposal_id = %s 
                AND DATE_FORMAT(tanggal, '%%Y-%%m') = %s
            ''', (proposal_id, bulan_tahun))
            total_biaya_operasional = cursor.fetchone()['total_biaya_operasional'] or 0
            
            cursor.execute('''
                SELECT COALESCE(SUM(total_biaya), 0) as total_biaya_non_operasional
                FROM biaya_non_operasional 
                WHERE proposal_id = %s 
                AND DATE_FORMAT(tanggal, '%%Y-%%m') = %s
            ''', (proposal_id, bulan_tahun))
            total_biaya_non_operasional = cursor.fetchone()['total_biaya_non_operasional'] or 0
            
            cursor.execute('''
                SELECT 
                    COALESCE(SUM(harga_alat), 0) as total_harga_alat,
                    COALESCE(SUM(harga_jual), 0) as total_harga_jual_alat
                FROM alat_produksi 
                WHERE proposal_id = %s 
                AND DATE_FORMAT(tanggal_beli, '%%Y-%%m') = %s
            ''', (proposal_id, bulan_tahun))
            alat_data = cursor.fetchone()
            total_harga_alat = alat_data['total_harga_alat'] or 0
            total_harga_jual_alat = alat_data['total_harga_jual_alat'] or 0
            
            # Hitung kas bersih
            kas_bersih_operasional = total_penjualan - total_biaya_produksi - total_biaya_operasional - total_biaya_non_operasional
            kas_bersih_investasi = total_harga_jual_alat - total_harga_alat
            kas_bersih_pembiayaan = 0  # Selalu 0 sesuai permintaan
            total_kas_bersih = kas_bersih_operasional + kas_bersih_investasi
        
        # Siapkan data untuk download
        arus_kas_data = {
            'total_penjualan': total_penjualan,
            'total_biaya_produksi': total_biaya_produksi,
            'total_biaya_operasional': total_biaya_operasional,
            'total_biaya_non_operasional': total_biaya_non_operasional,
            'kas_bersih_operasional': kas_bersih_operasional,
            'total_harga_jual_alat': total_harga_jual_alat,
            'total_harga_alat': total_harga_alat,
            'kas_bersih_investasi': kas_bersih_investasi,
            'kas_bersih_pembiayaan': 0,  # Selalu 0 sesuai permintaan
            'total_kas_bersih': total_kas_bersih
        }
        
        print(f"Prepared arus kas data: {arus_kas_data}")
        
        cursor.close()
        
        # Generate filename dengan pembersihan karakter
        import re
        safe_nama = re.sub(r'[<>:"/\\|?*]', '_', mahasiswa_info['nama_ketua'])
        safe_judul = re.sub(r'[<>:"/\\|?*]', '_', mahasiswa_info['judul_usaha'])
        filename = f"Laporan_Arus_Kas_{safe_nama.replace(' ', '_')}_{safe_judul.replace(' ', '_')}_{bulan_tahun}"
        
        print(f"Generated filename: {filename}")
        print(f"Selected month year: {selected_month_year}")
        
        if format_type == 'excel':
            result = get_app_functions()['generate_excel_arus_kas'](arus_kas_data, mahasiswa_info, bulan_tahun, selected_month_year, filename)
            print(f"Excel generation result: {result}")
            return result
        elif format_type == 'pdf':
            result = get_app_functions()['generate_pdf_arus_kas'](arus_kas_data, mahasiswa_info, bulan_tahun, selected_month_year, filename)
            print(f"PDF generation result: {result}")
            return result
        elif format_type == 'word':
            result = get_app_functions()['generate_word_arus_kas'](arus_kas_data, mahasiswa_info, bulan_tahun, selected_month_year, filename)
            print(f"Word generation result: {result}")
            return result
        else:
            return jsonify({'success': False, 'message': 'Format tidak didukung'})
            
    except Exception as e:
        print('--- ERROR DOWNLOAD ARUS KAS ADMIN ---')
        print(traceback.format_exc())
        return jsonify({'success': False, 'message': f'Error: {str(e)}', 'trace': traceback.format_exc()})

@admin_bp.route('/alat_produksi/<int:mahasiswa_id>')
def admin_alat_produksi(mahasiswa_id):
    if 'user_type' not in session or session['user_type'] != 'admin':
        flash('Anda harus login sebagai admin!', 'danger')
        return redirect(url_for('index'))
    
    if not hasattr(get_app_functions()['mysql'], 'connection') or get_app_functions()['mysql'].connection is None:
        flash('Koneksi ke database gagal. Cek konfigurasi database!', 'danger')
        return render_template('admin/alat_produksi.html', alat_produksi_list=[], mahasiswa_info={})
    
    try:
        cursor = get_app_functions()['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Ambil data mahasiswa
        cursor.execute('''
            SELECT m.*, p.judul_usaha, p.id as proposal_id
            FROM mahasiswa m
            JOIN proposal p ON m.nim = p.nim
            WHERE m.id = %s
        ''', (mahasiswa_id,))
        
        mahasiswa_info = cursor.fetchone()
        
        if not mahasiswa_info:
            flash('Data mahasiswa tidak ditemukan!', 'danger')
            cursor.close()
            return redirect(url_for('admin_monitoring_mahasiswa_alat_produksi'))
        
        proposal_id = mahasiswa_info['proposal_id']
        
        # Ambil data alat produksi
        cursor.execute('''
            SELECT * FROM alat_produksi 
            WHERE proposal_id = %s 
            ORDER BY tanggal_beli DESC
        ''', (proposal_id,))
        
        alat_produksi_list = cursor.fetchall()
        cursor.close()
        
        return render_template('admin/alat_produksi.html', 
                             alat_produksi_list=alat_produksi_list, 
                             mahasiswa_info=mahasiswa_info)
        
    except Exception as e:
        flash(f'Error: {str(e)}', 'danger')
        return render_template('admin/alat_produksi.html', alat_produksi_list=[], mahasiswa_info={})

@admin_bp.route('/biaya_operasional/<int:mahasiswa_id>')
def admin_biaya_operasional(mahasiswa_id):
    if 'user_type' not in session or session['user_type'] != 'admin':
        flash('Anda harus login sebagai admin!', 'danger')
        return redirect(url_for('index'))
    
    if not hasattr(get_app_functions()['mysql'], 'connection') or get_app_functions()['mysql'].connection is None:
        flash('Koneksi ke database gagal. Cek konfigurasi database!', 'danger')
        return render_template('admin/biaya_operasional.html', biaya_operasional_list=[], mahasiswa_info={})
    
    try:
        cursor = get_app_functions()['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Ambil data mahasiswa
        cursor.execute('''
            SELECT m.*, p.judul_usaha, p.id as proposal_id
            FROM mahasiswa m
            JOIN proposal p ON m.nim = p.nim
            WHERE m.id = %s
        ''', (mahasiswa_id,))
        
        mahasiswa_info = cursor.fetchone()
        
        if not mahasiswa_info:
            flash('Data mahasiswa tidak ditemukan!', 'danger')
            cursor.close()
            return redirect(url_for('admin_monitoring_mahasiswa_biaya_operasional'))
        
        proposal_id = mahasiswa_info['proposal_id']
        
        # Ambil data biaya operasional
        cursor.execute('''
            SELECT * FROM biaya_operasional 
            WHERE proposal_id = %s 
            ORDER BY created_at DESC
        ''', (proposal_id,))
        
        biaya_operasional_list = cursor.fetchall()
        cursor.close()
        
        return render_template('admin/biaya_operasional.html', 
                             biaya_operasional_list=biaya_operasional_list, 
                             mahasiswa_info=mahasiswa_info)
        
    except Exception as e:
        flash(f'Error: {str(e)}', 'danger')
        return render_template('admin/biaya_operasional.html', biaya_operasional_list=[], mahasiswa_info={})


@admin_bp.route('/laporan_penjualan/<int:mahasiswa_id>')
def admin_laporan_penjualan(mahasiswa_id):
    if 'user_type' not in session or session['user_type'] != 'admin':
        flash('Anda harus login sebagai admin!', 'danger')
        return redirect(url_for('index'))
    
    if not hasattr(get_app_functions()['mysql'], 'connection') or get_app_functions()['mysql'].connection is None:
        flash('Koneksi ke database gagal. Cek konfigurasi database!', 'danger')
        return render_template('admin/laporan_penjualan.html', mahasiswa_info=None, penjualan_list=[])
    
    try:
        cursor = get_app_functions()['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Ambil data mahasiswa
        cursor.execute('''
            SELECT m.*, p.judul_usaha, p.id as proposal_id
            FROM mahasiswa m
            JOIN proposal p ON m.nim = p.nim
            WHERE m.id = %s
        ''', (mahasiswa_id,))
        
        mahasiswa_info = cursor.fetchone()
        if not mahasiswa_info:
            flash('Data mahasiswa tidak ditemukan!', 'danger')
            cursor.close()
            return redirect(url_for('admin_monitoring_mahasiswa_penjualan_produk'))
        
        # Ambil data penjualan
        cursor.execute('''
            SELECT * FROM penjualan 
            WHERE proposal_id = %s
            ORDER BY tanggal_penjualan DESC
        ''', (mahasiswa_info['proposal_id'],))
        
        penjualan_list = cursor.fetchall()
        for penjualan in penjualan_list:
            if penjualan['tanggal_penjualan']:
                try:
                    penjualan['tanggal_penjualan_indo'] = format_tanggal_indonesia(penjualan['tanggal_penjualan'])
                except Exception:
                    penjualan['tanggal_penjualan_indo'] = penjualan['tanggal_penjualan']
            else:
                penjualan['tanggal_penjualan_indo'] = '-'
        cursor.close()
        
        return render_template('admin/laporan_penjualan.html', mahasiswa_info=mahasiswa_info, penjualan_list=penjualan_list)
        
    except Exception as e:
        flash(f'Error: {str(e)}', 'danger')
        return render_template('admin/laporan_penjualan.html', mahasiswa_info=None, penjualan_list=[])


@admin_bp.route('/laporan_laba_rugi/<int:mahasiswa_id>')
def admin_laporan_laba_rugi(mahasiswa_id):
    if 'user_type' not in session or session['user_type'] != 'admin':
        flash('Anda harus login sebagai admin!', 'danger')
        return redirect(url_for('index'))
    
    if not hasattr(get_app_functions()['mysql'], 'connection') or get_app_functions()['mysql'].connection is None:
        flash('Koneksi ke database gagal. Cek konfigurasi database!', 'danger')
        return render_template('admin/laporan_laba_rugi.html', proposal_data={})
    
    try:
        cursor = get_app_functions()['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Ambil data mahasiswa
        cursor.execute('''
            SELECT m.id, m.nama_ketua, m.nim, m.program_studi, p.id as proposal_id, 
                   p.judul_usaha, p.kategori, p.tahun_nib, p.dosen_pembimbing
            FROM mahasiswa m
            JOIN proposal p ON m.nim = p.nim
            WHERE m.id = %s
        ''', (mahasiswa_id,))
        
        mahasiswa_info = cursor.fetchone()
        
        if not mahasiswa_info:
            flash('Data mahasiswa tidak ditemukan!', 'danger')
            return redirect(url_for('admin_monitoring_mahasiswa_laporan_laba_rugi'))
        
        proposal_data = {}
        
        if mahasiswa_info['proposal_id']:
            # Ambil data laba rugi untuk proposal ini
            cursor.execute('''
                SELECT 
                    lr.id,
                    lr.tanggal_produksi,
                    lr.nama_produk,
                    lr.pendapatan as total_pendapatan,
                    lr.total_biaya_produksi,
                    lr.laba_rugi_kotor,
                    lr.biaya_operasional,
                    lr.laba_rugi_bersih
                FROM laba_rugi lr
                WHERE lr.proposal_id = %s
                ORDER BY lr.tanggal_produksi DESC
            ''', (mahasiswa_info['proposal_id'],))
            
            laba_rugi_data = cursor.fetchall()
            
            # Buat objek proposal dengan id yang benar
            proposal_obj = {
                'id': mahasiswa_info['proposal_id'],  # Tambahkan id proposal
                'nama_ketua': mahasiswa_info['nama_ketua'],
                'nim': mahasiswa_info['nim'],
                'program_studi': mahasiswa_info['program_studi'],
                'judul_usaha': mahasiswa_info['judul_usaha'],
                'kategori': mahasiswa_info['kategori'],
                'tahun_nib': mahasiswa_info['tahun_nib'],
                'dosen_pembimbing': mahasiswa_info['dosen_pembimbing']
            }
            
            proposal_data[mahasiswa_info['proposal_id']] = {
                'proposal': proposal_obj,
                'laba_rugi_data': laba_rugi_data
            }
        
        cursor.close()
        return render_template('admin/laporan_laba_rugi.html', proposal_data=proposal_data)
        
    except Exception as e:
        print(f"Error in admin_laporan_laba_rugi: {e}")
        flash('Terjadi kesalahan saat mengambil data laba rugi!', 'danger')
        return render_template('admin/laporan_laba_rugi.html', proposal_data={})


@admin_bp.route('/download_laba_rugi', methods=['POST'])
def admin_download_laba_rugi():
    import traceback
    if 'user_type' not in session or session['user_type'] != 'admin':
        return jsonify({'success': False, 'message': 'Unauthorized'})
    
    try:
        proposal_id = request.form.get('proposal_id')
        format_type = request.form.get('format')
        start_date = request.form.get('start_date')
        end_date = request.form.get('end_date')
        
        print(f"=== DEBUG ADMIN DOWNLOAD LABA RUGI ===")
        print(f"Proposal ID: {proposal_id}")
        print(f"Format: {format_type}")
        print(f"Start Date: {start_date}")
        print(f"End Date: {end_date}")
        print("=====================================")
        
        if not proposal_id or not format_type:
            return jsonify({'success': False, 'message': 'Parameter tidak lengkap'})
        
        if not hasattr(get_app_functions()['mysql'], 'connection') or get_app_functions()['mysql'].connection is None:
            return jsonify({'success': False, 'message': 'Koneksi database tidak tersedia'})
        
        cursor = get_app_functions()['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Ambil data proposal
        cursor.execute('''
            SELECT p.*, m.nama_ketua, m.nim, m.program_studi
            FROM proposal p
            JOIN mahasiswa m ON p.nim = m.nim
            WHERE p.id = %s
        ''', (proposal_id,))
        
        proposal = cursor.fetchone()
        
        print(f"Proposal found: {proposal is not None}")
        if proposal:
            print(f"Proposal data: {proposal}")
        
        if not proposal:
            cursor.close()
            return jsonify({'success': False, 'message': 'Proposal tidak ditemukan'})
        
        # Ambil data laba rugi dengan filter tanggal
        query = '''
            SELECT 
                lr.tanggal_produksi,
                lr.nama_produk,
                lr.pendapatan,
                lr.total_biaya_produksi,
                lr.laba_rugi_kotor,
                lr.biaya_operasional,
                lr.laba_rugi_bersih
            FROM laba_rugi lr
            WHERE lr.proposal_id = %s
        '''
        params = [proposal_id]
        
        if start_date and end_date:
            query += ' AND lr.tanggal_produksi BETWEEN %s AND %s'
            params.extend([start_date, end_date])
        
        query += ' ORDER BY lr.tanggal_produksi DESC'
        
        cursor.execute(query, params)
        laba_rugi_data = cursor.fetchall()
        
        print(f"Laba rugi data count: {len(laba_rugi_data)}")
        if laba_rugi_data:
            print(f"Sample laba rugi data: {laba_rugi_data[:2]}")
        
        if not laba_rugi_data:
            cursor.close()
            return jsonify({'success': False, 'message': 'Tidak ada data laba rugi untuk periode yang dipilih'})
        
        # Hitung total
        total_pendapatan = sum(item['pendapatan'] for item in laba_rugi_data)
        total_biaya_produksi = sum(item['total_biaya_produksi'] for item in laba_rugi_data)
        total_biaya_operasional = sum(item['biaya_operasional'] for item in laba_rugi_data)
        total_laba_kotor = sum(item['laba_rugi_kotor'] for item in laba_rugi_data)
        total_laba_bersih = sum(item['laba_rugi_bersih'] for item in laba_rugi_data)
        
        # Generate filename dengan pembersihan karakter
        import re
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        safe_nama = re.sub(r'[<>:"/\\|?*]', '_', proposal['nama_ketua'])
        filename = f"Laporan_Laba_Rugi_{safe_nama.replace(' ', '_')}_{timestamp}"
        
        if format_type == 'excel':
            result = get_app_functions()['generate_excel_laba_rugi'](
                laba_rugi_data, proposal, total_pendapatan, total_biaya_produksi,
                total_biaya_operasional, total_laba_kotor, total_laba_bersih,
                start_date, end_date, filename
            )
        elif format_type == 'pdf':
            result = get_app_functions()['generate_pdf_laba_rugi'](
                laba_rugi_data, proposal, total_pendapatan, total_biaya_produksi,
                total_biaya_operasional, total_laba_kotor, total_laba_bersih,
                start_date, end_date, filename
            )
        elif format_type == 'word':
            result = get_app_functions()['generate_word_laba_rugi'](
                laba_rugi_data, proposal, total_pendapatan, total_biaya_produksi,
                total_biaya_operasional, total_laba_kotor, total_laba_bersih,
                start_date, end_date, filename
            )
        else:
            cursor.close()
            return jsonify({'success': False, 'message': 'Format tidak didukung'})
        
        cursor.close()
        
        print(f"Generated filename: {filename}")
        print(f"Result type: {type(result)}")
        
        # Fungsi generate sudah mengembalikan jsonify response
        return result
        
    except Exception as e:
        print('--- ERROR ADMIN DOWNLOAD LABA RUGI ---')
        print(traceback.format_exc())
        return jsonify({'success': False, 'message': f'Error: {str(e)}', 'trace': traceback.format_exc()})

@admin_bp.route('/laporan_kemajuan_awal_bertumbuh/<int:mahasiswa_id>')
def admin_laporan_kemajuan_awal_bertumbuh(mahasiswa_id):
    if 'user_type' not in session or session['user_type'] != 'admin':
        flash('Anda harus login sebagai admin!', 'danger')
        
    if not hasattr(get_app_functions()['mysql'], 'connection') or get_app_functions()['mysql'].connection is None:
        flash('Koneksi ke database gagal. Cek konfigurasi database!', 'danger')
        return render_template('admin/laporan_kemajuan_awal_bertumbuh.html', anggaran_data=[], mahasiswa_info=None, proposal_id=None, total_nilai_bantuan=0, file_laporan=None)
        
    try:
        cursor = get_app_functions()['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Ambil data mahasiswa
        cursor.execute('''
            SELECT m.*, p.judul_usaha, p.tahapan_usaha, p.status as status_proposal, p.id as proposal_id
            FROM mahasiswa m 
            LEFT JOIN proposal p ON m.nim = p.nim
            WHERE m.id = %s
        ''', (mahasiswa_id,))
        
        mahasiswa_info = cursor.fetchone()
        if not mahasiswa_info:
            flash('Data mahasiswa tidak ditemukan!', 'danger')
            cursor.close()
            return redirect(url_for('admin_laporan_kemajuan'))
        
        # Tentukan jenis anggaran berdasarkan tahapan usaha
        tahapan_usaha = mahasiswa_info.get('tahapan_usaha', '').lower()
        if 'bertumbuh' in tahapan_usaha:
            tabel_laporan = 'laporan_kemajuan_bertumbuh'
        else:
            tabel_laporan = 'laporan_kemajuan_awal'
        
        proposal_id = mahasiswa_info.get('proposal_id')
        
        if not proposal_id:
            flash('Proposal tidak ditemukan!', 'danger')
            cursor.close()
            return render_template('admin/laporan_kemajuan_awal_bertumbuh.html', 
            anggaran_data=[], mahasiswa_info=mahasiswa_info, proposal_id=None, total_nilai_bantuan=0, file_laporan=None)
        
        # Tentukan tabel anggaran berdasarkan tahapan usaha
        if 'bertumbuh' in tahapan_usaha:
            tabel_anggaran = 'anggaran_bertumbuh'
        else:
            tabel_anggaran = 'anggaran_awal'
        
        # Buat tabel laporan kemajuan jika belum ada
        if tabel_laporan == 'laporan_kemajuan_awal':
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS laporan_kemajuan_awal (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    id_proposal INT NOT NULL,
                    kegiatan_utama VARCHAR(255),
                    kegiatan VARCHAR(255) NOT NULL,
                    penanggung_jawab VARCHAR(100) NOT NULL,
                    target_capaian TEXT NOT NULL,
                    nama_barang VARCHAR(255) NOT NULL,
                    kuantitas INT NOT NULL,
                    satuan VARCHAR(50) NOT NULL,
                    harga_satuan DECIMAL(15,2) NOT NULL,
                    jumlah DECIMAL(15,2) NOT NULL,
                    keterangan TEXT,
                    nilai_bantuan DECIMAL(15,2) DEFAULT 0.00,
                    tanggal_buat TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    status ENUM('draf','diajukan','disetujui','ditolak','revisi') DEFAULT 'draf',
                    tanggal_review TIMESTAMP NULL,
                    FOREIGN KEY (id_proposal) REFERENCES proposal(id) ON DELETE CASCADE
                )
            ''')
        else:
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS laporan_kemajuan_bertumbuh (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    id_proposal INT NOT NULL,
                    kegiatan_utama VARCHAR(255) NOT NULL,
                    kegiatan VARCHAR(255) NOT NULL,
                    penanggung_jawab VARCHAR(100) NOT NULL,
                    target_capaian TEXT NOT NULL,
                    nama_barang VARCHAR(255) NOT NULL,
                    kuantitas INT NOT NULL,
                    satuan VARCHAR(50) NOT NULL,
                    harga_satuan DECIMAL(15,2) NOT NULL,
                    jumlah DECIMAL(15,2) NOT NULL,
                    keterangan TEXT,
                    nilai_bantuan DECIMAL(15,2) DEFAULT 0.00,
                    tanggal_buat TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    status ENUM('draf','diajukan','disetujui','ditolak','revisi') DEFAULT 'draf',
                    tanggal_review TIMESTAMP NULL,
                    FOREIGN KEY (id_proposal) REFERENCES proposal(id) ON DELETE CASCADE
                )
            ''')
        
        # Cek apakah sudah ada data laporan kemajuan untuk proposal ini
        cursor.execute(f'''
            SELECT COUNT(*) as count FROM {tabel_laporan} WHERE id_proposal = %s
        ''', (proposal_id,))
        existing_count = cursor.fetchone()['count']
        
        # Jika belum ada data laporan kemajuan, salin data dari anggaran yang disetujui
        if existing_count == 0:
            cursor.execute(f'''
                SELECT id, kegiatan_utama, kegiatan, penanggung_jawab, target_capaian, 
                       nama_barang, satuan, status, nilai_bantuan
                FROM {tabel_anggaran} 
                WHERE id_proposal = %s AND status = 'disetujui'
                ORDER BY kegiatan_utama, kegiatan, nama_barang
            ''', (proposal_id,))
            
            anggaran_data = cursor.fetchall()
            
            if anggaran_data:
                for anggaran in anggaran_data:
                    cursor.execute(f'''
                        INSERT INTO {tabel_laporan} (
                            id_proposal, kegiatan_utama, kegiatan, penanggung_jawab, target_capaian,
                            nama_barang, kuantitas, satuan, harga_satuan, jumlah, keterangan, status, nilai_bantuan
                        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    ''', (
                        proposal_id, anggaran['kegiatan_utama'], anggaran['kegiatan'], 
                        anggaran['penanggung_jawab'], anggaran['target_capaian'], anggaran['nama_barang'],
                        0, anggaran['satuan'], 0, 
                        0, '', 'draf', anggaran.get('nilai_bantuan', 0)
                    ))
                
                get_app_functions()['mysql'].connection.commit()
        
        # Ambil data laporan kemajuan
        cursor.execute(f'''
            SELECT id, kegiatan_utama, kegiatan, penanggung_jawab, target_capaian, 
                   nama_barang, kuantitas, satuan, harga_satuan, jumlah, keterangan, status, nilai_bantuan
            FROM {tabel_laporan} 
            WHERE id_proposal = %s
            ORDER BY kegiatan_utama, kegiatan, nama_barang
        ''', (proposal_id,))
        
        laporan_data = cursor.fetchall()
        cursor.close()
        
        # Urutkan data sesuai urutan kegiatan utama dengan normalisasi teks
        if 'bertumbuh' in tahapan_usaha:
            urutan_kegiatan = [
                "Pengembangan Pasar dan Saluran Distribusi",
                "Pengembangan Produk/Riset",
                "Produksi",
                "Legalitas, perizinan, sertifikasi, dan standarisasi",
                "Belanja ATK dan Penunjang"
            ]
        else:
            urutan_kegiatan = [
                "Pengembangan Produk/Riset",
                "Produksi",
                "Legalitas, perizinan, sertifikasi, dan standarisasi",
                "Belanja ATK dan Penunjang"
            ]
        
        # Buat mapping untuk case-insensitive matching dan normalisasi teks
        urutan_kegiatan_lower = [k.strip().lower() for k in urutan_kegiatan]
        
        # Fungsi helper untuk normalisasi teks kegiatan utama
        def normalize_kegiatan_utama(text):
            if not text:
                return ""
            # Normalisasi: lowercase, strip whitespace, replace multiple spaces
            normalized = text.strip().lower()
            # Handle berbagai variasi teks
            replacements = {
                "legalitas, perijinan, sertifikasi, pengujian produk, dan standarisasi": "legalitas, perizinan, sertifikasi, dan standarisasi",
                "legalitas, perijinan, sertifikasi, dan standarisasi": "legalitas, perizinan, sertifikasi, dan standarisasi",
                "legalitas perijinan sertifikasi pengujian produk dan standarisasi": "legalitas, perizinan, sertifikasi, dan standarisasi",
                "legalitas perijinan sertifikasi dan standarisasi": "legalitas, perizinan, sertifikasi, dan standarisasi"
            }
            for old, new in replacements.items():
                if old in normalized:
                    normalized = normalized.replace(old, new)
            return normalized
        
        # Sort data berdasarkan urutan kegiatan utama dengan normalisasi
        laporan_data = sorted(
            laporan_data,
            key=lambda x: urutan_kegiatan_lower.index(normalize_kegiatan_utama(x['kegiatan_utama'])) if x['kegiatan_utama'] and normalize_kegiatan_utama(x['kegiatan_utama']) in urutan_kegiatan_lower else 99
        )
        
        grouped_data = group_anggaran_data(laporan_data)
        laporan_data_flat = flatten_anggaran_data(laporan_data)
        
        # Hitung total nilai bantuan dari anggaran yang sudah direview
        cursor = get_app_functions()['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        cursor.execute(f'''
            SELECT SUM(nilai_bantuan) as total_nilai_bantuan
            FROM {tabel_anggaran} 
            WHERE id_proposal = %s AND status_reviewer = 'sudah_direview'
        ''', (proposal_id,))
        
        nilai_bantuan_result = cursor.fetchone()
        total_nilai_bantuan = nilai_bantuan_result['total_nilai_bantuan'] or 0
        
        # Ambil data file laporan kemajuan
        cursor.execute('''
            SELECT nama_file, file_path, status, komentar_pembimbing 
            FROM file_laporan_kemajuan 
            WHERE id_proposal = %s
        ''', (proposal_id,))
        
        file_laporan = cursor.fetchone()
        cursor.close()
        
        return render_template('admin/laporan_kemajuan_awal_bertumbuh.html', 
                             anggaran_data=laporan_data, 
                             grouped_data=grouped_data,
                             anggaran_data_flat=laporan_data_flat,
                             mahasiswa_info=mahasiswa_info,
                             tabel_laporan=tabel_laporan,
                             proposal_id=proposal_id,
                             total_nilai_bantuan=total_nilai_bantuan,
                             file_laporan=file_laporan)
        
    except Exception as e:
        flash(f'Error saat mengambil data laporan kemajuan: {str(e)}', 'danger')
        return render_template('admin/laporan_kemajuan_awal_bertumbuh.html', 
                             anggaran_data=[], 
                             mahasiswa_info=None,
                             proposal_id=None,
                             total_nilai_bantuan=0,
                             file_laporan=None)

@admin_bp.route('/laporan_akhir_awal_bertumbuh/<int:mahasiswa_id>')
def admin_laporan_akhir_awal_bertumbuh(mahasiswa_id):
    if 'user_type' not in session or session['user_type'] != 'admin':
        flash('Anda harus login sebagai admin!', 'danger')
        return redirect(url_for('index'))
    
    if not hasattr(get_app_functions()['mysql'], 'connection') or get_app_functions()['mysql'].connection is None:
        flash('Koneksi ke database gagal. Cek konfigurasi database!', 'danger')
        return render_template('admin/laporan_akhir_awal_bertumbuh.html', 
                             anggaran_data=[], 
                             anggaran_data_flat=[],
                             grouped_data={},
                             mahasiswa_info=None, 
                             proposal_id=None,
                             total_anggaran_disetujui=0,
                             total_nilai_bantuan=0,
                             total_laporan_kemajuan_disetujui=0)
    
    try:
        cursor = get_app_functions()['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Ambil data mahasiswa
        cursor.execute('''
            SELECT m.*, p.judul_usaha, p.tahapan_usaha, p.status as status_proposal, p.id as proposal_id
            FROM mahasiswa m 
            LEFT JOIN proposal p ON m.nim = p.nim
            WHERE m.id = %s
        ''', (mahasiswa_id,))
        
        mahasiswa_info = cursor.fetchone()
        if not mahasiswa_info:
            flash('Data mahasiswa tidak ditemukan!', 'danger')
            cursor.close()
            return render_template('admin/laporan_akhir_awal_bertumbuh.html', 
                                 anggaran_data=[], 
                                 anggaran_data_flat=[],
                                 grouped_data={},
                                 mahasiswa_info=None,
                                 proposal_id=None,
                                 total_anggaran_disetujui=0,
                                 total_nilai_bantuan=0,
                                 total_laporan_kemajuan_disetujui=0)
        
        # Tentukan jenis anggaran berdasarkan tahapan usaha
        tahapan_usaha = mahasiswa_info.get('tahapan_usaha', '').lower()
        if 'bertumbuh' in tahapan_usaha:
            tabel_laporan = 'laporan_akhir_bertumbuh'
        else:
            tabel_laporan = 'laporan_akhir_awal'
        
        proposal_id = mahasiswa_info.get('proposal_id')
        
        if not proposal_id:
            flash('Proposal tidak ditemukan!', 'danger')
            cursor.close()
            return render_template('admin/laporan_akhir_awal_bertumbuh.html', 
                                 anggaran_data=[], 
                                 anggaran_data_flat=[],
                                 grouped_data={},
                                 mahasiswa_info=mahasiswa_info, 
                                 proposal_id=None,
                                 total_anggaran_disetujui=0,
                                 total_nilai_bantuan=0,
                                 total_laporan_kemajuan_disetujui=0)
        
        # Tentukan tabel anggaran berdasarkan tahapan usaha
        if 'bertumbuh' in tahapan_usaha:
            tabel_anggaran = 'anggaran_bertumbuh'
        else:
            tabel_anggaran = 'anggaran_awal'
        
        # Buat tabel laporan akhir jika belum ada
        if tabel_laporan == 'laporan_akhir_awal':
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS laporan_akhir_awal (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    id_proposal INT NOT NULL,
                    kegiatan_utama VARCHAR(255),
                    kegiatan VARCHAR(255) NOT NULL,
                    penanggung_jawab VARCHAR(100) NOT NULL,
                    target_capaian TEXT NOT NULL,
                    nama_barang VARCHAR(255) NOT NULL,
                    kuantitas INT NOT NULL,
                    satuan VARCHAR(50) NOT NULL,
                    harga_satuan DECIMAL(15,2) NOT NULL,
                    jumlah DECIMAL(15,2) NOT NULL,
                    keterangan TEXT,
                    nilai_bantuan DECIMAL(15,2) DEFAULT 0.00,
                    tanggal_buat TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    status ENUM('draf','diajukan','disetujui','ditolak','revisi') DEFAULT 'draf',
                    tanggal_review TIMESTAMP NULL,
                    FOREIGN KEY (id_proposal) REFERENCES proposal(id) ON DELETE CASCADE
                )
            ''')
        else:
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS laporan_akhir_bertumbuh (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    id_proposal INT NOT NULL,
                    kegiatan_utama VARCHAR(255) NOT NULL,
                    kegiatan VARCHAR(255) NOT NULL,
                    penanggung_jawab VARCHAR(100) NOT NULL,
                    target_capaian TEXT NOT NULL,
                    nama_barang VARCHAR(255) NOT NULL,
                    kuantitas INT NOT NULL,
                    satuan VARCHAR(50) NOT NULL,
                    harga_satuan DECIMAL(15,2) NOT NULL,
                    jumlah DECIMAL(15,2) NOT NULL,
                    keterangan TEXT,
                    nilai_bantuan DECIMAL(15,2) DEFAULT 0.00,
                    tanggal_buat TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    status ENUM('draf','diajukan','disetujui','ditolak','revisi') DEFAULT 'draf',
                    tanggal_review TIMESTAMP NULL,
                    FOREIGN KEY (id_proposal) REFERENCES proposal(id) ON DELETE CASCADE
                )
            ''')
        
        get_app_functions()['mysql'].connection.commit()
        
        # Tentukan tabel laporan kemajuan
        if 'bertumbuh' in tahapan_usaha:
            tabel_laporan_kemajuan = 'laporan_kemajuan_bertumbuh'
        else:
            tabel_laporan_kemajuan = 'laporan_kemajuan_awal'
        
        # Ambil data laporan akhir
        cursor.execute(f'''
            SELECT id, kegiatan_utama, kegiatan, penanggung_jawab, target_capaian, 
                   nama_barang, kuantitas, satuan, harga_satuan, jumlah, keterangan, status, nilai_bantuan
            FROM {tabel_laporan} 
            WHERE id_proposal = %s
            ORDER BY kegiatan_utama, kegiatan, nama_barang
        ''', (proposal_id,))
        
        laporan_data = cursor.fetchall()
        print(f"DEBUG: Data laporan akhir untuk proposal {proposal_id}: {len(laporan_data)} rows")
        
        # Debug: Log detail data laporan akhir
        for i, row in enumerate(laporan_data):
            print(f"DEBUG: Row {i+1}: kegiatan_utama='{row.get('kegiatan_utama', 'N/A')}', kegiatan='{row.get('kegiatan', 'N/A')}', nama_barang='{row.get('nama_barang', 'N/A')}', jumlah={row.get('jumlah', 'N/A')}, status='{row.get('status', 'N/A')}")
        
        # Urutkan data sesuai urutan kegiatan utama dengan normalisasi teks
        if 'bertumbuh' in tahapan_usaha:
            urutan_kegiatan = [
                "Pengembangan Pasar dan Saluran Distribusi",
                "Pengembangan Produk/Riset",
                "Produksi",
                "Legalitas, perizinan, sertifikasi, dan standarisasi",
                "Belanja ATK dan Penunjang"
            ]
        else:
            urutan_kegiatan = [
                "Pengembangan Produk/Riset",
                "Produksi",
                "Legalitas, perizinan, sertifikasi, dan standarisasi",
                "Belanja ATK dan Penunjang"
            ]
        
        # Buat mapping untuk case-insensitive matching dan normalisasi teks
        urutan_kegiatan_lower = [k.strip().lower() for k in urutan_kegiatan]
        
        # Fungsi helper untuk normalisasi teks kegiatan utama
        def normalize_kegiatan_utama(text):
            if not text:
                return ""
            # Normalisasi: lowercase, strip whitespace, replace multiple spaces
            normalized = text.strip().lower()
            # Handle berbagai variasi teks
            replacements = {
                "legalitas, perijinan, sertifikasi, pengujian produk, dan standarisasi": "legalitas, perizinan, sertifikasi, dan standarisasi",
                "legalitas, perijinan, sertifikasi, dan standarisasi": "legalitas, perizinan, sertifikasi, dan standarisasi",
                "legalitas perijinan sertifikasi pengujian produk dan standarisasi": "legalitas, perizinan, sertifikasi, dan standarisasi",
                "legalitas perijinan sertifikasi dan standarisasi": "legalitas, perizinan, sertifikasi, dan standarisasi"
            }
            for old, new in replacements.items():
                if old in normalized:
                    normalized = normalized.replace(old, new)
            return normalized
        
        # Sort data berdasarkan urutan kegiatan utama dengan normalisasi
        laporan_data = sorted(
            laporan_data,
            key=lambda x: urutan_kegiatan_lower.index(normalize_kegiatan_utama(x['kegiatan_utama'])) if x['kegiatan_utama'] and normalize_kegiatan_utama(x['kegiatan_utama']) in urutan_kegiatan_lower else 99
        )
        
        grouped_data = group_anggaran_data(laporan_data)
        laporan_data_flat = flatten_anggaran_data(laporan_data)
        
        print(f"DEBUG: Data berhasil di-flatten menjadi {len(laporan_data_flat)} baris")
        
        # Debug: Log detail data yang akan dikirim ke template
        print("=== DEBUG DETAIL DATA LAPORAN AKHIR FLAT ===")
        for i, row in enumerate(laporan_data_flat):
            print(f"Row {i+1}: kegiatan_utama='{row.get('kegiatan_utama', 'N/A')}', kegiatan='{row.get('kegiatan', 'N/A')}', nama_barang='{row.get('nama_barang', 'N/A')}', jumlah={row.get('jumlah', 'N/A')}, status='{row.get('status', 'N/A')}")
        print("============================================")
        
        # Hitung total anggaran yang disetujui
        if 'bertumbuh' in tahapan_usaha:
            tabel_anggaran = 'anggaran_bertumbuh'
        else:
            tabel_anggaran = 'anggaran_awal'
        
        print(f"DEBUG: Menggunakan tabel anggaran: {tabel_anggaran}")
        
        cursor.execute(f'''
            SELECT SUM(jumlah) as total_anggaran
            FROM {tabel_anggaran} 
            WHERE id_proposal = %s AND status = 'disetujui'
        ''', (proposal_id,))
        
        anggaran_result = cursor.fetchone()
        total_anggaran_disetujui = anggaran_result['total_anggaran'] or 0
        print(f"DEBUG: Total anggaran disetujui dari {tabel_anggaran}: {total_anggaran_disetujui}")
        
        # Hitung total nilai bantuan dari anggaran yang sudah direview
        cursor.execute(f'''
            SELECT SUM(nilai_bantuan) as total_nilai_bantuan
            FROM {tabel_anggaran} 
            WHERE id_proposal = %s AND status_reviewer = 'sudah_direview'
        ''', (proposal_id,))
        
        nilai_bantuan_result = cursor.fetchone()
        total_nilai_bantuan = nilai_bantuan_result['total_nilai_bantuan'] or 0
        print(f"DEBUG: Total nilai bantuan dari {tabel_anggaran}: {total_nilai_bantuan}")
        
        # Hitung total laporan kemajuan yang disetujui
        if 'bertumbuh' in tahapan_usaha:
            tabel_laporan_kemajuan = 'laporan_kemajuan_bertumbuh'
        else:
            tabel_laporan_kemajuan = 'laporan_kemajuan_awal'
            
        print(f"DEBUG: Menggunakan tabel laporan kemajuan: {tabel_laporan_kemajuan}")
            
        cursor.execute(f'''
            SELECT SUM(jumlah) as total_laporan_kemajuan
            FROM {tabel_laporan_kemajuan} 
            WHERE id_proposal = %s AND status = 'disetujui'
        ''', (proposal_id,))
        
        laporan_kemajuan_result = cursor.fetchone()
        total_laporan_kemajuan_disetujui = laporan_kemajuan_result['total_laporan_kemajuan'] or 0
        print(f"DEBUG: Total laporan kemajuan disetujui dari {tabel_laporan_kemajuan}: {total_laporan_kemajuan_disetujui}")
        
        # Hitung total laporan akhir (jumlah dari laporan_data)
        total_laporan_akhir = sum(row.get('jumlah', 0) for row in laporan_data) if laporan_data else 0
        print(f"DEBUG: Total laporan akhir (dihitung dari laporan_data): {total_laporan_akhir}")
        
        # Summary log
        print(f"=== SUMMARY LAPORAN AKHIR ADMIN PROPOSAL {proposal_id} ===")
        print(f"Total Nilai Bantuan: {total_nilai_bantuan}")
        print(f"Total Anggaran Disetujui: {total_anggaran_disetujui}")
        print(f"Total Laporan Kemajuan Disetujui: {total_laporan_kemajuan_disetujui}")
        print(f"Total Laporan Akhir: {total_laporan_akhir}")
        print(f"Jumlah Data Laporan: {len(laporan_data)}")
        print(f"Jumlah Data Flat: {len(laporan_data_flat)}")
        print(f"================================================")
        
        # Ambil data file laporan akhir
        cursor.execute('''
            SELECT id, nama_file, file_path, status, komentar_pembimbing, tanggal_upload, tanggal_update
            FROM file_laporan_akhir 
            WHERE id_proposal = %s
        ''', (proposal_id,))
        
        file_laporan_akhir = cursor.fetchone()
        
        cursor.close()
        
        return render_template('admin/laporan_akhir_awal_bertumbuh.html', 
                             anggaran_data=laporan_data, 
                             grouped_data=grouped_data,
                             anggaran_data_flat=laporan_data_flat,
                             mahasiswa_info=mahasiswa_info,
                             tabel_laporan=tabel_laporan,
                             proposal_id=proposal_id,
                             file_laporan_akhir=file_laporan_akhir,
                             total_anggaran_disetujui=total_anggaran_disetujui,
                             total_nilai_bantuan=total_nilai_bantuan,
                             total_laporan_kemajuan_disetujui=total_laporan_kemajuan_disetujui,
                             total_laporan_akhir=total_laporan_akhir)
        
    except Exception as e:
        print(f"Error in admin_laporan_akhir_awal_bertumbuh: {str(e)}")
        flash(f'Error saat mengambil data laporan akhir: {str(e)}', 'danger')
        return render_template('admin/laporan_akhir_awal_bertumbuh.html', 
                             anggaran_data=[], 
                             anggaran_data_flat=[],
                             grouped_data={},
                             mahasiswa_info=None,
                             proposal_id=None,
                             total_anggaran_disetujui=0,
                             total_nilai_bantuan=0,
                             total_laporan_kemajuan_disetujui=0)

@admin_bp.route('/data_admin')
def data_admin():
    if 'user_type' not in session or session['user_type'] != 'admin':
        flash('Anda harus login sebagai admin!', 'danger')
        return redirect(url_for('index'))
    if not hasattr(get_app_functions()['mysql'], 'connection') or get_app_functions()['mysql'].connection is None:
        flash('Koneksi ke database gagal. Cek konfigurasi database!', 'danger')
        return render_template('admin/admin.html', admin_list=[])
    try:
        cursor = get_app_functions()['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        cursor.execute('SELECT id, nama, nip, program_studi FROM admin ORDER BY nama')
        admin_list = cursor.fetchall()
        
        # Ambil data program studi untuk dropdown
        cursor.execute('SELECT id, nama_program_studi FROM program_studi ORDER BY nama_program_studi')
        program_studi_list = cursor.fetchall()
        
        cursor.close()
        return render_template('admin/admin.html', admin_list=admin_list, program_studi_list=program_studi_list)
    except Exception as e:
        flash(f'Error saat mengambil data admin: {str(e)}', 'danger')
        return render_template('admin/admin.html', admin_list=[])

@admin_bp.route('/get_admin_detail/<int:admin_id>')
def get_admin_detail(admin_id):
    if 'user_type' not in session or session['user_type'] != 'admin':
        return jsonify({'success': False, 'message': 'Anda harus login sebagai admin!'})
    if not hasattr(get_app_functions()['mysql'], 'connection') or get_app_functions()['mysql'].connection is None:
        return jsonify({'success': False, 'message': 'Koneksi ke database gagal!'})
    try:
        cursor = get_app_functions()['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        cursor.execute('SELECT id, nama, nip, program_studi FROM admin WHERE id = %s', (admin_id,))
        admin = cursor.fetchone()
        cursor.close()
        if admin:
            return jsonify({'success': True, 'admin': admin})
        else:
            return jsonify({'success': False, 'message': 'Data admin tidak ditemukan!'})
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@admin_bp.route('/add_admin', methods=['POST'])
def add_admin():
    if 'user_type' not in session or session['user_type'] != 'admin':
        return jsonify({'success': False, 'message': 'Anda harus login sebagai admin!'})
    if not hasattr(get_app_functions()['mysql'], 'connection') or get_app_functions()['mysql'].connection is None:
        return jsonify({'success': False, 'message': 'Koneksi ke database gagal!'})
    data = request.get_json()
    nama = data.get('nama')
    nip = data.get('nip')
    program_studi = data.get('program_studi')
    password = data.get('password')
    if not all([nama, nip, program_studi, password]):
        return jsonify({'success': False, 'message': 'Semua field wajib diisi!'})
    try:
        cursor = get_app_functions()['mysql'].connection.cursor()
        
        # Cek duplikasi komprehensif
        is_duplicate, message, table_name = check_comprehensive_duplicates(
            cursor, 
            nama=nama, 
            nip=nip
        )
        if is_duplicate:
            cursor.close()
            return jsonify({'success': False, 'message': message})
        
        from werkzeug.security import generate_password_hash
        password_hash = generate_password_hash(password)
        cursor.execute('INSERT INTO admin (nama, nip, program_studi, password) VALUES (%s, %s, %s, %s)', (nama, nip, program_studi, password_hash))
        get_app_functions()['mysql'].connection.commit()
        cursor.close()
        return jsonify({'success': True, 'message': 'Admin berhasil ditambahkan!'})
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@admin_bp.route('/update_admin_data', methods=['POST'])
def update_admin_data():
    if 'user_type' not in session or session['user_type'] != 'admin':
        return jsonify({'success': False, 'message': 'Anda harus login sebagai admin!'})
    if not hasattr(get_app_functions()['mysql'], 'connection') or get_app_functions()['mysql'].connection is None:
        return jsonify({'success': False, 'message': 'Koneksi ke database gagal!'})
    data = request.get_json()
    admin_id = data.get('id')
    nama = data.get('nama')
    nip = data.get('nip')
    program_studi = data.get('program_studi')
    password = data.get('password')
    if not all([admin_id, nama, nip, program_studi]):
        return jsonify({'success': False, 'message': 'Semua field wajib diisi!'})
    try:
        cursor = get_app_functions()['mysql'].connection.cursor()
        
        # Cek duplikasi komprehensif (exclude admin yang sedang diedit)
        is_duplicate, message, table_name = check_comprehensive_duplicates(
            cursor, 
            nama=nama, 
            nip=nip,
            exclude_id=admin_id,
            exclude_table='admin'
        )
        if is_duplicate:
            cursor.close()
            return jsonify({'success': False, 'message': message})
        
        if password:
            from werkzeug.security import generate_password_hash
            password_hash = generate_password_hash(password)
            cursor.execute('UPDATE admin SET nama=%s, nip=%s, program_studi=%s, password=%s WHERE id=%s', (nama, nip, program_studi, password_hash, admin_id))
        else:
            cursor.execute('UPDATE admin SET nama=%s, nip=%s, program_studi=%s WHERE id=%s', (nama, nip, program_studi, admin_id))
        get_app_functions()['mysql'].connection.commit()
        cursor.close()
        return jsonify({'success': True, 'message': 'Data admin berhasil diperbarui!'})
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

@admin_bp.route('/delete_admin', methods=['POST'])
def delete_admin():
    if 'user_type' not in session or session['user_type'] != 'admin':
        return jsonify({'success': False, 'message': 'Anda harus login sebagai admin!'})
    if not hasattr(get_app_functions()['mysql'], 'connection') or get_app_functions()['mysql'].connection is None:
        return jsonify({'success': False, 'message': 'Koneksi ke database gagal!'})
    data = request.get_json()
    admin_id = data.get('admin_id')
    if not admin_id:
        return jsonify({'success': False, 'message': 'ID admin tidak ditemukan!'})
    try:
        cursor = get_app_functions()['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        cursor.execute('SELECT nama FROM admin WHERE id = %s', (admin_id,))
        admin = cursor.fetchone()
        if not admin:
            cursor.close()
            return jsonify({'success': False, 'message': 'Data admin tidak ditemukan!'})
        if admin['nama'] == session.get('nama'):
            cursor.close()
            return jsonify({'success': False, 'message': 'Anda tidak dapat menghapus akun Anda sendiri!'})
        cursor.execute('DELETE FROM admin WHERE id = %s', (admin_id,))
        get_app_functions()['mysql'].connection.commit()
        cursor.close()
        return jsonify({'success': True, 'message': 'Admin berhasil dihapus!'})
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})


@admin_bp.route('/daftar_log_mahasiswa')
def admin_daftar_log_mahasiswa():
    if 'user_type' not in session or session['user_type'] != 'admin':
        flash('Anda harus login sebagai admin!', 'danger')
        return redirect(url_for('index'))
    
    if not hasattr(get_app_functions()['mysql'], 'connection') or get_app_functions()['mysql'].connection is None:
        flash('Koneksi ke database gagal. Cek konfigurasi database!', 'danger')
        return render_template('admin/daftar_log_ativitas_mahasiswa.html', mahasiswa_list=[])
    
    try:
        cursor = get_app_functions()['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Ambil data mahasiswa dengan informasi proposal
        cursor.execute('''
            SELECT m.id, m.nama_ketua, m.nim, m.program_studi, m.perguruan_tinggi, 
                   m.status, p.status_admin as proposal_status
            FROM mahasiswa m
            LEFT JOIN proposal p ON m.nim = p.nim
            ORDER BY m.nama_ketua ASC
        ''')
        
        mahasiswa_list = cursor.fetchall()
        cursor.close()
        
        return render_template('admin/daftar_log_ativitas_mahasiswa.html', mahasiswa_list=mahasiswa_list)
        
    except Exception as e:
        flash(f'Error saat mengambil data mahasiswa: {str(e)}', 'danger')
        return render_template('admin/daftar_log_ativitas_mahasiswa.html', mahasiswa_list=[])

@admin_bp.route('/daftar_log_pembimbing')
def admin_daftar_log_pembimbing():
    if 'user_type' not in session or session['user_type'] != 'admin':
        flash('Anda harus login sebagai admin!', 'danger')
        return redirect(url_for('index'))
    
    if not hasattr(get_app_functions()['mysql'], 'connection') or get_app_functions()['mysql'].connection is None:
        flash('Koneksi ke database gagal. Cek konfigurasi database!', 'danger')
        return render_template('admin/daftar_log_aktivitas_pembimbing.html', pembimbing_list=[])
    
    try:
        cursor = get_app_functions()['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Ambil data pembimbing dengan informasi bimbingan
        cursor.execute('''
            SELECT 
                p.id, p.nama, p.nip, p.program_studi, p.status, p.tanggal_dibuat, p.kuota_mahasiswa,
                COUNT(DISTINCT CASE WHEN pr.status != 'selesai' THEN pr.nim END) as jumlah_mahasiswa_bimbing
            FROM pembimbing p
            LEFT JOIN proposal pr ON p.nama = pr.dosen_pembimbing
            GROUP BY p.id, p.nama, p.nip, p.program_studi, p.status, p.tanggal_dibuat, p.kuota_mahasiswa
            ORDER BY p.nama ASC
        ''')
        
        pembimbing_list = cursor.fetchall()
        print(f"DEBUG: Ditemukan {len(pembimbing_list)} data pembimbing")
        for p in pembimbing_list:
            print(f"DEBUG: Pembimbing - ID: {p['id']}, Nama: {p['nama']}, Status: {p['status']}")
        
        cursor.close()
        
        return render_template('admin/daftar_log_aktivitas_pembimbing.html', 
                             pembimbing_list=pembimbing_list)
        
    except Exception as e:
        flash(f'Error saat mengambil data pembimbing: {str(e)}', 'danger')
        return render_template('admin/daftar_log_aktivitas_pembimbing.html', pembimbing_list=[])

@admin_bp.route('/log_pembimbing_detail/<int:pembimbing_id>')
def admin_log_pembimbing_detail(pembimbing_id):
    if 'user_type' not in session or session['user_type'] != 'admin':
        flash('Anda harus login sebagai admin!', 'danger')
        return redirect(url_for('index'))
    
    if not hasattr(get_app_functions()['mysql'], 'connection') or get_app_functions()['mysql'].connection is None:
        flash('Koneksi ke database gagal. Cek konfigurasi database!', 'danger')
        return render_template('admin/log_pembimbing.html', 
                             pembimbing_info=None, 
                             log_aktivitas=[], 
                             durasi_total=None, 
                             durasi_hari_ini=None)
    
    try:
        cursor = get_app_functions()['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Ambil informasi pembimbing
        cursor.execute('''
            SELECT id, nama, nip, program_studi, status, tanggal_dibuat, kuota_mahasiswa
            FROM pembimbing
            WHERE id = %s
        ''', (pembimbing_id,))
        
        pembimbing_info = cursor.fetchone()
        
        print(f"DEBUG: Info pembimbing ID {pembimbing_id}: {pembimbing_info}")
        
        if not pembimbing_info:
            flash('Data pembimbing tidak ditemukan!', 'danger')
            return redirect(url_for('admin_daftar_log_pembimbing'))
        
        # Hitung jumlah mahasiswa yang sedang dibimbing
        cursor.execute('''
            SELECT COUNT(DISTINCT p.nim) as jumlah_mahasiswa_bimbing,
                   GROUP_CONCAT(DISTINCT m.nama_ketua SEPARATOR ', ') as nama_mahasiswa
            FROM proposal p
            JOIN mahasiswa m ON p.nim = m.nim
            WHERE p.dosen_pembimbing = %s 
            AND p.status != 'selesai'
        ''', (pembimbing_info['nama'],))
        
        bimbingan_result = cursor.fetchone()
        jumlah_mahasiswa_bimbing = bimbingan_result['jumlah_mahasiswa_bimbing'] if bimbingan_result else 0
        nama_mahasiswa_bimbing = bimbingan_result['nama_mahasiswa'] if bimbingan_result and bimbingan_result['nama_mahasiswa'] else ''
        
        # Hitung persentase kuota yang terpakai
        kuota_terpakai = (jumlah_mahasiswa_bimbing / pembimbing_info['kuota_mahasiswa']) * 100 if pembimbing_info['kuota_mahasiswa'] > 0 else 0
        sisa_kuota = pembimbing_info['kuota_mahasiswa'] - jumlah_mahasiswa_bimbing
        
        print(f"DEBUG: Pembimbing {pembimbing_info['nama']} membimbing {jumlah_mahasiswa_bimbing}/{pembimbing_info['kuota_mahasiswa']} mahasiswa ({kuota_terpakai:.1f}%)")
        
        # Ambil data log aktivitas pembimbing dari database
        cursor.execute('''
            SELECT jenis_aktivitas, modul, detail_modul, deskripsi, created_at
            FROM log_aktivitas_pembimbing
            WHERE pembimbing_id = %s
            ORDER BY created_at DESC
            LIMIT 100
        ''', (pembimbing_id,))
        
        log_aktivitas = cursor.fetchall()
        
        print(f"DEBUG: Ditemukan {len(log_aktivitas)} log aktivitas untuk pembimbing ID {pembimbing_id}")
        
        # Hitung durasi total aktivitas pembimbing
        cursor.execute('''
            SELECT 
                COALESCE(SUM(durasi_detik), 0) as total_detik,
                COUNT(*) as total_session
            FROM session_pembimbing 
            WHERE pembimbing_id = %s AND status = 'ended'
        ''', (pembimbing_id,))
        
        durasi_result = cursor.fetchone()
        total_detik = durasi_result['total_detik'] if durasi_result else 0
        
        # Konversi ke hari, jam, menit, detik
        hari = total_detik // 86400
        jam = (total_detik % 86400) // 3600
        menit = (total_detik % 3600) // 60
        detik = total_detik % 60
        
        durasi_total = {
            'hari': hari,
            'jam': jam,
            'menit': menit,
            'detik': detik
        }
        
        # Hitung durasi aktivitas hari ini
        cursor.execute('''
            SELECT 
                COALESCE(SUM(durasi_detik), 0) as total_detik_hari_ini
            FROM session_pembimbing 
            WHERE pembimbing_id = %s 
            AND DATE(login_time) = CURDATE()
            AND status = 'ended'
        ''', (pembimbing_id,))
        
        durasi_hari_ini_result = cursor.fetchone()
        total_detik_hari_ini = durasi_hari_ini_result['total_detik_hari_ini'] if durasi_hari_ini_result else 0
        
        # Konversi ke jam, menit, detik (tidak perlu hari untuk hari ini)
        jam_hari_ini = total_detik_hari_ini // 3600
        menit_hari_ini = (total_detik_hari_ini % 3600) // 60
        detik_hari_ini = total_detik_hari_ini % 60
        
        durasi_hari_ini = {
            'jam': jam_hari_ini,
            'menit': menit_hari_ini,
            'detik': detik_hari_ini
        }
        
        cursor.close()
        
        return render_template('admin/log_pembimbing.html', 
                             pembimbing_info=pembimbing_info,
                             log_aktivitas=log_aktivitas,
                             durasi_total=durasi_total,
                             durasi_hari_ini=durasi_hari_ini,
                             jumlah_mahasiswa_bimbing=jumlah_mahasiswa_bimbing,
                             nama_mahasiswa_bimbing=nama_mahasiswa_bimbing,
                             kuota_terpakai=kuota_terpakai,
                             sisa_kuota=sisa_kuota)
        
    except Exception as e:
        flash(f'Error saat mengambil data log pembimbing: {str(e)}', 'danger')
        return render_template('admin/log_pembimbing.html', 
                             pembimbing_info=None, 
                             log_aktivitas=[], 
                             durasi_total=None, 
                             durasi_hari_ini=None)

@admin_bp.route('/clear_pembimbing_logs', methods=['POST'])
def clear_pembimbing_logs():
    """Hapus semua log aktivitas dan session pembimbing tertentu"""
    if 'user_type' not in session or session['user_type'] != 'admin':
        return jsonify({'success': False, 'message': 'Anda harus login sebagai admin!'})
    
    try:
        pembimbing_id = request.form.get('pembimbing_id')
        if not pembimbing_id:
            return jsonify({'success': False, 'message': 'ID Pembimbing tidak ditemukan!'})
        
        cursor = get_app_functions()['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Verifikasi pembimbing ada
        cursor.execute('SELECT id, nama FROM pembimbing WHERE id = %s', (pembimbing_id,))
        pembimbing = cursor.fetchone()
        
        if not pembimbing:
            return jsonify({'success': False, 'message': 'Data pembimbing tidak ditemukan!'})
        
        # Hapus log aktivitas pembimbing
        cursor.execute('DELETE FROM log_aktivitas_pembimbing WHERE pembimbing_id = %s', (pembimbing_id,))
        log_deleted = cursor.rowcount
        
        # Hapus session pembimbing
        cursor.execute('DELETE FROM session_pembimbing WHERE pembimbing_id = %s', (pembimbing_id,))
        session_deleted = cursor.rowcount
        
        # Commit perubahan
        get_app_functions()['mysql'].connection.commit()
        cursor.close()
        
        # Log aktivitas admin
        logger.info(f"Admin {session.get('username', 'Unknown')} menghapus {log_deleted} log aktivitas dan {session_deleted} session untuk pembimbing {pembimbing['nama']} (ID: {pembimbing_id})")
        
        return jsonify({
            'success': True, 
            'message': f'Berhasil menghapus {log_deleted} log aktivitas dan {session_deleted} session untuk pembimbing {pembimbing["nama"]}'
        })
        
    except Exception as e:
        logger.error(f"Error saat menghapus log pembimbing: {str(e)}")
        return jsonify({'success': False, 'message': f'Terjadi kesalahan: {str(e)}'})

@admin_bp.route('/log_mahasiswa_detail/<int:mahasiswa_id>')
def admin_log_mahasiswa_detail(mahasiswa_id):
    if 'user_type' not in session or session['user_type'] != 'admin':
        flash('Anda harus login sebagai admin!', 'danger')
        return redirect(url_for('index'))
    
    if not hasattr(get_app_functions()['mysql'], 'connection') or get_app_functions()['mysql'].connection is None:
        flash('Koneksi ke database gagal. Cek konfigurasi database!', 'danger')
        return render_template('admin/log_mahasiswa.html', 
                             mahasiswa_info=None, 
                             log_aktivitas=[], 
                             durasi_total=None, 
                             durasi_hari_ini=None)
    
    try:
        cursor = get_app_functions()['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Ambil informasi mahasiswa dengan pembimbing
        cursor.execute('''
            SELECT m.*, p.dosen_pembimbing as pembimbing
            FROM mahasiswa m
            LEFT JOIN proposal p ON m.nim = p.nim
            WHERE m.id = %s
            LIMIT 1
        ''', (mahasiswa_id,))
        
        mahasiswa_info = cursor.fetchone()
        
        if not mahasiswa_info:
            flash('Data mahasiswa tidak ditemukan!', 'danger')
            return redirect(url_for('admin_daftar_log_mahasiswa'))
        
        # Ambil data log aktivitas dari database (semua data untuk DataTable)
        cursor.execute('''
            SELECT jenis_aktivitas, modul, detail_modul, deskripsi, created_at
            FROM log_aktivitas_mahasiswa
            WHERE mahasiswa_id = %s
            ORDER BY created_at DESC
        ''', (mahasiswa_id,))
        
        log_aktivitas = cursor.fetchall()
        
        # Debug: Print log aktivitas untuk troubleshooting
        print(f"DEBUG: Found {len(log_aktivitas)} log activities for mahasiswa_id {mahasiswa_id}")
        for i, log in enumerate(log_aktivitas[:3]):  # Print first 3 logs
            print(f"DEBUG: Log {i+1}: {log}")
        
        # Hitung durasi total aktivitas mahasiswa
        cursor.execute('''
            SELECT 
                COALESCE(SUM(durasi_detik), 0) as total_detik,
                COUNT(*) as total_session
            FROM session_mahasiswa 
            WHERE mahasiswa_id = %s AND status = 'ended'
        ''', (mahasiswa_id,))
        
        durasi_data = cursor.fetchone()
        total_detik = durasi_data['total_detik'] if durasi_data else 0
        
        # Convert total detik ke hari, jam, menit, detik
        hari = total_detik // 86400
        jam = (total_detik % 86400) // 3600
        menit = (total_detik % 3600) // 60
        detik = total_detik % 60
        
        durasi_total = {
            'hari': hari,
            'jam': jam,
            'menit': menit,
            'detik': detik
        }
        
        # Hitung durasi hari ini
        cursor.execute('''
            SELECT COALESCE(SUM(durasi_detik), 0) as total_detik_hari_ini
            FROM session_mahasiswa 
            WHERE mahasiswa_id = %s 
            AND DATE(login_time) = CURDATE()
            AND status = 'ended'
        ''', (mahasiswa_id,))
        
        durasi_hari_ini_data = cursor.fetchone()
        total_detik_hari_ini = durasi_hari_ini_data['total_detik_hari_ini'] if durasi_hari_ini_data else 0
        
        # Convert detik hari ini ke jam, menit, detik
        jam_hari_ini = total_detik_hari_ini // 3600
        menit_hari_ini = (total_detik_hari_ini % 3600) // 60
        detik_hari_ini = total_detik_hari_ini % 60
        
        durasi_hari_ini = {
            'jam': jam_hari_ini,
            'menit': menit_hari_ini,
            'detik': detik_hari_ini
        }
        
        cursor.close()
        
        return render_template('admin/log_mahasiswa.html', 
                             mahasiswa_info=mahasiswa_info,
                             log_aktivitas=log_aktivitas,
                             durasi_total=durasi_total,
                             durasi_hari_ini=durasi_hari_ini)
        
    except Exception as e:
        flash(f'Error saat mengambil data log: {str(e)}', 'danger')
        return render_template('admin/log_mahasiswa.html', 
                             mahasiswa_info=None, 
                             log_aktivitas=[], 
                             durasi_total=None, 
                             durasi_hari_ini=None)

@admin_bp.route('/clear_mahasiswa_logs', methods=['POST'])
def clear_mahasiswa_logs():
    """Hapus semua log aktivitas dan session mahasiswa tertentu"""
    if 'user_type' not in session or session['user_type'] != 'admin':
        return jsonify({'success': False, 'message': 'Anda harus login sebagai admin!'})
    
    try:
        mahasiswa_id = request.form.get('mahasiswa_id')
        if not mahasiswa_id:
            return jsonify({'success': False, 'message': 'ID Mahasiswa tidak ditemukan!'})
        
        cursor = get_app_functions()['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Verifikasi mahasiswa ada
        cursor.execute('SELECT id, nama_ketua FROM mahasiswa WHERE id = %s', (mahasiswa_id,))
        mahasiswa = cursor.fetchone()
        
        if not mahasiswa:
            return jsonify({'success': False, 'message': 'Data mahasiswa tidak ditemukan!'})
        
        # Hapus log aktivitas mahasiswa
        cursor.execute('DELETE FROM log_aktivitas_mahasiswa WHERE mahasiswa_id = %s', (mahasiswa_id,))
        log_deleted = cursor.rowcount
        
        # Hapus session mahasiswa
        cursor.execute('DELETE FROM session_mahasiswa WHERE mahasiswa_id = %s', (mahasiswa_id,))
        session_deleted = cursor.rowcount
        
        # Commit perubahan
        get_app_functions()['mysql'].connection.commit()
        cursor.close()
        
        # Log aktivitas admin
        logger.info(f"Admin {session.get('username', 'Unknown')} menghapus {log_deleted} log aktivitas dan {session_deleted} session untuk mahasiswa {mahasiswa['nama_ketua']} (ID: {mahasiswa_id})")
        
        return jsonify({
            'success': True, 
            'message': f'Berhasil menghapus {log_deleted} log aktivitas dan {session_deleted} session untuk mahasiswa {mahasiswa["nama_ketua"]}'
        })
        
    except Exception as e:
        logger.error(f"Error saat menghapus log mahasiswa: {str(e)}")
        return jsonify({'success': False, 'message': f'Terjadi kesalahan: {str(e)}'})

@admin_bp.route('/monitoring_mahasiswa/laporan_neraca')
def admin_monitoring_mahasiswa_laporan_neraca():
    if 'user_type' not in session or session['user_type'] != 'admin':
        flash('Anda harus login sebagai admin!', 'danger')
        return redirect(url_for('index'))
    
    if not hasattr(get_app_functions()['mysql'], 'connection') or get_app_functions()['mysql'].connection is None:
        flash('Koneksi ke database gagal. Cek konfigurasi database!', 'danger')
        return render_template('admin/daftar_laporan_neraca.html', mahasiswa_list=[])
    
    try:
        cursor = get_app_functions()['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Ambil semua mahasiswa (tanpa filter pembimbing untuk admin)
        cursor.execute('''
            SELECT m.id as mahasiswa_id, m.nama_ketua, m.nim, p.judul_usaha, m.program_studi, p.id as proposal_id, p.dosen_pembimbing, p.kategori, p.tahapan_usaha
            FROM mahasiswa m
            JOIN proposal p ON m.nim = p.nim
            ORDER BY m.nama_ketua
        ''')
        
        mahasiswa_all = cursor.fetchall()
        mahasiswa_list = []
        
        for mhs in mahasiswa_all:
            proposal_id = mhs['proposal_id']
            
            # Cek apakah ada data neraca dari tabel laporan_neraca
            try:
                cursor.execute('SELECT COUNT(*) as cnt FROM laporan_neraca WHERE proposal_id = %s', (proposal_id,))
                neraca_count = cursor.fetchone()['cnt']
            except:
                # Jika tabel tidak ada, cek dari data lain
                cursor.execute('SELECT COUNT(*) as cnt FROM penjualan WHERE proposal_id = %s', (proposal_id,))
                penjualan_count = cursor.fetchone()['cnt']
                cursor.execute('SELECT COUNT(*) as cnt FROM biaya_operasional WHERE proposal_id = %s', (proposal_id,))
                biaya_operasional_count = cursor.fetchone()['cnt']
                neraca_count = 1 if (penjualan_count > 0 or biaya_operasional_count > 0) else 0
            
            # Tambahkan informasi data yang tersedia
            mhs['has_neraca'] = neraca_count > 0
            
            mahasiswa_list.append(mhs)
        
        cursor.close()
        
        return render_template('admin/daftar_laporan_neraca.html', mahasiswa_list=mahasiswa_list)
        
    except Exception as e:
        print(f"Error in admin_monitoring_mahasiswa_laporan_neraca: {e}")
        flash('Terjadi kesalahan saat mengambil data mahasiswa!', 'danger')
        return render_template('admin/daftar_laporan_neraca.html', mahasiswa_list=[])

@admin_bp.route('/laporan_neraca/<int:mahasiswa_id>')
def admin_laporan_neraca(mahasiswa_id):
    if 'user_type' not in session or session['user_type'] != 'admin':
        flash('Anda harus login sebagai admin!', 'danger')
        return redirect(url_for('index'))
    
    if not hasattr(get_app_functions()['mysql'], 'connection') or get_app_functions()['mysql'].connection is None:
        flash('Koneksi ke database gagal. Cek konfigurasi database!', 'danger')
        return render_template('admin/laporan_neraca.html', mahasiswa_info=None, neraca_data={})
    
    try:
        cursor = get_app_functions()['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Ambil informasi mahasiswa
        cursor.execute('''
            SELECT m.id, m.nama_ketua, m.nim, m.program_studi, p.judul_usaha, p.id as proposal_id, p.dosen_pembimbing, p.kategori, p.tahapan_usaha
            FROM mahasiswa m
            JOIN proposal p ON m.nim = p.nim
            WHERE m.id = %s
        ''', (mahasiswa_id,))
        
        mahasiswa_info = cursor.fetchone()
        
        if not mahasiswa_info:
            flash('Data mahasiswa tidak ditemukan!', 'danger')
            return redirect(url_for('admin.admin_monitoring_mahasiswa_laporan_neraca'))
        
        # Hitung laporan neraca secara real-time dari data transaksi
        app_funcs = get_app_functions()
        neraca_data = app_funcs['hitung_neraca_real_time'](mahasiswa_info['proposal_id'], cursor)
        
        cursor.close()
        
        return render_template('admin/laporan_neraca.html', 
                             mahasiswa_info=mahasiswa_info, 
                             neraca_data=neraca_data)
        
    except Exception as e:
        print(f"Error in admin_laporan_neraca: {e}")
        flash('Terjadi kesalahan saat mengambil data laporan neraca!', 'danger')
        return render_template('admin/laporan_neraca.html', mahasiswa_info=None, neraca_data={})

@admin_bp.route('/download_neraca', methods=['POST'])
def admin_download_neraca():
    if 'user_type' not in session or session['user_type'] != 'admin':
        flash('Anda harus login sebagai admin!', 'danger')
        return redirect(url_for('admin.admin_monitoring_mahasiswa_laporan_neraca'))
    
    app_funcs = get_app_functions()
    if not hasattr(app_funcs['mysql'], 'connection') or app_funcs['mysql'].connection is None:
        flash('Koneksi ke database gagal. Cek konfigurasi database!', 'danger')
        return redirect(url_for('admin.admin_monitoring_mahasiswa_laporan_neraca'))
    
    try:
        format_type = request.form.get('format', 'excel')
        proposal_id = request.form.get('proposal_id')
        
        if not proposal_id:
            flash('ID proposal tidak ditemukan!', 'danger')
            return redirect(url_for('admin.admin_monitoring_mahasiswa_laporan_neraca'))
        
        cursor = app_funcs['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Ambil data mahasiswa
        cursor.execute('''
            SELECT m.*, p.judul_usaha, p.id as proposal_id, p.tahapan_usaha
            FROM mahasiswa m
            JOIN proposal p ON m.nim = p.nim
            WHERE p.id = %s
        ''', (proposal_id,))
        
        mahasiswa_info = cursor.fetchone()
        if not mahasiswa_info:
            flash('Data mahasiswa tidak ditemukan!', 'danger')
            cursor.close()
            return redirect(url_for('admin.admin_monitoring_mahasiswa_laporan_neraca'))
        
        # Hitung laporan neraca secara real-time
        neraca_data = app_funcs['hitung_neraca_real_time'](proposal_id, cursor)
        
        cursor.close()
        
        # Generate file berdasarkan format
        if format_type == 'excel':
            file_buffer = app_funcs['generate_excel_neraca'](mahasiswa_info, neraca_data)
            if file_buffer:
                filename = f"Laporan_Neraca_{mahasiswa_info['nama_ketua']}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                return send_file(
                    file_buffer,
                    as_attachment=True,
                    download_name=filename,
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
            else:
                flash('Gagal membuat file Excel!', 'danger')
                return redirect(url_for('admin.admin_monitoring_mahasiswa_laporan_neraca'))
        elif format_type == 'pdf':
            file_buffer = app_funcs['generate_pdf_neraca'](mahasiswa_info, neraca_data)
            if file_buffer:
                filename = f"Laporan_Neraca_{mahasiswa_info['nama_ketua']}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
                return send_file(
                    file_buffer,
                    as_attachment=True,
                    download_name=filename,
                    mimetype='application/pdf'
                )
            else:
                flash('Gagal membuat file PDF!', 'danger')
                return redirect(url_for('admin.admin_monitoring_mahasiswa_laporan_neraca'))
        elif format_type == 'word':
            file_buffer = app_funcs['generate_word_neraca'](mahasiswa_info, neraca_data)
            if file_buffer:
                filename = f"Laporan_Neraca_{mahasiswa_info['nama_ketua']}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.docx"
                return send_file(
                    file_buffer,
                    as_attachment=True,
                    download_name=filename,
                    mimetype='application/vnd.openxmlformats-officedocument.wordprocessingml.document'
                )
            else:
                flash('Gagal membuat file Word!', 'danger')
                return redirect(url_for('admin.admin_monitoring_mahasiswa_laporan_neraca'))
        else:
            flash('Format file tidak valid!', 'danger')
            return redirect(url_for('admin.admin_monitoring_mahasiswa_laporan_neraca'))
        
    except Exception as e:
        print(f"Error in admin_download_neraca: {e}")
        flash(f'Error saat mengunduh file: {str(e)}', 'danger')
        return redirect(url_for('admin.admin_monitoring_mahasiswa_laporan_neraca'))

@admin_bp.route('/view_file_laporan_kemajuan/<int:proposal_id>')
def admin_view_file_laporan_kemajuan(proposal_id):
    if 'user_type' not in session or session['user_type'] != 'admin':
        flash('Anda harus login sebagai admin!', 'danger')
        return redirect(url_for('admin.admin_laporan_kemajuan'))
    
    try:
        app_funcs = get_app_functions()
        cursor = app_funcs['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        cursor.execute('''
            SELECT nama_file, file_path FROM file_laporan_kemajuan 
            WHERE id_proposal = %s
        ''', (proposal_id,))
        
        file_data = cursor.fetchone()
        cursor.close()
        
        if not file_data:
            flash('File laporan kemajuan tidak ditemukan!', 'danger')
            return redirect(url_for('admin.admin_laporan_kemajuan'))
        
        file_path = file_data['file_path']
        if not os.path.exists(file_path):
            flash('File tidak ditemukan di server!', 'danger')
            return redirect(url_for('admin.admin_laporan_kemajuan'))
        
        # Tampilkan file PDF di browser
        return send_file(file_path, as_attachment=False, mimetype='application/pdf')
        
    except Exception as e:
        flash(f'Error saat menampilkan file: {str(e)}', 'danger')
        return redirect(url_for('admin.admin_laporan_kemajuan'))

@admin_bp.route('/view_file_laporan_akhir/<int:proposal_id>')
def admin_view_file_laporan_akhir(proposal_id):
    if 'user_type' not in session or session['user_type'] != 'admin':
        flash('Anda harus login sebagai admin!', 'danger')
        return redirect(url_for('admin.admin_laporan_akhir'))
    
    try:
        app_funcs = get_app_functions()
        cursor = app_funcs['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        cursor.execute('''
            SELECT nama_file, file_path FROM file_laporan_akhir 
            WHERE id_proposal = %s
        ''', (proposal_id,))
        
        file_data = cursor.fetchone()
        cursor.close()
        
        if not file_data:
            flash('File laporan akhir tidak ditemukan!', 'danger')
            return redirect(url_for('admin.admin_laporan_akhir'))
        
        file_path = file_data['file_path']
        if not os.path.exists(file_path):
            flash('File tidak ditemukan di server!', 'danger')
            return redirect(url_for('admin.admin_laporan_akhir'))
        
        # Tampilkan file PDF di browser
        return send_file(file_path, as_attachment=False, mimetype='application/pdf')
        
    except Exception as e:
        flash(f'Error saat menampilkan file: {str(e)}', 'danger')
        return redirect(url_for('admin.admin_laporan_akhir'))

@admin_bp.route('/konfirmasi_pendanaan', methods=['POST'])
def admin_konfirmasi_pendanaan():
    if 'user_type' not in session or session['user_type'] != 'admin':
        return jsonify({'success': False, 'message': 'Anda harus login sebagai admin!'})
    
    if not hasattr(get_app_functions()['mysql'], 'connection') or get_app_functions()['mysql'].connection is None:
        return jsonify({'success': False, 'message': 'Koneksi ke database gagal!'})
    
    try:
        data = request.get_json()
        proposal_id = data.get('proposal_id')
        action = data.get('action')
        
        if not proposal_id or not action:
            return jsonify({'success': False, 'message': 'Data proposal_id dan action diperlukan!'})
        
        cursor = get_app_functions()['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Cek apakah proposal ada
        cursor.execute('SELECT * FROM proposal WHERE id = %s', (proposal_id,))
        proposal = cursor.fetchone()
        
        if not proposal:
            cursor.close()
            return jsonify({'success': False, 'message': 'Proposal tidak ditemukan!'})
        
        # Cek apakah reviewer sudah selesai review dan tidak ada anggaran revisi
        cursor.execute('''
            SELECT 
                (SELECT COUNT(*) FROM anggaran_awal aa 
                 WHERE aa.id_proposal = %s AND aa.status = 'revisi') as anggaran_awal_revisi,
                (SELECT COUNT(*) FROM anggaran_bertumbuh ab 
                 WHERE ab.id_proposal = %s AND ab.status = 'revisi') as anggaran_bertumbuh_revisi,
                (SELECT COUNT(*) FROM anggaran_awal aa 
                 WHERE aa.id_proposal = %s AND aa.status_reviewer = 'sudah_direview') as anggaran_awal_reviewed,
                (SELECT COUNT(*) FROM anggaran_bertumbuh ab 
                 WHERE ab.id_proposal = %s AND ab.status_reviewer = 'sudah_direview') as anggaran_bertumbuh_reviewed,
                (SELECT COUNT(*) FROM anggaran_awal aa 
                 WHERE aa.id_proposal = %s AND aa.status_reviewer = 'tolak_bantuan') as anggaran_awal_ditolak,
                (SELECT COUNT(*) FROM anggaran_bertumbuh ab 
                 WHERE ab.id_proposal = %s AND ab.status_reviewer = 'tolak_bantuan') as anggaran_bertumbuh_ditolak,
                (SELECT COUNT(*) FROM anggaran_awal aa 
                 WHERE aa.id_proposal = %s) as total_anggaran_awal,
                (SELECT COUNT(*) FROM anggaran_bertumbuh ab 
                 WHERE ab.id_proposal = %s) as total_anggaran_bertumbuh,
                (SELECT pr.status_review FROM proposal_reviewer pr 
                 WHERE pr.id_proposal = %s ORDER BY pr.tanggal_assign DESC LIMIT 1) as reviewer_status
        ''', (proposal_id, proposal_id, proposal_id, proposal_id, proposal_id, proposal_id, proposal_id, proposal_id, proposal_id))
        
        review_status = cursor.fetchone()
        
        # Validasi: tidak boleh ada anggaran revisi
        if review_status['anggaran_awal_revisi'] > 0 or review_status['anggaran_bertumbuh_revisi'] > 0:
            cursor.close()
            return jsonify({'success': False, 'message': 'Tidak dapat mengkonfirmasi pendanaan karena masih ada anggaran yang perlu direvisi!'})
        
        # Validasi: reviewer harus sudah selesai review atau sudah dinilai
        if review_status['reviewer_status'] not in ['selesai_review', 'sudah_dinilai']:
            cursor.close()
            return jsonify({'success': False, 'message': 'Reviewer belum menyelesaikan review proposal!'})
        
        # Validasi: semua anggaran harus sudah direview (termasuk yang ditolak)
        total_reviewed = (review_status['anggaran_awal_reviewed'] or 0) + (review_status['anggaran_bertumbuh_reviewed'] or 0)
        total_ditolak = (review_status['anggaran_awal_ditolak'] or 0) + (review_status['anggaran_bertumbuh_ditolak'] or 0)
        total_completed = total_reviewed + total_ditolak
        total_anggaran = (review_status['total_anggaran_awal'] or 0) + (review_status['total_anggaran_bertumbuh'] or 0)
        
        if total_anggaran > 0 and total_completed < total_anggaran:
            cursor.close()
            return jsonify({'success': False, 'message': f'Belum semua anggaran direview! ({total_completed}/{total_anggaran} direview)'})
        
        # Update status proposal berdasarkan action
        if action == 'lolos':
            status_admin_baru = 'lolos'
            message = 'Proposal berhasil diloloskan untuk pendanaan!'
        elif action == 'tidak_lolos':
            status_admin_baru = 'tidak_lolos'
            message = 'Proposal ditolak untuk pendanaan!'
        else:
            cursor.close()
            return jsonify({'success': False, 'message': 'Action tidak valid!'})
        
        # Update status admin proposal
        cursor.execute('''
            UPDATE proposal 
            SET status_admin = %s, tanggal_konfirmasi_admin = NOW()
            WHERE id = %s
        ''', (status_admin_baru, proposal_id))
        
        get_app_functions()['mysql'].connection.commit()
        cursor.close()
        
        return jsonify({
            'success': True,
            'message': message
        })
        
    except Exception as e:
        print(f"Error in admin_konfirmasi_pendanaan: {e}")
        return jsonify({'success': False, 'message': f'Error saat mengkonfirmasi pendanaan: {str(e)}'})

@admin_bp.route('/pengaturan')
def pengaturan():
    if 'user_type' not in session or session['user_type'] != 'admin':
        flash('Anda harus login sebagai admin!', 'danger')
        return redirect(url_for('index'))
    
    if not hasattr(get_app_functions()['mysql'], 'connection') or get_app_functions()['mysql'].connection is None:
        flash('Koneksi ke database gagal. Cek konfigurasi database!', 'danger')
        return render_template('admin/pengaturan.html', program_studi_list=[])
    
    try:
        cursor = get_app_functions()['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Ambil data program studi
        cursor.execute('SELECT * FROM program_studi ORDER BY nama_program_studi')
        program_studi_list = cursor.fetchall()
        
        cursor.close()
        
        return render_template('admin/pengaturan.html', program_studi_list=program_studi_list)
        
    except Exception as e:
        print(f"Error in pengaturan: {e}")
        flash(f'Error saat memuat data: {str(e)}', 'danger')
        return render_template('admin/pengaturan.html', program_studi_list=[])

@admin_bp.route('/add_program_studi', methods=['POST'])
def add_program_studi():
    if 'user_type' not in session or session['user_type'] != 'admin':
        return jsonify({'success': False, 'message': 'Anda harus login sebagai admin!'})
    
    if not hasattr(get_app_functions()['mysql'], 'connection') or get_app_functions()['mysql'].connection is None:
        return jsonify({'success': False, 'message': 'Koneksi ke database gagal!'})
    
    try:
        nama_program_studi = request.form.get('nama_program_studi')
        
        if not nama_program_studi:
            return jsonify({'success': False, 'message': 'Nama program studi harus diisi!'})
        
        cursor = get_app_functions()['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Cek apakah program studi sudah ada
        cursor.execute('SELECT * FROM program_studi WHERE nama_program_studi = %s', (nama_program_studi,))
        existing = cursor.fetchone()
        
        if existing:
            cursor.close()
            return jsonify({'success': False, 'message': 'Program studi sudah ada!'})
        
        # Tambah program studi baru
        cursor.execute('INSERT INTO program_studi (nama_program_studi) VALUES (%s)', (nama_program_studi,))
        get_app_functions()['mysql'].connection.commit()
        
        cursor.close()
        
        return jsonify({
            'success': True,
            'message': 'Program studi berhasil ditambahkan!'
        })
        
    except Exception as e:
        print(f"Error in add_program_studi: {e}")
        return jsonify({'success': False, 'message': f'Error saat menambah program studi: {str(e)}'})

@admin_bp.route('/update_program_studi', methods=['POST'])
def update_program_studi():
    if 'user_type' not in session or session['user_type'] != 'admin':
        return jsonify({'success': False, 'message': 'Anda harus login sebagai admin!'})
    
    if not hasattr(get_app_functions()['mysql'], 'connection') or get_app_functions()['mysql'].connection is None:
        return jsonify({'success': False, 'message': 'Koneksi ke database gagal!'})
    
    try:
        program_studi_id = request.form.get('program_studi_id')
        nama_program_studi = request.form.get('nama_program_studi')
        
        if not program_studi_id or not nama_program_studi:
            return jsonify({'success': False, 'message': 'ID dan nama program studi harus diisi!'})
        
        cursor = get_app_functions()['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Cek apakah program studi sudah ada (kecuali yang sedang diedit)
        cursor.execute('SELECT * FROM program_studi WHERE nama_program_studi = %s AND id != %s', (nama_program_studi, program_studi_id))
        existing = cursor.fetchone()
        
        if existing:
            cursor.close()
            return jsonify({'success': False, 'message': 'Program studi sudah ada!'})
        
        # Update program studi
        cursor.execute('UPDATE program_studi SET nama_program_studi = %s WHERE id = %s', (nama_program_studi, program_studi_id))
        get_app_functions()['mysql'].connection.commit()
        
        cursor.close()
        
        return jsonify({
            'success': True,
            'message': 'Program studi berhasil diperbarui!'
        })
        
    except Exception as e:
        print(f"Error in update_program_studi: {e}")
        return jsonify({'success': False, 'message': f'Error saat memperbarui program studi: {str(e)}'})

@admin_bp.route('/delete_program_studi', methods=['POST'])
def delete_program_studi():
    if 'user_type' not in session or session['user_type'] != 'admin':
        return jsonify({'success': False, 'message': 'Anda harus login sebagai admin!'})
    
    if not hasattr(get_app_functions()['mysql'], 'connection') or get_app_functions()['mysql'].connection is None:
        return jsonify({'success': False, 'message': 'Koneksi ke database gagal!'})
    
    try:
        program_studi_id = request.form.get('program_studi_id')
        
        if not program_studi_id:
            return jsonify({'success': False, 'message': 'ID program studi harus diisi!'})
        
        cursor = get_app_functions()['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Cek apakah program studi digunakan di tabel mahasiswa
        cursor.execute('SELECT COUNT(*) as count FROM mahasiswa WHERE program_studi = (SELECT nama_program_studi FROM program_studi WHERE id = %s)', (program_studi_id,))
        mahasiswa_count = cursor.fetchone()['count']
        
        if mahasiswa_count > 0:
            cursor.close()
            return jsonify({'success': False, 'message': f'Program studi tidak dapat dihapus karena digunakan oleh {mahasiswa_count} mahasiswa!'})
        
        # Cek apakah program studi digunakan di tabel pembimbing
        cursor.execute('SELECT COUNT(*) as count FROM pembimbing WHERE program_studi = (SELECT nama_program_studi FROM program_studi WHERE id = %s)', (program_studi_id,))
        pembimbing_count = cursor.fetchone()['count']
        
        if pembimbing_count > 0:
            cursor.close()
            return jsonify({'success': False, 'message': f'Program studi tidak dapat dihapus karena digunakan oleh {pembimbing_count} pembimbing!'})
        
        # Cek apakah program studi digunakan di tabel admin
        cursor.execute('SELECT COUNT(*) as count FROM admin WHERE program_studi = (SELECT nama_program_studi FROM program_studi WHERE id = %s)', (program_studi_id,))
        admin_count = cursor.fetchone()['count']
        
        if admin_count > 0:
            cursor.close()
            return jsonify({'success': False, 'message': f'Program studi tidak dapat dihapus karena digunakan oleh {admin_count} admin!'})
        
        # Hapus program studi
        cursor.execute('DELETE FROM program_studi WHERE id = %s', (program_studi_id,))
        get_app_functions()['mysql'].connection.commit()
        
        cursor.close()
        
        return jsonify({
            'success': True,
            'message': 'Program studi berhasil dihapus!'
        })
        
    except Exception as e:
        print(f"Error in delete_program_studi: {e}")
        return jsonify({'success': False, 'message': f'Error saat menghapus program studi: {str(e)}'})

@admin_bp.route('/get_program_studi_detail/<int:program_studi_id>')
def get_program_studi_detail(program_studi_id):
    if 'user_type' not in session or session['user_type'] != 'admin':
        return jsonify({'success': False, 'message': 'Anda harus login sebagai admin!'})
    
    if not hasattr(get_app_functions()['mysql'], 'connection') or get_app_functions()['mysql'].connection is None:
        return jsonify({'success': False, 'message': 'Koneksi ke database gagal!'})
    
    try:
        cursor = get_app_functions()['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        cursor.execute('SELECT * FROM program_studi WHERE id = %s', (program_studi_id,))
        program_studi = cursor.fetchone()
        
        cursor.close()
        
        if not program_studi:
            return jsonify({'success': False, 'message': 'Program studi tidak ditemukan!'})
        
        return jsonify({
            'success': True,
            'data': program_studi
        })
        
    except Exception as e:
        print(f"Error in get_program_studi_detail: {e}")
        return jsonify({'success': False, 'message': f'Error saat mengambil detail program studi: {str(e)}'})

@admin_bp.route('/get_program_studi_data')
def get_program_studi_data():
    if 'user_type' not in session or session['user_type'] != 'admin':
        return jsonify({'success': False, 'message': 'Akses ditolak'})
    
    try:
        cursor = get_app_functions()['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        cursor.execute('SELECT * FROM program_studi ORDER BY nama_program_studi ASC')
        program_studi_list = cursor.fetchall()
        cursor.close()
        
        # Format data untuk response
        formatted_data = []
        for i, program in enumerate(program_studi_list, 1):
            formatted_data.append({
                'no': i,
                'id': program['id'],
                'nama_program_studi': program['nama_program_studi'],
                'created_at': program['created_at'].strftime('%d/%m/%Y %H:%M') if program['created_at'] else '-'
            })
        
        return jsonify({
            'success': True,
            'data': formatted_data
        })
        
    except Exception as e:
        logger.error(f"Error getting program studi data: {str(e)}")
        return jsonify({'success': False, 'message': 'Terjadi kesalahan saat mengambil data program studi'})

@admin_bp.route('/reviewer')
def reviewer():
    logger.info("Halaman reviewer dipanggil")
    if 'user_type' not in session or session['user_type'] != 'admin':
        logger.warning(f"Akses ditolak untuk halaman reviewer: {session.get('user_type', 'None')}")
        flash('Anda harus login sebagai admin!', 'danger')
        return redirect(url_for('index'))
    
    try:
        cursor = get_app_functions()['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Ambil data reviewer beserta informasi kuota terpakai
        # Mahasiswa status 'selesai' tidak terhitung dalam kuota reviewer
        cursor.execute('''
            SELECT 
                r.id, 
                r.nama, 
                r.username, 
                r.kuota_review,
                COALESCE(COUNT(CASE WHEN m.status != 'selesai' THEN pr.id END), 0) AS kuota_terpakai
            FROM reviewer r
            LEFT JOIN proposal_reviewer pr ON pr.id_reviewer = r.id
            LEFT JOIN proposal p ON pr.id_proposal = p.id
            LEFT JOIN mahasiswa m ON p.nim = m.nim
            GROUP BY r.id, r.nama, r.username, r.kuota_review
            ORDER BY r.nama ASC
        ''')
        reviewers = cursor.fetchall()
        
        # Tambahkan informasi sisa kuota
        for reviewer in reviewers:
            reviewer['sisa_kuota'] = max(0, (reviewer.get('kuota_review') or 0) - (reviewer.get('kuota_terpakai') or 0))
            reviewer['status_kuota'] = 'tersedia' if reviewer['sisa_kuota'] > 0 else 'penuh'
        
        cursor.close()
        
        return render_template('admin/reviewer.html', reviewers=reviewers)
        
    except Exception as e:
        logger.error(f"Error saat mengambil data reviewer: {str(e)}")
        flash('Terjadi kesalahan saat mengambil data reviewer!', 'danger')
        return render_template('admin/reviewer.html', reviewers=[])

@admin_bp.route('/add_reviewer', methods=['POST'])
def add_reviewer():
    logger.info("Fungsi add_reviewer dipanggil")
    if 'user_type' not in session or session['user_type'] != 'admin':
        return jsonify({'success': False, 'message': 'Akses ditolak!'})
    
    try:
        data = request.get_json()
        nama = data.get('nama', '').strip()
        username = data.get('username', '').strip()
        password = data.get('password', '').strip()
        # Gunakan kuota dari request atau default 0
        kuota_review = data.get('kuota_review', 0)
        
        # Validasi input
        if not nama or not username or not password:
            return jsonify({'success': False, 'message': 'Nama, username, dan password harus diisi!'})
        
        # Validasi kuota review
        try:
            kuota_review = int(kuota_review) if kuota_review else 0
            if kuota_review < 0 or kuota_review > 100:
                return jsonify({'success': False, 'message': 'Kuota review harus antara 0-100!'})
        except (ValueError, TypeError):
            return jsonify({'success': False, 'message': 'Kuota review harus berupa angka!'})
        
        cursor = get_app_functions()['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Cek duplikasi komprehensif
        is_duplicate, message, table_name = check_comprehensive_duplicates(
            cursor, 
            nama=nama, 
            username=username
        )
        if is_duplicate:
            cursor.close()
            return jsonify({'success': False, 'message': message})
        
        # Hash password
        hashed_password = generate_password_hash(password)
        
        # Insert reviewer baru
        cursor.execute('''
            INSERT INTO reviewer (nama, username, password, kuota_review) 
            VALUES (%s, %s, %s, %s)
        ''', (nama, username, hashed_password, kuota_review))
        
        get_app_functions()['mysql'].connection.commit()
        cursor.close()
        
        logger.info(f"Reviewer baru berhasil ditambahkan: {nama}")
        return jsonify({'success': True, 'message': 'Reviewer berhasil ditambahkan!'})
        
    except Exception as e:
        logger.error(f"Error saat menambah reviewer: {str(e)}")
        return jsonify({'success': False, 'message': 'Terjadi kesalahan saat menambah reviewer!'})

@admin_bp.route('/get_reviewer_detail/<int:reviewer_id>')
def get_reviewer_detail(reviewer_id):
    logger.info(f"Fungsi get_reviewer_detail dipanggil untuk ID: {reviewer_id}")
    if 'user_type' not in session or session['user_type'] != 'admin':
        return jsonify({'success': False, 'message': 'Akses ditolak!'})
    
    try:
        cursor = get_app_functions()['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        cursor.execute('''
            SELECT id, nama, username, kuota_review
            FROM reviewer 
            WHERE id = %s
        ''', (reviewer_id,))
        
        reviewer = cursor.fetchone()
        cursor.close()
        
        if reviewer:
            return jsonify({'success': True, 'data': reviewer})
        else:
            return jsonify({'success': False, 'message': 'Reviewer tidak ditemukan!'})
            
    except Exception as e:
        logger.error(f"Error saat mengambil detail reviewer: {str(e)}")
        return jsonify({'success': False, 'message': 'Terjadi kesalahan saat mengambil data reviewer!'})

@admin_bp.route('/update_reviewer_data', methods=['POST'])
def update_reviewer_data():
    logger.info("Fungsi update_reviewer_data dipanggil")
    if 'user_type' not in session or session['user_type'] != 'admin':
        return jsonify({'success': False, 'message': 'Akses ditolak!'})
    
    try:
        data = request.get_json()
        reviewer_id = data.get('id')
        nama = data.get('nama', '').strip()
        username = data.get('username', '').strip()
        password = data.get('password', '').strip()
        kuota_review = data.get('kuota_review', 0)
        
        # Validasi input
        if not reviewer_id or not nama or not username:
            return jsonify({'success': False, 'message': 'ID, nama, dan username harus diisi!'})
        
        # Validasi kuota review
        try:
            kuota_review = int(kuota_review) if kuota_review else 0
            if kuota_review < 0 or kuota_review > 100:
                return jsonify({'success': False, 'message': 'Kuota review harus antara 0-100!'})
        except (ValueError, TypeError):
            return jsonify({'success': False, 'message': 'Kuota review harus berupa angka!'})
        
        cursor = get_app_functions()['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Cek duplikasi komprehensif (exclude reviewer yang sedang diedit)
        is_duplicate, message, table_name = check_comprehensive_duplicates(
            cursor, 
            nama=nama, 
            username=username,
            exclude_id=reviewer_id,
            exclude_table='reviewer'
        )
        if is_duplicate:
            cursor.close()
            return jsonify({'success': False, 'message': message})
        
        # Update data reviewer
        if password:  # Jika password diisi, update password juga
            hashed_password = generate_password_hash(password)
            cursor.execute('''
                UPDATE reviewer 
                SET nama = %s, username = %s, password = %s, kuota_review = %s
                WHERE id = %s
            ''', (nama, username, hashed_password, kuota_review, reviewer_id))
        else:  # Jika password kosong, tidak update password
            cursor.execute('''
                UPDATE reviewer 
                SET nama = %s, username = %s, kuota_review = %s
                WHERE id = %s
            ''', (nama, username, kuota_review, reviewer_id))
        
        get_app_functions()['mysql'].connection.commit()
        cursor.close()
        
        logger.info(f"Data reviewer berhasil diupdate: {nama}")
        return jsonify({'success': True, 'message': 'Data reviewer berhasil diupdate!'})
        
    except Exception as e:
        logger.error(f"Error saat update reviewer: {str(e)}")
        return jsonify({'success': False, 'message': 'Terjadi kesalahan saat update reviewer!'})

@admin_bp.route('/delete_reviewer', methods=['POST'])
def delete_reviewer():
    logger.info("Fungsi delete_reviewer dipanggil")
    if 'user_type' not in session or session['user_type'] != 'admin':
        return jsonify({'success': False, 'message': 'Akses ditolak!'})
    
    try:
        data = request.get_json()
        reviewer_id = data.get('id')
        
        if not reviewer_id:
            return jsonify({'success': False, 'message': 'ID reviewer tidak valid!'})
        
        cursor = get_app_functions()['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Ambil nama reviewer sebelum dihapus untuk log
        cursor.execute('SELECT nama FROM reviewer WHERE id = %s', (reviewer_id,))
        reviewer = cursor.fetchone()
        
        if not reviewer:
            cursor.close()
            return jsonify({'success': False, 'message': 'Reviewer tidak ditemukan!'})
        
        # Hapus semua assignment proposal reviewer
        cursor.execute('DELETE FROM proposal_reviewer WHERE id_reviewer = %s', (reviewer_id,))
        proposal_reviewer_deleted = cursor.rowcount
        
        # Hapus reviewer
        cursor.execute('DELETE FROM reviewer WHERE id = %s', (reviewer_id,))
        reviewer_deleted = cursor.rowcount
        
        get_app_functions()['mysql'].connection.commit()
        cursor.close()
        
        logger.info(f"Reviewer berhasil dihapus: {reviewer['nama']}")
        return jsonify({'success': True, 'message': f'Reviewer berhasil dihapus! Data yang dihapus: {reviewer_deleted} reviewer, {proposal_reviewer_deleted} assignment proposal.'})
        
    except Exception as e:
        logger.error(f"Error saat hapus reviewer: {str(e)}")
        return jsonify({'success': False, 'message': 'Terjadi kesalahan saat hapus reviewer!'})

@admin_bp.route('/get_reviewer_list')
def get_reviewer_list():
    logger.info("Fungsi get_reviewer_list dipanggil")
    if 'user_type' not in session or session['user_type'] != 'admin':
        return jsonify({'success': False, 'message': 'Akses ditolak!'})
    
    try:
        cursor = get_app_functions()['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)

        # Ambil reviewer beserta kuota terpakai dan sisa kuota
        # Mahasiswa status 'selesai' tidak terhitung dalam kuota reviewer
        cursor.execute('''
            SELECT 
                r.id,
                r.nama,
                r.username,
                r.kuota_review,
                COALESCE(COUNT(CASE WHEN m.status != 'selesai' THEN pr.id END), 0) AS kuota_terpakai
            FROM reviewer r
            LEFT JOIN proposal_reviewer pr ON pr.id_reviewer = r.id
            LEFT JOIN proposal p ON pr.id_proposal = p.id
            LEFT JOIN mahasiswa m ON p.nim = m.nim
            GROUP BY r.id, r.nama, r.username, r.kuota_review
            ORDER BY r.nama ASC
        ''')

        rows = cursor.fetchall()
        reviewers = []
        for row in rows:
            sisa = max(0, (row.get('kuota_review') or 0) - (row.get('kuota_terpakai') or 0))
            # Hanya masukkan reviewer dengan sisa kuota > 0
            if sisa > 0:
                reviewers.append({
                    'id': row['id'],
                    'nama': row['nama'],
                    'username': row['username'],
                    'kuota_review': row['kuota_review'],
                    'kuota_terpakai': row['kuota_terpakai'],
                    'sisa_kuota': sisa
                })
        cursor.close()
        
        return jsonify({
            'success': True, 
            'reviewers': reviewers
        })
        
    except Exception as e:
        logger.error(f"Error saat mengambil daftar reviewer: {str(e)}")
        return jsonify({'success': False, 'message': 'Terjadi kesalahan saat mengambil daftar reviewer!'})

@admin_bp.route('/assign_reviewer_to_proposal', methods=['POST'])
def assign_reviewer_to_proposal():
    logger.info("Fungsi assign_reviewer_to_proposal dipanggil")
    if 'user_type' not in session or session['user_type'] != 'admin':
        return jsonify({'success': False, 'message': 'Akses ditolak!'})
    
    try:
        data = request.get_json()
        proposal_id = data.get('proposal_id')
        reviewer_id = data.get('reviewer_id')
        
        if not proposal_id or not reviewer_id:
            return jsonify({'success': False, 'message': 'ID proposal dan reviewer harus diisi!'})
        
        cursor = get_app_functions()['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Cek apakah proposal sudah ada
        cursor.execute('SELECT id, judul_usaha, nim FROM proposal WHERE id = %s', (proposal_id,))
        proposal = cursor.fetchone()
        
        if not proposal:
            cursor.close()
            return jsonify({'success': False, 'message': 'Proposal tidak ditemukan!'})
        
        # Cek apakah reviewer sudah ada
        cursor.execute('SELECT id, nama FROM reviewer WHERE id = %s', (reviewer_id,))
        reviewer = cursor.fetchone()
        
        if not reviewer:
            cursor.close()
            return jsonify({'success': False, 'message': 'Reviewer tidak ditemukan!'})
        
        # Cek apakah proposal sudah memiliki reviewer
        cursor.execute('SELECT id FROM proposal_reviewer WHERE id_proposal = %s', (proposal_id,))
        existing_assignment = cursor.fetchone()
        
        if existing_assignment:
            # Update reviewer yang sudah ada
            cursor.execute('''
                UPDATE proposal_reviewer 
                SET id_reviewer = %s, tanggal_assign = NOW()
                WHERE id_proposal = %s
            ''', (reviewer_id, proposal_id))
        else:
            # Buat assignment baru
            cursor.execute('''
                INSERT INTO proposal_reviewer (id_proposal, id_reviewer, tanggal_assign)
                VALUES (%s, %s, NOW())
            ''', (proposal_id, reviewer_id))
        
        # Update status proposal menjadi 'diajukan' jika belum
        cursor.execute('''
            UPDATE proposal 
            SET status = 'diajukan', tanggal_review = NOW()
            WHERE id = %s AND status = 'draf'
        ''', (proposal_id,))
        
        get_app_functions()['mysql'].connection.commit()
        cursor.close()
        
        logger.info(f"Proposal {proposal['judul_usaha']} berhasil ditugaskan ke reviewer {reviewer['nama']}")
        return jsonify({
            'success': True, 
            'message': f'Proposal berhasil ditugaskan ke reviewer {reviewer["nama"]}!'
        })
        
    except Exception as e:
        logger.error(f"Error saat menugaskan reviewer: {str(e)}")
        return jsonify({'success': False, 'message': 'Terjadi kesalahan saat menugaskan reviewer!'})

@admin_bp.route('/get_proposal_reviewer/<int:proposal_id>')
def get_proposal_reviewer(proposal_id):
    logger.info(f"Fungsi get_proposal_reviewer dipanggil untuk proposal ID: {proposal_id}")
    if 'user_type' not in session or session['user_type'] != 'admin':
        return jsonify({'success': False, 'message': 'Akses ditolak!'})
    
    try:
        cursor = get_app_functions()['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        cursor.execute('''
            SELECT pr.id_reviewer, r.nama, r.username, pr.tanggal_assign
            FROM proposal_reviewer pr
            JOIN reviewer r ON pr.id_reviewer = r.id
            WHERE pr.id_proposal = %s
        ''', (proposal_id,))
        
        assignment = cursor.fetchone()
        cursor.close()
        
        if assignment:
            return jsonify({
                'success': True, 
                'data': assignment
            })
        else:
            return jsonify({
                'success': True, 
                'data': None
            })
            
    except Exception as e:
        logger.error(f"Error saat mengambil data reviewer proposal: {str(e)}")
        return jsonify({'success': False, 'message': 'Terjadi kesalahan saat mengambil data reviewer proposal!'})

# ========================================
# ROUTE UNTUK PENGELOLAAN PERTANYAAN PENILAIAN
# ========================================

@admin_bp.route('/add_pertanyaan_penilaian', methods=['POST'])
def add_pertanyaan_penilaian():
    """Tambah pertanyaan penilaian baru"""
    if 'user_type' not in session or session['user_type'] != 'admin':
        return jsonify({'success': False, 'message': 'Anda harus login sebagai admin!'})
    
    try:
        cursor = get_app_functions()['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Cek apakah ini request JSON untuk multiple pertanyaan mahasiswa
        if request.is_json:
            return add_multiple_pertanyaan_mahasiswa()
        
        # Ambil data dari form
        jenis_penilaian = request.form.get('jenis_penilaian')
        pertanyaan = request.form.get('pertanyaan')
        bobot = float(request.form.get('bobot'))
        status = request.form.get('status')
        
        # Ambil kategori untuk penilaian mahasiswa
        kategori = request.form.get('kategori') if jenis_penilaian == 'mahasiswa' else None
        
        # Validasi data
        if not pertanyaan or not bobot:
            return jsonify({'success': False, 'message': 'Pertanyaan dan bobot harus diisi!'})
        
        if bobot <= 0 or bobot > 100:
            return jsonify({'success': False, 'message': 'Bobot harus antara 1-100!'})
        
        # Validasi kategori untuk penilaian mahasiswa
        if jenis_penilaian == 'mahasiswa' and not kategori:
            return jsonify({'success': False, 'message': 'Kategori harus dipilih untuk penilaian mahasiswa!'})
        
        # Handle field khusus untuk proposal
        if jenis_penilaian == 'proposal':
            skor = int(request.form.get('skor', 100))
            if skor <= 0 or skor > 100:
                return jsonify({'success': False, 'message': 'Skor maksimal harus antara 1-100!'})
        else:
            skor = int(request.form.get('skor'))
            if not skor:
                return jsonify({'success': False, 'message': 'Skor maksimal harus diisi!'})
            if skor <= 0 or skor > 100:
                return jsonify({'success': False, 'message': 'Skor maksimal harus antara 1-100!'})
        
        # Tentukan tabel berdasarkan jenis penilaian
        tabel_map = {
            'mahasiswa': 'pertanyaan_penilaian_mahasiswa',
            'proposal': 'pertanyaan_penilaian_proposal',
            'laporan_kemajuan': 'pertanyaan_penilaian_laporan_kemajuan',
            'laporan_akhir': 'pertanyaan_penilaian_laporan_akhir'
        }
        
        if jenis_penilaian not in tabel_map:
            return jsonify({'success': False, 'message': 'Jenis penilaian tidak valid!'})
        
        tabel = tabel_map[jenis_penilaian]
        
        # Buat tabel jika belum ada
        if tabel == 'pertanyaan_penilaian_proposal':
            cursor.execute(f'''
                CREATE TABLE IF NOT EXISTS {tabel} (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    pertanyaan TEXT NOT NULL,
                    bobot DECIMAL(5,2) NOT NULL DEFAULT 0,
                    skor_maksimal INT NOT NULL DEFAULT 100,
                    is_active BOOLEAN DEFAULT TRUE,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    updated_at DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP
                )
            ''')
            
            # Insert data untuk proposal (dengan skor_maksimal)
            cursor.execute(f'''
                INSERT INTO {tabel} (pertanyaan, bobot, skor_maksimal, is_active)
                VALUES (%s, %s, %s, %s)
            ''', (pertanyaan, bobot, skor, status == 'Aktif'))
        elif tabel == 'pertanyaan_penilaian_mahasiswa':
            cursor.execute(f'''
                CREATE TABLE IF NOT EXISTS {tabel} (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    kategori VARCHAR(100) NOT NULL COMMENT 'Kategori penilaian mahasiswa',
                    pertanyaan TEXT NOT NULL,
                    bobot DECIMAL(5,2) NOT NULL COMMENT 'Bobot yang diinput admin (1-100)',
                    skor_maksimal INT NOT NULL COMMENT 'Skor maksimal yang bisa diberikan (1-100)',
                    status ENUM('Aktif', 'Nonaktif') DEFAULT 'Aktif',
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            ''')
            
            # Insert data untuk mahasiswa dengan kategori
            cursor.execute(f'''
                INSERT INTO {tabel} (kategori, pertanyaan, bobot, skor_maksimal, status)
                VALUES (%s, %s, %s, %s, %s)
            ''', (kategori, pertanyaan, bobot, skor, status))
        else:
            cursor.execute(f'''
                CREATE TABLE IF NOT EXISTS {tabel} (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    pertanyaan TEXT NOT NULL,
                    bobot DECIMAL(5,2) NOT NULL COMMENT 'Bobot yang diinput admin (1-100)',
                    skor_maksimal INT NOT NULL COMMENT 'Skor maksimal yang bisa diberikan (1-100)',
                    status ENUM('Aktif', 'Nonaktif') DEFAULT 'Aktif',
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            ''')
            
            # Insert data untuk tabel lain
            cursor.execute(f'''
                INSERT INTO {tabel} (pertanyaan, bobot, skor_maksimal, status)
                VALUES (%s, %s, %s, %s)
            ''', (pertanyaan, bobot, skor, status))
        
        get_app_functions()['mysql'].connection.commit()
        cursor.close()
        
        return jsonify({
            'success': True, 
            'message': f'Pertanyaan penilaian {jenis_penilaian} berhasil ditambahkan!'
        })
        
    except Exception as e:
        logger.error(f"Error adding pertanyaan penilaian: {str(e)}")
        return jsonify({'success': False, 'message': f'Terjadi kesalahan: {str(e)}'})


def add_multiple_pertanyaan_mahasiswa():
    """Tambah multiple pertanyaan penilaian mahasiswa berdasarkan kategori"""
    try:
        data = request.get_json()
        if not data or not isinstance(data, list):
            return jsonify({'success': False, 'message': 'Data tidak valid!'})
        
        cursor = get_app_functions()['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Buat tabel jika belum ada dengan kolom kategori
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS pertanyaan_penilaian_mahasiswa (
                id INT AUTO_INCREMENT PRIMARY KEY,
                kategori VARCHAR(100) NOT NULL COMMENT 'Kategori penilaian mahasiswa',
                pertanyaan TEXT NOT NULL,
                bobot DECIMAL(5,2) NOT NULL COMMENT 'Bobot yang diinput admin (1-100)',
                skor_maksimal INT NOT NULL COMMENT 'Skor maksimal yang bisa diberikan (1-100)',
                status ENUM('Aktif', 'Nonaktif') DEFAULT 'Aktif',
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        
        # Cek apakah kolom kategori sudah ada, jika tidak tambahkan
        try:
            cursor.execute("SELECT kategori FROM pertanyaan_penilaian_mahasiswa LIMIT 1")
        except:
            # Kolom kategori belum ada, tambahkan
            cursor.execute('''
                ALTER TABLE pertanyaan_penilaian_mahasiswa 
                ADD COLUMN kategori VARCHAR(100) NOT NULL COMMENT 'Kategori penilaian mahasiswa' AFTER id
            ''')
            # Set default kategori untuk data yang sudah ada
            cursor.execute('''
                UPDATE pertanyaan_penilaian_mahasiswa 
                SET kategori = 'umum' 
                WHERE kategori IS NULL OR kategori = ''
            ''')
        
        # Insert multiple pertanyaan
        for item in data:
            cursor.execute('''
                INSERT INTO pertanyaan_penilaian_mahasiswa (kategori, pertanyaan, bobot, skor_maksimal, status)
                VALUES (%s, %s, %s, %s, %s)
            ''', (item['kategori'], item['pertanyaan'], item['bobot'], item['skor'], item['status']))
        
        get_app_functions()['mysql'].connection.commit()
        cursor.close()
        
        return jsonify({
            'success': True, 
            'message': f'{len(data)} pertanyaan penilaian mahasiswa berhasil ditambahkan!'
        })
        
    except Exception as e:
        logger.error(f"Error adding multiple pertanyaan penilaian mahasiswa: {str(e)}")
        return jsonify({'success': False, 'message': f'Terjadi kesalahan: {str(e)}'})


@admin_bp.route('/update_pertanyaan_penilaian', methods=['POST'])
def update_pertanyaan_penilaian():
    """Update pertanyaan penilaian"""
    if 'user_type' not in session or session['user_type'] != 'admin':
        return jsonify({'success': False, 'message': 'Anda harus login sebagai admin!'})
    
    try:
        cursor = get_app_functions()['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Ambil data dari form
        jenis_penilaian = request.form.get('jenis_penilaian')
        pertanyaan_id = int(request.form.get('pertanyaan_id'))
        pertanyaan = request.form.get('pertanyaan')
        bobot = float(request.form.get('bobot'))
        status = request.form.get('status')
        
        # Ambil kategori untuk penilaian mahasiswa
        kategori = request.form.get('kategori') if jenis_penilaian == 'mahasiswa' else None
        
        # Validasi data
        if not pertanyaan or not bobot:
            return jsonify({'success': False, 'message': 'Pertanyaan dan bobot harus diisi!'})
        
        if bobot <= 0 or bobot > 100:
            return jsonify({'success': False, 'message': 'Bobot harus antara 1-100!'})
        
        # Validasi kategori untuk penilaian mahasiswa
        if jenis_penilaian == 'mahasiswa' and not kategori:
            return jsonify({'success': False, 'message': 'Kategori harus dipilih untuk penilaian mahasiswa!'})
        
        # Handle field khusus untuk proposal
        if jenis_penilaian == 'proposal':
            skor = int(request.form.get('skor', 100))
            if skor <= 0 or skor > 100:
                return jsonify({'success': False, 'message': 'Skor maksimal harus antara 1-100!'})
        else:
            skor = int(request.form.get('skor'))
            if not skor:
                return jsonify({'success': False, 'message': 'Skor maksimal harus diisi!'})
            if skor <= 0 or skor > 100:
                return jsonify({'success': False, 'message': 'Skor maksimal harus antara 1-100!'})
        
        # Tentukan tabel berdasarkan jenis penilaian
        tabel_map = {
            'mahasiswa': 'pertanyaan_penilaian_mahasiswa',
            'proposal': 'pertanyaan_penilaian_proposal',
            'laporan_kemajuan': 'pertanyaan_penilaian_laporan_kemajuan',
            'laporan_akhir': 'pertanyaan_penilaian_laporan_akhir'
        }
        
        if jenis_penilaian not in tabel_map:
            return jsonify({'success': False, 'message': 'Jenis penilaian tidak valid!'})
        
        tabel = tabel_map[jenis_penilaian]
        
        # Update data
        if tabel == 'pertanyaan_penilaian_proposal':
            cursor.execute(f'''
                UPDATE {tabel} 
                SET pertanyaan = %s, bobot = %s, skor_maksimal = %s, is_active = %s
                WHERE id = %s
            ''', (pertanyaan, bobot, skor, status == 'Aktif', pertanyaan_id))
        elif tabel == 'pertanyaan_penilaian_mahasiswa':
            cursor.execute(f'''
                UPDATE {tabel} 
                SET kategori = %s, pertanyaan = %s, bobot = %s, skor_maksimal = %s, status = %s
                WHERE id = %s
            ''', (kategori, pertanyaan, bobot, skor, status, pertanyaan_id))
        else:
            cursor.execute(f'''
                UPDATE {tabel} 
                SET pertanyaan = %s, bobot = %s, skor_maksimal = %s, status = %s
                WHERE id = %s
            ''', (pertanyaan, bobot, skor, status, pertanyaan_id))
        
        if cursor.rowcount == 0:
            return jsonify({'success': False, 'message': 'Pertanyaan tidak ditemukan!'})
        
        get_app_functions()['mysql'].connection.commit()
        cursor.close()
        
        return jsonify({
            'success': True, 
            'message': f'Pertanyaan penilaian {jenis_penilaian} berhasil diperbarui!'
        })
        
    except Exception as e:
        logger.error(f"Error updating pertanyaan penilaian: {str(e)}")
        return jsonify({'success': False, 'message': f'Terjadi kesalahan: {str(e)}'})

@admin_bp.route('/delete_pertanyaan_penilaian', methods=['POST'])
def delete_pertanyaan_penilaian():
    """Hapus pertanyaan penilaian"""
    if 'user_type' not in session or session['user_type'] != 'admin':
        return jsonify({'success': False, 'message': 'Anda harus login sebagai admin!'})
    
    try:
        cursor = get_app_functions()['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Ambil data dari form
        jenis_penilaian = request.form.get('jenis_penilaian')
        pertanyaan_id = int(request.form.get('pertanyaan_id'))
        
        # Tentukan tabel berdasarkan jenis penilaian
        tabel_map = {
            'mahasiswa': 'pertanyaan_penilaian_mahasiswa',
            'proposal': 'pertanyaan_penilaian_proposal',
            'laporan_kemajuan': 'pertanyaan_penilaian_laporan_kemajuan',
            'laporan_akhir': 'pertanyaan_penilaian_laporan_akhir'
        }
        
        if jenis_penilaian not in tabel_map:
            return jsonify({'success': False, 'message': 'Jenis penilaian tidak valid!'})
        
        tabel = tabel_map[jenis_penilaian]
        
        # Cek apakah pertanyaan sudah digunakan dalam penilaian
        detail_tabel_map = {
            'mahasiswa': 'detail_penilaian_mahasiswa',
            'proposal': 'detail_penilaian_proposal',
            'laporan_kemajuan': 'detail_penilaian_laporan_kemajuan',
            'laporan_akhir': 'detail_penilaian_laporan_akhir'
        }
        
        detail_tabel = detail_tabel_map[jenis_penilaian]
        
        # Cek apakah tabel detail ada dan ada data yang menggunakan pertanyaan ini
        cursor.execute(f"SHOW TABLES LIKE '{detail_tabel}'")
        if cursor.fetchone():
            cursor.execute(f"SELECT COUNT(*) as count FROM {detail_tabel} WHERE id_pertanyaan = %s", (pertanyaan_id,))
            count = cursor.fetchone()['count']
            if count > 0:
                return jsonify({
                    'success': False, 
                    'message': f'Pertanyaan tidak dapat dihapus karena sudah digunakan dalam {count} penilaian!'
                })
        
        # Hapus data
        cursor.execute(f'DELETE FROM {tabel} WHERE id = %s', (pertanyaan_id,))
        
        if cursor.rowcount == 0:
            return jsonify({'success': False, 'message': 'Pertanyaan tidak ditemukan!'})
        
        get_app_functions()['mysql'].connection.commit()
        cursor.close()
        
        return jsonify({
            'success': True, 
            'message': f'Pertanyaan penilaian {jenis_penilaian} berhasil dihapus!'
        })
        
    except Exception as e:
        logger.error(f"Error deleting pertanyaan penilaian: {str(e)}")
        return jsonify({'success': False, 'message': f'Terjadi kesalahan: {str(e)}'})

@admin_bp.route('/get_pengaturan_default')
def get_pengaturan_default():
    """Mendapatkan pengaturan default untuk frontend"""
    if 'user_type' not in session or session['user_type'] != 'admin':
        return jsonify({'success': False, 'message': 'Anda harus login sebagai admin!'})
    
    try:
        # Return pengaturan default (akan diambil dari localStorage di frontend)
        default_settings = {
            'default_skor_mahasiswa': 100,
            'default_skor_proposal': 100,
            'default_skor_laporan_kemajuan': 100,
            'default_skor_laporan_akhir': 100,
            'default_kuota_mahasiswa': 5,
            'default_kuota_review': 0
        }
        
        return jsonify({'success': True, 'data': default_settings})
        
    except Exception as e:
        logger.error(f"Error saat mengambil pengaturan default: {str(e)}")
        return jsonify({'success': False, 'message': 'Terjadi kesalahan saat mengambil pengaturan default!'})

@admin_bp.route('/get_pertanyaan_penilaian_data')
def get_pertanyaan_penilaian_data():
    """Ambil data pertanyaan penilaian untuk DataTable"""
    if 'user_type' not in session or session['user_type'] != 'admin':
        return jsonify({'success': False, 'message': 'Anda harus login sebagai admin!'})
    
    try:
        cursor = get_app_functions()['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Ambil parameter dari request
        jenis_penilaian = request.args.get('jenis_penilaian')
        
        # Tentukan tabel berdasarkan jenis penilaian
        tabel_map = {
            'mahasiswa': 'pertanyaan_penilaian_mahasiswa',
            'proposal': 'pertanyaan_penilaian_proposal',
            'laporan_kemajuan': 'pertanyaan_penilaian_laporan_kemajuan',
            'laporan_akhir': 'pertanyaan_penilaian_laporan_akhir'
        }
        
        if jenis_penilaian not in tabel_map:
            return jsonify({'success': False, 'message': 'Jenis penilaian tidak valid!'})
        
        tabel = tabel_map[jenis_penilaian]
        
        # Buat tabel jika belum ada
        if tabel == 'pertanyaan_penilaian_proposal':
            cursor.execute(f'''
                CREATE TABLE IF NOT EXISTS {tabel} (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    pertanyaan TEXT NOT NULL,
                    bobot DECIMAL(5,2) NOT NULL DEFAULT 0,
                    skor_maksimal INT NOT NULL DEFAULT 100,
                    is_active BOOLEAN DEFAULT TRUE,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    updated_at DATETIME DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP
                )
            ''')
            
            # Ambil data untuk proposal
            cursor.execute(f'''
                SELECT id, pertanyaan, bobot, skor_maksimal, is_active, created_at
                FROM {tabel}
                ORDER BY created_at ASC
            ''')
            
            data = cursor.fetchall()
            
            # Format data untuk DataTable
            formatted_data = []
            for i, item in enumerate(data, 1):
                formatted_data.append({
                    'no': i,
                    'id': item['id'],
                    'pertanyaan': item['pertanyaan'],
                    'bobot': f"{float(item['bobot']):.0f}",
                    'bobot_asli': float(item['bobot']),
                    'skor': item['skor_maksimal'],
                    'status': 'Aktif' if item['is_active'] else 'Nonaktif',
                    'created_at': item['created_at'].strftime('%d/%m/%Y %H:%M') if item['created_at'] else ''
                })
        elif tabel == 'pertanyaan_penilaian_mahasiswa':
            cursor.execute(f'''
                CREATE TABLE IF NOT EXISTS {tabel} (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    kategori VARCHAR(100) NOT NULL COMMENT 'Kategori penilaian mahasiswa',
                    pertanyaan TEXT NOT NULL,
                    bobot DECIMAL(5,2) NOT NULL COMMENT 'Bobot yang diinput admin (1-100)',
                    skor_maksimal INT NOT NULL COMMENT 'Skor maksimal yang bisa diberikan (1-100)',
                    status ENUM('Aktif', 'Nonaktif') DEFAULT 'Aktif',
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            ''')
            
            # Ambil data untuk mahasiswa dengan kategori
            cursor.execute(f'''
                SELECT id, kategori, pertanyaan, bobot, skor_maksimal, status, created_at
                FROM {tabel}
                ORDER BY FIELD(kategori, 'pondasi-bisnis', 'pelanggan', 'produk-jasa', 'pemasaran', 'penjualan', 'keuntungan', 'sdm', 'sistem-bisnis'), created_at ASC
            ''')
            
            data = cursor.fetchall()
            
            # Format data untuk DataTable
            formatted_data = []
            for i, item in enumerate(data, 1):
                formatted_data.append({
                    'no': i,
                    'id': item['id'],
                    'kategori': item['kategori'],
                    'pertanyaan': item['pertanyaan'],
                    'bobot': f"{float(item['bobot']):.0f}",
                    'bobot_asli': float(item['bobot']),
                    'skor': item['skor_maksimal'],
                    'status': item['status'],
                    'created_at': item['created_at'].strftime('%d/%m/%Y %H:%M') if item['created_at'] else ''
                })
        else:
            cursor.execute(f'''
                CREATE TABLE IF NOT EXISTS {tabel} (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    pertanyaan TEXT NOT NULL,
                    bobot DECIMAL(5,2) NOT NULL COMMENT 'Bobot yang diinput admin (1-100)',
                    skor_maksimal INT NOT NULL COMMENT 'Skor maksimal yang bisa diberikan (1-100)',
                    status ENUM('Aktif', 'Nonaktif') DEFAULT 'Aktif',
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                )
            ''')
            
            # Ambil data untuk tabel lain
            cursor.execute(f'''
                SELECT id, pertanyaan, bobot, skor_maksimal, status, created_at
                FROM {tabel}
                ORDER BY created_at ASC
            ''')
            
            data = cursor.fetchall()
            
            # Format data untuk DataTable
            formatted_data = []
            for i, item in enumerate(data, 1):
                formatted_data.append({
                    'no': i,
                    'id': item['id'],
                    'pertanyaan': item['pertanyaan'],
                    'bobot': f"{float(item['bobot']):.0f}",
                    'bobot_asli': float(item['bobot']),
                    'skor': item['skor_maksimal'],
                    'status': item['status'],
                    'created_at': item['created_at'].strftime('%d/%m/%Y %H:%M') if item['created_at'] else ''
                })
        
        cursor.close()
        
        return jsonify({
            'success': True,
            'data': formatted_data
        })
        
    except Exception as e:
        logger.error(f"Error getting pertanyaan penilaian data: {str(e)}")
        return jsonify({'success': False, 'message': f'Terjadi kesalahan: {str(e)}'})

# ========================================
# API ADMIN: DAFTAR NILAI & DETAIL PENILAIAN
# ========================================

@admin_bp.route('/daftar_nilai_mahasiswa')
def daftar_nilai_mahasiswa():
    if 'user_type' not in session or session['user_type'] != 'admin':
        flash('Anda harus login sebagai admin!', 'danger')
        return redirect(url_for('index'))
    try:
        return render_template('admin/daftar_nilai_mahasiswa.html')
    except Exception as e:
        logger.error(f"Error render admin daftar nilai: {str(e)}")
        flash('Terjadi kesalahan saat memuat halaman!', 'danger')
        return render_template('admin/daftar_nilai_mahasiswa.html')

@admin_bp.route('/get_daftar_nilai_mahasiswa')
def admin_get_daftar_nilai_mahasiswa():
    if 'user_type' not in session or session['user_type'] != 'admin':
        return jsonify({'success': False, 'message': 'Anda harus login sebagai admin!'})
    try:
        cursor = get_app_functions()['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Query untuk mendapatkan data nilai mahasiswa dengan perhitungan yang benar
        # Test query sederhana dulu untuk membandingkan nilai
        print("=== TEST QUERY NILAI MAHASISWA ===")
        cursor.execute('''
            SELECT 
                p.id AS proposal_id,
                m.nim,
                m.nama_ketua AS nama_mahasiswa,
                p.judul_usaha,
                -- Test: Hitung nilai dengan cara sederhana
                (
                    SELECT SUM(dpm.nilai)
                    FROM penilaian_mahasiswa pm
                    INNER JOIN detail_penilaian_mahasiswa dpm ON pm.id = dpm.id_penilaian_mahasiswa
                    WHERE pm.id_proposal = p.id
                ) AS nilai_mahasiswa_simple,
                -- Test: Hitung nilai dengan rumus yang kompleks
                (
                    SELECT SUM(
                        (dpm_avg.avg_skor / ppm.skor_maksimal) * ppm.bobot
                    )
                    FROM penilaian_mahasiswa pm
                    INNER JOIN detail_penilaian_mahasiswa dpm ON pm.id = dpm.id_penilaian_mahasiswa
                    INNER JOIN pertanyaan_penilaian_mahasiswa ppm ON dpm.id_pertanyaan = ppm.id
                    INNER JOIN (
                        SELECT 
                            dpm2.id_pertanyaan,
                            dpm2.id_penilaian_mahasiswa,
                            AVG(dpm2.skor) as avg_skor
                        FROM detail_penilaian_mahasiswa dpm2
                        GROUP BY dpm2.id_pertanyaan, dpm2.id_penilaian_mahasiswa
                    ) dpm_avg ON dpm_avg.id_pertanyaan = dpm.id_pertanyaan 
                        AND dpm_avg.id_penilaian_mahasiswa = dpm.id_penilaian_mahasiswa
                    WHERE pm.id_proposal = p.id
                ) AS nilai_mahasiswa_complex
            FROM proposal p
            INNER JOIN mahasiswa m ON m.nim = p.nim
            WHERE p.status IN ('disetujui', 'revisi', 'selesai')
            AND p.status_admin = 'lolos'
            ORDER BY p.tanggal_buat DESC
        ''')
        
        test_rows = cursor.fetchall()
        for test_row in test_rows:
            print(f"Proposal ID: {test_row['proposal_id']}")
            print(f"NIM: {test_row['nim']}")
            print(f"Nilai Simple: {test_row['nilai_mahasiswa_simple']}")
            print(f"Nilai Complex: {test_row['nilai_mahasiswa_complex']}")
            print("---")
        
        print("=== END TEST QUERY ===")
        
        # Query utama untuk mendapatkan data nilai mahasiswa dengan data tambahan
        cursor.execute('''
            SELECT 
                p.id AS proposal_id,
                m.nim,
                m.nama_ketua AS nama_mahasiswa,
                p.judul_usaha,
                -- Hitung nilai mahasiswa yang sudah dikonversi ke persentase (seperti di modal)
                COALESCE(
                    (
                        SELECT 
                            CASE 
                                WHEN SUM(ppm.bobot) > 0 THEN
                                    -- Konversi ke persentase: (total_nilai / total_bobot) * 100
                                    (SUM(
                                        (dpm_avg.avg_skor / ppm.skor_maksimal) * ppm.bobot
                                    ) / SUM(ppm.bobot)) * 100
                                ELSE 0
                            END
                        FROM penilaian_mahasiswa pm
                        INNER JOIN detail_penilaian_mahasiswa dpm ON pm.id = dpm.id_penilaian_mahasiswa
                        INNER JOIN pertanyaan_penilaian_mahasiswa ppm ON dpm.id_pertanyaan = ppm.id
                        INNER JOIN (
                            SELECT 
                                dpm2.id_pertanyaan,
                                dpm2.id_penilaian_mahasiswa,
                                AVG(dpm2.skor) as avg_skor
                            FROM detail_penilaian_mahasiswa dpm2
                            GROUP BY dpm2.id_pertanyaan, dpm2.id_penilaian_mahasiswa
                        ) dpm_avg ON dpm_avg.id_pertanyaan = dpm.id_pertanyaan 
                            AND dpm_avg.id_penilaian_mahasiswa = dpm.id_penilaian_mahasiswa
                        WHERE pm.id_proposal = p.id
                    ), 0
                ) AS nilai_mahasiswa,
                COALESCE(nilai_prop.avg_nilai, 0) AS nilai_proposal,
                COALESCE(nilai_prop.nilai_bantuan, 0) AS nilai_bantuan,
                COALESCE(nilai_kem.avg_nilai, 0) AS nilai_kemajuan,
                COALESCE(nilai_kem.rekomendasi_pendanaan, '') AS rekomendasi_pendanaan,
                COALESCE(nilai_akh.avg_nilai, 0) AS nilai_akhir,
                COALESCE(nilai_akh.rekomendasi_kelulusan, '') AS rekomendasi_kelulusan
            FROM proposal p
            INNER JOIN mahasiswa m ON m.nim = p.nim
            LEFT JOIN (
                SELECT 
                    id_proposal, 
                    AVG(nilai_akhir) AS avg_nilai,
                    AVG(nilai_bantuan) AS nilai_bantuan
                FROM penilaian_proposal 
                GROUP BY id_proposal
            ) nilai_prop ON nilai_prop.id_proposal = p.id
            LEFT JOIN (
                SELECT 
                    id_proposal, 
                    AVG(nilai_akhir) AS avg_nilai,
                    MAX(rekomendasi_pendanaan) AS rekomendasi_pendanaan
                FROM penilaian_laporan_kemajuan 
                GROUP BY id_proposal
            ) nilai_kem ON nilai_kem.id_proposal = p.id
            LEFT JOIN (
                SELECT 
                    id_proposal, 
                    AVG(nilai_akhir) AS avg_nilai,
                    MAX(rekomendasi_kelulusan) AS rekomendasi_kelulusan
                FROM penilaian_laporan_akhir 
                GROUP BY id_proposal
            ) nilai_akh ON nilai_akh.id_proposal = p.id
            WHERE p.status IN ('disetujui', 'revisi', 'selesai')
            AND p.status_admin = 'lolos'
            ORDER BY p.tanggal_buat DESC
        ''')
        rows = cursor.fetchall()
        
        data = []
        for r in rows:
            # Ambil nilai mahasiswa langsung dari hasil query
            nilai_mahasiswa = float(r['nilai_mahasiswa']) if r['nilai_mahasiswa'] is not None else 0.0
            
            # Debug: Log nilai untuk memastikan perhitungan benar
            print(f"=== DEBUG NILAI MAHASISWA ===")
            print(f"NIM: {r['nim']}")
            print(f"Nama: {r['nama_mahasiswa']}")
            print(f"Raw nilai_mahasiswa: {r['nilai_mahasiswa']}")
            print(f"Processed nilai_mahasiswa: {nilai_mahasiswa}")
            
            # Debug: Tampilkan detail perhitungan jika nilai tidak 0
            if nilai_mahasiswa > 0:
                print(f"✅ NILAI SUDAH DIPERBAIKI: {nilai_mahasiswa}")
                print(f"💡 Sekarang nilai dalam format persentase yang benar")
                print(f"🔍 Rumus: (total_nilai / total_bobot) * 100")
            
            print(f"===============================")
            
            data.append({
                'proposal_id': r['proposal_id'],
                'nim': r['nim'],
                'nama_mahasiswa': r['nama_mahasiswa'],
                'judul_usaha': r['judul_usaha'],
                'nilai_mahasiswa': nilai_mahasiswa,
                'nilai_proposal': float(r['nilai_proposal']) if r['nilai_proposal'] is not None else 0.0,
                'nilai_bantuan': float(r['nilai_bantuan']) if r['nilai_bantuan'] is not None else 0.0,
                'nilai_kemajuan': float(r['nilai_kemajuan']) if r['nilai_kemajuan'] is not None else 0.0,
                'rekomendasi_pendanaan': r['rekomendasi_pendanaan'] if r['rekomendasi_pendanaan'] else '',
                'nilai_akhir': float(r['nilai_akhir']) if r['nilai_akhir'] is not None else 0.0,
                'rekomendasi_kelulusan': r['rekomendasi_kelulusan'] if r['rekomendasi_kelulusan'] else '',
            })
        
        cursor.close()
        return jsonify({'success': True, 'data': data})
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})


@admin_bp.route('/get_admin_detail_penilaian_mahasiswa')
def get_admin_detail_penilaian_mahasiswa():
    if 'user_type' not in session or session['user_type'] != 'admin':
        return jsonify({'success': False, 'message': 'Anda harus login sebagai admin!'})
    try:
        nim = request.args.get('nim')
        if not nim:
            return jsonify({'success': False, 'message': 'NIM tidak ditemukan!'})
        cursor = get_app_functions()['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Helper function untuk nama kategori yang user-friendly dengan urutan yang benar
        def get_kategori_display_name(kategori):
            kategori_map = {
                'pondasi-bisnis': 'A. Pondasi Bisnis',
                'pelanggan': 'B. Pelanggan',
                'produk-jasa': 'C. Produk/Jasa',
                'pemasaran': 'D. Pemasaran',
                'penjualan': 'E. Penjualan',
                'keuntungan': 'F. Keuntungan',
                'sdm': 'G. Sumber Daya Manusia',
                'sistem-bisnis': 'H. Sistem Bisnis'
            }
            return kategori_map.get(kategori, kategori.title())
        
        # Helper function untuk urutan kategori yang benar
        def get_kategori_order(kategori):
            kategori_order = {
                'pondasi-bisnis': 1,
                'pelanggan': 2,
                'produk-jasa': 3,
                'pemasaran': 4,
                'penjualan': 5,
                'keuntungan': 6,
                'sdm': 7,
                'sistem-bisnis': 8
            }
            return kategori_order.get(kategori, 999)  # Default ke akhir jika tidak ada
        
        # Ambil data mahasiswa dan proposal
        cursor.execute('''
            SELECT p.id AS proposal_id, m.nim, m.nama_ketua AS nama_mahasiswa, p.judul_usaha, m.program_studi
            FROM mahasiswa m INNER JOIN proposal p ON m.nim = p.nim
            WHERE m.nim = %s
            ORDER BY p.tanggal_buat DESC LIMIT 1
        ''', (nim,))
        mhs = cursor.fetchone()
        if not mhs:
            cursor.close()
            return jsonify({'success': False, 'message': 'Data mahasiswa tidak ditemukan!'})
        
        # Ambil data penilaian mahasiswa
        cursor.execute('''
            SELECT id, nilai_akhir, komentar_pembimbing, status, created_at, updated_at
            FROM penilaian_mahasiswa
            WHERE id_proposal = %s
            ORDER BY updated_at DESC LIMIT 1
        ''', (mhs['proposal_id'],))
        pen = cursor.fetchone()
        if not pen:
            cursor.close()
            return jsonify({'success': False, 'message': 'Belum ada penilaian mahasiswa!'})
        
        # ✅ PERBAIKAN: Ambil data detail penilaian dengan metadata kolom yang sesuai
        cursor.execute('''
            SELECT 
                dpm.skor, 
                dpm.nilai, 
                dpm.sesi_penilaian,
                dpm.is_locked,
                dpm.tanggal_input,
                dpm.metadata_kolom_id,
                ppm.pertanyaan, 
                ppm.bobot, 
                ppm.skor_maksimal,
                ppm.kategori,
                ppm.id as id_pertanyaan
            FROM detail_penilaian_mahasiswa dpm
            INNER JOIN pertanyaan_penilaian_mahasiswa ppm ON dpm.id_pertanyaan = ppm.id
            WHERE dpm.id_penilaian_mahasiswa = %s
            ORDER BY ppm.kategori ASC, ppm.created_at ASC, dpm.sesi_penilaian ASC
        ''', (pen['id'],))
        details = cursor.fetchall()
        
        # Ambil data jadwal untuk informasi sesi dari database
        cursor.execute('''
            SELECT 
                pembimbing_interval_tipe as interval_tipe,
                pembimbing_interval_nilai as interval_nilai,
                pembimbing_nilai_mulai as jam_mulai,
                pembimbing_nilai_selesai as jam_selesai
            FROM pengaturan_jadwal
            ORDER BY created_at DESC LIMIT 1
        ''')
        jadwal_info = cursor.fetchone()
        
        # Generate jam_list berdasarkan interval dari database
        if jadwal_info and jadwal_info['interval_tipe'] and jadwal_info['jam_mulai'] and jadwal_info['jam_selesai']:
            jam_mulai = jadwal_info['jam_mulai']
            jam_selesai = jadwal_info['jam_selesai']
            interval = jadwal_info['interval_nilai'] or 1
            
            # Debug: Log data mentah dari database
            print(f"=== DEBUG JADWAL DATABASE ===")
            print(f"Raw jam_mulai from DB: {jam_mulai} (type: {type(jam_mulai)})")
            print(f"Raw jam_selesai from DB: {jam_selesai} (type: {type(jam_selesai)})")
            print(f"Raw interval from DB: {interval} (type: {type(interval)})")
            print(f"Raw interval_tipe from DB: {jadwal_info['interval_tipe']}")
            print(f"Database field pembimbing_nilai_mulai: {jadwal_info.get('jam_mulai')}")
            print(f"Database field pembimbing_nilai_selesai: {jadwal_info.get('jam_selesai')}")
            print(f"===============================")
            
            # Generate list jam dari jam_mulai sampai jam_selesai dengan interval
            jam_list = []
            
            try:
                # Pastikan jam_mulai dan jam_selesai adalah time object
                if isinstance(jam_mulai, str):
                    # Coba parse sebagai datetime lengkap dulu
                    try:
                        jam_mulai = datetime.strptime(jam_mulai, '%Y-%m-%d %H:%M:%S')
                        jam_mulai = jam_mulai.time()
                        print(f"Parsed datetime string jam_mulai: {jam_mulai}")
                    except:
                        # Jika gagal, coba parse sebagai time saja
                        jam_mulai = datetime.strptime(jam_mulai, '%H:%M:%S').time()
                        print(f"Parsed time string jam_mulai: {jam_mulai}")
                elif isinstance(jam_mulai, datetime):
                    # Jika sudah datetime object, ambil bagian time saja
                    jam_mulai = jam_mulai.time()
                    print(f"Extracted time from datetime jam_mulai: {jam_mulai}")
                elif isinstance(jam_mulai, timedelta):
                    # Jika timedelta, konversi ke time object
                    total_seconds = int(jam_mulai.total_seconds())
                    hours = total_seconds // 3600
                    minutes = (total_seconds % 3600) // 60
                    jam_mulai = datetime.strptime(f"{hours:02d}:{minutes:02d}:00", '%H:%M:%S').time()
                    print(f"Converted timedelta jam_mulai: {jam_mulai}")
                elif hasattr(jam_mulai, 'strftime'):
                    # Jika sudah time object, gunakan langsung
                    pass
                else:
                    # Jika tidak ada data jam dari database, return error
                    print(f"Error: jam_mulai tidak valid dari database: {jam_mulai} (type: {type(jam_mulai)})")
                    cursor.close()
                    return jsonify({'success': False, 'message': 'Data jam_mulai tidak valid dari database!'})
                
                if isinstance(jam_selesai, str):
                    # Coba parse sebagai datetime lengkap dulu
                    try:
                        jam_selesai = datetime.strptime(jam_selesai, '%Y-%m-%d %H:%M:%S')
                        jam_selesai = jam_selesai.time()
                        print(f"Parsed datetime string jam_selesai: {jam_selesai}")
                    except:
                        # Jika gagal, coba parse sebagai time saja
                        jam_selesai = datetime.strptime(jam_selesai, '%H:%M:%S').time()
                        print(f"Parsed time string jam_selesai: {jam_selesai}")
                elif isinstance(jam_selesai, datetime):
                    # Jika sudah datetime object, ambil bagian time saja
                    jam_selesai = jam_selesai.time()
                    print(f"Extracted time from datetime jam_selesai: {jam_selesai}")
                elif isinstance(jam_selesai, timedelta):
                    # Jika timedelta, konversi ke time object
                    total_seconds = int(jam_selesai.total_seconds())
                    hours = total_seconds // 3600
                    minutes = total_seconds % 3600 // 60
                    jam_selesai = datetime.strptime(f"{hours:02d}:{minutes:02d}:00", '%H:%M:%S').time()
                    print(f"Converted timedelta jam_selesai: {jam_selesai}")
                elif hasattr(jam_selesai, 'strftime'):
                    # Jika sudah time object, gunakan langsung
                    pass
                else:
                    # Jika tidak ada data jam dari database, return error
                    print(f"Error: jam_selesai tidak valid dari database: {jam_selesai} (type: {type(jam_selesai)})")
                    cursor.close()
                    return jsonify({'success': False, 'message': 'Data jam_selesai tidak valid dari database!'})
                
                # Konversi ke datetime untuk perhitungan
                today = datetime.now().date()
                current_dt = datetime.combine(today, jam_mulai)
                end_dt = datetime.combine(today, jam_selesai)
                
                print(f"=== DEBUG GENERATE JAM_LIST ===")
                print(f"Today: {today}")
                print(f"Jam mulai (time object): {jam_mulai}")
                print(f"Jam selesai (time object): {jam_selesai}")
                print(f"Current DT (start): {current_dt}")
                print(f"End DT: {end_dt}")
                print(f"Interval type: {jadwal_info['interval_tipe']}")
                print(f"Interval value: {interval}")
                print(f"===============================")
                
                # Generate jam_list berdasarkan interval yang sebenarnya dari database
                # PERBAIKAN: Handle semua jenis interval dengan benar
                if jadwal_info['interval_tipe'] == 'setiap_jam':
                    # PERBAIKAN: Handle kasus jam yang melewati tengah malam
                    while current_dt <= end_dt:
                        jam_list.append(current_dt.strftime('%H:%M'))
                        print(f"Added to jam_list: {current_dt.strftime('%H:%M')}")
                        current_dt += timedelta(hours=interval)
                        print(f"Added {interval} hours, new current_dt: {current_dt}")
                    
                    jadwal_info['jam_list'] = jam_list
                    
                    # PERBAIKAN: Jika jam_list kosong karena masalah tengah malam, generate manual
                    if not jam_list:
                        print("WARNING: jam_list kosong, generate manual berdasarkan metadata")
                        # Generate jam_list manual berdasarkan interval
                        current_time = jam_mulai
                        while current_time <= jam_selesai:
                            jam_list.append(current_time.strftime('%H:%M'))
                            current_time += timedelta(hours=interval)
                        jadwal_info['jam_list'] = jam_list
                
                elif jadwal_info['interval_tipe'] == 'harian':
                    # Generate label harian berdasarkan tanggal yang sebenarnya
                    hari_names = ['Senin', 'Selasa', 'Rabu', 'Kamis', 'Jumat', 'Sabtu', 'Minggu']
                    # Gunakan tanggal hari ini sebagai starting point
                    current_date = datetime.now().date()
                    urutan = 1
                    
                    # Generate 7 hari ke depan
                    for i in range(7):
                        hari_index = current_date.weekday()  # 0=Senin, 1=Selasa, dst
                        nama_hari = hari_names[hari_index]
                        jam_list.append(nama_hari)
                        current_date += timedelta(days=1)
                        urutan += 1
                    
                    jadwal_info['jam_list'] = jam_list
                
                elif jadwal_info['interval_tipe'] == 'mingguan':
                    # Generate label mingguan berdasarkan minggu dalam tahun
                    # Gunakan tanggal mulai dari jadwal sebagai starting point
                    if isinstance(jadwal_info['pembimbing_nilai_mulai'], str):
                        try:
                            start_date = datetime.strptime(jadwal_info['pembimbing_nilai_mulai'], '%Y-%m-%d %H:%M:%S').date()
                        except:
                            start_date = datetime.now().date()
                    elif isinstance(jadwal_info['pembimbing_nilai_mulai'], datetime):
                        start_date = jadwal_info['pembimbing_nilai_mulai'].date()
                    else:
                        start_date = datetime.now().date()
                    
                    current_date = start_date
                    urutan = 1
                    
                    # Generate minggu berdasarkan rentang jadwal yang sebenarnya
                    if isinstance(jadwal_info['pembimbing_nilai_selesai'], str):
                        try:
                            end_date = datetime.strptime(jadwal_info['pembimbing_nilai_selesai'], '%Y-%m-%d %H:%M:%S').date()
                        except:
                            end_date = start_date + timedelta(weeks=4)
                    elif isinstance(jadwal_info['pembimbing_nilai_selesai'], datetime):
                        end_date = jadwal_info['pembimbing_nilai_selesai'].date()
                    else:
                        end_date = start_date + timedelta(weeks=4)
                    
                    while current_date <= end_date:
                        # Hitung minggu ke berapa dalam tahun (1-52/53)
                        week_num = current_date.isocalendar()[1]  # ISO week number
                        jam_list.append(f"Minggu {week_num}")
                        current_date += timedelta(weeks=1)
                        urutan += 1
                    
                    jadwal_info['jam_list'] = jam_list
                
                elif jadwal_info['interval_tipe'] == 'bulanan':
                    # Generate label bulanan berdasarkan bulan yang sebenarnya
                    bulan_names = ['Januari', 'Februari', 'Maret', 'April', 'Mei', 'Juni', 
                                  'Juli', 'Agustus', 'September', 'Oktober', 'November', 'Desember']
                    
                    # Gunakan tanggal mulai dari jadwal sebagai starting point
                    if isinstance(jadwal_info['pembimbing_nilai_mulai'], str):
                        try:
                            start_date = datetime.strptime(jadwal_info['pembimbing_nilai_mulai'], '%Y-%m-%d %H:%M:%S').date()
                        except:
                            start_date = datetime.now().date()
                    elif isinstance(jadwal_info['pembimbing_nilai_mulai'], datetime):
                        start_date = jadwal_info['pembimbing_nilai_mulai'].date()
                    else:
                        start_date = datetime.now().date()
                    
                    current_date = start_date
                    urutan = 1
                    
                    # Generate bulan berdasarkan rentang jadwal yang sebenarnya
                    if isinstance(jadwal_info['pembimbing_nilai_selesai'], str):
                        try:
                            end_date = datetime.strptime(jadwal_info['pembimbing_nilai_selesai'], '%Y-%m-%d %H:%M:%S').date()
                        except:
                            end_date = start_date + timedelta(days=365)  # 1 tahun
                    elif isinstance(jadwal_info['pembimbing_nilai_selesai'], datetime):
                        end_date = jadwal_info['pembimbing_nilai_selesai'].date()
                    else:
                        end_date = start_date + timedelta(days=365)  # 1 tahun
                    
                    while current_date <= end_date:
                        bulan_index = current_date.month - 1  # 0-based index
                        nama_bulan = bulan_names[bulan_index]
                        jam_list.append(nama_bulan)
                        
                        # Pindah ke bulan berikutnya dengan aman
                        if current_date.month == 12:
                            current_date = current_date.replace(year=current_date.year + 1, month=1, day=1)
                        else:
                            # Gunakan tanggal 1 untuk menghindari masalah tanggal yang tidak valid
                            current_date = current_date.replace(month=current_date.month + 1, day=1)
                        urutan += 1
                    
                    jadwal_info['jam_list'] = jam_list
                
                else:
                    # Jika interval_tipe tidak dikenal, return error
                    cursor.close()
                    return jsonify({'success': False, 'message': f'Interval tipe tidak valid: {jadwal_info["interval_tipe"]}'})
                
                # Debug log untuk memastikan data benar
                print(f"=== FINAL RESULT ===")
                print(f"Generated jam_list from database: {jam_list}")
                print(f"Total jam in list: {len(jam_list)}")
                print(f"===============================")
                
            except Exception as e:
                print(f"Error generating jam_list: {e}")
                # Jika error, tidak ada fallback ke default, langsung return error
                cursor.close()
                return jsonify({'success': False, 'message': f'Error memproses jadwal dari database: {str(e)}'})
        else:
            # Jika jam_mulai atau jam_selesai None, gunakan jam_list kosong
            jadwal_info['jam_list'] = []
        
        # Pastikan semua data di jadwal_info bisa di-serialize ke JSON
        if jadwal_info:
            # Konversi semua field ke string atau tipe yang aman untuk JSON
            # TIDAK ADA FALLBACK KE DEFAULT - SEMUA HARUS DARI DATABASE
            if not jadwal_info['interval_tipe']:
                cursor.close()
                return jsonify({'success': False, 'message': 'Data interval_tipe tidak ditemukan di database!'})
            
            if not jadwal_info['jam_mulai']:
                cursor.close()
                return jsonify({'success': False, 'message': 'Data jam_mulai tidak ditemukan di database!'})
                
            if not jadwal_info['jam_selesai']:
                cursor.close()
                return jsonify({'success': False, 'message': 'Data jam_selesai tidak ditemukan di database!'})
            
            # Konversi ke tipe yang aman untuk JSON
            jadwal_info['interval_tipe'] = str(jadwal_info['interval_tipe'])
            jadwal_info['interval_nilai'] = int(jadwal_info['interval_nilai']) if jadwal_info['interval_nilai'] else 1
            jadwal_info['jam_mulai'] = str(jadwal_info['jam_mulai'])
            jadwal_info['jam_selesai'] = str(jadwal_info['jam_selesai'])
            jadwal_info['jam_list'] = jadwal_info.get('jam_list', [])
            
            # Hapus field yang tidak diperlukan dan bisa menyebabkan error JSON
            if 'jam_default_mulai' in jadwal_info:
                del jadwal_info['jam_default_mulai']
            if 'jam_default_selesai' in jadwal_info:
                del jadwal_info['jam_default_selesai']
            
            # Debug log final jadwal_info
            print(f"=== FINAL JSON CONVERSION ===")
            print(f"Final jadwal_info for JSON: {jadwal_info}")
            print(f"Jam mulai dari DB: {jadwal_info['jam_mulai']}")
            print(f"Jam selesai dari DB: {jadwal_info['jam_selesai']}")
            print(f"Interval tipe dari DB: {jadwal_info['interval_tipe']}")
            print(f"Interval nilai dari DB: {jadwal_info['interval_nilai']}")
            print(f"Jam list: {jadwal_info['jam_list']}")
            print(f"===============================")
        else:
            # Jika tidak ada data jadwal dari database, return error
            cursor.close()
            return jsonify({'success': False, 'message': 'Data jadwal tidak ditemukan di database!'})

        # ✅ PERBAIKAN: Ambil metadata kolom penilaian berdasarkan metadata_kolom_id yang tersimpan di detail_penilaian_mahasiswa
        try:
            # Buka ulang cursor karena ada data yang perlu diambil lagi
            cursor = get_app_functions()['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
            
            # ✅ PERBAIKAN: Ambil metadata_kolom_id yang digunakan dalam penilaian ini
            cursor.execute("""
                SELECT DISTINCT metadata_kolom_id 
                FROM detail_penilaian_mahasiswa 
                WHERE id_penilaian_mahasiswa = %s 
                AND metadata_kolom_id IS NOT NULL
                ORDER BY metadata_kolom_id
            """, (pen['id'],))
            
            metadata_ids = cursor.fetchall()
            print(f"=== DEBUG METADATA KOLOM ID ===")
            print(f"Found metadata_kolom_id: {[m['metadata_kolom_id'] for m in metadata_ids]}")
            
            if metadata_ids:
                # Ambil metadata kolom berdasarkan metadata_kolom_id yang tersimpan
                metadata_ids_list = [m['metadata_kolom_id'] for m in metadata_ids]
                cursor.execute("""
                    SELECT id, nama_kolom, urutan_kolom, interval_tipe 
                    FROM metadata_kolom_penilaian 
                    WHERE id IN ({})
                    ORDER BY urutan_kolom
                """.format(','.join(['%s'] * len(metadata_ids_list))), metadata_ids_list)
                
                metadata_kolom = cursor.fetchall()
                
                # Jika ada metadata kolom, gunakan sebagai jam_list
                if metadata_kolom:
                    jadwal_info['metadata_kolom'] = metadata_kolom
                    # Override jam_list dengan metadata kolom yang sebenarnya
                    jadwal_info['jam_list'] = [meta['nama_kolom'] for meta in metadata_kolom]
                    print(f"=== DEBUG METADATA KOLOM ===")
                    print(f"Found {len(metadata_kolom)} metadata columns for penilaian {pen['id']}")
                    print(f"Metadata kolom: {metadata_kolom}")
                print(f"Updated jam_list: {jadwal_info['jam_list']}")
                print(f"===============================")
            else:
                jadwal_info['metadata_kolom'] = []
                print(f"No metadata columns found for penilaian {pen['id']}, using generated jam_list")
        
        except Exception as e:
            print(f"Error getting metadata kolom: {e}")
            # Jika error, tetap lanjutkan dengan jam_list yang sudah digenerate
            jadwal_info['metadata_kolom'] = []
        
        cursor.close()
        
        # Hitung nilai akhir yang benar (rata-rata dari semua sesi)
        total_nilai = 0
        total_bobot = 0
        
        # ✅ PERBAIKAN: Organize data per kategori, pertanyaan, dan sesi dengan metadata yang benar
        kategori_data = {}
        for detail in details:
            kategori = detail['kategori']
            pertanyaan_id = detail['id_pertanyaan']
            metadata_kolom_id = detail['metadata_kolom_id']
            
            if kategori not in kategori_data:
                kategori_data[kategori] = {
                    'nama_kategori': get_kategori_display_name(kategori),
                    'pertanyaan_list': {},
                    'total_bobot_kategori': 0,
                    'total_nilai_kategori': 0
                }
            
            if pertanyaan_id not in kategori_data[kategori]['pertanyaan_list']:
                kategori_data[kategori]['pertanyaan_list'][pertanyaan_id] = {
                    'pertanyaan': detail['pertanyaan'],
                    'bobot': float(detail['bobot']),
                    'skor_maksimal': int(detail['skor_maksimal']),
                    'sesi_data': {},
                    'total_skor': 0,
                    'jumlah_sesi': 0,
                    'rata_rata_skor': 0,
                    'nilai_akhir': 0
                }
            
            # ✅ PERBAIKAN: Gunakan metadata_kolom_id sebagai key sesi untuk konsistensi
            sesi = detail['sesi_penilaian']
            kategori_data[kategori]['pertanyaan_list'][pertanyaan_id]['sesi_data'][sesi] = {
                'skor': int(detail['skor']),
                'nilai': float(detail['nilai']),
                'is_locked': detail['is_locked'],
                'tanggal_input': str(detail['tanggal_input']) if detail['tanggal_input'] else None,
                'metadata_kolom_id': metadata_kolom_id  # ✅ Tambahkan metadata_kolom_id untuk referensi
            }
            kategori_data[kategori]['pertanyaan_list'][pertanyaan_id]['total_skor'] += int(detail['skor'])
            kategori_data[kategori]['pertanyaan_list'][pertanyaan_id]['jumlah_sesi'] += 1
        
        # Hitung rata-rata dan nilai akhir untuk setiap pertanyaan dan kategori
        detail_penilaian = []
        for kategori, kategori_info in kategori_data.items():
            kategori_pertanyaan_list = []
            
            for pertanyaan_id, data in kategori_info['pertanyaan_list'].items():
                if data['jumlah_sesi'] > 0:
                    data['rata_rata_skor'] = data['total_skor'] / data['jumlah_sesi']
                    # Hitung nilai akhir berdasarkan rata-rata skor dan bobot
                    data['nilai_akhir'] = (data['rata_rata_skor'] / data['skor_maksimal']) * data['bobot']
                    total_nilai += data['nilai_akhir']
                    total_bobot += data['bobot']
                    
                    # Akumulasi untuk kategori
                    kategori_info['total_bobot_kategori'] += data['bobot']
                    kategori_info['total_nilai_kategori'] += data['nilai_akhir']
                
                kategori_pertanyaan_list.append({
                    'pertanyaan': data['pertanyaan'],
                    'bobot': data['bobot'],
                    'skor_maksimal': data['skor_maksimal'],
                    'rata_rata_skor': data['rata_rata_skor'],
                    'nilai_akhir': data['nilai_akhir'],
                    'sesi_data': data['sesi_data']
                })
            
            detail_penilaian.append({
                'kategori': kategori,
                'nama_kategori': kategori_info['nama_kategori'],
                'total_bobot_kategori': kategori_info['total_bobot_kategori'],
                'total_nilai_kategori': kategori_info['total_nilai_kategori'],
                'pertanyaan_list': kategori_pertanyaan_list
            })
        
        # Urutkan detail_penilaian berdasarkan urutan kategori yang benar
        detail_penilaian.sort(key=lambda x: get_kategori_order(x['kategori']))
        
        # Hitung nilai akhir total yang benar
        # Nilai akhir total = total_nilai (sudah dalam bentuk nilai akhir)
        nilai_akhir_total = total_nilai if total_bobot > 0 else 0
        
        # Debug: Log perhitungan untuk memastikan benar
        print(f"=== DEBUG PERHITUNGAN NILAI ===")
        print(f"Total bobot: {total_bobot}")
        print(f"Total nilai: {total_nilai}")
        print(f"Nilai akhir total: {nilai_akhir_total}")
        print(f"===============================")
        
        return jsonify({'success': True, 'data': {
            'mahasiswa': {
                'nim_mahasiswa': mhs['nim'],
                'nama_mahasiswa': mhs['nama_mahasiswa'],
                'judul_usaha': mhs['judul_usaha'],
                'program_studi': mhs['program_studi']
            },
            'penilaian': {
                'id': pen['id'],
                'nilai_akhir': nilai_akhir_total,
                'komentar_pembimbing': pen['komentar_pembimbing'],
                'status': pen['status'],
                'total_bobot': total_bobot,
                'total_nilai': total_nilai
            },
            'detail_penilaian': detail_penilaian,
            'jadwal_info': jadwal_info
        }})
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})


@admin_bp.route('/get_admin_detail_penilaian_proposal')
def get_admin_detail_penilaian_proposal():
    if 'user_type' not in session or session['user_type'] != 'admin':
        return jsonify({'success': False, 'message': 'Anda harus login sebagai admin!'})
    try:
        proposal_id = request.args.get('proposal_id')
        if not proposal_id:
            return jsonify({'success': False, 'message': 'ID proposal tidak ditemukan!'})
        cursor = get_app_functions()['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        cursor.execute('''
            SELECT m.nim, m.nama_ketua AS nama_mahasiswa, p.judul_usaha
            FROM proposal p INNER JOIN mahasiswa m ON m.nim = p.nim
            WHERE p.id = %s
        ''', (proposal_id,))
        info = cursor.fetchone()
        if not info:
            cursor.close()
            return jsonify({'success': False, 'message': 'Data proposal tidak ditemukan!'})
        cursor.execute('SELECT AVG(nilai_akhir) AS nilai_akhir FROM penilaian_proposal WHERE id_proposal = %s', (proposal_id,))
        nilai_row = cursor.fetchone()
        nilai = nilai_row['nilai_akhir'] or 0
        cursor.execute('''
            SELECT dpp.id_pertanyaan, AVG(dpp.skor) AS skor, AVG(dpp.nilai) AS nilai
            FROM detail_penilaian_proposal dpp
            WHERE dpp.id_proposal = %s
            GROUP BY dpp.id_pertanyaan
        ''', (proposal_id,))
        detail_rows = cursor.fetchall()
        cursor.execute('''
            SELECT id, pertanyaan, bobot, skor_maksimal
            FROM pertanyaan_penilaian_proposal
            WHERE is_active = TRUE
        ''')
        master = {row['id']: row for row in cursor.fetchall()}
        cursor.execute('SELECT catatan, nilai_bantuan, persentase_bantuan FROM penilaian_proposal WHERE id_proposal = %s', (proposal_id,))
        notes = cursor.fetchall()
        cursor.close()
        details = []
        total_bobot = total_skor = total_nilai = 0.0
        for d in detail_rows:
            m = master.get(d['id_pertanyaan'])
            if not m:
                continue
            details.append({
                'pertanyaan': m['pertanyaan'],
                'bobot': float(m['bobot']),
                'skor': float(d['skor'] or 0.0),
                'skor_maksimal': int(m['skor_maksimal']),
                'nilai': float(d['nilai'] or 0.0)
            })
            total_bobot += float(m['bobot'])
            total_skor += float(d['skor'] or 0.0)
            total_nilai += float(d['nilai'] or 0.0)
        return jsonify({'success': True, 'data': {
            'mahasiswa': info,
            'penilaian': {'nilai_akhir': float(nilai)},
            'detail_penilaian': details,
            'total': {
                'bobot': round(total_bobot, 2),
                'skor': round(total_skor, 2),
                'nilai': round(total_nilai, 2)
            },
            'catatan': [n['catatan'] for n in notes if n['catatan']],
            'rekomendasi': [
                {
                    'rekomendasi': f"Rp {round(sum((float(n['nilai_bantuan']) if n['nilai_bantuan'] is not None else 0.0) for n in notes) / max(len(notes), 1), 2):,.0f} ({round(sum((float(n['persentase_bantuan']) if n['persentase_bantuan'] is not None else 0.0) for n in notes) / max(len(notes), 1), 2)}%)"
                }
            ]
        }})
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})


@admin_bp.route('/get_admin_detail_penilaian_kemajuan')
def get_admin_detail_penilaian_kemajuan():
    if 'user_type' not in session or session['user_type'] != 'admin':
        return jsonify({'success': False, 'message': 'Anda harus login sebagai admin!'})
    try:
        proposal_id = request.args.get('proposal_id')
        if not proposal_id:
            return jsonify({'success': False, 'message': 'ID proposal tidak ditemukan!'})
        cursor = get_app_functions()['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        cursor.execute('''
            SELECT m.nim, m.nama_ketua AS nama_mahasiswa, p.judul_usaha
            FROM proposal p INNER JOIN mahasiswa m ON m.nim = p.nim
            WHERE p.id = %s
        ''', (proposal_id,))
        info = cursor.fetchone()
        cursor.execute('SELECT AVG(nilai_akhir) AS nilai_akhir FROM penilaian_laporan_kemajuan WHERE id_proposal = %s', (proposal_id,))
        nilai_row = cursor.fetchone()
        nilai = nilai_row['nilai_akhir'] or 0
        cursor.execute('''
            SELECT d.id_pertanyaan, AVG(d.skor_diberikan) AS skor, AVG(d.nilai_terbobot) AS nilai
            FROM detail_penilaian_laporan_kemajuan d
            INNER JOIN penilaian_laporan_kemajuan p ON d.id_penilaian_laporan_kemajuan = p.id
            WHERE p.id_proposal = %s
            GROUP BY d.id_pertanyaan
        ''', (proposal_id,))
        detail_rows = cursor.fetchall()
        cursor.execute('SELECT id, pertanyaan, bobot, skor_maksimal FROM pertanyaan_penilaian_laporan_kemajuan')
        master = {row['id']: row for row in cursor.fetchall()}
        cursor.execute('SELECT komentar_reviewer, rekomendasi_pendanaan, alasan_rekomendasi FROM penilaian_laporan_kemajuan WHERE id_proposal = %s', (proposal_id,))
        notes = cursor.fetchall()
        cursor.close()
        details = []
        total_bobot = total_skor = total_nilai = 0.0
        for d in detail_rows:
            m = master.get(d['id_pertanyaan'])
            if not m:
                continue
            details.append({
                'pertanyaan': m['pertanyaan'],
                'bobot': float(m['bobot']),
                'skor': float(d['skor'] or 0.0),
                'skor_maksimal': int(m['skor_maksimal']),
                'nilai': float(d['nilai'] or 0.0)
            })
            total_bobot += float(m['bobot'])
            total_skor += float(d['skor'] or 0.0)
            total_nilai += float(d['nilai'] or 0.0)
        return jsonify({'success': True, 'data': {
            'mahasiswa': info,
            'penilaian': {'nilai_akhir': float(nilai)},
            'detail_penilaian': details,
            'total': {
                'bobot': round(total_bobot, 2),
                'skor': round(total_skor, 2),
                'nilai': round(total_nilai, 2)
            },
            'catatan': [n['komentar_reviewer'] for n in notes if n['komentar_reviewer']],
            'rekomendasi': [
                {
                    'rekomendasi_pendanaan': n['rekomendasi_pendanaan'],
                    'alasan_rekomendasi': n['alasan_rekomendasi']
                } for n in notes if n['rekomendasi_pendanaan'] or n['alasan_rekomendasi']
            ]
        }})
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})


@admin_bp.route('/get_admin_detail_penilaian_akhir')
def get_admin_detail_penilaian_akhir():
    if 'user_type' not in session or session['user_type'] != 'admin':
        return jsonify({'success': False, 'message': 'Anda harus login sebagai admin!'})
    try:
        proposal_id = request.args.get('proposal_id')
        if not proposal_id:
            return jsonify({'success': False, 'message': 'ID proposal tidak ditemukan!'})
        cursor = get_app_functions()['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        cursor.execute('''
            SELECT m.nim, m.nama_ketua AS nama_mahasiswa, p.judul_usaha
            FROM proposal p INNER JOIN mahasiswa m ON m.nim = p.nim
            WHERE p.id = %s
        ''', (proposal_id,))
        info = cursor.fetchone()
        cursor.execute('SELECT AVG(nilai_akhir) AS nilai_akhir FROM penilaian_laporan_akhir WHERE id_proposal = %s', (proposal_id,))
        nilai_row = cursor.fetchone()
        nilai = nilai_row['nilai_akhir'] or 0
        cursor.execute('''
            SELECT d.id_pertanyaan, AVG(d.skor_diberikan) AS skor, AVG(d.nilai_terbobot) AS nilai
            FROM detail_penilaian_laporan_akhir d
            INNER JOIN penilaian_laporan_akhir p ON d.id_penilaian_laporan_akhir = p.id
            WHERE p.id_proposal = %s
            GROUP BY d.id_pertanyaan
        ''', (proposal_id,))
        detail_rows = cursor.fetchall()
        cursor.execute('SELECT id, pertanyaan, bobot, skor_maksimal FROM pertanyaan_penilaian_laporan_akhir')
        master = {row['id']: row for row in cursor.fetchall()}
        cursor.execute('SELECT komentar_reviewer, rekomendasi_kelulusan, alasan_rekomendasi FROM penilaian_laporan_akhir WHERE id_proposal = %s', (proposal_id,))
        notes = cursor.fetchall()
        cursor.close()
        details = []
        total_bobot = total_skor = total_nilai = 0.0
        for d in detail_rows:
            m = master.get(d['id_pertanyaan'])
            if not m:
                continue
            details.append({
                'pertanyaan': m['pertanyaan'],
                'bobot': float(m['bobot']),
                'skor': float(d['skor'] or 0.0),
                'skor_maksimal': int(m['skor_maksimal']),
                'nilai': float(d['nilai'] or 0.0)
            })
            total_bobot += float(m['bobot'])
            total_skor += float(d['skor'] or 0.0)
            total_nilai += float(d['nilai'] or 0.0)
        return jsonify({'success': True, 'data': {
            'mahasiswa': info,
            'penilaian': {'nilai_akhir': float(nilai)},
            'detail_penilaian': details,
            'total': {
                'bobot': round(total_bobot, 2),
                'skor': round(total_skor, 2),
                'nilai': round(total_nilai, 2)
            },
            'catatan': [n['komentar_reviewer'] for n in notes if n['komentar_reviewer']],
            'rekomendasi': [
                {
                    'rekomendasi_kelulusan': n['rekomendasi_kelulusan'],
                    'alasan_rekomendasi': n['alasan_rekomendasi']
                } for n in notes if n['rekomendasi_kelulusan'] or n['alasan_rekomendasi']
            ]
        }})
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

# ========================================
# FUNGSI HELPER UNTUK PENILAIAN
# ========================================

@admin_bp.route('/daftar_bimbingan_mahasiswa')
def admin_daftar_bimbingan_mahasiswa():
    if 'user_type' not in session or session['user_type'] != 'admin':
        return redirect(url_for('login'))
    
    try:
        cursor = get_app_functions()['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Ambil daftar mahasiswa dengan bimbingan
        cursor.execute('''
            SELECT DISTINCT m.id, m.nama_ketua, m.nim, m.program_studi, p.judul_usaha, p.dosen_pembimbing
            FROM mahasiswa m
            INNER JOIN proposal p ON m.nim = p.nim
            INNER JOIN bimbingan b ON m.nim = b.nim
            WHERE p.status_admin = 'lolos'
            ORDER BY m.nama_ketua ASC
        ''')
        
        daftar_mahasiswa = cursor.fetchall()
        cursor.close()
        
        return render_template('admin/daftar_bimbingan_mahasiswa.html', daftar_mahasiswa=daftar_mahasiswa)
        
    except Exception as e:
        return f"Error: {str(e)}"

@admin_bp.route('/bimbingan_mahasiswa/<int:mahasiswa_id>')
def admin_bimbingan_mahasiswa(mahasiswa_id):
    if 'user_type' not in session or session['user_type'] != 'admin':
        return redirect(url_for('login'))
    
    try:
        cursor = get_app_functions()['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Ambil info mahasiswa
        cursor.execute('''
            SELECT m.*, p.judul_usaha, p.dosen_pembimbing
            FROM mahasiswa m
            LEFT JOIN proposal p ON m.nim = p.nim
            WHERE m.id = %s AND p.status_admin = 'lolos'
        ''', (mahasiswa_id,))
        
        mahasiswa_info = cursor.fetchone()
        
        if not mahasiswa_info:
            return "Mahasiswa tidak ditemukan"
        
        # Ambil riwayat bimbingan
        cursor.execute('''
            SELECT b.*, p.judul_usaha, p.dosen_pembimbing
            FROM bimbingan b
            LEFT JOIN proposal p ON b.proposal_id = p.id
            WHERE b.nim = %s
            ORDER BY b.tanggal_buat DESC
        ''', (mahasiswa_info['nim'],))
        
        riwayat_bimbingan = cursor.fetchall()
        cursor.close()
        
        return render_template('admin/bimbingan_mahasiswa.html', 
                             mahasiswa_info=mahasiswa_info, 
                             riwayat_bimbingan=riwayat_bimbingan)
        
    except Exception as e:
        return f"Error: {str(e)}"

@admin_bp.route('/detail_bimbingan')
def admin_detail_bimbingan():
    if 'user_type' not in session or session['user_type'] != 'admin':
        return jsonify({'success': False, 'message': 'Akses ditolak!'})
    
    try:
        bimbingan_id = request.args.get('id')
        
        if not bimbingan_id:
            return jsonify({'success': False, 'message': 'ID bimbingan tidak ditemukan!'})
        
        cursor = get_app_functions()['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        cursor.execute('''
            SELECT b.*, p.judul_usaha, p.dosen_pembimbing, m.nama_ketua
            FROM bimbingan b
            LEFT JOIN proposal p ON b.proposal_id = p.id
            LEFT JOIN mahasiswa m ON b.nim = m.nim
            WHERE b.id = %s
        ''', (bimbingan_id,))
        
        bimbingan = cursor.fetchone()
        cursor.close()
        
        if not bimbingan:
            return jsonify({'success': False, 'message': 'Data bimbingan tidak ditemukan!'})
        
        # Render HTML untuk modal
        html_content = f'''
        <div class="row">
            <div class="col-12">
                <div class="mb-3">
                    <h6 class="text-primary"><i class="bi bi-person-circle"></i> Informasi Mahasiswa</h6>
                    <p><strong>Nama:</strong> {bimbingan['nama_ketua']}</p>
                    <p><strong>Judul Usaha:</strong> {bimbingan['judul_usaha']}</p>
                    <p><strong>Dosen Pembimbing:</strong> {bimbingan['dosen_pembimbing']}</p>
                </div>
                
                <div class="mb-3">
                    <h6 class="text-primary"><i class="bi bi-chat-dots"></i> Detail Bimbingan</h6>
                    <p><strong>Judul/Topik:</strong> {bimbingan['judul_bimbingan']}</p>
                    <p><strong>Tanggal:</strong> {bimbingan['tanggal_buat'].strftime('%d/%m/%Y %H:%M') if bimbingan['tanggal_buat'] else '-'}</p>
                </div>
                
                <div class="mb-3">
                    <h6 class="text-primary"><i class="bi bi-file-text"></i> Hasil Bimbingan</h6>
                    <div class="hasil-bimbingan-content">
                        {bimbingan['hasil_bimbingan']}
                    </div>
                </div>
            </div>
        </div>
        '''
        
        return jsonify({'success': True, 'html': html_content})
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})


@admin_bp.route('/admin_hapus_bimbingan', methods=['POST'])
def admin_hapus_bimbingan():
    """API untuk menghapus bimbingan"""
    if 'user_type' not in session or session['user_type'] != 'admin':
        return jsonify({'success': False, 'message': 'Akses ditolak!'})
    
    try:
        bimbingan_id = request.form.get('id')
        
        if not bimbingan_id:
            return jsonify({'success': False, 'message': 'ID bimbingan tidak ditemukan!'})
        
        cursor = get_app_functions()['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Cek apakah bimbingan ada
        cursor.execute('SELECT * FROM bimbingan WHERE id = %s', (bimbingan_id,))
        bimbingan = cursor.fetchone()
        
        if not bimbingan:
            return jsonify({'success': False, 'message': 'Data bimbingan tidak ditemukan!'})
        
        # Hapus bimbingan
        cursor.execute('DELETE FROM bimbingan WHERE id = %s', (bimbingan_id,))
        get_app_functions()['mysql'].connection.commit()
        
        cursor.close()
        
        return jsonify({
            'success': True, 
            'message': 'Bimbingan berhasil dihapus!'
        })
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})


# ========================================
# PENGATURAN BIMBINGAN
# ========================================

@admin_bp.route('/get_pengaturan_bimbingan')
def get_pengaturan_bimbingan():
    """Ambil pengaturan bimbingan"""
    if 'user_type' not in session or session['user_type'] != 'admin':
        return jsonify({'success': False, 'message': 'Anda harus login sebagai admin!'})
    
    try:
        cursor = get_app_functions()['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)

        # Pastikan tabel ada (tanpa akses information_schema)
        cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS pengaturan_bimbingan (
                id INT AUTO_INCREMENT PRIMARY KEY,
                maksimal_bimbingan_per_hari INT DEFAULT 3,
                pesan_batasan TEXT DEFAULT 'Anda telah mencapai batas maksimal bimbingan hari ini. Silakan coba lagi besok.',
                status_aktif ENUM('aktif', 'nonaktif') DEFAULT 'aktif',
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP
            )
            """
        )
        get_app_functions()['mysql'].connection.commit()

        # Jika belum ada baris konfigurasi, sisipkan default
        cursor.execute('SELECT COUNT(*) as count FROM pengaturan_bimbingan')
        if cursor.fetchone()['count'] == 0:
            cursor.execute(
                """
                INSERT INTO pengaturan_bimbingan (maksimal_bimbingan_per_hari, pesan_batasan, status_aktif)
                VALUES (3, 'Anda telah mencapai batas maksimal bimbingan hari ini. Silakan coba lagi besok.', 'aktif')
                """
            )
            get_app_functions()['mysql'].connection.commit()
        
        # Ambil pengaturan bimbingan
        cursor.execute('SELECT * FROM pengaturan_bimbingan ORDER BY id DESC LIMIT 1')
        pengaturan = cursor.fetchone()
        
        cursor.close()
        
        if not pengaturan:
            return jsonify({
                'success': True,
                'data': {
                    'maksimal_bimbingan_per_hari': 3,
                    'pesan_batasan': 'Anda telah mencapai batas maksimal bimbingan hari ini. Silakan coba lagi besok.',
                    'status_aktif': 'aktif'
                }
            })
        
        return jsonify({
            'success': True,
            'data': {
                'maksimal_bimbingan_per_hari': pengaturan['maksimal_bimbingan_per_hari'],
                'pesan_batasan': pengaturan['pesan_batasan'],
                'status_aktif': pengaturan['status_aktif']
            }
        })
        
    except Exception as e:
        print(f"Error in get_pengaturan_bimbingan: {e}")
        return jsonify({'success': False, 'message': f'Error saat mengambil pengaturan bimbingan: {str(e)}'})

@admin_bp.route('/update_pengaturan_bimbingan', methods=['POST'])
def update_pengaturan_bimbingan():
    """Update pengaturan bimbingan"""
    print(f"DEBUG: Session data: {session}")
    print(f"DEBUG: User type: {session.get('user_type', 'Not set')}")
    
    if 'user_type' not in session or session['user_type'] != 'admin':
        print(f"DEBUG: Access denied - user_type: {session.get('user_type', 'Not set')}")
        return jsonify({'success': False, 'message': 'Anda harus login sebagai admin!'})
    
    try:
        print(f"DEBUG: Form data received: {request.form}")
        maksimal_bimbingan_per_hari = request.form.get('maksimal_bimbingan_per_hari')
        status_aktif = request.form.get('status_aktif')
        
        print(f"DEBUG: maksimal_bimbingan_per_hari: {maksimal_bimbingan_per_hari}")
        print(f"DEBUG: status_aktif: {status_aktif}")
        
        if not maksimal_bimbingan_per_hari or not status_aktif:
            print(f"DEBUG: Missing required fields")
            return jsonify({'success': False, 'message': 'Semua field harus diisi!'})
        
        # Validasi input
        try:
            maksimal_bimbingan_per_hari = int(maksimal_bimbingan_per_hari)
            if maksimal_bimbingan_per_hari < 1 or maksimal_bimbingan_per_hari > 10:
                print(f"DEBUG: Invalid range: {maksimal_bimbingan_per_hari}")
                return jsonify({'success': False, 'message': 'Maksimal bimbingan per hari harus antara 1-10!'})
        except ValueError:
            print(f"DEBUG: Invalid number format: {maksimal_bimbingan_per_hari}")
            return jsonify({'success': False, 'message': 'Maksimal bimbingan per hari harus berupa angka!'})
        
        if status_aktif not in ['aktif', 'nonaktif']:
            print(f"DEBUG: Invalid status: {status_aktif}")
            return jsonify({'success': False, 'message': 'Status aktif tidak valid!'})
        
        print(f"DEBUG: Starting database operations")
        cursor = get_app_functions()['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)

        # Pastikan tabel ada (tanpa akses information_schema)
        print(f"DEBUG: Creating table if not exists")
        cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS pengaturan_bimbingan (
                id INT AUTO_INCREMENT PRIMARY KEY,
                maksimal_bimbingan_per_hari INT DEFAULT 3,
                pesan_batasan TEXT DEFAULT 'Anda telah mencapai batas maksimal bimbingan hari ini. Silakan coba lagi besok.',
                status_aktif ENUM('aktif', 'nonaktif') DEFAULT 'aktif',
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP
            )
            """
        )
        get_app_functions()['mysql'].connection.commit()
        print(f"DEBUG: Table created/verified")
        
        # Cek apakah sudah ada pengaturan
        cursor.execute('SELECT COUNT(*) as count FROM pengaturan_bimbingan')
        count = cursor.fetchone()['count']
        
        # Generate pesan_batasan otomatis berdasarkan maksimal_bimbingan_per_hari
        auto_pesan = f"Anda telah mencapai batas maksimal bimbingan hari ini. Maksimal {maksimal_bimbingan_per_hari} kali per hari. Silakan coba lagi besok."

        if count == 0:
            # Insert pengaturan baru
            cursor.execute("""
                INSERT INTO pengaturan_bimbingan (maksimal_bimbingan_per_hari, pesan_batasan, status_aktif) 
                VALUES (%s, %s, %s)
            """, (maksimal_bimbingan_per_hari, auto_pesan, status_aktif))
        else:
            # Update pengaturan yang ada
            cursor.execute("""
                UPDATE pengaturan_bimbingan 
                SET maksimal_bimbingan_per_hari = %s, pesan_batasan = %s, status_aktif = %s 
                WHERE id = (SELECT id FROM (SELECT id FROM pengaturan_bimbingan ORDER BY id DESC LIMIT 1) as temp)
            """, (maksimal_bimbingan_per_hari, auto_pesan, status_aktif))
        
        get_app_functions()['mysql'].connection.commit()
        cursor.close()
        
        return jsonify({
            'success': True,
            'message': 'Pengaturan bimbingan berhasil diperbarui!'
        })
        
    except Exception as e:
        print(f"Error in update_pengaturan_bimbingan: {e}")
        return jsonify({'success': False, 'message': f'Error saat memperbarui pengaturan bimbingan: {str(e)}'})

# ========================================
# END PENGATURAN BIMBINGAN
# ========================================


# ========================================
# PENGATURAN JADWAL (Proposal, Kemajuan, Akhir)
# ========================================

@admin_bp.route('/get_pengaturan_jadwal')
def get_pengaturan_jadwal():
    """Ambil pengaturan jadwal (mulai/selesai) untuk proposal, laporan kemajuan, dan laporan akhir"""
    if 'user_type' not in session or session['user_type'] != 'admin':
        return jsonify({'success': False, 'message': 'Anda harus login sebagai admin!'})

    try:
        cursor = get_app_functions()['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)

        cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS `pengaturan_jadwal` (
            `id` int(11) NOT NULL AUTO_INCREMENT PRIMARY KEY,
            `proposal_mulai` datetime DEFAULT NULL,
            `proposal_selesai` datetime DEFAULT NULL,
            `kemajuan_mulai` datetime DEFAULT NULL,
            `kemajuan_selesai` datetime DEFAULT NULL,
            `akhir_mulai` datetime DEFAULT NULL,
            `akhir_selesai` datetime DEFAULT NULL,
            `created_at` timestamp NOT NULL DEFAULT current_timestamp(),
            `updated_at` timestamp NOT NULL DEFAULT current_timestamp() ON UPDATE current_timestamp(),
            `pembimbing_nilai_mulai` datetime DEFAULT NULL,
            `pembimbing_nilai_selesai` datetime DEFAULT NULL,
            `pembimbing_interval_tipe` enum('harian','mingguan','bulanan','setiap_jam') DEFAULT 'harian',
            `pembimbing_interval_nilai` int(11),
            `pembimbing_hari_aktif` varchar(20),
            `reviewer_proposal_mulai` datetime DEFAULT NULL,
            `reviewer_proposal_selesai` datetime DEFAULT NULL,
            `reviewer_kemajuan_mulai` datetime DEFAULT NULL,
            `reviewer_kemajuan_selesai` datetime DEFAULT NULL,
            `reviewer_akhir_mulai` datetime DEFAULT NULL,
            `reviewer_akhir_selesai` datetime DEFAULT NULL
            )
            """
        )
        get_app_functions()['mysql'].connection.commit()

        # Pastikan kolom-kolom jadwal penilaian tersedia (untuk kompatibilitas versi lama)
        def ensure_column(column_name: str, add_sql: str) -> None:
            cursor.execute("SHOW COLUMNS FROM pengaturan_jadwal LIKE %s", (column_name,))
            if cursor.fetchone() is None:
                cursor.execute(f"ALTER TABLE pengaturan_jadwal ADD COLUMN {add_sql}")
                get_app_functions()['mysql'].connection.commit()

        ensure_column('pembimbing_nilai_mulai', 'pembimbing_nilai_mulai DATETIME NULL')
        ensure_column('pembimbing_nilai_selesai', 'pembimbing_nilai_selesai DATETIME NULL')
        ensure_column('pembimbing_interval_tipe', "pembimbing_interval_tipe ENUM('harian','mingguan','bulanan','setiap_jam') DEFAULT 'harian'")
        ensure_column('pembimbing_interval_nilai', 'pembimbing_interval_nilai INT(11)')

        ensure_column('pembimbing_hari_aktif', "pembimbing_hari_aktif VARCHAR(20)")
        ensure_column('reviewer_proposal_mulai', 'reviewer_proposal_mulai DATETIME NULL')
        ensure_column('reviewer_proposal_selesai', 'reviewer_proposal_selesai DATETIME NULL')
        ensure_column('reviewer_kemajuan_mulai', 'reviewer_kemajuan_mulai DATETIME NULL')
        ensure_column('reviewer_kemajuan_selesai', 'reviewer_kemajuan_selesai DATETIME NULL')
        ensure_column('reviewer_akhir_mulai', 'reviewer_akhir_mulai DATETIME NULL')
        ensure_column('reviewer_akhir_selesai', 'reviewer_akhir_selesai DATETIME NULL')

        cursor.execute('SELECT * FROM pengaturan_jadwal ORDER BY id DESC LIMIT 1')
        row = cursor.fetchone()

        if not row:
            cursor.execute('INSERT INTO pengaturan_jadwal (proposal_mulai, proposal_selesai, kemajuan_mulai, kemajuan_selesai, akhir_mulai, akhir_selesai) VALUES (NULL, NULL, NULL, NULL, NULL, NULL)')
            get_app_functions()['mysql'].connection.commit()
            cursor.execute('SELECT * FROM pengaturan_jadwal ORDER BY id DESC LIMIT 1')
            row = cursor.fetchone()

        cursor.close()

        def fmt(dt):
            if not dt:
                return None
            # Kembalikan dalam format yang mudah diisi ke input datetime-local (YYYY-MM-DDTHH:MM)
            try:
                return dt.strftime('%Y-%m-%dT%H:%M')
            except Exception:
                return None



        return jsonify({
            'success': True,
            'data': {
                'proposal_mulai': fmt(row.get('proposal_mulai')),
                'proposal_selesai': fmt(row.get('proposal_selesai')),
                'kemajuan_mulai': fmt(row.get('kemajuan_mulai')),
                'kemajuan_selesai': fmt(row.get('kemajuan_selesai')),
                'akhir_mulai': fmt(row.get('akhir_mulai')),
                'akhir_selesai': fmt(row.get('akhir_selesai')),
                # Jadwal Penilaian
                'pembimbing_nilai_mulai': fmt(row.get('pembimbing_nilai_mulai')),
                'pembimbing_nilai_selesai': fmt(row.get('pembimbing_nilai_selesai')),
                'pembimbing_interval_tipe': row.get('pembimbing_interval_tipe', 'harian'),
                'pembimbing_interval_nilai': row.get('pembimbing_interval_nilai', ''),

                'pembimbing_hari_aktif': row.get('pembimbing_hari_aktif', ''),
                'reviewer_proposal_mulai': fmt(row.get('reviewer_proposal_mulai')),
                'reviewer_proposal_selesai': fmt(row.get('reviewer_proposal_selesai')),
                'reviewer_kemajuan_mulai': fmt(row.get('reviewer_kemajuan_mulai')),
                'reviewer_kemajuan_selesai': fmt(row.get('reviewer_kemajuan_selesai')),
                'reviewer_akhir_mulai': fmt(row.get('reviewer_akhir_mulai')),
                'reviewer_akhir_selesai': fmt(row.get('reviewer_akhir_selesai')),
            }
        })
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error saat mengambil pengaturan jadwal: {str(e)}'})


@admin_bp.route('/update_pengaturan_jadwal', methods=['POST'])
def update_pengaturan_jadwal():
    """Update pengaturan jadwal mulai/selesai"""
    if 'user_type' not in session or session['user_type'] != 'admin':
        return jsonify({'success': False, 'message': 'Anda harus login sebagai admin!'})

    try:
        # Deteksi grup form yang disubmit supaya field yang tidak terkirim di-grup itu tetap di-set NULL
        schedule_keys = ['proposal_mulai','proposal_selesai','kemajuan_mulai','kemajuan_selesai','akhir_mulai','akhir_selesai']
        penilaian_keys = [
            'pembimbing_nilai_mulai','pembimbing_nilai_selesai',
            'pembimbing_interval_tipe','pembimbing_interval_nilai',
            'pembimbing_hari_aktif',
            'reviewer_proposal_mulai','reviewer_proposal_selesai',
            'reviewer_kemajuan_mulai','reviewer_kemajuan_selesai',
            'reviewer_akhir_mulai','reviewer_akhir_selesai'
        ]

        schedule_group_present = any(k in request.form for k in schedule_keys)
        penilaian_group_present = any(k in request.form for k in penilaian_keys)

        def get_val_force(key: str, force_group: bool):
            # Jika form grup terkait disubmit, paksa key ada ('' jika tidak diisi) agar bisa di-set NULL
            if force_group:
                return request.form.get(key, '')
            # Jika bukan bagian grup yang disubmit, jangan sentuh kolom tersebut
            return None

        proposal_mulai = get_val_force('proposal_mulai', schedule_group_present)
        proposal_selesai = get_val_force('proposal_selesai', schedule_group_present)
        kemajuan_mulai = get_val_force('kemajuan_mulai', schedule_group_present)
        kemajuan_selesai = get_val_force('kemajuan_selesai', schedule_group_present)
        akhir_mulai = get_val_force('akhir_mulai', schedule_group_present)
        akhir_selesai = get_val_force('akhir_selesai', schedule_group_present)
        # Jadwal Penilaian
        pembimbing_nilai_mulai = get_val_force('pembimbing_nilai_mulai', penilaian_group_present)
        pembimbing_nilai_selesai = get_val_force('pembimbing_nilai_selesai', penilaian_group_present)
        pembimbing_interval_tipe = get_val_force('pembimbing_interval_tipe', penilaian_group_present)
        pembimbing_interval_nilai = get_val_force('pembimbing_interval_nilai', penilaian_group_present)

        pembimbing_hari_aktif = get_val_force('pembimbing_hari_aktif', penilaian_group_present)
        reviewer_proposal_mulai = get_val_force('reviewer_proposal_mulai', penilaian_group_present)
        reviewer_proposal_selesai = get_val_force('reviewer_proposal_selesai', penilaian_group_present)
        reviewer_kemajuan_mulai = get_val_force('reviewer_kemajuan_mulai', penilaian_group_present)
        reviewer_kemajuan_selesai = get_val_force('reviewer_kemajuan_selesai', penilaian_group_present)
        reviewer_akhir_mulai = get_val_force('reviewer_akhir_mulai', penilaian_group_present)
        reviewer_akhir_selesai = get_val_force('reviewer_akhir_selesai', penilaian_group_present)

        cursor = get_app_functions()['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)

        cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS pengaturan_jadwal (
                id INT AUTO_INCREMENT PRIMARY KEY,
                proposal_mulai DATETIME NULL,
                proposal_selesai DATETIME NULL,
                kemajuan_mulai DATETIME NULL,
                kemajuan_selesai DATETIME NULL,
                akhir_mulai DATETIME NULL,
                akhir_selesai DATETIME NULL,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP
            )
            """
        )
        get_app_functions()['mysql'].connection.commit()

        # Pastikan kolom-kolom jadwal penilaian tersedia
        def ensure_column(column_name: str, add_sql: str) -> None:
            cursor.execute("SHOW COLUMNS FROM pengaturan_jadwal LIKE %s", (column_name,))
            if cursor.fetchone() is None:
                cursor.execute(f"ALTER TABLE pengaturan_jadwal ADD COLUMN {add_sql}")
                get_app_functions()['mysql'].connection.commit()

        ensure_column('pembimbing_nilai_mulai', 'pembimbing_nilai_mulai DATETIME NULL')
        ensure_column('pembimbing_nilai_selesai', 'pembimbing_nilai_selesai DATETIME NULL')
        ensure_column('pembimbing_interval_tipe', "pembimbing_interval_tipe ENUM('harian','mingguan','bulanan','setiap_jam') DEFAULT 'harian'")
        ensure_column('pembimbing_interval_nilai', 'pembimbing_interval_nilai INT(11)')

        ensure_column('pembimbing_hari_aktif', "pembimbing_hari_aktif VARCHAR(20)")
        ensure_column('reviewer_proposal_mulai', 'reviewer_proposal_mulai DATETIME NULL')
        ensure_column('reviewer_proposal_selesai', 'reviewer_proposal_selesai DATETIME NULL')
        ensure_column('reviewer_kemajuan_mulai', 'reviewer_kemajuan_mulai DATETIME NULL')
        ensure_column('reviewer_kemajuan_selesai', 'reviewer_kemajuan_selesai DATETIME NULL')
        ensure_column('reviewer_akhir_mulai', 'reviewer_akhir_mulai DATETIME NULL')
        ensure_column('reviewer_akhir_selesai', 'reviewer_akhir_selesai DATETIME NULL')

        # Pastikan kolom jadwal_id ada di metadata_kolom_penilaian
        cursor.execute("SHOW COLUMNS FROM metadata_kolom_penilaian LIKE 'jadwal_id'")
        if cursor.fetchone() is None:
            cursor.execute("ALTER TABLE metadata_kolom_penilaian ADD COLUMN jadwal_id INT")
            get_app_functions()['mysql'].connection.commit()

        # Buat tabel metadata kolom penilaian jika belum ada
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS metadata_kolom_penilaian (
                id INT PRIMARY KEY AUTO_INCREMENT,
                jadwal_id INT,
                periode_interval_id INT DEFAULT 1,
                nama_kolom VARCHAR(100) NOT NULL,
                urutan_kolom INT NOT NULL,
                interval_tipe ENUM('harian','mingguan','bulanan','setiap_jam') NOT NULL,
                interval_nilai INT NOT NULL,
                tanggal_mulai DATE,
                tanggal_selesai DATE,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                INDEX idx_interval_tipe (interval_tipe),
                INDEX idx_periode (periode_interval_id),
                INDEX idx_jadwal (jadwal_id)
            )
        """)
        get_app_functions()['mysql'].connection.commit()
        
        # Tambah kolom periode_metadata_aktif ke pengaturan_jadwal jika belum ada
        cursor.execute("""
            ALTER TABLE pengaturan_jadwal 
            ADD COLUMN IF NOT EXISTS periode_metadata_aktif INT DEFAULT 1
        """)
        get_app_functions()['mysql'].connection.commit()

        cursor.execute('SELECT COUNT(*) as count FROM pengaturan_jadwal')
        has_row = cursor.fetchone()['count'] > 0

        # Gunakan STR_TO_DATE untuk konversi aman (MySQL format)
        # Penting: gandakan '%' agar tidak diinterpretasikan sebagai placeholder Python DB-API

        if not has_row:
            # Insert hanya kolom yang dikirim oleh form; kolom lain akan NULL default
            columns = []
            values_sql = []
            params = []
            for col, val in [
                ('proposal_mulai', proposal_mulai),
                ('proposal_selesai', proposal_selesai),
                ('kemajuan_mulai', kemajuan_mulai),
                ('kemajuan_selesai', kemajuan_selesai),
                ('akhir_mulai', akhir_mulai),
                ('akhir_selesai', akhir_selesai),
                ('pembimbing_nilai_mulai', pembimbing_nilai_mulai),
                ('pembimbing_nilai_selesai', pembimbing_nilai_selesai),
                ('pembimbing_interval_tipe', pembimbing_interval_tipe),
                ('pembimbing_interval_nilai', pembimbing_interval_nilai),

                ('pembimbing_hari_aktif', pembimbing_hari_aktif),
                ('reviewer_proposal_mulai', reviewer_proposal_mulai),
                ('reviewer_proposal_selesai', reviewer_proposal_selesai),
                ('reviewer_kemajuan_mulai', reviewer_kemajuan_mulai),
                ('reviewer_kemajuan_selesai', reviewer_kemajuan_selesai),
                ('reviewer_akhir_mulai', reviewer_akhir_mulai),
                ('reviewer_akhir_selesai', reviewer_akhir_selesai),
            ]:
                if val is None:
                    continue  # kolom tidak dikirim oleh form
                columns.append(col)
                if str(val).strip() == '':
                    values_sql.append('NULL')
                else:
                    # Handle datetime columns vs regular columns
                    if col.endswith('_mulai') or col.endswith('_selesai'):
                        values_sql.append("STR_TO_DATE(%s, '%%Y-%%m-%%dT%%H:%%i')")
                    else:
                        values_sql.append("%s")
                    params.append(val)

            if not columns:
                # Tidak ada kolom dikirim; buat baris kosong agar konsisten dengan get_pengaturan_jadwal
                cursor.execute("INSERT INTO pengaturan_jadwal () VALUES ()")
            else:
                cursor.execute(
                    f"INSERT INTO pengaturan_jadwal ({', '.join(columns)}) VALUES ({', '.join(values_sql)})",
                    tuple(params)
                )
        else:
            # Update parsial: hanya kolom yang dikirim oleh form yang di-set
            set_clauses = []
            params = []
            for col, val in [
                ('proposal_mulai', proposal_mulai),
                ('proposal_selesai', proposal_selesai),
                ('kemajuan_mulai', kemajuan_mulai),
                ('kemajuan_selesai', kemajuan_selesai),
                ('akhir_mulai', akhir_mulai),
                ('akhir_selesai', akhir_selesai),
                ('pembimbing_nilai_mulai', pembimbing_nilai_mulai),
                ('pembimbing_nilai_selesai', pembimbing_nilai_selesai),
                ('pembimbing_interval_tipe', pembimbing_interval_tipe),
                ('pembimbing_interval_nilai', pembimbing_interval_nilai),

                ('pembimbing_hari_aktif', pembimbing_hari_aktif),
                ('reviewer_proposal_mulai', reviewer_proposal_mulai),
                ('reviewer_proposal_selesai', reviewer_proposal_selesai),
                ('reviewer_kemajuan_mulai', reviewer_kemajuan_mulai),
                ('reviewer_kemajuan_selesai', reviewer_kemajuan_selesai),
                ('reviewer_akhir_mulai', reviewer_akhir_mulai),
                ('reviewer_akhir_selesai', reviewer_akhir_selesai),
            ]:
                if val is None:
                    continue  # kolom tidak dikirim oleh form, lewati
                if str(val).strip() == '':
                    set_clauses.append(f"{col} = NULL")
                else:
                    # Handle datetime columns vs regular columns
                    if col.endswith('_mulai') or col.endswith('_selesai'):
                        set_clauses.append(f"{col} = STR_TO_DATE(%s, '%%Y-%%m-%%dT%%H:%%i')")
                    else:
                        set_clauses.append(f"{col} = %s")
                    params.append(val)

            if set_clauses:
                cursor.execute(
                    f"""
                    UPDATE pengaturan_jadwal SET {', '.join(set_clauses)}
                    WHERE id = (SELECT id FROM (SELECT id FROM pengaturan_jadwal ORDER BY id DESC LIMIT 1) as t)
                    """,
                    tuple(params)
                )

        # Update metadata kolom penilaian jika ada perubahan interval
        if penilaian_group_present and pembimbing_interval_tipe and pembimbing_interval_nilai and pembimbing_nilai_mulai and pembimbing_nilai_selesai:
            try:
                # PERBAIKAN: Cek apakah jadwal sedang berjalan
                cursor.execute("""
                    SELECT pembimbing_nilai_mulai, pembimbing_nilai_selesai 
                    FROM pengaturan_jadwal 
                    ORDER BY id DESC LIMIT 1
                """)
                current_jadwal = cursor.fetchone()
                
                jadwal_sedang_berjalan = False
                if current_jadwal and current_jadwal['pembimbing_nilai_mulai'] and current_jadwal['pembimbing_nilai_selesai']:
                    from datetime import datetime
                    now = datetime.now()
                    jadwal_mulai = current_jadwal['pembimbing_nilai_mulai']
                    jadwal_selesai = current_jadwal['pembimbing_nilai_selesai']
                    
                    if jadwal_mulai <= now <= jadwal_selesai:
                        jadwal_sedang_berjalan = True
                        logger.warning("PERINGATAN: Admin mengupdate jadwal yang sedang berjalan!")
                
                # PERBAIKAN: Jangan hapus metadata lama, buat versi baru dengan periode_id
                # Ambil periode_id terbaru
                cursor.execute("""
                    SELECT COALESCE(MAX(periode_interval_id), 0) + 1 as next_periode_id
                    FROM metadata_kolom_penilaian
                """)
                periode_result = cursor.fetchone()
                periode_id = periode_result['next_periode_id'] if periode_result else 1
                
                # Generate metadata baru dengan periode_id baru
                # Handle NULL or empty interval_nilai
                interval_nilai_int = 1  # default value
                if pembimbing_interval_nilai and str(pembimbing_interval_nilai).strip():
                    try:
                        interval_nilai_int = int(pembimbing_interval_nilai)
                    except (ValueError, TypeError):
                        interval_nilai_int = 1  # fallback to default
                
                # PERBAIKAN: Validasi dan konversi tanggal dengan lebih baik
                tanggal_mulai_dt = None
                tanggal_selesai_dt = None
                
                if pembimbing_nilai_mulai and str(pembimbing_nilai_mulai).strip():
                    try:
                        # Coba parse berbagai format tanggal
                        if 'T' in str(pembimbing_nilai_mulai):
                            tanggal_mulai_dt = datetime.strptime(str(pembimbing_nilai_mulai), '%Y-%m-%dT%H:%M')
                        else:
                            tanggal_mulai_dt = datetime.strptime(str(pembimbing_nilai_mulai), '%Y-%m-%d %H:%M:%S')
                    except Exception as e:
                        print(f"Error parsing tanggal_mulai '{pembimbing_nilai_mulai}': {str(e)}")
                        tanggal_mulai_dt = None
                
                if pembimbing_nilai_selesai and str(pembimbing_nilai_selesai).strip():
                    try:
                        # Coba parse berbagai format tanggal
                        if 'T' in str(pembimbing_nilai_selesai):
                            tanggal_selesai_dt = datetime.strptime(str(pembimbing_nilai_selesai), '%Y-%m-%dT%H:%M')
                        else:
                            tanggal_selesai_dt = datetime.strptime(str(pembimbing_nilai_selesai), '%Y-%m-%d %H:%M:%S')
                    except Exception as e:
                        print(f"Error parsing tanggal_selesai '{pembimbing_nilai_selesai}': {str(e)}")
                        tanggal_selesai_dt = None
                
                # Validate tanggal_mulai and tanggal_selesai
                if not tanggal_mulai_dt or not tanggal_selesai_dt:
                    print(f"Warning: tanggal_mulai ({pembimbing_nilai_mulai}) or tanggal_selesai ({pembimbing_nilai_selesai}) is invalid, skipping metadata generation")
                    metadata_kolom = []
                elif tanggal_mulai_dt >= tanggal_selesai_dt:
                    print(f"Warning: tanggal_mulai ({tanggal_mulai_dt}) >= tanggal_selesai ({tanggal_selesai_dt}), skipping metadata generation")
                    metadata_kolom = []
                else:
                    print(f"Generating metadata for interval {pembimbing_interval_tipe} with nilai {interval_nilai_int}")
                    print(f"Date range: {tanggal_mulai_dt} to {tanggal_selesai_dt}")
                    
                    metadata_kolom = generate_metadata_kolom_penilaian(
                        pembimbing_interval_tipe,
                        interval_nilai_int,
                        tanggal_mulai_dt,
                        tanggal_selesai_dt
                    )
                
                # Dapatkan jadwal_id yang baru saja diupdate/insert
                cursor.execute("SELECT id FROM pengaturan_jadwal ORDER BY id DESC LIMIT 1")
                jadwal_result = cursor.fetchone()
                jadwal_id = jadwal_result['id'] if jadwal_result else None
                
                # Simpan metadata baru dengan periode_id baru dan jadwal_id
                if metadata_kolom and len(metadata_kolom) > 0:
                    print(f"Inserting {len(metadata_kolom)} metadata columns...")
                    for kolom in metadata_kolom:
                        try:
                            cursor.execute("""
                                INSERT INTO metadata_kolom_penilaian 
                                (jadwal_id, periode_interval_id, nama_kolom, urutan_kolom, interval_tipe, interval_nilai, tanggal_mulai, tanggal_selesai)
                                VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                            """, (jadwal_id, periode_id, kolom['nama_kolom'], kolom['urutan_kolom'], kolom['interval_tipe'], 
                                  kolom['interval_nilai'], kolom['tanggal_mulai'], kolom['tanggal_selesai']))
                            print(f"Inserted metadata: {kolom['nama_kolom']}")
                        except Exception as insert_error:
                            print(f"Error inserting metadata {kolom['nama_kolom']}: {str(insert_error)}")
                            continue
                else:
                    print("Warning: No metadata columns generated, skipping insert")
                
                # Update pengaturan jadwal dengan periode_id aktif
                cursor.execute("""
                    UPDATE pengaturan_jadwal 
                    SET periode_metadata_aktif = %s
                    WHERE id = (SELECT id FROM (SELECT id FROM pengaturan_jadwal ORDER BY id DESC LIMIT 1) as t)
                """, (periode_id,))
                
                if metadata_kolom and len(metadata_kolom) > 0:
                    print(f"Metadata kolom berhasil diupdate: {len(metadata_kolom)} kolom untuk interval {pembimbing_interval_tipe} dengan periode_id {periode_id}")
                else:
                    print(f"Metadata kolom tidak diupdate karena tidak ada kolom yang dihasilkan untuk interval {pembimbing_interval_tipe}")
                
            except Exception as e:
                print(f"Error saat update metadata kolom: {str(e)}")
                import traceback
                traceback.print_exc()
                # Jangan gagalkan update jadwal jika metadata gagal
        
        get_app_functions()['mysql'].connection.commit()
        cursor.close()

        return jsonify({'success': True, 'message': 'Pengaturan jadwal berhasil disimpan!'})
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error saat menyimpan pengaturan jadwal: {str(e)}'})

# ========================================
# END PENGATURAN JADWAL
# ========================================

@admin_bp.route('/get_metadata_kolom_penilaian')
def get_metadata_kolom_penilaian():
    """Ambil metadata kolom penilaian yang sudah dibuat"""
    if 'user_type' not in session or session['user_type'] != 'admin':
        return jsonify({'success': False, 'message': 'Anda harus login sebagai admin!'})

    try:
        cursor = get_app_functions()['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Ambil periode aktif terbaru
        cursor.execute("""
            SELECT periode_metadata_aktif 
            FROM pengaturan_jadwal 
            ORDER BY id DESC LIMIT 1
        """)
        periode_result = cursor.fetchone()
        periode_aktif = periode_result['periode_metadata_aktif'] if periode_result else 1
        
        # Ambil metadata berdasarkan periode aktif
        cursor.execute("""
            SELECT * FROM metadata_kolom_penilaian 
            WHERE periode_interval_id = %s
            ORDER BY interval_tipe, urutan_kolom
        """, (periode_aktif,))
        
        metadata_aktif = cursor.fetchall()
        
        # Ambil semua metadata untuk history (opsional)
        cursor.execute("""
            SELECT periode_interval_id, interval_tipe, interval_nilai, COUNT(*) as jumlah_kolom,
                   MIN(created_at) as tanggal_dibuat
            FROM metadata_kolom_penilaian 
            GROUP BY periode_interval_id, interval_tipe, interval_nilai
            ORDER BY periode_interval_id DESC
        """)
        
        metadata_history = cursor.fetchall()
        cursor.close()
        
        return jsonify({
            'success': True,
            'data': metadata_aktif,
            'periode_aktif': periode_aktif,
            'history': metadata_history
        })
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error saat mengambil metadata kolom: {str(e)}'})

@admin_bp.route('/preview_metadata_kolom', methods=['POST'])
def preview_metadata_kolom():
    """Preview metadata kolom berdasarkan interval yang dipilih (tanpa menyimpan)"""
    if 'user_type' not in session or session['user_type'] != 'admin':
        return jsonify({'success': False, 'message': 'Anda harus login sebagai admin!'})

    try:
        interval_tipe = request.form.get('interval_tipe')
        interval_nilai = request.form.get('interval_nilai')
        tanggal_mulai = request.form.get('tanggal_mulai')
        tanggal_selesai = request.form.get('tanggal_selesai')
        
        if not interval_tipe or not interval_nilai or not tanggal_mulai:
            return jsonify({'success': False, 'message': 'Data interval tidak lengkap'})
        
        # Generate preview metadata
        metadata_preview = generate_metadata_kolom_penilaian(
            interval_tipe,
            int(interval_nilai),
            tanggal_mulai,
            tanggal_selesai
        )
        
        return jsonify({
            'success': True,
            'data': metadata_preview
        })
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error saat generate preview: {str(e)}'})

@admin_bp.route('/debug_metadata_kolom')
def debug_metadata_kolom():
    """Debug endpoint untuk melihat metadata kolom yang tersimpan"""
    if 'user_type' not in session or session['user_type'] != 'admin':
        return jsonify({'success': False, 'message': 'Anda harus login sebagai admin!'})

    try:
        cursor = get_app_functions()['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Ambil semua metadata kolom
        cursor.execute("""
            SELECT * FROM metadata_kolom_penilaian 
            ORDER BY periode_interval_id DESC, urutan_kolom ASC
        """)
        
        metadata_list = cursor.fetchall()
        
        # Ambil pengaturan jadwal terbaru
        cursor.execute("""
            SELECT periode_metadata_aktif, pembimbing_interval_tipe, pembimbing_interval_nilai,
                   pembimbing_nilai_mulai, pembimbing_nilai_selesai
            FROM pengaturan_jadwal 
            ORDER BY id DESC LIMIT 1
        """)
        
        jadwal_info = cursor.fetchone()
        
        cursor.close()
        
        return jsonify({
            'success': True,
            'metadata_count': len(metadata_list),
            'metadata_list': metadata_list,
            'jadwal_info': jadwal_info
        })
        
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error: {str(e)}'})

# ========================================
# PENGATURAN TOTAL ANGGARAN (Min/Max)
# ========================================

@admin_bp.route('/get_pengaturan_anggaran')
def get_pengaturan_anggaran():
    """Ambil pengaturan total anggaran (min/max) untuk anggaran awal dan bertumbuh"""
    if 'user_type' not in session or session['user_type'] != 'admin':
        return jsonify({'success': False, 'message': 'Anda harus login sebagai admin!'})

    try:
        cursor = get_app_functions()['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)

        cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS pengaturan_anggaran (
                id INT AUTO_INCREMENT PRIMARY KEY,
                min_total_awal BIGINT DEFAULT 0,
                max_total_awal BIGINT DEFAULT 0,
                min_total_bertumbuh BIGINT DEFAULT 0,
                max_total_bertumbuh BIGINT DEFAULT 0,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP
            )
            """
        )
        get_app_functions()['mysql'].connection.commit()

        # Pastikan kolom-kolom ada (untuk menangani versi lama tanpa kolom ini)
        def ensure_column(column_name: str, add_sql: str) -> None:
            cursor.execute("SHOW COLUMNS FROM pengaturan_anggaran LIKE %s", (column_name,))
            if cursor.fetchone() is None:
                cursor.execute(f"ALTER TABLE pengaturan_anggaran ADD COLUMN {add_sql}")
                get_app_functions()['mysql'].connection.commit()

        ensure_column('min_total_awal', 'min_total_awal BIGINT DEFAULT 0')
        ensure_column('max_total_awal', 'max_total_awal BIGINT DEFAULT 0')
        ensure_column('min_total_bertumbuh', 'min_total_bertumbuh BIGINT DEFAULT 0')
        ensure_column('max_total_bertumbuh', 'max_total_bertumbuh BIGINT DEFAULT 0')

        cursor.execute('SELECT * FROM pengaturan_anggaran ORDER BY id DESC LIMIT 1')
        row = cursor.fetchone()

        # Jika belum ada baris, buat default tanpa batas (0 berarti tidak dibatasi)
        if not row:
            cursor.execute(
                'INSERT INTO pengaturan_anggaran (min_total_awal, max_total_awal, min_total_bertumbuh, max_total_bertumbuh) VALUES (0, 0, 0, 0)'
            )
            get_app_functions()['mysql'].connection.commit()
            cursor.execute('SELECT * FROM pengaturan_anggaran ORDER BY id DESC LIMIT 1')
            row = cursor.fetchone()

        cursor.close()

        return jsonify({
            'success': True,
            'data': {
                'min_total_awal': int(row.get('min_total_awal') or 0),
                'max_total_awal': int(row.get('max_total_awal') or 0),
                'min_total_bertumbuh': int(row.get('min_total_bertumbuh') or 0),
                'max_total_bertumbuh': int(row.get('max_total_bertumbuh') or 0)
            }
        })
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error saat mengambil pengaturan anggaran: {str(e)}'})


@admin_bp.route('/update_pengaturan_anggaran', methods=['POST'])
def update_pengaturan_anggaran():
    """Update pengaturan total anggaran (min/max)"""
    if 'user_type' not in session or session['user_type'] != 'admin':
        return jsonify({'success': False, 'message': 'Anda harus login sebagai admin!'})

    try:
        # Ambil nilai dari form; kosong/None ditangani sebagai 0 (tidak dibatasi)
        def to_int_or_zero(val):
            try:
                if val is None or str(val).strip() == '':
                    return 0
                n = int(str(val).strip())
                return max(0, n)
            except Exception:
                return None

        min_awal = to_int_or_zero(request.form.get('min_total_awal'))
        max_awal = to_int_or_zero(request.form.get('max_total_awal'))
        min_bert = to_int_or_zero(request.form.get('min_total_bertumbuh'))
        max_bert = to_int_or_zero(request.form.get('max_total_bertumbuh'))

        # Validasi angka
        for v, name in [
            (min_awal, 'min_total_awal'), (max_awal, 'max_total_awal'),
            (min_bert, 'min_total_bertumbuh'), (max_bert, 'max_total_bertumbuh')
        ]:
            if v is None:
                return jsonify({'success': False, 'message': f'Nilai {name} harus berupa bilangan bulat >= 0!'})

        # Validasi relasi min <= max (kecuali max=0 artinya tidak dibatasi)
        if max_awal not in (0, None) and min_awal > max_awal:
            return jsonify({'success': False, 'message': 'Min Total Anggaran Awal tidak boleh melebihi Max Total Anggaran Awal!'})
        if max_bert not in (0, None) and min_bert > max_bert:
            return jsonify({'success': False, 'message': 'Min Total Anggaran Bertumbuh tidak boleh melebihi Max Total Anggaran Bertumbuh!'})

        cursor = get_app_functions()['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        cursor.execute(
            """
            CREATE TABLE IF NOT EXISTS pengaturan_anggaran (
                id INT AUTO_INCREMENT PRIMARY KEY,
                min_total_awal BIGINT DEFAULT 0,
                max_total_awal BIGINT DEFAULT 0,
                min_total_bertumbuh BIGINT DEFAULT 0,
                max_total_bertumbuh BIGINT DEFAULT 0,
                created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP
            )
            """
        )
        get_app_functions()['mysql'].connection.commit()

        # Pastikan kolom-kolom ada sebelum insert/update (untuk handle skema lama)
        def ensure_column(column_name: str, add_sql: str) -> None:
            cursor.execute("SHOW COLUMNS FROM pengaturan_anggaran LIKE %s", (column_name,))
            if cursor.fetchone() is None:
                cursor.execute(f"ALTER TABLE pengaturan_anggaran ADD COLUMN {add_sql}")
                get_app_functions()['mysql'].connection.commit()

        ensure_column('min_total_awal', 'min_total_awal BIGINT DEFAULT 0')
        ensure_column('max_total_awal', 'max_total_awal BIGINT DEFAULT 0')
        ensure_column('min_total_bertumbuh', 'min_total_bertumbuh BIGINT DEFAULT 0')
        ensure_column('max_total_bertumbuh', 'max_total_bertumbuh BIGINT DEFAULT 0')

        cursor.execute('SELECT COUNT(*) as cnt FROM pengaturan_anggaran')
        has_row = cursor.fetchone()['cnt'] > 0

        if not has_row:
            cursor.execute(
                'INSERT INTO pengaturan_anggaran (min_total_awal, max_total_awal, min_total_bertumbuh, max_total_bertumbuh) VALUES (%s,%s,%s,%s)',
                (min_awal, max_awal, min_bert, max_bert)
            )
        else:
            cursor.execute(
                """
                UPDATE pengaturan_anggaran
                SET min_total_awal=%s, max_total_awal=%s,
                    min_total_bertumbuh=%s, max_total_bertumbuh=%s
                WHERE id = (SELECT id FROM (SELECT id FROM pengaturan_anggaran ORDER BY id DESC LIMIT 1) t)
                """,
                (min_awal, max_awal, min_bert, max_bert)
            )

        get_app_functions()['mysql'].connection.commit()
        cursor.close()

        return jsonify({'success': True, 'message': 'Pengaturan total anggaran berhasil disimpan!'})
    except Exception as e:
        return jsonify({'success': False, 'message': f'Error saat menyimpan pengaturan anggaran: {str(e)}'})

@admin_bp.route('/download_excel_anggaran')
def admin_download_excel_anggaran():
    """Download Excel anggaran awal/bertumbuh dengan format kop surat untuk Admin"""
    try:
        # Cek login
        if 'user_type' not in session or session.get('user_type') != 'admin':
            return jsonify({'success': False, 'message': 'Unauthorized'}), 401
        
        id_proposal = request.args.get('id_proposal')
        jenis = request.args.get('jenis')  # 'awal' atau 'bertumbuh'
        
        if not id_proposal or not jenis:
            return jsonify({'success': False, 'message': 'Parameter tidak lengkap'}), 400
        
        # Koneksi database
        app_funcs = get_app_functions()
        cursor = app_funcs['mysql'].connection.cursor(MySQLdb.cursors.DictCursor)
        
        # Ambil data proposal
        cursor.execute("""
            SELECT p.id, p.judul_usaha, m.nama_ketua as nama, m.nim 
            FROM proposal p
            JOIN mahasiswa m ON p.nim = m.nim
            WHERE p.id = %s
        """, (id_proposal,))
        proposal_data = cursor.fetchone()
        
        if not proposal_data:
            return jsonify({'success': False, 'message': 'Proposal tidak ditemukan'}), 404
        
        # Tentukan tabel berdasarkan jenis
        tabel_anggaran = f'anggaran_{jenis}'
        
        # Ambil data anggaran
        cursor.execute(f"""
            SELECT kegiatan_utama, kegiatan, nama_barang, kuantitas, satuan, 
                   harga_satuan, jumlah, keterangan, target_capaian, penanggung_jawab
            FROM {tabel_anggaran}
            WHERE id_proposal = %s
            ORDER BY kegiatan_utama, kegiatan, nama_barang
        """, (id_proposal,))
        anggaran_data = cursor.fetchall()
        
        # Ambil tahapan usaha untuk menentukan urutan kegiatan utama
        cursor.execute("""
            SELECT tahapan_usaha FROM proposal WHERE id = %s
        """, (id_proposal,))
        proposal_info = cursor.fetchone()
        tahapan_usaha = proposal_info.get('tahapan_usaha', '').lower() if proposal_info else ''
        
        # Urutkan data sesuai urutan kegiatan utama yang standardisasi
        if 'bertumbuh' in tahapan_usaha:
            urutan_kegiatan = [
                "Pengembangan Pasar dan Saluran Distribusi",
                "Pengembangan Produk/Riset",
                "Produksi",
                "Legalitas, perizinan, sertifikasi, dan standarisasi",
                "Belanja ATK dan Penunjang"
            ]
        else:
            urutan_kegiatan = [
                "Pengembangan Produk/Riset",
                "Produksi",
                "Legalitas, perizinan, sertifikasi, dan standarisasi",
                "Belanja ATK dan Penunjang"
            ]
        
        # Buat mapping untuk case-insensitive matching dan normalisasi teks
        urutan_kegiatan_lower = [k.strip().lower() for k in urutan_kegiatan]
        
        # Fungsi helper untuk normalisasi teks kegiatan utama
        def normalize_kegiatan_utama(text):
            if not text:
                return ""
            # Normalisasi: lowercase, strip whitespace, replace multiple spaces
            normalized = text.strip().lower()
            # Handle berbagai variasi teks
            replacements = {
                "legalitas, perijinan, sertifikasi, pengujian produk, dan standarisasi": "legalitas, perizinan, sertifikasi, dan standarisasi",
                "legalitas, perijinan, sertifikasi, dan standarisasi": "legalitas, perizinan, sertifikasi, dan standarisasi",
                "legalitas perijinan sertifikasi pengujian produk dan standarisasi": "legalitas, perizinan, sertifikasi, dan standarisasi",
                "legalitas perijinan sertifikasi dan standarisasi": "legalitas, perizinan, sertifikasi, dan standarisasi"
            }
            for old, new in replacements.items():
                if old in normalized:
                    normalized = normalized.replace(old, new)
            return normalized
        
        # Sort data berdasarkan urutan kegiatan utama dengan normalisasi
        anggaran_data = sorted(
            anggaran_data,
            key=lambda x: urutan_kegiatan_lower.index(normalize_kegiatan_utama(x['kegiatan_utama'])) if x['kegiatan_utama'] and normalize_kegiatan_utama(x['kegiatan_utama']) in urutan_kegiatan_lower else 99
        )
        
        if not anggaran_data:
            return jsonify({'success': False, 'message': 'Data anggaran tidak ditemukan'}), 404
        
        # Buat workbook Excel
        wb = Workbook()
        ws = wb.active
        ws.title = f"Anggaran {jenis.title()}"
        
        # Styling
        header_font = Font(bold=True, color="000000", size=12)
        title_font = Font(bold=True, color="000000", size=14)
        border_thin = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        # Header kop surat (sama dengan laporan laba rugi)
        current_row = 1
        
        # Menggunakan format kop surat yang sama dengan laporan laba rugi
        from utils import create_pdf_header_from_kop_surat_pdf
        try:
            header_template = create_pdf_header_from_kop_surat_pdf()
            header_data = header_template['header_data']
            
            # Gunakan data dari kop surat yang sama
            for i, header_row in enumerate(header_data[:6]):  # Ambil 6 baris pertama
                if len(header_row) > 1 and header_row[1].strip():  # Jika ada teks
                    ws.merge_cells(f'A{current_row}:K{current_row}')
                    ws[f'A{current_row}'] = header_row[1].strip()
                    if i < 2:  # Baris 1-2 bold dan lebih besar
                        ws[f'A{current_row}'].font = Font(bold=True, size=12, name='Times New Roman')
                    else:  # Baris lainnya normal
                        ws[f'A{current_row}'].font = Font(bold=False, size=10, name='Times New Roman')
                    ws[f'A{current_row}'].alignment = center_alignment
                    current_row += 1
                elif not header_row[1].strip():  # Baris kosong
                    current_row += 1
        except Exception as e:
            # Fallback ke format default jika gagal
            print(f"Warning: Menggunakan fallback kop surat: {e}")
            ws.merge_cells(f'A{current_row}:K{current_row}')
            ws[f'A{current_row}'] = "YAYASAN KARTIKA EKA PAKSI"
            ws[f'A{current_row}'].font = Font(bold=True, size=12, name='Times New Roman')
            ws[f'A{current_row}'].alignment = center_alignment
            current_row += 1
            
            ws.merge_cells(f'A{current_row}:K{current_row}')
            ws[f'A{current_row}'] = "UNIVERSITAS JENDERAL ACHMAD YANI YOGYAKARTA"
            ws[f'A{current_row}'].font = Font(bold=True, size=12, name='Times New Roman')
            ws[f'A{current_row}'].alignment = center_alignment
            current_row += 1
            
            ws.merge_cells(f'A{current_row}:K{current_row}')
            ws[f'A{current_row}'] = "Jl Siliwangi Ringroad Barat, Banyuraden, Gamping, Sleman, Yogyakarta 55293"
            ws[f'A{current_row}'].font = Font(bold=False, size=10, name='Times New Roman')
            ws[f'A{current_row}'].alignment = center_alignment
            current_row += 1
            
            ws.merge_cells(f'A{current_row}:K{current_row}')
            ws[f'A{current_row}'] = "Telp. (0274) 552489, 552851 Fax. (0274) 557228 Website: www.unjaya.ac.id"
            ws[f'A{current_row}'].font = Font(bold=False, size=10, name='Times New Roman')
            ws[f'A{current_row}'].alignment = center_alignment
            current_row += 1
            
            ws.merge_cells(f'A{current_row}:K{current_row}')
            ws[f'A{current_row}'] = "Email: info@unjaya.ac.id"
            ws[f'A{current_row}'].font = Font(bold=False, size=10, name='Times New Roman')
            ws[f'A{current_row}'].alignment = center_alignment
            current_row += 1
        
        # Tambahkan spasi setelah kop surat
        current_row += 1
        
        # Judul laporan
        ws.merge_cells(f'A{current_row}:K{current_row}')
        ws[f'A{current_row}'] = f"RENCANA ANGGARAN {jenis.upper()}"
        ws[f'A{current_row}'].font = title_font
        ws[f'A{current_row}'].alignment = center_alignment
        current_row += 2
        
        # Info mahasiswa
        ws[f'A{current_row}'] = "Nama Mahasiswa:"
        ws[f'A{current_row}'].font = header_font
        ws[f'B{current_row}'] = proposal_data['nama']
        current_row += 1
        
        ws[f'A{current_row}'] = "NIM:"
        ws[f'A{current_row}'].font = header_font
        ws[f'B{current_row}'] = proposal_data['nim']
        current_row += 1
        
        ws[f'A{current_row}'] = "Nama Usaha:"
        ws[f'A{current_row}'].font = header_font
        ws[f'B{current_row}'] = proposal_data['judul_usaha']
        current_row += 1
        
        ws[f'A{current_row}'] = "Judul Usaha:"
        ws[f'A{current_row}'].font = header_font
        ws[f'B{current_row}'] = proposal_data['judul_usaha']
        current_row += 2
        
        # Header tabel dengan merged cells
        # Row 1: Header utama
        ws.cell(row=current_row, column=1, value="Kegiatan Utama").font = header_font
        ws.cell(row=current_row, column=1).fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
        ws.cell(row=current_row, column=1).border = border_thin
        ws.cell(row=current_row, column=1).alignment = center_alignment
        
        # Merge cells untuk "Rencana" (kolom B-G)
        ws.merge_cells(f'B{current_row}:G{current_row}')
        ws[f'B{current_row}'] = "Rencana"
        ws[f'B{current_row}'].font = header_font
        ws[f'B{current_row}'].fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
        ws[f'B{current_row}'].border = border_thin
        ws[f'B{current_row}'].alignment = center_alignment
        
        ws.cell(row=current_row, column=8, value="Keterangan/Ref. Harga").font = header_font
        ws.cell(row=current_row, column=8).fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
        ws.cell(row=current_row, column=8).border = border_thin
        ws.cell(row=current_row, column=8).alignment = center_alignment
        
        ws.cell(row=current_row, column=9, value="Target Capaian H=Output A").font = header_font
        ws.cell(row=current_row, column=9).fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
        ws.cell(row=current_row, column=9).border = border_thin
        ws.cell(row=current_row, column=9).alignment = center_alignment
        
        ws.cell(row=current_row, column=10, value="Penanggung Jawab").font = header_font
        ws.cell(row=current_row, column=10).fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
        ws.cell(row=current_row, column=10).border = border_thin
        ws.cell(row=current_row, column=10).alignment = center_alignment
        
        current_row += 1
        
        # Row 2: Sub-header untuk Rencana
        ws.cell(row=current_row, column=1, value="").border = border_thin  # Empty cell untuk alignment
        
        # Sub-header untuk kolom Rencana
        sub_headers = ["Kegiatan A", "Nama Barang B", "Kuantitas C", "Satuan D", "Harga Satuan E", "Jumlah (Rp) F=C×E"]
        for col, sub_header in enumerate(sub_headers, 2):
            cell = ws.cell(row=current_row, column=col, value=sub_header)
            cell.font = header_font
            cell.fill = PatternFill(start_color="E6E6FA", end_color="E6E6FA", fill_type="solid")
            cell.border = border_thin
            cell.alignment = center_alignment
        
        # Empty cells untuk kolom lainnya
        for col in range(8, 11):
            ws.cell(row=current_row, column=col, value="").border = border_thin
        
        current_row += 1
        
        # Tulis data anggaran dengan grouping berdasarkan kegiatan utama
        total_jumlah = 0
        current_kegiatan_utama = None
        kegiatan_utama_start_row = None
        kegiatan_utama_row_count = 0
        
        current_kegiatan = None
        kegiatan_start_row = None
        kegiatan_row_count = 0
        
        current_target_capaian = None
        target_capaian_start_row = None
        target_capaian_row_count = 0
        
        current_penanggung_jawab = None
        penanggung_jawab_start_row = None
        penanggung_jawab_row_count = 0
        
        for i, row_data in enumerate(anggaran_data):
            # Cek jika kegiatan utama berubah
            if current_kegiatan_utama != row_data['kegiatan_utama']:
                # Merge cells untuk kegiatan utama sebelumnya jika ada
                if current_kegiatan_utama is not None and kegiatan_utama_row_count > 1:
                    ws.merge_cells(f'A{kegiatan_utama_start_row}:A{current_row-1}')
                    ws[f'A{kegiatan_utama_start_row}'].alignment = Alignment(horizontal='center', vertical='center')
                
                # Set kegiatan utama baru
                current_kegiatan_utama = row_data['kegiatan_utama']
                kegiatan_utama_start_row = current_row
                kegiatan_utama_row_count = 0
            
            # Cek jika kegiatan berubah
            if current_kegiatan != row_data['kegiatan']:
                # Merge cells untuk kegiatan sebelumnya jika ada
                if current_kegiatan is not None and kegiatan_row_count > 1:
                    ws.merge_cells(f'B{kegiatan_start_row}:B{current_row-1}')
                    ws[f'B{kegiatan_start_row}'].alignment = Alignment(horizontal='center', vertical='center')
                
                # Set kegiatan baru
                current_kegiatan = row_data['kegiatan']
                kegiatan_start_row = current_row
                kegiatan_row_count = 0
            
            # Cek jika target capaian berubah
            if current_target_capaian != row_data['target_capaian']:
                # Merge cells untuk target capaian sebelumnya jika ada
                if current_target_capaian is not None and target_capaian_row_count > 1:
                    ws.merge_cells(f'I{target_capaian_start_row}:I{current_row-1}')
                    ws[f'I{target_capaian_start_row}'].alignment = Alignment(horizontal='center', vertical='center')
                
                # Set target capaian baru
                current_target_capaian = row_data['target_capaian']
                target_capaian_start_row = current_row
                target_capaian_row_count = 0
            
            # Cek jika penanggung jawab berubah
            if current_penanggung_jawab != row_data['penanggung_jawab']:
                # Merge cells untuk penanggung jawab sebelumnya jika ada
                if current_penanggung_jawab is not None and penanggung_jawab_row_count > 1:
                    ws.merge_cells(f'J{penanggung_jawab_start_row}:J{current_row-1}')
                    ws[f'J{penanggung_jawab_start_row}'].alignment = Alignment(horizontal='center', vertical='center')
                
                # Set penanggung jawab baru
                current_penanggung_jawab = row_data['penanggung_jawab']
                penanggung_jawab_start_row = current_row
                penanggung_jawab_row_count = 0
            
            # Tulis data
            ws.cell(row=current_row, column=1, value=row_data['kegiatan_utama']).alignment = left_alignment
            ws.cell(row=current_row, column=2, value=row_data['kegiatan']).alignment = left_alignment
            ws.cell(row=current_row, column=3, value=row_data['nama_barang']).alignment = left_alignment
            ws.cell(row=current_row, column=4, value=row_data['kuantitas']).alignment = center_alignment
            ws.cell(row=current_row, column=5, value=row_data['satuan']).alignment = center_alignment
            ws.cell(row=current_row, column=6, value=f"Rp {row_data['harga_satuan']:,}").alignment = center_alignment
            ws.cell(row=current_row, column=7, value=f"Rp {row_data['jumlah']:,}").alignment = center_alignment
            ws.cell(row=current_row, column=8, value=row_data['keterangan'] or '-').alignment = left_alignment
            ws.cell(row=current_row, column=9, value=row_data['target_capaian']).alignment = left_alignment
            ws.cell(row=current_row, column=10, value=row_data['penanggung_jawab']).alignment = left_alignment
            
            # Apply borders
            for col in range(1, 11):
                ws.cell(row=current_row, column=col).border = border_thin
            
            total_jumlah += row_data['jumlah']
            current_row += 1
            kegiatan_utama_row_count += 1
            kegiatan_row_count += 1
            target_capaian_row_count += 1
            penanggung_jawab_row_count += 1
        
        # Merge cells untuk kegiatan utama terakhir jika ada lebih dari 1 row
        if kegiatan_utama_row_count > 1:
            ws.merge_cells(f'A{kegiatan_utama_start_row}:A{current_row-1}')
            ws[f'A{kegiatan_utama_start_row}'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Merge cells untuk kegiatan terakhir jika ada lebih dari 1 row
        if kegiatan_row_count > 1:
            ws.merge_cells(f'B{kegiatan_start_row}:B{current_row-1}')
            ws[f'B{kegiatan_start_row}'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Merge cells untuk target capaian terakhir jika ada lebih dari 1 row
        if target_capaian_row_count > 1:
            ws.merge_cells(f'I{target_capaian_start_row}:I{current_row-1}')
            ws[f'I{target_capaian_start_row}'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Merge cells untuk penanggung jawab terakhir jika ada lebih dari 1 row
        if penanggung_jawab_row_count > 1:
            ws.merge_cells(f'J{penanggung_jawab_start_row}:J{current_row-1}')
            ws[f'J{penanggung_jawab_start_row}'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Total row
        ws.merge_cells(f'A{current_row}:F{current_row}')
        ws[f'A{current_row}'] = "TOTAL ANGGARAN:"
        ws[f'A{current_row}'].font = Font(bold=True)
        ws[f'A{current_row}'].alignment = Alignment(horizontal='right', vertical='center')
        ws[f'A{current_row}'].border = border_thin
        
        ws[f'G{current_row}'] = f"Rp {total_jumlah:,}"
        ws[f'G{current_row}'].font = Font(bold=True)
        ws[f'G{current_row}'].fill = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")
        ws[f'G{current_row}'].alignment = center_alignment
        ws[f'G{current_row}'].border = border_thin
        
        # Merge remaining cells in total row
        for col in range(8, 11):
            ws.cell(row=current_row, column=col).border = border_thin
        
        current_row += 2
        
        # Footer dengan tanggal
        ws[f'H{current_row}'] = f"Yogyakarta, {datetime.now().strftime('%d %B %Y')}"
        ws[f'H{current_row}'].alignment = center_alignment
        current_row += 3
        
        ws[f'H{current_row}'] = "Mahasiswa,"
        ws[f'H{current_row}'].alignment = center_alignment
        current_row += 4
        
        ws[f'H{current_row}'] = proposal_data['nama']
        ws[f'H{current_row}'].font = Font(bold=True)
        ws[f'H{current_row}'].alignment = center_alignment
        current_row += 1
        
        ws[f'H{current_row}'] = f"NIM: {proposal_data['nim']}"
        ws[f'H{current_row}'].alignment = center_alignment
        
        # Auto-adjust column widths
        for col in range(1, 11):
            column_letter = get_column_letter(col)
            max_length = 0
            for row in range(1, current_row + 1):
                try:
                    cell_value = str(ws[f'{column_letter}{row}'].value)
                    if len(cell_value) > max_length:
                        max_length = len(cell_value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Max width 50
            ws.column_dimensions[column_letter].width = max(adjusted_width, 15)  # Min width 15
        
        # Simpan ke temporary file
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
        wb.save(temp_file.name)
        temp_file.close()
        
        # Buat response untuk download
        filename = f"Anggaran_{jenis.title()}_{proposal_data['judul_usaha']}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            temp_file.name,
            as_attachment=True,
            download_name=filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        
    except Exception as e:
        print(f"Error in admin_download_excel_anggaran: {e}")
        traceback.print_exc()
        return jsonify({'success': False, 'message': 'Terjadi kesalahan saat membuat file Excel'}), 500
    finally:
        if 'cursor' in locals():
            cursor.close()

# ========================================
# END PENGATURAN TOTAL ANGGARAN
# ========================================